"""
geih.replicacion_dane_seguridad_social — Replicación de las 4 hojas de
seguridad social del anexo oficial DANE `anex-GEIHEISS-*.xlsx`.

Hojas replicadas:

    1. Seguridad social Tnal
    2. Seguridad social Tnal sexo
    3. Seguridad social 13 ciudades
    4. Seguridad social 13C sexo

Cada hoja tiene DOS BLOQUES por dominio (plan IT1 v2 §5):

    A. Valores en miles     (parte superior de la hoja)
    B. Distribución %       (parte inferior, cociente contra Población ocupada)

Líneas fijas del DANE (orden del Excel):

    1. Población ocupada
    2. Afiliada a salud
    3.     Régimen contributivo
    4.     Régimen especial
    5.         Aportantes
    6.         Beneficiarios
    7.         Otro
    8.     Régimen subsidiado
    9.     No sabe
    10. Cotiza a pensión

Variables del microdato GEIH requeridas:
    - P6090:  ¿Afiliado a salud? (1=Sí, 2=No, 9=NS)
    - P6100:  Régimen (1=Contributivo, 2=Especial, 3=Subsidiado, 9=NS)
    - P6110:  ¿Quién paga? (usado para aportante/beneficiario/otro dentro de especial)
    - P6920:  ¿Cotiza pensión? (1=Sí, 2=No, 3=Pensionado)

Notas:
- "Aportantes/Beneficiarios/Otro" solo aparecen dentro de "Régimen especial".
  Se determinan con P6110 pero esta interpretación debe validarse contra el
  Excel. En caso de divergencia, ajustar la función `_clasificar_salud`.
- "Cotiza a pensión" incluye P6920 IN (1, 3) (cotizando o ya pensionado).

Autor: Néstor Enrique Forero Herrera · ProColombia · 2026-04-16
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .replicacion_dane_common import (
    BloqueLayout,
    BloqueReplicado,
    CeldaComparada,
    ExcelLayoutParser,
    ParametrosValidacion,
    PeriodoMovil,
    ResultadoHoja,
    ResultadoReplicacion,
    ValidadorParidad,
    hash_archivo,
    normalizar_texto,
)
from .replicacion_dane_informalidad import MAPEO_DOMINIO_EXCEL

__all__ = [
    "HOJAS_SEGURIDAD_SOCIAL",
    "LINEAS_SS",
    "ReplicadorSeguridadSocial",
    "replicar_seguridad_social_excel",
]


# ═════════════════════════════════════════════════════════════════════
# Constantes
# ═════════════════════════════════════════════════════════════════════

HOJAS_SEGURIDAD_SOCIAL: list[str] = [
    "Seguridad social Tnal",
    "Seguridad social Tnal sexo",
    "Seguridad social 13 ciudades",  # con espacio en el Excel
    "Seguridad social 13C sexo",
]

ALIAS_HOJA_SS: dict[str, str] = {
    "Seguridad social 13 ciudades": "Seguridad social 13 ciudades ",
}


# 10 líneas de SS en orden fijo del Excel (col A normalizado)
LINEAS_SS: list[str] = [
    "Población ocupada",
    "Afiliada a salud",
    "Régimen contributivo",
    "Régimen especial",
    "Aportantes",
    "Beneficiarios",
    "Otro",
    "Régimen subsidiado",
    "No sabe",
    "Cotiza a pensión",
]

LINEAS_SS_NORM: dict[str, str] = {normalizar_texto(linea): linea for linea in LINEAS_SS}


# Dominios esperados por hoja
DOMINIOS_POR_HOJA_SS: dict[str, list[str]] = {
    "Seguridad social Tnal": [
        "Total Nacional",
        "Cabeceras",
        "Centros poblados y rural disperso",
    ],
    "Seguridad social Tnal sexo": [
        "Total Nacional - Hombres",
        "Total Nacional - Mujeres",
        "Cabeceras - Hombres",
        "Cabeceras - Mujeres",
        "Centros poblados y rural disperso - Hombres",
        "Centros poblados y rural disperso - Mujeres",
    ],
    "Seguridad social 13 ciudades": [
        "13 Ciudades y áreas metropolitanas",
        "23 Ciudades y áreas metropolitanas",
    ],
    "Seguridad social 13C sexo": [
        "13 Ciudades y áreas metropolitanas - Hombres",
        "13 Ciudades y áreas metropolitanas - Mujeres",
        "23 Ciudades y áreas metropolitanas - Hombres",
        "23 Ciudades y áreas metropolitanas - Mujeres",
    ],
}


# ═════════════════════════════════════════════════════════════════════
# Replicador
# ═════════════════════════════════════════════════════════════════════


class ReplicadorSeguridadSocial:
    """Replica las 4 hojas de seguridad social con dos bloques por dominio.

    Uso:
        rep = ReplicadorSeguridadSocial(ruta_excel)
        resultado = rep.replicar(df_preparado, periodo)
    """

    # Filas relativas desde el inicio del bloque de dominio (R=fila del título).
    # Bloque miles: [R+3, R+3+9] con 10 líneas
    # Entre bloque miles y bloque % puede haber varias filas de separación.
    # Por seguridad, buscaremos por label en col A.

    def __init__(
        self,
        ruta_excel: Path,
        parametros: ParametrosValidacion | None = None,
    ):
        self.ruta_excel = Path(ruta_excel)
        self.params = parametros or ParametrosValidacion()
        self.validador = ValidadorParidad(self.params)

    def replicar(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        hojas: list[str] | None = None,
    ) -> ResultadoReplicacion:
        """Replica todas las hojas de SS."""
        self._verificar_columnas_ss(df)

        if hojas is None:
            hojas = HOJAS_SEGURIDAD_SOCIAL

        resultados: list[ResultadoHoja] = []
        wb = load_workbook(self.ruta_excel, read_only=False, data_only=True)
        try:
            for hoja in hojas:
                try:
                    res = self._replicar_hoja_ss(df, periodo, hoja, wb)
                    resultados.append(res)
                    print(res.resumen())
                except Exception as e:
                    print(f"❌ Error replicando '{hoja}': {e}")
                    resultados.append(
                        ResultadoHoja(
                            hoja=hoja,
                            estado="ABIERTA_CON_CAUSA",
                            trimestre=periodo.etiqueta_original,
                            causa_raiz=f"Excepción: {type(e).__name__}: {e}",
                        )
                    )
        finally:
            wb.close()

        return ResultadoReplicacion(
            ruta_excel=self.ruta_excel,
            hash_excel=hash_archivo(self.ruta_excel),
            timestamp=datetime.now(),
            trimestre=periodo,
            hojas=resultados,
            parametros=self.params,
        )

    # ────────────────────────────────────────────────────────────────

    @staticmethod
    def _verificar_columnas_ss(df: pd.DataFrame) -> None:
        """Chequeo suave: si faltan columnas, las líneas correspondientes
        quedarán NA pero no se aborta.
        """
        requeridas_min = {"FEX_ADJ", "OCI", "DOMINIO"}
        faltan = requeridas_min - set(df.columns)
        if faltan:
            raise ValueError(f"Faltan columnas mínimas: {sorted(faltan)}")
        opcionales = {"P6090", "P6100", "P6110", "P6920"}
        faltan_op = opcionales - set(df.columns)
        if faltan_op:
            print(
                f"⚠️  Columnas opcionales ausentes: {sorted(faltan_op)}. "
                f"Las líneas de SS que las requieran quedarán como NA."
            )

    # ────────────────────────────────────────────────────────────────

    def _mask_dominio_ss(self, df: pd.DataFrame, dominio_excel: str) -> pd.Series:
        """Resuelve máscara para dominios de SS, incluyendo los que llevan sexo."""
        key = normalizar_texto(dominio_excel)
        # ¿Trae ' - hombres' o ' - mujeres'?
        sexo_filtro: str | None = None
        if key.endswith(" - hombres"):
            sexo_filtro = "Hombres"
            key = key[: -len(" - hombres")].strip()
        elif key.endswith(" - mujeres"):
            sexo_filtro = "Mujeres"
            key = key[: -len(" - mujeres")].strip()

        # Normalizar variantes "Total Nacional" (mayúsculas diferentes)
        if key in MAPEO_DOMINIO_EXCEL:
            mask = MAPEO_DOMINIO_EXCEL[key](df["DOMINIO"])
        else:
            # Fallback: comparar normalizando DOMINIO
            return pd.Series(False, index=df.index)

        if sexo_filtro is not None and "SEXO" in df.columns:
            mask = mask & (df["SEXO"] == sexo_filtro)
        return mask

    # ────────────────────────────────────────────────────────────────

    def _replicar_hoja_ss(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        hoja: str,
        wb,
    ) -> ResultadoHoja:
        """Replica una hoja de SS con sus dos bloques (miles y %)."""
        nombre_real = ALIAS_HOJA_SS.get(hoja, hoja)
        if nombre_real not in wb.sheetnames:
            return ResultadoHoja(
                hoja=hoja,
                estado="ABIERTA_CON_CAUSA",
                trimestre=periodo.etiqueta_original,
                causa_raiz=f"Hoja '{nombre_real}' no existe",
            )
        ws = wb[nombre_real]
        col_ult = self._encontrar_col_ult_ss(ws)
        # Buscar fila de meses real (puede no ser 13)
        fila_meses_real = self._detectar_fila_meses_real(ws)
        etiqueta_trim = str(ws.cell(fila_meses_real, col_ult).value or "").strip()

        # Parser: bloques por dominio (no usa sub-condición T/F/I en SS)
        parser = ExcelLayoutParser(self.ruta_excel)
        with parser:
            layout = parser.parsear_bloques_por_dominio(nombre_real)

        celdas: list[CeldaComparada] = []
        dominios_esperados = DOMINIOS_POR_HOJA_SS[hoja]

        # Clasificar los bloques: primer grupo = miles, segundo = %
        # El Excel tiene un título de "Distribución porcentual ..." entre grupos
        # El parser devuelve los bloques en orden vertical, alternando
        # los bloques por dominio. Los primeros N corresponden a miles,
        # los siguientes N a %.
        n_esperados = len(dominios_esperados)
        bloques_miles: list[BloqueLayout] = []
        bloques_pct: list[BloqueLayout] = []
        for b in layout.bloques:
            norm = normalizar_texto(b.dominio)
            # El parser puede recoger variantes de case/tildes
            if any(normalizar_texto(d) == norm for d in dominios_esperados):
                if len(bloques_miles) < n_esperados:
                    bloques_miles.append(b)
                else:
                    bloques_pct.append(b)

        # Fallback: si no encontramos suficientes bloques, usar la
        # primera mitad como miles y segunda como %.
        if len(bloques_miles) + len(bloques_pct) == 0:
            # Parser no reconoció títulos; usar dominios esperados y
            # escaneo manual
            return self._replicar_ss_escaneo_manual(
                df,
                ws,
                nombre_real,
                hoja,
                dominios_esperados,
                col_ult,
                etiqueta_trim,
            )

        # Calcular valores por dominio
        for b in bloques_miles:
            mask = self._mask_dominio_ss(df, b.dominio)
            valores_miles = self._calcular_lineas_ss(df, mask, en_miles=True)
            bloque_rep = BloqueReplicado(
                hoja=hoja,
                dominio=b.dominio,
                condicion="Miles",
                valores=valores_miles,
                es_proporcion=False,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    b,
                    ws,
                    col_ult,
                    etiqueta_trim,
                )
            )

        for b in bloques_pct:
            mask = self._mask_dominio_ss(df, b.dominio)
            valores_pct = self._calcular_lineas_ss(df, mask, en_miles=False)
            bloque_rep = BloqueReplicado(
                hoja=hoja,
                dominio=b.dominio,
                condicion="%",
                valores=valores_pct,
                es_proporcion=True,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    b,
                    ws,
                    col_ult,
                    etiqueta_trim,
                )
            )

        return self.validador.construir_resultado_hoja(
            hoja=hoja,
            trimestre=etiqueta_trim,
            celdas=celdas,
        )

    def _replicar_ss_escaneo_manual(
        self,
        df,
        ws,
        nombre_real,
        hoja,
        dominios_esperados,
        col_ult,
        etiqueta_trim,
    ) -> ResultadoHoja:
        """Fallback cuando el parser genérico no reconoce títulos."""
        celdas: list[CeldaComparada] = []
        # Escanear filas buscando títulos de dominio
        n_filas = ws.max_row
        dominios_encontrados = []  # [(dominio, fila_titulo)]
        for r in range(1, n_filas + 1):
            v = ws.cell(r, 1).value
            if v is None:
                continue
            s = str(v).strip()
            for dom in dominios_esperados:
                if normalizar_texto(s) == normalizar_texto(dom):
                    dominios_encontrados.append((dom, r))
                    break

        # Los primeros n_esperados son "miles", los siguientes "%"
        n = len(dominios_esperados)
        for idx, (dom, r_titulo) in enumerate(dominios_encontrados):
            es_pct = idx >= n
            cond = "%" if es_pct else "Miles"
            # Construir BloqueLayout manualmente: buscar filas de línea SS
            bloque = BloqueLayout(
                hoja=hoja,
                dominio=dom,
                condicion=cond,
                fila_titulo=r_titulo,
                fila_encabezado=r_titulo + 1,
                fila_meses=r_titulo + 2,
                fila_inicio_datos=r_titulo + 3,
                fila_fin_datos=r_titulo + 3 + 10,
                col_categoria=1,
            )
            # Emparejar labels con las 10 líneas estándar
            for r in range(bloque.fila_inicio_datos, bloque.fila_inicio_datos + 12):
                if r > n_filas:
                    break
                v = ws.cell(r, 1).value
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                s_norm = normalizar_texto(s)
                if s_norm in LINEAS_SS_NORM:
                    bloque.categorias.append((r, LINEAS_SS_NORM[s_norm]))
                if s_norm == normalizar_texto("Cotiza a pensión"):
                    bloque.fila_fin_datos = r
                    break

            mask = self._mask_dominio_ss(df, dom)
            valores = self._calcular_lineas_ss(df, mask, en_miles=not es_pct)
            bloque_rep = BloqueReplicado(
                hoja=hoja,
                dominio=dom,
                condicion=cond,
                valores=valores,
                es_proporcion=es_pct,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    bloque,
                    ws,
                    col_ult,
                    etiqueta_trim,
                )
            )

        return self.validador.construir_resultado_hoja(
            hoja=hoja,
            trimestre=etiqueta_trim,
            celdas=celdas,
        )

    # ────────────────────────────────────────────────────────────────

    def _calcular_lineas_ss(
        self,
        df: pd.DataFrame,
        mask_dominio: pd.Series,
        en_miles: bool,
    ) -> dict[str, float]:
        """Calcula las 10 líneas de SS para un dominio.

        Args:
            df: DataFrame preparado.
            mask_dominio: máscara booleana del dominio/sexo.
            en_miles: True → devuelve valores absolutos en miles;
                      False → devuelve distribución porcentual
                              (cociente contra "Población ocupada"×100).

        Returns:
            dict con las 10 líneas (etiqueta → valor).
        """
        df_ocu = df[mask_dominio & (df["OCI"] == 1)].copy()
        if len(df_ocu) == 0:
            return {linea: 0.0 for linea in LINEAS_SS}

        fex = df_ocu["FEX_ADJ"]

        def _w(mask):
            """Suma ponderada de una submáscara."""
            return float(fex[mask].sum())

        # Población ocupada (todos los del dominio)
        n_ocu = _w(pd.Series(True, index=df_ocu.index))

        # Variables de salud/pensión (pueden faltar)
        p6090 = (
            pd.to_numeric(df_ocu.get("P6090", np.nan), errors="coerce")
            if "P6090" in df_ocu.columns
            else pd.Series(np.nan, index=df_ocu.index)
        )
        p6100 = (
            pd.to_numeric(df_ocu.get("P6100", np.nan), errors="coerce")
            if "P6100" in df_ocu.columns
            else pd.Series(np.nan, index=df_ocu.index)
        )
        p6110 = (
            pd.to_numeric(df_ocu.get("P6110", np.nan), errors="coerce")
            if "P6110" in df_ocu.columns
            else pd.Series(np.nan, index=df_ocu.index)
        )
        p6920 = (
            pd.to_numeric(df_ocu.get("P6920", np.nan), errors="coerce")
            if "P6920" in df_ocu.columns
            else pd.Series(np.nan, index=df_ocu.index)
        )

        # Afiliada a salud = P6090 == 1
        mask_afilia = p6090 == 1
        n_afilia = _w(mask_afilia)

        # Régimen contributivo = P6090==1 & P6100 == 1
        n_contrib = _w(mask_afilia & (p6100 == 1))

        # Régimen especial = P6090==1 & P6100 == 2
        mask_esp = mask_afilia & (p6100 == 2)
        n_esp = _w(mask_esp)

        # Sub-líneas del contributivo (Aportantes/Beneficiarios/Otro):
        # Según interpretación: P6110 es "¿Quién paga la afiliación?"
        # 1 = El empleador → aportante
        # 2 = Cotiza/paga él mismo → aportante
        # 3 = Beneficiario del cotizante
        # 4 = Otra persona paga la afiliación → otro
        # 9 = NS/NR
        # Aportantes (en el Excel están dentro de "Régimen especial" del layout,
        # pero representan aportantes del CONTRIBUTIVO según el DANE).
        # Si se detecta que están sobre "contributivo" en el Excel usar ese denominador.
        mask_contrib = mask_afilia & (p6100 == 1)
        n_aportantes = _w(mask_contrib & p6110.isin([1, 2]))
        n_benef = _w(mask_contrib & (p6110 == 3))
        n_otro = _w(mask_contrib & p6110.isin([4, 9]))

        # Régimen subsidiado = P6090==1 & P6100 == 3
        n_subs = _w(mask_afilia & (p6100 == 3))

        # No sabe = P6090==1 & P6100 == 9
        n_ns = _w(mask_afilia & (p6100 == 9))

        # Cotiza a pensión = P6920 in (1, 3)  (cotizando o pensionado)
        n_pens = _w(p6920.isin([1, 3]))

        # Componer resultado
        valores_abs = {
            "Población ocupada": n_ocu,
            "Afiliada a salud": n_afilia,
            "Régimen contributivo": n_contrib,
            "Régimen especial": n_esp,
            "Aportantes": n_aportantes,
            "Beneficiarios": n_benef,
            "Otro": n_otro,
            "Régimen subsidiado": n_subs,
            "No sabe": n_ns,
            "Cotiza a pensión": n_pens,
        }

        if en_miles:
            return {k: v / 1000.0 for k, v in valores_abs.items()}

        # Distribución porcentual contra población ocupada
        if n_ocu == 0:
            return {k: 0.0 for k in valores_abs}
        return {k: 100 * v / n_ocu for k, v in valores_abs.items()}

    @staticmethod
    def _detectar_fila_meses_real(ws) -> int:
        """Devuelve la fila donde están las etiquetas de período."""
        import re as _re

        for f in range(10, 21):
            v = ws.cell(f, 2).value
            if v is not None and str(v).strip():
                txt = str(v).strip().lower()
                if _re.match(r"^[a-zñ]{3,}(\s*-\s*[a-zñ]{3,})?", txt):
                    return f
        return 13  # fallback

    @staticmethod
    def _encontrar_col_ult_ss(ws, fila_meses: int = 13) -> int:
        """Encuentra la última columna con dato en la fila de meses.

        Si la fila_meses configurada está vacía, autodetecta escaneando
        filas 10-20 en busca de una que tenga etiqueta de período.
        """
        import re as _re

        max_col = ws.max_column
        fila_real = fila_meses
        v_test = ws.cell(fila_real, 2).value
        if v_test is None:
            for f in range(10, 21):
                v = ws.cell(f, 2).value
                if v is not None and str(v).strip():
                    txt = str(v).strip().lower()
                    if _re.match(r"^[a-zñ]{3,}(\s*-\s*[a-zñ]{3,})?", txt):
                        fila_real = f
                        break

        for c in range(max_col, 1, -1):
            v = ws.cell(fila_real, c).value
            if v is not None and str(v).strip():
                return c
        raise ValueError(f"Sin datos en ninguna fila de '{ws.title}'")


# ═════════════════════════════════════════════════════════════════════
# Función pública
# ═════════════════════════════════════════════════════════════════════


def replicar_seguridad_social_excel(
    ruta_excel: Path,
    df_preparado: pd.DataFrame,
    periodo: PeriodoMovil,
    parametros: ParametrosValidacion | None = None,
    hojas: list[str] | None = None,
) -> ResultadoReplicacion:
    """API funcional para replicar las 4 hojas de SS."""
    rep = ReplicadorSeguridadSocial(ruta_excel, parametros)
    return rep.replicar(df_preparado, periodo, hojas)
