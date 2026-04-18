"""
tests/test_layout_excel_dane.py — Blindaje del parser de layout
contra cambios silenciosos del DANE.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from geih.replicacion_dane_common import ExcelLayoutParser, normalizar_texto
from geih.replicacion_dane_informalidad import ALIAS_HOJA, HOJAS_INFORMALIDAD
from geih.replicacion_dane_seguridad_social import (
    ALIAS_HOJA_SS,
    HOJAS_SEGURIDAD_SOCIAL,
)

EXCEL_OFICIAL = Path("/mnt/user-data/uploads/1776399661404_anex-GEIHEISS-dic2025-feb2026.xlsx")


pytestmark = pytest.mark.skipif(
    not EXCEL_OFICIAL.exists(),
    reason="Excel oficial no disponible en esta corrida",
)


class TestLayoutExcel:
    """Validaciones del layout del Excel oficial."""

    def test_todas_hojas_informalidad_existen(self):
        from openpyxl import load_workbook

        wb = load_workbook(EXCEL_OFICIAL, read_only=True)
        try:
            for hoja in HOJAS_INFORMALIDAD:
                nombre = ALIAS_HOJA.get(hoja, hoja)
                assert nombre in wb.sheetnames, f"Hoja '{nombre}' (alias de '{hoja}') no existe"
        finally:
            wb.close()

    def test_todas_hojas_ss_existen(self):
        from openpyxl import load_workbook

        wb = load_workbook(EXCEL_OFICIAL, read_only=True)
        try:
            for hoja in HOJAS_SEGURIDAD_SOCIAL:
                nombre = ALIAS_HOJA_SS.get(hoja, hoja)
                assert nombre in wb.sheetnames
        finally:
            wb.close()

    def test_total_nacional_tiene_3_dominios(self):
        with ExcelLayoutParser(EXCEL_OFICIAL) as parser:
            layout = parser.parsear_bloques_por_dominio("Total nacional")
        assert len(layout.bloques) == 3
        normalizados = [normalizar_texto(b.dominio) for b in layout.bloques]
        assert "total nacional" in normalizados
        assert "cabeceras" in normalizados
        assert "centros poblados y rural disperso" in normalizados

    def test_total_nacional_cada_bloque_tiene_3_categorias(self):
        with ExcelLayoutParser(EXCEL_OFICIAL) as parser:
            layout = parser.parsear_bloques_por_dominio("Total nacional")
        for b in layout.bloques:
            assert len(b.categorias) == 3, (
                f"{b.dominio}: esperaba 3 cats (Pob, Formal, Informal), got {len(b.categorias)}"
            )

    def test_grandes_dominios_tiene_5_dominios(self):
        with ExcelLayoutParser(EXCEL_OFICIAL) as parser:
            layout = parser.parsear_bloques_por_dominio("Grandes dominios ")
        # Plan v2: TN, 13 A.M., 10 ciudades, 23 A.M., Rural
        assert len(layout.bloques) == 5

    def test_educacion_bloques_por_dominio_y_condicion(self):
        """Educación tiene 3 dominios × 3 condiciones = 9 bloques, 7 niveles cada uno."""
        with ExcelLayoutParser(EXCEL_OFICIAL) as parser:
            layout = parser.parsear_bloques_por_dominio_y_condicion("Educación ")
        assert len(layout.bloques) == 9
        for b in layout.bloques:
            assert len(b.categorias) == 7, (
                f"{b.dominio}/{b.condicion}: esperaba 7 niveles, got {len(b.categorias)}"
            )

    def test_posicion_ocupacional_listas_distintas_por_condicion(self):
        """Plan v2 riesgo 7: Formal/Informal tienen distintos listados."""
        with ExcelLayoutParser(EXCEL_OFICIAL) as parser:
            layout = parser.parsear_bloques_por_dominio_y_condicion("Posición ocupacional")
        # Encontrar bloques de Total Nacional
        tn_bloques = [b for b in layout.bloques if normalizar_texto(b.dominio) == "total nacional"]
        assert len(tn_bloques) == 3  # Total, Formal, Informal
        # Los tres deben tener tamaños distintos
        tamaños = {b.condicion: len(b.categorias) for b in tn_bloques}
        # Total=8, Formal=6, Informal=7 en el Excel vigente
        assert tamaños["Total"] >= 7
        assert tamaños["Formal"] < tamaños["Total"]
        assert tamaños["Informal"] < tamaños["Total"]

    def test_ss_tnal_detecta_bloques_miles_y_pct(self):
        """Seguridad social Tnal: 3 dominios × 2 bloques (miles, %) = 6."""
        with ExcelLayoutParser(EXCEL_OFICIAL) as parser:
            layout = parser.parsear_bloques_por_dominio("Seguridad social Tnal")
        assert len(layout.bloques) >= 6

    def test_ss_13c_sexo_detecta_4x2_bloques(self):
        """SS 13C sexo: 4 dominios (13H/13M/23H/23M) × 2 bloques = 8."""
        with ExcelLayoutParser(EXCEL_OFICIAL) as parser:
            layout = parser.parsear_bloques_por_dominio("Seguridad social 13C sexo")
        assert len(layout.bloques) >= 8

    def test_ultima_etiqueta_trim_es_dic25_feb26(self):
        """El Excel vigente (abril 2026) termina en 'Dic 25 - feb 26'."""
        with ExcelLayoutParser(EXCEL_OFICIAL) as parser:
            _col, etiqueta, _ano = parser.encontrar_ultima_columna("Total nacional")
        assert "dic" in etiqueta.lower() and "feb" in etiqueta.lower()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
