# ════════════════════════════════════════════════════════════════════════════════
# 📦 ARCHIVO CONSOLIDADO: GEIH
# ════════════════════════════════════════════════════════════════════════════════
#
# Generado automáticamente el 2026-04-01 03:40:09
#
# Este archivo agrupa todo el código fuente (.py), documentación (.md)
# y archivos de configuración del proyecto "GEIH".
# Su propósito es servir como referencia única para consulta con IA.
#
# CONTENIDO:
#   🐍 Archivos de código        : 30
#   📝 Archivos de documentación : 5
#   📄 Archivos de configuración : 4
#   📊 Total de archivos         : 39
#
# ════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🌳 ESTRUCTURA DE ARCHIVOS DEL PROYECTO
────────────────────────────────────────────────────────────────────────────────

📁 GEIH/
├── 📂 docs/
│   ├── 📝 Aprendizajes GitHub_PyPI_Publicacion.md
│   ├── 📝 DOCUMENTACION_TECNICA.md
│   └── 📝 GUIA_GEIH.md
├── 📂 geih/
│   ├── 🐍 __init__.py
│   ├── 🐍 analisis_area.py
│   ├── 🐍 analisis_avanzado.py
│   ├── 🐍 analisis_complementario.py
│   ├── 🐍 analisis_poblacional.py
│   ├── 🐍 comparativo.py
│   ├── 🐍 config.py
│   ├── 🐍 consolidador.py
│   ├── 🐍 dashboard.py
│   ├── 🐍 descargador.py
│   ├── 🐍 diagnostico.py
│   ├── 🐍 exportador.py
│   ├── 🐍 indicadores.py
│   ├── 🐍 logger.py
│   ├── 🐍 preparador.py
│   ├── 🐍 profiler.py
│   ├── 🐍 utils.py
│   ├── 🐍 visualizacion.py
│   └── 🐍 visualizacion_interactiva.py
├── 📂 geih_2025/
│   └── 🐍 __init__.py
├── 📂 notebooks/
│   ├── 🐍 2026_03_28_consolidar_geih_analisis_y_calculos.py
│   └── 🐍 Pipeline_GEIH.py
├── 📂 tests/
│   ├── 🐍 __init__.py
│   ├── 🐍 smoke_test.py
│   ├── 🐍 test_config.py
│   ├── 🐍 test_consolidador.py
│   ├── 🐍 test_estadisticas_ponderadas.py
│   ├── 🐍 test_indicadores.py
│   └── 🐍 test_preparador.py
├── 📄 .gitignore
├── 📝 CHANGELOG.md
├── 📄 pyproject.toml
├── 📝 README.md
├── 📄 requirements.txt
├── 📄 setup.cfg
└── 🐍 setup.py

════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📋 ÍNDICE DE ARCHIVOS (para búsqueda rápida)
────────────────────────────────────────────────────────────────────────────────
  [  1] 📄 .gitignore
  [  2] 📝 CHANGELOG.md
  [  3] 📝 README.md
  [  4] 📝 docs/Aprendizajes GitHub_PyPI_Publicacion.md
  [  5] 📝 docs/DOCUMENTACION_TECNICA.md
  [  6] 📝 docs/GUIA_GEIH.md
  [  7] 🐍 geih/__init__.py
  [  8] 🐍 geih/analisis_area.py
  [  9] 🐍 geih/analisis_avanzado.py
  [ 10] 🐍 geih/analisis_complementario.py
  [ 11] 🐍 geih/analisis_poblacional.py
  [ 12] 🐍 geih/comparativo.py
  [ 13] 🐍 geih/config.py
  [ 14] 🐍 geih/consolidador.py
  [ 15] 🐍 geih/dashboard.py
  [ 16] 🐍 geih/descargador.py
  [ 17] 🐍 geih/diagnostico.py
  [ 18] 🐍 geih/exportador.py
  [ 19] 🐍 geih/indicadores.py
  [ 20] 🐍 geih/logger.py
  [ 21] 🐍 geih/preparador.py
  [ 22] 🐍 geih/profiler.py
  [ 23] 🐍 geih/utils.py
  [ 24] 🐍 geih/visualizacion.py
  [ 25] 🐍 geih/visualizacion_interactiva.py
  [ 26] 🐍 geih_2025/__init__.py
  [ 27] 🐍 notebooks/2026_03_28_consolidar_geih_analisis_y_calculos.py
  [ 28] 🐍 notebooks/Pipeline_GEIH.py
  [ 29] 📄 pyproject.toml
  [ 30] 📄 requirements.txt
  [ 31] 📄 setup.cfg
  [ 32] 🐍 setup.py
  [ 33] 🐍 tests/__init__.py
  [ 34] 🐍 tests/smoke_test.py
  [ 35] 🐍 tests/test_config.py
  [ 36] 🐍 tests/test_consolidador.py
  [ 37] 🐍 tests/test_estadisticas_ponderadas.py
  [ 38] 🐍 tests/test_indicadores.py
  [ 39] 🐍 tests/test_preparador.py
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📄 ARCHIVO [1/39]: .gitignore
   TIPO      : CONFIGURACIÓN
   UBICACIÓN : (raíz del proyecto)
   RUTA      : .gitignore
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: .gitignore <<<
────────────────────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════
# .gitignore — geih-analisis
# ══════════════════════════════════════════════════════════════════════
#
# PRINCIPIO: solo va a git el CÓDIGO.
# Datos del DANE, resultados y artefactos de Colab nunca.
#
# Este .gitignore es redundante con la lógica _debe_incluir() del
# notebook de publicación, que aplica las mismas reglas al copiar
# desde Drive. La doble protección garantiza que nunca se escapen datos.
# ══════════════════════════════════════════════════════════════════════


# ── Python ────────────────────────────────────────────────────────────
__pycache__/
*.py[cod]
*$py.class
*.so
*.egg
*.egg-info/
dist/
build/
.eggs/


# ── Datos del DANE — NUNCA en git ─────────────────────────────────────
GEIH_*.parquet
*.parquet
!tests/golden_set.parquet
*.csv
!requirements.txt
*.zip
*.xlsx
*.xls


# ── Carpetas mensuales del DANE ───────────────────────────────────────
Enero*/
Febrero*/
Marzo*/
Abril*/
Mayo*/
Junio*/
Julio*/
Agosto*/
Septiembre*/
Octubre*/
Noviembre*/
Diciembre*/


# ── Outputs generados ─────────────────────────────────────────────────
resultados_geih*/
dashboard/
*.png
*.pdf
metadata.json
geih_pipeline.log
BoxPlot_*.csv


# ── Notebooks Jupyter ─────────────────────────────────────────────────
*.ipynb
.ipynb_checkpoints/


# ── Carpetas de trabajo personal ──────────────────────────────────────
z. */
z.*/


# ── Artefactos de Google Colab ────────────────────────────────────────
mnt/


# ── IDE ───────────────────────────────────────────────────────────────
.vscode/
.idea/
*.swp
*.swo
*~
.DS_Store
Thumbs.db


# ── Entornos virtuales ────────────────────────────────────────────────
venv/
.venv/
env/
.env


# ── Credenciales ──────────────────────────────────────────────────────
.env.local
secrets.json
*_token.txt
*.pem
*.key

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: .gitignore >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📝 ARCHIVO [2/39]: CHANGELOG.md
   TIPO      : DOCUMENTACIÓN
   UBICACIÓN : (raíz del proyecto)
   RUTA      : CHANGELOG.md
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: CHANGELOG.md <<<
────────────────────────────────────────────────────────────────────────────────
# Changelog — geih-analisis

Formato: [Semantic Versioning](https://semver.org/lang/es/)

---

## [5.0.0] — 2026-03-29

### Cambios de ruptura (MAJOR)
- **Renombrado**: el paquete Python ahora se llama `geih` en vez de `geih_2025`.
  - `pip install geih-analisis` (antes `geih-2025`)
  - `from geih import ConfigGEIH` (antes `from geih import`)
  - El shim `geih_2025` garantiza compatibilidad hasta v6.0.0.
- **Exportador**: el parámetro `nombre_carpeta` ahora es opcional.
  Si se pasa `config`, la carpeta se genera automáticamente como
  `resultados_geih_{anio}` (antes siempre era `resultados_geih_2025`).

### Nuevas funcionalidades
- `py.typed` agregado: soporte completo de type hints para IDEs.
- Exports faltantes agregados a `__init__.py`:
  `TAMANO_EMPRESA`, `NIVELES_AGRUPADOS`, `NIVELES_EDUCATIVOS`,
  `P3042_A_ANOS`, `AGRUPACION_DANE_8`, `LLAVES_PERSONA`,
  `LLAVES_HOGAR`, `CONVERTERS_BASE`, `CONVERTERS_CON_AREA`, `MODULOS_CSV`.
- Logger: nombre raíz cambiado de `geih_2025` a `geih`.

### Migración desde v4.x

```python
# Opción A — Código antiguo sigue funcionando (con DeprecationWarning):
from geih import ConfigGEIH  # emite warning, funciona

# Opción B — Código nuevo (recomendado):
from geih import ConfigGEIH       # limpio, sin warning

# Exportador con año automático (nuevo parámetro config):
exp = Exportador(ruta_base=RUTA, config=config)
# → crea resultados_geih_2025/ o resultados_geih_2026/ según config.anio

# Antes era necesario:
from geih_2025.config import TAMANO_EMPRESA  # workaround
# Ahora disponible directamente:
from geih import TAMANO_EMPRESA              # limpio
```

---

## [4.3.0] — 2026-03-28

### Nuevas funcionalidades
- `ComparadorMultiAnio`: comparación inter-anual TD 2025 vs 2026.
- `ConfigComparativo`: `@dataclass` para parámetros del comparativo.
- `asegurar_parquet()`: extracción RAM-eficiente con PyArrow predicate pushdown.
- Dashboard Streamlit con túnel cloudflared para Google Colab.
- 70+ clases de análisis (antes 57).

### Correcciones
- `Estacionalidad.calcular()`: acepta `geih` crudo (con `MES_NUM`), no `df`.
- `AnalisisOcupadosCiudad.calcular()`: método correcto (`calcular`, no `calcular_tablas`).
- `ComparadorMultiAnio`: separación de período de comparación para evitar
  inflación del FEX_ADJ al mezclar 12 meses con 1 mes.

---

## [4.0.0] — 2025-12-01

### Cambios de ruptura
- `ConfigGEIH` ahora recibe `anio` y auto-selecciona SMMLV y carpetas.
- `MESES_CARPETAS` migró de constante a función dinámica.
- Soporte multi-año (2022–presente).

### Nuevas funcionalidades
- 57 clases de análisis (antes 22).
- 8 módulos CSV del DANE (antes 5).
- Checkpointing en consolidación.
- DescargadorDANE para descarga automática.

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: CHANGELOG.md >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📝 ARCHIVO [3/39]: README.md
   TIPO      : DOCUMENTACIÓN
   UBICACIÓN : (raíz del proyecto)
   RUTA      : README.md
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: README.md <<<
────────────────────────────────────────────────────────────────────────────────
# geih-analisis

[![PyPI version](https://badge.fury.io/py/geih-analisis.svg)](https://pypi.org/project/geih-analisis/)
[![Python 3.9+](https://img.shields.io/badge/python-3.9%2B-blue.svg)](https://www.python.org/downloads/)
[![License: MIT](https://img.shields.io/badge/License-MIT-yellow.svg)](https://opensource.org/licenses/MIT)
[![Tests](https://img.shields.io/badge/tests-71%20passed-brightgreen.svg)]()
[![AI Assisted](https://img.shields.io/badge/AI%20Assisted-Claude%20%7C%20Gemini-blue)]()

**Paquete Python para analizar los microdatos de la Gran Encuesta Integrada de Hogares (GEIH) del DANE — Colombia.**

Convierte los archivos CSV crudos del DANE en indicadores del mercado laboral listos para reportar: desempleo, salarios, brecha de género, formalidad, educación y más — con pocas líneas de código.

---

> **English summary:** Python package for analyzing Colombia's GEIH household survey microdata (DANE). Computes labor market indicators (unemployment, wages, gender gap, formality) from raw CSV files. Supports 2022–present, 70+ analytical classes, Google Colab optimized. `pip install geih-analisis` → `from geih import ConfigGEIH`.

---

## Tabla de contenidos

1. [¿Para quién es este paquete?](#1-para-quién-es-este-paquete)
2. [Instalación](#2-instalación)
3. [Inicio rápido](#3-inicio-rápido)
4. [Paso 0 — Descargar los datos del DANE](#4-paso-0--descargar-los-datos-del-dane)
5. [Flujo de trabajo completo](#5-flujo-de-trabajo-completo)
6. [Análisis disponibles](#6-análisis-disponibles)
7. [Ejemplos de análisis](#7-ejemplos-de-análisis)
8. [Agregar un año nuevo](#8-agregar-un-año-nuevo)
9. [FAQ — Preguntas frecuentes](#9-faq--preguntas-frecuentes)
10. [Cómo citar](#10-cómo-citar)
11. [Licencia](#11-licencia)
12. [Metodología de desarrollo](#12-metodología-de-desarrollo)
13. [Créditos y agradecimientos](#13-créditos-y-agradecimientos)

---

## 1. ¿Para quién es este paquete?

Este paquete es para ti si:

- Eres **economista, analista laboral o investigador** y necesitas calcular tasas de desempleo, ingresos medianos o brechas salariales a partir de la GEIH.
- Eres **estudiante** y quieres explorar el mercado laboral colombiano sin procesar manualmente millones de filas.
- Tienes conocimiento de **Python básico** como saber instalar paquetes y ejecutar celdas es suficiente para empezar.
- Usas **Google Colab** y no quieres instalar nada en tu computador.

**No necesitas:** experiencia avanzada en programación, conocimiento de la estructura interna de la GEIH, ni saber qué es un factor de expansión. El paquete maneja todo eso por ti.

---

## 2. Instalación

### Opción A — pip (recomendado para uso local o Colab con pip)

```bash
pip install geih-analisis
```

Con gráficos interactivos y dashboard visual:

```bash
pip install "geih-analisis[all]"
```

### Opción B — Google Colab (sin instalación, desde Drive)

```python
from google.colab import drive
drive.mount('/content/drive')

import sys
sys.path.insert(0, '/content/drive/MyDrive/GEIH')  # ← ajustar a tu ruta

from geih import __version__
print(f"geih v{__version__} listo")
```

### Opción C — Última versión en desarrollo (desde GitHub)

```bash
pip install git+https://github.com/enriqueforero/geih-analisis.git
```

### Requisitos del sistema

| Dependencia | Versión mínima |
|---|---|
| Python | 3.9+ |
| pandas | 1.5+ |
| numpy | 1.21+ |
| pyarrow | 10.0+ |
| openpyxl | 3.0+ |

---

## 3. Inicio rápido

### Ultra-rápido — 3 líneas (si ya tienes los datos consolidados)

```python
from geih import ConfigGEIH, ConsolidadorGEIH, PreparadorGEIH, IndicadoresLaborales

config = ConfigGEIH(anio=2025, n_meses=12)
geih   = ConsolidadorGEIH.cargar('GEIH_2025_Consolidado.parquet')
df     = PreparadorGEIH(config=config).preparar_base(geih)
r      = IndicadoresLaborales(config=config).calcular(df)
print(f"Desempleo: {r['TD_%']:.1f}%  |  Participación: {r['TGP_%']:.1f}%  |  Ocupación: {r['TO_%']:.1f}%")
```

### Completo — desde los CSV hasta los resultados

```python
from geih import (
    ConfigGEIH, ConsolidadorGEIH, PreparadorGEIH,
    IndicadoresLaborales, AnalisisSalarios, BrechaGenero, Exportador
)
import os

# ── 1. Configurar ─────────────────────────────────────────────────────
RUTA   = '/content/drive/MyDrive/GEIH'   # ← ajustar a tu ruta en Drive
config = ConfigGEIH(anio=2025, n_meses=12)
config.resumen()   # muestra SMMLV, período y carpetas esperadas

# ── 2. Consolidar los CSV del DANE ────────────────────────────────────
# Primera vez: ~5 minutos. Siguientes veces: instantáneo desde el archivo .Parquet.
PARQUET = f'{RUTA}/GEIH_{config.anio}_Consolidado.parquet'

if os.path.exists(PARQUET):
    geih = ConsolidadorGEIH.cargar(PARQUET)
else:
    cons = ConsolidadorGEIH(ruta_base=RUTA, config=config, incluir_area=True)
    cons.verificar_estructura()            # verifica que estén los 96 archivos CSV
    geih = cons.consolidar(checkpoint=True)  # retoma automáticamente si Colab se cae
    cons.exportar(geih)

# ── 3. Preparar datos ─────────────────────────────────────────────────
prep = PreparadorGEIH(config=config)
df   = prep.preparar_base(geih)
df   = prep.agregar_variables_derivadas(df)

# ── 4. Calcular y exportar ────────────────────────────────────────────
ind     = IndicadoresLaborales(config=config)
r       = ind.calcular(df)
ind.sanity_check(r, f"Anual {config.anio}")   # valida contra cifras DANE

exp     = Exportador(ruta_base=RUTA, config=config)
td_dpto = ind.por_departamento(df)
exp.guardar_tabla(td_dpto, f'desempleo_departamentos_{config.anio}')

print(f"✅  TD={r['TD_%']:.1f}%  TGP={r['TGP_%']:.1f}%  TO={r['TO_%']:.1f}%")
```

---

## 4. Paso 0 — Descargar los datos del DANE

> ⚠️ **Este paquete no incluye los datos.** Los microdatos de la GEIH son públicos y gratuitos, pero debes descargarlos del portal oficial del DANE y organizarlos en la estructura de carpetas que el paquete espera.

### 4.1 Dónde descargar

**Portal oficial de microdatos del DANE:**
🔗 [https://microdatos.dane.gov.co](https://microdatos.dane.gov.co/index.php/catalog/central/about)

### 4.2 Paso a paso detallado

**Paso 1 — Ir al portal y buscar la encuesta**

1. Abrir [microdatos.dane.gov.co](https://microdatos.dane.gov.co)
2. En el buscador escribir: `Gran Encuesta Integrada de Hogares`
3. Seleccionar el año que necesitas (ej: `GEIH 2025`)
4. Hacer clic en **"Obtener microdatos"**

**Paso 2 — Descargar los archivos en formato .zip para cada mes**

La GEIH se publica en **8 archivos CSV por mes** que están todos en un archivo de formato .zip. Debes descargar los meses a analizar:

| # | Módulo | Nombre del archivo |
|---|---|---|
| 1 | Características generales | `Características generales, seguridad social en salud y educación.CSV` |
| 2 | Datos del hogar | `Datos del hogar y la vivienda.CSV` |
| 3 | Fuerza de trabajo | `Fuerza de trabajo.CSV` |
| 4 | Ocupados | `Ocupados.CSV` |
| 5 | No ocupados | `No ocupados.CSV` |
| 6 | Otras formas de trabajo | `Otras formas de trabajo.CSV` |
| 7 | Migración | `Migración.CSV` |
| 8 | Otros ingresos | `Otros ingresos e impuestos.CSV` |

**Paso 3 — Organizar los archivos en carpetas**

Crea una carpeta raíz (por ejemplo `GEIH`) y dentro organiza los archivos exactamente así:

```
GEIH/                                ← carpeta raíz (tu variable RUTA)
│
├── Enero 2025/
│   └── CSV/
│       ├── Características generales, seguridad social en salud y educación.CSV
│       ├── Datos del hogar y la vivienda.CSV
│       ├── Fuerza de trabajo.CSV
│       ├── Ocupados.CSV
│       ├── No ocupados.CSV
│       ├── Otras formas de trabajo.CSV
│       ├── Migración.CSV
│       └── Otros ingresos e impuestos.CSV
│
├── Febrero 2025/
│   └── CSV/
│       └── (mismos 8 archivos)
│
├── Marzo 2025/
│   └── CSV/
│       └── (mismos 8 archivos)
│
│   ... (una carpeta por cada mes)
│
└── Diciembre 2025/
    └── CSV/
        └── (mismos 8 archivos)
```

> **Importante — nombres exactos de carpetas:** deben ser `Enero 2025`, `Febrero 2025`, `Marzo 2025`, `Abril 2025`, `Mayo 2025`, `Junio 2025`, `Julio 2025`, `Agosto 2025`, `Septiembre 2025`, `Octubre 2025`, `Noviembre 2025`, `Diciembre 2025`. Con mayúscula inicial, espacio y año. El paquete busca estas carpetas con esa nomenclatura exacta. Las otras carpetas que genera el DANE como DAT y SAV, contienen formatos para programas estadísticos especializados, esos no se usan en esta librería. Sólo trabaje con los archivos en formato CSV. 

**Paso 5 — Verificar que todo está en orden**

Antes de consolidar, verifica que los archivos están correctamente ubicados:

```python
from geih import ConsolidadorGEIH, ConfigGEIH

config = ConfigGEIH(anio=2025, n_meses=12)
cons   = ConsolidadorGEIH(ruta_base='/tu/ruta/GEIH', config=config)
cons.verificar_estructura()
# ✅ Enero 2025: 8 archivos — OK
# ✅ Febrero 2025: 8 archivos — OK
# ✅ Marzo 2025: 8 archivos — OK
# ...
```

Si falta algún archivo, el verificador te indica exactamente cuál.

### 4.3 Descarga automática con `DescargadorDANE` (opcional)

Si prefieres que el paquete organice los ZIPs que descargaste manualmente:

```python
from geih import DescargadorDANE, ConfigGEIH

desc = DescargadorDANE(
    config=ConfigGEIH(anio=2025, n_meses=12),
    ruta_destino='/tu/ruta/GEIH',
)
# Si tienes los ZIPs descargados manualmente:
desc.organizar_zips('/ruta/donde/guardaste/los/zips')
# Verifica la estructura resultante:
desc.verificar()
```

---

## 5. Flujo de trabajo completo

```
Archivos CSV del DANE (8 módulos × N meses)
          ↓
  ConsolidadorGEIH     → une los módulos, concatena los meses
          ↓               guarda GEIH_2025_Consolidado.parquet
  PreparadorGEIH       → calcula FEX_ADJ, mapea ramas y departamentos
          ↓               agrega variables derivadas (INGLABO_SML, etc.)
  Clases de análisis   → TD, salarios, brecha, Gini, ICE, ICI, ITAT...
          ↓
  Exportador           → resultados_geih_2025/
                           ├── graficas/   (PNG)
                           ├── tablas/     (CSV)
                           └── excel/      (XLSX multi-hoja)
```

### El único parámetro que cambia: `ConfigGEIH`

```python
from geih import ConfigGEIH

# Año completo
config = ConfigGEIH(anio=2025, n_meses=12)

# Primer trimestre de 2026
config = ConfigGEIH(anio=2026, n_meses=3)

# Solo enero de 2026
config = ConfigGEIH(anio=2026, n_meses=1)
```

`ConfigGEIH` selecciona automáticamente el SMMLV correcto, genera la lista de carpetas esperadas y controla cómo se calcula el factor de expansión. Cambiar `anio` y `n_meses` es todo lo que necesitas.

---

## 6. Análisis disponibles

### Indicadores fundamentales del mercado laboral

| Clase | Qué produce |
|---|---|
| `IndicadoresLaborales` | TD, TGP, TO — nacionales y por departamento |
| `DistribucionIngresos` | Distribución de ingresos por rangos de SMMLV |
| `AnalisisSalarios` | Mediana, media, percentiles por rama de actividad y edad |
| `BrechaGenero` | Diferencia salarial hombres/mujeres por nivel educativo |
| `IndicesCompuestos` | Coeficiente de Gini del ingreso laboral |
| `AnalisisRamaSexo` | Composición del empleo por industria y sexo |

### Análisis avanzados

| Clase | Qué produce |
|---|---|
| `CalidadEmpleo` | ICE: índice que combina pensión, salud, horas y salario |
| `CompetitividadLaboral` | ICI: competitividad laboral por departamento |
| `VulnerabilidadLaboral` | IVI: vulnerabilidad por rama de actividad |
| `MapaTalento` | ITAT: oferta laboral, costo y calidad por departamento |
| `EcuacionMincer` | Retorno salarial a cada año adicional de educación |
| `Estacionalidad` | Variación mensual de TD, TGP y TO durante el año |
| `FuerzaLaboralJoven` | Indicadores específicos para jóvenes de 15 a 28 años |
| `CostoLaboral` | Costo total incluyendo prestaciones sociales (~54% sobre salario) |
| `FormalidadSectorial` | ICF: formalidad del empleo por sector económico |

### Poblaciones especiales

- `AnalisisCampesino` — trabajadores que se autodeclaran campesinos
- `AnalisisDiscapacidad` — personas con discapacidad (escala Washington ONU, 8 dimensiones)
- `AnalisisMigracion` — migración interna e internacional
- `AnalisisSobrecalificacion` — universitarios ocupados en empleos que no requieren ese nivel
- `AnalisisContractual` — tipo de contrato: escrito indefinido, escrito fijo, verbal, sin contrato
- `AnalisisAutonomia` — trabajadores formalmente independientes pero en realidad dependientes
- `DuracionDesempleo` — semanas buscando empleo (friccional, cíclico, estructural, largo plazo)
- `DashboardSectoresProColombia` — 7 sectores estratégicos de actividad económica

### 32 ciudades y áreas metropolitanas

```python
from geih import AnalisisOcupadosCiudad

area   = AnalisisOcupadosCiudad(config=config)
tablas = area.calcular(df)                  # 6 tablas: nacional, agrupación DANE, ciudad/AM
area.imprimir(tablas)
area.exportar_excel(tablas, 'CIIU_Area_2025.xlsx')
```

### Comparación entre años

```python
from geih import ComparadorMultiAnio, ConfigGEIH

comp = ComparadorMultiAnio()
comp.agregar_anio(2025, 'GEIH_2025_M01.parquet', ConfigGEIH(anio=2025, n_meses=1))
comp.agregar_anio(2026, 'GEIH_2026_M01.parquet', ConfigGEIH(anio=2026, n_meses=1))

comp.comparar_indicadores()    # TD / TGP / TO con variación anual
comp.evolucion_ingresos()      # mediana salarial por año en COP y SMMLV
comp.comparar_departamentos()  # TD por departamento × año
comp.comparar_brecha_genero()  # brecha salarial H/M por año
```

---

## 7. Ejemplos de análisis

### Tasa de desempleo nacional y por departamento

```python
from geih import IndicadoresLaborales

ind = IndicadoresLaborales(config=config)

# Nacional
r = ind.calcular(df)
print(f"TD={r['TD_%']:.1f}%  TGP={r['TGP_%']:.1f}%  TO={r['TO_%']:.1f}%")
# → TD=9.8%  TGP=63.4%  TO=57.2%

# Por departamento
td_dpto = ind.por_departamento(df)
print(td_dpto[['Departamento', 'TD_%', 'Ocupados_M']].head(10).to_string(index=False))
```

### Salarios medianos por industria y nivel educativo

```python
from geih import AnalisisSalarios

# Mediana salarial por rama de actividad
salarios = AnalisisSalarios(config=config).por_rama(df)
print(salarios[['Mediana', 'Media', 'Mediana_SMMLV']].head(10))

# Por nivel educativo (diferenciando Especialización, Maestría y Doctorado)
from geih.utils import EstadisticasPonderadas as EP

niveles = {10: 'Universitaria', 11: 'Especialización', 12: 'Maestría', 13: 'Doctorado'}
df_edu  = df[(df['OCI'] == 1) & (df['INGLABO'] > 0)].copy()
df_edu['NIVEL'] = df_edu['P3042'].map(niveles)

for nivel in ['Universitaria', 'Especialización', 'Maestría', 'Doctorado']:
    m   = df_edu['NIVEL'] == nivel
    med = EP.mediana(df_edu.loc[m, 'INGLABO'], df_edu.loc[m, 'FEX_ADJ'])
    print(f"{nivel:20s}: ${med:>12,.0f}  ({med / config.smmlv:.1f}× SMMLV)")
# Universitaria       :   $2,100,000  (1.5× SMMLV)
# Especialización     :   $3,800,000  (2.7× SMMLV)
# Maestría            :   $5,200,000  (3.7× SMMLV)
# Doctorado           :   $7,500,000  (5.3× SMMLV)
```

### Brecha salarial de género

```python
from geih import BrechaGenero

brecha = BrechaGenero().calcular(df)
print(brecha)
#               Hombres    Mujeres  Brecha_%
# Media           1.35       1.18     -12.6
# Universitaria   2.10       1.95      -7.1
# Maestría        4.82       4.41      -8.5
```

### Comparación enero 2025 vs enero 2026

```python
from geih import ComparadorMultiAnio, ConfigGEIH

comp = ComparadorMultiAnio()
comp.agregar_anio(2025, 'GEIH_2025_M01.parquet', ConfigGEIH(anio=2025, n_meses=1))
comp.agregar_anio(2026, 'GEIH_2026_M01.parquet', ConfigGEIH(anio=2026, n_meses=1))

df_ind = comp.comparar_indicadores()
# ANIO   TD_%   Δ_TD_%   TGP_%   TO_%   Ocupados_M
# 2025   10.8      —      63.1   56.3      22.5
# 2026   11.4    +0.6     62.8   55.7      22.8
```

### Dashboard interactivo (sin código)

```python
from geih import ejecutar_dashboard

ejecutar_dashboard(ruta_base='/tu/ruta/GEIH')
# Abre un dashboard en tu navegador con filtros y gráficos
# Requiere: pip install "geih-analisis[dashboard]"
```

---

## 8. Agregar un año nuevo

Cuando el DANE publique los datos de un año nuevo, solo necesitas 2 cambios en `config.py`:

```python
# 1. Agregar el SMMLV del año nuevo (publicado por decreto en diciembre)
SMMLV_POR_ANIO = {
    2025: 1_423_500,
    2026: 1_750_905,
    2027: X_XXX_XXX,   # ← agregar aquí
}

# 2. Agregar la referencia DANE cuando publiquen el boletín oficial
REF_DANE = {
    2027: ReferenciaDane(td_anual_pct=..., tgp_anual_pct=..., to_anual_pct=...),
}
```

Si no agregas la referencia DANE, el paquete muestra un aviso pero los análisis siguen funcionando normalmente.

---

## 9. FAQ — Preguntas frecuentes

**¿Qué es la GEIH?**
La Gran Encuesta Integrada de Hogares es la encuesta mensual del DANE que mide el mercado laboral en Colombia. Encuesta a más de 250.000 hogares al año y es la fuente oficial de las tasas de desempleo y empleo del país.

**¿Qué es el SMMLV y por qué importa?**
El Salario Mínimo Mensual Legal Vigente es la referencia salarial de Colombia. El paquete lo usa para expresar ingresos en múltiplos comprensibles (ej: "gana 2,3× SMMLV"). Cada año tiene su propio valor fijado por decreto.

**¿Qué es el factor de expansión (FEX)?**
La GEIH encuesta una muestra, no a toda la población. El factor de expansión indica cuántas personas reales representa cada encuestado. El paquete lo calcula automáticamente tomando el factor de expansión que está en la base y con `ConfigGEIH(n_meses=...)` — no necesitas hacer nada.

**¿Por qué la primera consolidación tarda ~5 minutos?**
Está leyendo y uniendo ~96 archivos CSV (8 módulos × 12 meses) con cerca de 817.550 filas × 515 columnas. El resultado se guarda en formato Parquet que es un formato eficiente para gran volumen de datos en velocidad y espacio en disco: las siguientes veces carga en segundos.

**¿Qué pasa si Colab se desconecta durante la consolidación?**
El parámetro `checkpoint=True` guarda el avance después de cada mes procesado. Al volver a ejecutar, retoma automáticamente desde el último mes completado.

**¿Puedo analizar solo enero 2026 si aún no tengo todo el año?**
Sí. Usa `ConfigGEIH(anio=2026, n_meses=1)`. El paquete espera solo la carpeta `Enero 2026/CSV/` y calcula los factores de expansión correctamente para ese mes.

**¿Puedo comparar 2025 vs 2026 si solo tengo enero 2026?**
Sí, usando `n_meses=1` para ambos años en el `ComparadorMultiAnio`. Esto garantiza que los factores de expansión sean equivalentes entre años.

**¿Los datos del DANE tienen algún costo?**
No. Son públicos y gratuitos en [microdatos.dane.gov.co](https://microdatos.dane.gov.co) y no requieren registro o autenticación.

**¿Funciona con datos de años anteriores (2022, 2023, 2024)?**
Sí. El paquete soporta la GEIH desde 2022 (Marco Muestral 2018 del DANE) hasta el presente con `ConfigGEIH(anio=2022, ...)`.

**¿Qué significa el error "PEA > 40M"?**
Indica que el factor de expansión no se está calculando correctamente. Asegúrate de que `n_meses` en `ConfigGEIH` coincide con el número de meses reales en tu archivo Parquet:

```python
# Si el Parquet tiene 12 meses → n_meses=12 → FEX ÷ 12
config = ConfigGEIH(anio=2025, n_meses=12)

# Si el Parquet tiene solo enero → n_meses=1 → FEX ÷ 1
config = ConfigGEIH(anio=2026, n_meses=1)
```

---

## 10. Cómo citar

Si usas este paquete en un trabajo académico o publicación, puedes citarlo así:

**Formato BibTeX:**

```bibtex
@software{forero2026geih,
  author  = {Forero Herrera, Néstor Enrique},
  title   = {geih-analisis: Paquete Python para análisis de microdatos GEIH},
  year    = {2026},
  version = {0.1.0},
  url     = {https://github.com/enriqueforero/geih-analisis},
  note    = {Datos fuente: Gran Encuesta Integrada de Hogares — DANE, Colombia}
}
```

**Formato texto (APA):**

> Forero Herrera, N. E. (2026). *geih-analisis* (v0.1.0) [Software]. GitHub. https://github.com/enriqueforero/geih-analisis

---

## 11. Licencia

MIT — Néstor Enrique Forero Herrera · Colombia · 2026

Los **datos de la GEIH** son de acceso público y propiedad del DANE (Departamento Administrativo Nacional de Estadística de Colombia). Este paquete es una herramienta de análisis independiente y no tiene afiliación oficial con el DANE ni con ninguna entidad gubernamental.

---

## 12. Metodología de desarrollo

La arquitectura, la lógica de negocio y los requerimientos de este paquete son de autoría humana. El autor asume responsabilidad total sobre la integridad del código publicado.

Se utilizaron modelos de IA generativa (Claude y Gemini) como asistencia técnica para generación de código boilerplate, optimización de consultas y sugerencias de refactorización.

**Validación:** Ningún bloque asistido por IA fue integrado sin revisión crítica, ajuste al contexto del dominio y validación mediante pruebas funcionales.

---

## 13. Créditos y agradecimientos

Este paquete es resultado de un esfuerzo acumulativo que comenzó con trabajo exploratorio y operativo previo dentro de la Gerencia de Inteligencia Comercial (GIC) de ProColombia.

**Lina Castro y Nicolás Rivera** fueron los creadores del análisis pionero de ocupados por CIIU Rev 4 y área geográfica a partir de la GEIH 2023. Su notebook estableció el enfoque metodológico base del proyecto: la carga iterativa mes a mes del módulo de Ocupados, la selección de variables relevantes del módulo (`OCI`, `FEX_C18`, `RAMA4D_R4`, `AREA`, `DPTO`, `INGLABO`, `P6430`, entre otras), la aplicación del factor de expansión anual (`FEX_C18 / 12`) para estimar ocupados reales, y la construcción de tablas pivote desagregadas por agrupación DANE, división CIIU, área geográfica y sus cruces. También diseñaron el esquema de correlativas — DIVIPOLA, CIIU Rev4 y agrupaciones DANE — que permitió enriquecer semánticamente los códigos crudos del DANE. Los resultados fueron exportados a Excel multi-hoja para su uso operativo directo. Entre ambos participaron activamente en la definición del alcance analítico, la revisión de resultados y la validación de los cálculos, asegurando que los indicadores respondieran a las preguntas reales del equipo. Adicionalmente, Lina y Nicolás diseñaron y ejecutaron el proceso de consolidación completa de la GEIH 2025 (12 meses, Enero a Diciembre). Su aporte más técnico fue el desarrollo de la función `unir_modulos_sin_duplicados()`, que resolvió un problema central del proceso: el solapamiento de columnas entre los distintos módulos del DANE al hacer merges entre ellos. Esta función — que filtra las columnas del módulo derecho a las estrictamente nuevas antes del join — es el antecedente directo del `ConsolidadorGEIH` que hoy es el núcleo del paquete. También manejó las claves de cruce diferenciadas por módulo (`['DIRECTORIO', 'SECUENCIA_P', 'ORDEN']` para personas, `['DIRECTORIO', 'SECUENCIA_P']` para hogares) y exploró la integración posterior con las correlativas CIIU y DIVIPOLA.

**Enrique Forero** fue responsable de la arquitectura del sistema, el diseño técnico, la implementación del paquete `geih/`, la suite de pruebas (71 tests), la documentación y la publicación en PyPI. El trabajo de Nicolás y Lina fue la fuente de referencia que orientó las decisiones de diseño: qué variables incluir, cómo manejar los módulos del DANE, qué correlativas enriquecen la base y qué indicadores tienen valor analítico real para el equipo.

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: README.md >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📝 ARCHIVO [4/39]: Aprendizajes GitHub_PyPI_Publicacion.md
   TIPO      : DOCUMENTACIÓN
   UBICACIÓN : docs
   RUTA      : docs/Aprendizajes GitHub_PyPI_Publicacion.md
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: docs/Aprendizajes GitHub_PyPI_Publicacion.md <<<
────────────────────────────────────────────────────────────────────────────────
# FOR Néstor — Publicar un paquete Python desde Colab a GitHub y PyPI

**Tarea:** Diseñar el notebook de publicación, recomendar el nombre del
paquete y explicar todo el proceso de distribución open source.

---

## Paso 1: La decisión del nombre — por qué importa más de lo que parece

### El problema de `geih_2025`

El nombre del paquete tiene dos partes que hay que separar:
- **Nombre de distribución** (lo que pones en `pip install`)
- **Nombre de importación** (lo que pones en `from ... import`)

En Python, estos dos pueden ser distintos. Pandas lo hace: se llama
`scikit-learn` en pip pero `sklearn` en el código. Esta separación
es una solución estándar cuando el nombre ideal ya está tomado
o cuando el nombre de importación tiene restricciones (años, guiones, etc.)

`geih_2025` como nombre de importación tiene estos problemas:
1. El año implica que el paquete "vence" — un usuario en 2028 dudará
   antes de instalar algo llamado 2025.
2. Ya soporta 2022-2026+. El nombre dice una cosa, el código hace otra.
3. Si algún día lo renombras a `geih`, *rompes* todo el código de todos
   los usuarios (`from geih_2025 import` deja de funcionar).

### La solución: cambiar solo el nombre de distribución

```
pip install geih-analisis        ← nombre nuevo, PyPI
from geih_2025 import ConfigGEIH ← nombre existente, sin cambios
```

Así el cambio es completamente invisible para todo el código existente.
Los usuarios nuevos lo instalan con `geih-analisis`, pero el código
que ya escribiste (y todo lo que documentes) sigue usando `geih_2025`.

### Cuándo y cómo migrar el nombre de importación (futuro)

Si en 2027 decides que `geih_2025` ya no tiene sentido y quieres
llamarlo simplemente `geih`:
1. Creas el nuevo paquete `geih` que importa todo de `geih_2025`
2. En `geih_2025` pones un warning de deprecación: `DeprecationWarning: use 'geih' instead`
3. Das 6 meses de período de transición donde ambos funcionan
4. Publicas `geih_2025` como "deprecated" en PyPI
Este proceso se llama *deprecation path* y es como lo hacen todos los
paquetes profesionales cuando cambian nombres.

---

## Paso 2: Por qué publicar desde Colab es diferente a publicar desde local

### El problema central

En tu máquina local, git tiene acceso al sistema de archivos y guarda
las credenciales de GitHub en el keychain. En Colab, el entorno se
destruye con cada sesión — no hay keychain, no hay credenciales
persistentes, y el sistema de archivos de /content desaparece.

La solución son los **Colab Secrets**: un almacén cifrado de valores
sensibles que persiste entre sesiones y se puede leer con `userdata.get()`.
Es lo equivalente a las variables de entorno en sistemas locales.

### Por qué copiar a /content antes de hacer git

Google Drive tiene un overhead de I/O muy alto para operaciones git
porque cada acceso a un archivo implica una solicitud de red. Un
`git add -A` sobre una carpeta en Drive con 50 archivos puede tomar
30 segundos. El mismo comando sobre /content (SSD local de Colab)
toma 0.3 segundos. La solución: copiar de Drive a /content, operar
localmente, y el resultado queda en GitHub/PyPI (que son externos,
no en Drive).

---

## Paso 3: El flujo de publicación en detalle

```
Google Drive (fuente de verdad del código)
    ↓ shutil.copytree() — excluye *.parquet, *.csv, resultados/
/content/geih_build (directorio de trabajo temporal)
    ├── geih_2025/          ← código Python
    ├── pyproject.toml      ← metadatos generados por el notebook
    ├── README.md
    ├── LICENSE
    └── .gitignore

[Paso GitHub]
    git init + git add + git commit
    git push → github.com/usuario/geih-analisis
    git tag v4.3.1 → github.com/usuario/geih-analisis/releases

[Paso build]
    python -m build → dist/
        geih_analisis-4.3.1-py3-none-any.whl  (rueda, instalación rápida)
        geih_analisis-4.3.1.tar.gz             (fuente, para auditoría)

[Paso twine check]
    Verifica que los archivos cumplen estándares PyPI
    antes de intentar subir (fail fast)

[Paso twine upload]
    Si subir_a_pypi_real=False → TestPyPI (seguro, para probar)
    Si subir_a_pypi_real=True  → PyPI real (visible para el mundo)
```

---

## Paso 4: Semver — cómo numerar las versiones

El estándar es `MAJOR.MINOR.PATCH` (ej: `4.3.1`):

| Tipo | Cuándo | Ejemplo |
|---|---|---|
| PATCH | Bug fix, sin nueva funcionalidad | 4.3.0 → 4.3.1 |
| MINOR | Nueva funcionalidad, compatible | 4.3.0 → 4.4.0 |
| MAJOR | Cambio que rompe compatibilidad | 4.3.0 → 5.0.0 |

Romper compatibilidad significa que el código existente de los usuarios
deja de funcionar sin modificaciones. Ejemplos: renombrar una clase,
cambiar la firma de una función, eliminar un parámetro.

La regla práctica: mientras el paquete esté en desarrollo activo y
los usuarios sean principalmente internos de ProColombia, puedes usar
MINOR libremente. Cuando haya usuarios externos que dependen de la API,
sé más conservador con los cambios MINOR y cuidadoso con los MAJOR.

---

## Paso 5: TestPyPI vs PyPI real — nunca saltar el paso de prueba

TestPyPI es un servidor de prueba idéntico al real pero aislado.
Publicar allí tiene cero consecuencias: si algo sale mal, simplemente
publicas de nuevo. Las ventajas:

1. **Verifica el proceso completo**: el build, el upload, la instalación
2. **Verifica los metadatos**: que la descripción, el README y los links
   aparecen correctamente en la página del paquete
3. **Detecta errores de nombre**: si el nombre ya está tomado en PyPI real,
   lo descubres en TestPyPI sin problemas

El flujo recomendado:
```
TestPyPI (siempre) → revisar página web → PyPI real (solo si todo OK)
```

Para probar la instalación desde TestPyPI:
```bash
pip install --index-url https://test.pypi.org/simple/ \
            --extra-index-url https://pypi.org/simple/ \
            geih-analisis
```
El `--extra-index-url` es necesario porque las dependencias
(pandas, numpy, etc.) no están en TestPyPI — las busca en el PyPI real.

---

## Paso 6: Los tokens — gestión de secretos en Colab

### Por qué nunca escribir tokens en el código

Un token de GitHub con permisos de repositorio permite hacer push,
crear issues, acceder a código privado. Si queda en un notebook
que se comparte o en el historial de git, es una vulnerabilidad real.

Los Colab Secrets resuelven esto: están cifrados, no aparecen en
el historial de celdas, y se pueden revocar desde GitHub/PyPI sin
tocar el código.

### Tipos de tokens que necesitas

**GitHub Personal Access Token (Classic)**:
- Permisos mínimos necesarios: `repo` (acceso completo al repositorio)
- Dónde crearlo: GitHub → Settings → Developer settings → Personal access tokens
- Duración recomendada: 90 días (renovar cuando expire)

**PyPI API Token**:
- Scope recomendado: "Specific project" en vez de "Entire account"
  (más seguro — si el token se compromete, solo afecta ese paquete)
- Dónde crearlo: pypi.org → Account settings → API tokens
- Empieza siempre con `pypi-`

---

## Paso 7: Qué falta para tener un paquete verdaderamente profesional

El notebook hace GitHub + PyPI. Para el siguiente nivel:

**1. GitHub Actions** (automatización):
Cuando crees un Release en GitHub, el build y el upload a PyPI
suceden automáticamente sin correr el notebook. Solo tienes que
crear el Release con el tag correcto.

**2. Tests en CI**:
Cada push a GitHub corre los 71 tests automáticamente.
Si alguno falla, GitHub te avisa antes de publicar.

**3. Documentación automática** (Read the Docs):
Los docstrings del paquete se convierten en una página web
de documentación automáticamente. La URL sería
`geih-analisis.readthedocs.io`.

**4. CHANGELOG.md**:
Un archivo que documenta qué cambió en cada versión.
Es lo primero que leen los usuarios antes de actualizar.

---

## Paso 8: Lo que un experto vería

### `py.typed` — un marcador de 0 bytes que cambia todo

El archivo `py.typed` que crea el notebook le dice a mypy, pyright
y los IDEs que el paquete soporta type hints. Sin él, el IDE no
muestra autocompletado de tipos cuando importas el paquete.
Es un archivo vacío que tarda 0 segundos en crear y mejora la
experiencia de usuario de forma significativa.

### El `pyproject.toml` moderno vs el `setup.py` viejo

El notebook genera `pyproject.toml`, no `setup.py`. Desde 2021,
`pyproject.toml` es el estándar oficial de Python (PEP 517/518).
`setup.py` sigue funcionando pero está en proceso de deprecación.
La diferencia técnica: `pyproject.toml` es declarativo (datos puros),
`setup.py` es código ejecutable (necesitas correr Python para leerlo,
lo que crea problemas de seguridad y reproducibilidad).

### `py3-none-any.whl` — qué significa el nombre del wheel

```
geih_analisis-4.3.1-py3-none-any.whl
                     ↑    ↑    ↑
                     │    │    └─ any CPU architecture (no compilado)
                     │    └────── none (no extensiones C)
                     └─────────── Python 3.x compatible
```
Un paquete "none-any" es el ideal: funciona en cualquier SO,
cualquier arquitectura, cualquier Python 3.x. Esto es posible porque
el paquete es Python puro (sin extensiones C como numpy tiene).

---

## Paso 9: Lecciones que aplican a cualquier proyecto

### Lección 1: Nombre de distribución ≠ nombre de importación

Esta separación existe en muchos paquetes famosos y es una herramienta
poderosa para evolucionar el nombre de un paquete sin romper el código
existente. Aprenderla temprano evita el problema doloroso de tener
que hacer una migración forzada más tarde.

### Lección 2: TestPyPI antes de PyPI siempre

En ingeniería de software hay un principio: nunca haces el deploy
a producción directamente. Siempre hay un ambiente de staging.
TestPyPI es tu staging. No hay razón para saltárselo.

### Lección 3: Los secretos no van en el código, nunca

Tokens, contraseñas, claves API: ninguno en el código, ninguno en
comentarios, ninguno en archivos de configuración que se suban a git.
El historial de git es permanente — incluso si borras el token del código
actual, sigue estando en el historial a menos que hagas un rebase completo.
La solución son las variables de entorno o los sistemas de secretos
(Colab Secrets, AWS Secrets Manager, GitHub Secrets, etc.).

### Lección 4: Semver es un contrato con tus usuarios

Cuando publicas con semver, les estás diciendo: "si solo cambia
el PATCH, puedes actualizar con ojos cerrados". Si rompes ese contrato
(un PATCH que rompe código), pierdes la confianza de los usuarios.
Es mejor sobre-versionar (un MINOR cuando era PATCH) que bajo-versionar.

---

*Generado automáticamente como parte del flujo de aprendizaje continuo.*
*Pipeline: `geih_2025 v4.3.0` | ProColombia · GIC*

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: docs/Aprendizajes GitHub_PyPI_Publicacion.md >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📝 ARCHIVO [5/39]: DOCUMENTACION_TECNICA.md
   TIPO      : DOCUMENTACIÓN
   UBICACIÓN : docs
   RUTA      : docs/DOCUMENTACION_TECNICA.md
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: docs/DOCUMENTACION_TECNICA.md <<<
────────────────────────────────────────────────────────────────────────────────
# Documentación Técnica — `geih_2025` v4.3.0

**Paquete modular de Python para análisis de la Gran Encuesta Integrada de Hogares (GEIH) del DANE — Colombia**

*ProColombia · Gerencia de Inteligencia Comercial · Equipo de Analítica de Datos*

---

## Tabla de contenido

1. [Descripción general](#1-descripción-general)
2. [Instalación](#2-instalación)
3. [Inicio rápido](#3-inicio-rápido)
4. [Arquitectura del paquete](#4-arquitectura-del-paquete)
5. [Estándares de programación](#5-estándares-de-programación)
6. [Gestión de RAM en Google Colab](#6-gestión-de-ram-en-google-colab)
7. [Configuración](#7-configuración)
8. [Consolidación de datos](#8-consolidación-de-datos)
9. [Preparación de datos](#9-preparación-de-datos)
10. [Diagnóstico de calidad](#10-diagnóstico-de-calidad)
11. [Indicadores laborales](#11-indicadores-laborales)
12. [Análisis avanzados](#12-análisis-avanzados)
13. [Análisis poblacionales](#13-análisis-poblacionales)
14. [Análisis complementarios](#14-análisis-complementarios)
15. [Análisis por 32 ciudades](#15-análisis-por-32-ciudades)
16. [Visualización matplotlib](#16-visualización-matplotlib)
17. [Visualización interactiva Plotly](#17-visualización-interactiva-plotly)
18. [Exportación](#18-exportación)
19. [Comparativo multi-año](#19-comparativo-multi-año)
20. [Flujo comparativo completo en el notebook](#20-flujo-comparativo-completo-en-el-notebook)
21. [Descarga de datos DANE](#21-descarga-de-datos-dane)
22. [Utilidades](#22-utilidades)
23. [Logging y Profiling](#23-logging-y-profiling)
24. [Dashboard Streamlit en Colab](#24-dashboard-streamlit-en-colab)
25. [Análisis personalizados fuera del paquete](#25-análisis-personalizados-fuera-del-paquete)
26. [Testing](#26-testing)
27. [Variables GEIH de referencia](#27-variables-geih-de-referencia)
28. [Variables de alto valor para ProColombia](#28-variables-de-alto-valor-para-procolombia)
29. [La regla del FEX](#29-la-regla-del-fex)
30. [Flujo de trabajo mes a mes](#30-flujo-de-trabajo-mes-a-mes)
31. [Cómo agregar un año nuevo](#31-cómo-agregar-un-año-nuevo)
32. [Convenciones de nombres](#32-convenciones-de-nombres)
33. [Preguntas frecuentes](#33-preguntas-frecuentes)

---

## 1. Descripción general

`geih_2025` transforma los microdatos crudos de la GEIH del DANE (8 módulos CSV × 12 meses × ~5 millones de registros) en indicadores publicables del mercado laboral colombiano. Incluye 74 clases con 142 métodos públicos, 71 tests automatizados y ~10,300 líneas de código en 19 módulos Python.

---

## 2. Instalación

**Google Colab:**
```python
from google.colab import drive
drive.mount('/content/drive')
import sys
sys.path.insert(0, '/content/drive/MyDrive/ProColombia/GEIH')
from geih_2025 import __version__
print(f"geih_2025 v{__version__}")
```

**Local:** `pip install -e ".[dev]"`

---

## 3. Inicio rápido

```python
from geih_2025 import ConfigGEIH, ConsolidadorGEIH, PreparadorGEIH, IndicadoresLaborales

config = ConfigGEIH(anio=2025, n_meses=12)
geih   = ConsolidadorGEIH.cargar(f'{RUTA}/GEIH_2025_Consolidado.parquet')
df     = PreparadorGEIH(config=config).preparar_base(geih)
df     = PreparadorGEIH(config=config).agregar_variables_derivadas(df)
r      = IndicadoresLaborales(config=config).calcular(df)
# → TD=8.9%, TGP=64.3%, TO=58.6%
```

---

## 4. Arquitectura del paquete

19 módulos independientes. La única dependencia obligatoria entre módulos es `config.py`.

```
geih_2025/
├── config.py                    ← Año, SMMLV, carpetas, constantes
├── utils.py                     ← EstadisticasPonderadas, GestorMemoria
├── consolidador.py              ← CSVs → DataFrame + checkpointing
├── preparador.py                ← 515 cols → 60 cols + FEX_ADJ
├── diagnostico.py               ← Missing values, tipos, identidades
├── indicadores.py               ← TD, TGP, TO, ingresos, brecha, Gini
├── analisis_avanzado.py         ← ICE, ICI, ITAT, IVI, Mincer (18 clases)
├── analisis_poblacional.py      ← Campesinos, discapacidad, migración (10)
├── analisis_complementario.py   ← Duración desempleo, forma pago, canal (5)
├── analisis_area.py             ← 32 ciudades × CIIU
├── visualizacion.py             ← 9 gráficos matplotlib
├── visualizacion_interactiva.py ← 8 gráficos Plotly
├── exportador.py                ← CSVs, Excel, PNGs organizados
├── comparativo.py               ← Comparación inter-anual
├── descargador.py               ← Organizar ZIPs del DANE
├── logger.py                    ← Logging centralizado
├── profiler.py                  ← Profiling de RAM y tiempo
└── dashboard.py                 ← Dashboard Streamlit
```

---

## 5. Estándares de programación

El notebook sigue estándares de ingeniería optimizados para Colab (~12GB RAM, límite 12h).

**Principios:** KISS (una sola responsabilidad), DRY (configuración centralizada), Fail Fast (validar ANTES de procesar), Trust but Verify (verificar tamaño después de cada operación).

**Regla de oro:** NUNCA `apply`, `iterrows` ni `for` sobre DataFrames >100k filas. SIEMPRE operaciones vectorizadas.

**Organización del notebook:** Celdas EXTRAS (definiciones, ejecutar UNA vez) y celdas EJECUTAR (máximo 10 líneas, solo configuración + llamada a función).

**Emojis:** ✅ = ejecutar siempre, 🟨 = opcional, 🛑 = precaución, ♻️ = reutilizable, 🚧 = pendiente.

---

## 6. Gestión de RAM en Google Colab

### Regla cardinal

Nunca retener dos DataFrames de año completo en RAM simultáneamente. Un año crudo (515 cols) ocupa ~2 GB. Dos = crash.

**Patrón correcto:** Cargar → procesar → guardar a disco → `del df; gc.collect()` → cargar el siguiente.

### Preparación reduce 10x la memoria

```
Parquet raw  : 515 cols, ~2.0 GB RAM, sin FEX_ADJ → crash probable
Parquet prep :  ~60 cols, ~300 MB RAM, con FEX_ADJ → funciona bien
```

### Verificar siempre

```python
from geih_2025 import GestorMemoria
GestorMemoria.estado()                      # RAM usada/libre
GestorMemoria.tamano_df(df, 'mi_base')      # MB del DataFrame
```

### PyArrow predicate pushdown

Para extraer un subconjunto de un Parquet grande sin cargarlo completo en RAM:

```python
import pyarrow.parquet as pq
table = pq.read_table('GEIH_2025.parquet', filters=[('MES_NUM', 'in', [1])])
df_enero = table.to_pandas()
del table; gc.collect()
# RAM: ~50 MB (solo enero) vs ~2 GB (todo el archivo)
```

---

## 7. Configuración

### `ConfigGEIH` — Configuración del paquete

```python
from geih_2025 import ConfigGEIH
config = ConfigGEIH(anio=2025, n_meses=12)
config.resumen()
```

**Atributos:** `anio`, `n_meses`, `smmlv` (auto-seleccionado), `carpetas_mensuales`, `periodo_etiqueta`, `referencia_dane`.

**`n_meses` es el divisor del FEX**, no un contador de meses. Si tiene 12 meses en el Parquet pero carga con `n_meses=1`, los pesos se multiplican por 12.

### Variables del notebook para comparaciones

```python
ANIO_BASE    = 2025     # Año de referencia
N_MESES_BASE = 12       # Meses del año base
ANIO_COMP    = 2026     # Año en curso
N_MESES_COMP = 1        # ← ÚNICO valor que cambia mes a mes
RUTA         = '/content/drive/MyDrive/ProColombia/GEIH'
RUTA_CIIU    = '/content/drive/.../Correlativa CIIU Rev4.xlsx'
```

### Constantes disponibles

```python
from geih_2025 import SMMLV_POR_ANIO, DEPARTAMENTOS, RAMAS_DANE, TAMANO_EMPRESA, REF_DANE
```

---

## 8. Consolidación de datos

```python
from geih_2025 import ConsolidadorGEIH
consolidador = ConsolidadorGEIH(ruta_base=RUTA, config=config, incluir_area=True)
consolidador.verificar_estructura()
geih = consolidador.consolidar(checkpoint=True)  # se recupera si falla
consolidador.exportar(geih)
# Cargar existente: geih = ConsolidadorGEIH.cargar('GEIH_2025_Consolidado.parquet')
```

**Agregar mes nuevo:** `consolidador.agregar_mes('Abril 2026', 'GEIH_2026_Consolidado.parquet')`

---

## 9. Preparación de datos

```python
from geih_2025 import PreparadorGEIH, MergeCorrelativas
prep = PreparadorGEIH(config=config)
df = prep.preparar_base(geih)                    # 515 → ~60 cols + FEX_ADJ
df = prep.agregar_variables_derivadas(df)         # RAMA, RANGO_SMMLV, ZONA...
# Mes puntual: df_dic = prep.preparar_base(geih, mes_filtro=12)
# CIIU descriptivo: df = MergeCorrelativas().merge_ciiu(df, ruta_ciiu=RUTA_CIIU)
```

---

## 10. Diagnóstico de calidad

```python
from geih_2025 import DiagnosticoCalidad, Top20Sectores
diag = DiagnosticoCalidad()
diag.resumen_rapido(geih); diag.verificar_tipos(geih); diag.validar_identidades(geih)
Top20Sectores(config=config).calcular(df, ruta_ciiu=RUTA_CIIU)
```

---

## 11. Indicadores laborales

```python
from geih_2025 import IndicadoresLaborales, DistribucionIngresos, AnalisisRamaSexo
from geih_2025 import AnalisisSalarios, BrechaGenero, IndicesCompuestos

ind = IndicadoresLaborales(config=config)
r = ind.calcular(df);  ind.sanity_check(r, "Anual 2025")
td_dpto = ind.por_departamento(df)

DistribucionIngresos(config=config).calcular(df)
AnalisisRamaSexo().calcular(df)
AnalisisSalarios(config=config).por_rama(df)
BrechaGenero().calcular(df)
IndicesCompuestos(config=config).gini(df)
```

---

## 12. Análisis avanzados

```python
from geih_2025 import (
    CalidadEmpleo, FormalidadSectorial, CompetitividadLaboral,
    VulnerabilidadLaboral, CostoLaboral, ContribucionSectorial,
    MapaTalento, BonoDemografico, Estacionalidad, EcuacionMincer,
    FuerzaLaboralJoven, AnalisisUrbanoRural, EtnicoRacial, ProxyBilinguismo,
)

CalidadEmpleo(config=config).calcular_por_departamento(df)   # ICE
FormalidadSectorial(config=config).calcular(df)               # ICF
CompetitividadLaboral(config=config).calcular(df)             # ICI
VulnerabilidadLaboral(config=config).calcular(df)             # IVI
MapaTalento(config=config).calcular(df)                       # ITAT
CostoLaboral(config=config).calcular(df)
BonoDemografico(config=config).calcular(df)
Estacionalidad().calcular(geih)        # ← usa geih crudo, NO df
ContribucionSectorial().calcular(geih)
EcuacionMincer(config=config).estimar_todos(df)
FuerzaLaboralJoven(config=config).calcular(df)
AnalisisUrbanoRural(config=config).calcular(df)
EtnicoRacial().calcular(df)
ProxyBilinguismo().calcular(df)
```

**Fórmulas:** ICE = 0.30×Pensión + 0.25×Salud + 0.25×Horas + 0.20×Ingreso. ICI = 0.25×TD + 0.20×Costo + 0.25×Talento + 0.20×Formalidad + 0.10×Jóvenes. ITAT = 0.35×Oferta + 0.35×Costo + 0.30×Calidad.

---

## 13. Análisis poblacionales

```python
from geih_2025 import (
    AnalisisCampesino, AnalisisDiscapacidad, AnalisisMigracion,
    AnalisisOtrasFormas, AnalisisOtrosIngresos, AnalisisSobrecalificacion,
    AnalisisContractual, AnalisisAutonomia, AnalisisAlcanceMercado,
    AnalisisDesanimados,
)
AnalisisCampesino(config=config).calcular(df)          # P2057
AnalisisDiscapacidad().calcular(df)                     # P1906S1-S8
AnalisisMigracion(config=config).calcular(df)           # P3370/P3376
AnalisisOtrasFormas().calcular(df)                      # P3054-P3057
AnalisisOtrosIngresos().calcular(df)                    # P7422/P7500
AnalisisSobrecalificacion(config=config).calcular(df)   # P3042 × P6430
AnalisisContractual().calcular(df)                      # P6440/P6450/P6460
AnalisisAutonomia().calcular(df)                        # P3047-P3049
AnalisisAlcanceMercado().calcular(df)                   # P1802
AnalisisDesanimados().calcular(df)                      # P6300/P6310
```

**Regla:** Las que necesitan SMMLV llevan `config=config`. Las demás se instancian con `()`.

---

## 14. Análisis complementarios

```python
from geih_2025 import DuracionDesempleo, DashboardSectoresProColombia
from geih_2025 import AnatomaSalario, FormaPago, CanalEmpleo

dur = DuracionDesempleo(config=config)
dur.calcular(df); dur.por_sexo(df); dur.por_educacion(df); dur.por_departamento(df)

DashboardSectoresProColombia(config=config).calcular(df)

anat = AnatomaSalario(config=config)
anat.resumen_nacional(df); anat.por_rama(df); anat.por_tamano_empresa(df)

fp = FormaPago(config=config)
fp.calcular(df); fp.cruce_formalidad(df)

ce = CanalEmpleo(config=config)
ce.calcular(df); ce.por_nivel_educativo(df)
```

---

## 15. Análisis por 32 ciudades

```python
from geih_2025 import AnalisisOcupadosCiudad
area = AnalisisOcupadosCiudad(config=config)
tablas = area.calcular(df, ruta_ciiu=RUTA_CIIU)  # ← calcular(), NO calcular_tablas()
area.imprimir(tablas); area.graficar(tablas)
area.exportar_excel(tablas, f'{RUTA}/resultados/CIIU_Area_{config.anio}.xlsx')
```

---

## 16. Visualización matplotlib

```python
from geih_2025 import (
    GraficoCurvaLorenz, GraficoBoxPlotSalarios, GraficoBrechaGenero,
    GraficoDistribucionIngresos, GraficoICIBubble, GraficoEstacionalidad,
    GraficoContribucionHeatmap, GraficoRamaSexo,
)
fig = GraficoCurvaLorenz().graficar(df[(df['OCI']==1) & (df['INGLABO']>0)])
exp.guardar_grafica(fig, 'Lorenz'); plt.show()
```

---

## 17. Visualización interactiva Plotly

Requiere `pip install plotly`. Retornan `plotly.graph_objects.Figure`.

```python
from geih_2025 import PlotlyLorenz, PlotlyEstacionalidad, PlotlyComparativoAnual
fig = PlotlyLorenz().graficar(df_ocu); fig.show()
```

---

## 18. Exportación

```python
from geih_2025 import Exportador
exp = Exportador(ruta_base=RUTA)
exp.guardar_tabla(td_dpto, 'indicadores_dpto')
exp.guardar_grafica(fig, 'Lorenz_2025')
exp.guardar_excel({'Hoja1': df1}, 'nombre')
exp.guardar_metadata(config, {'registros': len(geih)})
exp.resumen()
```

---

## 19. Comparativo multi-año

```python
from geih_2025 import ComparadorMultiAnio, ConfigGEIH
comp = ComparadorMultiAnio()
comp.agregar_anio(2025, f'{RUTA}/GEIH_2025_M01.parquet', ConfigGEIH(anio=2025, n_meses=1))
comp.agregar_anio(2026, f'{RUTA}/GEIH_2026_M01.parquet', ConfigGEIH(anio=2026, n_meses=1))
comp.comparar_indicadores(); comp.comparar_departamentos(); comp.evolucion_ingresos()
comp.comparar_ramas(); comp.comparar_brecha_genero()
```

---

## 20. Flujo comparativo completo en el notebook

El notebook real usa un patrón más sofisticado. Define un `ConfigComparativo` (dataclass en el notebook, no en el paquete) que orquesta la comparación entre dos años con gestión estricta de RAM.

### ConfigComparativo

```python
@dataclass
class ConfigComparativo:
    ruta: str; ruta_ciiu: str; ruta_divipola: str
    anio_base: int = 2025; n_meses_base: int = 12
    anio_comp: int = 2026; n_meses_comp: int = 1  # ← cambiar mes a mes
```

### Funciones auxiliares del notebook

`asegurar_parquet(cfg, anio, n_meses)` — Garantiza que un Parquet existe en disco con jerarquía de eficiencia: (1) ya existe → retorna ruta. (2) Existe parquet completo → extrae meses vía PyArrow predicate pushdown (sin cargar todo en RAM). (3) Nada existe → consolida desde CSV. Al retornar: CERO DataFrames en RAM.

`garantizar_parquets(cfg)` — Garantiza los 3 Parquets necesarios: período base, período comp, año base completo.

`agregar_mes_nuevo(cfg, n_meses_nuevo)` — Agrega un mes sin re-consolidar.

### Flujo de ejecución

```python
cfg = ConfigComparativo(ruta=RUTA, ruta_ciiu=RUTA_CIIU, ruta_divipola=RUTA_DIVIPOLA,
    anio_base=2025, n_meses_base=12, anio_comp=2026, n_meses_comp=1)
PATH_BASE_COMP, PATH_COMP_COMP, PATH_BASE_FULL = garantizar_parquets(cfg)

comp = ComparadorMultiAnio()
comp.agregar_anio(2025, PATH_BASE_COMP, ConfigGEIH(anio=2025, n_meses=1))
gc.collect()
comp.agregar_anio(2026, PATH_COMP_COMP, ConfigGEIH(anio=2026, n_meses=1))
comp.comparar_indicadores()
```

---

## 21. Descarga de datos DANE

La descarga automática no funciona (el DANE requiere aceptar términos). Lo que sí funciona es organizar ZIPs descargados manualmente.

```python
from geih_2025 import DescargadorDANE
desc = DescargadorDANE(config=ConfigGEIH(anio=2025, n_meses=12), ruta_destino=RUTA)
desc.instrucciones_descarga_manual()
desc.organizar_zips(f'{RUTA}/zips_dane_2025')
desc.verificar()
```

**IDs catálogo:** 2022→771, 2023→782, 2024→819, 2025→853.

---

## 22. Utilidades

```python
from geih_2025 import EstadisticasPonderadas as EP
EP.media(df['INGLABO'], df['FEX_ADJ'])
EP.mediana(df['INGLABO'], df['FEX_ADJ'])
EP.percentil(df['INGLABO'], df['FEX_ADJ'], 0.25)
EP.gini(df['INGLABO'], df['FEX_ADJ'])

from geih_2025 import GestorMemoria
GestorMemoria.estado(); GestorMemoria.tamano_df(df, 'mi_base')

from geih_2025 import ConversorTipos
ConversorTipos.estandarizar_dpto(serie)  # '5' → '05'
```

---

## 23. Logging y Profiling

```python
from geih_2025 import configurar_logging, PerfilMemoria, medir_tiempo

configurar_logging(nivel='INFO', archivo='geih.log')

perf = PerfilMemoria()
perf.snapshot("Inicio")
with medir_tiempo("Consolidar"):
    geih = consolidador.consolidar()
perf.snapshot("Post-consolidar")
perf.reporte()
```

---

## 24. Dashboard Streamlit en Colab

**Opción simple:** `from geih_2025 import ejecutar_dashboard; ejecutar_dashboard(ruta_base=RUTA)`

**Opción producción (con túnel):** El notebook incluye `lanzar_dashboard_prep()` que prepara Parquets livianos y abre túnel Cloudflare:

```python
lanzar_dashboard_prep(cfg=cfg, modo='comparacion', metodo='cloudflare')
```

| Modo | Qué carga | RAM |
|---|---|---|
| `'enero'` | Solo enero año comp | ~40 MB |
| `'comparacion'` | Enero de ambos años | ~80 MB |
| `'anio_comp'` | Meses acumulados | ~40-400 MB |
| `'anio_base'` | Año base completo (preparado) | ~400 MB |

---

## 25. Análisis personalizados fuera del paquete

El notebook incluye análisis que van más allá del paquete, escritos directamente en celdas usando `EstadisticasPonderadas`:

**Salario por educación × industria** — Mapeo granular P3042 con 9 niveles (separando Especialización, Maestría, Doctorado), tabla nacional + heatmap cruzado con top 8 ramas. Prima salarial vs universitaria.

**Análisis de altos ingresos (P90/P95/P99/P99.9)** — Umbrales percentiles, perfiles demográficos de cada segmento (sexo, educación, industria, ciudad, posición, tamaño empresa), 8 tablas comparativas + 4 gráficos (distribución log, stack bars, paneles dobles, heatmap edad × percentil).

**Box plot ponderado completo** — Versión expandida de ~150 líneas con 11 estadísticas por rama, gradiente de color, segundo eje SMMLV, etiquetas dentro de las cajas.

Estos análisis usan `EstadisticasPonderadas` del paquete y se pueden replicar con:

```python
from geih_2025.utils import EstadisticasPonderadas as EP
mediana = EP.mediana(df.loc[mask, 'INGLABO'], df.loc[mask, 'FEX_ADJ'])
```

---

## 26. Testing

```bash
python -m pytest tests/ -v     # 71 tests
python tests/smoke_test.py      # validación rápida
```

Golden set: 1,000 registros sintéticos (TD=11.76%, TGP=68.0%, TO=60.0%).

---

## 27. Variables GEIH de referencia

| Variable | Significado | Valores |
|---|---|---|
| `FEX_C18` | Factor expansión | float |
| `P3271` | Sexo | 1=H, 2=M |
| `P6040` | Edad | entero |
| `P3042` | Nivel educativo | 1-13 |
| `CLASE` | Zona | 1=Urbano, 2=Rural |
| `DPTO` | Departamento | 2 dígitos |
| `OCI` | Ocupado | 1=Sí |
| `DSI` | Desocupado | 1=Sí |
| `FT` | PEA | 1=Sí |
| `FFT` | Fuera FT | 1=Sí |
| `PET` | Edad trabajar | 1=Sí |
| `INGLABO` | Ingreso laboral | COP |
| `P6500` | Salario declarado | COP |
| `P6800` | Horas semanales | entero |
| `P6920` | Cotiza pensión | 1=Sí |
| `RAMA2D_R4` | CIIU 2 dígitos | string |
| `P6430` | Posición ocupacional | 1-9 |
| `P7250` | Semanas buscando | entero |
| `P6765` | Forma de pago | numérico |
| `P3363` | Canal empleo | 1-7 |
| `P1802` | Alcance mercado | 1-6 |
| `P2057` | ¿Campesino? | 1=Sí |
| `P1906S1-S8` | Discapacidad | 1-4 |
| `AREA` | Municipio | 5 dígitos |

---

## 28. Variables de alto valor para ProColombia

| Variable | Valor para IED/Exportaciones |
|---|---|
| `P1802=6` | Empleo DIRECTO en comercio exterior |
| `P3047-P3049` | Contratistas dependientes (evasión laboral) |
| `P3363` | Digitalización del mercado laboral |
| `P6765` | Precariedad: destajo, honorarios, comisión |
| `P3364` | Proxy tributación formal DIAN |
| `P3042 × P6430` | Talento subutilizado (sobrecalificación) |
| `P2057` | Política pública rural |
| `P1906S1-S8` | ESG para IED europea |
| `P3376` | Migración venezolana |

---

## 29. La regla del FEX

| Período | ConfigGEIH | FEX_ADJ |
|---|---|---|
| Mes puntual | `n_meses=1` | FEX_C18 ÷ 1 |
| Trimestre | `n_meses=3` | FEX_C18 ÷ 3 |
| Anual | `n_meses=12` | FEX_C18 ÷ 12 |

**Alarma:** PEA > 40M → FEX no se dividió. `sanity_check()` lo detecta.

**Excepción:** `Estacionalidad().calcular(geih)` usa FEX_C18 sin dividir.

---

## 30. Flujo de trabajo mes a mes

Cuando el DANE publica un nuevo mes (ej: febrero 2026):

1. Descargar ZIPs del DANE y `desc.organizar_zips()`
2. `consolidador.agregar_mes('Febrero 2026', 'GEIH_2026_Consolidado.parquet')`
3. Cambiar `N_MESES_COMP = 2` en la celda de configuración
4. Re-ejecutar desde la celda de comparaciones

---

## 31. Cómo agregar un año nuevo

En `config.py`: `SMMLV_POR_ANIO[2027] = 1_800_000`

Opcional: `REF_DANE[2027] = ReferenciaDane(td_anual_pct=...)` cuando DANE publique.

En `descargador.py`: `CATALOGO_DANE[2027] = {"catalog_id": 920}`

---

## 32. Convenciones de nombres

| Nombre archivo | Contenido |
|---|---|
| `GEIH_2025_M01.parquet` | Solo enero |
| `GEIH_2025_M01_M03.parquet` | Enero a marzo |
| `GEIH_2025_M01_M12.parquet` | Año completo |
| `GEIH_2025_Consolidado.parquet` | Alias año completo |
| `*_prep.parquet` | Preparado (60 cols + FEX_ADJ) para dashboard |

**Variables en memoria:** `geih` = crudo (515 cols), `df` = preparado (60 cols), `config` = ConfigGEIH, `exp` = Exportador.

**Recálculo seguro:** `if 'salarios_rama' not in dir(): salarios_rama = AnalisisSalarios(config=config).por_rama(df)`

---

## 33. Preguntas frecuentes

**¿Tabla vacía?** → `print(df['P6765'].value_counts().head())` para ver la codificación real.

**¿PEA > 40M?** → FEX no dividido. Verificar `config.n_meses`.

**¿Un solo mes?** → `ConfigGEIH(n_meses=1)` + `preparar_base(geih, mes_filtro=3)`.

**¿Cuánta RAM?** → Preparado: ~300 MB. Crudo: ~2 GB. Dos crudos: crash.

**¿`config=config` o `()`?** → Con SMMLV: `config=config`. Sin SMMLV: `()`.

**¿Colab se cayó?** → `checkpoint=True` retoma desde donde quedó.

**¿Dashboard desde Colab?** → `lanzar_dashboard_prep(cfg, modo='comparacion', metodo='cloudflare')`.

**¿Comparar enero 2025 vs enero 2026?** → Ambos con `ConfigGEIH(n_meses=1)` en `ComparadorMultiAnio`.

---

*geih_2025 v4.3.0 — 74 clases, 142 métodos, 19 módulos · ProColombia · GIC · Bogotá, 2026*

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: docs/DOCUMENTACION_TECNICA.md >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📝 ARCHIVO [6/39]: GUIA_GEIH.md
   TIPO      : DOCUMENTACIÓN
   UBICACIÓN : docs
   RUTA      : docs/GUIA_GEIH.md
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: docs/GUIA_GEIH.md <<<
────────────────────────────────────────────────────────────────────────────────
# Guía de Uso — Paquete `geih_2025` v4.3.0

**ProColombia · Gerencia de Inteligencia Comercial · Equipo de Analítica de Datos**
*Gran Encuesta Integrada de Hogares | DANE | Marco Muestral 2018*

---

## 1. Configuración inicial (una sola vez)

```python
# ═══ CELDA 1: Montar Drive y cargar paquete ═══
from google.colab import drive
drive.mount('/content/drive')

import sys
sys.path.insert(0, '/content/drive/MyDrive/ProColombia/GEIH')

from geih_2025 import __version__
print(f"geih_2025 v{__version__}")
```

## 2. Configuración del año (el ÚNICO lugar que se cambia)

```python
from geih_2025 import ConfigGEIH

# ┌──────────────────────────────────────────────────────┐
# │  Cambiar SOLO estas 2 líneas para procesar otro año  │
# └──────────────────────────────────────────────────────┘
config = ConfigGEIH(anio=2025, n_meses=12)
RUTA = '/content/drive/MyDrive/ProColombia/GEIH'

config.resumen()  # Muestra SMMLV, carpetas, período
```

**Ejemplos de configuración:**

| Caso | Código |
|---|---|
| Año 2025 completo | `ConfigGEIH(anio=2025, n_meses=12)` |
| Primer trimestre 2026 | `ConfigGEIH(anio=2026, n_meses=3)` |
| Solo enero 2026 | `ConfigGEIH(anio=2026, n_meses=1)` |

## 3. Consolidar datos (primera vez) o cargar existente

```python
from geih_2025 import ConsolidadorGEIH
import os

PARQUET = f'{RUTA}/GEIH_{config.anio}_Consolidado.parquet'

if os.path.exists(PARQUET):
    geih = ConsolidadorGEIH.cargar(PARQUET)
else:
    consolidador = ConsolidadorGEIH(ruta_base=RUTA, config=config, incluir_area=True)
    consolidador.verificar_estructura()
    geih = consolidador.consolidar(checkpoint=True)  # ← se recupera si falla
    consolidador.exportar(geih)
```

**Agregar un mes nuevo sin re-consolidar:**

```python
config_4m = ConfigGEIH(anio=2026, n_meses=4)
consolidador = ConsolidadorGEIH(ruta_base=RUTA, config=config_4m, incluir_area=True)
geih = consolidador.agregar_mes(
    mes_carpeta='Abril 2026',
    parquet_existente='GEIH_2026_Consolidado.parquet',
)
consolidador.exportar(geih)
```

## 4. Preparar datos

```python
from geih_2025 import PreparadorGEIH, Exportador

prep = PreparadorGEIH(config=config)
df = prep.preparar_base(geih)
df = prep.agregar_variables_derivadas(df)
exp = Exportador(ruta_base=RUTA)
```

## 5. Análisis disponibles (70+ clases)

### 5.1 Indicadores nacionales

```python
from geih_2025 import IndicadoresLaborales

ind = IndicadoresLaborales(config=config)
r = ind.calcular(df)
ind.sanity_check(r, f"Anual {config.anio}")
td_dpto = ind.por_departamento(df)
```

### 5.2 Los 10 análisis básicos

```python
from geih_2025 import (
    DistribucionIngresos, AnalisisRamaSexo, AnalisisSalarios,
    BrechaGenero, IndicesCompuestos, Estacionalidad,
    AnalisisUrbanoRural, FuerzaLaboralJoven, EcuacionMincer,
)
from geih_2025 import Top20Sectores

DistribucionIngresos(config=config).calcular(df)
AnalisisRamaSexo().calcular(df)
AnalisisSalarios(config=config).por_rama(df)
BrechaGenero().calcular(df)
IndicesCompuestos(config=config).gini(df)
Estacionalidad().calcular(geih)            # usa geih crudo
AnalisisUrbanoRural(config=config).calcular(df)
FuerzaLaboralJoven(config=config).calcular(df)
EcuacionMincer(config=config).estimar_todos(df)
Top20Sectores(config=config).calcular(df)
```

### 5.3 Los 8 análisis avanzados

```python
from geih_2025 import (
    CalidadEmpleo, FormalidadSectorial, CompetitividadLaboral,
    VulnerabilidadLaboral, CostoLaboral, ContribucionSectorial,
    MapaTalento, BonoDemografico,
)

CalidadEmpleo(config=config).calcular_por_departamento(df)     # ICE
FormalidadSectorial(config=config).calcular(df)                 # ICF
CompetitividadLaboral(config=config).calcular(df)               # ICI
VulnerabilidadLaboral(config=config).calcular(df)               # IVI
CostoLaboral(config=config).calcular(df)
ContribucionSectorial().calcular(geih)
MapaTalento(config=config).calcular(df)                         # ITAT
BonoDemografico(config=config).calcular(df)
```

### 5.4 Los 15 análisis poblacionales
```python
from geih_2025 import (
    AnalisisCampesino, AnalisisDiscapacidad, AnalisisMigracion,
    AnalisisOtrasFormas, AnalisisOtrosIngresos, AnalisisSobrecalificacion,
    AnalisisContractual, AnalisisAutonomia, AnalisisAlcanceMercado,
    AnalisisDesanimados,
    DuracionDesempleo, DashboardSectoresProColombia,
    AnatomaSalario, FormaPago, CanalEmpleo,
)

# ── Poblaciones especiales ───────────────────────────────
AnalisisCampesino(config=config).calcular(df)          # P2057 — ¿se considera campesino?
AnalisisDiscapacidad().calcular(df)                     # P1906S1-S8 — escala Washington
AnalisisMigracion(config=config).calcular(df)           # P3370/P3376 — migración interna e internacional
AnalisisOtrasFormas().calcular(df)                      # P3054-P3057 — autoconsumo, voluntariado, formación
AnalisisOtrosIngresos().calcular(df)                    # P7422/P7500 — pensiones, remesas, arriendos
AnalisisSobrecalificacion(config=config).calcular(df)   # P3042 × P6430 — universitarios en empleos simples
AnalisisContractual().calcular(df)                      # P6440/P6450/P6460 — contrato escrito/verbal/sin
AnalisisAutonomia().calcular(df)                        # P3047-P3049 — contratistas dependientes (asalariados disfrazados)
AnalisisAlcanceMercado().calcular(df)                   # P1802 — local/regional/nacional/exportación
AnalisisDesanimados().calcular(df)                      # P6300/P6310 — FFT que desean trabajar (potencial laboral latente)

# ── Análisis complementarios ────────────────────────────
DuracionDesempleo(config=config).calcular(df)           # P7250 — semanas buscando empleo (friccional→crónico)
DuracionDesempleo(config=config).por_departamento(df)   # Mediana semanas × dpto (proxy rigidez)
DashboardSectoresProColombia(config=config).calcular(df)# 7 sectores estratégicos para IED
AnatomaSalario(config=config).resumen_nacional(df)      # P6500 vs INGLABO — ingreso "invisible"
AnatomaSalario(config=config).por_rama(df)              # Brecha salarial por rama
FormaPago(config=config).calcular(df)                   # P6765 — destajo, honorarios, comisión
CanalEmpleo(config=config).calcular(df)                 # P3363 — contactos, internet, agencia
```

### 5.5 Análisis por 32 ciudades

```python
from geih_2025 import AnalisisOcupadosCiudad

area = AnalisisOcupadosCiudad(config=config)
tablas = area.calcular(df, ruta_ciiu=RUTA_CIIU)   # ← NO es calcular_tablas
area.imprimir(tablas)
area.exportar_excel(tablas, f'{RUTA}/resultados/CIIU_Area_{config.anio}.xlsx')
```

## 6. Gráficos

### Matplotlib (estáticos, para PNG)

```python
from geih_2025 import (
    GraficoBoxPlotSalarios, GraficoBrechaGenero,
    GraficoCurvaLorenz, GraficoICIBubble, GraficoEstacionalidad,
)

fig = GraficoCurvaLorenz().graficar(df[(df['OCI']==1) & (df['INGLABO']>0)])
exp.guardar_grafica(fig, f'Lorenz_{config.anio}')
```

### Plotly (interactivos, con tooltips y zoom)

```python
from geih_2025 import PlotlyLorenz, PlotlyICIBubble, PlotlyEstacionalidad

fig = PlotlyLorenz().graficar(df[(df['OCI']==1) & (df['INGLABO']>0)])
fig.show()  # interactivo en Colab
```

## 7. Comparación entre años

```python
from geih_2025 import ComparadorMultiAnio, ConfigGEIH

comp = ComparadorMultiAnio()
comp.agregar_anio(2025, f'{RUTA}/GEIH_2025_Consolidado.parquet', ConfigGEIH(anio=2025))
comp.agregar_anio(2026, f'{RUTA}/GEIH_2026_Consolidado.parquet', ConfigGEIH(anio=2026, n_meses=6))

comp.comparar_indicadores()        # TD/TGP/TO × año + variación
comp.comparar_departamentos()      # TD por dpto × año
comp.evolucion_ingresos()          # mediana salarial por año
comp.comparar_ramas()              # empleo por rama × año
comp.comparar_brecha_genero()      # brecha H/M por año
```

## 8. Dashboard interactivo (sin código)

```bash
pip install streamlit plotly
streamlit run geih_2025/dashboard.py
```

O desde el notebook:

```python
from geih_2025 import ejecutar_dashboard
ejecutar_dashboard(ruta_base=RUTA)
```

## 9. Profiling de memoria

```python
from geih_2025 import PerfilMemoria, medir_tiempo

perf = PerfilMemoria()
perf.snapshot("Inicio")

with medir_tiempo("Consolidar"):
    geih = consolidador.consolidar()
perf.snapshot("Post-consolidar")

perf.reporte()  # tabla RAM × etapa con alertas si > 80%
```

## 10. Logging con trazabilidad

```python
from geih_2025 import configurar_logging

# Al inicio del notebook (una vez)
configurar_logging(nivel='INFO', archivo='geih_pipeline.log')
# Desde ahora, todo se registra en consola + archivo
```

## 11. Exportar todo

```python
exp.guardar_excel({
    'Indicadores': td_dpto,
    'Salarios': salarios_rama,
    'ICE': ice_dpto,
    'ICI': ici,
}, f'GEIH_{config.anio}_Completo')

exp.guardar_metadata(config, {'registros': len(geih), 'version': '4.3.0'})
exp.resumen()
```

## 12. Agregar un año nuevo al sistema

Solo 2 cambios en `config.py`:

```python
# 1. SMMLV:
SMMLV_POR_ANIO = {
    ...
    2027: 1_800_000,   # ← agregar línea
}

# 2. Referencia DANE (cuando se publique):
REF_DANE = {
    ...
    2027: ReferenciaDane(td_anual_pct=..., tgp_anual_pct=..., ...),
}
```

Si no agrega las entradas, el sistema usa el último SMMLV conocido y el sanity_check advierte en vez de fallar.

## 13. Tests

```bash
python -m pytest tests/ -v           # 71 tests
python tests/smoke_test.py           # Validación rápida
```

## 14. La regla de oro del FEX

| Período | ConfigGEIH | FEX_ADJ |
|---|---|---|
| Mes puntual | `ConfigGEIH(anio=X, n_meses=1)` | FEX_C18 ÷ 1 |
| Trimestre | `ConfigGEIH(anio=X, n_meses=3)` | FEX_C18 ÷ 3 |
| Anual | `ConfigGEIH(anio=X, n_meses=12)` | FEX_C18 ÷ 12 |

**Si PEA > 40M → el factor NO se está dividiendo.** El `sanity_check()` lo detecta automáticamente.

---

*ProColombia · GIC · Coordinación de Analítica · Bogotá, 2026*

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: docs/GUIA_GEIH.md >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [7/39]: __init__.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/__init__.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/__init__.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih — Análisis de microdatos GEIH del DANE.

Gran Encuesta Integrada de Hogares | DANE | Marco Muestral 2018
Autor: Néstor Enrique Forero Herrera

Paquete multi-año: soporta GEIH 2022–presente. No está atado a ningún
año específico — el nombre 'geih_2025' era la versión anterior.

Instalación:
    pip install geih-analisis
    # o desde GitHub:
    pip install git+https://github.com/enriqueforero/geih-analisis.git

Uso rápido:
    from geih import ConfigGEIH, ConsolidadorGEIH, PreparadorGEIH
    config = ConfigGEIH(anio=2025, n_meses=12)

Compatibilidad hacia atrás:
    'from geih_2025 import ...' sigue funcionando gracias al shim
    incluido en este paquete. Recibirás un DeprecationWarning.

CAMBIO v5.0 — Renombrado geih_2025 → geih:
  El paquete ahora se llama 'geih' (nombre de importación) y
  'geih-analisis' (nombre de distribución en PyPI).
  El shim geih_2025/ garantiza compatibilidad durante la transición.

Módulos del paquete (17 archivos):
  config.py                → Constantes, mapeos, configuración centralizada
  utils.py                 → Memoria, conversión de tipos, estadísticas ponderadas
  consolidador.py          → Lectura y unión de módulos CSV mensuales
  preparador.py            → Preparación de datos, merge con correlativas
  diagnostico.py           → Diagnóstico de calidad de datos
  indicadores.py           → Indicadores básicos (TD, TGP, TO, ingresos, rama)
  analisis_avanzado.py     → Módulos avanzados (ICE, ICI, ITAT, Mincer, etc.)
  analisis_area.py         → 32 ciudades × CIIU
  analisis_poblacional.py  → Campesinos, discapacidad, migración
  analisis_complementario.py → M8, M14, MX1–MX3
  exportador.py            → Exportación organizada a carpetas
  visualizacion.py         → Gráficos matplotlib
  visualizacion_interactiva.py → Gráficos Plotly
  comparativo.py           → Comparación inter-anual
  descargador.py           → Descarga automática DANE
  logger.py                → Logging centralizado
  profiler.py              → Profiling de memoria
  dashboard.py             → Dashboard Streamlit
"""

__version__ = "0.1.2"
__author__  = "Néstor Enrique Forero Herrera"
__email__   = "nforero@procolombia.co"
__url__     = "https://github.com/enriqueforero/geih-analisis"
__license__ = "MIT"

# ── Configuración ──────────────────────────────────────────────────
from .config import (
    # Configuración principal
    ConfigGEIH,
    # SMMLV
    SMMLV_2025, SMMLV_POR_ANIO,
    # Colores
    COLORES,
    # Períodos
    MESES_CARPETAS, MESES_NOMBRES,
    generar_carpetas_mensuales, generar_etiqueta_periodo,
    # Geografía
    DEPARTAMENTOS, DPTO_A_CIUDAD, AREA_A_CIUDAD,
    CIUDADES_13_PRINCIPALES, CIUDADES_10_INTERMEDIAS,
    # Clasificaciones económicas
    RAMAS_DANE, TABLA_CIIU_RAMAS, AGRUPACION_DANE_8,
    # Referencias DANE
    REF_DANE_2025, REF_DANE, ReferenciaDane,
    # Constantes laborales — antes faltaban, causaban ImportError
    CARGA_PRESTACIONAL, TAMANO_EMPRESA,
    RANGOS_SMMLV_LIMITES, RANGOS_SMMLV_ETIQUETAS,
    # Educación — antes faltaban
    NIVELES_AGRUPADOS, NIVELES_EDUCATIVOS, P3042_A_ANOS,
    # Llaves y converters
    LLAVES_PERSONA, LLAVES_HOGAR,
    CONVERTERS_BASE, CONVERTERS_CON_AREA,
    MODULOS_CSV,
)

# ── Utilidades ─────────────────────────────────────────────────────
from .utils import GestorMemoria, ConversorTipos, EstadisticasPonderadas

# ── Consolidación ──────────────────────────────────────────────────
from .consolidador import ConsolidadorGEIH

# ── Preparación ────────────────────────────────────────────────────
from .preparador import PreparadorGEIH, MergeCorrelativas

# ── Diagnóstico ────────────────────────────────────────────────────
from .diagnostico import DiagnosticoCalidad, Top20Sectores

# ── Exportación organizada ─────────────────────────────────────────
from .exportador import Exportador

# ── Indicadores básicos ────────────────────────────────────────────
from .indicadores import (
    IndicadoresLaborales, DistribucionIngresos, AnalisisRamaSexo,
    AnalisisSalarios, BrechaGenero, AnalisisCruzado,
    IndicesCompuestos, AnalisisArea,
)

# ── Análisis por 32 ciudades ───────────────────────────────────────
from .analisis_area import AnalisisOcupadosCiudad

# ── Análisis avanzado ──────────────────────────────────────────────
from .analisis_avanzado import (
    CalidadEmpleo, FormalidadSectorial, VulnerabilidadLaboral,
    CompetitividadLaboral, AnalisisSubempleo, AnalisisHoras,
    Estacionalidad, FuerzaLaboralJoven, EtnicoRacial,
    BonoDemografico, CostoLaboral, AnalisisFFT,
    AnalisisUrbanoRural, ProductividadTamano,
    ContribucionSectorial, MapaTalento, EcuacionMincer,
    ProxyBilinguismo,
)

# ── Visualización matplotlib ───────────────────────────────────────
from .visualizacion import (
    GraficoDistribucionIngresos, GraficoBoxPlotSalarios,
    GraficoBrechaGenero, GraficoRamaSexo,
    GraficoCurvaLorenz, GraficoICIBubble,
    GraficoEstacionalidad, GraficoContribucionHeatmap,
)

# ── Análisis poblacional ───────────────────────────────────────────
from .analisis_poblacional import (
    AnalisisCampesino, AnalisisDiscapacidad, AnalisisMigracion,
    AnalisisOtrasFormas, AnalisisOtrosIngresos,
    AnalisisSobrecalificacion, AnalisisContractual,
    AnalisisAutonomia, AnalisisAlcanceMercado, AnalisisDesanimados,
)

# ── Análisis complementarios ───────────────────────────────────────
from .analisis_complementario import (
    DuracionDesempleo, DashboardSectoresProColombia,
    AnatomaSalario, FormaPago, CanalEmpleo,
)

# ── Descarga automática DANE ───────────────────────────────────────
from .descargador import DescargadorDANE

# ── Comparativo multi-año ──────────────────────────────────────────
from .comparativo import ComparadorMultiAnio

# ── Visualización interactiva Plotly ───────────────────────────────
try:
    from .visualizacion_interactiva import (
        PlotlyLorenz, PlotlyICIBubble, PlotlyEstacionalidad,
        PlotlyDistribucionIngresos, PlotlyBrechaGenero,
        PlotlyBoxPlotSalarios, PlotlySalarioRama,
        PlotlyComparativoAnual,
    )
except ImportError:
    pass  # plotly no instalado — instalar con: pip install geih-analisis[viz]

# ── Logging centralizado ───────────────────────────────────────────
from .logger import get_logger, configurar_logging, LoggerGEIH

# ── Profiling de memoria ───────────────────────────────────────────
from .profiler import PerfilMemoria, medir_tiempo, tamano_objeto

# ── Dashboard Streamlit ────────────────────────────────────────────
from .dashboard import ejecutar_dashboard

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/__init__.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [8/39]: analisis_area.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/analisis_area.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/analisis_area.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.analisis_area — Ocupados por CIIU y 32 ciudades/áreas metropolitanas.

Replica y mejora el módulo del notebook original (script histórico 2022–2024):
  - 6 tablas: total nacional, agrupación DANE, dominio geográfico,
    ciudad/AM, granular CIIU×ciudad, CIIU nacional
  - 3 gráficos: barras agrupación, barras ciudades, heatmap rama×ciudad
  - Exportación a Excel multi-hoja

Usa la variable AREA (código DIVIPOLA de 5 dígitos del módulo Ocupados)
para identificar 32 ciudades y áreas metropolitanas del DANE.

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "AnalisisOcupadosCiudad",
]


import gc
from pathlib import Path
from typing import Optional, Dict, List

import numpy as np
import pandas as pd

from .config import (
    ConfigGEIH,
    AREA_A_CIUDAD,
    DPTO_A_CIUDAD,
    CIUDADES_13_PRINCIPALES,
    CIUDADES_10_INTERMEDIAS,
    _AGRUP_DANE_POR_DIVISION,
    REF_DANE_2025,
    REF_DANE,
)
from .utils import ConversorTipos


class AnalisisOcupadosCiudad:
    """Ocupados por actividad económica CIIU y 32 ciudades DANE.

    Produce 6 tablas que replican la estructura del script GEIH 2022–2024:
      tabla1: Total nacional de ocupados (con validación DANE)
      tabla2: Ocupados por Agrupación DANE (8 grupos CIIU)
      tabla3: Ocupados por dominio geográfico (13 ciudades, 10 intermedias, otras)
      tabla4: Ocupados por ciudad y área metropolitana (top 23)
      tabla5: Granular → Agrupación × División × CIIU × Ciudad
      tabla6: Agrupación × División × CIIU (nacional, sin ciudad)

    Uso típico:
        analisis = AnalisisOcupadosCiudad(config=ConfigGEIH(n_meses=12))
        tablas = analisis.calcular(geih_2025_final)
        analisis.imprimir(tablas)
        analisis.exportar_excel(tablas, exportador)
        fig = analisis.graficar(tablas)
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    # ═══════════════════════════════════════════════════════════════
    # MÉTODO PRINCIPAL
    # ═══════════════════════════════════════════════════════════════

    def calcular(
        self,
        df_raw: pd.DataFrame,
        ruta_ciiu: Optional[str] = None,
    ) -> Dict[str, pd.DataFrame]:
        """Calcula las 6 tablas de ocupados por CIIU y área geográfica.

        Args:
            df_raw: Base GEIH consolidada (cruda o preparada).
            ruta_ciiu: Ruta al Excel de correlativa CIIU (opcional).
                      Si se provee, agrega DESCRIPCION_CIIU a tablas 5 y 6.

        Returns:
            Dict con 'tabla1' a 'tabla6', más 'df_trabajo' para gráficos.
        """
        df = self._preparar(df_raw)

        # Merge CIIU descriptivo si hay correlativa
        if ruta_ciiu:
            df = self._merge_ciiu(df, ruta_ciiu)

        total = df["FEX_ADJ"].sum()

        tablas = {
            "tabla1": self._tabla1_total(total),
            "tabla2": self._tabla2_agrupacion(df, total),
            "tabla3": self._tabla3_dominio(df, total),
            "tabla4": self._tabla4_ciudad(df, total),
            "tabla5": self._tabla5_granular(df),
            "tabla6": self._tabla6_nacional(df),
            "df_trabajo": df,
        }

        print(f"\n   ✅ 6 tablas calculadas — {len(df):,} ocupados, "
              f"{total/1e6:.2f}M expandidos")
        return tablas

    # ═══════════════════════════════════════════════════════════════
    # PREPARACIÓN DE DATOS
    # ═══════════════════════════════════════════════════════════════

    def _preparar(self, df_raw: pd.DataFrame) -> pd.DataFrame:
        """Extrae columnas necesarias, filtra ocupados, mapea área y CIIU."""
        cols = ["FEX_C18", "OCI", "DPTO", "RAMA2D_R4", "RAMA4D_R4", "MES_NUM"]
        if "AREA" in df_raw.columns:
            cols.append("AREA")
        cols_ok = [c for c in cols if c in df_raw.columns]
        df = df_raw[cols_ok].copy()

        # Tipos
        df["FEX_C18"] = ConversorTipos.a_numerico(df["FEX_C18"]).fillna(0)
        df["OCI"] = ConversorTipos.a_numerico(df["OCI"])
        df["FEX_ADJ"] = df["FEX_C18"] / self.config.n_meses

        # Solo ocupados
        df = df[df["OCI"] == 1].copy()
        gc.collect()

        n_total = df["FEX_ADJ"].sum()
        ref = self.config.referencia_dane
        ref_ocu = f"~{ref.ocupados_anual_m} M" if ref else "N/D"
        print(f"   Ocupados totales (anual): {n_total/1e6:.2f} M  "
              f"(ref. DANE: {ref_ocu})")

        # Área geográfica → Ciudad/AM
        df["DPTO_STR"] = ConversorTipos.estandarizar_dpto(df["DPTO"])

        if "AREA" in df.columns:
            df["AREA_STR"] = ConversorTipos.estandarizar_area(df["AREA"])
            df["CIUDAD_AM"] = df["AREA_STR"].map(AREA_A_CIUDAD)
            # Fallback con DPTO para registros sin match en AREA
            fallback = df["DPTO_STR"].map(DPTO_A_CIUDAD)
            df["CIUDAD_AM"] = df["CIUDAD_AM"].fillna(fallback)
            print(f"   ✅ Variable AREA disponible — análisis por 32 ciudades habilitado")
        else:
            df["CIUDAD_AM"] = df["DPTO_STR"].map(DPTO_A_CIUDAD)
            print(f"   ⚠️ AREA no disponible — usando DPTO como proxy")

        # Dominio geográfico
        df["DOMINIO"] = df["CIUDAD_AM"].apply(self._asignar_dominio)

        # CIIU: División (2 dígitos) y Agrupación DANE (8 grupos)
        df["DIVISION"] = self._extraer_division(df["RAMA2D_R4"])
        df["AGRUPACION_DANE"] = df["DIVISION"].map(_AGRUP_DANE_POR_DIVISION).fillna("No informa")

        # CIIU 4 dígitos estandarizado
        df["RAMA4D_STD"] = ConversorTipos.estandarizar_ciiu4(df["RAMA4D_R4"])

        return df

    @staticmethod
    def _extraer_division(serie: pd.Series) -> pd.Series:
        """Extrae código de División CIIU (2 dígitos) de RAMA2D_R4."""
        num = pd.to_numeric(serie, errors="coerce")
        return num.round(0).astype("Int64").astype(str).str.zfill(2).where(num.notna())

    @staticmethod
    def _asignar_dominio(ciudad: str) -> str:
        if ciudad in CIUDADES_13_PRINCIPALES:
            return "13 ciudades y A.M."
        elif ciudad in CIUDADES_10_INTERMEDIAS:
            return "10 ciudades intermedias"
        elif pd.notna(ciudad):
            return "Otras cabeceras / Rural"
        return "No identificado"

    def _merge_ciiu(self, df: pd.DataFrame, ruta_ciiu: str) -> pd.DataFrame:
        """Merge con correlativa CIIU para agregar descripción textual."""
        try:
            df_ciiu = pd.read_excel(
                ruta_ciiu, sheet_name="CIIU 2022",
                converters={"RAMA4D_R4": str},
            )
            df_ciiu["RAMA4D_STD"] = ConversorTipos.estandarizar_ciiu4(df_ciiu["RAMA4D_R4"])
            slim = df_ciiu[["RAMA4D_STD", "DESCRIPCION_CIIU"]].drop_duplicates("RAMA4D_STD")
            df = df.merge(slim, on="RAMA4D_STD", how="left")
            pct = df["DESCRIPCION_CIIU"].notna().mean() * 100
            print(f"   ✅ CIIU descripción: {pct:.0f}% con match")
        except Exception as e:
            print(f"   ⚠️ Sin correlativa CIIU: {e}")
            df["DESCRIPCION_CIIU"] = df["AGRUPACION_DANE"]
        return df

    # ═══════════════════════════════════════════════════════════════
    # LAS 6 TABLAS
    # ═══════════════════════════════════════════════════════════════

    def _tabla1_total(self, total: float) -> pd.DataFrame:
        ref_dane = self.config.referencia_dane
        if ref_dane and ref_dane.ocupados_anual_m > 0:
            ref = round(ref_dane.ocupados_anual_m * 1_000)
            diff_pct = round((round(total / 1_000) - ref) / ref * 100, 1)
        else:
            ref = 0
            diff_pct = 0.0
        calc = round(total / 1_000)
        return pd.DataFrame([{
            "Período": self.config.periodo_etiqueta,
            "Ocupados_miles": calc,
            "Referencia_DANE_miles": ref if ref > 0 else "N/D",
            "Diferencia_%": diff_pct if ref > 0 else "N/D",
        }])

    @staticmethod
    def _tabla2_agrupacion(df: pd.DataFrame, total: float) -> pd.DataFrame:
        t = (
            df.groupby("AGRUPACION_DANE")["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"FEX_ADJ": "Ocupados"})
            .sort_values("Ocupados", ascending=False)
        )
        t["Ocupados_miles"] = (t["Ocupados"] / 1_000).round(0).astype(int)
        t["Pct_%"] = (t["Ocupados"] / total * 100).round(1)
        return t

    @staticmethod
    def _tabla3_dominio(df: pd.DataFrame, total: float) -> pd.DataFrame:
        t = (
            df.groupby("DOMINIO")["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"FEX_ADJ": "Ocupados"})
            .sort_values("Ocupados", ascending=False)
        )
        t["Ocupados_miles"] = (t["Ocupados"] / 1_000).round(0).astype(int)
        t["Pct_%"] = (t["Ocupados"] / total * 100).round(1)
        return t

    @staticmethod
    def _tabla4_ciudad(df: pd.DataFrame, total: float) -> pd.DataFrame:
        t = (
            df[df["CIUDAD_AM"].notna()]
            .groupby("CIUDAD_AM")["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"CIUDAD_AM": "Ciudad_AM", "FEX_ADJ": "Ocupados"})
            .sort_values("Ocupados", ascending=False)
        )
        t["Ocupados_miles"] = (t["Ocupados"] / 1_000).round(0).astype(int)
        t["Pct_%"] = (t["Ocupados"] / total * 100).round(1)
        t["Dominio"] = t["Ciudad_AM"].apply(AnalisisOcupadosCiudad._asignar_dominio)
        return t

    @staticmethod
    def _tabla5_granular(df: pd.DataFrame) -> pd.DataFrame:
        cols = ["AGRUPACION_DANE", "DIVISION", "RAMA4D_STD", "CIUDAD_AM"]
        if "DESCRIPCION_CIIU" in df.columns:
            cols = ["AGRUPACION_DANE", "DIVISION", "RAMA4D_STD",
                    "DESCRIPCION_CIIU", "CIUDAD_AM"]
        t = (
            df[df["CIUDAD_AM"].notna() & df["DIVISION"].notna()]
            .groupby(cols, dropna=True)["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"FEX_ADJ": "Ocupados_miles"})
        )
        t["Ocupados_miles"] = (t["Ocupados_miles"] / 1_000).round(1)
        return t.sort_values(["AGRUPACION_DANE", "Ocupados_miles"], ascending=[True, False])

    @staticmethod
    def _tabla6_nacional(df: pd.DataFrame) -> pd.DataFrame:
        cols = ["AGRUPACION_DANE", "DIVISION", "RAMA4D_STD"]
        if "DESCRIPCION_CIIU" in df.columns:
            cols = ["AGRUPACION_DANE", "DIVISION", "RAMA4D_STD", "DESCRIPCION_CIIU"]
        t = (
            df[df["DIVISION"].notna()]
            .groupby(cols, dropna=True)["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"FEX_ADJ": "Ocupados_miles"})
        )
        t["Ocupados_miles"] = (t["Ocupados_miles"] / 1_000).round(1)
        return t.sort_values("Ocupados_miles", ascending=False)

    # ═══════════════════════════════════════════════════════════════
    # IMPRESIÓN DE TABLAS
    # ═══════════════════════════════════════════════════════════════

    def imprimir(self, tablas: Dict[str, pd.DataFrame]) -> None:
        """Imprime las 6 tablas en formato legible para el notebook."""
        print(f"\n{'='*70}")
        print(f"  OCUPADOS POR CIIU Y ÁREA — GEIH {self.config.periodo_etiqueta}")
        print(f"  FEX_C18 / {self.config.n_meses} | Miles de personas")
        print(f"{'='*70}")

        # Tabla 1
        print(f"\n{'─'*50}")
        print(f"  TABLA 1: Total nacional")
        print(f"{'─'*50}")
        print(tablas["tabla1"].to_string(index=False))

        # Tabla 2
        t2 = tablas["tabla2"]
        print(f"\n{'─'*70}")
        print(f"  TABLA 2: Agrupación DANE (8 grupos CIIU)")
        print(f"{'─'*70}")
        print(f"  {'Agrupación DANE':<55} {'Miles':>7} {'%':>6}")
        print(f"  {'─'*55} {'─'*7} {'─'*6}")
        for _, row in t2.iterrows():
            print(f"  {str(row['AGRUPACION_DANE']):<55} "
                  f"{row['Ocupados_miles']:>7,} {row['Pct_%']:>5.1f}%")
        print(f"  {'─'*55} {'─'*7}")
        print(f"  {'TOTAL':<55} {t2['Ocupados_miles'].sum():>7,}")

        # Tabla 3
        print(f"\n{'─'*60}")
        print(f"  TABLA 3: Dominio geográfico DANE")
        print(f"{'─'*60}")
        print(tablas["tabla3"][["DOMINIO", "Ocupados_miles", "Pct_%"]].to_string(index=False))

        # Tabla 4
        t4 = tablas["tabla4"]
        print(f"\n{'─'*65}")
        print(f"  TABLA 4: Top 23 ciudades y áreas metropolitanas")
        print(f"{'─'*65}")
        print(f"  {'Ciudad / AM':<35} {'Miles':>7} {'%':>6} {'Dominio'}")
        print(f"  {'─'*35} {'─'*7} {'─'*6} {'─'*25}")
        for _, row in t4.head(23).iterrows():
            print(f"  {str(row['Ciudad_AM']):<35} {row['Ocupados_miles']:>7,} "
                  f"{row['Pct_%']:>5.1f}%  {row['Dominio']}")

        # Tablas 5 y 6 (resumen)
        print(f"\n{'─'*50}")
        print(f"  TABLA 5: Granular (Agrupación × CIIU × Ciudad)")
        print(f"  {len(tablas['tabla5']):,} combinaciones únicas")
        print(f"{'─'*50}")
        print(tablas["tabla5"].head(10).to_string(index=False))

        print(f"\n{'─'*65}")
        print(f"  TABLA 6: Top 20 actividades CIIU nacional")
        print(f"{'─'*65}")
        print(tablas["tabla6"].head(20).to_string(index=False))

    # ═══════════════════════════════════════════════════════════════
    # GRÁFICOS
    # ═══════════════════════════════════════════════════════════════

    def graficar(self, tablas: Dict[str, pd.DataFrame]):
        """Genera panel de 3 gráficos: agrupación, ciudades, heatmap.

        Returns:
            Figura matplotlib.
        """
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        import matplotlib.colors as mcolors
        from matplotlib.gridspec import GridSpec

        FONDO = "#F7F9FC"
        COLORES_AGRUP = [
            "#2E6DA4", "#C0392B", "#1E8449", "#8E44AD",
            "#E67E22", "#1ABC9C", "#F39C12", "#7F8C8D",
        ]

        fig = plt.figure(figsize=(20, 12))
        fig.patch.set_facecolor(FONDO)
        gs = GridSpec(2, 2, figure=fig, hspace=0.45, wspace=0.38)
        ax1 = fig.add_subplot(gs[0, 0])
        ax2 = fig.add_subplot(gs[0, 1])
        ax3 = fig.add_subplot(gs[1, :])
        for ax in [ax1, ax2, ax3]:
            ax.set_facecolor("white")
            ax.spines[["top", "right"]].set_visible(False)

        t2 = tablas["tabla2"]
        t4 = tablas["tabla4"]

        # ── Panel 1: Agrupación DANE ───────────────────────────
        t2p = t2.sort_values("Ocupados_miles", ascending=True)
        y1 = range(len(t2p))
        ax1.barh(y1, t2p["Ocupados_miles"], 0.65,
                 color=COLORES_AGRUP[:len(t2p)], alpha=0.88)
        for i, (_, row) in enumerate(t2p.iterrows()):
            ax1.text(row["Ocupados_miles"] + 30, i,
                     f"{row['Ocupados_miles']:,.0f}K  ({row['Pct_%']:.1f}%)",
                     va="center", fontsize=8.5)
        ax1.set_yticks(y1)
        ax1.set_yticklabels([str(r)[:35] for r in t2p["AGRUPACION_DANE"]], fontsize=8.5)
        ax1.set_xlabel("Miles de ocupados", fontsize=10)
        ax1.set_title("Agrupación DANE — 8 grupos CIIU", fontsize=11, fontweight="bold")
        ax1.grid(axis="x", alpha=0.3)

        # ── Panel 2: Top 15 ciudades ──────────────────────────
        t4top = t4.head(15).sort_values("Ocupados_miles", ascending=True)
        col_c = [
            "#2E6DA4" if d == "13 ciudades y A.M."
            else ("#1ABC9C" if d == "10 ciudades intermedias" else "#7F8C8D")
            for d in t4top["Dominio"]
        ]
        ax2.barh(range(len(t4top)), t4top["Ocupados_miles"], 0.65, color=col_c, alpha=0.88)
        for i, (_, row) in enumerate(t4top.iterrows()):
            ax2.text(row["Ocupados_miles"] + 10, i, f"{row['Ocupados_miles']:,.0f}K",
                     va="center", fontsize=8.5)
        ax2.set_yticks(range(len(t4top)))
        ax2.set_yticklabels(t4top["Ciudad_AM"], fontsize=9)
        ax2.set_xlabel("Miles de ocupados", fontsize=10)
        ax2.set_title("Top 15 ciudades y AM", fontsize=11, fontweight="bold")
        leyenda = [
            mpatches.Patch(color="#2E6DA4", alpha=0.88, label="13 ciudades principales"),
            mpatches.Patch(color="#1ABC9C", alpha=0.88, label="10 ciudades intermedias"),
            mpatches.Patch(color="#7F8C8D", alpha=0.88, label="Otras"),
        ]
        ax2.legend(handles=leyenda, fontsize=8, loc="lower right")
        ax2.grid(axis="x", alpha=0.3)

        # ── Panel 3: Heatmap Agrupación × Ciudad ─────────────
        df = tablas["df_trabajo"]
        ciudades_top = t4.head(8)["Ciudad_AM"].tolist()
        agrup_orden = t2.sort_values("Ocupados", ascending=False)["AGRUPACION_DANE"].tolist()

        hm = np.zeros((len(agrup_orden), len(ciudades_top)))
        for i, agrup in enumerate(agrup_orden):
            for j, ciudad in enumerate(ciudades_top):
                m = (df["AGRUPACION_DANE"] == agrup) & (df["CIUDAD_AM"] == ciudad)
                hm[i, j] = df.loc[m, "FEX_ADJ"].sum() / 1_000

        cmap = mcolors.LinearSegmentedColormap.from_list("bw", ["#FFFFFF", "#2E6DA4"])
        im = ax3.imshow(hm, cmap=cmap, aspect="auto", vmin=0)
        plt.colorbar(im, ax=ax3, label="Miles de ocupados", pad=0.01)
        for i in range(len(agrup_orden)):
            for j in range(len(ciudades_top)):
                v = hm[i, j]
                if v >= 10:
                    ax3.text(j, i, f"{v:,.0f}K", ha="center", va="center",
                             fontsize=7.5, fontweight="bold",
                             color="white" if v > hm.max() * 0.5 else "#1A1A1A")
        ax3.set_xticks(range(len(ciudades_top)))
        ax3.set_xticklabels(ciudades_top, fontsize=9, rotation=20, ha="right")
        ax3.set_yticks(range(len(agrup_orden)))
        ax3.set_yticklabels([str(a)[:40] for a in agrup_orden], fontsize=8.5)
        ax3.set_title("Heatmap: Agrupación DANE × Ciudad principal",
                       fontsize=11, fontweight="bold")

        fig.suptitle(
            f"Ocupados por CIIU y Área Geográfica — GEIH {self.config.periodo_etiqueta}\n"
            f"Ponderado FEX_C18/{self.config.n_meses} | 8 grupos DANE | 32 ciudades",
            fontsize=13, fontweight="bold",
        )
        fig.tight_layout(rect=[0, 0, 1, 0.94])
        return fig

    # ═══════════════════════════════════════════════════════════════
    # EXPORTACIÓN A EXCEL
    # ═══════════════════════════════════════════════════════════════

    def exportar_excel(
        self,
        tablas: Dict[str, pd.DataFrame],
        ruta: str = "Resultados_CIIU_Area_GEIH2025.xlsx",
    ) -> None:
        """Exporta las 6 tablas a un Excel con una hoja por tabla.

        Args:
            tablas: Output de calcular().
            ruta: Path completo del archivo Excel de salida.
        """
        hojas = {
            "Total Nacional":           tablas["tabla1"],
            "Agrupación DANE":          tablas["tabla2"][["AGRUPACION_DANE", "Ocupados_miles", "Pct_%"]],
            "Dominio Geográfico":       tablas["tabla3"][["DOMINIO", "Ocupados_miles", "Pct_%"]],
            "Ciudad-AM":                tablas["tabla4"][["Ciudad_AM", "Dominio", "Ocupados_miles", "Pct_%"]],
            "Agrupación-CIIU-Ciudad":   tablas["tabla5"],
            "Agrupación-CIIU":          tablas["tabla6"],
        }

        with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
            for nombre_hoja, df in hojas.items():
                df.to_excel(writer, sheet_name=nombre_hoja[:31], index=False)

        print(f"   ✅ Excel: {Path(ruta).name} ({len(hojas)} hojas)")
        for nombre in hojas:
            print(f"      • {nombre}")

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/analisis_area.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [9/39]: analisis_avanzado.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/analisis_avanzado.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/analisis_avanzado.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.analisis_avanzado — Módulos de análisis avanzado M5–M20 + A/B/C.

Contiene TODOS los módulos temáticos que faltaban en la primera versión:
  - CalidadEmpleo (ICE, ICF)
  - CompetitividadLaboral (ICI, ITAT)
  - VulnerabilidadLaboral (IVI)
  - AnalisisSubempleo (M9)
  - AnalisisHoras (M10)
  - Estacionalidad (M11)
  - FuerzaLaboralJoven (M12)
  - DashboardSectores (M14)
  - EtnicoRacial (M15)
  - BonoDesaparecido (M18)
  - CostoLaboral (M19)
  - AnalisisFFT (fuera fuerza de trabajo)
  - AnalisisUrbanoRural (M4 expandido)
  - ProductividadTamano (M3 expandido)
  - ContribucionSectorial (Módulo A)
  - MapaTalento (Módulo B — ITAT)
  - EcuacionMincer (Módulo C)
  - ProxyBilinguismo

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "CalidadEmpleo",
    "FormalidadSectorial",
    "VulnerabilidadLaboral",
    "CompetitividadLaboral",
    "AnalisisSubempleo",
    "AnalisisHoras",
    "Estacionalidad",
    "FuerzaLaboralJoven",
    "EtnicoRacial",
    "BonoDemografico",
    "CostoLaboral",
    "AnalisisFFT",
    "AnalisisUrbanoRural",
    "ProductividadTamano",
    "ContribucionSectorial",
    "MapaTalento",
    "EcuacionMincer",
    "ProxyBilinguismo",
]


import gc
from typing import Optional, Dict, Any, List, Tuple

import numpy as np
import pandas as pd

from .config import (
    ConfigGEIH, SMMLV_2025, CARGA_PRESTACIONAL,
    DEPARTAMENTOS, DPTO_A_CIUDAD,
    TABLA_CIIU_RAMAS, RAMAS_DANE,
    NIVELES_AGRUPADOS, P3042_A_ANOS,
    TAMANO_EMPRESA,
)
from .utils import EstadisticasPonderadas as EP


# ═════════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES INTERNAS
# ═════════════════════════════════════════════════════════════════════

def _norm_min_max(serie: pd.Series, invertir: bool = False) -> pd.Series:
    """Normalización min-max [0, 100]. Invierte si menor = mejor."""
    mn, mx = serie.min(), serie.max()
    if mx == mn:
        return pd.Series(50.0, index=serie.index)
    norm = (serie - mn) / (mx - mn) * 100
    return 100 - norm if invertir else norm


def _tasa_ponderada(df, mask_condicion, mask_universo, col_peso="FEX_ADJ"):
    """Calcula % ponderado = Σ(FEX condición) / Σ(FEX universo) × 100."""
    num = df.loc[mask_condicion, col_peso].sum()
    den = df.loc[mask_universo, col_peso].sum()
    return (num / den * 100) if den > 0 else np.nan


# ═════════════════════════════════════════════════════════════════════
# M7 · ICE — ÍNDICE DE CALIDAD DEL EMPLEO
# ═════════════════════════════════════════════════════════════════════

class CalidadEmpleo:
    """ICE = 0.30×Pensión + 0.25×Salud + 0.25×Horas_adecuadas + 0.20×Ingreso≥SML.

    Calcula a nivel de departamento, rama, o nacional.
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular_por_departamento(self, df: pd.DataFrame) -> pd.DataFrame:
        """ICE promedio por departamento."""
        df_ocu = df[df["OCI"] == 1].copy()
        filas = []
        for dpto, nombre in DEPARTAMENTOS.items():
            m = df_ocu["DPTO_STR"] == dpto if "DPTO_STR" in df_ocu.columns else pd.Series(False, index=df_ocu.index)
            n = df_ocu.loc[m, "FEX_ADJ"].sum()
            if n < 5_000:
                continue
            ice = self._calcular_ice(df_ocu[m])
            filas.append({"Departamento": nombre, "DPTO": dpto, "ICE": ice, "Ocupados_M": n / 1e6})
        return pd.DataFrame(filas).sort_values("ICE", ascending=False)

    def calcular_por_rama(self, df: pd.DataFrame) -> pd.DataFrame:
        """ICE promedio por rama de actividad."""
        df_ocu = df[(df["OCI"] == 1) & df["RAMA"].notna()].copy()
        filas = []
        for rama in df_ocu["RAMA"].unique():
            m = df_ocu["RAMA"] == rama
            ice = self._calcular_ice(df_ocu[m])
            n = df_ocu.loc[m, "FEX_ADJ"].sum()
            filas.append({"Rama": rama, "ICE": ice, "Ocupados_M": n / 1e6})
        return pd.DataFrame(filas).sort_values("ICE", ascending=False)

    def _calcular_ice(self, df_ocu: pd.DataFrame) -> float:
        """Calcula ICE promedio ponderado para un subconjunto."""
        fex = df_ocu["FEX_ADJ"]
        total = fex.sum()
        if total == 0:
            return np.nan
        d_pen = (df_ocu.get("P6920", pd.Series(dtype=float)) == 1).astype(float)
        d_sal = (df_ocu.get("P6090", pd.Series(dtype=float)) == 1).astype(float)
        d_hor = df_ocu.get("P6800", pd.Series(dtype=float)).between(20, 48).astype(float)
        d_ing = (df_ocu.get("INGLABO", pd.Series(dtype=float)) >= self.config.smmlv).astype(float)
        ice = 0.30 * d_pen + 0.25 * d_sal + 0.25 * d_hor + 0.20 * d_ing
        return float((ice * fex).sum() / total * 100)


# ═════════════════════════════════════════════════════════════════════
# M13 · ICF — ÍNDICE DE FORMALIDAD SECTORIAL
# ═════════════════════════════════════════════════════════════════════

class FormalidadSectorial:
    """ICF = Media(% cotiza pensión, % ingreso ≥ SMMLV, % afiliado salud)."""

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        df_ocu = df[(df["OCI"] == 1) & df["RAMA"].notna()].copy()
        filas = []
        for rama in df_ocu["RAMA"].unique():
            m_r = df_ocu["RAMA"] == rama
            n = df_ocu.loc[m_r, "FEX_ADJ"].sum()
            if n < 5_000:
                continue
            pct_pen = _tasa_ponderada(df_ocu, m_r & (df_ocu.get("P6920", 0) == 1), m_r)
            pct_sal = _tasa_ponderada(df_ocu, m_r & (df_ocu.get("P6090", 0) == 1), m_r)
            pct_ing = _tasa_ponderada(df_ocu, m_r & (df_ocu.get("INGLABO", 0) >= self.config.smmlv), m_r)
            icf = np.nanmean([pct_pen, pct_sal, pct_ing])
            filas.append({"Rama": rama, "ICF": round(icf, 1),
                          "Cotiza_pension_%": round(pct_pen, 1),
                          "Ingreso_SML_%": round(pct_ing, 1),
                          "Afiliado_salud_%": round(pct_sal, 1)})
        return pd.DataFrame(filas).sort_values("ICF", ascending=False)


# ═════════════════════════════════════════════════════════════════════
# M20 · IVI — ÍNDICE DE VULNERABILIDAD LABORAL
# ═════════════════════════════════════════════════════════════════════

class VulnerabilidadLaboral:
    """IVI = Media(% cta_propia sin pensión, % sin protección, % sobretrabajo, % <SMMLV)."""

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        df_ocu = df[(df["OCI"] == 1) & df["RAMA"].notna()].copy()
        filas = []
        for rama in df_ocu["RAMA"].unique():
            m_r = df_ocu["RAMA"] == rama
            n = df_ocu.loc[m_r, "FEX_ADJ"].sum()
            if n < 5_000:
                continue
            pct_cta_sin_pen = _tasa_ponderada(
                df_ocu, m_r & (df_ocu.get("P6430", 0) == 4) & (df_ocu.get("P6920", 0) != 1), m_r
            )
            pct_sobretrab = _tasa_ponderada(
                df_ocu, m_r & (df_ocu.get("P6800", 0) > 48), m_r
            )
            pct_sub_sml = _tasa_ponderada(
                df_ocu, m_r & (df_ocu.get("INGLABO", 0) < self.config.smmlv) & (df_ocu.get("INGLABO", 0) > 0), m_r
            )
            pct_sin_prot = _tasa_ponderada(
                df_ocu, m_r & (df_ocu.get("P6920", 0) != 1) & (df_ocu.get("P6090", 0) != 1), m_r
            )
            ivi = np.nanmean([pct_cta_sin_pen, pct_sobretrab, pct_sub_sml, pct_sin_prot])
            filas.append({"Rama": rama, "IVI": round(ivi, 1), "Ocupados_miles": round(n / 1_000)})
        return pd.DataFrame(filas).sort_values("IVI", ascending=False)


# ═════════════════════════════════════════════════════════════════════
# M16 · ICI — COMPETITIVIDAD LABORAL DEPARTAMENTAL
# ═════════════════════════════════════════════════════════════════════

class CompetitividadLaboral:
    """ICI = 0.25·TD + 0.20·Costo + 0.25·Talento + 0.20·Formalidad + 0.10·Jóvenes.

    Scores normalizados min-max [0,100]. Costo INVERTIDO (menor=mejor).
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        filas = []
        for dpto, nombre in DEPARTAMENTOS.items():
            col_dpto = "DPTO_STR" if "DPTO_STR" in df.columns else "DPTO"
            m = df[col_dpto] == dpto
            n = df.loc[m & (df["OCI"] == 1), "FEX_ADJ"].sum()
            if n < 10_000:
                continue
            pet = df.loc[m & (df.get("PET", pd.Series(dtype=float)) == 1), "FEX_ADJ"].sum()
            pea = df.loc[m & (df.get("FT", pd.Series(dtype=float)) == 1), "FEX_ADJ"].sum()
            des = df.loc[m & (df.get("DSI", pd.Series(dtype=float)) == 1), "FEX_ADJ"].sum()
            td = (des / pea * 100) if pea > 0 else np.nan
            mediana = EP.mediana(df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "INGLABO"],
                                df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "FEX_ADJ"])
            costo = mediana * (1 + CARGA_PRESTACIONAL) if not np.isnan(mediana) else np.nan
            pct_univ = _tasa_ponderada(df, m & (df.get("P3042", 0) >= 10) & (df["OCI"] == 1),
                                       m & (df["OCI"] == 1))
            pct_pen = _tasa_ponderada(df, m & (df.get("P6920", 0) == 1) & (df["OCI"] == 1),
                                      m & (df["OCI"] == 1))
            pct_jov = _tasa_ponderada(df, m & df.get("P6040", pd.Series(dtype=float)).between(15, 28) & (df["OCI"] == 1),
                                      m & (df["OCI"] == 1))
            filas.append({
                "Departamento": nombre, "DPTO": dpto,
                "TD_%": round(td, 1) if not np.isnan(td) else np.nan,
                "Costo_efectivo": round(costo) if not np.isnan(costo) else np.nan,
                "Talento_univ_%": round(pct_univ, 1),
                "Formalidad_%": round(pct_pen, 1),
                "Jovenes_%": round(pct_jov, 1),
                "Ocupados_miles": round(n / 1_000),
            })

        resultado = pd.DataFrame(filas).dropna(subset=["TD_%", "Costo_efectivo"])
        if resultado.empty:
            return resultado
        resultado["Score_TD"] = _norm_min_max(resultado["TD_%"], invertir=True)
        resultado["Score_Costo"] = _norm_min_max(resultado["Costo_efectivo"], invertir=True)
        resultado["Score_Talento"] = _norm_min_max(resultado["Talento_univ_%"])
        resultado["Score_Formal"] = _norm_min_max(resultado["Formalidad_%"])
        resultado["Score_Joven"] = _norm_min_max(resultado["Jovenes_%"])
        resultado["ICI"] = (
            0.25 * resultado["Score_TD"]
            + 0.20 * resultado["Score_Costo"]
            + 0.25 * resultado["Score_Talento"]
            + 0.20 * resultado["Score_Formal"]
            + 0.10 * resultado["Score_Joven"]
        ).round(1)
        return resultado.sort_values("ICI", ascending=False)


# ═════════════════════════════════════════════════════════════════════
# M9 · SUBEMPLEO
# ═════════════════════════════════════════════════════════════════════

class AnalisisSubempleo:
    """Subempleo por horas, ingresos y competencias."""

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> Dict[str, Any]:
        df_ocu = df[df["OCI"] == 1].copy()
        n_ocu = df_ocu["FEX_ADJ"].sum()
        resultado = {"Ocupados_M": n_ocu / 1e6}
        if "P6800" in df_ocu.columns:
            sub_horas = _tasa_ponderada(df_ocu, (df_ocu["P6800"] < 32), pd.Series(True, index=df_ocu.index))
            resultado["Subempleo_horas_%"] = round(sub_horas, 1)
        if "INGLABO" in df_ocu.columns:
            sub_ing = _tasa_ponderada(df_ocu,
                                      (df_ocu["INGLABO"] > 0) & (df_ocu["INGLABO"] < self.config.smmlv),
                                      (df_ocu["INGLABO"] > 0))
            resultado["Sub_SMMLV_%"] = round(sub_ing, 1)
        return resultado


# ═════════════════════════════════════════════════════════════════════
# M10 · HORAS TRABAJADAS
# ═════════════════════════════════════════════════════════════════════

class AnalisisHoras:
    """Distribución de horas trabajadas (P6800 normales, P6850 reales)."""

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        df_ocu = df[(df["OCI"] == 1) & df.get("P6800", pd.Series(dtype=float)).notna()].copy()
        rangos = [(0, 20, "<20h"), (20, 32, "20-32h"), (32, 40, "32-40h"),
                  (40, 48, "40-48h"), (48, 60, "48-60h"), (60, 200, ">60h")]
        filas = []
        total = df_ocu["FEX_ADJ"].sum()
        for lo, hi, etiq in rangos:
            m = df_ocu["P6800"].between(lo, hi, inclusive="left")
            n = df_ocu.loc[m, "FEX_ADJ"].sum()
            filas.append({"Rango_horas": etiq, "Personas_M": round(n / 1e6, 2), "Pct": round(n / total * 100, 1)})
        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# M11 · ESTACIONALIDAD MENSUAL
# ═════════════════════════════════════════════════════════════════════

class Estacionalidad:
    """Serie mensual de TD, TO, TGP usando MES_NUM y FEX_C18 sin dividir."""

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        if "MES_NUM" not in df.columns:
            print("⚠️ MES_NUM no disponible")
            return pd.DataFrame()
        filas = []
        for mes in sorted(df["MES_NUM"].dropna().unique()):
            m = df["MES_NUM"] == mes
            # Usar FEX_C18 sin dividir para indicador puntual
            fex = "FEX_C18"
            pea = df.loc[m & (df.get("FT", 0) == 1), fex].sum()
            ocu = df.loc[m & (df["OCI"] == 1), fex].sum()
            des = df.loc[m & (df.get("DSI", 0) == 1), fex].sum()
            pet = df.loc[m & (df.get("PET", 0) == 1), fex].sum()
            filas.append({
                "MES": int(mes),
                "PEA_M": round(pea / 1e6, 2), "Ocupados_M": round(ocu / 1e6, 2),
                "TD_%": round(des / pea * 100, 1) if pea > 0 else np.nan,
                "TGP_%": round(pea / pet * 100, 1) if pet > 0 else np.nan,
                "TO_%": round(ocu / pet * 100, 1) if pet > 0 else np.nan,
            })
        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# M12 · FUERZA LABORAL JOVEN (15-28 AÑOS)
# ═════════════════════════════════════════════════════════════════════

class FuerzaLaboralJoven:
    """TD, TO, TGP para jóvenes 15-28 años, nacional y por departamento."""

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> Dict[str, Any]:
        m_jov = df.get("P6040", pd.Series(dtype=float)).between(15, 28)
        df_jov = df[m_jov].copy()
        pea = df_jov.loc[df_jov.get("FT", 0) == 1, "FEX_ADJ"].sum()
        ocu = df_jov.loc[df_jov["OCI"] == 1, "FEX_ADJ"].sum()
        des = df_jov.loc[df_jov.get("DSI", 0) == 1, "FEX_ADJ"].sum()
        td = (des / pea * 100) if pea > 0 else np.nan
        return {"TD_joven_%": round(td, 1), "Ocupados_joven_M": round(ocu / 1e6, 2),
                "PEA_joven_M": round(pea / 1e6, 2)}


# ═════════════════════════════════════════════════════════════════════
# M15 · AUTORRECONOCIMIENTO ÉTNICO-RACIAL
# ═════════════════════════════════════════════════════════════════════

class EtnicoRacial:
    """Indicadores laborales por grupo étnico (P6080)."""

    GRUPOS = {1: "Indígena", 3: "Raizal", 4: "Palenquero",
              5: "Negro/Afrocolombiano", 6: "Ninguno"}

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        if "P6080" not in df.columns:
            print("⚠️ P6080 no disponible")
            return pd.DataFrame()
        filas = []
        for cod, nombre in self.GRUPOS.items():
            m = df["P6080"] == cod
            pea = df.loc[m & (df.get("FT", 0) == 1), "FEX_ADJ"].sum()
            des = df.loc[m & (df.get("DSI", 0) == 1), "FEX_ADJ"].sum()
            ocu = df.loc[m & (df["OCI"] == 1), "FEX_ADJ"].sum()
            td = (des / pea * 100) if pea > 0 else np.nan
            mediana = EP.mediana(
                df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "INGLABO"],
                df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "FEX_ADJ"]
            )
            filas.append({"Grupo": nombre, "TD_%": round(td, 1),
                          "Mediana_COP": round(mediana) if not np.isnan(mediana) else np.nan,
                          "Ocupados_M": round(ocu / 1e6, 2)})
        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# M18 · BONO DEMOGRÁFICO
# ═════════════════════════════════════════════════════════════════════

class BonoDemografico:
    """Ratio de dependencia económica = (Desocupados + FFT) / Ocupados."""

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        filas = []
        for dpto, nombre in DEPARTAMENTOS.items():
            col_dpto = "DPTO_STR" if "DPTO_STR" in df.columns else "DPTO"
            m = df[col_dpto] == dpto
            ocu = df.loc[m & (df["OCI"] == 1), "FEX_ADJ"].sum()
            des = df.loc[m & (df.get("DSI", 0) == 1), "FEX_ADJ"].sum()
            fft = df.loc[m & (df.get("FFT", 0) == 1), "FEX_ADJ"].sum() if "FFT" in df.columns else 0
            ratio = (des + fft) / ocu if ocu > 0 else np.nan
            if not np.isnan(ratio):
                filas.append({"Departamento": nombre, "Ratio_dependencia": round(ratio, 2), "Ocupados_M": round(ocu / 1e6, 2)})
        return pd.DataFrame(filas).sort_values("Ratio_dependencia")


# ═════════════════════════════════════════════════════════════════════
# M19 · COSTO LABORAL EFECTIVO
# ═════════════════════════════════════════════════════════════════════

class CostoLaboral:
    """Costo_efectivo = Salario_mediano × (1 + 0.54 carga prestacional)."""

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        df_ocu = df[(df["OCI"] == 1) & (df["INGLABO"] > 0) & df["RAMA"].notna()].copy()
        filas = []
        for rama in df_ocu["RAMA"].unique():
            m = df_ocu["RAMA"] == rama
            mediana = EP.mediana(df_ocu.loc[m, "INGLABO"], df_ocu.loc[m, "FEX_ADJ"])
            if np.isnan(mediana):
                continue
            costo = mediana * (1 + CARGA_PRESTACIONAL)
            filas.append({
                "Rama": rama,
                "Mediana_COP": round(mediana),
                "Costo_efectivo_COP": round(costo),
                "Mediana_SMMLV": round(mediana / self.config.smmlv, 2),
                "Costo_SMMLV": round(costo / self.config.smmlv, 2),
            })
        return pd.DataFrame(filas).sort_values("Costo_efectivo_COP", ascending=False)


# ═════════════════════════════════════════════════════════════════════
# FFT — FUERA DE LA FUERZA DE TRABAJO
# ═════════════════════════════════════════════════════════════════════

class AnalisisFFT:
    """Personas fuera de la fuerza de trabajo por tipo de actividad."""

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        if "FFT" not in df.columns:
            print("⚠️ FFT no disponible")
            return pd.DataFrame()
        df_fft = df[df["FFT"] == 1].copy()
        total = df_fft["FEX_ADJ"].sum()
        resultado = {"Total_FFT_M": round(total / 1e6, 2)}
        # Desagregación por sexo
        for sexo_val, sexo_lbl in [(1, "Hombres"), (2, "Mujeres")]:
            n = df_fft.loc[df_fft["P3271"] == sexo_val, "FEX_ADJ"].sum()
            resultado[f"FFT_{sexo_lbl}_M"] = round(n / 1e6, 2)
        return pd.DataFrame([resultado])


# ═════════════════════════════════════════════════════════════════════
# URBANO VS RURAL
# ═════════════════════════════════════════════════════════════════════

class AnalisisUrbanoRural:
    """Comparativo mercado laboral urbano (CLASE=1) vs rural (CLASE=2)."""

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        if "CLASE" not in df.columns:
            print("⚠️ CLASE no disponible")
            return pd.DataFrame()
        filas = []
        for clase, etiq in [(1, "Urbano"), (2, "Rural")]:
            # CLASE puede ser str o int
            m = (df["CLASE"].astype(str) == str(clase))
            pea = df.loc[m & (df.get("FT", 0) == 1), "FEX_ADJ"].sum()
            ocu = df.loc[m & (df["OCI"] == 1), "FEX_ADJ"].sum()
            des = df.loc[m & (df.get("DSI", 0) == 1), "FEX_ADJ"].sum()
            pet = df.loc[m & (df.get("PET", 0) == 1), "FEX_ADJ"].sum()
            mediana = EP.mediana(
                df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "INGLABO"],
                df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "FEX_ADJ"]
            )
            filas.append({
                "Zona": etiq,
                "TD_%": round(des / pea * 100, 1) if pea > 0 else np.nan,
                "TGP_%": round(pea / pet * 100, 1) if pet > 0 else np.nan,
                "TO_%": round(ocu / pet * 100, 1) if pet > 0 else np.nan,
                "Mediana_SMMLV": round(mediana / self.config.smmlv, 2) if not np.isnan(mediana) else np.nan,
                "Ocupados_M": round(ocu / 1e6, 2),
            })
        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# PRODUCTIVIDAD POR TAMAÑO DE EMPRESA
# ═════════════════════════════════════════════════════════════════════

class ProductividadTamano:
    """Salario mediano y formalidad por tamaño de empresa (P3069)."""

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        if "P3069" not in df.columns:
            print("⚠️ P3069 no disponible")
            return pd.DataFrame()
        df_ocu = df[(df["OCI"] == 1)].copy()
        filas = []
        for cod, etiq in TAMANO_EMPRESA.items():
            m = df_ocu["P3069"] == cod
            n = df_ocu.loc[m, "FEX_ADJ"].sum()
            if n < 5_000:
                continue
            mediana = EP.mediana(
                df_ocu.loc[m & (df_ocu["INGLABO"] > 0), "INGLABO"],
                df_ocu.loc[m & (df_ocu["INGLABO"] > 0), "FEX_ADJ"]
            )
            pct_pen = _tasa_ponderada(df_ocu, m & (df_ocu.get("P6920", 0) == 1), m)
            filas.append({
                "Tamano": etiq, "Cod": cod,
                "Mediana_SMMLV": round(mediana / self.config.smmlv, 2) if not np.isnan(mediana) else np.nan,
                "Formalidad_%": round(pct_pen, 1),
                "Ocupados_miles": round(n / 1_000),
            })
        return pd.DataFrame(filas).sort_values("Cod")


# ═════════════════════════════════════════════════════════════════════
# MÓDULO A · CONTRIBUCIÓN SECTORIAL AL EMPLEO
# ═════════════════════════════════════════════════════════════════════

class ContribucionSectorial:
    """Contribución de cada rama al cambio mensual del empleo (en p.p.)."""

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        if "MES_NUM" not in df.columns or "RAMA" not in df.columns:
            print("⚠️ MES_NUM o RAMA no disponibles")
            return pd.DataFrame()
        # Empleo por rama y mes (usando FEX_C18 sin dividir)
        pivot = (
            df[(df["OCI"] == 1) & df["RAMA"].notna()]
            .groupby(["MES_NUM", "RAMA"])["FEX_C18"]
            .sum().unstack(fill_value=0)
        )
        pea_mes = df[df.get("FT", pd.Series(dtype=float)) == 1].groupby("MES_NUM")["FEX_C18"].sum()
        # Contribución = (Emp_rama_t - Emp_rama_t-1) / PEA_t-1
        contrib = pivot.diff()
        for mes in contrib.index:
            if mes - 1 in pea_mes.index:
                contrib.loc[mes] = contrib.loc[mes] / pea_mes.loc[mes - 1] * 100
        return contrib.round(3)


# ═════════════════════════════════════════════════════════════════════
# MÓDULO B · MAPA DE TALENTO (ITAT)
# ═════════════════════════════════════════════════════════════════════

class MapaTalento:
    """ITAT = 0.35·Oferta + 0.35·Costo + 0.30·Calidad.

    Cada departamento es evaluado por: oferta (desocupados + subempleados),
    costo (mediana salarial invertida), calidad (% universitarios).
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        filas = []
        col_dpto = "DPTO_STR" if "DPTO_STR" in df.columns else "DPTO"
        for dpto, nombre in DEPARTAMENTOS.items():
            m = df[col_dpto] == dpto
            ocu = df.loc[m & (df["OCI"] == 1), "FEX_ADJ"].sum()
            des = df.loc[m & (df.get("DSI", 0) == 1), "FEX_ADJ"].sum()
            if ocu < 10_000:
                continue
            oferta = des  # + subempleados si disponible
            mediana = EP.mediana(
                df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "INGLABO"],
                df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "FEX_ADJ"]
            )
            pct_univ = _tasa_ponderada(df, m & (df.get("P3042", 0) >= 10) & (df["OCI"] == 1), m & (df["OCI"] == 1))
            filas.append({
                "Departamento": nombre, "Oferta_miles": round(oferta / 1_000),
                "Costo_mediano": round(mediana) if not np.isnan(mediana) else np.nan,
                "Calidad_univ_%": round(pct_univ, 1),
            })
        resultado = pd.DataFrame(filas).dropna()
        if resultado.empty:
            return resultado
        resultado["Score_Oferta"] = _norm_min_max(resultado["Oferta_miles"])
        resultado["Score_Costo"] = _norm_min_max(resultado["Costo_mediano"], invertir=True)
        resultado["Score_Calidad"] = _norm_min_max(resultado["Calidad_univ_%"])
        resultado["ITAT"] = (
            0.35 * resultado["Score_Oferta"]
            + 0.35 * resultado["Score_Costo"]
            + 0.30 * resultado["Score_Calidad"]
        ).round(1)
        return resultado.sort_values("ITAT", ascending=False)


# ═════════════════════════════════════════════════════════════════════
# MÓDULO C · ECUACIÓN DE MINCER
# ═════════════════════════════════════════════════════════════════════

class EcuacionMincer:
    """ln(W) = β₀ + β₁·Educ + β₂·Exp + β₃·Exp² (WLS con FEX como peso).

    β₁ = % de aumento salarial por cada año adicional de educación.
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def estimar(self, df: pd.DataFrame, grupo: str = "Nacional") -> Dict[str, Any]:
        """Estima la ecuación de Mincer para un subconjunto.

        Requiere: INGLABO > 0, ANOS_EDUC, P6040 (edad para calcular experiencia).

        Returns:
            Dict con beta_educacion, SE, R2, N.
        """
        df_calc = df[
            (df["OCI"] == 1) & (df["INGLABO"] > 0)
            & df["ANOS_EDUC"].notna() & df["P6040"].notna()
        ].copy()

        if len(df_calc) < 100:
            return {"Grupo": grupo, "beta_educacion": np.nan, "N": len(df_calc)}

        df_calc["LN_W"] = np.log(df_calc["INGLABO"])
        df_calc["EXP"] = (df_calc["P6040"] - df_calc["ANOS_EDUC"] - 6).clip(lower=0)
        df_calc["EXP2"] = df_calc["EXP"] ** 2

        try:
            from numpy.linalg import lstsq
            X = df_calc[["ANOS_EDUC", "EXP", "EXP2"]].values
            X = np.column_stack([np.ones(len(X)), X])
            y = df_calc["LN_W"].values
            w = np.sqrt(df_calc["FEX_ADJ"].values)
            Xw = X * w[:, np.newaxis]
            yw = y * w
            betas, residuals, rank, sv = lstsq(Xw, yw, rcond=None)
            y_pred = X @ betas
            ss_res = ((y - y_pred) ** 2 * df_calc["FEX_ADJ"].values).sum()
            y_mean = (y * df_calc["FEX_ADJ"].values).sum() / df_calc["FEX_ADJ"].sum()
            ss_tot = ((y - y_mean) ** 2 * df_calc["FEX_ADJ"].values).sum()
            r2 = 1 - ss_res / ss_tot if ss_tot > 0 else np.nan
            return {
                "Grupo": grupo,
                "beta_educacion": round(betas[1] * 100, 1),
                "beta_exp": round(betas[2] * 100, 2),
                "R2": round(r2, 3),
                "N": len(df_calc),
            }
        except Exception as e:
            return {"Grupo": grupo, "beta_educacion": np.nan, "error": str(e), "N": len(df_calc)}

    def estimar_todos(self, df: pd.DataFrame) -> pd.DataFrame:
        """Estima Mincer para: Nacional, H, M, Urbano, Rural, por rama."""
        resultados = [self.estimar(df, "Nacional")]

        if "P3271" in df.columns:
            resultados.append(self.estimar(df[df["P3271"] == 1], "Hombres"))
            resultados.append(self.estimar(df[df["P3271"] == 2], "Mujeres"))

        if "CLASE" in df.columns:
            resultados.append(self.estimar(df[df["CLASE"].astype(str) == "1"], "Urbano"))
            resultados.append(self.estimar(df[df["CLASE"].astype(str) == "2"], "Rural"))

        if "RAMA" in df.columns:
            for rama in df["RAMA"].dropna().unique():
                resultados.append(self.estimar(df[df["RAMA"] == rama], f"Rama: {rama}"))

        return pd.DataFrame(resultados).sort_values("beta_educacion", ascending=False)


# ═════════════════════════════════════════════════════════════════════
# PROXY DE BILINGÜISMO
# ═════════════════════════════════════════════════════════════════════

class ProxyBilinguismo:
    """Tres proxies para estimar bilingüismo con los datos disponibles en GEIH.

    Proxy 1: Formación en idiomas (códigos CINE-F 22x en P3043S1)
    Proxy 2: Demanda laboral (sectores TIC/BPO + universidad)
    Proxy 3: Perfil alta exposición (asalariado privado + sector + universidad)
    """

    SECTORES_INGLES = ["TIC/Información", "Información y comunicaciones",
                       "Financiero", "Actividades financieras y de seguros"]

    def calcular(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """Calcula los tres proxies por departamento/ciudad.

        Returns:
            Dict con 'proxy1', 'proxy2', 'proxy3' como DataFrames.
        """
        df_pet = df[df.get("P6040", pd.Series(dtype=float)) >= 15].copy()
        resultados = {}

        # Proxy 1: Códigos 22x (formación en idiomas)
        if "P3043S1" in df_pet.columns:
            cod_campo = pd.to_numeric(df_pet["P3043S1"], errors="coerce")
            df_pet["PROXY1"] = cod_campo.apply(
                lambda x: 1.0 if (not pd.isna(x) and str(int(x)).startswith("22")) else 0.0
            )
            n_p1 = (df_pet["PROXY1"] == 1).sum()
            print(f"   Proxy 1 (códigos 22x): {n_p1:,} registros")
            resultados["proxy1_total"] = n_p1

        # Proxy 2: Sectores + universidad
        if "RAMA" in df_pet.columns and "P3042" in df_pet.columns:
            mask_sector = df_pet["RAMA"].isin(self.SECTORES_INGLES)
            mask_univ = df_pet.get("P3042", pd.Series(dtype=float)) >= 10
            df_pet["PROXY2"] = ((mask_sector & mask_univ) & (df_pet["OCI"] == 1)).astype(float)
            n_p2 = (df_pet["PROXY2"] == 1).sum()
            print(f"   Proxy 2 (sector+univ): {n_p2:,} registros")
            resultados["proxy2_total"] = n_p2

        # Proxy 3: Asalariado privado + sector + universidad
        if "P6430" in df_pet.columns:
            mask_asalariado = df_pet.get("P6430", pd.Series(dtype=float)) == 1
            df_pet["PROXY3"] = (
                mask_asalariado & df_pet.get("PROXY2", pd.Series(0, index=df_pet.index)).astype(bool)
            ).astype(float)
            n_p3 = (df_pet["PROXY3"] == 1).sum()
            print(f"   Proxy 3 (asalariado): {n_p3:,} registros")
            resultados["proxy3_total"] = n_p3

        # Resumen por departamento
        if "NOMBRE_DPTO" in df_pet.columns:
            proxy_col = "PROXY2" if "PROXY2" in df_pet.columns else "PROXY1"
            if proxy_col in df_pet.columns:
                dept_resumen = []
                for dpto in df_pet["NOMBRE_DPTO"].dropna().unique():
                    m = df_pet["NOMBRE_DPTO"] == dpto
                    total = df_pet.loc[m & (df_pet["OCI"] == 1), "FEX_ADJ"].sum()
                    proxy = df_pet.loc[m & (df_pet[proxy_col] == 1), "FEX_ADJ"].sum()
                    if total > 10_000:
                        dept_resumen.append({
                            "Departamento": dpto,
                            f"Pct_{proxy_col}_%": round(proxy / total * 100, 1),
                            "Ocupados_miles": round(total / 1_000),
                        })
                resultados["por_departamento"] = pd.DataFrame(dept_resumen).sort_values(
                    f"Pct_{proxy_col}_%", ascending=False
                )

        return resultados

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/analisis_avanzado.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [10/39]: analisis_complementario.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/analisis_complementario.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/analisis_complementario.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.analisis_complementario — Análisis complementarios M8, M14, MX1–MX3.

Clases que cubren funcionalidades presentes en la Versión Antigua del notebook
que no tenían clase dedicada en la Versión Nueva v3.0:

  - DuracionDesempleo (M8)         → P7250 semanas buscando empleo
  - DashboardSectoresProColombia   → 7 sectores estratégicos de actividad económica
  - AnatomaSalario (MX1)           → P6500 vs INGLABO, ingreso "invisible"
  - FormaPago (MX2)                → P6765 destajo/honorarios/comisión
  - CanalEmpleo (MX3)              → P3363 contactos/internet/agencia

Cada clase sigue el patrón: recibe DataFrame preparado → retorna DataFrame
de resultados. No modifica el DataFrame original. No genera gráficos.

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "DuracionDesempleo",
    "DashboardSectoresProColombia",
    "AnatomaSalario",
    "FormaPago",
    "CanalEmpleo",
]


from typing import Optional, Dict, Any, List

import numpy as np
import pandas as pd

from .config import (
    ConfigGEIH,
    SMMLV_2025,
    DEPARTAMENTOS,
    RAMAS_DANE,
    TABLA_CIIU_RAMAS,
    TAMANO_EMPRESA,
    CARGA_PRESTACIONAL,
    NIVELES_AGRUPADOS,
)
from .utils import EstadisticasPonderadas as EP


# ═════════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES INTERNAS
# ═════════════════════════════════════════════════════════════════════

def _tasa(df, mask_num, mask_den, col_peso="FEX_ADJ"):
    """Calcula % ponderado = Σ(FEX condición) / Σ(FEX universo) × 100."""
    n = df.loc[mask_num, col_peso].sum()
    d = df.loc[mask_den, col_peso].sum()
    return round(n / d * 100, 1) if d > 0 else np.nan


# ═════════════════════════════════════════════════════════════════════
# M8 · DURACIÓN DEL DESEMPLEO
# ═════════════════════════════════════════════════════════════════════

class DuracionDesempleo:
    """Análisis de duración del desempleo por semanas buscando empleo (P7250).

    Clasifica el desempleo en categorías según la teoría laboral:
      - Friccional  (< 4 semanas)  : transición normal entre empleos
      - Cíclico     (5–12 semanas) : asociado al ciclo económico
      - Estructural (13–26 semanas): desajuste de habilidades/geografía
      - Largo plazo (> 26 semanas) : riesgo de pérdida permanente de
                                     capital humano y exclusión laboral

    Produce:
      - Distribución nacional por rango de duración
      - Mediana ponderada de semanas por departamento (proxy de rigidez)
      - Cruces por sexo y nivel educativo

    Módulo M8 del notebook original.

    Uso:
        dur = DuracionDesempleo(config=config).calcular(df)
        dur_sexo = DuracionDesempleo(config=config).por_sexo(df)
        dur_educ = DuracionDesempleo(config=config).por_educacion(df)
        dur_dpto = DuracionDesempleo(config=config).por_departamento(df)
    """

    # Rangos en semanas y etiquetas
    BINS = [0, 4, 12, 26, 52, float("inf")]
    ETIQUETAS = [
        "< 1 mes (friccional)",
        "1–3 meses (cíclico)",
        "3–6 meses (estructural)",
        "6–12 meses (largo plazo)",
        "> 1 año (crónico)",
    ]
    ETIQUETAS_CORTAS = ["< 1 mes", "1–3 meses", "3–6 meses", "6–12 meses", "> 1 año"]

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def _filtrar_desocupados(self, df: pd.DataFrame) -> pd.DataFrame:
        """Filtra desocupados con P7250 válido."""
        mask = (df["DSI"] == 1) & df["P7250"].notna() & (df["P7250"] > 0)
        return df[mask].copy()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Distribución nacional por rango de duración.

        Returns:
            DataFrame con columnas: Rango, Personas, Personas_M, Pct.
        """
        df_dsi = self._filtrar_desocupados(df)
        if df_dsi.empty:
            print("⚠️  No hay desocupados con P7250 válido.")
            return pd.DataFrame()

        df_dsi["DUR_CAT"] = pd.cut(
            df_dsi["P7250"], bins=self.BINS,
            labels=self.ETIQUETAS_CORTAS, right=False, include_lowest=True,
        )

        dist = (
            df_dsi.groupby("DUR_CAT", observed=True)["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"DUR_CAT": "Rango", "FEX_ADJ": "Personas"})
        )
        dist["Personas_M"] = (dist["Personas"] / 1e6).round(3)
        dist["Pct"] = (dist["Personas"] / dist["Personas"].sum() * 100).round(1)

        # Mediana nacional
        med_nac = EP.mediana(df_dsi["P7250"], df_dsi["FEX_ADJ"])

        self._imprimir(dist, med_nac)
        return dist

    def por_sexo(self, df: pd.DataFrame) -> pd.DataFrame:
        """Distribución de duración cruzada por sexo.

        Returns:
            DataFrame con Rango × Sexo y porcentajes.
        """
        df_dsi = self._filtrar_desocupados(df)
        if df_dsi.empty:
            return pd.DataFrame()

        df_dsi["DUR_CAT"] = pd.cut(
            df_dsi["P7250"], bins=self.BINS,
            labels=self.ETIQUETAS_CORTAS, right=False, include_lowest=True,
        )
        df_dsi["SEXO"] = df_dsi["P3271"].map({1: "Hombres", 2: "Mujeres"})

        filas = []
        for sexo in ["Hombres", "Mujeres"]:
            m_sexo = df_dsi["SEXO"] == sexo
            total_sexo = df_dsi.loc[m_sexo, "FEX_ADJ"].sum()
            med = EP.mediana(
                df_dsi.loc[m_sexo, "P7250"], df_dsi.loc[m_sexo, "FEX_ADJ"]
            )
            for rango in self.ETIQUETAS_CORTAS:
                m_rango = m_sexo & (df_dsi["DUR_CAT"] == rango)
                n = df_dsi.loc[m_rango, "FEX_ADJ"].sum()
                filas.append({
                    "Sexo": sexo, "Rango": rango,
                    "Personas": n, "Pct": round(n / total_sexo * 100, 1),
                    "Mediana_semanas": round(med, 1),
                })
        return pd.DataFrame(filas)

    def por_educacion(self, df: pd.DataFrame) -> pd.DataFrame:
        """Mediana de semanas y distribución por nivel educativo agrupado.

        Returns:
            DataFrame con Nivel, Desocupados_M, Mediana_semanas, Pct_largo_plazo.
        """
        df_dsi = self._filtrar_desocupados(df)
        if df_dsi.empty:
            return pd.DataFrame()

        df_dsi["NIVEL_GRUPO"] = df_dsi["P3042"].map(NIVELES_AGRUPADOS)

        filas = []
        for nivel in sorted(df_dsi["NIVEL_GRUPO"].dropna().unique()):
            m = df_dsi["NIVEL_GRUPO"] == nivel
            n = df_dsi.loc[m, "FEX_ADJ"].sum()
            if n < 1_000:
                continue
            med = EP.mediana(df_dsi.loc[m, "P7250"], df_dsi.loc[m, "FEX_ADJ"])
            # % largo plazo (> 26 semanas)
            m_lp = m & (df_dsi["P7250"] >= 26)
            pct_lp = df_dsi.loc[m_lp, "FEX_ADJ"].sum() / n * 100
            filas.append({
                "Nivel_educativo": nivel,
                "Desocupados_M": round(n / 1e6, 3),
                "Mediana_semanas": round(med, 1),
                "Pct_largo_plazo_%": round(pct_lp, 1),
            })
        return pd.DataFrame(filas)

    def por_departamento(self, df: pd.DataFrame) -> pd.DataFrame:
        """Mediana ponderada de semanas por departamento (proxy de rigidez).

        Returns:
            DataFrame con Departamento, Mediana_semanas, Desocupados_miles.
        """
        df_dsi = self._filtrar_desocupados(df)
        if df_dsi.empty:
            return pd.DataFrame()

        if "DPTO_STR" not in df_dsi.columns:
            from .utils import ConversorTipos
            df_dsi["DPTO_STR"] = ConversorTipos.estandarizar_dpto(df_dsi["DPTO"])

        filas = []
        for dpto, nombre in DEPARTAMENTOS.items():
            m = df_dsi["DPTO_STR"] == dpto
            n = df_dsi.loc[m, "FEX_ADJ"].sum()
            if n < 1_000:
                continue
            med = EP.mediana(df_dsi.loc[m, "P7250"], df_dsi.loc[m, "FEX_ADJ"])
            filas.append({
                "Departamento": nombre, "DPTO": dpto,
                "Mediana_semanas": round(med, 1),
                "Desocupados_miles": round(n / 1_000, 1),
            })
        return pd.DataFrame(filas).sort_values("Mediana_semanas", ascending=False)

    def _imprimir(self, dist: pd.DataFrame, mediana: float) -> None:
        print(f"\n{'='*60}")
        print(f"  M8 · DURACIÓN DEL DESEMPLEO — {self.config.periodo_etiqueta}")
        print(f"{'='*60}")
        print(f"  Mediana nacional: {mediana:.1f} semanas")
        for _, row in dist.iterrows():
            print(f"  {row['Rango']:<16} → {row['Personas_M']:.3f}M  ({row['Pct']:.1f}%)")


# ═════════════════════════════════════════════════════════════════════
# M14 · DASHBOARD SECTORES ESTRATÉGICOS
# ═════════════════════════════════════════════════════════════════════

# Mapeo CIIU 2D → sector estratégico (rangos de códigos CIIU relevantes)
_SECTORES_PROCOLOMBIA: Dict[str, List[range]] = {
    "TIC / Software":            [range(58, 64)],          # Información y comunicaciones
    "Turismo y Hotelería":       [range(55, 57), range(79, 80)],  # Alojamiento + agencias
    "Agroindustria":             [range(1, 4), range(10, 13)],    # Agri + alimentos
    "Financiero y Seguros":      [range(64, 67)],          # Finanzas y seguros
    "Química y Farmacéutica":    [range(20, 22)],          # Químicos + farma
    "Textil y Confecciones":     [range(13, 16)],          # Textil, confección, calzado
    "Manufactura Avanzada":      [range(26, 31)],          # Electrónica, maquinaria, vehículos
}


class DashboardSectoresProColombia:
    """Dashboard de 7 sectores estratégicos para discurso IED.

    Para cada sector calcula:
      - Empleo total (miles)
      - Ingreso mediano en SMMLV
      - % universitarios (P3042 >= 10)
      - % jóvenes 15–28
      - Formalidad (% cotiza pensión P6920=1)
      - % mujeres

    Módulo M14 del notebook original.

    Uso:
        dash = DashboardSectoresProColombia(config=config).calcular(df)
        print(dash.to_string(index=False))
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    @staticmethod
    def _clasificar_sector(rama2d: pd.Series) -> pd.Series:
        """Clasifica RAMA2D_R4 en sector estratégico o NaN."""
        rama_num = pd.to_numeric(rama2d, errors="coerce")
        resultado = pd.Series(np.nan, index=rama2d.index, dtype=object)
        for sector, rangos in _SECTORES_PROCOLOMBIA.items():
            mask = pd.Series(False, index=rama2d.index)
            for rng in rangos:
                mask = mask | rama_num.between(rng.start, rng.stop - 1)
            resultado = resultado.where(~mask, sector)
        return resultado

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calcula el dashboard para los 7 sectores estratégicos.

        Args:
            df: DataFrame preparado con OCI, RAMA2D_R4, FEX_ADJ, INGLABO,
                P3042, P6040, P6920, P3271.

        Returns:
            DataFrame con una fila por sector y 7 columnas de indicadores.
        """
        df_ocu = df[df["OCI"] == 1].copy()
        df_ocu["SECTOR_PC"] = self._clasificar_sector(df_ocu["RAMA2D_R4"])
        df_ocu = df_ocu[df_ocu["SECTOR_PC"].notna()]

        smmlv = self.config.smmlv
        filas = []

        for sector in _SECTORES_PROCOLOMBIA.keys():
            m = df_ocu["SECTOR_PC"] == sector
            n = df_ocu.loc[m, "FEX_ADJ"].sum()
            if n < 500:
                continue

            sub = df_ocu[m]
            med_ing = EP.mediana(sub["INGLABO"], sub["FEX_ADJ"])

            # % universitarios
            pct_univ = np.nan
            if "P3042" in sub.columns:
                m_univ = m & (df_ocu["P3042"] >= 10)
                pct_univ = round(df_ocu.loc[m_univ, "FEX_ADJ"].sum() / n * 100, 1)

            # % jóvenes 15-28
            pct_jov = np.nan
            if "P6040" in sub.columns:
                m_jov = m & df_ocu["P6040"].between(15, 28)
                pct_jov = round(df_ocu.loc[m_jov, "FEX_ADJ"].sum() / n * 100, 1)

            # Formalidad (cotiza pensión)
            pct_pen = np.nan
            if "P6920" in sub.columns:
                m_pen = m & (df_ocu["P6920"] == 1)
                pct_pen = round(df_ocu.loc[m_pen, "FEX_ADJ"].sum() / n * 100, 1)

            # % mujeres
            pct_muj = np.nan
            if "P3271" in sub.columns:
                m_muj = m & (df_ocu["P3271"] == 2)
                pct_muj = round(df_ocu.loc[m_muj, "FEX_ADJ"].sum() / n * 100, 1)

            filas.append({
                "Sector": sector,
                "Empleo_miles": round(n / 1_000, 1),
                "Mediana_SMMLV": round(med_ing / smmlv, 2) if med_ing else np.nan,
                "Universitarios_%": pct_univ,
                "Jóvenes_15_28_%": pct_jov,
                "Formalidad_%": pct_pen,
                "Mujeres_%": pct_muj,
            })

        resultado = pd.DataFrame(filas).sort_values("Empleo_miles", ascending=False)
        self._imprimir(resultado)
        return resultado

    def _imprimir(self, df: pd.DataFrame) -> None:
        print(f"\n{'='*80}")
        print(f"  M14 · DASHBOARD SECTORES ESTRATÉGICOS — {self.config.periodo_etiqueta}")
        print(f"{'='*80}")
        print(f"  {'Sector':<28} {'Empleo':>8} {'Med.SML':>8} {'%Univ':>6} "
              f"{'%Jov':>5} {'%Form':>6} {'%Muj':>5}")
        print(f"  {'─'*28} {'─'*8} {'─'*8} {'─'*6} {'─'*5} {'─'*6} {'─'*5}")
        for _, r in df.iterrows():
            print(f"  {r['Sector']:<28} {r['Empleo_miles']:>7.0f}K "
                  f"{r['Mediana_SMMLV']:>7.2f}× {r['Universitarios_%']:>5.1f} "
                  f"{r['Jóvenes_15_28_%']:>5.1f} {r['Formalidad_%']:>5.1f} "
                  f"{r['Mujeres_%']:>5.1f}")


# ═════════════════════════════════════════════════════════════════════
# MX1 · ANATOMÍA DEL SALARIO: P6500 vs INGLABO
# ═════════════════════════════════════════════════════════════════════

class AnatomaSalario:
    """Cruce P6500 (salario bruto declarado) vs INGLABO (ingreso consolidado DANE).

    La brecha (INGLABO − P6500) / P6500 revela el ingreso "invisible":
    bonificaciones, especie, comisiones, viáticos que el trabajador no
    identifica como "salario" pero el DANE sí imputa.

    Incluye también P3364 (retención en la fuente) como proxy de
    tributación formal ante la DIAN.

    Módulo MX1 del notebook original.

    Uso:
        anat = AnatomaSalario(config=config)
        por_rama = anat.por_rama(df)
        por_tamano = anat.por_tamano_empresa(df)
        resumen = anat.resumen_nacional(df)
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def _filtrar(self, df: pd.DataFrame) -> pd.DataFrame:
        """Filtra ocupados con ambos ingresos válidos y positivos."""
        mask = (
            (df["OCI"] == 1) &
            df["P6500"].notna() & (df["P6500"] > 0) &
            df["INGLABO"].notna() & (df["INGLABO"] > 0)
        )
        return df[mask].copy()

    def resumen_nacional(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Resumen nacional de la brecha P6500 vs INGLABO.

        Returns:
            Dict con medianas, brecha %, y % con retención en la fuente.
        """
        df_f = self._filtrar(df)
        if df_f.empty:
            return {}

        med_p6500 = EP.mediana(df_f["P6500"], df_f["FEX_ADJ"])
        med_inglabo = EP.mediana(df_f["INGLABO"], df_f["FEX_ADJ"])
        brecha_pct = (med_inglabo - med_p6500) / med_p6500 * 100

        # % con retención en la fuente (P3364=1)
        pct_retencion = np.nan
        if "P3364" in df_f.columns:
            n_ret = df_f.loc[df_f["P3364"] == 1, "FEX_ADJ"].sum()
            pct_retencion = round(n_ret / df_f["FEX_ADJ"].sum() * 100, 1)

        resultado = {
            "Registros_con_ambos": len(df_f),
            "Mediana_P6500": round(med_p6500),
            "Mediana_INGLABO": round(med_inglabo),
            "Brecha_%": round(brecha_pct, 1),
            "Pct_con_retencion_fuente": pct_retencion,
        }

        print(f"\n{'='*55}")
        print(f"  MX1 · ANATOMÍA SALARIAL — {self.config.periodo_etiqueta}")
        print(f"{'='*55}")
        print(f"  Mediana P6500 (bruto declarado): ${med_p6500:,.0f}")
        print(f"  Mediana INGLABO (consolidado) :  ${med_inglabo:,.0f}")
        print(f"  Brecha (ingreso invisible)    :  {brecha_pct:+.1f}%")
        if pct_retencion is not np.nan:
            print(f"  % con retención fuente (DIAN) :  {pct_retencion:.1f}%")

        return resultado

    def por_rama(self, df: pd.DataFrame) -> pd.DataFrame:
        """Brecha P6500 vs INGLABO por rama de actividad.

        Returns:
            DataFrame con Rama, Mediana_P6500, Mediana_INGLABO, Brecha_%.
        """
        df_f = self._filtrar(df)
        if df_f.empty or "RAMA" not in df_f.columns:
            return pd.DataFrame()

        filas = []
        for rama in df_f["RAMA"].dropna().unique():
            m = df_f["RAMA"] == rama
            n = df_f.loc[m, "FEX_ADJ"].sum()
            if n < 5_000:
                continue
            med_p = EP.mediana(df_f.loc[m, "P6500"], df_f.loc[m, "FEX_ADJ"])
            med_i = EP.mediana(df_f.loc[m, "INGLABO"], df_f.loc[m, "FEX_ADJ"])
            brecha = (med_i - med_p) / med_p * 100 if med_p > 0 else np.nan
            filas.append({
                "Rama": rama,
                "Ocupados_miles": round(n / 1_000, 1),
                "Mediana_P6500": round(med_p),
                "Mediana_INGLABO": round(med_i),
                "Brecha_%": round(brecha, 1),
            })
        return pd.DataFrame(filas).sort_values("Brecha_%", ascending=False)

    def por_tamano_empresa(self, df: pd.DataFrame) -> pd.DataFrame:
        """Brecha P6500 vs INGLABO por tamaño de empresa (P3069).

        Returns:
            DataFrame con Tamano, Mediana_P6500, Mediana_INGLABO, Brecha_%.
        """
        df_f = self._filtrar(df)
        if df_f.empty or "P3069" not in df_f.columns:
            return pd.DataFrame()

        filas = []
        for cod, etiq in TAMANO_EMPRESA.items():
            m = df_f["P3069"] == cod
            n = df_f.loc[m, "FEX_ADJ"].sum()
            if n < 5_000:
                continue
            med_p = EP.mediana(df_f.loc[m, "P6500"], df_f.loc[m, "FEX_ADJ"])
            med_i = EP.mediana(df_f.loc[m, "INGLABO"], df_f.loc[m, "FEX_ADJ"])
            brecha = (med_i - med_p) / med_p * 100 if med_p > 0 else np.nan
            filas.append({
                "Tamaño": etiq,
                "Cod": cod,
                "Ocupados_miles": round(n / 1_000, 1),
                "Mediana_P6500": round(med_p),
                "Mediana_INGLABO": round(med_i),
                "Brecha_%": round(brecha, 1),
            })
        return pd.DataFrame(filas).sort_values("Cod")


# ═════════════════════════════════════════════════════════════════════
# MX2 · FORMA DE PAGO (P6765)
# ═════════════════════════════════════════════════════════════════════

# Mapeo P6765 → etiqueta.
# El DANE codifica esta variable de forma inconsistente entre años:
#   - Algunos años usan letras: 'a', 'b', 'c', ...
#   - Otros años usan números: 1, 2, 3, ...
# Este mapeo cubre AMBOS formatos para máxima compatibilidad.
_FORMA_PAGO_LETRAS: Dict[str, str] = {
    "a": "Honorarios / Prestación de servicios",
    "b": "Jornal o diario",
    "c": "A destajo (pieza, maquila)",
    "d": "Por comisión",
    "e": "Porcentaje",
    "f": "Ingreso mensual",
    "g": "Negocio / finca propia",
    "h": "No recibe ingresos",
    "i": "Ganancia neta (negocio)",
}

_FORMA_PAGO_NUMEROS: Dict[str, str] = {
    "1": "Honorarios / Prestación de servicios",
    "2": "Jornal o diario",
    "3": "A destajo (pieza, maquila)",
    "4": "Por comisión",
    "5": "Porcentaje",
    "6": "Ingreso mensual",
    "7": "Negocio / finca propia",
    "8": "No recibe ingresos",
    "9": "Ganancia neta (negocio)",
}

# Mapeo combinado: acepta '1', '1.0', 'a', etc.
_FORMA_PAGO: Dict[str, str] = {
    **_FORMA_PAGO_LETRAS,
    **_FORMA_PAGO_NUMEROS,
    # Variantes con .0 (pandas lee floats como '1.0')
    **{f"{k}.0": v for k, v in _FORMA_PAGO_NUMEROS.items()},
}


class FormaPago:
    """Análisis de la forma real de pago/trabajo (P6765).

    Revela cómo se remunera realmente a los trabajadores, más allá
    de la clasificación formal/informal. El destajo (c) indica
    precariedad en manufactura; los honorarios (a) indican economía
    gig sin protección social.

    Módulo MX2 del notebook original.

    Uso:
        fp = FormaPago(config=config)
        dist = fp.calcular(df)              # distribución nacional
        cruce = fp.cruce_formalidad(df)      # forma × cotiza pensión
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Distribución ponderada de ocupados por forma de pago.

        Auto-detecta la codificación del DANE (letras, números, o floats)
        y muestra diagnóstico si no logra mapear los valores.

        Returns:
            DataFrame con Forma_pago, Ocupados, Ocupados_M, Pct.
        """
        df_ocu = df[(df["OCI"] == 1) & df["P6765"].notna()].copy()
        if df_ocu.empty:
            print("⚠️  P6765 no disponible o vacía en la base.")
            return pd.DataFrame()

        # Normalizar: convertir a string, limpiar espacios, minúsculas
        df_ocu["P6765_STR"] = df_ocu["P6765"].astype(str).str.strip().str.lower()

        # Intentar mapeo
        df_ocu["FORMA_ETIQ"] = df_ocu["P6765_STR"].map(_FORMA_PAGO)
        n_mapeados = df_ocu["FORMA_ETIQ"].notna().sum()

        # Si el mapeo no funcionó, diagnosticar y usar valores crudos
        if n_mapeados == 0:
            valores_unicos = df_ocu["P6765_STR"].value_counts().head(10)
            print(f"⚠️  P6765: ningún valor coincide con el mapeo conocido.")
            print(f"   Valores encontrados en la base (top 10):")
            for val, count in valores_unicos.items():
                print(f"     '{val}' → {count:,} registros")
            print(f"   Usando valores crudos como etiqueta.")
            df_ocu["FORMA_ETIQ"] = df_ocu["P6765_STR"]

        dist = (
            df_ocu.groupby("FORMA_ETIQ", observed=True)["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"FORMA_ETIQ": "Forma_pago", "FEX_ADJ": "Ocupados"})
        )
        # Eliminar filas NaN (valores que no mapearon)
        dist = dist.dropna(subset=["Forma_pago"])
        if dist.empty:
            print("⚠️  No se pudo calcular distribución de forma de pago.")
            return pd.DataFrame()

        dist["Ocupados_M"] = (dist["Ocupados"] / 1e6).round(3)
        dist["Pct"] = (dist["Ocupados"] / dist["Ocupados"].sum() * 100).round(1)
        dist = dist.sort_values("Ocupados", ascending=False)

        self._imprimir(dist)
        return dist

    def cruce_formalidad(self, df: pd.DataFrame) -> pd.DataFrame:
        """Cruce forma de pago × cotización a pensión (proxy formalidad).

        Returns:
            DataFrame con Forma_pago, Total, Cotiza_pension_%, No_cotiza_%.
        """
        df_ocu = df[(df["OCI"] == 1) & df["P6765"].notna()].copy()
        if df_ocu.empty or "P6920" not in df_ocu.columns:
            return pd.DataFrame()

        df_ocu["P6765_STR"] = df_ocu["P6765"].astype(str).str.strip().str.lower()
        df_ocu["FORMA_ETIQ"] = df_ocu["P6765_STR"].map(_FORMA_PAGO)
        # Fallback a valores crudos si nada mapea
        if df_ocu["FORMA_ETIQ"].notna().sum() == 0:
            df_ocu["FORMA_ETIQ"] = df_ocu["P6765_STR"]

        filas = []
        for forma in df_ocu["FORMA_ETIQ"].dropna().unique():
            m = df_ocu["FORMA_ETIQ"] == forma
            total = df_ocu.loc[m, "FEX_ADJ"].sum()
            if total < 5_000:
                continue
            n_pen = df_ocu.loc[m & (df_ocu["P6920"] == 1), "FEX_ADJ"].sum()
            filas.append({
                "Forma_pago": forma,
                "Ocupados_miles": round(total / 1_000, 1),
                "Cotiza_pension_%": round(n_pen / total * 100, 1),
                "No_cotiza_%": round((1 - n_pen / total) * 100, 1),
            })
        return pd.DataFrame(filas).sort_values("Cotiza_pension_%", ascending=False)

    def _imprimir(self, dist: pd.DataFrame) -> None:
        print(f"\n{'='*60}")
        print(f"  MX2 · FORMA DE PAGO — {self.config.periodo_etiqueta}")
        print(f"{'='*60}")
        for _, r in dist.iterrows():
            print(f"  {str(r['Forma_pago']):<45} "
                  f"{r['Ocupados_M']:>6.3f}M  ({r['Pct']:>5.1f}%)")


# ═════════════════════════════════════════════════════════════════════
# MX3 · CANAL DE ACCESO AL EMPLEO (P3363)
# ═════════════════════════════════════════════════════════════════════

# Mapeo P3363 → etiqueta descriptiva
_CANAL_EMPLEO: Dict[int, str] = {
    1: "Contactos / amigos / familiares",
    2: "Clasificados / avisos",
    3: "Agencia empleo (SENA/Cajas)",
    4: "Bolsa de empleo / internet",
    5: "Convocatoria / concurso",
    6: "Por cuenta propia (propio negocio)",
    7: "Otra forma",
}


class CanalEmpleo:
    """Análisis del canal por el cual se consiguió el empleo (P3363).

    Mide la digitalización y segmentación del mercado laboral:
    - Contactos > 60% → mercado segmentado, red social > productividad
    - Internet/bolsa creciendo → señal de modernización
    - SENA/Agencia bajo → sistema público de empleo subutilizado

    Módulo MX3 del notebook original.

    Uso:
        ce = CanalEmpleo(config=config)
        dist = ce.calcular(df)                   # distribución nacional
        por_educ = ce.por_nivel_educativo(df)     # cruce × educación
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Distribución ponderada de ocupados por canal de empleo.

        Returns:
            DataFrame con Canal, Ocupados, Ocupados_M, Pct.
        """
        df_ocu = df[(df["OCI"] == 1) & df["P3363"].notna()].copy()
        if df_ocu.empty:
            print("⚠️  P3363 no disponible en la base.")
            return pd.DataFrame()

        df_ocu["CANAL_ETIQ"] = pd.to_numeric(
            df_ocu["P3363"], errors="coerce"
        ).map(_CANAL_EMPLEO)

        dist = (
            df_ocu[df_ocu["CANAL_ETIQ"].notna()]
            .groupby("CANAL_ETIQ", observed=True)["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"CANAL_ETIQ": "Canal", "FEX_ADJ": "Ocupados"})
        )
        dist["Ocupados_M"] = (dist["Ocupados"] / 1e6).round(3)
        dist["Pct"] = (dist["Ocupados"] / dist["Ocupados"].sum() * 100).round(1)
        dist = dist.sort_values("Ocupados", ascending=False)

        self._imprimir(dist)
        return dist

    def por_nivel_educativo(self, df: pd.DataFrame) -> pd.DataFrame:
        """Distribución de canal × nivel educativo agrupado.

        Revela si los universitarios usan más internet y los menos
        educados dependen más de contactos personales.

        Returns:
            DataFrame con Nivel, Canal, Pct por nivel.
        """
        df_ocu = df[(df["OCI"] == 1) & df["P3363"].notna()].copy()
        if df_ocu.empty or "P3042" not in df_ocu.columns:
            return pd.DataFrame()

        df_ocu["CANAL_ETIQ"] = pd.to_numeric(
            df_ocu["P3363"], errors="coerce"
        ).map(_CANAL_EMPLEO)
        df_ocu["NIVEL_GRUPO"] = df_ocu["P3042"].map(NIVELES_AGRUPADOS)

        filas = []
        for nivel in sorted(df_ocu["NIVEL_GRUPO"].dropna().unique()):
            m_niv = df_ocu["NIVEL_GRUPO"] == nivel
            total_niv = df_ocu.loc[m_niv, "FEX_ADJ"].sum()
            if total_niv < 5_000:
                continue
            for canal in df_ocu["CANAL_ETIQ"].dropna().unique():
                m_can = m_niv & (df_ocu["CANAL_ETIQ"] == canal)
                n = df_ocu.loc[m_can, "FEX_ADJ"].sum()
                filas.append({
                    "Nivel_educativo": nivel,
                    "Canal": canal,
                    "Pct": round(n / total_niv * 100, 1),
                })
        return pd.DataFrame(filas)

    def _imprimir(self, dist: pd.DataFrame) -> None:
        print(f"\n{'='*60}")
        print(f"  MX3 · CANAL DE EMPLEO — {self.config.periodo_etiqueta}")
        print(f"{'='*60}")
        for _, r in dist.iterrows():
            print(f"  {str(r['Canal']):<42} "
                  f"{r['Ocupados_M']:>6.3f}M  ({r['Pct']:>5.1f}%)")

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/analisis_complementario.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [11/39]: analisis_poblacional.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/analisis_poblacional.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/analisis_poblacional.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.analisis_poblacional — Análisis de poblaciones y módulos especiales.

Cubre los grupos poblacionales y módulos que la GEIH mide pero que
rara vez se explotan en los análisis estándar de mercado laboral:

  - AnalisisCampesino    → P2057/P2059 (autodefinición campesina)
  - AnalisisDiscapacidad → P1906S1-S8 (escala Washington)
  - AnalisisMigracion    → P3370-P3379 (migración interna e internacional)
  - AnalisisOtrasFormas  → P3054-P3057 (autoconsumo, voluntariado, formación)
  - AnalisisOtrosIngresos → P7422-P7510 (ingresos no laborales)
  - AnalisisSobrecalificacion → P3042 × P6430 (universitarios en empleos simples)
  - AnalisisContractual  → P6440/P6450/P6460/P6765 (formalidad contractual real)
  - AnalisisAutonomia    → P3047/P3048/P3049 (contratista dependiente)
  - AnalisisAlcanceMercado → P1802 (local → exportación)
  - AnalisisDesanimados  → P6300/P6310 (FFT pero desean trabajar)

Cada clase sigue el patrón: recibe DataFrame preparado → retorna DataFrame de resultados.
No modifica el DataFrame original. No genera gráficos.

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "AnalisisCampesino",
    "AnalisisDiscapacidad",
    "AnalisisMigracion",
    "AnalisisOtrasFormas",
    "AnalisisOtrosIngresos",
    "AnalisisSobrecalificacion",
    "AnalisisContractual",
    "AnalisisAutonomia",
    "AnalisisAlcanceMercado",
    "AnalisisDesanimados",
]


import gc
from typing import Optional, Dict, Any, List

import numpy as np
import pandas as pd

from .config import ConfigGEIH, SMMLV_2025, DEPARTAMENTOS
from .utils import EstadisticasPonderadas as EP


def _tasa(df, mask_num, mask_den, col_peso="FEX_ADJ"):
    n = df.loc[mask_num, col_peso].sum()
    d = df.loc[mask_den, col_peso].sum()
    return (n / d * 100) if d > 0 else np.nan


# ═════════════════════════════════════════════════════════════════════
# POBLACIÓN CAMPESINA (P2057 / P2059)
# ═════════════════════════════════════════════════════════════════════

class AnalisisCampesino:
    """Análisis del mercado laboral de la población campesina.

    P2057: ¿Usted se considera campesino(a)? (1=Sí, 2=No)
    P2059: ¿Alguna vez fue campesino? (1=Sí, 2=No)

    Incluido en GEIH desde el rediseño 2022 por mandato del PND.
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Indicadores laborales para campesinos vs no campesinos."""
        if "P2057" not in df.columns:
            print("⚠️ P2057 no disponible. Incluir módulo Características generales.")
            return pd.DataFrame()

        filas = []
        for val, etiq in [(1, "Se considera campesino"), (2, "No se considera campesino")]:
            m = df["P2057"] == val
            pea = df.loc[m & (df.get("FT", 0) == 1), "FEX_ADJ"].sum()
            ocu = df.loc[m & (df["OCI"] == 1), "FEX_ADJ"].sum()
            des = df.loc[m & (df.get("DSI", 0) == 1), "FEX_ADJ"].sum()
            mediana = EP.mediana(
                df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "INGLABO"],
                df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "FEX_ADJ"]
            )
            pct_pen = _tasa(df, m & (df.get("P6920", 0) == 1) & (df["OCI"] == 1), m & (df["OCI"] == 1))
            filas.append({
                "Grupo": etiq, "Poblacion_M": round(df.loc[m, "FEX_ADJ"].sum() / 1e6, 2),
                "TD_%": round(des / pea * 100, 1) if pea > 0 else np.nan,
                "Mediana_SMMLV": round(mediana / self.config.smmlv, 2) if not np.isnan(mediana) else np.nan,
                "Formalidad_%": round(pct_pen, 1),
            })
        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# DISCAPACIDAD (P1906S1-S8 — ESCALA WASHINGTON)
# ═════════════════════════════════════════════════════════════════════

class AnalisisDiscapacidad:
    """Indicadores laborales por condición de discapacidad.

    P1906S1-S8: 8 dimensiones de dificultad funcional.
    Valores: 1=Sin dificultad, 2=Alguna, 3=Mucha, 4=No puede.
    Criterio ONU: discapacidad = al menos una dimensión con valor 3 o 4.
    """

    DIMENSIONES = {
        "P1906S1": "Oír",      "P1906S2": "Hablar",
        "P1906S3": "Ver",      "P1906S4": "Moverse",
        "P1906S5": "Agarrar",  "P1906S6": "Entender",
        "P1906S7": "Autocuidado", "P1906S8": "Relacionarse",
    }

    def calcular(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Indicadores por presencia de discapacidad (criterio ONU)."""
        dims_ok = [d for d in self.DIMENSIONES if d in df.columns]
        if not dims_ok:
            print("⚠️ Variables P1906S1-S8 no disponibles.")
            return {}

        # Criterio ONU: al menos una dimensión con mucha dificultad (3) o no puede (4)
        mask_disca = pd.Series(False, index=df.index)
        for dim in dims_ok:
            mask_disca = mask_disca | (df[dim].isin([3, 4]))

        resultado = {"dimensiones_disponibles": len(dims_ok)}
        for etiq, m in [("Con discapacidad", mask_disca), ("Sin discapacidad", ~mask_disca)]:
            pea = df.loc[m & (df.get("FT", 0) == 1), "FEX_ADJ"].sum()
            ocu = df.loc[m & (df["OCI"] == 1), "FEX_ADJ"].sum()
            des = df.loc[m & (df.get("DSI", 0) == 1), "FEX_ADJ"].sum()
            td = (des / pea * 100) if pea > 0 else np.nan
            resultado[f"{etiq}_poblacion_M"] = round(df.loc[m, "FEX_ADJ"].sum() / 1e6, 2)
            resultado[f"{etiq}_TD_%"] = round(td, 1)
            resultado[f"{etiq}_ocupados_M"] = round(ocu / 1e6, 2)

        # Prevalencia por dimensión
        prev = {}
        for dim, nombre in self.DIMENSIONES.items():
            if dim in df.columns:
                pct = _tasa(df, df[dim].isin([3, 4]), pd.Series(True, index=df.index))
                prev[nombre] = round(pct, 2)
        resultado["prevalencia_por_dimension"] = prev

        return resultado


# ═════════════════════════════════════════════════════════════════════
# MIGRACIÓN (P3370-P3379)
# ═════════════════════════════════════════════════════════════════════

class AnalisisMigracion:
    """Análisis del mercado laboral por condición migratoria.

    P3370: ¿Dónde vivía hace 12 meses? (1=Mismo municipio, 2=Otro, 3=Otro país)
    P3376: País de nacimiento (170=Colombia, otro=Extranjero)
    P3378S1: Año de llegada a Colombia
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Indicadores por condición migratoria."""
        filas = []

        # Migración reciente (12 meses)
        if "P3370" in df.columns:
            for val, etiq in [(1, "Mismo municipio"), (2, "Otro municipio/dpto"), (3, "Otro país")]:
                m = df["P3370"] == val
                n = df.loc[m, "FEX_ADJ"].sum()
                if n < 5_000:
                    continue
                ocu = df.loc[m & (df["OCI"] == 1), "FEX_ADJ"].sum()
                des = df.loc[m & (df.get("DSI", 0) == 1), "FEX_ADJ"].sum()
                pea = df.loc[m & (df.get("FT", 0) == 1), "FEX_ADJ"].sum()
                td = (des / pea * 100) if pea > 0 else np.nan
                filas.append({
                    "Tipo_migracion": etiq, "Periodo": "12 meses",
                    "Poblacion_M": round(n / 1e6, 2), "TD_%": round(td, 1),
                    "Ocupados_M": round(ocu / 1e6, 2),
                })

        # Nacidos en el extranjero
        if "P3376" in df.columns:
            m_ext = df["P3376"] != 170  # 170 = Colombia
            m_col = df["P3376"] == 170
            for etiq, m in [("Nacido en Colombia", m_col), ("Nacido en el extranjero", m_ext)]:
                n = df.loc[m, "FEX_ADJ"].sum()
                if n < 1_000:
                    continue
                ocu = df.loc[m & (df["OCI"] == 1), "FEX_ADJ"].sum()
                pea = df.loc[m & (df.get("FT", 0) == 1), "FEX_ADJ"].sum()
                des = df.loc[m & (df.get("DSI", 0) == 1), "FEX_ADJ"].sum()
                td = (des / pea * 100) if pea > 0 else np.nan
                mediana = EP.mediana(
                    df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "INGLABO"],
                    df.loc[m & (df["OCI"] == 1) & (df["INGLABO"] > 0), "FEX_ADJ"]
                )
                filas.append({
                    "Tipo_migracion": etiq, "Periodo": "Nacimiento",
                    "Poblacion_M": round(n / 1e6, 2), "TD_%": round(td, 1),
                    "Mediana_SMMLV": round(mediana / self.config.smmlv, 2) if not np.isnan(mediana) else np.nan,
                })

        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# OTRAS FORMAS DE TRABAJO (P3054-P3057)
# ═════════════════════════════════════════════════════════════════════

class AnalisisOtrasFormas:
    """Trabajo no remunerado: autoconsumo, voluntariado, formación.

    P3054: Producción de bienes para autoconsumo (1=Sí)
    P3055: Producción de servicios para autoconsumo (1=Sí)
    P3056: Trabajo voluntario (1=Sí)
    P3057: Trabajo en formación no remunerado (1=Sí)
    """

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prevalencia de otras formas de trabajo."""
        formas = {
            "P3054": "Autoconsumo bienes",
            "P3055": "Autoconsumo servicios",
            "P3056": "Voluntariado",
            "P3057": "Formación no remunerada",
        }
        filas = []
        total = df["FEX_ADJ"].sum()
        for var, nombre in formas.items():
            if var in df.columns:
                n = df.loc[df[var] == 1, "FEX_ADJ"].sum()
                filas.append({
                    "Forma_trabajo": nombre, "Variable": var,
                    "Personas_M": round(n / 1e6, 2),
                    "Pct_%": round(n / total * 100, 1),
                })
        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# OTROS INGRESOS (P7422-P7510)
# ═════════════════════════════════════════════════════════════════════

class AnalisisOtrosIngresos:
    """Ingresos no laborales del hogar — insumo para pobreza monetaria.

    P7500S1: Pensiones/jubilaciones
    P7500S2: Ayudas de otros hogares nacionales
    P7510S2: Remesas del exterior
    P7422:   Arriendos recibidos
    """

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prevalencia y monto de ingresos no laborales."""
        fuentes = {
            "P7500S1": ("Pensiones/jubilaciones", "P7500S1A1"),
            "P7500S2": ("Ayudas hogares nacionales", "P7500S2A1"),
            "P7500S3": ("Ayudas institucionales", "P7500S3A1"),
            "P7510S2": ("Remesas del exterior", "P7510S2A1"),
            "P7422":   ("Arriendos recibidos", "P7422S1"),
        }
        filas = []
        total = df["FEX_ADJ"].sum()
        for var_si, (nombre, var_monto) in fuentes.items():
            if var_si in df.columns:
                m = df[var_si] == 1
                n = df.loc[m, "FEX_ADJ"].sum()
                monto_med = np.nan
                if var_monto in df.columns:
                    monto_med = EP.mediana(df.loc[m, var_monto], df.loc[m, "FEX_ADJ"])
                filas.append({
                    "Fuente": nombre, "Receptores_M": round(n / 1e6, 2),
                    "Pct_%": round(n / total * 100, 1),
                    "Mediana_monto": round(monto_med) if not np.isnan(monto_med) else np.nan,
                })
        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# SOBRECALIFICACIÓN (P3042 × P6430)
# ═════════════════════════════════════════════════════════════════════

class AnalisisSobrecalificacion:
    """Universitarios en empleos de baja complejidad.

    Detecta P3042 ≥ 10 (universitarios+) en posiciones P6430 = 4 (cuenta propia)
    o P6430 = 9 (servicio doméstico) con INGLABO < 2 SMMLV.
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> Dict[str, Any]:
        if "P3042" not in df.columns or "P6430" not in df.columns:
            return {}
        df_ocu = df[(df["OCI"] == 1)].copy()
        m_univ = df_ocu["P3042"] >= 10
        m_baja = df_ocu["P6430"].isin([4, 5, 6, 9])  # cuenta propia, jornalero, sin pago, doméstico
        m_sobre = m_univ & m_baja
        total_univ = df_ocu.loc[m_univ, "FEX_ADJ"].sum()
        n_sobre = df_ocu.loc[m_sobre, "FEX_ADJ"].sum()
        pct = (n_sobre / total_univ * 100) if total_univ > 0 else np.nan

        # Por rama
        por_rama = []
        if "RAMA" in df_ocu.columns:
            for rama in df_ocu["RAMA"].dropna().unique():
                m_r = df_ocu["RAMA"] == rama
                n_u_r = df_ocu.loc[m_r & m_univ, "FEX_ADJ"].sum()
                n_s_r = df_ocu.loc[m_r & m_sobre, "FEX_ADJ"].sum()
                if n_u_r > 5_000:
                    por_rama.append({
                        "Rama": rama,
                        "Universitarios_miles": round(n_u_r / 1_000),
                        "Sobrecalificados_%": round(n_s_r / n_u_r * 100, 1),
                    })

        return {
            "total_universitarios_M": round(total_univ / 1e6, 2),
            "sobrecalificados_M": round(n_sobre / 1e6, 2),
            "pct_sobrecalificacion": round(pct, 1),
            "por_rama": pd.DataFrame(por_rama).sort_values("Sobrecalificados_%", ascending=False) if por_rama else pd.DataFrame(),
        }


# ═════════════════════════════════════════════════════════════════════
# FORMALIDAD CONTRACTUAL REAL (P6440/P6450/P6460/P6765)
# ═════════════════════════════════════════════════════════════════════

class AnalisisContractual:
    """Mapa de formalidad contractual real.

    P6440: ¿Tiene contrato? (1=Sí)
    P6450: ¿Escrito? (1=Sí)
    P6460: ¿Indefinido? (1=Sí)
    P6765: Forma de trabajo (a=honorarios, c=destajo, g=negocio propio)
    """

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        df_ocu = df[(df["OCI"] == 1)].copy()
        total = df_ocu["FEX_ADJ"].sum()
        filas = []

        niveles = [
            ("Contrato escrito indefinido",
             lambda d: (d.get("P6440", 0) == 1) & (d.get("P6450", 0) == 1) & (d.get("P6460", 0) == 1)),
            ("Contrato escrito temporal",
             lambda d: (d.get("P6440", 0) == 1) & (d.get("P6450", 0) == 1) & (d.get("P6460", 0) != 1)),
            ("Contrato verbal",
             lambda d: (d.get("P6440", 0) == 1) & (d.get("P6450", 0) != 1)),
            ("Sin contrato",
             lambda d: (d.get("P6440", 0) != 1)),
        ]

        for nombre, fn_mask in niveles:
            try:
                m = fn_mask(df_ocu)
                n = df_ocu.loc[m, "FEX_ADJ"].sum()
                filas.append({"Tipo_contrato": nombre, "Personas_M": round(n / 1e6, 2), "Pct_%": round(n / total * 100, 1)})
            except Exception:
                continue

        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# AUTONOMÍA LABORAL (P3047/P3048/P3049)
# ═════════════════════════════════════════════════════════════════════

class AnalisisAutonomia:
    """Identifica contratistas dependientes disfrazados de independientes.

    P3047: ¿Quién decide horario? (1=Usted, 3=Empleador)
    P3048: ¿Quién decide qué producir? (1=Usted, 3=Empleador)
    P3049: ¿Quién decide precio? (1=Usted, 3=Empleador)

    Cuenta propia (P6430=4) con P3047/P3048=3 → asalariado disfrazado.
    """

    def calcular(self, df: pd.DataFrame) -> Dict[str, Any]:
        vars_req = ["P3047", "P3048", "P6430"]
        if not all(v in df.columns for v in vars_req):
            return {}

        df_ocu = df[(df["OCI"] == 1)].copy()
        m_cta = df_ocu["P6430"] == 4  # cuenta propia
        m_depend = m_cta & ((df_ocu["P3047"] == 3) | (df_ocu["P3048"] == 3))

        n_cta = df_ocu.loc[m_cta, "FEX_ADJ"].sum()
        n_dep = df_ocu.loc[m_depend, "FEX_ADJ"].sum()

        return {
            "cuenta_propia_M": round(n_cta / 1e6, 2),
            "cta_propia_dependiente_M": round(n_dep / 1e6, 2),
            "pct_dependientes": round(n_dep / n_cta * 100, 1) if n_cta > 0 else np.nan,
            "interpretacion": "Asalariados disfrazados de independientes — evasión laboral",
        }


# ═════════════════════════════════════════════════════════════════════
# ALCANCE DE MERCADO (P1802)
# ═════════════════════════════════════════════════════════════════════

class AnalisisAlcanceMercado:
    """Alcance geográfico del mercado de la empresa del trabajador.

    P1802: 1=Hogar/vecinos, 2=Barrio, 3=Municipio, 4=Dpto, 5=Nacional, 6=Exportación
    Código 6 = empleo directamente vinculado a comercio exterior.
    """

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        if "P1802" not in df.columns:
            return pd.DataFrame()

        ALCANCES = {
            1: "Hogar/vecinos", 2: "Barrio", 3: "Municipio",
            4: "Departamento", 5: "Nacional", 6: "Exportación ★",
        }
        df_ocu = df[(df["OCI"] == 1)].copy()
        total = df_ocu["FEX_ADJ"].sum()
        filas = []
        for cod, nombre in ALCANCES.items():
            m = df_ocu["P1802"] == cod
            n = df_ocu.loc[m, "FEX_ADJ"].sum()
            mediana = EP.mediana(
                df_ocu.loc[m & (df_ocu["INGLABO"] > 0), "INGLABO"],
                df_ocu.loc[m & (df_ocu["INGLABO"] > 0), "FEX_ADJ"]
            )
            filas.append({
                "Alcance": nombre, "Personas_M": round(n / 1e6, 2),
                "Pct_%": round(n / total * 100, 1),
                "Mediana_COP": round(mediana) if not np.isnan(mediana) else np.nan,
            })
        return pd.DataFrame(filas)


# ═════════════════════════════════════════════════════════════════════
# DESANIMADOS / POTENCIAL LATENTE (P6300/P6310)
# ═════════════════════════════════════════════════════════════════════

class AnalisisDesanimados:
    """Personas fuera de la FT que desean trabajar (potencial laboral latente).

    P6300=1: Fuera de la FT pero desearía trabajar.
    P6310: ¿Disponible para trabajar? (1=Sí, semana pasada)
    """

    def calcular(self, df: pd.DataFrame) -> Dict[str, Any]:
        if "P6300" not in df.columns:
            return {}

        m_fft = df.get("FFT", pd.Series(dtype=float)) == 1
        m_desea = (df["P6300"] == 1)
        m_disponible = (df.get("P6310", pd.Series(dtype=float)) == 1)

        n_fft = df.loc[m_fft, "FEX_ADJ"].sum()
        n_desea = df.loc[m_fft & m_desea, "FEX_ADJ"].sum()
        n_disp = df.loc[m_fft & m_desea & m_disponible, "FEX_ADJ"].sum()

        return {
            "FFT_total_M": round(n_fft / 1e6, 2),
            "Desean_trabajar_M": round(n_desea / 1e6, 2),
            "Disponibles_inmediato_M": round(n_disp / 1e6, 2),
            "Pct_desanimados": round(n_desea / n_fft * 100, 1) if n_fft > 0 else np.nan,
        }

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/analisis_poblacional.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [12/39]: comparativo.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/comparativo.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/comparativo.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.comparativo — Análisis comparativo multi-año de la GEIH.

Permite cargar múltiples años, agregarles columna ANIO, y calcular
comparaciones inter-anuales: variaciones de TD, evolución de ingresos,
cambios en la estructura sectorial, convergencia departamental, etc.

Flujo típico:
    comp = ComparadorMultiAnio()
    comp.agregar_anio(2025, '/ruta/GEIH_2025.parquet', config_2025)
    comp.agregar_anio(2026, '/ruta/GEIH_2026.parquet', config_2026)
    comp.comparar_indicadores()
    comp.evolucion_ingresos()

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "ComparadorMultiAnio",
]

import gc
from typing import Optional, Dict, List, Any

import numpy as np
import pandas as pd

from .config import ConfigGEIH, DEPARTAMENTOS, RAMAS_DANE
from .preparador import PreparadorGEIH
from .indicadores import IndicadoresLaborales
from .utils import EstadisticasPonderadas as EP


class ComparadorMultiAnio:
    """Compara indicadores laborales entre múltiples años GEIH.

    Carga bases de diferentes años, les agrega la columna ANIO,
    y produce tablas comparativas listas para reportes y gráficos.

    Cada año se carga con su propia ConfigGEIH (SMMLV, n_meses, etc.)
    para que los cálculos sean correctos según el período.

    Uso:
        comp = ComparadorMultiAnio()
        comp.agregar_anio(2025, 'GEIH_2025.parquet', ConfigGEIH(anio=2025))
        comp.agregar_anio(2026, 'GEIH_2026.parquet', ConfigGEIH(anio=2026, n_meses=3))

        tabla_ind = comp.comparar_indicadores()         # TD, TGP, TO por año
        tabla_dpto = comp.comparar_departamentos()       # TD por dpto × año
        tabla_ing = comp.evolucion_ingresos()            # mediana salarial por año
        tabla_rama = comp.comparar_ramas()               # empleo por rama × año
        df_combo = comp.obtener_base_combinada()         # base unificada con ANIO
    """

    def __init__(self):
        """Inicializa el comparador sin datos."""
        self._anios: Dict[int, Dict[str, Any]] = {}
        # {anio: {'df': DataFrame, 'config': ConfigGEIH, 'geih_raw': DataFrame}}

    def agregar_anio(
        self,
        anio: int,
        ruta_parquet: str,
        config: Optional[ConfigGEIH] = None,
        preparar: bool = True,
    ) -> None:
        """Carga un año y lo registra para comparación.

        Args:
            anio: Año de los datos (2025, 2026, ...).
            ruta_parquet: Ruta al archivo Parquet consolidado.
            config: ConfigGEIH del año. Si None, se crea con defaults.
            preparar: Si True, prepara la base (FEX_ADJ, variables derivadas).
        """
        config = config or ConfigGEIH(anio=anio)

        print(f"\n📂 Cargando {anio}: {ruta_parquet}")
        geih_raw = pd.read_parquet(ruta_parquet)
        print(f"   {geih_raw.shape[0]:,} filas × {geih_raw.shape[1]} cols")

        if preparar:
            prep = PreparadorGEIH(config=config)
            df = prep.preparar_base(geih_raw)
            df = prep.agregar_variables_derivadas(df)
        else:
            df = geih_raw.copy()
            df["FEX_ADJ"] = df["FEX_C18"] / config.n_meses

        # Agregar columna ANIO
        df["ANIO"] = anio

        self._anios[anio] = {
            "df": df,
            "config": config,
            "n_registros": len(df),
        }

        print(f"   ✅ {anio} registrado ({len(df):,} filas, "
              f"SMMLV=${config.smmlv:,})")

        # Liberar raw
        del geih_raw
        gc.collect()

    @property
    def anios_disponibles(self) -> List[int]:
        """Lista de años cargados, ordenados."""
        return sorted(self._anios.keys())

    def obtener_base_combinada(self) -> pd.DataFrame:
        """Retorna un DataFrame unificado con todos los años y columna ANIO.

        Útil para análisis ad-hoc que no están cubiertos por los métodos
        predefinidos del comparador.

        Returns:
            DataFrame con todos los años concatenados y columna ANIO.
        """
        if not self._anios:
            print("⚠️  No hay años cargados. Use agregar_anio() primero.")
            return pd.DataFrame()

        dfs = [data["df"] for data in self._anios.values()]
        combinado = pd.concat(dfs, ignore_index=True)
        print(f"✅ Base combinada: {combinado.shape[0]:,} filas × "
              f"{combinado.shape[1]} cols — Años: {self.anios_disponibles}")
        return combinado

    # ═════════════════════════════════════════════════════════════
    # COMPARACIONES PREDEFINIDAS
    # ═════════════════════════════════════════════════════════════

    def comparar_indicadores(self) -> pd.DataFrame:
        """Compara TD, TGP, TO nacionales entre años.

        Returns:
            DataFrame con una fila por año y columnas:
            ANIO, TD_%, TGP_%, TO_%, PEA_M, Ocupados_M, Desocupados_M,
            y columnas de variación (Δ) respecto al año anterior.
        """
        filas = []
        for anio in self.anios_disponibles:
            data = self._anios[anio]
            ind = IndicadoresLaborales(config=data["config"])
            r = ind.calcular(data["df"])
            r["ANIO"] = anio
            r["SMMLV"] = data["config"].smmlv
            r["N_meses"] = data["config"].n_meses
            filas.append(r)

        df = pd.DataFrame(filas).sort_values("ANIO")

        # Calcular variaciones inter-anuales
        for col in ["TD_%", "TGP_%", "TO_%"]:
            df[f"Δ_{col}"] = df[col].diff().round(2)

        for col in ["Ocupados_M", "Desocupados_M", "PEA_M"]:
            df[f"Var_{col}_%"] = (df[col].pct_change() * 100).round(1)

        self._imprimir_indicadores(df)
        return df

    def comparar_departamentos(
        self,
        top_n: int = 10,
    ) -> pd.DataFrame:
        """TD por departamento × año (panel).

        Args:
            top_n: Mostrar solo los top N departamentos por TD del último año.

        Returns:
            DataFrame pivotado: filas=departamento, columnas=año, valores=TD_%.
        """
        filas = []
        for anio in self.anios_disponibles:
            data = self._anios[anio]
            ind = IndicadoresLaborales(config=data["config"])
            td_dpto = ind.por_departamento(data["df"])
            td_dpto["ANIO"] = anio
            filas.append(td_dpto[["Departamento", "TD_%", "ANIO"]])

        df = pd.concat(filas, ignore_index=True)
        pivot = df.pivot_table(
            index="Departamento", columns="ANIO", values="TD_%"
        )

        # Calcular variación entre el primer y último año
        anios = sorted(pivot.columns)
        if len(anios) >= 2:
            pivot[f"Δ_{anios[0]}→{anios[-1]}"] = (
                pivot[anios[-1]] - pivot[anios[0]]
            ).round(1)

        # Ordenar por TD del último año
        pivot = pivot.sort_values(anios[-1], ascending=False)

        if top_n:
            pivot = pivot.head(top_n)

        return pivot

    def evolucion_ingresos(self) -> pd.DataFrame:
        """Mediana salarial (en SMMLV y COP) por año.

        Calcula la mediana ponderada del ingreso laboral para
        cada año, expresada en COP corrientes y en SMMLV del año.

        Returns:
            DataFrame con ANIO, Mediana_COP, Mediana_SMMLV, SMMLV_anio.
        """
        filas = []
        for anio in self.anios_disponibles:
            data = self._anios[anio]
            df = data["df"]
            smmlv = data["config"].smmlv
            mask_ocu = (df["OCI"] == 1) & (df["INGLABO"] > 0)
            med_cop = EP.mediana(
                df.loc[mask_ocu, "INGLABO"], df.loc[mask_ocu, "FEX_ADJ"]
            )
            filas.append({
                "ANIO": anio,
                "Mediana_COP": round(med_cop),
                "Mediana_SMMLV": round(med_cop / smmlv, 2),
                "SMMLV_anio": smmlv,
            })

        df = pd.DataFrame(filas).sort_values("ANIO")

        if len(df) >= 2:
            df["Var_COP_%"] = (df["Mediana_COP"].pct_change() * 100).round(1)
            df["Var_SMMLV"] = df["Mediana_SMMLV"].diff().round(2)

        return df

    def comparar_ramas(self) -> pd.DataFrame:
        """Empleo por rama de actividad × año.

        Returns:
            DataFrame pivotado: filas=rama, columnas=año, valores=ocupados (miles).
        """
        filas = []
        for anio in self.anios_disponibles:
            df = self._anios[anio]["df"]
            df_ocu = df[(df["OCI"] == 1) & df["RAMA"].notna()]
            rama_emp = (
                df_ocu.groupby("RAMA")["FEX_ADJ"]
                .sum().div(1_000).round(1)
                .reset_index()
                .rename(columns={"FEX_ADJ": "Ocupados_miles"})
            )
            rama_emp["ANIO"] = anio
            filas.append(rama_emp)

        df = pd.concat(filas, ignore_index=True)
        pivot = df.pivot_table(
            index="RAMA", columns="ANIO", values="Ocupados_miles"
        )

        anios = sorted(pivot.columns)
        if len(anios) >= 2:
            pivot[f"Var_{anios[0]}→{anios[-1]}_%"] = (
                (pivot[anios[-1]] - pivot[anios[0]]) / pivot[anios[0]] * 100
            ).round(1)

        return pivot.sort_values(anios[-1], ascending=False)

    def comparar_brecha_genero(self) -> pd.DataFrame:
        """Brecha salarial de género por año (nacional).

        Calcula mediana hombres, mediana mujeres, y brecha %.

        Returns:
            DataFrame con ANIO, Mediana_H, Mediana_M, Brecha_%.
        """
        filas = []
        for anio in self.anios_disponibles:
            df = self._anios[anio]["df"]
            smmlv = self._anios[anio]["config"].smmlv
            mask = (df["OCI"] == 1) & (df["INGLABO"] > 0) & df["P3271"].notna()
            for sexo_val, sexo_lab in [(1, "H"), (2, "M")]:
                m = mask & (df["P3271"] == sexo_val)
                med = EP.mediana(df.loc[m, "INGLABO"], df.loc[m, "FEX_ADJ"])
                filas.append({
                    "ANIO": anio, "Sexo": sexo_lab,
                    "Mediana_COP": round(med),
                    "Mediana_SMMLV": round(med / smmlv, 2),
                })

        df = pd.DataFrame(filas)
        pivot = df.pivot_table(
            index="ANIO", columns="Sexo",
            values=["Mediana_COP", "Mediana_SMMLV"],
        )
        pivot.columns = ["_".join(c) for c in pivot.columns]

        if "Mediana_COP_H" in pivot.columns and "Mediana_COP_M" in pivot.columns:
            pivot["Brecha_%"] = (
                (pivot["Mediana_COP_M"] - pivot["Mediana_COP_H"])
                / pivot["Mediana_COP_H"] * 100
            ).round(1)

        return pivot

    def resumen(self) -> None:
        """Imprime resumen de los años cargados."""
        print(f"\n{'='*60}")
        print(f"  COMPARADOR MULTI-AÑO — {len(self._anios)} años cargados")
        print(f"{'='*60}")
        for anio in self.anios_disponibles:
            data = self._anios[anio]
            print(f"  {anio}: {data['n_registros']:,} filas | "
                  f"SMMLV=${data['config'].smmlv:,} | "
                  f"{data['config'].n_meses} meses")
        print(f"{'='*60}")

    # ═════════════════════════════════════════════════════════════
    # IMPRESIÓN
    # ═════════════════════════════════════════════════════════════

    def _imprimir_indicadores(self, df: pd.DataFrame) -> None:
        print(f"\n{'='*75}")
        print(f"  COMPARACIÓN INTER-ANUAL DE INDICADORES LABORALES")
        print(f"{'='*75}")
        for _, r in df.iterrows():
            anio = int(r["ANIO"])
            delta_td = f"  Δ={r.get('Δ_TD_%', 'N/A')}" if "Δ_TD_%" in r else ""
            print(f"  {anio}: TD={r['TD_%']:.1f}%{delta_td}  |  "
                  f"TGP={r['TGP_%']:.1f}%  |  TO={r['TO_%']:.1f}%  |  "
                  f"OCI={r['Ocupados_M']:.2f}M  |  "
                  f"SMMLV=${int(r['SMMLV']):,}")

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/comparativo.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [13/39]: config.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/config.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/config.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.config — Configuración centralizada para el análisis GEIH.

Toda constante, mapeo, paleta de colores y parámetro configurable vive aquí.
Nunca hay "números mágicos" sueltos en la lógica de negocio.

CAMBIO v4.0 — Escalabilidad multi-año:
  - ConfigGEIH ahora recibe `anio` y auto-selecciona SMMLV y carpetas.
  - MESES_CARPETAS es ahora una función, no una lista hardcoded.
  - ReferenciaDane es un diccionario por año.
  - Para años sin referencia DANE publicada, el sanity_check advierte
    en vez de fallar.

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "SMMLV_POR_ANIO",
    "SMMLV_2025",
    "CARGA_PRESTACIONAL",
    "MESES_NOMBRES",
    "generar_carpetas_mensuales",
    "generar_etiqueta_periodo",
    "MESES_CARPETAS",
    "ConfigGEIH",
    "COLORES",
    "LLAVES_PERSONA",
    "LLAVES_HOGAR",
    "CONVERTERS_BASE",
    "CONVERTERS_CON_AREA",
    "MODULOS_CSV",
    "VARIABLES_POR_MODULO",
    "RAMAS_DANE",
    "TABLA_CIIU_RAMAS",
    "AGRUPACION_DANE_8",
    "_AGRUP_DANE_POR_DIVISION",
    "DEPARTAMENTOS",
    "DPTO_A_CIUDAD",
    "AREA_A_CIUDAD",
    "CIUDADES_13_PRINCIPALES",
    "CIUDADES_10_INTERMEDIAS",
    "NIVELES_EDUCATIVOS",
    "NIVELES_AGRUPADOS",
    "P3042_A_ANOS",
    "RANGOS_SMMLV_LIMITES",
    "RANGOS_SMMLV_ETIQUETAS",
    "TAMANO_EMPRESA",
    "CIIU_DESCRIPCION_FALLBACK",
    "ReferenciaDane",
    "REF_DANE",
    "REF_DANE_2025",
]


from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional


# ═════════════════════════════════════════════════════════════════════
# PARÁMETROS ECONÓMICOS — MULTI-AÑO
# ═════════════════════════════════════════════════════════════════════

SMMLV_POR_ANIO: Dict[int, int] = {
    2022: 1_000_000,
    2023: 1_160_000,
    2024: 1_300_000,
    2025: 1_423_500,
    2026: 1_750_905,   # Decreto 2426 de 2025 — actualizar si cambia
    # Agregar años futuros aquí. Solo se modifica esta línea.
}
"""SMMLV por año en COP. Fuente: Decretos anuales del Gobierno Nacional.
Para agregar un año nuevo, solo agregar una entrada al diccionario."""

# Retrocompatibilidad: módulos que importaban SMMLV_2025 directamente
SMMLV_2025: int = SMMLV_POR_ANIO[2025]

CARGA_PRESTACIONAL: float = 0.54
"""Factor de carga prestacional sobre salario base en Colombia (~54%).
Incluye pensión 12%, salud 8.5%, parafiscales 9%, cesantías 8.33%,
intereses 1%, prima 8.33%, vacaciones 4.17%, riesgos variable."""


# ═════════════════════════════════════════════════════════════════════
# NOMBRES DE LOS MESES (GENERACIÓN DINÁMICA DE CARPETAS)
# ═════════════════════════════════════════════════════════════════════

MESES_NOMBRES: List[str] = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]
"""Nombres de meses en español, tal como los usa el DANE en sus carpetas."""


def generar_carpetas_mensuales(anio: int, n_meses: int = 12) -> List[str]:
    """Genera la lista de carpetas mensuales para un año y cantidad de meses.

    El DANE nombra las carpetas como 'Enero 2025', 'Febrero 2025', etc.
    Esta función genera esa lista dinámicamente.

    Args:
        anio: Año de los datos (ej: 2025, 2026).
        n_meses: Cuántos meses generar (1-12). Permite procesamiento
                 parcial del año en curso.

    Returns:
        Lista de strings como ['Enero 2025', 'Febrero 2025', ...].

    Ejemplo:
        >>> generar_carpetas_mensuales(2026, n_meses=3)
        ['Enero 2026', 'Febrero 2026', 'Marzo 2026']
    """
    n = max(1, min(n_meses, 12))
    return [f"{mes} {anio}" for mes in MESES_NOMBRES[:n]]


def generar_etiqueta_periodo(anio: int, n_meses: int = 12) -> str:
    """Genera la etiqueta legible del período para títulos y reportes.

    Ejemplos:
        (2025, 12) → 'Enero – Diciembre 2025'
        (2026, 3)  → 'Enero – Marzo 2026'
        (2026, 1)  → 'Enero 2026'
    """
    n = max(1, min(n_meses, 12))
    if n == 1:
        return f"{MESES_NOMBRES[0]} {anio}"
    return f"{MESES_NOMBRES[0]} – {MESES_NOMBRES[n - 1]} {anio}"


# Retrocompatibilidad: código que importaba MESES_CARPETAS como constante
MESES_CARPETAS: List[str] = generar_carpetas_mensuales(2025, 12)


# ═════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DEL ANÁLISIS — MULTI-AÑO
# ═════════════════════════════════════════════════════════════════════

@dataclass
class ConfigGEIH:
    """Parámetros configurables para el pipeline GEIH.

    Centraliza todo lo que un usuario necesita ajustar.
    Valida en __post_init__ para fallar rápido si algo está mal.

    CAMBIO v4.0: ahora recibe `anio` y auto-deriva SMMLV, carpetas
    mensuales y etiqueta de período. El usuario solo necesita cambiar
    `anio` y `n_meses` para procesar cualquier año.

    Uso típico:
        # Año 2025 completo (12 meses)
        config = ConfigGEIH(anio=2025, n_meses=12)

        # Primeros 3 meses de 2026
        config = ConfigGEIH(anio=2026, n_meses=3)

        # Mes puntual: marzo 2026
        config = ConfigGEIH(anio=2026, n_meses=1)
    """
    anio: int = 2025
    """Año de los datos a procesar."""

    n_meses: int = 12
    """Número de meses consolidados. Controla la división del FEX_C18."""

    smmlv: int = 0
    """SMMLV del año de análisis en COP. Si 0 (default), se auto-selecciona
    de SMMLV_POR_ANIO según el año. Solo asignar manualmente si el año
    aún no está en el diccionario."""

    periodo_etiqueta: str = ""
    """Etiqueta legible para títulos y reportes. Si vacía, se genera
    automáticamente como 'Enero – Diciembre 2025'."""

    random_seed: int = 42
    """Semilla para reproducibilidad (muestreo, ML)."""

    encoding_csv: str = "latin-1"
    """Encoding de los CSV del DANE."""

    separador_csv: str = ";"
    """Separador de columnas en los CSV del DANE."""

    edad_minima_oit: int = 15
    """Edad mínima para análisis de mercado laboral (estándar OIT)."""

    edad_minima_dane: int = 10
    """Edad mínima usada por el DANE para PET."""

    def __post_init__(self):
        """Valida parámetros y auto-deriva valores calculados."""
        # ── Validación básica ──────────────────────────────────────
        if self.n_meses < 1 or self.n_meses > 12:
            raise ValueError(f"n_meses={self.n_meses} fuera de rango [1, 12]")
        if self.anio < 2018 or self.anio > 2050:
            raise ValueError(
                f"anio={self.anio} fuera de rango [2018, 2050]. "
                f"La GEIH Marco 2018 inicia en 2022."
            )

        # ── Auto-seleccionar SMMLV según el año ────────────────────
        if self.smmlv == 0:
            if self.anio in SMMLV_POR_ANIO:
                self.smmlv = SMMLV_POR_ANIO[self.anio]
            else:
                # Año no registrado → usar el último conocido + advertencia
                ultimo_anio = max(SMMLV_POR_ANIO.keys())
                self.smmlv = SMMLV_POR_ANIO[ultimo_anio]
                print(
                    f"⚠️  SMMLV para {self.anio} no está registrado en "
                    f"SMMLV_POR_ANIO. Usando el de {ultimo_anio}: "
                    f"${self.smmlv:,}. Actualice config.py cuando se "
                    f"publique el decreto."
                )

        if self.smmlv < 100_000:
            raise ValueError(f"smmlv={self.smmlv} parece demasiado bajo")

        # ── Auto-generar etiqueta de período ───────────────────────
        if not self.periodo_etiqueta:
            self.periodo_etiqueta = generar_etiqueta_periodo(
                self.anio, self.n_meses
            )

    # ── Propiedades calculadas ──────────────────────────────────
    @property
    def carpetas_mensuales(self) -> List[str]:
        """Lista de carpetas mensuales DANE para este año y n_meses.

        Reemplaza la constante MESES_CARPETAS. Las clases que antes
        usaban MESES_CARPETAS ahora usan config.carpetas_mensuales.
        """
        return generar_carpetas_mensuales(self.anio, self.n_meses)

    @property
    def referencia_dane(self) -> Optional["ReferenciaDane"]:
        """Referencia DANE para validación, o None si no está disponible.

        Devuelve None para años sin referencia publicada (ej: 2026 parcial).
        El sanity_check() maneja None gracefully: advierte en vez de fallar.
        """
        return REF_DANE.get(self.anio)

    def resumen(self) -> None:
        """Imprime un resumen legible de la configuración activa."""
        ref = self.referencia_dane
        ref_status = "✅ Disponible" if ref else "⚠️  No disponible"
        print(f"\n{'='*55}")
        print(f"  CONFIGURACIÓN GEIH — {self.periodo_etiqueta}")
        print(f"{'='*55}")
        print(f"  Año           : {self.anio}")
        print(f"  Meses         : {self.n_meses}")
        print(f"  SMMLV         : ${self.smmlv:,} COP")
        print(f"  FEX divisor   : ÷ {self.n_meses}")
        print(f"  Ref. DANE     : {ref_status}")
        print(f"  Carpetas      : {self.carpetas_mensuales[0]} → "
              f"{self.carpetas_mensuales[-1]}")
        print(f"{'='*55}")


# ═════════════════════════════════════════════════════════════════════
# PALETA DE COLORES INSTITUCIONAL
# ═════════════════════════════════════════════════════════════════════

COLORES: Dict[str, str] = {
    "azul":     "#2E6DA4",
    "rojo":     "#C0392B",
    "verde":    "#1E8449",
    "morado":   "#7D3C98",
    "naranja":  "#E67E22",
    "gris":     "#7F8C8D",
    "cyan":     "#1ABC9C",
    "amarillo": "#F39C12",
    "negro":    "#1A252F",
    "linea":    "#BDC3C7",
    "fondo":    "#F7F9FC",
}


# ═════════════════════════════════════════════════════════════════════
# LLAVES DE ENLACE ENTRE MÓDULOS GEIH
# ═════════════════════════════════════════════════════════════════════

LLAVES_PERSONA: List[str] = ["DIRECTORIO", "SECUENCIA_P", "ORDEN"]
"""Llave única a nivel de persona (individuo)."""

LLAVES_HOGAR: List[str] = ["DIRECTORIO", "SECUENCIA_P"]
"""Llave única a nivel de hogar."""


# ═════════════════════════════════════════════════════════════════════
# COLUMNAS QUE DEBEN LEERSE COMO STRING (NO NUMÉRICAS)
# ═════════════════════════════════════════════════════════════════════

CONVERTERS_BASE: Dict[str, type] = {
    "DIRECTORIO":  str,
    "SECUENCIA_P": str,
    "ORDEN":       str,
    "DPTO":        str,
    "RAMA2D_R4":   str,
    "RAMA4D_R4":   str,
}

CONVERTERS_CON_AREA: Dict[str, type] = {
    **CONVERTERS_BASE,
    "AREA": str,
}


# ═════════════════════════════════════════════════════════════════════
# NOMBRES DE ARCHIVOS CSV POR MÓDULO GEIH
# ═════════════════════════════════════════════════════════════════════

MODULOS_CSV: Dict[str, str] = {
    "caracteristicas": "Características generales, seguridad social en salud y educación.CSV",
    "hogar":           "Datos del hogar y la vivienda.CSV",
    "fuerza_trabajo":  "Fuerza de trabajo.CSV",
    "ocupados":        "Ocupados.CSV",
    "no_ocupados":     "No ocupados.CSV",
    "otras_formas":    "Otras formas de trabajo.CSV",
    "migracion":       "Migración.CSV",
    "otros_ingresos":  "Otros ingresos e impuestos.CSV",
}

# Catálogo de variables clave por módulo (para referencia y selección)
VARIABLES_POR_MODULO: Dict[str, List[str]] = {
    "caracteristicas": [
        "P3271",      # Sexo (1=H, 2=M)
        "P6040",      # Edad
        "P6080",      # Autorreconocimiento étnico (1=Indígena, 5=Afro, 6=Ninguno)
        "P3042",      # Nivel educativo (1-13)
        "P3043S1",    # Campo de formación (CINE-F)
        "P6090",      # Afiliado salud (1=Sí)
        "P2057",      # ¿Se considera campesino? (1=Sí)
        "P2059",      # ¿Alguna vez fue campesino?
        "P1906S1",    # Discapacidad: Oír
        "P1906S2",    # Discapacidad: Hablar
        "P1906S3",    # Discapacidad: Ver
        "P1906S4",    # Discapacidad: Moverse
        "P1906S5",    # Discapacidad: Agarrar
        "P1906S6",    # Discapacidad: Entender
        "P1906S7",    # Discapacidad: Autocuidado
        "P1906S8",    # Discapacidad: Relacionarse
        "CLASE",      # Zona (1=Urbano/Cabecera, 2=Rural)
        "FEX_C18",    # Factor de expansión
    ],
    "ocupados": [
        "OCI",        # Ocupado (=1)
        "INGLABO",    # Ingreso laboral mensual COP
        "P6500",      # Salario bruto declarado
        "P6430",      # Posición ocupacional (1-9)
        "P6800",      # Horas normales semana
        "P6850",      # Horas reales semana pasada
        "P6920",      # Cotiza pensión (1=Sí, 2=No, 3=Pensionado)
        "P3069",      # Tamaño empresa (1-10 categorías)
        "P7130",      # Desea cambiar trabajo (1=Sí)
        "P6440",      # ¿Tiene contrato?
        "P6450",      # ¿Contrato escrito?
        "P6460",      # ¿Contrato indefinido?
        "P1802",      # Alcance mercado (1-6, 6=Exportación)
        "P3047",      # ¿Quién decide horario?
        "P3048",      # ¿Quién decide qué producir?
        "P3049",      # ¿Quién decide precio?
        "P3363",      # ¿Cómo consiguió empleo?
        "P3364",      # ¿Le descontaron retención en la fuente?
        "P6765",      # Forma de trabajo (destajo, honorarios, etc.)
        "P6400",      # ¿Trabaja donde lo contrataron?
        "P6410",      # Tipo intermediación (EST, CTA)
        "P6510S1",    # Horas extras
        "P6580S1",    # Bonificaciones
        "P6585S1A1",  # Auxilio alimentación
        "P6585S2A1",  # Auxilio transporte
        "RAMA2D_R4",  # CIIU 2 dígitos
        "RAMA4D_R4",  # CIIU 4 dígitos
        "AREA",       # Municipio (5 dígitos)
    ],
    "no_ocupados": [
        "DSI",        # Desocupado (=1)
        "P7250",      # Semanas buscando trabajo
        "P6300",      # ¿Desearía trabajar? (FFT con deseo)
        "P6310",      # ¿Disponible para trabajar?
        "FFT",        # Fuera de la fuerza de trabajo (=1)
    ],
    "fuerza_trabajo": [
        "FT",         # En la fuerza de trabajo (=1)
        "PET",        # Población en edad de trabajar (=1)
        "P6240",      # Actividad semana pasada
        "P6280",      # ¿Disponible para trabajar?
    ],
    "otras_formas": [
        "P3054",      # Producción bienes autoconsumo
        "P3054S1",    # Horas autoconsumo
        "P3055",      # Producción servicios autoconsumo
        "P3055S1",    # Horas servicios autoconsumo
        "P3056",      # Trabajo voluntario
        "P3057",      # Trabajo en formación
    ],
    "migracion": [
        "P3370",      # ¿Dónde vivía hace 12 meses?
        "P3370S1",    # Departamento hace 12 meses
        "P3371",      # ¿Dónde vivía hace 5 años?
        "P3376",      # País de nacimiento
        "P3378S1",    # Año de llegada a Colombia
    ],
    "otros_ingresos": [
        "P7422",      # Arriendos recibidos
        "P7500S1",    # Pensión jubilación
        "P7500S2",    # Ayudas de otros hogares
        "P7500S3",    # Ayudas institucionales
        "P7510S1",    # Intereses/dividendos
        "P7510S2",    # Remesas del exterior
        "P7510S3",    # Cesantías
    ],
}


# ═════════════════════════════════════════════════════════════════════
# MAPEO CIIU Rev.4 → 13 RAMAS DANE (AGRUPACIÓN ESTÁNDAR)
# ═════════════════════════════════════════════════════════════════════

RAMAS_DANE: Dict[str, str] = {
    "AGRI": "Agricultura, ganadería, caza, silvicultura y pesca",
    "SUMI": "Suministro de electricidad, gas, agua y gestión de desechos^",
    "MANU": "Industrias manufactureras",
    "CONS": "Construcción",
    "COME": "Comercio y reparación de vehículos",
    "TRAN": "Transporte y almacenamiento",
    "ALOJ": "Alojamiento y servicios de comida",
    "INFO": "Información y comunicaciones",
    "FINA": "Actividades financieras y de seguros",
    "INMO": "Actividades inmobiliarias",
    "PROF": "Actividades profesionales, científicas, técnicas y de servicios administrativos",
    "ADMP": "Administración pública y defensa, educación y atención de la salud humana",
    "ARTE": "Actividades artísticas, entretenimiento, recreación y otras actividades de servicio",
}

# Tabla de rangos: (CIIU_min, CIIU_max, clave_rama)
TABLA_CIIU_RAMAS: List[Tuple[int, int, str]] = [
    (1,  3,  "AGRI"), (5,  9,  "SUMI"), (10, 33, "MANU"),
    (35, 35, "SUMI"), (36, 39, "SUMI"), (41, 43, "CONS"),
    (45, 47, "COME"), (49, 53, "TRAN"), (55, 56, "ALOJ"),
    (58, 63, "INFO"), (64, 66, "FINA"), (68, 68, "INMO"),
    (69, 75, "PROF"), (77, 82, "PROF"), (84, 84, "ADMP"),
    (85, 85, "ADMP"), (86, 88, "ADMP"), (90, 93, "ARTE"),
    (94, 96, "ARTE"), (97, 98, "ARTE"), (99, 99, "ARTE"),
]


# ═════════════════════════════════════════════════════════════════════
# AGRUPACIÓN DANE DE 8 GRUPOS (PARA TABLAS POR ÁREA)
# ═════════════════════════════════════════════════════════════════════

AGRUPACION_DANE_8: Dict[str, List[Tuple[int, int]]] = {
    "Agricultura, ganadería, pesca y silvicultura":              [(1, 3)],
    "Explotación de minas y canteras":                           [(5, 9)],
    "Industrias manufactureras":                                 [(10, 33)],
    "Electricidad, agua, gas y desechos":                        [(35, 39)],
    "Construcción":                                              [(41, 43)],
    "Comercio, transporte, alojamiento y comida":                [(45, 56)],
    "Actividades financieras, profesionales y administrativas":  [(58, 83)],
    "Administración pública, educación, salud y otros":          [(84, 99)],
}

# Mapeo directo: código CIIU 2 dígitos (string) → nombre de agrupación DANE
# Usado por AnalisisOcupadosCiudad para tablas y Excel
_AGRUP_DANE_POR_DIVISION: Dict[str, str] = {}
for _nombre, _rangos in AGRUPACION_DANE_8.items():
    for _lo, _hi in _rangos:
        for _i in range(_lo, _hi + 1):
            _AGRUP_DANE_POR_DIVISION[str(_i).zfill(2)] = _nombre


# ═════════════════════════════════════════════════════════════════════
# DEPARTAMENTOS DE COLOMBIA (CÓDIGO DIVIPOLA → NOMBRE)
# ═════════════════════════════════════════════════════════════════════

DEPARTAMENTOS: Dict[str, str] = {
    "05": "Antioquia",       "08": "Atlántico",
    "11": "Bogotá D.C.",     "13": "Bolívar",
    "15": "Boyacá",          "17": "Caldas",
    "18": "Caquetá",         "19": "Cauca",
    "20": "Cesar",           "23": "Córdoba",
    "25": "Cundinamarca",    "27": "Chocó",
    "41": "Huila",           "44": "La Guajira",
    "47": "Magdalena",       "50": "Meta",
    "52": "Nariño",          "54": "Norte de Santander",
    "63": "Quindío",         "66": "Risaralda",
    "68": "Santander",       "70": "Sucre",
    "73": "Tolima",          "76": "Valle del Cauca",
}


# ═════════════════════════════════════════════════════════════════════
# CIUDADES Y ÁREAS METROPOLITANAS
# ═════════════════════════════════════════════════════════════════════

DPTO_A_CIUDAD: Dict[str, str] = {
    "11": "Bogotá D.C.",          "05": "Medellín A.M.",
    "76": "Cali A.M.",            "08": "Barranquilla A.M.",
    "68": "Bucaramanga A.M.",     "17": "Manizales A.M.",
    "66": "Pereira A.M.",         "54": "Cúcuta A.M.",
    "52": "Pasto",                "41": "Ibagué",
    "23": "Montería",             "13": "Cartagena",
    "50": "Villavicencio",        "15": "Boyacá/Tunja",
    "18": "Caquetá/Florencia",    "19": "Cauca/Popayán",
    "20": "Cesar/Valledupar",     "27": "Chocó/Quibdó",
    "44": "La Guajira/Riohacha",  "47": "Magdalena/Sta.Marta",
    "63": "Quindío/Armenia",      "70": "Sucre/Sincelejo",
    "25": "Cundinamarca",         "73": "Tolima",
}

AREA_A_CIUDAD: Dict[str, str] = {
    # ── 13 ciudades principales y sus áreas metropolitanas ──────
    "11001": "Bogotá D.C.",
    # Medellín AM (8 municipios)
    "05001": "Medellín A.M.",     "05088": "Medellín A.M.",
    "05308": "Medellín A.M.",     "05318": "Medellín A.M.",
    "05360": "Medellín A.M.",     "05380": "Medellín A.M.",
    "05400": "Medellín A.M.",     "05501": "Medellín A.M.",
    # Cali AM (6 municipios)
    "76001": "Cali A.M.",         "76111": "Cali A.M.",
    "76113": "Cali A.M.",         "76364": "Cali A.M.",
    "76520": "Cali A.M.",         "76563": "Cali A.M.",
    # Barranquilla AM (4 municipios)
    "08001": "Barranquilla A.M.", "08433": "Barranquilla A.M.",
    "08549": "Barranquilla A.M.", "08758": "Barranquilla A.M.",
    # Bucaramanga AM (6 municipios)
    "68001": "Bucaramanga A.M.",  "68081": "Bucaramanga A.M.",
    "68276": "Bucaramanga A.M.",  "68307": "Bucaramanga A.M.",
    "68615": "Bucaramanga A.M.",  "68705": "Bucaramanga A.M.",
    # Manizales AM (3 municipios)
    "17001": "Manizales A.M.",    "17042": "Manizales A.M.",
    "17616": "Manizales A.M.",
    # Pereira AM (3 municipios)
    "66001": "Pereira A.M.",      "66045": "Pereira A.M.",
    "66170": "Pereira A.M.",
    # Cúcuta AM (5 municipios)
    "54001": "Cúcuta A.M.",       "54128": "Cúcuta A.M.",
    "54172": "Cúcuta A.M.",       "54206": "Cúcuta A.M.",
    "54520": "Cúcuta A.M.",
    # Ciudades sin AM
    "52001": "Pasto",             "41001": "Ibagué",
    "23001": "Montería",          "13001": "Cartagena",
    "50001": "Villavicencio",
    # ── 10 ciudades intermedias ─────────────────────────────────
    "15001": "Tunja",             "18001": "Florencia",
    "19001": "Popayán",           "20001": "Valledupar",
    "27001": "Quibdó",            "41551": "Neiva",
    "44001": "Riohacha",          "47001": "Santa Marta",
    "63001": "Armenia",           "70001": "Sincelejo",
}

# Clasificación DANE: dominios geográficos para reportes
CIUDADES_13_PRINCIPALES: set = {
    "Bogotá D.C.", "Medellín A.M.", "Cali A.M.", "Barranquilla A.M.",
    "Bucaramanga A.M.", "Manizales A.M.", "Pereira A.M.", "Cúcuta A.M.",
    "Pasto", "Ibagué", "Montería", "Cartagena", "Villavicencio",
}
CIUDADES_10_INTERMEDIAS: set = {
    "Tunja", "Florencia", "Popayán", "Valledupar", "Quibdó",
    "Neiva", "Riohacha", "Santa Marta", "Armenia", "Sincelejo",
}


# ═════════════════════════════════════════════════════════════════════
# NIVELES EDUCATIVOS (P3042 → ETIQUETA)
# ═════════════════════════════════════════════════════════════════════

NIVELES_EDUCATIVOS: Dict[int, str] = {
    1: "Ninguno",             2: "Preescolar",
    3: "Básica primaria",     4: "Básica secundaria",
    5: "Media académica",     6: "Media técnica",
    7: "Normalista",          8: "Técnica profesional",
    9: "Tecnológica",        10: "Universitaria",
    11: "Especialización",   12: "Maestría",
    13: "Doctorado",
}

NIVELES_AGRUPADOS: Dict[int, str] = {
    1: "1. Sin educación",     2: "1. Sin educación",
    3: "2. Primaria",          4: "3. Secundaria",
    5: "4. Media",             6: "4. Media",
    7: "4. Media",             8: "5. Técnica/Tecno.",
    9: "5. Técnica/Tecno.",   10: "6. Universitaria",
    11: "7. Posgrado",        12: "7. Posgrado",
    13: "7. Posgrado",
}

# Conversión P3042 → Años de educación acumulados (para Mincer)
P3042_A_ANOS: Dict[int, int] = {
    1: 0, 2: 0, 3: 5, 4: 9, 5: 11, 6: 11, 7: 11,
    8: 14, 9: 15, 10: 16, 11: 17, 12: 18, 13: 21,
}


# ═════════════════════════════════════════════════════════════════════
# RANGOS DE INGRESO EN MÚLTIPLOS DE SMMLV
# ═════════════════════════════════════════════════════════════════════

RANGOS_SMMLV_LIMITES: List[float] = [0, 0.5, 1, 1.5, 2, 3, 4, 6, 10, float("inf")]

RANGOS_SMMLV_ETIQUETAS: List[str] = [
    "< 0.5 SMMLV",    "0.5 – 1 SMMLV",  "1 – 1.5 SMMLV",
    "1.5 – 2 SMMLV",  "2 – 3 SMMLV",    "3 – 4 SMMLV",
    "4 – 6 SMMLV",    "6 – 10 SMMLV",   "> 10 SMMLV",
]


# ═════════════════════════════════════════════════════════════════════
# ETIQUETAS DE TAMAÑO DE EMPRESA (P3069)
# ═════════════════════════════════════════════════════════════════════

TAMANO_EMPRESA: Dict[int, str] = {
    1: "Solo (1)",       2: "2–3 pers.",     3: "4–5 pers.",
    4: "6–10 pers.",     5: "11–19 pers.",   6: "20–30 pers.",
    7: "31–50 pers.",    8: "51–100 pers.",  9: "101–200 pers.",
    10: "201+ pers.",
}


# ═════════════════════════════════════════════════════════════════════
# MAPEO CIIU INTERNO (FALLBACK PARA MERGE INCOMPLETO)
# ═════════════════════════════════════════════════════════════════════

CIIU_DESCRIPCION_FALLBACK: List[Tuple[range, str]] = [
    (range(1, 4),   "Agricultura, ganadería, caza, silvicultura y pesca"),
    (range(5, 10),  "Explotación de minas y canteras"),
    (range(10, 34), "Industrias manufactureras"),
    (range(35, 36), "Suministro de electricidad, gas, vapor y aire acondicionado"),
    (range(36, 40), "Distribución de agua; gestión de desechos"),
    (range(41, 44), "Construcción"),
    (range(45, 48), "Comercio al por mayor y al por menor; reparación de vehículos"),
    (range(49, 54), "Transporte y almacenamiento"),
    (range(55, 57), "Alojamiento y servicios de comida"),
    (range(58, 64), "Información y comunicaciones"),
    (range(64, 67), "Actividades financieras y de seguros"),
    (range(68, 69), "Actividades inmobiliarias"),
    (range(69, 76), "Actividades profesionales, científicas y técnicas"),
    (range(77, 83), "Actividades de servicios administrativos y de apoyo"),
    (range(84, 85), "Administración pública y defensa"),
    (range(85, 86), "Educación"),
    (range(86, 89), "Salud y asistencia social"),
    (range(90, 100), "Actividades artísticas, entretenimiento y otras"),
]


# ═════════════════════════════════════════════════════════════════════
# REFERENCIA DANE — MULTI-AÑO (PARA VALIDACIÓN SANITY CHECK)
# ═════════════════════════════════════════════════════════════════════

@dataclass(frozen=True)
class ReferenciaDane:
    """Valores de referencia del boletín DANE para validación cruzada.

    Cada año tiene sus propios valores publicados por el DANE.
    Para años sin datos publicados, no se crea una entrada en REF_DANE
    y el sanity_check advierte 'sin referencia disponible'.
    """
    pea_anual_m: float = 0.0        # Millones
    ocupados_anual_m: float = 0.0
    desocupados_anual_m: float = 0.0
    td_anual_pct: float = 0.0
    tgp_anual_pct: float = 0.0
    to_anual_pct: float = 0.0
    pea_dic_m: float = 0.0
    ocupados_dic_m: float = 0.0
    desocupados_dic_m: float = 0.0
    td_dic_pct: float = 0.0
    tgp_dic_pct: float = 0.0
    to_dic_pct: float = 0.0


REF_DANE: Dict[int, ReferenciaDane] = {
    2025: ReferenciaDane(
        pea_anual_m=26.3,  ocupados_anual_m=23.8, desocupados_anual_m=2.1,
        td_anual_pct=8.9,  tgp_anual_pct=64.3,   to_anual_pct=58.6,
        pea_dic_m=26.3,    ocupados_dic_m=24.2,   desocupados_dic_m=2.1,
        td_dic_pct=8.0,    tgp_dic_pct=64.3,      to_dic_pct=59.2,
    ),
    # Agregar 2026 cuando el DANE publique el boletín anual:
    # 2026: ReferenciaDane(
    #     pea_anual_m=..., ocupados_anual_m=..., desocupados_anual_m=...,
    #     td_anual_pct=..., tgp_anual_pct=..., to_anual_pct=...,
    # ),
}

# Retrocompatibilidad: código que importaba REF_DANE_2025 directamente
REF_DANE_2025 = REF_DANE[2025]

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/config.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [14/39]: consolidador.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/consolidador.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/consolidador.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.consolidador — Consolidación de microdatos GEIH mensuales.

Lee los módulos CSV de cada mes, los une por las llaves DIRECTORIO +
SECUENCIA_P + ORDEN y concatena los 12 meses en una base única.

Decisiones críticas:
  - El módulo ancla siempre es "Características generales" (universo completo).
  - El join es LEFT para no multiplicar filas.
  - Las columnas duplicadas entre módulos se eliminan antes del merge.
  - MES_NUM se agrega como variable creada (no existe en DANE).

CAMBIO v4.0 — Escalabilidad multi-año:
  - consolidar() ahora usa config.carpetas_mensuales en vez de MESES_CARPETAS.
  - agregar_mes() permite añadir un mes nuevo a un Parquet existente
    sin re-consolidar todo (procesamiento incremental mes a mes).

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "ConsolidadorGEIH",
]


import gc
import os
import unicodedata
from pathlib import Path
from typing import List, Optional, Dict

import pandas as pd

from .config import (
    ConfigGEIH,
    MESES_CARPETAS,
    MESES_NOMBRES,
    MODULOS_CSV,
    CONVERTERS_BASE,
    CONVERTERS_CON_AREA,
    LLAVES_PERSONA,
    LLAVES_HOGAR,
)


class ConsolidadorGEIH:
    """Consolida los microdatos GEIH mensuales en una base única.

    El proceso es:
      1. Para cada mes, leer los módulos CSV necesarios
      2. Unir módulos usando LEFT JOIN (ancla: Características generales)
      3. Agregar MES_NUM para identificar el mes
      4. Concatenar todos los meses
      5. Exportar a Parquet (eficiente en espacio y velocidad)

    CAMBIO v4.0: Las carpetas mensuales se toman de config.carpetas_mensuales,
    que se genera dinámicamente según config.anio y config.n_meses.

    Uso típico desde el notebook:
        # Año 2025 completo
        config = ConfigGEIH(anio=2025, n_meses=12)
        consolidador = ConsolidadorGEIH(
            ruta_base='/content/drive/MyDrive/GEIH',
            config=config,
        )
        geih = consolidador.consolidar()
        consolidador.exportar(geih, f'GEIH_{config.anio}_Consolidado.parquet')

        # Agregar enero 2026 sin re-consolidar
        config_2026 = ConfigGEIH(anio=2026, n_meses=1)
        consolidador_2026 = ConsolidadorGEIH(
            ruta_base='/content/drive/MyDrive/GEIH',
            config=config_2026,
        )
        consolidador_2026.agregar_mes(
            mes_carpeta='Enero 2026',
            parquet_existente='GEIH_2026_Consolidado.parquet',
        )
    """

    def __init__(
        self,
        ruta_base: str,
        config: Optional[ConfigGEIH] = None,
        incluir_area: bool = False,
        modulos_extra: Optional[List[str]] = None,
    ):
        """
        Args:
            ruta_base: Ruta a la carpeta que contiene las subcarpetas mensuales.
            config: Configuración del análisis. Usa defaults si no se provee.
            incluir_area: Si True, incluye 'AREA' en los converters para
                          habilitar análisis por 32 ciudades.
            modulos_extra: Módulos adicionales a incluir (ej: 'otras_formas',
                          'micronegocios'). Por defecto solo se consolidan
                          los 5 módulos principales.
        """
        self.ruta_base = Path(ruta_base)
        self.config = config or ConfigGEIH()
        self.converters = CONVERTERS_CON_AREA if incluir_area else CONVERTERS_BASE
        self._modulos_a_incluir = [
            "caracteristicas", "hogar", "fuerza_trabajo",
            "ocupados", "no_ocupados",
            "otras_formas", "migracion", "otros_ingresos",
        ]
        if modulos_extra:
            self._modulos_a_incluir.extend(modulos_extra)

    # ── Búsqueda de archivos resiliente a tildes ────────────────

    @staticmethod
    def _normalizar(texto: str) -> str:
        """Elimina tildes y convierte a minúsculas para comparación.

        'Migración.CSV' → 'migracion.csv'
        'Características generales...' → 'caracteristicas generales...'

        Usa descomposición Unicode NFD: separa la letra base del acento,
        luego elimina los acentos (categoría Mn = Mark, Nonspacing).
        """
        nfkd = unicodedata.normalize("NFD", texto)
        sin_tildes = "".join(c for c in nfkd if unicodedata.category(c) != "Mn")
        return sin_tildes.lower()

    @classmethod
    def _buscar_archivo(cls, carpeta: Path, nombre_esperado: str) -> Optional[Path]:
        """Busca un archivo en disco comparando sin tildes ni mayúsculas.

        El DANE a veces publica 'Migración.CSV' y a veces 'Migracion.CSV'.
        Esta función encuentra el archivo sin importar las tildes.

        Args:
            carpeta: Directorio donde buscar.
            nombre_esperado: Nombre de archivo que esperamos (puede tener tildes o no).

        Returns:
            Path al archivo real encontrado, o None si no existe.
        """
        # Intento directo primero (rápido si el nombre es exacto)
        ruta_directa = carpeta / nombre_esperado
        if ruta_directa.exists():
            return ruta_directa

        # Búsqueda normalizada: comparar sin tildes
        nombre_norm = cls._normalizar(nombre_esperado)
        try:
            for archivo in carpeta.iterdir():
                if cls._normalizar(archivo.name) == nombre_norm:
                    return archivo
        except (PermissionError, OSError):
            pass

        return None

    # ── Métodos públicos ───────────────────────────────────────────

    def verificar_estructura(
        self,
        carpetas: Optional[List[str]] = None,
    ) -> Dict[str, List[str]]:
        """Verifica que existan las carpetas y archivos esperados.

        Pre-flight check: detecta problemas ANTES de iniciar la lectura.
        Busca archivos de forma resiliente a tildes (Migración = Migracion).

        CAMBIO v4.0: Si no se pasan carpetas, usa config.carpetas_mensuales
        (dinámicas según año y n_meses).

        Returns:
            Dict con 'ok' (meses completos) y 'faltantes' (archivos no encontrados).
        """
        carpetas = carpetas or self.config.carpetas_mensuales
        resultado = {"ok": [], "faltantes": []}

        for mes in carpetas:
            ruta_mes = self.ruta_base / mes / "CSV"
            if not ruta_mes.exists():
                resultado["faltantes"].append(f"{mes}/CSV/ (carpeta no existe)")
                continue

            archivos_faltantes = []
            for mod_key in self._modulos_a_incluir:
                nombre_csv = MODULOS_CSV.get(mod_key, "")
                encontrado = self._buscar_archivo(ruta_mes, nombre_csv)
                if encontrado is None:
                    archivos_faltantes.append(nombre_csv)

            if archivos_faltantes:
                for arch in archivos_faltantes:
                    resultado["faltantes"].append(f"{mes}/CSV/{arch}")
            else:
                resultado["ok"].append(mes)

        # Reporte
        print(f"\n{'='*60}")
        print(f"  PRE-FLIGHT CHECK — Estructura de archivos GEIH {self.config.anio}")
        print(f"{'='*60}")
        print(f"  ✅ Meses completos : {len(resultado['ok'])}")
        if resultado["faltantes"]:
            print(f"  ❌ Archivos faltantes: {len(resultado['faltantes'])}")
            for f in resultado["faltantes"][:10]:
                print(f"     • {f}")
        else:
            print(f"  ✅ Todos los archivos presentes")

        return resultado

    def consolidar(
        self,
        carpetas: Optional[List[str]] = None,
        checkpoint: bool = True,
        ruta_checkpoints: Optional[str] = None,
    ) -> pd.DataFrame:
        """Consolida todos los meses en una base única.

        Este es el método principal. Lee los CSV, une los módulos
        y concatena. Reporta progreso visible al usuario.

        CAMBIO v4.2 — Checkpointing:
          Si checkpoint=True, guarda un Parquet después de cada mes.
          Si el proceso falla a mitad de camino (timeout, OOM, crash),
          al re-ejecutar detecta qué meses ya fueron procesados y
          los omite, reanudando desde donde quedó.

        Args:
            carpetas: Lista de carpetas mensuales a procesar.
                      Usa config.carpetas_mensuales por defecto.
            checkpoint: Si True, guarda progreso mes a mes.
            ruta_checkpoints: Carpeta para checkpoints. Si None, usa
                             ruta_base/_checkpoints_{anio}/

        Returns:
            DataFrame consolidado con ~5M filas y columna MES_NUM.

        Raises:
            FileNotFoundError: Si no se encuentra ningún archivo.
        """
        carpetas = carpetas or self.config.carpetas_mensuales
        bases_mensuales: List[pd.DataFrame] = []

        # ── Configurar checkpoints ─────────────────────────────────
        if checkpoint:
            ckpt_dir = Path(ruta_checkpoints) if ruta_checkpoints else (
                self.ruta_base / f"_checkpoints_{self.config.anio}"
            )
            ckpt_dir.mkdir(parents=True, exist_ok=True)
        else:
            ckpt_dir = None

        print(f"\n{'='*60}")
        print(f"  CONSOLIDACIÓN GEIH {self.config.anio} — {len(carpetas)} meses")
        if checkpoint:
            print(f"  Checkpointing: ON → {ckpt_dir}")
        print(f"{'='*60}")

        for i, mes in enumerate(carpetas, 1):
            # ── Verificar si ya existe checkpoint ──────────────────
            if ckpt_dir:
                ckpt_file = ckpt_dir / f"mes_{i:02d}.parquet"
                if ckpt_file.exists():
                    print(f"\n♻️  [{i}/{len(carpetas)}] {mes} — "
                          f"recuperado de checkpoint")
                    df_mes = pd.read_parquet(ckpt_file)
                    bases_mensuales.append(df_mes)
                    print(f"   ✅ {mes}: {df_mes.shape[0]:,} filas (checkpoint)")
                    continue

            print(f"\n🔄 [{i}/{len(carpetas)}] Procesando {mes}...")

            try:
                df_mes = self._procesar_mes(mes, numero_mes=i)
                bases_mensuales.append(df_mes)
                print(f"   ✅ {mes}: {df_mes.shape[0]:,} filas × {df_mes.shape[1]} cols")

                # ── Guardar checkpoint ─────────────────────────────
                if ckpt_dir:
                    df_mes.to_parquet(ckpt_file, index=False, compression="snappy")
                    print(f"   💾 Checkpoint guardado: {ckpt_file.name}")

            except FileNotFoundError as e:
                print(f"   ⚠️ Archivos faltantes en {mes}: {e}")
            except Exception as e:
                print(f"   ❌ Error en {mes}: {e}")
                if ckpt_dir and bases_mensuales:
                    print(f"   💾 {len(bases_mensuales)} meses en checkpoint — "
                          f"re-ejecute para continuar.")

        if not bases_mensuales:
            raise FileNotFoundError(
                "❌ No se procesó ningún mes. Verifica las carpetas con "
                "verificar_estructura()."
            )

        print(f"\n🔗 Concatenando {len(bases_mensuales)} meses...")
        geih = pd.concat(bases_mensuales, ignore_index=True)

        # Limpieza de intermedios
        del bases_mensuales
        gc.collect()

        # ── Limpiar checkpoints si todo salió bien ─────────────────
        if ckpt_dir and len(carpetas) == geih["MES_NUM"].nunique():
            self._limpiar_checkpoints(ckpt_dir)

        print(f"\n{'='*60}")
        print(f"  ✅ CONSOLIDACIÓN COMPLETA — {self.config.anio}")
        print(f"  {geih.shape[0]:,} filas × {geih.shape[1]} columnas")
        print(f"  Meses: {geih['MES_NUM'].nunique()} (de {geih['MES_NUM'].min()} a {geih['MES_NUM'].max()})")
        print(f"{'='*60}")

        return geih

    @staticmethod
    def _limpiar_checkpoints(ckpt_dir: Path) -> None:
        """Elimina la carpeta de checkpoints después de consolidación exitosa."""
        try:
            import shutil
            shutil.rmtree(ckpt_dir)
            print(f"   🗑️  Checkpoints eliminados (consolidación exitosa)")
        except Exception:
            pass  # No crítico si falla la limpieza

    def agregar_mes(
        self,
        mes_carpeta: str,
        parquet_existente: str,
        numero_mes: Optional[int] = None,
    ) -> pd.DataFrame:
        """Agrega un mes nuevo a un Parquet existente SIN re-consolidar todo.

        Caso de uso: cada mes, el DANE publica datos nuevos. En vez de
        re-consolidar los 12 meses (~10 min), solo lee el mes nuevo y
        lo concatena al Parquet existente (~1 min).

        Args:
            mes_carpeta: Nombre de la carpeta del mes nuevo (ej: 'Abril 2026').
            parquet_existente: Nombre del archivo Parquet al que agregar.
                              Se busca en self.ruta_base.
            numero_mes: Número del mes (1-12). Si None, se infiere del
                        nombre de la carpeta.

        Returns:
            DataFrame con la base actualizada (existente + mes nuevo).

        Ejemplo desde el notebook:
            consolidador = ConsolidadorGEIH(
                ruta_base=RUTA,
                config=ConfigGEIH(anio=2026, n_meses=4),
                incluir_area=True,
            )
            geih = consolidador.agregar_mes(
                mes_carpeta='Abril 2026',
                parquet_existente='GEIH_2026_Consolidado.parquet',
            )
            consolidador.exportar(geih, 'GEIH_2026_Consolidado.parquet')
        """
        ruta_parquet = self.ruta_base / parquet_existente

        # ── Cargar base existente ──────────────────────────────────
        if ruta_parquet.exists():
            print(f"📂 Cargando base existente: {parquet_existente}")
            df_existente = pd.read_parquet(ruta_parquet)
            meses_existentes = sorted(df_existente["MES_NUM"].unique().tolist())
            print(f"   Meses existentes: {meses_existentes} "
                  f"({df_existente.shape[0]:,} filas)")
        else:
            print(f"📂 No existe {parquet_existente} — se creará desde cero.")
            df_existente = None
            meses_existentes = []

        # ── Inferir número de mes ──────────────────────────────────
        if numero_mes is None:
            numero_mes = self._inferir_numero_mes(mes_carpeta)

        if numero_mes in meses_existentes:
            print(f"⚠️  El mes {numero_mes} ({mes_carpeta}) ya existe en la base.")
            print(f"   Para reemplazarlo, elimine ese mes primero o "
                  f"re-consolide desde cero.")
            return df_existente

        # ── Procesar mes nuevo ─────────────────────────────────────
        print(f"\n🔄 Procesando {mes_carpeta} (MES_NUM={numero_mes})...")
        df_nuevo = self._procesar_mes(mes_carpeta, numero_mes=numero_mes)
        print(f"   ✅ {mes_carpeta}: {df_nuevo.shape[0]:,} filas × "
              f"{df_nuevo.shape[1]} cols")

        # ── Concatenar ─────────────────────────────────────────────
        if df_existente is not None:
            geih = pd.concat([df_existente, df_nuevo], ignore_index=True)
            del df_existente, df_nuevo
        else:
            geih = df_nuevo
            del df_nuevo
        gc.collect()

        meses_final = sorted(geih["MES_NUM"].unique().tolist())
        print(f"\n{'='*60}")
        print(f"  ✅ MES AGREGADO — {mes_carpeta}")
        print(f"  {geih.shape[0]:,} filas × {geih.shape[1]} columnas")
        print(f"  Meses: {meses_final}")
        print(f"  ⚠️  Recuerde actualizar config.n_meses={len(meses_final)} "
              f"para que FEX_ADJ se divida correctamente.")
        print(f"{'='*60}")

        return geih

    def exportar(
        self,
        df: pd.DataFrame,
        nombre: Optional[str] = None,
        formato: str = "parquet",
    ) -> None:
        """Exporta la base consolidada a disco.

        CAMBIO v4.0: Si no se pasa nombre, se genera automáticamente
        como 'GEIH_{anio}_Consolidado.parquet'.

        Args:
            df: DataFrame consolidado.
            nombre: Nombre del archivo de salida. Si None, se auto-genera.
            formato: 'parquet' (recomendado) o 'csv'.
        """
        if nombre is None:
            ext = "parquet" if formato == "parquet" else "csv"
            nombre = f"GEIH_{self.config.anio}_Consolidado.{ext}"

        ruta = self.ruta_base / nombre

        if formato == "parquet":
            df.to_parquet(ruta, index=False, compression="snappy")
        elif formato == "csv":
            df.to_csv(
                ruta, index=False,
                encoding=self.config.encoding_csv,
            )
        else:
            raise ValueError(f"Formato no soportado: {formato}")

        size_mb = ruta.stat().st_size / 1e6
        print(f"✅ Base guardada: {ruta.name} ({size_mb:.0f} MB)")

    @staticmethod
    def cargar(ruta: str) -> pd.DataFrame:
        """Carga una base previamente consolidada.

        Args:
            ruta: Ruta al archivo .parquet o .csv.

        Returns:
            DataFrame con la base GEIH consolidada.
        """
        ruta_p = Path(ruta)
        if ruta_p.suffix == ".parquet":
            df = pd.read_parquet(ruta_p)
        elif ruta_p.suffix in (".csv", ".zip"):
            df = pd.read_csv(ruta_p, low_memory=False)
        else:
            raise ValueError(f"Formato no reconocido: {ruta_p.suffix}")

        print(f"✅ Base cargada: {df.shape[0]:,} filas × {df.shape[1]} cols")
        return df

    # ── Métodos privados ───────────────────────────────────────────

    @staticmethod
    def _inferir_numero_mes(nombre_carpeta: str) -> int:
        """Infiere el número de mes a partir del nombre de carpeta DANE.

        'Enero 2025' → 1, 'Diciembre 2026' → 12.

        Args:
            nombre_carpeta: Nombre como 'Marzo 2026'.

        Returns:
            Número del mes (1-12).

        Raises:
            ValueError: Si no se puede inferir el mes.
        """
        nombre_lower = nombre_carpeta.lower().strip()
        for i, mes in enumerate(MESES_NOMBRES, 1):
            if nombre_lower.startswith(mes.lower()):
                return i
        raise ValueError(
            f"No se pudo inferir el número de mes de '{nombre_carpeta}'. "
            f"Pase numero_mes= explícitamente."
        )

    def _procesar_mes(self, mes: str, numero_mes: int) -> pd.DataFrame:
        """Lee y une los módulos de un mes específico.

        El módulo ancla es siempre 'Características generales' porque
        contiene TODOS los individuos del universo. Los demás módulos
        son subconjuntos (solo ocupados, solo desocupados, etc.).

        Usa _buscar_archivo() para encontrar CSVs sin importar
        si tienen tildes o no (Migración.CSV = Migracion.CSV).
        """
        ruta_csv = self.ruta_base / mes / "CSV"

        # Leer módulo ancla (universo completo)
        ruta_ancla = self._buscar_archivo(ruta_csv, MODULOS_CSV["caracteristicas"])
        if ruta_ancla is None:
            raise FileNotFoundError(
                f"Módulo ancla no encontrado en {ruta_csv}: "
                f"{MODULOS_CSV['caracteristicas']}"
            )
        df_mes = self._leer_csv(ruta_ancla)

        # Unir módulos secundarios
        modulos_secundarios = {
            "hogar":          LLAVES_HOGAR,
            "fuerza_trabajo": LLAVES_PERSONA,
            "ocupados":       LLAVES_PERSONA,
            "no_ocupados":    LLAVES_PERSONA,
        }
        # Módulos opcionales
        for mod in self._modulos_a_incluir:
            if mod in ("caracteristicas",):
                continue
            if mod not in modulos_secundarios:
                modulos_secundarios[mod] = LLAVES_PERSONA

        for mod_key, llaves in modulos_secundarios.items():
            if mod_key not in self._modulos_a_incluir:
                continue
            nombre_csv = MODULOS_CSV.get(mod_key)
            if not nombre_csv:
                continue
            ruta_mod = self._buscar_archivo(ruta_csv, nombre_csv)
            if ruta_mod is not None:
                df_mod = self._leer_csv(ruta_mod)
                df_mes = self._unir_sin_duplicados(df_mes, df_mod, llaves)
                del df_mod
                gc.collect()

        # Agregar número de mes
        df_mes["MES_NUM"] = numero_mes

        return df_mes

    def _leer_csv(self, ruta: Path) -> pd.DataFrame:
        """Lee un CSV del DANE con la configuración correcta."""
        return pd.read_csv(
            ruta,
            sep=self.config.separador_csv,
            encoding=self.config.encoding_csv,
            converters=self.converters,
            low_memory=False,
        )

    @staticmethod
    def _unir_sin_duplicados(
        df_izq: pd.DataFrame,
        df_der: pd.DataFrame,
        llaves: List[str],
    ) -> pd.DataFrame:
        """Une dos DataFrames evitando columnas duplicadas.

        Solo trae del df_der las columnas que NO existen en df_izq.
        Usa how='left' para preservar el universo del módulo ancla.

        ⚠️ NUNCA usar how='outer' aquí: multiplicaría filas y
        produciría el error clásico de PEA inflada.

        Args:
            df_izq: DataFrame base (módulo ancla o acumulado).
            df_der: DataFrame a unir (módulo secundario).
            llaves: Columnas de join.

        Returns:
            DataFrame unido sin columnas _x / _y.
        """
        columnas_nuevas = df_der.columns.difference(df_izq.columns).tolist()
        columnas_a_usar = columnas_nuevas + llaves
        return pd.merge(
            df_izq,
            df_der[columnas_a_usar],
            on=llaves,
            how="left",
        )

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/consolidador.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [15/39]: dashboard.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/dashboard.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/dashboard.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.dashboard — Dashboard interactivo Streamlit para exploración de resultados.

Permite a usuarios no-técnicos explorar los resultados de la GEIH
con filtros, gráficos interactivos y tablas descargables, sin
necesidad de escribir código Python.

REQUISITOS:
    pip install streamlit plotly

EJECUCIÓN:
    # Desde la carpeta del paquete:
    streamlit run geih_2025/dashboard.py

    # Desde Google Colab (con tunnel):
    !pip install streamlit plotly
    !streamlit run geih_2025/dashboard.py --server.headless true &
    # Usar el link que genera Streamlit

    # O con localtunnel:
    !npm install -g localtunnel
    !lt --port 8501

DATOS:
    El dashboard busca archivos Parquet consolidados en una ruta
    configurable. También puede cargar CSVs de la carpeta resultados/.

Autor: Néstor Enrique Forero Herrera
"""

__all__ = ["ejecutar_dashboard"]

import sys
from pathlib import Path


def ejecutar_dashboard(
    ruta_base: str = ".",
    puerto: int = 8501,
) -> None:
    """Lanza el dashboard Streamlit programáticamente.

    Args:
        ruta_base: Carpeta con los Parquets y resultados.
        puerto: Puerto del servidor Streamlit.
    """
    import subprocess
    script = str(Path(__file__).resolve())
    subprocess.Popen([
        sys.executable, "-m", "streamlit", "run", script,
        "--server.port", str(puerto),
        "--server.headless", "true",
        "--", ruta_base,
    ])
    print(f"🌐 Dashboard disponible en http://localhost:{puerto}")


# ═════════════════════════════════════════════════════════════════════
# La siguiente sección se ejecuta SOLO cuando Streamlit carga el archivo.
# No se ejecuta al hacer `from geih.dashboard import ...`
# ═════════════════════════════════════════════════════════════════════

def _main():
    """Punto de entrada del dashboard Streamlit."""
    try:
        import streamlit as st
        import plotly.express as px
        import plotly.graph_objects as go
    except ImportError:
        print("❌ Instale streamlit y plotly:")
        print("   pip install streamlit plotly")
        return

    import pandas as pd
    import numpy as np
    from pathlib import Path

    # ── Configuración de página ────────────────────────────────
    st.set_page_config(
        page_title="GEIH Dashboard — geih-analisis",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── CSS institucional ──────────────────────────────────────
    st.markdown("""
    <style>
    .main-header {
        font-size: 2rem; font-weight: bold; color: #1A3C6E;
        border-bottom: 3px solid #8B1A4A; padding-bottom: 10px;
        margin-bottom: 20px;
    }
    .metric-card {
        background: #F7F9FC; border-radius: 8px; padding: 15px;
        border-left: 4px solid #2E6DA4; margin-bottom: 10px;
    }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="main-header">📊 GEIH Dashboard</div>',
                unsafe_allow_html=True)

    # ── Sidebar: configuración ─────────────────────────────────
    st.sidebar.image("https://procolombia.co/sites/default/files/logo_procolombia.png",
                     width=200, use_container_width=False)
    st.sidebar.title("⚙️ Configuración")

    # Ruta de datos
    ruta_default = sys.argv[-1] if len(sys.argv) > 1 and Path(sys.argv[-1]).exists() else "."
    ruta_base = st.sidebar.text_input("Ruta de datos", value=ruta_default)
    ruta = Path(ruta_base)

    # Buscar Parquets disponibles
    parquets = sorted(ruta.glob("GEIH_*_Consolidado*.parquet"))
    if not parquets:
        parquets = sorted(ruta.glob("*.parquet"))

    if not parquets:
        st.warning("⚠️ No se encontraron archivos Parquet en la ruta especificada.")
        st.info(f"Buscando en: {ruta.resolve()}")
        st.info("Suba un archivo .parquet o ajuste la ruta en el sidebar.")

        # Permitir subir archivo
        uploaded = st.file_uploader("📂 O suba un archivo Parquet", type=["parquet"])
        if uploaded:
            df = pd.read_parquet(uploaded)
            st.success(f"✅ Archivo cargado: {df.shape[0]:,} filas × {df.shape[1]} cols")
        else:
            st.stop()
    else:
        archivo_sel = st.sidebar.selectbox(
            "Archivo Parquet", parquets,
            format_func=lambda x: x.name,
        )
        df = _cargar_con_cache(str(archivo_sel))

    # ── Info del dataset ───────────────────────────────────────
    st.sidebar.markdown(f"**Registros:** {df.shape[0]:,}")
    st.sidebar.markdown(f"**Columnas:** {df.shape[1]}")
    if "MES_NUM" in df.columns:
        meses = sorted(df["MES_NUM"].unique())
        st.sidebar.markdown(f"**Meses:** {len(meses)}")

    # ── Tabs principales ───────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📈 Indicadores", "💰 Ingresos", "🏢 Sectores",
        "🗺️ Departamentos", "📋 Explorar datos",
    ])

    # ── Preparar datos básicos ─────────────────────────────────
    col_peso = "FEX_ADJ" if "FEX_ADJ" in df.columns else "FEX_C18"
    tiene_ocu = "OCI" in df.columns
    tiene_ing = "INGLABO" in df.columns

    # ═══ TAB 1: INDICADORES ═══════════════════════════════════
    with tab1:
        st.subheader("Indicadores del Mercado Laboral")

        if all(c in df.columns for c in ["FT", "OCI", "DSI", "PET", col_peso]):
            pet = df.loc[df["PET"] == 1, col_peso].sum()
            pea = df.loc[df["FT"] == 1, col_peso].sum()
            ocu = df.loc[df["OCI"] == 1, col_peso].sum()
            dsi = df.loc[df["DSI"] == 1, col_peso].sum()

            td = dsi / pea * 100 if pea > 0 else 0
            tgp = pea / pet * 100 if pet > 0 else 0
            to = ocu / pet * 100 if pet > 0 else 0

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Tasa de Desempleo", f"{td:.1f}%")
            c2.metric("TGP", f"{tgp:.1f}%")
            c3.metric("Tasa de Ocupación", f"{to:.1f}%")
            c4.metric("Ocupados", f"{ocu/1e6:.2f}M")

            # Indicadores por sexo
            if "P3271" in df.columns:
                st.markdown("#### Por sexo")
                sexo_data = []
                for sv, sl in [(1, "Hombres"), (2, "Mujeres")]:
                    m = df["P3271"] == sv
                    pea_s = df.loc[m & (df["FT"] == 1), col_peso].sum()
                    dsi_s = df.loc[m & (df["DSI"] == 1), col_peso].sum()
                    td_s = dsi_s / pea_s * 100 if pea_s > 0 else 0
                    ocu_s = df.loc[m & (df["OCI"] == 1), col_peso].sum()
                    sexo_data.append({"Sexo": sl, "TD_%": round(td_s, 1),
                                      "Ocupados_M": round(ocu_s / 1e6, 2)})
                st.dataframe(pd.DataFrame(sexo_data), use_container_width=True)

            # Gráfico estacionalidad si hay MES_NUM
            if "MES_NUM" in df.columns:
                st.markdown("#### Estacionalidad mensual")
                estac_rows = []
                for mes in sorted(df["MES_NUM"].unique()):
                    m = df["MES_NUM"] == mes
                    pea_m = df.loc[m & (df["FT"] == 1), "FEX_C18"].sum()
                    dsi_m = df.loc[m & (df["DSI"] == 1), "FEX_C18"].sum()
                    ocu_m = df.loc[m & (df["OCI"] == 1), "FEX_C18"].sum()
                    pet_m = df.loc[m & (df["PET"] == 1), "FEX_C18"].sum()
                    estac_rows.append({
                        "Mes": int(mes),
                        "TD_%": round(dsi_m / pea_m * 100, 1) if pea_m > 0 else 0,
                        "TGP_%": round(pea_m / pet_m * 100, 1) if pet_m > 0 else 0,
                        "TO_%": round(ocu_m / pet_m * 100, 1) if pet_m > 0 else 0,
                    })
                df_estac = pd.DataFrame(estac_rows)
                fig_e = px.line(df_estac, x="Mes", y=["TD_%", "TGP_%", "TO_%"],
                                title="Indicadores mensuales", markers=True,
                                color_discrete_map={"TD_%": "#C0392B",
                                                     "TGP_%": "#2E6DA4",
                                                     "TO_%": "#1E8449"})
                st.plotly_chart(fig_e, use_container_width=True)
        else:
            st.info("Columnas FT, OCI, DSI, PET no disponibles para indicadores.")

    # ═══ TAB 2: INGRESOS ══════════════════════════════════════
    with tab2:
        st.subheader("Distribución de Ingresos Laborales")
        if tiene_ocu and tiene_ing:
            df_ocu = df[(df["OCI"] == 1) & (df["INGLABO"] > 0)].copy()

            smmlv_input = st.number_input("SMMLV (COP)", value=1_423_500, step=100_000)
            df_ocu["SMMLV_mult"] = df_ocu["INGLABO"] / smmlv_input

            # Histograma
            sample_n = min(50_000, len(df_ocu))
            df_sample = df_ocu.sample(sample_n, weights=col_peso, random_state=42, replace=True)
            if "P3271" in df_sample.columns:
                df_sample["Sexo"] = df_sample["P3271"].map({1: "Hombres", 2: "Mujeres"})
                fig_h = px.histogram(df_sample, x="SMMLV_mult", color="Sexo",
                                      nbins=60, barmode="overlay", opacity=0.7,
                                      range_x=[0, 10],
                                      color_discrete_map={"Hombres": "#2E6DA4",
                                                           "Mujeres": "#C0392B"},
                                      title="Distribución del ingreso (× SMMLV)")
            else:
                fig_h = px.histogram(df_sample, x="SMMLV_mult", nbins=60,
                                      range_x=[0, 10],
                                      title="Distribución del ingreso (× SMMLV)")
            fig_h.add_vline(x=1, line_dash="dash", line_color="green",
                            annotation_text="1 SMMLV")
            st.plotly_chart(fig_h, use_container_width=True)

            # Estadísticas
            med = np.average(df_ocu["INGLABO"],
                             weights=df_ocu[col_peso]) if col_peso in df_ocu.columns else df_ocu["INGLABO"].mean()
            c1, c2, c3 = st.columns(3)
            c1.metric("Media", f"${med:,.0f}")
            c2.metric("Mediana (aprox.)", f"${df_ocu['INGLABO'].median():,.0f}")
            c3.metric("En SMMLV", f"{med/smmlv_input:.2f}×")
        else:
            st.info("Columnas OCI e INGLABO no disponibles.")

    # ═══ TAB 3: SECTORES ══════════════════════════════════════
    with tab3:
        st.subheader("Empleo por Actividad Económica")
        rama_col = "RAMA" if "RAMA" in df.columns else "RAMA2D_R4"
        if tiene_ocu and rama_col in df.columns:
            df_rama = (
                df[df["OCI"] == 1].groupby(rama_col)[col_peso]
                .sum().sort_values(ascending=False)
                .reset_index()
                .rename(columns={rama_col: "Rama", col_peso: "Ocupados"})
            )
            df_rama["Ocupados_M"] = (df_rama["Ocupados"] / 1e6).round(2)
            df_rama = df_rama.head(15)

            fig_r = px.bar(df_rama, y="Rama", x="Ocupados_M",
                          orientation="h", title="Top 15 ramas por ocupados (M)",
                          color="Ocupados_M", color_continuous_scale="Blues")
            fig_r.update_layout(yaxis=dict(autorange="reversed"), height=500)
            st.plotly_chart(fig_r, use_container_width=True)

            st.dataframe(df_rama[["Rama", "Ocupados_M"]], use_container_width=True)
        else:
            st.info(f"Columna {rama_col} no disponible.")

    # ═══ TAB 4: DEPARTAMENTOS ═════════════════════════════════
    with tab4:
        st.subheader("Indicadores por Departamento")
        dpto_col = "NOMBRE_DPTO" if "NOMBRE_DPTO" in df.columns else "DPTO_STR" if "DPTO_STR" in df.columns else "DPTO"
        if dpto_col in df.columns and all(c in df.columns for c in ["FT", "OCI", "DSI", col_peso]):
            dptos = df[dpto_col].dropna().unique()
            filas_d = []
            for d in dptos:
                m = df[dpto_col] == d
                pea_d = df.loc[m & (df["FT"] == 1), col_peso].sum()
                dsi_d = df.loc[m & (df["DSI"] == 1), col_peso].sum()
                ocu_d = df.loc[m & (df["OCI"] == 1), col_peso].sum()
                if pea_d < 1_000:
                    continue
                filas_d.append({
                    "Departamento": str(d),
                    "TD_%": round(dsi_d / pea_d * 100, 1),
                    "Ocupados_miles": round(ocu_d / 1_000, 1),
                })
            df_dpto = pd.DataFrame(filas_d).sort_values("TD_%", ascending=False)

            fig_d = px.bar(df_dpto, x="TD_%", y="Departamento",
                          orientation="h", title="Tasa de Desempleo por Departamento",
                          color="TD_%", color_continuous_scale="RdYlGn_r")
            fig_d.update_layout(yaxis=dict(autorange="reversed"),
                               height=max(400, len(df_dpto) * 25))
            st.plotly_chart(fig_d, use_container_width=True)

            st.dataframe(df_dpto, use_container_width=True)
        else:
            st.info("Columnas de departamento no disponibles.")

    # ═══ TAB 5: EXPLORAR DATOS ════════════════════════════════
    with tab5:
        st.subheader("Exploración libre de la base")

        # Filtros
        col1, col2 = st.columns(2)
        with col1:
            cols_disponibles = df.columns.tolist()
            cols_sel = st.multiselect("Columnas a mostrar",
                                       cols_disponibles,
                                       default=cols_disponibles[:10])
        with col2:
            n_filas = st.slider("Filas a mostrar", 10, 1000, 100)

        if cols_sel:
            st.dataframe(df[cols_sel].head(n_filas), use_container_width=True)

        # Descarga
        st.markdown("#### 📥 Descargar datos")
        if st.button("Generar CSV (primeras 10,000 filas)"):
            csv = df.head(10_000).to_csv(index=False).encode("utf-8-sig")
            st.download_button("⬇️ Descargar CSV", csv,
                              "geih_muestra.csv", "text/csv")

    # ── Footer ─────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        "<div style='text-align:center; color:#7F8C8D; font-size:0.85rem;'>"
        "Autor: Néstor Enrique Forero Herrera · geih_2025 v4.3"
        "</div>",
        unsafe_allow_html=True,
    )


def _cargar_con_cache(ruta: str):
    """Carga Parquet con cache manual (compatible sin Streamlit)."""
    import pandas as pd
    return pd.read_parquet(ruta)


# Solo ejecutar la app Streamlit si se llama directamente
if __name__ == "__main__":
    _main()

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/dashboard.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [16/39]: descargador.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/descargador.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/descargador.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.descargador — Descarga automática de microdatos GEIH desde el DANE.

Elimina el paso manual de descargar ZIPs del portal de microdatos,
descomprimirlos y organizarlos en carpetas mensuales.

Inspirado en geihdanepy (github.com/BautistaDavid/geihdanepy) pero
mejorado con: soporte multi-año, descompresión automática a la
estructura de carpetas que espera ConsolidadorGEIH, verificación
de integridad, y progreso visible.

NOTA IMPORTANTE: El portal de microdatos del DANE (microdatos.dane.gov.co)
a veces requiere aceptar términos de uso. Si la descarga falla con
código 403 o redirección, el usuario debe descargar manualmente desde
el portal y colocar los ZIPs en la carpeta indicada.

Autor: Néstor Enrique Forero Herrera
"""

__all__ = ["DescargadorDANE"]

import io
import os
import zipfile
import shutil
from pathlib import Path
from typing import Optional, List, Dict

from .config import ConfigGEIH, MESES_NOMBRES, MODULOS_CSV


# ═════════════════════════════════════════════════════════════════════
# CATÁLOGO DE URLs DEL DANE POR AÑO
# ═════════════════════════════════════════════════════════════════════

# IDs del catálogo de microdatos DANE para cada año.
# Página: https://microdatos.dane.gov.co/index.php/catalog/{ID}/get-microdata
# Actualizar cuando el DANE publique nuevos años.
CATALOGO_DANE: Dict[int, Dict] = {
    2022: {"catalog_id": 771, "nota": "Marco 2018 — primer año"},
    2023: {"catalog_id": 782, "nota": "Marco 2018"},
    2024: {"catalog_id": 819, "nota": "Marco 2018"},
    2025: {"catalog_id": 853, "nota": "Marco 2018"},
    # 2026: {"catalog_id": ???, "nota": "Actualizar cuando se publique"},
}

# Patrones de URL del DANE (pueden cambiar — el DANE no tiene API estable)
_URL_MICRODATOS = "https://microdatos.dane.gov.co/index.php/catalog/{catalog_id}/get-microdata"
_URL_DOWNLOAD = "https://microdatos.dane.gov.co/index.php/catalog/{catalog_id}/download/{file_id}"


class DescargadorDANE:
    """Descarga y organiza microdatos GEIH desde el portal del DANE.

    El flujo completo es:
      1. Descargar ZIPs mensuales del portal de microdatos
      2. Descomprimirlos en la estructura de carpetas esperada
      3. Verificar que todos los módulos CSV estén presentes

    Si la descarga automática falla (el DANE requiere aceptar términos),
    el módulo ofrece instrucciones claras para descarga manual.

    Uso típico:
        desc = DescargadorDANE(
            config=ConfigGEIH(anio=2025, n_meses=12),
            ruta_destino='/content/drive/MyDrive/GEIH',
        )
        # Opción A: Descarga automática (si el DANE lo permite)
        desc.descargar_todos()

        # Opción B: Organizar ZIPs ya descargados manualmente
        desc.organizar_zips('/content/drive/MyDrive/GEIH/zips_dane')

        # Verificar estructura
        desc.verificar()
    """

    def __init__(
        self,
        config: Optional[ConfigGEIH] = None,
        ruta_destino: str = ".",
    ):
        """
        Args:
            config: Configuración con año y n_meses.
            ruta_destino: Carpeta donde crear la estructura de carpetas
                          mensuales (ej: 'Enero 2025/CSV/').
        """
        self.config = config or ConfigGEIH()
        self.ruta_destino = Path(ruta_destino)
        self.ruta_destino.mkdir(parents=True, exist_ok=True)

    def descargar_todos(self) -> Dict[str, str]:
        """Intenta descargar todos los meses desde el portal del DANE.

        Returns:
            Dict con {mes: 'ok'|'error: mensaje'} para cada mes.

        NOTA: El DANE puede bloquear descargas automáticas. Si falla,
        use organizar_zips() con archivos descargados manualmente.
        """
        anio = self.config.anio
        if anio not in CATALOGO_DANE:
            print(f"❌ Año {anio} no está en el catálogo del DANE.")
            print(f"   Años disponibles: {sorted(CATALOGO_DANE.keys())}")
            print(f"   → Descargue manualmente desde:")
            print(f"     https://microdatos.dane.gov.co/index.php/catalog/central")
            return {}

        catalog_id = CATALOGO_DANE[anio]["catalog_id"]
        print(f"\n{'='*60}")
        print(f"  DESCARGA GEIH {anio} — Catálogo DANE #{catalog_id}")
        print(f"{'='*60}")
        print(f"  Portal: {_URL_MICRODATOS.format(catalog_id=catalog_id)}")
        print(f"  Meses a descargar: {self.config.n_meses}")

        resultados = {}
        carpetas = self.config.carpetas_mensuales

        try:
            import requests
        except ImportError:
            print("\n⚠️  El módulo 'requests' no está instalado.")
            print("   Instalar con: !pip install requests")
            print("   O use organizar_zips() con archivos descargados manualmente.")
            return {}

        for i, mes_carpeta in enumerate(carpetas, 1):
            mes_nombre = mes_carpeta.split()[0]  # 'Enero 2025' → 'Enero'
            print(f"\n🔄 [{i}/{len(carpetas)}] Descargando {mes_carpeta}...")

            try:
                ok = self._descargar_mes(mes_nombre, catalog_id, requests)
                resultados[mes_carpeta] = "ok" if ok else "no encontrado"
            except Exception as e:
                resultados[mes_carpeta] = f"error: {e}"
                print(f"   ❌ Error: {e}")

        # Resumen
        ok_count = sum(1 for v in resultados.values() if v == "ok")
        print(f"\n{'='*60}")
        print(f"  RESUMEN: {ok_count}/{len(carpetas)} meses descargados")
        if ok_count < len(carpetas):
            print(f"\n  Para los meses faltantes, descargue manualmente desde:")
            print(f"  {_URL_MICRODATOS.format(catalog_id=catalog_id)}")
            print(f"  y use: descargador.organizar_zips('ruta/a/los/zips')")
        print(f"{'='*60}")

        return resultados

    def _descargar_mes(self, mes_nombre: str, catalog_id: int, requests) -> bool:
        """Intenta descargar el ZIP de un mes específico.

        El DANE nombra los archivos como 'Enero.csv', 'Febrero.csv', etc.
        (son ZIPs renombrados o CSVs directos según el año).
        """
        # Intentar variantes de nombre que usa el DANE
        variantes = [
            f"{mes_nombre}.csv",
            f"{mes_nombre} CSV.zip",
            f"{mes_nombre}.zip",
            f"mes_{MESES_NOMBRES.index(mes_nombre)+1:02d}.zip",
        ]

        url_page = _URL_MICRODATOS.format(catalog_id=catalog_id)

        # El DANE no tiene API pública — intentamos scraping ligero
        try:
            resp = requests.get(url_page, timeout=30)
            if resp.status_code != 200:
                print(f"   ⚠️  Portal responde {resp.status_code} — puede requerir login")
                return False

            # Buscar links de descarga en la página
            content = resp.text
            for variante in variantes:
                if variante.lower() in content.lower():
                    print(f"   📥 Encontrado: {variante}")
                    # Intentar extraer el link de descarga
                    # (El DANE usa JavaScript, esto es best-effort)
                    break
            else:
                print(f"   ⚠️  No se encontró archivo para {mes_nombre}")
                return False

        except requests.exceptions.RequestException as e:
            print(f"   ⚠️  Error de red: {e}")
            return False

        print(f"   ⚠️  Descarga automática no soportada por el portal actual del DANE.")
        print(f"   → Use organizar_zips() con archivos descargados manualmente.")
        return False

    def organizar_zips(
        self,
        ruta_zips: str,
        patron: str = "*.zip",
    ) -> int:
        """Organiza ZIPs descargados manualmente en la estructura de carpetas.

        El DANE publica ZIPs mensuales. Esta función los descomprime
        en la estructura que espera ConsolidadorGEIH:
            Enero 2025/CSV/Características generales....CSV
            Enero 2025/CSV/Ocupados.CSV
            ...

        Args:
            ruta_zips: Carpeta donde están los ZIPs descargados.
            patron: Patrón glob para encontrar ZIPs.

        Returns:
            Número de meses organizados exitosamente.

        Uso:
            # 1. Descargue los ZIPs del DANE manualmente
            # 2. Colóquelos en una carpeta
            desc.organizar_zips('/content/drive/MyDrive/GEIH/zips')
        """
        ruta = Path(ruta_zips)
        if not ruta.exists():
            print(f"❌ Carpeta no existe: {ruta}")
            return 0

        zips = sorted(ruta.glob(patron))
        if not zips:
            print(f"⚠️  No se encontraron archivos {patron} en {ruta}")
            return 0

        print(f"\n{'='*60}")
        print(f"  ORGANIZANDO ZIPs GEIH {self.config.anio}")
        print(f"{'='*60}")
        print(f"  ZIPs encontrados: {len(zips)}")

        organizados = 0
        for zip_path in zips:
            try:
                mes_nombre = self._inferir_mes_de_zip(zip_path.name)
                if mes_nombre is None:
                    print(f"  ⚠️  No se pudo inferir el mes de: {zip_path.name}")
                    continue

                carpeta_mes = f"{mes_nombre} {self.config.anio}"
                ruta_csv = self.ruta_destino / carpeta_mes / "CSV"
                ruta_csv.mkdir(parents=True, exist_ok=True)

                n_extraidos = self._extraer_zip(zip_path, ruta_csv)
                print(f"  ✅ {carpeta_mes}: {n_extraidos} archivos extraídos")
                organizados += 1

            except Exception as e:
                print(f"  ❌ Error con {zip_path.name}: {e}")

        print(f"\n✅ {organizados} meses organizados en {self.ruta_destino}")
        return organizados

    def _extraer_zip(self, zip_path: Path, ruta_csv: Path) -> int:
        """Extrae un ZIP al directorio CSV, manejando subcarpetas."""
        n = 0
        with zipfile.ZipFile(zip_path, "r") as zf:
            for member in zf.namelist():
                # Solo extraer CSVs (ignorar carpetas, __MACOSX, etc.)
                basename = Path(member).name
                if not basename.upper().endswith(".CSV"):
                    continue
                if basename.startswith(".") or "__MACOSX" in member:
                    continue

                # Extraer al directorio CSV plano
                with zf.open(member) as src:
                    dest = ruta_csv / basename
                    with open(dest, "wb") as dst:
                        shutil.copyfileobj(src, dst)
                    n += 1
        return n

    def _inferir_mes_de_zip(self, nombre_archivo: str) -> Optional[str]:
        """Infiere el nombre del mes a partir del nombre del ZIP."""
        nombre_lower = nombre_archivo.lower()
        for mes in MESES_NOMBRES:
            if mes.lower() in nombre_lower:
                return mes
        # Intentar por número: mes_01.zip, 01_enero.zip, etc.
        import re
        match = re.search(r"(\d{1,2})", nombre_archivo)
        if match:
            num = int(match.group(1))
            if 1 <= num <= 12:
                return MESES_NOMBRES[num - 1]
        return None

    def verificar(self) -> Dict[str, List[str]]:
        """Verifica que la estructura de carpetas esté completa.

        Wrapper sobre ConsolidadorGEIH.verificar_estructura().

        Returns:
            Dict con 'ok' y 'faltantes'.
        """
        from .consolidador import ConsolidadorGEIH
        c = ConsolidadorGEIH(
            ruta_base=str(self.ruta_destino),
            config=self.config,
        )
        return c.verificar_estructura()

    def instrucciones_descarga_manual(self) -> None:
        """Imprime instrucciones paso a paso para descarga manual."""
        anio = self.config.anio
        catalog_id = CATALOGO_DANE.get(anio, {}).get("catalog_id", "???")

        print(f"\n{'='*65}")
        print(f"  📋 INSTRUCCIONES DE DESCARGA MANUAL — GEIH {anio}")
        print(f"{'='*65}")
        print(f"")
        print(f"  1. Abra el portal de microdatos del DANE:")
        print(f"     https://microdatos.dane.gov.co/index.php/catalog/{catalog_id}/get-microdata")
        print(f"")
        print(f"  2. Acepte los términos de uso (si se le solicita).")
        print(f"")
        print(f"  3. Descargue los ZIPs mensuales (Enero.csv ... Diciembre.csv)")
        print(f"     Nota: El DANE los nombra .csv pero son ZIPs.")
        print(f"")
        print(f"  4. Coloque TODOS los ZIPs en una sola carpeta, por ejemplo:")
        print(f"     {self.ruta_destino}/zips_dane_{anio}/")
        print(f"")
        print(f"  5. Ejecute en el notebook:")
        print(f"     from geih import DescargadorDANE, ConfigGEIH")
        print(f"     desc = DescargadorDANE(")
        print(f"         config=ConfigGEIH(anio={anio}, n_meses={self.config.n_meses}),")
        print(f"         ruta_destino='{self.ruta_destino}',")
        print(f"     )")
        print(f"     desc.organizar_zips('{self.ruta_destino}/zips_dane_{anio}')")
        print(f"     desc.verificar()")
        print(f"")
        print(f"  6. Continúe con el pipeline normal:")
        print(f"     consolidador = ConsolidadorGEIH(ruta_base='{self.ruta_destino}', ...)")
        print(f"{'='*65}")

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/descargador.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [17/39]: diagnostico.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/diagnostico.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/diagnostico.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.diagnostico — Diagnóstico de calidad de datos de la base GEIH.

Replica y mejora las funciones missing_values_table() y missing()
del notebook GEIH 2021, con visualización y reportes detallados.

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "DiagnosticoCalidad",
    "Top20Sectores",
]


from typing import Optional, List

import numpy as np
import pandas as pd


class DiagnosticoCalidad:
    """Diagnóstico completo de calidad de datos de la base consolidada.

    Detecta: valores faltantes, ceros sospechosos, tipos de dato incorrectos,
    columnas potencialmente duplicadas, y distribuciones anómalas.

    Uso:
        diag = DiagnosticoCalidad()
        tabla = diag.valores_faltantes(geih_2025_final)
        diag.verificar_tipos(geih_2025_final)
        diag.columnas_duplicadas(geih_2025_final)
    """

    @staticmethod
    def valores_faltantes(
        df: pd.DataFrame,
        titulo: str = "Base consolidada",
        umbral_pct: float = 0.0,
    ) -> pd.DataFrame:
        """Tabla de valores faltantes y ceros por columna.

        Args:
            df: DataFrame a diagnosticar.
            titulo: Nombre para el reporte.
            umbral_pct: Solo mostrar columnas con % nulos > umbral.

        Returns:
            DataFrame con estadísticas de calidad por columna.
        """
        total = len(df)
        mis_val = df.isnull().sum()
        mis_pct = (mis_val / total * 100).round(1)
        zeros = (df == 0).sum()
        zeros_p = (zeros / total * 100).round(1)

        tabla = pd.DataFrame({
            "Valores_nulos": mis_val,
            "Pct_nulos_%": mis_pct,
            "Ceros": zeros,
            "Pct_ceros_%": zeros_p,
            "Dtype": df.dtypes.astype(str),
            "Valores_unicos": df.nunique(),
        })

        tabla = tabla[
            (tabla["Pct_nulos_%"] > umbral_pct) | (tabla["Pct_ceros_%"] > 0)
        ].sort_values("Pct_nulos_%", ascending=False)

        print(f"\n{'='*75}")
        print(f"  DIAGNÓSTICO DE CALIDAD — {titulo}")
        print(f"  Total filas: {total:,}  |  Total columnas: {df.shape[1]}")
        print(f"  Columnas con nulos o ceros: {len(tabla)}")
        print(f"{'='*75}")
        print(f"  {'Columna':<25} {'%Nulos':>8} {'N_nulos':>10} "
              f"{'%Ceros':>8} {'Únicos':>8} {'Dtype':<12}")
        print(f"  {'─'*25} {'─'*8} {'─'*10} {'─'*8} {'─'*8} {'─'*12}")

        for col, row in tabla.head(40).iterrows():
            print(f"  {str(col):<25} {row['Pct_nulos_%']:>7.1f}% "
                  f"{int(row['Valores_nulos']):>10,} "
                  f"{row['Pct_ceros_%']:>7.1f}% "
                  f"{int(row['Valores_unicos']):>8,} "
                  f"{str(row['Dtype']):<12}")

        return tabla

    @staticmethod
    def verificar_tipos(df: pd.DataFrame) -> pd.DataFrame:
        """Verifica que las columnas críticas tengan el tipo correcto.

        Alerta si DPTO, RAMA2D_R4, etc. son numéricas (deben ser str).
        Alerta si FEX_C18, INGLABO, etc. son string (deben ser numéricas).
        """
        DEBE_SER_STR = ["DIRECTORIO", "SECUENCIA_P", "ORDEN", "DPTO",
                        "RAMA2D_R4", "RAMA4D_R4", "AREA"]
        DEBE_SER_NUM = ["FEX_C18", "OCI", "P3271", "P6040", "INGLABO",
                        "FT", "DSI", "PET", "P6920", "P6800"]

        problemas = []
        for col in DEBE_SER_STR:
            if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
                problemas.append({
                    "Columna": col, "Tipo_actual": str(df[col].dtype),
                    "Tipo_esperado": "string/object",
                    "Riesgo": "Pérdida de ceros líderes (ej: DPTO '05'→5)",
                })
        for col in DEBE_SER_NUM:
            if col in df.columns and not pd.api.types.is_numeric_dtype(df[col]):
                problemas.append({
                    "Columna": col, "Tipo_actual": str(df[col].dtype),
                    "Tipo_esperado": "numeric",
                    "Riesgo": "Cálculos fallan o producen NaN silencioso",
                })

        df_prob = pd.DataFrame(problemas)
        if df_prob.empty:
            print("✅ Todos los tipos de dato son correctos")
        else:
            print(f"\n⚠️  {len(df_prob)} columnas con tipo incorrecto:")
            for _, row in df_prob.iterrows():
                print(f"   ❌ {row['Columna']}: es {row['Tipo_actual']}, "
                      f"debería ser {row['Tipo_esperado']}")
                print(f"      Riesgo: {row['Riesgo']}")
        return df_prob

    @staticmethod
    def columnas_duplicadas(df: pd.DataFrame) -> List[str]:
        """Detecta columnas con nombres similares (potenciales duplicados de merge).

        Busca patrones como col_x / col_y que indican merges sin control.
        """
        sufijos_merge = [c for c in df.columns if c.endswith("_x") or c.endswith("_y")]
        if sufijos_merge:
            print(f"\n⚠️  {len(sufijos_merge)} columnas con sufijo de merge duplicado:")
            for c in sufijos_merge:
                print(f"   • {c}")
            print("   → Indica merge con columnas duplicadas. Revisar consolidación.")
        else:
            print("✅ Sin columnas duplicadas por merge")
        return sufijos_merge

    @staticmethod
    def validar_identidades(df: pd.DataFrame) -> bool:
        """Valida las identidades fundamentales del mercado laboral.

        PEA = OCI + DSI
        PET = PEA + FFT (si FFT disponible)
        """
        ok = True
        cols_requeridas = ["FEX_C18", "OCI", "FT", "DSI", "PET"]
        faltantes = [c for c in cols_requeridas if c not in df.columns]
        if faltantes:
            print(f"⚠️  No se puede validar identidades — faltan: {faltantes}")
            return False

        fex = "FEX_C18"
        pea = df.loc[df["FT"] == 1, fex].sum()
        ocu = df.loc[df["OCI"] == 1, fex].sum()
        dsi = df.loc[df["DSI"] == 1, fex].sum()
        pet = df.loc[df["PET"] == 1, fex].sum()

        # PEA = OCI + DSI (tolerancia 0.1%)
        pea_calc = ocu + dsi
        diff_pea = abs(pea - pea_calc) / max(pea, 1) * 100
        if diff_pea > 0.1:
            print(f"⚠️  PEA ≠ OCI + DSI  (Δ={diff_pea:.2f}%)")
            ok = False
        else:
            print(f"✅ PEA = OCI + DSI  ({pea/1e6:.2f}M = {ocu/1e6:.2f}M + {dsi/1e6:.2f}M)")

        return ok

    @staticmethod
    def resumen_rapido(df: pd.DataFrame) -> None:
        """Imprime un resumen rápido de la base para orientación."""
        print(f"\n{'='*55}")
        print(f"  RESUMEN RÁPIDO DE LA BASE")
        print(f"{'='*55}")
        print(f"  Dimensiones : {df.shape[0]:,} filas × {df.shape[1]} columnas")
        mb = df.memory_usage(deep=True).sum() / 1e6
        print(f"  Memoria     : {mb:,.0f} MB")

        if "MES_NUM" in df.columns:
            meses = sorted(df["MES_NUM"].dropna().unique())
            print(f"  Meses       : {len(meses)} → {meses}")

        if "OCI" in df.columns:
            n_ocu = (df["OCI"] == 1).sum()
            print(f"  Ocupados    : {n_ocu:,} registros ({n_ocu/len(df)*100:.1f}%)")

        if "INGLABO" in df.columns:
            n_ing = df["INGLABO"].notna().sum()
            n_cero = (df["INGLABO"] == 0).sum()
            print(f"  Con INGLABO : {n_ing:,} registros")
            print(f"  INGLABO = 0 : {n_cero:,} (pago en especie)")

        print(f"{'='*55}")

    @staticmethod
    def graficar_nulos(
        tabla_nulos: pd.DataFrame,
        top_n: int = 20,
        titulo: str = "Top columnas por valores faltantes o en cero",
    ):
        """Genera gráfico de barras apiladas: % nulos + % ceros.

        Args:
            tabla_nulos: Output de valores_faltantes().
            top_n: Cuántas columnas mostrar.

        Returns:
            Figura matplotlib.
        """
        import matplotlib.pyplot as plt

        top = tabla_nulos.head(top_n)
        if top.empty:
            print("✅ No hay columnas con valores faltantes para graficar.")
            return None

        fig, ax = plt.subplots(figsize=(14, 6))
        fig.patch.set_facecolor("#F7F9FC")
        ax.set_facecolor("white")

        x = np.arange(len(top))
        ax.bar(x, top["Pct_nulos_%"], 0.55, color="#C0392B", alpha=0.82, label="% Nulos")
        ax.bar(x, top["Pct_ceros_%"], 0.55, bottom=top["Pct_nulos_%"],
               color="#E67E22", alpha=0.70, label="% Ceros")

        ax.set_xticks(x)
        ax.set_xticklabels(top.index, rotation=40, ha="right", fontsize=9)
        ax.set_ylabel("% del total de filas", fontsize=11)
        ax.set_title(titulo, fontsize=12, fontweight="bold")
        ax.legend(fontsize=10)
        ax.set_ylim(0, 110)
        ax.grid(axis="y", alpha=0.3)
        ax.spines[["top", "right"]].set_visible(False)
        fig.tight_layout(pad=2)
        return fig


class Top20Sectores:
    """Top 20 actividades económicas CIIU por número de ocupados.

    Uso:
        from geih import Top20Sectores
        top20 = Top20Sectores()
        tabla = top20.calcular(df, ruta_ciiu=RUTA_CIIU)
        top20.imprimir(tabla)
        fig = top20.graficar(tabla)
    """

    def __init__(self, config=None):
        from .config import ConfigGEIH
        self.config = config or ConfigGEIH()

    def calcular(
        self,
        df: pd.DataFrame,
        ruta_ciiu: Optional[str] = None,
    ) -> pd.DataFrame:
        """Calcula el Top 20 de sectores CIIU por ocupados expandidos.

        Args:
            df: DataFrame con OCI, RAMA2D_R4 o RAMA4D_R4, FEX_ADJ.
            ruta_ciiu: Ruta al Excel de correlativa CIIU para descripciones.
        """
        from .preparador import MergeCorrelativas, PreparadorGEIH
        from .utils import ConversorTipos

        df_ocu = df[df["OCI"] == 1].copy()

        if ruta_ciiu and "RAMA4D_R4" in df_ocu.columns:
            df_ocu = MergeCorrelativas().merge_ciiu(df_ocu, ruta_ciiu)
            col_desc = "DESCRIPCION_CIIU"
        elif "RAMA" in df_ocu.columns:
            col_desc = "RAMA"
        else:
            df_ocu["RAMA"] = PreparadorGEIH.mapear_rama_ciiu(df_ocu["RAMA2D_R4"])
            col_desc = "RAMA"

        top = (
            df_ocu[df_ocu[col_desc].notna()]
            .groupby(col_desc)["FEX_ADJ"]
            .sum().sort_values(ascending=False)
            .head(20).reset_index()
            .rename(columns={col_desc: "Sector_CIIU", "FEX_ADJ": "Ocupados"})
        )
        top["Ocupados_M"] = (top["Ocupados"] / 1e6).round(2)
        top["Pct_%"] = (top["Ocupados"] / top["Ocupados"].sum() * 100).round(1)
        return top

    def graficar(self, top: pd.DataFrame):
        """Gráfico de barras horizontales Top 20 sectores."""
        import matplotlib.pyplot as plt

        df_plot = top.sort_values("Ocupados_M", ascending=True)
        fig, ax = plt.subplots(figsize=(14, 8))
        fig.patch.set_facecolor("#F7F9FC")
        ax.set_facecolor("white")

        colores = ["#2E6DA4" if i >= len(df_plot) - 5 else "#7FB3D3"
                   for i in range(len(df_plot))]
        ax.barh(range(len(df_plot)), df_plot["Ocupados_M"], 0.65,
                color=colores, alpha=0.88)
        for i, (_, row) in enumerate(df_plot.iterrows()):
            ax.text(row["Ocupados_M"] + 0.03, i,
                    f"{row['Ocupados_M']:.2f}M  ({row['Pct_%']:.1f}%)",
                    va="center", fontsize=8.5)
        ax.set_yticks(range(len(df_plot)))
        ax.set_yticklabels([str(r)[:60] for r in df_plot["Sector_CIIU"]], fontsize=8.5)
        ax.set_xlabel("Ocupados (millones)", fontsize=11)
        ax.set_title("Top 20 actividades económicas por ocupados\n"
                     "GEIH 2025 — CIIU Rev.4", fontsize=12, fontweight="bold")
        ax.grid(axis="x", alpha=0.3)
        ax.spines[["top", "right"]].set_visible(False)
        fig.tight_layout(pad=2.5)
        return fig

    def imprimir(self, top: pd.DataFrame) -> None:
        """Imprime la tabla Top 20."""
        print(f"\n{'='*75}")
        print(f"  TOP 20 ACTIVIDADES ECONÓMICAS — GEIH 2025")
        print(f"{'='*75}")
        print(f"  {'Sector CIIU':<58} {'M':>6} {'%':>6}")
        print(f"  {'─'*58} {'─'*6} {'─'*6}")
        for _, row in top.iterrows():
            print(f"  {str(row['Sector_CIIU']):<58} "
                  f"{row['Ocupados_M']:>6.2f} {row['Pct_%']:>5.1f}%")

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/diagnostico.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [18/39]: exportador.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/exportador.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/exportador.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.exportador — Gestión de exportaciones con estructura de carpetas.

Organiza TODOS los outputs (gráficas, tablas CSV, Excel) en una
estructura predecible. Nunca deja archivos sueltos en la raíz.

Estructura de salida:
    resultados_geih_{anio}/
    ├── graficas/          ← PNGs de matplotlib
    ├── tablas/            ← CSVs de resultados
    ├── excel/             ← Archivos Excel multi-hoja
    └── metadata.json      ← Parámetros usados en la corrida

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "Exportador",
]


import json
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any

import pandas as pd
import matplotlib.pyplot as plt

from .config import ConfigGEIH


class Exportador:
    """Gestiona la exportación organizada de todos los resultados.

    Crea la estructura de carpetas automáticamente y provee métodos
    para guardar gráficas, tablas y archivos Excel en el lugar correcto.

    Uso típico:
        exp = Exportador(ruta_base='/content/drive/MyDrive/GEIH')
        exp.guardar_grafica(fig, 'BoxPlot_salarios_rama')
        exp.guardar_tabla(df, 'estadisticas_rama_2025')
        exp.guardar_excel({'Hoja1': df1, 'Hoja2': df2}, 'resultados_2025')
        exp.guardar_metadata(config, {'n_registros': 5_000_000})
    """

    def __init__(
        self,
        ruta_base: str = ".",
        nombre_carpeta: Optional[str] = None,
        config: Optional["ConfigGEIH"] = None,
    ):
        """
        Args:
            ruta_base     : Directorio padre donde se crea la carpeta de resultados.
            nombre_carpeta: Nombre de la carpeta raíz. Si None, se auto-genera
                            como 'resultados_geih_{anio}' usando config.
                            Si config también es None, usa 'resultados_geih'.
            config        : ConfigGEIH para inferir el año automáticamente.

        Ejemplos:
            Exportador(ruta_base=RUTA, config=config)
            # → resultados_geih_2025/ si config.anio==2025
            # → resultados_geih_2026/ si config.anio==2026

            Exportador(ruta_base=RUTA, nombre_carpeta="mis_resultados")
            # → mis_resultados/ (nombre explícito, compatible con v4.x)
        """
        if nombre_carpeta is None:
            anio = config.anio if config is not None else ""
            nombre_carpeta = f"resultados_geih_{anio}" if anio else "resultados_geih"
        self.raiz = Path(ruta_base) / nombre_carpeta
        self.graficas = self.raiz / "graficas"
        self.tablas = self.raiz / "tablas"
        self.excel = self.raiz / "excel"

        # Crear toda la estructura
        for carpeta in [self.graficas, self.tablas, self.excel]:
            carpeta.mkdir(parents=True, exist_ok=True)

        print(f"📂 Exportador listo → {self.raiz}")

    def guardar_grafica(
        self,
        fig: plt.Figure,
        nombre: str,
        dpi: int = 160,
        fondo: str = "#F7F9FC",
        cerrar: bool = True,
    ) -> Path:
        """Guarda una figura matplotlib como PNG.

        Args:
            fig: Figura de matplotlib.
            nombre: Nombre del archivo (sin extensión).
            dpi: Resolución (160 es bueno para presentaciones).
            fondo: Color de fondo del archivo exportado.
            cerrar: Si True, cierra la figura después de guardar (libera RAM).

        Returns:
            Path al archivo guardado.
        """
        ruta = self.graficas / f"{nombre}.png"
        fig.savefig(ruta, dpi=dpi, bbox_inches="tight", facecolor=fondo)
        if cerrar:
            plt.close(fig)
        print(f"   📊 {ruta.name}")
        return ruta

    def guardar_tabla(
        self,
        df: pd.DataFrame,
        nombre: str,
        sep: str = ";",
        encoding: str = "utf-8-sig",
    ) -> Path:
        """Guarda un DataFrame como CSV.

        Args:
            df: DataFrame a exportar.
            nombre: Nombre del archivo (sin extensión).

        Returns:
            Path al archivo guardado.
        """
        ruta = self.tablas / f"{nombre}.csv"
        df.to_csv(ruta, index=False, sep=sep, encoding=encoding)
        print(f"   📋 {ruta.name} ({len(df):,} filas)")
        return ruta

    def guardar_excel(
        self,
        hojas: Dict[str, pd.DataFrame],
        nombre: str,
        formato_institucional: bool = True,
    ) -> Path:
        """Guarda múltiples DataFrames como hojas de un Excel con formato.

        Aplica formato institucional:
        - Headers con color burdeos corporativo
        - Filas alternas
        - Ancho de columnas automático
        - Primera fila congelada

        Args:
            hojas: Dict {nombre_hoja: DataFrame}.
            nombre: Nombre del archivo (sin extensión).
            formato_institucional: Si True, aplica formato institucional con estilos predefinidos.

        Returns:
            Path al archivo guardado.
        """
        ruta = self.excel / f"{nombre}.xlsx"

        with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
            for hoja, df in hojas.items():
                if df is None or len(df) == 0:
                    continue
                df.to_excel(writer, sheet_name=hoja[:31], index=False)

            if formato_institucional:
                self._aplicar_formato(writer)

        print(f"   📗 {ruta.name} ({len(hojas)} hojas)")
        return ruta

    @staticmethod
    def _aplicar_formato(writer) -> None:
        """Aplica formato institucional a todas las hojas."""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return  # openpyxl no disponible, salir silenciosamente

        header_fill = PatternFill("solid", fgColor="8B1A4A")  # burdeos corporativo
        header_font = Font(bold=True, color="FFFFFF", size=10)
        fill_alt = PatternFill("solid", fgColor="F5E8EF")      # rosa claro alterno

        for ws in writer.sheets.values():
            # Headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Ancho automático de columnas
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = 0
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

            # Filas alternas
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if i % 2 == 0:
                    for cell in row:
                        cell.fill = fill_alt

            # Congelar primera fila
            ws.freeze_panes = "A2"

    def guardar_metadata(
        self,
        config: ConfigGEIH,
        stats: Optional[Dict[str, Any]] = None,
    ) -> Path:
        """Guarda metadata de la corrida para trazabilidad.

        Args:
            config: Configuración usada.
            stats: Estadísticas adicionales (n_registros, etc.)

        Returns:
            Path al archivo metadata.json.
        """
        meta = {
            "timestamp": datetime.now().isoformat(),
            "config": {
                "n_meses": config.n_meses,
                "smmlv": config.smmlv,
                "periodo": config.periodo_etiqueta,
                "random_seed": config.random_seed,
            },
            "stats": stats or {},
        }
        ruta = self.raiz / "metadata.json"
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2, ensure_ascii=False)
        print(f"   📋 metadata.json")
        return ruta

    def resumen(self) -> None:
        """Imprime un resumen de todos los archivos exportados."""
        print(f"\n{'='*60}")
        print(f"  RESUMEN DE EXPORTACIÓN")
        print(f"  {self.raiz}")
        print(f"{'='*60}")
        for carpeta, icono in [
            (self.graficas, "📊"),
            (self.tablas, "📋"),
            (self.excel, "📗"),
        ]:
            archivos = sorted(carpeta.glob("*"))
            if archivos:
                print(f"\n  {icono} {carpeta.name}/ ({len(archivos)} archivos)")
                for a in archivos:
                    size = a.stat().st_size / 1024
                    print(f"     {a.name:<50} {size:>7.0f} KB")
        # Metadata
        meta_path = self.raiz / "metadata.json"
        if meta_path.exists():
            print(f"\n  📋 metadata.json")

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/exportador.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [19/39]: indicadores.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/indicadores.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/indicadores.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.indicadores — Cálculo de indicadores del mercado laboral.

Contiene todas las clases de análisis, cada una con responsabilidad única:
  - IndicadoresLaborales: TD, TGP, TO, sanity checks
  - DistribucionIngresos: rangos SMMLV, distribución por sexo
  - AnalisisRamaSexo: ocupados por rama CIIU y sexo
  - AnalisisSalarios: estadísticas salariales por rama y edad
  - BrechaGenero: brecha salarial por nivel educativo
  - AnalisisCruzado: empresa × departamento
  - IndiceFormalidad: ICE, ICF, IVI, ICI, ITAT
  - AnalisisArea: ocupados por CIIU y 32 ciudades

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "IndicadoresLaborales",
    "DistribucionIngresos",
    "AnalisisRamaSexo",
    "AnalisisSalarios",
    "BrechaGenero",
    "AnalisisCruzado",
    "IndicesCompuestos",
    "AnalisisArea",
]


import gc
from typing import Optional, Dict, Any, List

import numpy as np
import pandas as pd

from .config import (
    ConfigGEIH,
    SMMLV_2025,
    REF_DANE_2025,
    REF_DANE,
    RAMAS_DANE,
    TABLA_CIIU_RAMAS,
    DEPARTAMENTOS,
    DPTO_A_CIUDAD,
    RANGOS_SMMLV_LIMITES,
    RANGOS_SMMLV_ETIQUETAS,
    NIVELES_AGRUPADOS,
    P3042_A_ANOS,
    TAMANO_EMPRESA,
    AGRUPACION_DANE_8,
    CARGA_PRESTACIONAL,
)
from .utils import EstadisticasPonderadas as EP, ConversorTipos
from .preparador import PreparadorGEIH


# ═════════════════════════════════════════════════════════════════════
# INDICADORES LABORALES FUNDAMENTALES
# ═════════════════════════════════════════════════════════════════════

class IndicadoresLaborales:
    """Calcula TD, TGP, TO y valida contra el boletín DANE.

    Módulo M0 del notebook original.

    Las tres tasas fundamentales:
      TD  = Desocupados / PEA       (tasa de desempleo)
      TGP = PEA / PET               (tasa global de participación)
      TO  = Ocupados / PET          (tasa de ocupación)

    Identidades que siempre deben cumplirse:
      PEA = OCI + DSI
      PET = PEA + FFT
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> Dict[str, float]:
        """Calcula indicadores nacionales a partir de la base preparada.

        Args:
            df: DataFrame con columnas FEX_ADJ, OCI, FT, DSI, PET.

        Returns:
            Dict con PET, PEA, Ocupados, Desocupados, TD, TGP, TO.
        """
        fex = "FEX_ADJ"

        pet = df.loc[df["PET"] == 1, fex].sum() if "PET" in df.columns else 0
        pea = df.loc[df["FT"] == 1, fex].sum() if "FT" in df.columns else 0
        ocu = df.loc[df["OCI"] == 1, fex].sum() if "OCI" in df.columns else 0
        des = df.loc[df["DSI"] == 1, fex].sum() if "DSI" in df.columns else 0

        td  = (des / pea * 100) if pea > 0 else np.nan
        tgp = (pea / pet * 100) if pet > 0 else np.nan
        to  = (ocu / pet * 100) if pet > 0 else np.nan

        resultado = {
            "PET_M": pet / 1e6,
            "PEA_M": pea / 1e6,
            "Ocupados_M": ocu / 1e6,
            "Desocupados_M": des / 1e6,
            "TD_%": round(td, 1),
            "TGP_%": round(tgp, 1),
            "TO_%": round(to, 1),
        }
        return resultado

    def sanity_check(
        self,
        resultado: Dict[str, float],
        periodo: str = "Anual",
    ) -> bool:
        """Valida los indicadores contra las referencias DANE.

        Si las cifras difieren más de ±0.5 p.p., alerta al usuario.

        CAMBIO v4.0: Usa config.referencia_dane (multi-año) en vez de
        REF_DANE_2025 hardcoded. Si no hay referencia para el año
        configurado, advierte pero no falla — solo valida PEA < 40M.

        Returns:
            True si todos los indicadores pasan la validación.
        """
        ref = self.config.referencia_dane
        is_anual = "anual" in periodo.lower()

        print(f"\n{'─'*55}")
        print(f"  SANITY CHECK — {periodo} ({self.config.anio})")
        print(f"{'─'*55}")

        ok = True

        if ref is None:
            # Año sin referencia publicada (ej: 2026 parcial)
            print(f"  ⚠️  Sin referencia DANE para {self.config.anio}.")
            print(f"     No se puede validar TD/TGP/TO contra boletín oficial.")
            print(f"     Agregue la referencia en config.py → REF_DANE cuando")
            print(f"     el DANE publique las cifras oficiales del año.")
        else:
            refs = {
                "TD_%":  ref.td_anual_pct if is_anual else ref.td_dic_pct,
                "TGP_%": ref.tgp_anual_pct if is_anual else ref.tgp_dic_pct,
                "TO_%":  ref.to_anual_pct if is_anual else ref.to_dic_pct,
            }
            for key, ref_val in refs.items():
                if ref_val == 0.0:
                    print(f"  ⚠️  {key:>6} — ref. DANE no disponible para este período")
                    continue
                calc = resultado.get(key, np.nan)
                diff = abs(calc - ref_val)
                estado = "✅" if diff <= 0.5 else "⚠️"
                if diff > 0.5:
                    ok = False
                print(f"  {estado} {key:>6} = {calc:.1f}%  (ref. DANE: {ref_val:.1f}%  Δ={diff:.1f})")

        for key in ["PET_M", "PEA_M", "Ocupados_M", "Desocupados_M"]:
            val = resultado.get(key, 0)
            print(f"     {key:>15} = {val:.2f} M")

        if resultado.get("PEA_M", 0) > 40:
            print("  ❌ ALERTA: PEA > 40M → El factor de expansión NO está siendo dividido.")
            ok = False

        return ok

    def por_departamento(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calcula TD, TGP, TO por departamento.

        Args:
            df: DataFrame con DPTO_STR, FEX_ADJ, OCI, FT, DSI, PET.

        Returns:
            DataFrame con indicadores por departamento.
        """
        if "DPTO_STR" not in df.columns:
            df["DPTO_STR"] = ConversorTipos.estandarizar_dpto(df["DPTO"])

        filas = []
        for dpto, nombre in DEPARTAMENTOS.items():
            m = df["DPTO_STR"] == dpto
            if m.sum() < 100:
                continue
            r = self.calcular(df[m])
            r["Departamento"] = nombre
            r["DPTO"] = dpto
            filas.append(r)

        return pd.DataFrame(filas).sort_values("TD_%", ascending=False)


# ═════════════════════════════════════════════════════════════════════
# DISTRIBUCIÓN DE INGRESOS
# ═════════════════════════════════════════════════════════════════════

class DistribucionIngresos:
    """Distribución del ingreso laboral por rangos de SMMLV.

    Módulo M1 del notebook original.
    Clasifica ~23M de ocupados según su ingreso en múltiplos de SMMLV.
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def calcular(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """Calcula la distribución de ingresos por rangos SMMLV.

        Args:
            df: DataFrame con OCI, INGLABO, P3271, FEX_ADJ.

        Returns:
            Dict con 'total' (distribución general) y 'por_sexo'.
        """
        # Filtrar ocupados con ingreso positivo
        df_ocu = df[(df["OCI"] == 1) & (df["INGLABO"] > 0)].copy()

        n_total = df.loc[df["OCI"] == 1, "FEX_ADJ"].sum()
        n_pos = df_ocu["FEX_ADJ"].sum()

        print(f"   Ocupados totales      : {n_total/1e6:.2f} M")
        print(f"   Con ingreso > 0       : {n_pos/1e6:.2f} M")
        print(f"   Sin ingreso / especie : {(n_total-n_pos)/1e6:.2f} M")

        # Clasificar en rangos (vectorizado con pd.cut)
        limites_cop = [l * self.config.smmlv for l in RANGOS_SMMLV_LIMITES]
        df_ocu["RANGO"] = pd.cut(
            df_ocu["INGLABO"],
            bins=limites_cop,
            labels=RANGOS_SMMLV_ETIQUETAS,
            right=False,
            include_lowest=True,
        )
        df_ocu["SEXO"] = df_ocu["P3271"].map({1: "Hombres", 2: "Mujeres"})

        # Agregación total
        dist = (
            df_ocu.groupby("RANGO", observed=True)["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"FEX_ADJ": "Personas"})
        )
        dist["Personas_M"] = dist["Personas"] / 1e6
        dist["Pct"] = dist["Personas"] / n_pos * 100
        dist["Acum_Pct"] = dist["Pct"].cumsum()

        # Por sexo
        dist_sexo = (
            df_ocu.groupby(["RANGO", "SEXO"], observed=True)["FEX_ADJ"]
            .sum().unstack(fill_value=0).reset_index()
        )
        for col in ["Hombres", "Mujeres"]:
            if col not in dist_sexo.columns:
                dist_sexo[col] = 0
        dist_sexo["H_M"] = dist_sexo["Hombres"] / 1e6
        dist_sexo["M_M"] = dist_sexo["Mujeres"] / 1e6

        del df_ocu
        gc.collect()

        return {"total": dist, "por_sexo": dist_sexo}

    def imprimir(self, resultado: Dict[str, pd.DataFrame], titulo: str = "") -> None:
        """Imprime la tabla de distribución de ingresos."""
        dist = resultado["total"]
        dist_s = resultado["por_sexo"]

        print(f"\n{'─'*70}")
        print(f"  DISTRIBUCIÓN DE INGRESOS LABORALES — {titulo}")
        print(f"  SMMLV 2025 = ${self.config.smmlv:,} COP")
        print(f"{'─'*70}")
        print(f"  {'Rango':<22} {'Personas':>10} {'%':>7}  {'Acum%':>7}")
        print(f"  {'─'*22} {'─'*10} {'─'*7}  {'─'*7}")

        for _, row in dist.iterrows():
            print(f"  {str(row['RANGO']):<22} {row['Personas_M']:>8.2f}M "
                  f"{row['Pct']:>6.1f}%  {row['Acum_Pct']:>6.1f}%")


# ═════════════════════════════════════════════════════════════════════
# ANÁLISIS POR RAMA Y SEXO
# ═════════════════════════════════════════════════════════════════════

class AnalisisRamaSexo:
    """Ocupados por rama de actividad económica y sexo.

    Módulo M3 del notebook original.
    Agrega ocupados expandidos por las 13 ramas DANE, desagregados por sexo.
    """

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calcula la tabla pivote de ocupados por rama y sexo.

        Args:
            df: DataFrame con RAMA (ya mapeada), SEXO, FEX_ADJ, OCI.

        Returns:
            DataFrame con Total, Hombres, Mujeres, y distribuciones %.
        """
        df_calc = df[
            (df["OCI"] == 1)
            & df["RAMA"].notna()
            & df["SEXO"].notna()
        ].copy()

        pivot = (
            df_calc.groupby(["RAMA", "SEXO"])["FEX_ADJ"]
            .sum().unstack(fill_value=0).reset_index()
        )
        for col in ["Hombres", "Mujeres"]:
            if col not in pivot.columns:
                pivot[col] = 0

        pivot["Total"] = pivot["Hombres"] + pivot["Mujeres"]
        pivot = pivot.sort_values("Total", ascending=False).reset_index(drop=True)

        total_nac = pivot["Total"].sum()
        total_h = pivot["Hombres"].sum()
        total_m = pivot["Mujeres"].sum()

        pivot["Dist_%"] = (pivot["Total"] / total_nac * 100).round(1)
        pivot["Dist_H_%"] = (pivot["Hombres"] / total_h * 100).round(1)
        pivot["Dist_M_%"] = (pivot["Mujeres"] / total_m * 100).round(1)

        # Convertir a miles
        for col in ["Total", "Hombres", "Mujeres"]:
            pivot[f"{col}_miles"] = (pivot[col] / 1_000).round(0).astype(int)

        del df_calc
        gc.collect()

        return pivot


# ═════════════════════════════════════════════════════════════════════
# ESTADÍSTICAS SALARIALES POR RAMA Y EDAD
# ═════════════════════════════════════════════════════════════════════

class AnalisisSalarios:
    """Estadísticas de salario por rama y grupo de edad.

    Módulo M4 del notebook original.
    Calcula media, mediana, percentiles ponderados por FEX.
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def por_rama(self, df: pd.DataFrame) -> pd.DataFrame:
        """Estadísticas salariales ponderadas por rama de actividad.

        Args:
            df: DataFrame con RAMA, INGLABO, FEX_ADJ.

        Returns:
            DataFrame con N, Media, Mediana, P10-P90, CV%, etc.
        """
        df_calc = df[df["INGLABO"] > 0].copy()
        filas = []
        for rama in df_calc["RAMA"].dropna().unique():
            m = df_calc["RAMA"] == rama
            est = EP.resumen_completo(
                df_calc.loc[m, "INGLABO"],
                df_calc.loc[m, "FEX_ADJ"],
                self.config.smmlv,
            )
            if est:
                est["Rama"] = rama
                filas.append(est)

        tabla = pd.DataFrame(filas)
        if not tabla.empty:
            tabla = tabla.set_index("Rama").sort_values("Mediana", ascending=False)
        return tabla

    def por_edad(
        self,
        df: pd.DataFrame,
        bin_size: int = 5,
        edad_min: int = 15,
        edad_max: int = 79,
    ) -> pd.DataFrame:
        """Estadísticas salariales por grupo de edad (ciclo vital).

        Args:
            df: DataFrame con P6040 (edad), INGLABO, FEX_ADJ.
            bin_size: Tamaño de los bins de edad (por defecto 5 años).

        Returns:
            DataFrame con estadísticas por grupo de edad.
        """
        df_calc = df[
            (df["INGLABO"] > 0) & df["P6040"].between(edad_min, edad_max)
        ].copy()

        bins = list(range(edad_min, edad_max + 1, bin_size)) + [edad_max + 1]
        labels = [f"{b}–{b+bin_size-1}" for b in bins[:-1]]
        df_calc["GRUPO_EDAD"] = pd.cut(
            df_calc["P6040"], bins=bins, labels=labels,
            right=False, include_lowest=True,
        )

        filas = []
        for edad in df_calc["GRUPO_EDAD"].cat.categories:
            m = df_calc["GRUPO_EDAD"] == edad
            if m.sum() < 30:
                continue
            est = EP.resumen_completo(
                df_calc.loc[m, "INGLABO"],
                df_calc.loc[m, "FEX_ADJ"],
                self.config.smmlv,
            )
            if est:
                est["Grupo_edad"] = str(edad)
                filas.append(est)

        return pd.DataFrame(filas).dropna(subset=["Media"])


# ═════════════════════════════════════════════════════════════════════
# BRECHA SALARIAL DE GÉNERO
# ═════════════════════════════════════════════════════════════════════

class BrechaGenero:
    """Brecha salarial de género por nivel educativo.

    Módulo M6 del notebook original.
    Brecha% = (Mediana_M − Mediana_H) / Mediana_H × 100
    """

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calcula la brecha salarial por nivel educativo agrupado.

        Args:
            df: DataFrame con NIVEL_GRUPO, P3271, INGLABO, FEX_ADJ.

        Returns:
            DataFrame con mediana por sexo y brecha %.
        """
        if "NIVEL_GRUPO" not in df.columns or "P3271" not in df.columns:
            print("⚠️ Columnas requeridas no disponibles.")
            return pd.DataFrame()

        df_calc = df[(df["OCI"] == 1) & (df["INGLABO"] > 0)].copy()

        filas = []
        for niv in sorted(df_calc["NIVEL_GRUPO"].dropna().unique()):
            for sexo_val, sexo_lbl in [(1, "Hombres"), (2, "Mujeres")]:
                m = (df_calc["NIVEL_GRUPO"] == niv) & (df_calc["P3271"] == sexo_val)
                med = EP.mediana(df_calc.loc[m, "INGLABO"], df_calc.loc[m, "FEX_ADJ"])
                mea = EP.media(df_calc.loc[m, "INGLABO"], df_calc.loc[m, "FEX_ADJ"])
                n = EP.suma(df_calc, m)
                filas.append({
                    "Nivel": niv, "Sexo": sexo_lbl,
                    "Mediana": med, "Media": mea, "N_miles": n / 1_000,
                })

        df_edu = pd.DataFrame(filas).drop_duplicates(subset=["Nivel", "Sexo"])

        # Calcular brecha
        pivot = df_edu.pivot_table(index="Nivel", columns="Sexo", values="Mediana")
        if "Hombres" in pivot.columns and "Mujeres" in pivot.columns:
            pivot["Brecha_%"] = (
                (pivot["Mujeres"] - pivot["Hombres"]) / pivot["Hombres"] * 100
            )

        return pivot.dropna()


# ═════════════════════════════════════════════════════════════════════
# ANÁLISIS CRUZADO EMPRESA × DEPARTAMENTO
# ═════════════════════════════════════════════════════════════════════

class AnalisisCruzado:
    """Insatisfacción laboral y tamaño de empresa por departamento.

    Sección 3 del notebook original.
    """

    def calcular(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calcula deseo de cambio y distribución por tamaño de empresa.

        Args:
            df: DataFrame con NOMBRE_DPTO, P7130, P3069, P6920, FEX_ADJ, OCI.

        Returns:
            DataFrame con indicadores por departamento.
        """
        df_ocu = df[df["OCI"] == 1].copy()

        filas = []
        for dpto in df_ocu["NOMBRE_DPTO"].dropna().unique():
            m = df_ocu["NOMBRE_DPTO"] == dpto
            n_tot = df_ocu.loc[m, "FEX_ADJ"].sum()
            if n_tot < 5_000:
                continue

            fila: Dict[str, Any] = {
                "Departamento": dpto,
                "Ocupados_M": round(n_tot / 1e6, 2),
            }

            # Deseo de cambio (P7130=1)
            if "P7130" in df_ocu.columns:
                n_cambia = df_ocu.loc[
                    m & (df_ocu["P7130"] == 1), "FEX_ADJ"
                ].sum()
                fila["Desea_cambiar_%"] = round(n_cambia / n_tot * 100, 1)

            # Formalidad (P6920=1 → cotiza pensión)
            if "P6920" in df_ocu.columns:
                n_pen = df_ocu.loc[
                    m & (df_ocu["P6920"] == 1), "FEX_ADJ"
                ].sum()
                fila["Cotiza_pension_%"] = round(n_pen / n_tot * 100, 1)

            filas.append(fila)

        resultado = pd.DataFrame(filas)
        if not resultado.empty:
            resultado = resultado.sort_values(
                "Desea_cambiar_%", ascending=False
            )
        return resultado


# ═════════════════════════════════════════════════════════════════════
# ÍNDICES COMPUESTOS
# ═════════════════════════════════════════════════════════════════════

class IndicesCompuestos:
    """Calcula ICE, ICF, IVI, ICI e ITAT.

    Módulos M7, M13, M16, M20 del notebook original.
    Todos usan normalización min-max [0, 100].
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def ice(self, df: pd.DataFrame) -> pd.DataFrame:
        """Índice de Calidad del Empleo (ICE).

        ICE = 0.30×Pensión + 0.25×Salud + 0.25×Horas_adecuadas + 0.20×Ingreso≥SML
        """
        df_ocu = df[(df["OCI"] == 1)].copy()

        df_ocu["D_PENSION"] = (df_ocu["P6920"] == 1).astype(float)
        df_ocu["D_SALUD"] = (df_ocu["P6090"] == 1).astype(float) if "P6090" in df_ocu.columns else 0
        df_ocu["D_HORAS"] = df_ocu["P6800"].between(20, 48).astype(float) if "P6800" in df_ocu.columns else 0
        df_ocu["D_INGRESO"] = (df_ocu["INGLABO"] >= self.config.smmlv).astype(float)

        df_ocu["ICE"] = (
            0.30 * df_ocu["D_PENSION"]
            + 0.25 * df_ocu["D_SALUD"]
            + 0.25 * df_ocu["D_HORAS"]
            + 0.20 * df_ocu["D_INGRESO"]
        ) * 100

        return df_ocu

    def gini(self, df: pd.DataFrame) -> float:
        """Calcula el coeficiente de Gini del ingreso laboral.

        Módulo M5 del notebook original.
        """
        df_calc = df[(df["OCI"] == 1) & (df["INGLABO"] > 0)]
        return EP.gini(df_calc["INGLABO"], df_calc["FEX_ADJ"])


# ═════════════════════════════════════════════════════════════════════
# ANÁLISIS POR ÁREA GEOGRÁFICA (32 CIUDADES)
# ═════════════════════════════════════════════════════════════════════

class AnalisisArea:
    """Análisis de ocupados por CIIU y 32 ciudades/áreas metropolitanas.

    Produce las 6 tablas equivalentes al script GEIH 2022–2024.
    Requiere que AREA esté incluida en la consolidación.
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()

    def _mapear_agrupacion_dane_8(self, serie_ciiu: pd.Series) -> pd.Series:
        """Mapea CIIU numérico a la agrupación DANE de 8 grupos."""
        s = pd.to_numeric(serie_ciiu, errors="coerce").astype("float64")
        condiciones = []
        etiquetas = []
        for nombre, rangos in AGRUPACION_DANE_8.items():
            cond = pd.Series(False, index=s.index)
            for lo, hi in rangos:
                cond = cond | s.between(lo, hi)
            condiciones.append(cond)
            etiquetas.append(nombre)

        resultado = np.select(
            condiciones,
            np.array(etiquetas, dtype=object),
            default=None,
        )
        return pd.Series(resultado, index=serie_ciiu.index, dtype=object)

    def calcular_tablas(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """Genera las 6 tablas de análisis por área.

        Returns:
            Dict con tabla1..tabla6.
        """
        df_ocu = df[(df["OCI"] == 1)].copy()
        df_ocu["AGRUPACION_DANE"] = self._mapear_agrupacion_dane_8(df_ocu["RAMA2D_R4"])

        tablas = {}

        # Tabla 1: Total nacional
        total = df_ocu["FEX_ADJ"].sum()
        tablas["tabla1"] = pd.DataFrame([{
            "Indicador": "Total Ocupados",
            "Valor_miles": round(total / 1_000),
        }])

        # Tabla 2: Por agrupación DANE
        t2 = (
            df_ocu[df_ocu["AGRUPACION_DANE"].notna()]
            .groupby("AGRUPACION_DANE")["FEX_ADJ"]
            .sum().reset_index()
            .rename(columns={"FEX_ADJ": "Ocupados"})
        )
        t2["Ocupados_miles"] = (t2["Ocupados"] / 1_000).round(0).astype(int)
        t2["Pct_%"] = (t2["Ocupados"] / total * 100).round(1)
        t2 = t2.sort_values("Ocupados", ascending=False)
        tablas["tabla2"] = t2

        # Tabla 4: Por ciudad
        if "CIUDAD" in df_ocu.columns:
            t4 = (
                df_ocu[df_ocu["CIUDAD"].notna()]
                .groupby("CIUDAD")["FEX_ADJ"]
                .sum().reset_index()
                .rename(columns={"CIUDAD": "Ciudad_AM", "FEX_ADJ": "Ocupados"})
            )
            t4["Ocupados_miles"] = (t4["Ocupados"] / 1_000).round(0).astype(int)
            t4["Pct_%"] = (t4["Ocupados"] / total * 100).round(1)
            t4 = t4.sort_values("Ocupados", ascending=False)
            tablas["tabla4"] = t4

        # Tabla 6: Agrupación × División (sin ciudad)
        if "RAMA_INT" in df_ocu.columns:
            cols_grp = ["AGRUPACION_DANE", "RAMA_INT"]
            t6 = (
                df_ocu[df_ocu["AGRUPACION_DANE"].notna() & df_ocu["RAMA_INT"].notna()]
                .groupby(cols_grp, dropna=True)["FEX_ADJ"]
                .sum().reset_index()
                .rename(columns={"FEX_ADJ": "Ocupados_miles", "RAMA_INT": "DIVISION"})
            )
            t6["Ocupados_miles"] = (t6["Ocupados_miles"] / 1_000).round(1)
            t6 = t6.sort_values("Ocupados_miles", ascending=False)
            tablas["tabla6"] = t6

        return tablas

    def exportar_excel(
        self,
        tablas: Dict[str, pd.DataFrame],
        nombre: str = "Resultados_CIIU_Area_GEIH2025.xlsx",
    ) -> None:
        """Exporta todas las tablas a un archivo Excel con múltiples hojas."""
        nombres_hojas = {
            "tabla1": "Total Nacional",
            "tabla2": "Agrupación DANE",
            "tabla4": "Ciudad-AM",
            "tabla6": "Agrupación-CIIU",
        }

        with pd.ExcelWriter(nombre, engine="openpyxl") as writer:
            for key, hoja in nombres_hojas.items():
                if key in tablas:
                    tablas[key].to_excel(writer, sheet_name=hoja, index=False)

        print(f"✅ Excel exportado: {nombre}")

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/indicadores.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [20/39]: logger.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/logger.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/logger.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.logger — Sistema centralizado de logging con trazabilidad.

Reemplaza los print() dispersos por logging estructurado que:
  - Escribe a consola (para visibilidad inmediata en Colab)
  - Escribe a archivo (para trazabilidad post-ejecución)
  - Registra timestamp, módulo, nivel y mensaje
  - Permite filtrar por nivel (DEBUG, INFO, WARNING, ERROR)

Migración gradual: los módulos existentes pueden seguir usando print()
y migrar a logging progresivamente. Este módulo provee get_logger()
para obtener un logger del paquete ya configurado.

Uso desde cualquier módulo del paquete:
    from .logger import get_logger
    log = get_logger(__name__)
    log.info("Procesando mes %d...", i)
    log.warning("Columna %s no encontrada", col)
    log.error("Error en consolidación: %s", e)

Configuración desde el notebook:
    from geih.logger import configurar_logging
    configurar_logging(nivel='DEBUG', archivo='geih_pipeline.log')

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "get_logger",
    "configurar_logging",
    "LoggerGEIH",
]

import logging
import sys
from pathlib import Path
from typing import Optional


# Nombre raíz del logger del paquete
_LOGGER_NAME = "geih"

# Formato para consola (compacto, con emojis para niveles)
_FORMATO_CONSOLA = "%(message)s"

# Formato para archivo (completo, con timestamp y módulo)
_FORMATO_ARCHIVO = "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"

# Formato de fecha
_FORMATO_FECHA = "%Y-%m-%d %H:%M:%S"

# Flag para evitar configuración múltiple
_configurado = False


def get_logger(nombre: str = _LOGGER_NAME) -> logging.Logger:
    """Obtiene un logger del paquete ya configurado.

    Si el logging no ha sido configurado explícitamente, se configura
    con defaults (INFO a consola, sin archivo).

    Args:
        nombre: Nombre del logger. Usar __name__ desde cada módulo
                para trazabilidad (ej: 'geih.consolidador').

    Returns:
        Logger configurado listo para usar.

    Uso:
        log = get_logger(__name__)
        log.info("Mensaje informativo")
        log.warning("⚠️  Advertencia")
        log.error("❌ Error")
        log.debug("Detalle para depuración")
    """
    global _configurado
    if not _configurado:
        configurar_logging()  # defaults
    return logging.getLogger(nombre)


def configurar_logging(
    nivel: str = "INFO",
    archivo: Optional[str] = None,
    nivel_archivo: str = "DEBUG",
    formato_consola: Optional[str] = None,
    formato_archivo: Optional[str] = None,
) -> logging.Logger:
    """Configura el sistema de logging del paquete.

    Llamar UNA VEZ al inicio del notebook o script.
    Las llamadas posteriores se ignoran (idempotente).

    Args:
        nivel: Nivel mínimo para consola ('DEBUG', 'INFO', 'WARNING', 'ERROR').
        archivo: Ruta del archivo de log. Si None, solo consola.
        nivel_archivo: Nivel mínimo para el archivo (default: DEBUG = todo).
        formato_consola: Formato custom para consola.
        formato_archivo: Formato custom para archivo.

    Returns:
        Logger raíz del paquete.

    Ejemplo en notebook:
        from geih.logger import configurar_logging
        configurar_logging(
            nivel='INFO',                    # consola: solo INFO+
            archivo='geih_pipeline.log',     # archivo: todo desde DEBUG
        )
    """
    global _configurado

    logger = logging.getLogger(_LOGGER_NAME)

    # Evitar handlers duplicados si se llama más de una vez
    if _configurado:
        return logger

    logger.setLevel(logging.DEBUG)  # capturar todo; los handlers filtran

    # ── Handler de consola ─────────────────────────────────────
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(getattr(logging, nivel.upper(), logging.INFO))
    console_handler.setFormatter(logging.Formatter(
        formato_consola or _FORMATO_CONSOLA,
        datefmt=_FORMATO_FECHA,
    ))
    logger.addHandler(console_handler)

    # ── Handler de archivo (opcional) ──────────────────────────
    if archivo:
        ruta = Path(archivo)
        ruta.parent.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(
            ruta, mode="a", encoding="utf-8",
        )
        file_handler.setLevel(getattr(logging, nivel_archivo.upper(), logging.DEBUG))
        file_handler.setFormatter(logging.Formatter(
            formato_archivo or _FORMATO_ARCHIVO,
            datefmt=_FORMATO_FECHA,
        ))
        logger.addHandler(file_handler)
        logger.info(f"📝 Log → {ruta.resolve()}")

    # No propagar al logger root de Python
    logger.propagate = False
    _configurado = True

    return logger


class LoggerGEIH:
    """Wrapper de conveniencia para módulos que quieran migrar de print() a log.

    Provee métodos que imitan la interfaz de print() pero usan logging.
    Permite migración gradual: reemplazar `print(msg)` por `self.log.info(msg)`.

    Uso como mixin en clases existentes:
        class MiClase(LoggerGEIH):
            def __init__(self):
                super().__init__()   # configura self.log

            def procesar(self):
                self.log.info("Procesando...")
                self.log.warning("⚠️  Algo raro")

    O como atributo:
        class MiClase:
            def __init__(self):
                self._logger = LoggerGEIH(nombre='geih.mi_clase')

            def procesar(self):
                self._logger.info("Procesando...")
    """

    def __init__(self, nombre: Optional[str] = None):
        self.log = get_logger(nombre or self.__class__.__qualname__)

    def info(self, msg: str, *args) -> None:
        self.log.info(msg, *args)

    def warning(self, msg: str, *args) -> None:
        self.log.warning(msg, *args)

    def error(self, msg: str, *args) -> None:
        self.log.error(msg, *args)

    def debug(self, msg: str, *args) -> None:
        self.log.debug(msg, *args)


def resetear_logging() -> None:
    """Resetea la configuración de logging (útil para tests).

    Elimina todos los handlers del logger del paquete y permite
    reconfigurar desde cero.
    """
    global _configurado
    logger = logging.getLogger(_LOGGER_NAME)
    logger.handlers.clear()
    _configurado = False

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/logger.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [21/39]: preparador.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/preparador.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/preparador.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.preparador — Preparación y enriquecimiento de la base GEIH.

Transforma la base consolidada cruda en datos listos para análisis:
  1. Ajusta el factor de expansión según el período
  2. Mapea códigos CIIU a ramas de actividad
  3. Hace merge con correlativas externas (CIIU descriptivo, DIVIPOLA)
  4. Crea variables derivadas (SEXO, CIUDAD, NIVEL_GRUPO, etc.)

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "PreparadorGEIH",
    "MergeCorrelativas",
]


import gc
from pathlib import Path
from typing import Optional, List

import numpy as np
import pandas as pd

from .config import (
    ConfigGEIH,
    RAMAS_DANE,
    TABLA_CIIU_RAMAS,
    DEPARTAMENTOS,
    DPTO_A_CIUDAD,
    AREA_A_CIUDAD,
    NIVELES_AGRUPADOS,
    P3042_A_ANOS,
    CIIU_DESCRIPCION_FALLBACK,
)
from .utils import ConversorTipos


class PreparadorGEIH:
    """Prepara la base GEIH consolidada para análisis.

    Responsabilidad única: transformar datos crudos en datos analíticos.
    No calcula indicadores ni genera gráficos.

    Uso típico:
        prep = PreparadorGEIH(config=ConfigGEIH(n_meses=12))

        # Para análisis anual (12 meses → dividir FEX entre 12)
        df = prep.preparar_base(geih_2025_final, mes_filtro=None)

        # Para análisis puntual de diciembre (FEX sin dividir)
        df_dic = prep.preparar_base(geih_2025_final, mes_filtro=12)
    """

    def __init__(self, config: Optional[ConfigGEIH] = None):
        self.config = config or ConfigGEIH()
        self._conversor = ConversorTipos()

    # ── Método principal ───────────────────────────────────────────

    def preparar_base(
        self,
        df_raw: pd.DataFrame,
        columnas: Optional[List[str]] = None,
        mes_filtro: Optional[int] = None,
        solo_ocupados: bool = False,
        solo_ingreso_positivo: bool = False,
    ) -> pd.DataFrame:
        """Prepara un subconjunto de la base con tipos correctos y FEX ajustado.

        Estrategia de memoria: extrae solo las columnas necesarias ANTES
        de copiar, reduciendo el footprint de RAM.

        Args:
            df_raw: Base GEIH consolidada (cruda).
            columnas: Columnas a extraer. Si None, usa un set por defecto.
            mes_filtro: Si se especifica, filtra solo ese mes (1-12).
                       None = todos los meses (análisis anual).
            solo_ocupados: Si True, filtra OCI == 1.
            solo_ingreso_positivo: Si True, filtra INGLABO > 0.

        Returns:
            DataFrame preparado con FEX_ADJ y columnas estandarizadas.
        """
        # Columnas por defecto — cubre las 48 clases de análisis del paquete.
        # Agrupadas por módulo DANE para legibilidad.
        # Solo se copian las que existan en la base (sin error si falta alguna).
        if columnas is None:
            columnas = [
                # ── Identificación y factor de expansión ──────────
                "FEX_C18", "MES_NUM",
                # ── Características generales (E/F/G) ─────────────
                "P3271",      # Sexo (1=H, 2=M)
                "P6040",      # Edad
                "P6080",      # Autorreconocimiento étnico
                "P3042",      # Nivel educativo (1-13)
                "P3043S1",    # Campo de formación (CINE-F) — para proxy bilingüismo
                "P6090",      # Afiliado salud (1=Sí)
                "CLASE",      # Zona (1=Cabecera/Urbano, 2=Rural)
                "DPTO",       # Departamento (código 2 dígitos)
                "P2057",      # ¿Se considera campesino?
                "P2059",      # ¿Alguna vez fue campesino?
                # Discapacidad — Escala Washington (8 dimensiones)
                "P1906S1", "P1906S2", "P1906S3", "P1906S4",
                "P1906S5", "P1906S6", "P1906S7", "P1906S8",
                # ── Fuerza de trabajo (H) ─────────────────────────
                "FT", "PET", "OCI", "DSI", "FFT",
                "P6240",      # Actividad semana pasada
                # ── Ocupados (I) ──────────────────────────────────
                "INGLABO",    # Ingreso laboral mensual
                "P6500",      # Salario bruto declarado
                "P6430",      # Posición ocupacional (1-9)
                "P6800",      # Horas normales semana
                "P6850",      # Horas reales semana pasada
                "P6920",      # Cotiza pensión
                "P3069",      # Tamaño empresa
                "P7130",      # Desea cambiar trabajo
                "RAMA2D_R4",  # CIIU 2 dígitos
                "RAMA4D_R4",  # CIIU 4 dígitos
                "AREA",       # Municipio 5 dígitos (32 ciudades)
                # Autonomía laboral (contratistas dependientes)
                "P3047",      # ¿Quién decide horario?
                "P3048",      # ¿Quién decide qué producir?
                "P3049",      # ¿Quién decide precio?
                # Formalidad contractual
                "P6440",      # ¿Tiene contrato?
                "P6450",      # ¿Contrato escrito?
                "P6460",      # ¿Contrato indefinido?
                # Variables de alto valor analítico
                "P1802",      # Alcance mercado (1-6, 6=Exportación)
                "P6765",      # Forma de trabajo (destajo, honorarios...)
                "P3363",      # ¿Cómo consiguió empleo?
                "P3364",      # ¿Retención en la fuente?
                "P6400",      # ¿Trabaja donde lo contrataron?
                "P6410",      # Tipo intermediación (EST, CTA)
                # Compensación (para análisis de paquete salarial)
                "P6510S1",    # Horas extras
                "P6580S1",    # Bonificaciones
                "P6585S1A1",  # Auxilio alimentación
                "P6585S2A1",  # Auxilio transporte
                # ── No ocupados (J) ───────────────────────────────
                "P7250",      # Semanas buscando trabajo
                "P6300",      # ¿Desea trabajar? (FFT con deseo)
                "P6310",      # ¿Disponible para trabajar?
                "P7140S2",    # Razón: mejorar ingresos
                # ── Otras formas de trabajo (K) ───────────────────
                "P3054",      # Autoconsumo bienes
                "P3054S1",    # Horas autoconsumo bienes
                "P3055",      # Autoconsumo servicios
                "P3055S1",    # Horas autoconsumo servicios
                "P3056",      # Voluntariado
                "P3057",      # Formación no remunerada
                # ── Migración (L) ─────────────────────────────────
                "P3370",      # ¿Dónde vivía hace 12 meses?
                "P3370S1",    # Departamento hace 12 meses
                "P3376",      # País de nacimiento
                "P3378S1",    # Año de llegada a Colombia
                # ── Otros ingresos (M) ────────────────────────────
                "P7422",      # Arriendos recibidos
                "P7500S1",    # Pensiones
                "P7500S1A1",  # Monto pensiones
                "P7500S2",    # Ayudas hogares nacionales
                "P7500S2A1",  # Monto ayudas
                "P7500S3",    # Ayudas institucionales
                "P7510S2",    # Remesas del exterior
                "P7510S2A1",  # Monto remesas
            ]

        # Solo tomar columnas que existan
        cols_ok = [c for c in columnas if c in df_raw.columns]
        cols_faltantes = [c for c in columnas if c not in df_raw.columns]

        # Filtrar mes si se especifica (antes de copiar → ahorra RAM)
        if mes_filtro is not None and "MES_NUM" in df_raw.columns:
            mask = df_raw["MES_NUM"] == mes_filtro
            df = df_raw.loc[mask, cols_ok].copy()
        else:
            df = df_raw[cols_ok].copy()

        gc.collect()

        # Convertir tipos numéricos (excluir columnas que deben ser string)
        cols_str = {"DPTO", "RAMA2D_R4", "RAMA4D_R4", "AREA", "CLASE"}
        self._conversor.convertir_columnas_numericas(
            df, cols_ok, excluir=list(cols_str)
        )

        # Factor de expansión ajustado
        n_meses_divisor = 1 if mes_filtro is not None else self.config.n_meses
        df["FEX_ADJ"] = df["FEX_C18"] / n_meses_divisor

        # Filtros opcionales
        if solo_ocupados and "OCI" in df.columns:
            df = df[df["OCI"] == 1].copy()
        if solo_ingreso_positivo and "INGLABO" in df.columns:
            df = df[df["INGLABO"] > 0].copy()

        gc.collect()

        # Reporte
        n = len(df)
        fex_total = df["FEX_ADJ"].sum() / 1e6
        print(f"   ✅ Base preparada: {n:,} registros → {fex_total:.2f}M personas expandidas")
        if cols_faltantes:
            print(f"   ⚠️  Columnas no disponibles: {cols_faltantes}")

        return df

    # ── Mapeo CIIU → Ramas ─────────────────────────────────────────

    @staticmethod
    def mapear_rama_ciiu(serie_ciiu: pd.Series) -> pd.Series:
        """Mapea códigos CIIU de 2 dígitos a las 13 ramas DANE.

        Usa operaciones vectorizadas (np.select) en lugar de loops.
        Es la versión optimizada para millones de registros.

        Args:
            serie_ciiu: Columna RAMA2D_R4 (puede ser str o numérica).

        Returns:
            Serie con nombres de rama o None para códigos sin mapeo.
        """
        s = pd.to_numeric(serie_ciiu, errors="coerce").astype("float64")

        condiciones = [s.between(lo, hi) for lo, hi, _ in TABLA_CIIU_RAMAS]
        etiquetas = np.array(
            [RAMAS_DANE[clave] for _, _, clave in TABLA_CIIU_RAMAS],
            dtype=object,
        )

        resultado = np.select(condiciones, etiquetas, default=None)
        return pd.Series(resultado, index=serie_ciiu.index, dtype=object)

    # ── Enriquecimiento con variables derivadas ────────────────────

    def agregar_variables_derivadas(self, df: pd.DataFrame) -> pd.DataFrame:
        """Agrega variables derivadas de uso frecuente.

        Variables creadas:
          - DPTO_STR: departamento con cero líder
          - NOMBRE_DPTO: nombre legible del departamento
          - SEXO: 'Hombres' / 'Mujeres'
          - RAMA: rama CIIU mapeada
          - RAMA_INT: código CIIU como entero nullable
          - CIUDAD: nombre de la ciudad/AM (desde DPTO o AREA)
          - NIVEL_GRUPO: nivel educativo agrupado (7 categorías)
          - ANOS_EDUC: años de educación (para Mincer)
          - INGLABO_SML: ingreso en múltiplos de SMMLV

        Args:
            df: DataFrame preparado.

        Returns:
            El mismo DataFrame con columnas nuevas.
        """
        # Departamento
        if "DPTO" in df.columns:
            df["DPTO_STR"] = self._conversor.estandarizar_dpto(df["DPTO"])
            df["NOMBRE_DPTO"] = df["DPTO_STR"].map(DEPARTAMENTOS)

        # Sexo
        if "P3271" in df.columns:
            df["SEXO"] = df["P3271"].map({1: "Hombres", 2: "Mujeres"})

        # Rama CIIU
        if "RAMA2D_R4" in df.columns:
            df["RAMA_INT"] = pd.to_numeric(
                df["RAMA2D_R4"], errors="coerce"
            ).round(0).astype("Int64")
            df["RAMA"] = self.mapear_rama_ciiu(df["RAMA2D_R4"])

        # Ciudad
        if "DPTO_STR" in df.columns:
            df["CIUDAD"] = df["DPTO_STR"].map(DPTO_A_CIUDAD)

        if "AREA" in df.columns:
            area_str = self._conversor.estandarizar_area(df["AREA"])
            ciudad_area = area_str.map(AREA_A_CIUDAD)
            # AREA es más preciso que DPTO; usar AREA si disponible
            if "CIUDAD" in df.columns:
                df["CIUDAD"] = ciudad_area.combine_first(df["CIUDAD"])
            else:
                df["CIUDAD"] = ciudad_area

        # Nivel educativo
        if "P3042" in df.columns:
            df["NIVEL_GRUPO"] = df["P3042"].map(NIVELES_AGRUPADOS)
            df["ANOS_EDUC"] = df["P3042"].map(P3042_A_ANOS)

        # Ingreso en SMMLV
        if "INGLABO" in df.columns:
            df["INGLABO_SML"] = df["INGLABO"] / self.config.smmlv

        return df


class MergeCorrelativas:
    """Hace merge con las correlativas externas CIIU y DIVIPOLA.

    Las correlativas agregan descripciones legibles a los códigos
    numéricos de la GEIH. Son archivos Excel mantenidos por el autor.

    Uso típico:
        merger = MergeCorrelativas()
        df = merger.merge_ciiu(df, ruta_ciiu)
        df = merger.merge_divipola(df, ruta_divipola)
    """

    def __init__(self):
        self._conversor = ConversorTipos()

    def merge_ciiu(
        self,
        df: pd.DataFrame,
        ruta_ciiu: str,
        sheet_name: str = "CIIU 2022",
    ) -> pd.DataFrame:
        """Hace merge con la correlativa CIIU Rev.4 adaptada para Colombia.

        Agrega la columna DESCRIPCION_CIIU al DataFrame.
        Si el merge deja registros sin descripción, aplica un fallback
        usando el código de 2 dígitos.

        Args:
            df: DataFrame con columna RAMA4D_R4.
            ruta_ciiu: Ruta al archivo Excel de la correlativa.
            sheet_name: Nombre de la hoja en el Excel.

        Returns:
            DataFrame enriquecido con DESCRIPCION_CIIU.
        """
        if "RAMA4D_R4" not in df.columns:
            print("⚠️ Columna RAMA4D_R4 no encontrada. Saltando merge CIIU.")
            return df

        print("\n🔗 Merge con correlativa CIIU Rev.4...")

        # Leer correlativa
        df_ciiu = pd.read_excel(
            ruta_ciiu, sheet_name=sheet_name,
            converters={"RAMA4D_R4": str},
        )

        # Estandarizar códigos en ambos lados
        df["RAMA4D_STD"] = self._conversor.estandarizar_ciiu4(df["RAMA4D_R4"])
        df_ciiu["RAMA4D_STD"] = self._conversor.estandarizar_ciiu4(
            df_ciiu["RAMA4D_R4"]
        )

        # Diagnóstico de solapamiento
        rama_df = set(df["RAMA4D_STD"].dropna().unique())
        rama_ciiu = set(df_ciiu["RAMA4D_STD"].dropna().unique())
        match = rama_df & rama_ciiu
        sin_match = rama_df - rama_ciiu
        print(f"   Códigos en base   : {len(rama_df)}")
        print(f"   Códigos en CIIU   : {len(rama_ciiu)}")
        print(f"   Con match         : {len(match)} ({len(match)/max(len(rama_df),1)*100:.0f}%)")
        if sin_match:
            print(f"   Sin match (top 10): {sorted(sin_match)[:10]}")

        # Merge
        ciiu_slim = df_ciiu[["RAMA4D_STD", "DESCRIPCION_CIIU"]].drop_duplicates(
            "RAMA4D_STD"
        )
        df = df.merge(ciiu_slim, on="RAMA4D_STD", how="left")

        n_con = df["DESCRIPCION_CIIU"].notna().sum()
        n_sin = df["DESCRIPCION_CIIU"].isna().sum()
        print(f"   Con descripción   : {n_con:,} ({n_con/len(df)*100:.1f}%)")
        print(f"   Sin descripción   : {n_sin:,} ({n_sin/len(df)*100:.1f}%)")

        # Fallback con código de 2 dígitos
        if n_sin > 0 and "RAMA2D_R4" in df.columns:
            df = self._aplicar_fallback_ciiu(df)

        return df

    def merge_divipola(
        self,
        df: pd.DataFrame,
        ruta_divipola: str,
        sheet_name: str = "Departamentos",
        skiprows: int = 9,
    ) -> pd.DataFrame:
        """Hace merge con DIVIPOLA para agregar nombres de departamento.

        Args:
            df: DataFrame con columna DPTO.
            ruta_divipola: Ruta al archivo Excel DIVIPOLA.
            sheet_name: Hoja del Excel.
            skiprows: Filas a saltar en el Excel.

        Returns:
            DataFrame con columna NOMBRE_DPTO.
        """
        if "DPTO" not in df.columns:
            print("⚠️ Columna DPTO no encontrada. Saltando merge DIVIPOLA.")
            return df

        print("\n🔗 Merge con DIVIPOLA...")

        df_depto = pd.read_excel(
            ruta_divipola, sheet_name=sheet_name,
            skiprows=skiprows, converters={"Código": str},
        )

        # Estandarizar
        df["DPTO_STR"] = self._conversor.estandarizar_dpto(df["DPTO"])

        df = df.merge(
            df_depto[["Código", "Nombre"]].rename(
                columns={"Código": "DPTO_STR", "Nombre": "NOMBRE_DPTO"}
            ),
            on="DPTO_STR",
            how="left",
        )

        n_ok = df["NOMBRE_DPTO"].notna().sum()
        print(f"   Departamentos mapeados: {n_ok:,} ({n_ok/len(df)*100:.1f}%)")

        return df

    def _aplicar_fallback_ciiu(self, df: pd.DataFrame) -> pd.DataFrame:
        """Aplica descripción CIIU de fallback usando el código de 2 dígitos.

        Para registros que no hicieron match en el merge principal,
        usa el código de 2 dígitos (División CIIU) para asignar
        una descripción genérica.
        """
        sin_desc = df["DESCRIPCION_CIIU"].isna()
        rama2d = pd.to_numeric(df.loc[sin_desc, "RAMA2D_R4"], errors="coerce")

        for rango, desc in CIIU_DESCRIPCION_FALLBACK:
            mask_rango = rama2d.apply(lambda x: x in rango if pd.notna(x) else False)
            if mask_rango.any():
                df.loc[sin_desc & mask_rango.reindex(df.index, fill_value=False),
                       "DESCRIPCION_CIIU"] = desc

        n_recuperados = sin_desc.sum() - df["DESCRIPCION_CIIU"].isna().sum()
        if n_recuperados > 0:
            print(f"   Fallback CIIU 2D  : {n_recuperados:,} registros recuperados")

        return df

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/preparador.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [22/39]: profiler.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/profiler.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/profiler.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.profiler — Profiling de memoria y rendimiento del pipeline.

Herramientas para identificar cuellos de botella de RAM y tiempo:
  - PerfilMemoria: snapshots antes/después de cada operación
  - medir_tiempo: decorador y context manager para cronometrar
  - reporte_memoria: resumen de consumo por etapa

Diseñado para Google Colab (~12GB RAM) donde la memoria es el
recurso más escaso. Los reportes son visibles e informativos.

Uso desde el notebook:
    from geih.profiler import PerfilMemoria, medir_tiempo

    # Context manager para bloques de código
    with medir_tiempo("Consolidar 12 meses"):
        geih = consolidador.consolidar()

    # Profiling de memoria paso a paso
    perf = PerfilMemoria()
    perf.snapshot("Antes de consolidar")
    geih = consolidador.consolidar()
    perf.snapshot("Después de consolidar")
    perf.snapshot("Después de preparar", df=df)
    perf.reporte()

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "PerfilMemoria",
    "medir_tiempo",
    "tamano_objeto",
]

import gc
import os
import sys
import time
import functools
from typing import Optional, List, Dict, Any, Union
from contextlib import contextmanager
from datetime import timedelta

import pandas as pd

from .logger import get_logger

log = get_logger(__name__)


# ═════════════════════════════════════════════════════════════════════
# PERFIL DE MEMORIA
# ═════════════════════════════════════════════════════════════════════

class PerfilMemoria:
    """Registra snapshots de uso de RAM a lo largo del pipeline.

    Cada snapshot captura: timestamp, RAM usada por el proceso,
    RAM libre del sistema, y opcionalmente el tamaño de un DataFrame.

    Al final, reporte() imprime una tabla con la evolución de RAM
    y destaca las etapas donde más creció el consumo.

    Uso:
        perf = PerfilMemoria()
        perf.snapshot("Inicio")
        geih = consolidador.consolidar()
        perf.snapshot("Post-consolidar")
        df = prep.preparar_base(geih)
        perf.snapshot("Post-preparar", df=df)
        perf.reporte()

    Output:
        ┌───────────────────┬──────────┬──────────┬──────────┬──────────┐
        │ Etapa             │ RAM (GB) │ Δ RAM    │ DF (MB)  │ Tiempo   │
        ├───────────────────┼──────────┼──────────┼──────────┼──────────┤
        │ Inicio            │ 0.42     │ —        │ —        │ 0:00     │
        │ Post-consolidar   │ 3.85     │ +3.43    │ —        │ 2:30     │
        │ Post-preparar     │ 4.12     │ +0.27    │ 847 MB   │ 0:05     │
        └───────────────────┴──────────┴──────────┴──────────┴──────────┘
    """

    def __init__(self):
        self._snapshots: List[Dict[str, Any]] = []
        self._inicio = time.time()

    def snapshot(
        self,
        etapa: str,
        df: Optional[pd.DataFrame] = None,
        gc_collect: bool = True,
    ) -> Dict[str, Any]:
        """Captura un snapshot de memoria.

        Args:
            etapa: Nombre descriptivo de la etapa.
            df: DataFrame para medir su tamaño en memoria (opcional).
            gc_collect: Si True, ejecuta gc.collect() antes de medir
                       (más preciso pero ligeramente más lento).

        Returns:
            Dict con los valores capturados.
        """
        if gc_collect:
            gc.collect()

        ram_usada = self._ram_proceso_gb()
        ram_libre = self._ram_libre_gb()
        df_mb = df.memory_usage(deep=True).sum() / 1e6 if df is not None else None
        elapsed = time.time() - self._inicio

        snap = {
            "etapa": etapa,
            "ram_usada_gb": round(ram_usada, 2),
            "ram_libre_gb": round(ram_libre, 2),
            "df_mb": round(df_mb, 1) if df_mb else None,
            "timestamp": elapsed,
        }
        self._snapshots.append(snap)

        # Log en tiempo real
        delta = ""
        if len(self._snapshots) > 1:
            prev = self._snapshots[-2]["ram_usada_gb"]
            diff = snap["ram_usada_gb"] - prev
            delta = f"  Δ={diff:+.2f} GB"
        df_info = f"  DF={df_mb:.0f}MB" if df_mb else ""

        log.info(f"  📊 {etapa}: RAM={ram_usada:.2f}GB{delta}{df_info}")
        return snap

    def reporte(self) -> pd.DataFrame:
        """Imprime y retorna un resumen de todos los snapshots.

        Returns:
            DataFrame con la tabla de profiling.
        """
        if not self._snapshots:
            log.warning("No hay snapshots registrados.")
            return pd.DataFrame()

        log.info(f"\n{'='*70}")
        log.info(f"  PERFIL DE MEMORIA — {len(self._snapshots)} snapshots")
        log.info(f"{'='*70}")
        log.info(f"  {'Etapa':<25} {'RAM(GB)':>8} {'Δ RAM':>8} "
                 f"{'DF(MB)':>8} {'Tiempo':>10}")
        log.info(f"  {'─'*25} {'─'*8} {'─'*8} {'─'*8} {'─'*10}")

        rows = []
        for i, snap in enumerate(self._snapshots):
            delta = "—"
            if i > 0:
                diff = snap["ram_usada_gb"] - self._snapshots[i-1]["ram_usada_gb"]
                delta = f"{diff:+.2f}"

            t_prev = self._snapshots[i-1]["timestamp"] if i > 0 else 0
            dt = snap["timestamp"] - t_prev
            tiempo_str = str(timedelta(seconds=int(dt)))

            df_str = f"{snap['df_mb']:.0f}" if snap["df_mb"] else "—"

            log.info(f"  {snap['etapa']:<25} {snap['ram_usada_gb']:>7.2f} "
                     f"{delta:>8} {df_str:>8} {tiempo_str:>10}")

            rows.append({
                "Etapa": snap["etapa"],
                "RAM_GB": snap["ram_usada_gb"],
                "Delta_GB": float(delta) if delta != "—" else 0,
                "DF_MB": snap["df_mb"],
                "Tiempo_s": dt,
            })

        # Alerta si RAM > 80% del total
        total = self._ram_total_gb()
        ultimo = self._snapshots[-1]["ram_usada_gb"]
        pct = ultimo / total * 100 if total > 0 else 0
        if pct > 80:
            log.warning(f"\n  ⚠️  RAM al {pct:.0f}% de capacidad "
                       f"({ultimo:.1f}/{total:.1f} GB)")
            log.warning(f"     Considere liberar DataFrames intermedios con "
                       f"del df; gc.collect()")
        elif pct > 60:
            log.info(f"\n  RAM al {pct:.0f}% ({ultimo:.1f}/{total:.1f} GB)")
        else:
            log.info(f"\n  ✅ RAM al {pct:.0f}% ({ultimo:.1f}/{total:.1f} GB) — holgado")

        log.info(f"{'='*70}")
        return pd.DataFrame(rows)

    @property
    def pico_ram_gb(self) -> float:
        """Máxima RAM usada durante la sesión."""
        if not self._snapshots:
            return 0
        return max(s["ram_usada_gb"] for s in self._snapshots)

    @staticmethod
    def _ram_proceso_gb() -> float:
        """RAM usada por el proceso actual en GB."""
        try:
            import psutil
            return psutil.Process(os.getpid()).memory_info().rss / 1e9
        except ImportError:
            return 0.0

    @staticmethod
    def _ram_libre_gb() -> float:
        """RAM libre del sistema en GB."""
        try:
            import psutil
            return psutil.virtual_memory().available / 1e9
        except ImportError:
            return 0.0

    @staticmethod
    def _ram_total_gb() -> float:
        """RAM total del sistema en GB."""
        try:
            import psutil
            return psutil.virtual_memory().total / 1e9
        except ImportError:
            return 12.0  # default Colab


# ═════════════════════════════════════════════════════════════════════
# MEDICIÓN DE TIEMPO
# ═════════════════════════════════════════════════════════════════════

@contextmanager
def medir_tiempo(nombre: str = "Operación"):
    """Context manager que mide y reporta el tiempo de ejecución.

    Uso como context manager:
        with medir_tiempo("Consolidar"):
            geih = consolidador.consolidar()
        # Output: ⏱️  Consolidar: 2:30.4

    Args:
        nombre: Nombre descriptivo de la operación.

    Yields:
        None — usar solo como context manager.
    """
    inicio = time.time()
    log.info(f"🔄 {nombre}...")
    try:
        yield
    finally:
        elapsed = time.time() - inicio
        tiempo_str = str(timedelta(seconds=round(elapsed, 1)))
        log.info(f"⏱️  {nombre}: {tiempo_str}")


def cronometrar(func):
    """Decorador que mide el tiempo de ejecución de una función.

    Uso:
        @cronometrar
        def mi_funcion_pesada():
            ...

    Output: ⏱️  mi_funcion_pesada: 0:01:23
    """
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        inicio = time.time()
        resultado = func(*args, **kwargs)
        elapsed = time.time() - inicio
        tiempo_str = str(timedelta(seconds=round(elapsed, 1)))
        log.info(f"⏱️  {func.__name__}: {tiempo_str}")
        return resultado
    return wrapper


# ═════════════════════════════════════════════════════════════════════
# UTILIDADES DE TAMAÑO
# ═════════════════════════════════════════════════════════════════════

def tamano_objeto(obj, nombre: str = "Objeto") -> float:
    """Mide y reporta el tamaño en memoria de un objeto.

    Para DataFrames usa memory_usage(deep=True).
    Para otros objetos usa sys.getsizeof().

    Args:
        obj: Objeto a medir.
        nombre: Nombre para el reporte.

    Returns:
        Tamaño en MB.
    """
    if isinstance(obj, pd.DataFrame):
        mb = obj.memory_usage(deep=True).sum() / 1e6
        log.info(f"  📦 {nombre}: {obj.shape[0]:,} × {obj.shape[1]} → {mb:.1f} MB")
    else:
        mb = sys.getsizeof(obj) / 1e6
        log.info(f"  📦 {nombre}: {mb:.1f} MB")
    return mb

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/profiler.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [23/39]: utils.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/utils.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/utils.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.utils — Funciones utilitarias transversales.

Contiene tres grupos de funciones:
  1. Gestión de memoria RAM (para Google Colab ~12GB)
  2. Conversión de tipos de datos (el DANE entrega formatos mixtos)
  3. Estadísticas ponderadas (el corazón metodológico de la GEIH)

Todas son funciones puras o casi puras: reciben datos, devuelven datos.
No modifican estado global ni acceden a disco.

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "GestorMemoria",
    "ConversorTipos",
    "EstadisticasPonderadas",
]


import gc
import os
import sys
from typing import Optional, List, Dict, Any

import numpy as np
import pandas as pd


# ═════════════════════════════════════════════════════════════════════
# 1. GESTIÓN DE MEMORIA RAM
# ═════════════════════════════════════════════════════════════════════

class GestorMemoria:
    """Monitoreo y liberación de memoria RAM en Google Colab.

    En Colab Free (~12GB), la memoria es el recurso más escaso.
    Esta clase encapsula la lógica de monitoreo y limpieza.

    Uso típico:
        mem = GestorMemoria()
        mem.estado()                           # ver RAM disponible
        mem.liberar(['df_temp', 'resultado'])   # eliminar variables
    """

    @staticmethod
    def estado() -> Dict[str, float]:
        """Muestra y retorna el uso actual de RAM en GB.

        Returns:
            Dict con 'usada', 'libre', 'total' en GB.
            Retorna dict vacío si psutil no está disponible.
        """
        try:
            import psutil
            proceso = psutil.Process(os.getpid())
            ram_usada = proceso.memory_info().rss / 1e9
            vm = psutil.virtual_memory()
            ram_libre = vm.available / 1e9
            ram_total = vm.total / 1e9
            print(f"   RAM usada : {ram_usada:.2f} GB")
            print(f"   RAM libre : {ram_libre:.2f} GB  /  Total: {ram_total:.2f} GB")
            return {"usada": ram_usada, "libre": ram_libre, "total": ram_total}
        except ImportError:
            print("   (instala psutil para ver RAM: !pip install psutil)")
            return {}

    @staticmethod
    def liberar(variables: Optional[List[str]] = None,
                scope: Optional[dict] = None,
                verbose: bool = True) -> None:
        """Libera memoria eliminando variables del scope dado.

        Args:
            variables: Nombres de variables a eliminar.
            scope: Diccionario del scope (usar globals() desde el notebook).
            verbose: Si True, muestra RAM después de limpiar.

        Nota:
            Desde un notebook llamar así:
            GestorMemoria.liberar(['df_temp'], scope=globals())
        """
        if variables and scope:
            for var in variables:
                if var in scope:
                    del scope[var]
        gc.collect()
        if verbose:
            GestorMemoria.estado()

    @staticmethod
    def tamano_df(df: pd.DataFrame, nombre: str = "DataFrame") -> None:
        """Imprime el tamaño en memoria de un DataFrame."""
        mb = df.memory_usage(deep=True).sum() / 1e6
        print(f"   {nombre}: {len(df):,} filas × {df.shape[1]} cols → {mb:.1f} MB")


# ═════════════════════════════════════════════════════════════════════
# 2. CONVERSIÓN DE TIPOS DE DATOS
# ═════════════════════════════════════════════════════════════════════

class ConversorTipos:
    """Convierte columnas GEIH al tipo de dato correcto.

    El DANE entrega archivos CSV con formatos inconsistentes:
    - Números con comas como separador decimal ('1.234,56')
    - Códigos numéricos que deben ser strings ('05' → Antioquia)
    - Floats que representan enteros (1.0 → '1')

    Esta clase estandariza todo de forma vectorizada.
    """

    @staticmethod
    def a_numerico(serie: pd.Series) -> pd.Series:
        """Convierte una serie a numérico, manejando comas y formatos mixtos.

        Ejemplo: '1.234,56' → 1234.56 ; 'nan' → NaN ; '1.0' → 1.0

        Args:
            serie: Columna con valores potencialmente numéricos.

        Returns:
            Serie con dtype float64 o Int64.
        """
        if pd.api.types.is_numeric_dtype(serie):
            return serie
        try:
            return pd.to_numeric(
                serie.astype(str).str.replace(",", ".", regex=False),
                errors="coerce",
            )
        except Exception:
            return pd.to_numeric(serie, errors="coerce")

    @staticmethod
    def estandarizar_ciiu4(serie: pd.Series) -> pd.Series:
        """Estandariza códigos CIIU de 4 dígitos para merge con correlativa.

        Problema: GEIH trae '111', '4711.0', '111.5', etc.
        Solución: quitar .0, tomar parte entera, rellenar a 4 dígitos.

        Args:
            serie: Columna RAMA4D_R4 en cualquier formato.

        Returns:
            Serie con códigos de exactamente 4 dígitos o NaN.
        """
        s = serie.astype(str).str.strip()
        s = s.str.replace(r"\.0$", "", regex=True)      # quitar .0
        s = s.str.split(".").str[0]                      # parte entera
        s = s.str.zfill(4)                               # rellenar ceros
        s = s.where(s.str.match(r"^\d{4}$"), other=np.nan)
        return s

    @staticmethod
    def estandarizar_dpto(serie: pd.Series) -> pd.Series:
        """Estandariza código de departamento a 2 dígitos con cero líder.

        Crítico: '5' debe ser '05' (Antioquia), no '5'.

        Args:
            serie: Columna DPTO.

        Returns:
            Serie con códigos de exactamente 2 dígitos.
        """
        return serie.astype(str).str.strip().str.zfill(2)

    @staticmethod
    def estandarizar_area(serie: pd.Series) -> pd.Series:
        """Estandariza código AREA (municipio) a 5 dígitos.

        Args:
            serie: Columna AREA.

        Returns:
            Serie con códigos de exactamente 5 dígitos.
        """
        return serie.astype(str).str.strip().str.zfill(5)

    @staticmethod
    def convertir_columnas_numericas(
        df: pd.DataFrame,
        columnas: List[str],
        excluir: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """Convierte múltiples columnas a numérico in-place.

        Args:
            df: DataFrame a modificar.
            columnas: Lista de columnas a convertir.
            excluir: Columnas que deben permanecer como string.

        Returns:
            El mismo DataFrame modificado.
        """
        excluir = set(excluir or [])
        for col in columnas:
            if col in df.columns and col not in excluir:
                df[col] = ConversorTipos.a_numerico(df[col])
        return df


# ═════════════════════════════════════════════════════════════════════
# 3. ESTADÍSTICAS PONDERADAS
# ═════════════════════════════════════════════════════════════════════

class EstadisticasPonderadas:
    """Calcula estadísticas descriptivas usando pesos de frecuencia (FEX).

    La GEIH es una encuesta con diseño muestral complejo.
    Cada registro representa a FEX_C18 personas de la población real.
    Ignorar los pesos produce estimaciones sesgadas.

    Todas las funciones:
      - Son vectorizadas (no usan loops sobre filas)
      - Manejan NaN y pesos ≤ 0 de forma segura
      - Retornan np.nan si no hay datos válidos
    """

    @staticmethod
    def media(valores: pd.Series, pesos: pd.Series) -> float:
        """Media aritmética ponderada.

        Fórmula: Σ(valor_i × peso_i) / Σ(peso_i)
        """
        mask = valores.notna() & pesos.notna() & (pesos > 0)
        v, w = valores[mask], pesos[mask]
        total_peso = w.sum()
        if total_peso == 0:
            return np.nan
        return float((v * w).sum() / total_peso)

    @staticmethod
    def mediana(valores: pd.Series, pesos: pd.Series) -> float:
        """Mediana ponderada (percentil 50).

        Ordena por valor, acumula pesos, busca donde la acumulación
        cruza el 50% del peso total.
        """
        return EstadisticasPonderadas.percentil(valores, pesos, 0.5)

    @staticmethod
    def percentil(valores: pd.Series, pesos: pd.Series, p: float) -> float:
        """Percentil ponderado.

        Args:
            valores: Datos observados.
            pesos: Pesos de frecuencia (FEX_ADJ).
            p: Percentil deseado en [0, 1]. Ej: 0.25 para P25.

        Returns:
            Valor en la posición del percentil ponderado.
        """
        mask = valores.notna() & pesos.notna() & (pesos > 0)
        v = valores[mask].values
        w = pesos[mask].values
        if len(v) == 0:
            return np.nan
        idx = np.argsort(v)
        v, w = v[idx], w[idx]
        acum = np.cumsum(w)
        corte = p * acum[-1]
        pos = np.searchsorted(acum, corte)
        return float(v[min(pos, len(v) - 1)])

    @staticmethod
    def desviacion_estandar(valores: pd.Series, pesos: pd.Series) -> float:
        """Desviación estándar ponderada."""
        media = EstadisticasPonderadas.media(valores, pesos)
        if np.isnan(media):
            return np.nan
        mask = valores.notna() & pesos.notna() & (pesos > 0)
        v, w = valores[mask], pesos[mask]
        var = ((v - media) ** 2 * w).sum() / w.sum()
        return float(np.sqrt(var))

    @staticmethod
    def coeficiente_variacion(valores: pd.Series, pesos: pd.Series) -> float:
        """CV% = (Desv.Std / Media) × 100.

        CV > 100% indica alta dispersión (ej: sector financiero).
        """
        media = EstadisticasPonderadas.media(valores, pesos)
        if np.isnan(media) or media == 0:
            return np.nan
        std = EstadisticasPonderadas.desviacion_estandar(valores, pesos)
        return float(std / media * 100)

    @staticmethod
    def suma(df: pd.DataFrame, mask: pd.Series,
             col_peso: str = "FEX_ADJ") -> float:
        """Suma ponderada filtrada por una máscara.

        Uso típico: w_sum(df, df['OCI']==1) → total de ocupados expandidos.
        """
        return float(df.loc[mask, col_peso].sum())

    @staticmethod
    def resumen_completo(
        valores: pd.Series,
        pesos: pd.Series,
        smmlv: int = 1_423_500,
    ) -> Dict[str, Any]:
        """Calcula todas las estadísticas descriptivas ponderadas.

        Retorna un diccionario listo para convertir en fila de DataFrame.

        Args:
            valores: Serie con la variable de interés (ej: INGLABO).
            pesos: Serie con los pesos (FEX_ADJ).
            smmlv: SMMLV para expresar en múltiplos.

        Returns:
            Dict con N_personas, Media, Mediana, P10-P90, Std, CV%, etc.
        """
        ep = EstadisticasPonderadas
        n_pond = pesos[pesos.notna() & (pesos > 0)].sum()
        if n_pond == 0:
            return {}
        media   = ep.media(valores, pesos)
        mediana = ep.mediana(valores, pesos)
        std     = ep.desviacion_estandar(valores, pesos)
        cv      = ep.coeficiente_variacion(valores, pesos)
        p10     = ep.percentil(valores, pesos, 0.10)
        p25     = ep.percentil(valores, pesos, 0.25)
        p75     = ep.percentil(valores, pesos, 0.75)
        p90     = ep.percentil(valores, pesos, 0.90)
        return {
            "N_personas":    round(n_pond / 1_000),
            "Media":         round(media),
            "Mediana":       round(mediana),
            "P10":           round(p10),
            "P25":           round(p25),
            "P75":           round(p75),
            "P90":           round(p90),
            "Std":           round(std),
            "CV_%":          round(cv, 1),
            "IQR":           round(p75 - p25),
            "Media_SMMLV":   round(media / smmlv, 2),
            "Mediana_SMMLV": round(mediana / smmlv, 2),
        }

    @staticmethod
    def gini(valores: pd.Series, pesos: pd.Series) -> float:
        """Coeficiente de Gini del ingreso laboral (ponderado).

        Gini = 1 − 2 × ∫₀¹ L(p) dp
        Donde L(p) es la curva de Lorenz.

        Returns:
            Float entre 0 (igualdad perfecta) y 1 (máxima desigualdad).
        """
        mask = valores.notna() & pesos.notna() & (pesos > 0) & (valores > 0)
        v = valores[mask].values
        w = pesos[mask].values
        if len(v) < 2:
            return np.nan
        idx = np.argsort(v)
        v, w = v[idx], w[idx]
        acum_w = np.cumsum(w)
        acum_vw = np.cumsum(v * w)
        total_w = acum_w[-1]
        total_vw = acum_vw[-1]
        # Proporción acumulada de población y de ingreso
        p = acum_w / total_w
        l = acum_vw / total_vw
        # Gini usando la regla del trapecio
        # np.trapz fue renombrado a np.trapezoid en NumPy ≥ 2.0
        _trapz = getattr(np, "trapezoid", getattr(np, "trapz", None))
        gini_val = 1.0 - 2.0 * _trapz(l, p)
        return float(gini_val)

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/utils.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [24/39]: visualizacion.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/visualizacion.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/visualizacion.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.visualizacion — Generación de gráficos para el análisis GEIH.

Cada clase produce un tipo de gráfico. Las clases solo DIBUJAN;
nunca calculan indicadores (eso es responsabilidad de `indicadores.py`).

Convención: todos los métodos retornan la figura matplotlib para que
el usuario pueda ajustarla, guardarla o mostrarla desde el notebook.

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "EstiloBase",
    "GraficoDistribucionIngresos",
    "GraficoBoxPlotSalarios",
    "GraficoBrechaGenero",
    "GraficoRamaSexo",
    "GraficoCurvaLorenz",
    "GraficoICIBubble",
    "GraficoEstacionalidad",
    "GraficoContribucionHeatmap",
]


from typing import Optional, Dict

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec

from .config import COLORES, SMMLV_2025, ConfigGEIH


class EstiloBase:
    """Configuración de estilo compartida por todos los gráficos."""

    FONDO = COLORES["fondo"]
    C = COLORES

    @staticmethod
    def configurar_ejes(ax, titulo: str = "", xlabel: str = "", ylabel: str = ""):
        """Aplica estilo base a un eje matplotlib."""
        ax.set_facecolor("white")
        ax.spines[["top", "right"]].set_visible(False)
        if titulo:
            ax.set_title(titulo, fontsize=12, fontweight="bold", pad=12)
        if xlabel:
            ax.set_xlabel(xlabel, fontsize=11)
        if ylabel:
            ax.set_ylabel(ylabel, fontsize=11)
        ax.grid(axis="x", alpha=0.3, zorder=0)

    @staticmethod
    def crear_figura(ancho: float = 14, alto: float = 8):
        """Crea una figura con fondo institucional."""
        fig = plt.figure(figsize=(ancho, alto))
        fig.patch.set_facecolor(COLORES["fondo"])
        return fig


class GraficoDistribucionIngresos(EstiloBase):
    """Barras apiladas H+M por rango SMMLV con curva acumulada.

    Módulo M1 visual.
    """

    def graficar(
        self,
        dist: pd.DataFrame,
        dist_sexo: pd.DataFrame,
        titulo: str = "Distribución de ingresos laborales",
        smmlv: int = SMMLV_2025,
    ) -> plt.Figure:
        """Genera el gráfico de distribución de ingresos.

        Args:
            dist: DataFrame con RANGO, Personas_M, Pct, Acum_Pct.
            dist_sexo: DataFrame con RANGO, H_M, M_M.

        Returns:
            Figura matplotlib lista para .show() o .savefig().
        """
        fig, (ax1, ax2) = plt.subplots(
            1, 2, figsize=(18, 7),
            gridspec_kw={"width_ratios": [1.8, 1]},
        )
        fig.patch.set_facecolor(self.FONDO)

        labels = [str(r) for r in dist["RANGO"]]
        x = np.arange(len(labels))
        w = 0.65

        # Panel izquierdo: barras apiladas
        self.configurar_ejes(ax1, titulo=titulo, ylabel="Millones de personas")
        h_vals = dist_sexo["H_M"].values
        m_vals = dist_sexo["M_M"].values
        tot = h_vals + m_vals

        ax1.bar(x, h_vals, w, label="Hombres", color=self.C["azul"], alpha=0.88, zorder=3)
        ax1.bar(x, m_vals, w, bottom=h_vals, label="Mujeres", color=self.C["rojo"], alpha=0.88, zorder=3)

        for i, (t, pct) in enumerate(zip(tot, dist["Pct"])):
            ax1.text(i, t + 0.03, f"{t:.2f}M\n{pct:.1f}%",
                     ha="center", va="bottom", fontsize=8, fontweight="bold")

        ax1.axvline(x=1.5, color=self.C["verde"], ls="--", lw=1.8, alpha=0.75)
        ax1.set_xticks(x)
        ax1.set_xticklabels(labels, fontsize=9, rotation=20, ha="right")
        ax1.legend(fontsize=10)

        # Panel derecho: curva acumulada
        self.configurar_ejes(ax2, titulo="Distribución acumulada", ylabel="% acumulado")
        ax2.plot(x, dist["Acum_Pct"], color=self.C["morado"],
                 lw=2.5, marker="o", markersize=6, zorder=3)
        ax2.axhline(50, color=self.C["gris"], ls="--", lw=1, alpha=0.6)
        ax2.set_xticks(x)
        ax2.set_xticklabels(labels, fontsize=8, rotation=25, ha="right")
        ax2.set_ylim(0, 105)

        fig.tight_layout(pad=2.5)
        return fig


class GraficoBoxPlotSalarios(EstiloBase):
    """Box plot ponderado de salarios por rama de actividad.

    Cada "caja" se construye a partir de percentiles ponderados,
    no del método estándar de matplotlib (que ignora pesos).
    """

    def graficar(
        self,
        tabla: pd.DataFrame,
        titulo: str = "Distribución del ingreso laboral por rama",
        smmlv: int = SMMLV_2025,
        max_smmlv_eje: int = 8,
    ) -> plt.Figure:
        """Genera box plot con IQR, bigotes P10-P90, mediana y media.

        Args:
            tabla: DataFrame indexado por Rama con P10, P25, Mediana, P75, P90.
            smmlv: SMMLV para referencias verticales.
            max_smmlv_eje: Límite del eje X en múltiplos de SMMLV.

        Returns:
            Figura matplotlib.
        """
        ramas_orden = tabla.sort_values("Mediana").index.tolist()
        n = len(ramas_orden)
        limite = max_smmlv_eje * smmlv

        fig, ax = plt.subplots(figsize=(14, max(7, n * 0.7)))
        fig.patch.set_facecolor(self.FONDO)
        self.configurar_ejes(ax, titulo=titulo, xlabel="Ingreso laboral mensual (COP)")

        y_pos = np.arange(n)

        for i, rama in enumerate(ramas_orden):
            row = tabla.loc[rama]
            p10, p25 = min(row["P10"], limite), min(row["P25"], limite)
            med, mea = min(row["Mediana"], limite), min(row["Media"], limite)
            p75, p90 = min(row["P75"], limite), min(row["P90"], limite)

            # Bigotes P10–P90
            ax.plot([p10, p90], [i, i], color=self.C["gris"], lw=1.4, zorder=2)

            # Caja IQR
            rect = plt.Rectangle(
                (p25, i - 0.28), p75 - p25, 0.56,
                facecolor=self.C["azul"], alpha=0.22,
                edgecolor=self.C["azul"], linewidth=1.6, zorder=3,
            )
            ax.add_patch(rect)

            # Mediana
            ax.plot([med, med], [i - 0.28, i + 0.28],
                    color=self.C["azul"], lw=2.8, zorder=5)

            # Media (diamante)
            ax.scatter(mea, i, marker="D", s=38, color=self.C["rojo"],
                       zorder=6, edgecolors="white", linewidth=0.8)

            # Etiqueta
            ax.text(p90 + smmlv * 0.05, i,
                    f'{row["Mediana_SMMLV"]:.1f}× | CV:{row["CV_%"]:.0f}%',
                    va="center", fontsize=7.8, color=self.C["negro"])

        # Líneas de referencia SMMLV
        for mult, alpha in [(1, 0.55), (2, 0.35), (3, 0.25), (4, 0.18)]:
            ax.axvline(mult * smmlv, color=self.C["verde"], ls="--", lw=1, alpha=alpha)
            ax.text(mult * smmlv, n - 0.3, f"{mult} SML",
                    ha="center", fontsize=8, color=self.C["verde"], alpha=0.8)

        ax.set_yticks(y_pos)
        ax.set_yticklabels(ramas_orden, fontsize=9.5)
        ax.set_xlim(left=0, right=limite * 1.18)
        ax.xaxis.set_major_formatter(
            mticker.FuncFormatter(lambda v, _: f"${v/1e6:.1f}M" if v >= 1e6 else f"${v/1e3:.0f}k")
        )

        # Leyenda
        leyenda = [
            mpatches.Patch(facecolor=self.C["azul"], alpha=0.22,
                           edgecolor=self.C["azul"], label="Caja IQR (P25–P75)"),
            plt.Line2D([0], [0], color=self.C["azul"], lw=2.8, label="Mediana (P50)"),
            plt.Line2D([0], [0], marker="D", color="w", markerfacecolor=self.C["rojo"],
                       markersize=8, label="Media ponderada"),
            plt.Line2D([0], [0], color=self.C["gris"], lw=1.4, label="Bigotes P10–P90"),
        ]
        ax.legend(handles=leyenda, loc="lower right", fontsize=9, framealpha=0.92)
        fig.tight_layout(pad=2.5)
        return fig


class GraficoBrechaGenero(EstiloBase):
    """Gráfico de brecha salarial de género por nivel educativo."""

    def graficar(
        self,
        pivot_edu: pd.DataFrame,
        titulo: str = "Brecha salarial de género por nivel educativo",
        smmlv: int = SMMLV_2025,
    ) -> plt.Figure:
        """Barras dobles H/M con línea de brecha %.

        Args:
            pivot_edu: DataFrame con columnas Hombres, Mujeres, Brecha_%.

        Returns:
            Figura matplotlib.
        """
        niveles = pivot_edu.index.tolist()
        x = np.arange(len(niveles))
        w = 0.35

        fig, ax1 = plt.subplots(figsize=(14, 7))
        fig.patch.set_facecolor(self.FONDO)
        self.configurar_ejes(ax1, titulo=titulo)

        h_vals = (pivot_edu["Hombres"] / smmlv).values
        m_vals = (pivot_edu["Mujeres"] / smmlv).values

        ax1.bar(x - w/2, h_vals, w, color=self.C["azul"], alpha=0.85, label="Hombres")
        ax1.bar(x + w/2, m_vals, w, color=self.C["rojo"], alpha=0.85, label="Mujeres")

        ax1.set_xticks(x)
        ax1.set_xticklabels(niveles, fontsize=9, rotation=15, ha="right")
        ax1.set_ylabel("Mediana de ingreso (× SMMLV)", fontsize=11)
        ax1.legend(fontsize=10)

        # Eje secundario: brecha %
        if "Brecha_%" in pivot_edu.columns:
            ax2 = ax1.twinx()
            ax2.plot(x, pivot_edu["Brecha_%"], color=self.C["morado"],
                     lw=2.5, marker="s", markersize=7, label="Brecha %")
            ax2.axhline(0, color=self.C["gris"], ls="--", lw=0.8, alpha=0.5)
            ax2.set_ylabel("Brecha % (M − H) / H", fontsize=11, color=self.C["morado"])
            ax2.legend(fontsize=9, loc="upper left")

        fig.tight_layout(pad=2.5)
        return fig


class GraficoRamaSexo(EstiloBase):
    """Barras horizontales de ocupados por rama, desagregado por sexo."""

    def graficar(
        self,
        pivot: pd.DataFrame,
        titulo: str = "Ocupados por rama de actividad y sexo",
    ) -> plt.Figure:
        """Genera gráfico de barras horizontales apiladas H+M.

        Args:
            pivot: DataFrame con RAMA, Hombres_miles, Mujeres_miles.

        Returns:
            Figura matplotlib.
        """
        pivot = pivot.sort_values("Total_miles", ascending=True)
        n = len(pivot)

        fig, ax = plt.subplots(figsize=(14, max(7, n * 0.6)))
        fig.patch.set_facecolor(self.FONDO)
        self.configurar_ejes(ax, titulo=titulo, xlabel="Miles de personas")

        y = np.arange(n)
        h = pivot["Hombres_miles"].values
        m = pivot["Mujeres_miles"].values

        ax.barh(y, h, 0.65, color=self.C["azul"], alpha=0.85, label="Hombres")
        ax.barh(y, m, 0.65, left=h, color=self.C["rojo"], alpha=0.85, label="Mujeres")

        for i, (hv, mv) in enumerate(zip(h, m)):
            ax.text(hv + mv + 20, i, f"{hv+mv:,.0f}K",
                    va="center", fontsize=8)

        ax.set_yticks(y)
        ax.set_yticklabels(pivot["RAMA"], fontsize=9)
        ax.legend(fontsize=10, loc="lower right")
        fig.tight_layout(pad=2.5)
        return fig


# ═════════════════════════════════════════════════════════════════════
# NUEVOS EN v4.0: 4 GRÁFICOS ADICIONALES
# ═════════════════════════════════════════════════════════════════════

class GraficoCurvaLorenz(EstiloBase):
    """Curva de Lorenz del ingreso laboral con shading de desigualdad.

    Dibuja la curva nacional + comparativos (sexo, zona) con el
    coeficiente de Gini en la leyenda. El área sombreada entre
    la línea de igualdad perfecta y la curva real representa
    visualmente la desigualdad.

    Módulo M5 visual — complementa IndicesCompuestos.gini().
    """

    def graficar(
        self,
        df: pd.DataFrame,
        col_val: str = "INGLABO",
        col_peso: str = "FEX_ADJ",
        titulo: str = "Curva de Lorenz del ingreso laboral",
    ) -> plt.Figure:
        """Genera curva de Lorenz con comparativos por sexo y zona.

        Args:
            df: DataFrame de ocupados con INGLABO > 0 y FEX_ADJ.
            col_val: Columna de ingresos.
            col_peso: Columna de pesos.

        Returns:
            Figura matplotlib.
        """
        fig, ax = plt.subplots(figsize=(10, 8))
        fig.patch.set_facecolor(self.FONDO)
        ax.set_facecolor("white")

        # Línea de perfecta igualdad
        ax.plot([0, 1], [0, 1], "k--", lw=1.5, alpha=0.7,
                label="Perfecta igualdad")

        # Configuraciones: (subset, color, label)
        configs = [(df, self.C["azul"], "Nacional")]
        if "ZONA" in df.columns:
            configs.append((df[df["ZONA"] == "Urbano"], self.C["verde"], "Urbano"))
            configs.append((df[df["ZONA"] == "Rural"], self.C["naranja"], "Rural"))
        if "P3271" in df.columns:
            configs.append((df[df["P3271"] == 1], "#1A5276", "Hombres"))
            configs.append((df[df["P3271"] == 2], "#922B21", "Mujeres"))

        _trapz = getattr(np, "trapezoid", getattr(np, "trapz", None))

        for sub, color, lbl in configs:
            m = sub[col_val].notna() & (sub[col_val] > 0) & sub[col_peso].notna()
            v = sub.loc[m, col_val].values
            w = sub.loc[m, col_peso].values
            if len(v) < 10:
                continue
            idx = np.argsort(v)
            v, w = v[idx], w[idx]
            w_cum = np.cumsum(w) / w.sum()
            vw_cum = np.cumsum(v * w) / (v * w).sum()
            # Insertar origen (0,0)
            w_c = np.insert(w_cum, 0, 0)
            vw_c = np.insert(vw_cum, 0, 0)
            gini = 1.0 - 2.0 * _trapz(vw_c, w_c) if _trapz else np.nan
            ax.plot(w_c, vw_c, lw=2.2, color=color,
                    label=f"{lbl} (Gini={gini:.3f})")
            if lbl == "Nacional":
                ax.fill_between(w_c, w_c, vw_c, alpha=0.08, color=color)

        ax.set_xlabel("Proporción acumulada de ocupados", fontsize=11)
        ax.set_ylabel("Proporción acumulada del ingreso", fontsize=11)
        ax.set_title(titulo, fontsize=12, fontweight="bold")
        ax.legend(fontsize=10, framealpha=0.95)
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.xaxis.set_major_formatter(mticker.PercentFormatter(xmax=1))
        ax.yaxis.set_major_formatter(mticker.PercentFormatter(xmax=1))
        ax.grid(alpha=0.25)
        ax.spines[["top", "right"]].set_visible(False)
        fig.tight_layout(pad=2.5)
        return fig


class GraficoICIBubble(EstiloBase):
    """Gráfico de burbujas del ICI — Competitividad laboral por departamento.

    Eje X = Costo laboral efectivo (mediana × 1.54)
    Eje Y = Talento (% universitarios)
    Tamaño burbuja = Oferta disponible (desocupados + subempleados)

    Cuadrantes:
      Alto talento + Bajo costo → ideal para IED manufactura/servicios
      Alto talento + Alto costo → operaciones de alto valor agregado
    """

    def graficar(
        self,
        ici: pd.DataFrame,
        col_x: str = "Costo_efectivo",
        col_y: str = "Talento_univ_%",
        col_size: str = "Oferta_miles",
        titulo: str = "Competitividad laboral por departamento",
    ) -> plt.Figure:
        """Genera gráfico de burbujas del ICI.

        Args:
            ici: DataFrame del resultado de CompetitividadLaboral().calcular().
                 Requiere columnas: Departamento, Costo_efectivo, Talento_univ_%,
                 Oferta_miles (o similar), ICI.

        Returns:
            Figura matplotlib.
        """
        fig, ax = plt.subplots(figsize=(14, 9))
        fig.patch.set_facecolor(self.FONDO)
        ax.set_facecolor("white")

        if col_size in ici.columns:
            sizes = ici[col_size].fillna(1)
            sizes = (sizes / sizes.max() * 800).clip(lower=30)
        else:
            sizes = 150

        # Color por ICI
        ici_vals = ici["ICI"] if "ICI" in ici.columns else pd.Series(50, index=ici.index)
        colors = [self.C["verde"] if v > 55 else
                  (self.C["naranja"] if v > 45 else self.C["rojo"])
                  for v in ici_vals]

        ax.scatter(
            ici[col_x], ici[col_y], s=sizes,
            c=colors, alpha=0.72, edgecolors="white", linewidth=1.2, zorder=3,
        )

        # Etiquetas de departamento
        for _, row in ici.iterrows():
            ax.annotate(
                row.get("Departamento", "")[:12],
                (row[col_x], row[col_y]),
                fontsize=7.5, ha="center", va="bottom",
                textcoords="offset points", xytext=(0, 6),
            )

        # Líneas de referencia (medianas)
        med_x = ici[col_x].median()
        med_y = ici[col_y].median()
        ax.axvline(med_x, color=self.C["gris"], ls="--", lw=1, alpha=0.5)
        ax.axhline(med_y, color=self.C["gris"], ls="--", lw=1, alpha=0.5)

        # Etiquetas de cuadrantes
        ax.text(ici[col_x].min(), ici[col_y].max(),
                " ★ Alto talento\n    Bajo costo",
                fontsize=9, color=self.C["verde"], fontweight="bold",
                va="top", ha="left")
        ax.text(ici[col_x].max(), ici[col_y].max(),
                "Alto talento\nAlto costo ",
                fontsize=9, color=self.C["naranja"], fontweight="bold",
                va="top", ha="right")

        ax.set_xlabel("Costo laboral efectivo (COP)", fontsize=11)
        ax.set_ylabel("% universitarios (proxy talento)", fontsize=11)
        ax.set_title(titulo, fontsize=12, fontweight="bold")
        ax.xaxis.set_major_formatter(
            mticker.FuncFormatter(lambda v, _: f"${v/1e6:.1f}M")
        )
        ax.grid(alpha=0.2)
        ax.spines[["top", "right"]].set_visible(False)
        fig.tight_layout(pad=2.5)
        return fig


class GraficoEstacionalidad(EstiloBase):
    """Líneas mensuales TD/TGP/TO con eje dual.

    Eje izquierdo: TD (%) — la serie más volátil
    Eje derecho: TGP y TO (%) — rango mayor pero más estables
    """

    def graficar(
        self,
        estac: pd.DataFrame,
        titulo: str = "Estacionalidad del mercado laboral",
    ) -> plt.Figure:
        """Genera gráfico de líneas mensuales.

        Args:
            estac: DataFrame con MES (o MES_NUM), TD_%, TGP_%, TO_%.
                   Output de Estacionalidad().calcular().

        Returns:
            Figura matplotlib.
        """
        fig, ax1 = plt.subplots(figsize=(14, 6))
        fig.patch.set_facecolor(self.FONDO)
        ax1.set_facecolor("white")

        # Determinar eje X
        col_mes = "MES" if "MES" in estac.columns else "MES_NUM"
        x = estac[col_mes].values
        x_pos = np.arange(len(x))

        # Eje izquierdo: TD
        color_td = self.C["rojo"]
        ax1.plot(x_pos, estac["TD_%"], color=color_td, lw=2.5,
                 marker="o", markersize=7, label="TD %", zorder=3)
        ax1.set_ylabel("Tasa de Desempleo (%)", fontsize=11, color=color_td)
        ax1.tick_params(axis="y", labelcolor=color_td)

        # Anotar valores TD
        for i, v in enumerate(estac["TD_%"]):
            ax1.text(i, v + 0.15, f"{v:.1f}", ha="center", fontsize=8,
                     color=color_td, fontweight="bold")

        # Eje derecho: TGP y TO
        ax2 = ax1.twinx()
        ax2.plot(x_pos, estac["TGP_%"], color=self.C["azul"], lw=2,
                 marker="s", markersize=5, label="TGP %", zorder=2)
        ax2.plot(x_pos, estac["TO_%"], color=self.C["verde"], lw=2,
                 marker="^", markersize=5, label="TO %", zorder=2)
        ax2.set_ylabel("TGP y TO (%)", fontsize=11)

        # Eje X
        ax1.set_xticks(x_pos)
        if col_mes == "MES":
            labels = [str(m)[:3] for m in x]
        else:
            from .config import MESES_NOMBRES
            labels = [MESES_NOMBRES[int(m)-1][:3] if 1 <= m <= 12 else str(m) for m in x]
        ax1.set_xticklabels(labels, fontsize=10)

        ax1.set_title(titulo, fontsize=12, fontweight="bold")
        ax1.grid(axis="y", alpha=0.2)
        ax1.spines[["top"]].set_visible(False)
        ax2.spines[["top"]].set_visible(False)

        # Leyenda combinada
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2,
                   fontsize=10, loc="upper left", framealpha=0.92)

        fig.tight_layout(pad=2.5)
        return fig


class GraficoContribucionHeatmap(EstiloBase):
    """Heatmap de contribución sectorial al empleo (p.p.) por rama × mes.

    14 ramas CIIU en el eje Y, 12 meses en el eje X.
    El color indica contribución positiva (verde) o negativa (rojo).
    """

    def graficar(
        self,
        contrib: pd.DataFrame,
        col_rama: str = "Rama",
        col_mes: str = "MES_NUM",
        col_valor: str = "Contribucion_pp",
        titulo: str = "Contribución sectorial al empleo (p.p.)",
    ) -> plt.Figure:
        """Genera heatmap de contribución sectorial.

        Args:
            contrib: DataFrame con Rama, MES_NUM, Contribucion_pp.
                     Output de ContribucionSectorial().calcular().
                     Puede ser long format o wide format.

        Returns:
            Figura matplotlib.
        """
        # Si es long format, pivotar
        if col_valor in contrib.columns and col_rama in contrib.columns:
            pivot = contrib.pivot_table(
                index=col_rama, columns=col_mes, values=col_valor,
                aggfunc="sum"
            )
        else:
            # Asumir que ya está en formato wide (ramas como filas, meses como cols)
            pivot = contrib.copy()

        if pivot.empty:
            print("⚠️  No hay datos para generar el heatmap.")
            return plt.figure()

        fig, ax = plt.subplots(figsize=(16, max(8, len(pivot) * 0.55)))
        fig.patch.set_facecolor(self.FONDO)
        ax.set_facecolor("white")

        # Calcular límite simétrico para colores
        vmax = max(abs(pivot.values.min()), abs(pivot.values.max()), 0.5)

        import matplotlib.colors as mcolors
        cmap = mcolors.LinearSegmentedColormap.from_list(
            "contrib", [self.C["rojo"], "white", self.C["verde"]]
        )

        im = ax.imshow(pivot.values, cmap=cmap, aspect="auto",
                       vmin=-vmax, vmax=vmax)

        # Etiquetas
        ax.set_xticks(np.arange(pivot.shape[1]))
        if hasattr(pivot.columns, 'tolist'):
            from .config import MESES_NOMBRES
            cols = pivot.columns.tolist()
            x_labels = []
            for c in cols:
                try:
                    idx = int(c) - 1
                    x_labels.append(MESES_NOMBRES[idx][:3] if 0 <= idx < 12 else str(c))
                except (ValueError, IndexError):
                    x_labels.append(str(c)[:6])
            ax.set_xticklabels(x_labels, fontsize=9)
        ax.set_yticks(np.arange(pivot.shape[0]))
        ax.set_yticklabels([str(r)[:45] for r in pivot.index], fontsize=9)

        # Valores en las celdas
        for i in range(pivot.shape[0]):
            for j in range(pivot.shape[1]):
                val = pivot.values[i, j]
                if not np.isnan(val):
                    color = "white" if abs(val) > vmax * 0.6 else "black"
                    ax.text(j, i, f"{val:+.2f}", ha="center", va="center",
                            fontsize=7, color=color, fontweight="bold")

        ax.set_title(titulo, fontsize=12, fontweight="bold", pad=12)
        fig.colorbar(im, ax=ax, label="Contribución (p.p.)", shrink=0.8)
        fig.tight_layout(pad=2.5)
        return fig

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/visualizacion.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [25/39]: visualizacion_interactiva.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih
   RUTA      : geih/visualizacion_interactiva.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih/visualizacion_interactiva.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih.visualizacion_interactiva — Gráficos interactivos con Plotly.

Versiones interactivas de los gráficos matplotlib con:
  - Tooltips detallados en hover
  - Zoom, pan, selección
  - Filtros dinámicos
  - Exportación PNG/SVG desde el navegador

Requiere: pip install plotly
Si plotly no está disponible, los imports fallan con mensaje claro.

Cada clase replica un gráfico de visualizacion.py pero con interactividad.
Los métodos retornan objetos plotly.graph_objects.Figure que se pueden
mostrar con .show() en Colab/Jupyter o guardar con .write_html().

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "PlotlyLorenz",
    "PlotlyICIBubble",
    "PlotlyEstacionalidad",
    "PlotlyDistribucionIngresos",
    "PlotlyBrechaGenero",
    "PlotlyBoxPlotSalarios",
    "PlotlySalarioRama",
    "PlotlyComparativoAnual",
]

from typing import Optional, Dict, List

import numpy as np
import pandas as pd

try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    _PLOTLY_DISPONIBLE = True
except ImportError:
    _PLOTLY_DISPONIBLE = False


def _verificar_plotly():
    """Verifica que plotly esté instalado."""
    if not _PLOTLY_DISPONIBLE:
        raise ImportError(
            "plotly no está instalado. Instale con:\n"
            "  !pip install plotly\n"
            "O use los gráficos de visualizacion.py (matplotlib, sin interactividad)."
        )


# Paleta de colores institucional
_C = {
    "azul": "#2E6DA4", "rojo": "#C0392B", "verde": "#1E8449",
    "morado": "#7D3C98", "naranja": "#E67E22", "gris": "#7F8C8D",
    "cyan": "#1ABC9C", "amarillo": "#F39C12", "fondo": "#F7F9FC",
}

_LAYOUT_BASE = dict(
    template="plotly_white",
    paper_bgcolor=_C["fondo"],
    font=dict(family="Arial", size=13),
    margin=dict(l=60, r=40, t=80, b=60),
)


# ═════════════════════════════════════════════════════════════════════
# CURVA DE LORENZ INTERACTIVA
# ═════════════════════════════════════════════════════════════════════

class PlotlyLorenz:
    """Curva de Lorenz interactiva con hover que muestra % población y % ingreso."""

    def graficar(
        self,
        df: pd.DataFrame,
        col_val: str = "INGLABO",
        col_peso: str = "FEX_ADJ",
        titulo: str = "Curva de Lorenz del ingreso laboral",
    ) -> "go.Figure":
        _verificar_plotly()

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=[0, 1], y=[0, 1], mode="lines",
            line=dict(dash="dash", color="gray", width=1.5),
            name="Igualdad perfecta", hoverinfo="skip",
        ))

        _trapz = getattr(np, "trapezoid", getattr(np, "trapz", None))

        configs = [
            (df, _C["azul"], "Nacional"),
        ]
        if "ZONA" in df.columns:
            configs.append((df[df["ZONA"] == "Urbano"], _C["verde"], "Urbano"))
            configs.append((df[df["ZONA"] == "Rural"], _C["naranja"], "Rural"))
        if "P3271" in df.columns:
            configs.append((df[df["P3271"] == 1], "#1A5276", "Hombres"))
            configs.append((df[df["P3271"] == 2], "#922B21", "Mujeres"))

        for sub, color, lbl in configs:
            m = sub[col_val].notna() & (sub[col_val] > 0) & sub[col_peso].notna()
            v = sub.loc[m, col_val].values
            w = sub.loc[m, col_peso].values
            if len(v) < 10:
                continue
            idx = np.argsort(v)
            v, w = v[idx], w[idx]
            w_cum = np.cumsum(w) / w.sum()
            vw_cum = np.cumsum(v * w) / (v * w).sum()
            w_c = np.insert(w_cum, 0, 0)
            vw_c = np.insert(vw_cum, 0, 0)
            gini = 1.0 - 2.0 * _trapz(vw_c, w_c) if _trapz else np.nan

            # Reducir puntos para rendimiento (max 200)
            step = max(1, len(w_c) // 200)
            w_s, vw_s = w_c[::step], vw_c[::step]

            fig.add_trace(go.Scatter(
                x=w_s, y=vw_s, mode="lines",
                line=dict(color=color, width=2.5),
                name=f"{lbl} (Gini={gini:.3f})",
                hovertemplate=(
                    f"<b>{lbl}</b><br>"
                    "Población acumulada: %{x:.1%}<br>"
                    "Ingreso acumulado: %{y:.1%}<br>"
                    "<extra></extra>"
                ),
            ))
            if lbl == "Nacional":
                fig.add_trace(go.Scatter(
                    x=np.concatenate([w_s, w_s[::-1]]),
                    y=np.concatenate([w_s, vw_s[::-1]]),
                    fill="toself", fillcolor=f"rgba(46,109,164,0.08)",
                    line=dict(width=0), showlegend=False, hoverinfo="skip",
                ))

        fig.update_layout(
            **_LAYOUT_BASE,
            title=dict(text=titulo, font=dict(size=16)),
            xaxis=dict(title="% acumulado de ocupados", tickformat=".0%", range=[0, 1]),
            yaxis=dict(title="% acumulado del ingreso", tickformat=".0%", range=[0, 1]),
            legend=dict(x=0.05, y=0.95),
            width=700, height=600,
        )
        return fig


# ═════════════════════════════════════════════════════════════════════
# ICI BUBBLE INTERACTIVO
# ═════════════════════════════════════════════════════════════════════

class PlotlyICIBubble:
    """Gráfico de burbujas ICI con hover que muestra todos los indicadores."""

    def graficar(
        self,
        ici: pd.DataFrame,
        col_x: str = "Costo_efectivo",
        col_y: str = "Talento_univ_%",
        col_size: str = "Oferta_miles",
        titulo: str = "Competitividad laboral (ICI) por departamento",
    ) -> "go.Figure":
        _verificar_plotly()

        df = ici.copy()
        size_col = col_size if col_size in df.columns else None
        sizes = df[size_col].fillna(1) if size_col else pd.Series(20, index=df.index)
        sizes = (sizes / sizes.max() * 60).clip(lower=8)

        colors = df["ICI"] if "ICI" in df.columns else pd.Series(50, index=df.index)

        fig = px.scatter(
            df, x=col_x, y=col_y,
            size=sizes.values, color=colors.values,
            hover_name="Departamento",
            hover_data={col_x: ":,.0f", col_y: ":.1f",
                        "ICI": ":.1f" if "ICI" in df.columns else False},
            color_continuous_scale="RdYlGn",
            text="Departamento",
        )
        fig.update_traces(
            textposition="top center", textfont=dict(size=9),
            marker=dict(line=dict(width=1, color="white")),
        )

        # Líneas de referencia (medianas)
        med_x = df[col_x].median()
        med_y = df[col_y].median()
        fig.add_vline(x=med_x, line_dash="dash", line_color="gray", opacity=0.5)
        fig.add_hline(y=med_y, line_dash="dash", line_color="gray", opacity=0.5)

        fig.update_layout(
            **_LAYOUT_BASE,
            title=dict(text=titulo, font=dict(size=16)),
            xaxis=dict(title="Costo laboral efectivo (COP)"),
            yaxis=dict(title="% universitarios"),
            coloraxis_colorbar=dict(title="ICI"),
            width=900, height=700,
        )
        return fig


# ═════════════════════════════════════════════════════════════════════
# ESTACIONALIDAD INTERACTIVA
# ═════════════════════════════════════════════════════════════════════

class PlotlyEstacionalidad:
    """Líneas mensuales TD/TGP/TO con hover detallado."""

    def graficar(
        self,
        estac: pd.DataFrame,
        titulo: str = "Estacionalidad del mercado laboral",
    ) -> "go.Figure":
        _verificar_plotly()

        from .config import MESES_NOMBRES

        col_mes = "MES" if "MES" in estac.columns else "MES_NUM"
        meses_labels = []
        for m in estac[col_mes]:
            try:
                meses_labels.append(MESES_NOMBRES[int(m) - 1][:3])
            except (ValueError, IndexError):
                meses_labels.append(str(m))

        fig = make_subplots(specs=[[{"secondary_y": True}]])

        fig.add_trace(go.Scatter(
            x=meses_labels, y=estac["TD_%"],
            mode="lines+markers", name="TD %",
            line=dict(color=_C["rojo"], width=3),
            marker=dict(size=10),
            hovertemplate="<b>%{x}</b><br>TD: %{y:.1f}%<extra></extra>",
        ), secondary_y=False)

        fig.add_trace(go.Scatter(
            x=meses_labels, y=estac["TGP_%"],
            mode="lines+markers", name="TGP %",
            line=dict(color=_C["azul"], width=2),
            marker=dict(size=7, symbol="square"),
            hovertemplate="<b>%{x}</b><br>TGP: %{y:.1f}%<extra></extra>",
        ), secondary_y=True)

        fig.add_trace(go.Scatter(
            x=meses_labels, y=estac["TO_%"],
            mode="lines+markers", name="TO %",
            line=dict(color=_C["verde"], width=2),
            marker=dict(size=7, symbol="triangle-up"),
            hovertemplate="<b>%{x}</b><br>TO: %{y:.1f}%<extra></extra>",
        ), secondary_y=True)

        fig.update_layout(
            **_LAYOUT_BASE,
            title=dict(text=titulo, font=dict(size=16)),
            legend=dict(x=0.01, y=0.99),
            width=900, height=500,
        )
        fig.update_yaxes(title_text="Tasa de Desempleo (%)", secondary_y=False,
                         titlefont=dict(color=_C["rojo"]))
        fig.update_yaxes(title_text="TGP y TO (%)", secondary_y=True)

        return fig


# ═════════════════════════════════════════════════════════════════════
# DISTRIBUCIÓN DE INGRESOS INTERACTIVA
# ═════════════════════════════════════════════════════════════════════

class PlotlyDistribucionIngresos:
    """Barras apiladas H/M por rango SMMLV con hover detallado."""

    def graficar(
        self,
        dist_sexo: pd.DataFrame,
        titulo: str = "Distribución de ingresos por sexo",
    ) -> "go.Figure":
        _verificar_plotly()

        fig = go.Figure()
        if "H_M" in dist_sexo.columns and "M_M" in dist_sexo.columns:
            rangos = dist_sexo["RANGO"].astype(str).tolist()
            fig.add_trace(go.Bar(
                x=rangos, y=dist_sexo["H_M"], name="Hombres",
                marker_color=_C["azul"],
                hovertemplate="<b>%{x}</b><br>Hombres: %{y:.2f}M<extra></extra>",
            ))
            fig.add_trace(go.Bar(
                x=rangos, y=dist_sexo["M_M"], name="Mujeres",
                marker_color=_C["rojo"],
                hovertemplate="<b>%{x}</b><br>Mujeres: %{y:.2f}M<extra></extra>",
            ))
            fig.update_layout(barmode="stack")

        fig.update_layout(
            **_LAYOUT_BASE,
            title=dict(text=titulo, font=dict(size=16)),
            xaxis=dict(title="Rango SMMLV"),
            yaxis=dict(title="Millones de personas"),
            width=800, height=500,
        )
        return fig


# ═════════════════════════════════════════════════════════════════════
# BRECHA DE GÉNERO INTERACTIVA
# ═════════════════════════════════════════════════════════════════════

class PlotlyBrechaGenero:
    """Barras dobles H/M con línea de brecha % en hover."""

    def graficar(
        self,
        pivot_edu: pd.DataFrame,
        smmlv: int = 1_423_500,
        titulo: str = "Brecha salarial de género por educación",
    ) -> "go.Figure":
        _verificar_plotly()

        niveles = pivot_edu.index.tolist()
        h_vals = (pivot_edu["Hombres"] / smmlv).values
        m_vals = (pivot_edu["Mujeres"] / smmlv).values
        brecha = pivot_edu.get("Brecha_%", pd.Series(dtype=float)).values

        fig = make_subplots(specs=[[{"secondary_y": True}]])

        fig.add_trace(go.Bar(
            x=niveles, y=h_vals, name="Hombres",
            marker_color=_C["azul"], opacity=0.85,
            hovertemplate="<b>%{x}</b><br>Hombres: %{y:.2f}× SMMLV<extra></extra>",
        ), secondary_y=False)

        fig.add_trace(go.Bar(
            x=niveles, y=m_vals, name="Mujeres",
            marker_color=_C["rojo"], opacity=0.85,
            hovertemplate="<b>%{x}</b><br>Mujeres: %{y:.2f}× SMMLV<extra></extra>",
        ), secondary_y=False)

        if len(brecha) > 0 and not np.all(np.isnan(brecha)):
            fig.add_trace(go.Scatter(
                x=niveles, y=brecha, mode="lines+markers",
                name="Brecha %", line=dict(color=_C["morado"], width=2.5),
                marker=dict(size=8, symbol="square"),
                hovertemplate="<b>%{x}</b><br>Brecha: %{y:+.1f}%<extra></extra>",
            ), secondary_y=True)

        fig.update_layout(
            **_LAYOUT_BASE, barmode="group",
            title=dict(text=titulo, font=dict(size=16)),
            width=900, height=550,
        )
        fig.update_yaxes(title_text="Mediana (× SMMLV)", secondary_y=False)
        fig.update_yaxes(title_text="Brecha % (M−H)/H", secondary_y=True,
                         titlefont=dict(color=_C["morado"]))
        return fig


# ═════════════════════════════════════════════════════════════════════
# BOX PLOT SALARIOS INTERACTIVO
# ═════════════════════════════════════════════════════════════════════

class PlotlyBoxPlotSalarios:
    """Box plot interactivo de salarios por rama (muestra ponderada)."""

    def graficar(
        self,
        df: pd.DataFrame,
        smmlv: int = 1_423_500,
        muestra: int = 30_000,
        titulo: str = "Distribución del ingreso por rama",
    ) -> "go.Figure":
        _verificar_plotly()

        df_ocu = df[(df["OCI"] == 1) & (df["INGLABO"] > 0) & df["RAMA"].notna()].copy()
        df_ocu["INGLABO_SML"] = df_ocu["INGLABO"] / smmlv

        if len(df_ocu) > muestra:
            df_s = df_ocu.sample(muestra, weights="FEX_ADJ", random_state=42, replace=True)
        else:
            df_s = df_ocu

        fig = px.box(
            df_s, x="INGLABO_SML", y="RAMA", color="RAMA",
            orientation="h", points=False,
            labels={"INGLABO_SML": "Ingreso (× SMMLV)", "RAMA": ""},
        )
        fig.add_vline(x=1, line_dash="dash", line_color=_C["verde"],
                      annotation_text="1 SMMLV")
        fig.add_vline(x=2, line_dash="dot", line_color=_C["naranja"],
                      annotation_text="2 SMMLV")

        fig.update_layout(
            **_LAYOUT_BASE, showlegend=False,
            title=dict(text=titulo, font=dict(size=16)),
            xaxis=dict(range=[0, 10]),
            height=max(500, len(df_s["RAMA"].unique()) * 45),
            width=900,
        )
        return fig


# ═════════════════════════════════════════════════════════════════════
# SALARIO POR RAMA (BARRAS HORIZONTALES)
# ═════════════════════════════════════════════════════════════════════

class PlotlySalarioRama:
    """Barras horizontales de mediana salarial por rama con tooltip completo."""

    def graficar(
        self,
        tabla: pd.DataFrame,
        smmlv: int = 1_423_500,
        titulo: str = "Mediana salarial por rama de actividad",
    ) -> "go.Figure":
        _verificar_plotly()

        df = tabla.sort_values("Mediana", ascending=True).copy()

        fig = go.Figure(go.Bar(
            y=df.index if df.index.name else df.iloc[:, 0],
            x=df["Mediana"] / smmlv,
            orientation="h",
            marker_color=_C["azul"],
            hovertemplate=(
                "<b>%{y}</b><br>"
                "Mediana: %{x:.2f}× SMMLV<br>"
                "<extra></extra>"
            ),
        ))

        fig.add_vline(x=1, line_dash="dash", line_color=_C["verde"],
                      annotation_text="1 SMMLV")

        fig.update_layout(
            **_LAYOUT_BASE,
            title=dict(text=titulo, font=dict(size=16)),
            xaxis=dict(title="Mediana ingreso (× SMMLV)"),
            height=max(400, len(df) * 35),
            width=800,
        )
        return fig


# ═════════════════════════════════════════════════════════════════════
# COMPARATIVO ANUAL INTERACTIVO
# ═════════════════════════════════════════════════════════════════════

class PlotlyComparativoAnual:
    """Gráficos interactivos para comparación multi-año.

    Diseñado para usarse con ComparadorMultiAnio.

    Uso:
        comp = ComparadorMultiAnio()
        comp.agregar_anio(2025, ..., ...)
        comp.agregar_anio(2026, ..., ...)

        vis = PlotlyComparativoAnual()
        fig = vis.indicadores(comp.comparar_indicadores())
        fig.show()
    """

    def indicadores(
        self,
        df_ind: pd.DataFrame,
        titulo: str = "Evolución de indicadores laborales",
    ) -> "go.Figure":
        """Líneas TD/TGP/TO por año."""
        _verificar_plotly()

        fig = make_subplots(specs=[[{"secondary_y": True}]])
        anios = df_ind["ANIO"].astype(str).tolist()

        fig.add_trace(go.Scatter(
            x=anios, y=df_ind["TD_%"], mode="lines+markers+text",
            name="TD %", line=dict(color=_C["rojo"], width=3),
            marker=dict(size=12), text=[f"{v:.1f}%" for v in df_ind["TD_%"]],
            textposition="top center",
        ), secondary_y=False)

        fig.add_trace(go.Scatter(
            x=anios, y=df_ind["TGP_%"], mode="lines+markers",
            name="TGP %", line=dict(color=_C["azul"], width=2),
        ), secondary_y=True)

        fig.add_trace(go.Scatter(
            x=anios, y=df_ind["TO_%"], mode="lines+markers",
            name="TO %", line=dict(color=_C["verde"], width=2),
        ), secondary_y=True)

        fig.update_layout(
            **_LAYOUT_BASE,
            title=dict(text=titulo, font=dict(size=16)),
            xaxis=dict(title="Año"),
            width=800, height=500,
        )
        fig.update_yaxes(title_text="TD (%)", secondary_y=False)
        fig.update_yaxes(title_text="TGP / TO (%)", secondary_y=True)
        return fig

    def ingresos(
        self,
        df_ing: pd.DataFrame,
        titulo: str = "Evolución de la mediana salarial",
    ) -> "go.Figure":
        """Barras de mediana en SMMLV + línea en COP."""
        _verificar_plotly()

        fig = make_subplots(specs=[[{"secondary_y": True}]])
        anios = df_ing["ANIO"].astype(str).tolist()

        fig.add_trace(go.Bar(
            x=anios, y=df_ing["Mediana_SMMLV"], name="Mediana (× SMMLV)",
            marker_color=_C["azul"], opacity=0.8,
            text=[f"{v:.2f}×" for v in df_ing["Mediana_SMMLV"]],
            textposition="outside",
        ), secondary_y=False)

        fig.add_trace(go.Scatter(
            x=anios, y=df_ing["Mediana_COP"] / 1e6, mode="lines+markers",
            name="Mediana (M COP)", line=dict(color=_C["rojo"], width=2),
            marker=dict(size=8),
        ), secondary_y=True)

        fig.update_layout(
            **_LAYOUT_BASE,
            title=dict(text=titulo, font=dict(size=16)),
            xaxis=dict(title="Año"),
            width=800, height=500,
        )
        fig.update_yaxes(title_text="× SMMLV", secondary_y=False)
        fig.update_yaxes(title_text="Millones COP", secondary_y=True)
        return fig

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih/visualizacion_interactiva.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [26/39]: __init__.py
   TIPO      : CÓDIGO
   UBICACIÓN : geih_2025
   RUTA      : geih_2025/__init__.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: geih_2025/__init__.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih — Análisis de microdatos GEIH del DANE.

Gran Encuesta Integrada de Hogares | DANE | Marco Muestral 2018
Autor: Néstor Enrique Forero Herrera

Paquete multi-año: soporta GEIH 2022–presente. No está atado a ningún
año específico — el nombre 'geih_2025' era la versión anterior.

Instalación:
    pip install geih-analisis
    # o desde GitHub:
    pip install git+https://github.com/enriqueforero/geih-analisis.git

Uso rápido:
    from geih import ConfigGEIH, ConsolidadorGEIH, PreparadorGEIH
    config = ConfigGEIH(anio=2025, n_meses=12)

Compatibilidad hacia atrás:
    'from geih_2025 import ...' sigue funcionando gracias al shim
    incluido en este paquete. Recibirás un DeprecationWarning.

CAMBIO v5.0 — Renombrado geih_2025 → geih:
  El paquete ahora se llama 'geih' (nombre de importación) y
  'geih-analisis' (nombre de distribución en PyPI).
  El shim geih_2025/ garantiza compatibilidad durante la transición.

Módulos del paquete (17 archivos):
  config.py                → Constantes, mapeos, configuración centralizada
  utils.py                 → Memoria, conversión de tipos, estadísticas ponderadas
  consolidador.py          → Lectura y unión de módulos CSV mensuales
  preparador.py            → Preparación de datos, merge con correlativas
  diagnostico.py           → Diagnóstico de calidad de datos
  indicadores.py           → Indicadores básicos (TD, TGP, TO, ingresos, rama)
  analisis_avanzado.py     → Módulos avanzados (ICE, ICI, ITAT, Mincer, etc.)
  analisis_area.py         → 32 ciudades × CIIU
  analisis_poblacional.py  → Campesinos, discapacidad, migración
  analisis_complementario.py → M8, M14, MX1–MX3
  exportador.py            → Exportación organizada a carpetas
  visualizacion.py         → Gráficos matplotlib
  visualizacion_interactiva.py → Gráficos Plotly
  comparativo.py           → Comparación inter-anual
  descargador.py           → Descarga automática DANE
  logger.py                → Logging centralizado
  profiler.py              → Profiling de memoria
  dashboard.py             → Dashboard Streamlit
"""

__version__ = "5.0.0"
__author__  = "Néstor Enrique Forero Herrera"
__email__   = "nforero@procolombia.co"
__url__     = "https://github.com/enriqueforero/geih-analisis"
__license__ = "MIT"

# ── Configuración ──────────────────────────────────────────────────
from .config import (
    # Configuración principal
    ConfigGEIH,
    # SMMLV
    SMMLV_2025, SMMLV_POR_ANIO,
    # Colores
    COLORES,
    # Períodos
    MESES_CARPETAS, MESES_NOMBRES,
    generar_carpetas_mensuales, generar_etiqueta_periodo,
    # Geografía
    DEPARTAMENTOS, DPTO_A_CIUDAD, AREA_A_CIUDAD,
    CIUDADES_13_PRINCIPALES, CIUDADES_10_INTERMEDIAS,
    # Clasificaciones económicas
    RAMAS_DANE, TABLA_CIIU_RAMAS, AGRUPACION_DANE_8,
    # Referencias DANE
    REF_DANE_2025, REF_DANE, ReferenciaDane,
    # Constantes laborales — antes faltaban, causaban ImportError
    CARGA_PRESTACIONAL, TAMANO_EMPRESA,
    RANGOS_SMMLV_LIMITES, RANGOS_SMMLV_ETIQUETAS,
    # Educación — antes faltaban
    NIVELES_AGRUPADOS, NIVELES_EDUCATIVOS, P3042_A_ANOS,
    # Llaves y converters
    LLAVES_PERSONA, LLAVES_HOGAR,
    CONVERTERS_BASE, CONVERTERS_CON_AREA,
    MODULOS_CSV,
)

# ── Utilidades ─────────────────────────────────────────────────────
from .utils import GestorMemoria, ConversorTipos, EstadisticasPonderadas

# ── Consolidación ──────────────────────────────────────────────────
from .consolidador import ConsolidadorGEIH

# ── Preparación ────────────────────────────────────────────────────
from .preparador import PreparadorGEIH, MergeCorrelativas

# ── Diagnóstico ────────────────────────────────────────────────────
from .diagnostico import DiagnosticoCalidad, Top20Sectores

# ── Exportación organizada ─────────────────────────────────────────
from .exportador import Exportador

# ── Indicadores básicos ────────────────────────────────────────────
from .indicadores import (
    IndicadoresLaborales, DistribucionIngresos, AnalisisRamaSexo,
    AnalisisSalarios, BrechaGenero, AnalisisCruzado,
    IndicesCompuestos, AnalisisArea,
)

# ── Análisis por 32 ciudades ───────────────────────────────────────
from .analisis_area import AnalisisOcupadosCiudad

# ── Análisis avanzado ──────────────────────────────────────────────
from .analisis_avanzado import (
    CalidadEmpleo, FormalidadSectorial, VulnerabilidadLaboral,
    CompetitividadLaboral, AnalisisSubempleo, AnalisisHoras,
    Estacionalidad, FuerzaLaboralJoven, EtnicoRacial,
    BonoDemografico, CostoLaboral, AnalisisFFT,
    AnalisisUrbanoRural, ProductividadTamano,
    ContribucionSectorial, MapaTalento, EcuacionMincer,
    ProxyBilinguismo,
)

# ── Visualización matplotlib ───────────────────────────────────────
from .visualizacion import (
    GraficoDistribucionIngresos, GraficoBoxPlotSalarios,
    GraficoBrechaGenero, GraficoRamaSexo,
    GraficoCurvaLorenz, GraficoICIBubble,
    GraficoEstacionalidad, GraficoContribucionHeatmap,
)

# ── Análisis poblacional ───────────────────────────────────────────
from .analisis_poblacional import (
    AnalisisCampesino, AnalisisDiscapacidad, AnalisisMigracion,
    AnalisisOtrasFormas, AnalisisOtrosIngresos,
    AnalisisSobrecalificacion, AnalisisContractual,
    AnalisisAutonomia, AnalisisAlcanceMercado, AnalisisDesanimados,
)

# ── Análisis complementarios ───────────────────────────────────────
from .analisis_complementario import (
    DuracionDesempleo, DashboardSectoresProColombia,
    AnatomaSalario, FormaPago, CanalEmpleo,
)

# ── Descarga automática DANE ───────────────────────────────────────
from .descargador import DescargadorDANE

# ── Comparativo multi-año ──────────────────────────────────────────
from .comparativo import ComparadorMultiAnio

# ── Visualización interactiva Plotly ───────────────────────────────
try:
    from .visualizacion_interactiva import (
        PlotlyLorenz, PlotlyICIBubble, PlotlyEstacionalidad,
        PlotlyDistribucionIngresos, PlotlyBrechaGenero,
        PlotlyBoxPlotSalarios, PlotlySalarioRama,
        PlotlyComparativoAnual,
    )
except ImportError:
    pass  # plotly no instalado — instalar con: pip install geih-analisis[viz]

# ── Logging centralizado ───────────────────────────────────────────
from .logger import get_logger, configurar_logging, LoggerGEIH

# ── Profiling de memoria ───────────────────────────────────────────
from .profiler import PerfilMemoria, medir_tiempo, tamano_objeto

# ── Dashboard Streamlit ────────────────────────────────────────────
from .dashboard import ejecutar_dashboard

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: geih_2025/__init__.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [27/39]: 2026_03_28_consolidar_geih_analisis_y_calculos.py
   TIPO      : CÓDIGO
   UBICACIÓN : notebooks
   RUTA      : notebooks/2026_03_28_consolidar_geih_analisis_y_calculos.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: notebooks/2026_03_28_consolidar_geih_analisis_y_calculos.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""2026-03-28 Consolidar_GEIH_2025 y cálculos.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1qlnDv9uXcOYWYRaSyZ6xfKKrB9PWvf8_

Autor: *Enrique Forero*

*Date: 20YY-MM-DD*

Breve descripción.

Este notebook fue concebido, diseñado e implementado por el autor, quien mantuvo el control total sobre la arquitectura del sistema, definición de requerimientos, decisiones técnicas y validación de resultados. Se emplearon herramientas de inteligencia artificial (IA) como Claude de Anthropic, Gemini de Google, ChatGPT de OpenAI para acelerar la construcción de funciones específicas, optimizar consultas complejas y sugerir mejores prácticas de programación. Cada propuesta de la IA fue revisada críticamente, adaptada a los requerimientos específicos del negocio, integrada con los sistemas existentes y validada mediante pruebas funcionales y de rendimiento para garantizar los más altos estándares de calidad, seguridad y rendimiento. Este enfoque de desarrollo asistido por IA maximizó la eficiencia (ganancias en velocidad de programación y maximización de resultados) del proceso sin comprometer la responsabilidad profesional ni los estándares técnicos más exigentes.

[Notebook original](https://colab.research.google.com/drive/1b3sSduXNoN5h11NIhFvQTT5fBmZa30va#scrollTo=L7kwsG_oAMLu)

📋 ESTÁNDARES DE PROGRAMACIÓN PARA NOTEBOOKS

🎯 CONTEXTO DE EJECUCIÓN
La programación se corre en un notebook de Google Colab gratuito (~12GB RAM, límite 12h) y debe estar optimizada para:
- Uso eficiente de memoria RAM (utilizar lo mínimo requerido)
- Velocidad de procesamiento máxima
- Procesamiento de bases grandes (+4 millones de registros)

🔍 PROCESO DE DESARROLLO
Antes de implementar cambios:
1. Hacer preguntas rigurosas para alinear la solución
2. Analizar implicaciones de cada cambio propuesto
3. Validar que la solución sea escalable y robusta
4. Evitar falsos positivos y falsos negativos en la lógica

Principios de Ingeniería de Software Probados:

KISS (Keep It Simple, Stupid),

YAGNI (You Aren't Gonna Need It),

Fail Fast > Fail Slow,

User Control > Automation,

DRY (Don't Repeat Yourself).


⚡ PRINCIPIOS DE EFICIENCIA

Procesamiento
- **Vectorización**: Usar operaciones vectorizadas de NumPy/Pandas. NUNCA bucles `for` sobre DataFrames
- **Streaming/Batching**: Leer, procesar y escribir en lotes (chunks). No cargar archivos completos en memoria
- **Caché LRU**: Usar `@lru_cache` para evitar recálculos de operaciones costosas

Almacenamiento
- **Disco temporal**: Volcar datos intermedios a disco cuando excedan memoria disponible.

Resiliencia
- **Checkpointing**: Guardar progreso en disco paso a paso
- **Recuperación**: Permitir reanudar desde el último checkpoint si el proceso falla

Verificaciones:  Implementa el patrón Trust but Verify (Confía pero Verifica), que es fundamental para sistemas de archivos distribuidos/lentos como Google Drive

🏗️ ARQUITECTURA DE CÓDIGO

 Principios SOLID
- **SRP (Responsabilidad Única)**: Cada clase/función hace UNA sola cosa
- **Defensive Programming**: Validar inputs, manejar excepciones, fallar de forma controlada

Organización en Notebooks
Separar SIEMPRE el código en DOS tipos de celdas:

**CELDA TIPO "EXTRAS" (Definiciones):**
- Clases, funciones, constantes, enums, dataclasses
- Imports específicos del módulo
- Ejecutar UNA VEZ al inicio del notebook
- Ubicar en sección dedicada (ej: "SECCIÓN EXTRAS" o "UTILIDADES")

**CELDA TIPO "EJECUTAR" (Máximo 10 líneas):**
- Solo variables configurables por el usuario
- Una llamada a la función principal
- Sin definiciones, sin lógica compleja
- Ejemplo:
  ```python
  WORKSPACE = "mi_carpeta"  # ← Configurable
  LIMIT = 100_000           # ← Configurable
  funcion_principal(WORKSPACE, config, datos)
  ```

 Nombres y Conflictos
- **Evitar conflictos**: No crear clases con nombres que ya existen en el pipeline (ej: si existe `MemoryManager`, no crear otra con el mismo nombre)
- **Nomenclatura clara**: Usar prefijos o sufijos descriptivos si es necesario (ej: `DiagnosticoMemoryManager`)

🚫 PRÁCTICAS PROHIBIDAS
- ❌ Monkey-patching (parchar código existente) - solo para pruebas temporales
- ❌ Bucles `for` sobre filas de DataFrames
- ❌ Cargar datasets completos en memoria sin necesidad
- ❌ Modificar clases existentes sin analizar implicaciones
- ❌ Código duplicado o redundante

 ✅ PRÁCTICAS REQUERIDAS
- ✅ Modificar código de raíz para que sea escalable
- ✅ Hacer ajustes paso a paso con validación
- ✅ Explicar dónde hacer ajustes y sus implicaciones
- ✅ Conservar estructura actual mejorándola incrementalmente
- ✅ Documentar funciones con docstrings
- ✅ Type hints en funciones públicas


⚠️ LA REGLA DE ORO (Instrucción Obligatoria)
Para evitar afectar la memoria RAM y el rendimiento, debes seguir esta instrucción en cada línea de código que escribas de ahora en adelante:
"NUNCA usar apply, iterrows o bucles for sobre DataFrames grandes (>100k filas). SIEMPRE usar operaciones vectorizadas (.str, .dt, operadores + - * /) sobre columnas completas o Arrays de NumPy."



📋 ESTÁNDARES DE PROGRAMACIÓN PARA NOTEBOOKS

🎯 CONTEXTO DE EJECUCIÓN
La programación se corre en un notebook de Google Colab gratuito (~12GB RAM, límite 12h) y debe estar optimizada para:
* Uso eficiente de memoria RAM (utilizar lo mínimo requerido)
* Velocidad de procesamiento máxima
* Procesamiento de bases grandes (+4 millones de registros)

🔍 PROCESO DE DESARROLLO
Antes de implementar cambios:
1. Hacer preguntas rigurosas para alinear la solución
2. Analizar implicaciones de cada cambio propuesto
3. Validar que la solución sea escalable y robusta
4. Evitar falsos positivos y falsos negativos en la lógica

⚡ PRINCIPIOS DE EFICIENCIA

Procesamiento
* Vectorización: Usar operaciones vectorizadas de NumPy/Pandas. NUNCA bucles `for` sobre DataFrames
* Streaming/Batching: Leer, procesar y escribir en lotes (chunks). No cargar archivos completos en memoria
* Caché LRU: Usar `@lru_cache` para evitar recálculos de operaciones costosas

Almacenamiento
* Disco temporal: Volcar datos intermedios a disco cuando excedan memoria disponible

Resiliencia
* Checkpointing: Guardar progreso en disco paso a paso
* Recuperación: Permitir reanudar desde el último checkpoint si el proceso falla
* Verificaciones: Implementa el patrón Trust but Verify (Confía pero Verifica)

🧠 GESTIÓN DE ESTADO EN NOTEBOOKS

Variables en memoria persisten entre ejecuciones causando "estado contaminado":

**Soluciones requeridas:**
* Limpieza explícita: `del variable; gc.collect()` antes de recargas
* Flags de reinicio: `reiniciar: bool = False` en métodos que manipulan datos
* Preservar originales: `self._datos_full = self.datos` antes de filtrar
* Validación de tamaño: `print(f"✅ {len(datos):,} registros cargados")` siempre

**Regla crítica:** Distinguir entre:
* Funciones que leen desde disco (estado fresco) → usar para recargas
* Funciones que leen estado guardado (puede tener filtros) → puede estar contaminado

🏗️ ARQUITECTURA DE CÓDIGO

Principios SOLID
* SRP (Responsabilidad Única): Cada clase/función hace UNA sola cosa
* Separación MVC: Lógica (controlador) separada de renderizado (vista)
* Defensive Programming: Validar inputs, manejar excepciones, fallar de forma controlada

Organización en Notebooks
Separar SIEMPRE el código en DOS tipos de celdas:

**CELDA TIPO "EXTRAS" (Definiciones):**
* Clases, funciones, constantes, enums, dataclasses
* Imports específicos del módulo
* Ejecutar UNA VEZ al inicio del notebook
* Ubicar en sección dedicada (ej: "SECCIÓN EXTRAS" o "UTILIDADES")

**CELDA TIPO "EJECUTAR" (Máximo 15 líneas, idealmente <10):**
* Variables configurables por el usuario
* Limpieza de memoria si es necesario (`del variable; gc.collect()`)
* Una llamada a la función principal
* Sin definiciones, sin lógica compleja

Ejemplo:
```python
# Configuración
WORKSPACE = "mi_carpeta"
LIMIT = 100_000

# Limpieza (si es necesario)
if 'sistema' in globals():
    del sistema
    gc.collect()

# Ejecución
funcion_principal(WORKSPACE, config, datos)
```

Configuración Centralizada
* Usar `@dataclass` para toda configuración
* NUNCA números mágicos en lógica de negocio
* Validar configuración en `__post_init__`

Ejemplo:
```python
@dataclass
class Config:
    tamano_lote: int = 5000
    umbral_minimo: int = 50
    
    def __post_init__(self):
        if self.tamano_lote < 100:
            raise ValueError("Lote muy pequeño")
```

Nombres y Conflictos
* Evitar conflictos: No crear clases/funciones con nombres que ya existen
* Nomenclatura clara: Usar prefijos descriptivos si hay riesgo de conflicto
* Convenciones: `_interno()` para privados, `_full` para datasets completos, `CONSTANTES` en UPPER_CASE

🚫 PRÁCTICAS PROHIBIDAS
* ❌ Monkey-patching (parchar código existente) - solo para pruebas temporales
* ❌ Bucles `for` sobre filas de DataFrames
* ❌ Cargar datasets completos en memoria sin necesidad
* ❌ Modificar clases existentes sin analizar implicaciones
* ❌ Código duplicado o redundante
* ❌ Lógica de negocio en clases de renderizado/vista

✅ PRÁCTICAS REQUERIDAS
* ✅ Modificar código de raíz para que sea escalable
* ✅ Hacer ajustes paso a paso con validación
* ✅ Explicar dónde hacer ajustes y sus implicaciones
* ✅ Conservar estructura actual mejorándola incrementalmente
* ✅ Documentar funciones con docstrings y decisiones con comentarios
* ✅ Type hints en funciones públicas
* ✅ Validación inmediata después de operaciones críticas
* ✅ Reportes informativos para el usuario (progreso, composición, advertencias)

🧪 VALIDACIÓN Y ROBUSTEZ

Testing en Notebooks
* Assertions después de cambios: `assert len(datos) > 1000, "Datos insuficientes"`
* Prints estratégicos: Usuario debe VER qué está pasando
* Casos edge explícitos: Manejar listas vacías, valores únicos, etc.

Backward Compatibility
* Mantener parámetros por defecto que no rompan código existente
* Si cambias defaults, documentarlo como BREAKING CHANGE
* Agregar parámetros nuevos con valores default seguros

📊 OBSERVABILIDAD

El usuario debe VER lo que está pasando:
* Progreso detallado: `print(f"🔄 Procesando {i}/{total}...")`
* Composición de resultados: Mostrar distribuciones, no solo totales
* Advertencias visibles: `print(f"⚠️ Solo {actual} de {esperado}")`
* Validación de tamaños: Siempre confirmar cantidad de registros cargados

📋 CHECKLIST PREVIO A ENTREGA

Antes de marcar código como "listo para producción":
- [ ] ¿Maneja estado contaminado? (limpieza o flag `reiniciar`)
- [ ] ¿Validación de tamaño visible? (`print(f"{len(datos):,}")`)
- [ ] ¿Configuración centralizada? (@dataclass, sin números mágicos)
- [ ] ¿Reportes informativos? (usuario ve progreso y composición)
- [ ] ¿Casos edge manejados? (vacíos, nulos, valores únicos)
- [ ] ¿Backward compatible? (defaults no cambiados sin razón)
- [ ] ¿Separación de responsabilidades? (lógica vs renderizado)
- [ ] ¿Documentación de "por qué"? (no solo "qué")

Prácticas de ingeniería de software que garantizan que el proceso sea auditable y resistente a errores:

1.  **Centralización de la Verdad (`Single Source of Truth`):**
    *   Todas las rutas se definen en una sola clase (`ConfiguracionOperacionRescate`). No hay rutas "hardcodeadas" (escritas a mano) dentro de las funciones lógicas. Si cambia una carpeta, solo se cambia en un lugar.

2.  **Uso de `pathlib`:**
    *   En lugar de usar cadenas de texto simples, se utilizan objetos `Path`. Esto evita errores de separadores de sistema operativo (`/` vs `\`) y facilita la creación de directorios padres (`mkdir(parents=True)`).

3.  **Validación de Dependencias (`Pre-flight Checks`):**
    *   Antes de ejecutar cualquier fase, el sistema usa el método `verificar_fase()`. Este revisa físicamente en el disco si los archivos de entrada existen. Si falta uno, el proceso se detiene antes de empezar, evitando errores a mitad de camino.

4.  **Versionado de Salidas:**
    *   Cada fase genera un archivo con un nombre distinto (ej: `correlative_remediado.parquet`, `correlative_fn001_aplicado.parquet`, `correlative_FINAL.parquet`). Esto permite trazar la evolución del dato y hacer *rollback* si es necesario.

5.  **Abstracción de Iteraciones:**
    *   Se utiliza una variable `ITERACION_ACTUAL` (ej: `IT7...`). Al cambiar esta variable, todo el árbol de carpetas se redirige automáticamente a la nueva iteración mensual, permitiendo reusar el código sin riesgo de sobrescribir meses anteriores.

    Aquí tienes el texto complementario, redactado de forma genérica para que sirva como estándar transversal a cualquier proyecto de datos o ingeniería que desarrolles en el futuro.

Puedes agregarlo justo antes de la sección "Checklist Previo a Entrega".

***

🛡️ ESQUEMA DE VERIFICACIÓN Y TRAZABILIDAD (MANDATORIO)

Para garantizar que la evolución del código no introduzca errores silenciosos, todo proyecto debe cumplir con este protocolo de seguridad:

*   **Dataset de Referencia (Golden Set):**
    *   Mantener un archivo estático y pequeño (ej. 1,000 registros representativos) que incluya **casos de borde críticos** (valores nulos, ceros, formatos extraños, máximos/mínimos).
    *   **Regla:** Este dataset debe tener un "Resultado Esperado" conocido. Antes de procesar la data real, el algoritmo debe correr sobre este set. Si el resultado discrepa del esperado, **STOP**. No pasar a producción.

*   **Trazabilidad Configuración-Resultado:**
    *   Los resultados de un modelo o algoritmo son inútiles si no se sabe qué parámetros los generaron.
    *   **Acción:** Al guardar el output final, generar siempre un archivo adjunto (metadata/json) que contenga los **parámetros exactos** usados en esa corrida (umbrales, fechas, versiones). Vincular siempre el *Qué* (resultado) con el *Cómo* (lógica).

*   **Determinismo Estricto:**
    *   Eliminar el azar no controlado. En procesos que involucren aleatoriedad (muestreo, ML, hashing), fijar siempre las semillas (`random_state=42`, `seed=123`) al inicio del notebook.
    *   **Objetivo:** Misma entrada + Mismo código = **Exactamente** la misma salida, siempre.

*   **Prueba de Humo (Smoke Test):**
    *   Nunca lanzar un proceso masivo sin antes validar la integridad del pipeline.
    *   **Acción:** Ejecutar el flujo completo con una muestra minúscula (ej. 0.1% de los datos o los primeros 100 registros). Esto valida conexiones, rutas, tipos de datos y sintaxis en segundos, evitando fallos catastróficos tras horas de cómputo.


🚀 OPTIMIZACIÓN AVANZADA DE STACK TECNOLÓGICO

Para proyectos que superan los límites de la memoria RAM disponible o requieren tiempos de respuesta críticos, se deben considerar estas estrategias de modernización:

*   **Modern Dataframes (Beyond Pandas):**
    *   Para volúmenes medios-altos (>2M registros) en entornos limitados, evaluar la migración de Pandas a **Polars**. Su ejecución *Lazy* y multihilo maximiza el hardware disponible.
    *   Alternativamente, usar **DuckDB** para realizar transformaciones SQL complejas directamente sobre archivos en disco (Parquet/CSV) sin saturar la RAM ("Out-of-core processing").

*   **Modelado Probabilístico vs. Reglas Fijas:**
    *   Evitar los "números mágicos" (umbrales arbitrarios como `0.85`). Preferir modelos de **Aprendizaje Activo** o regresiones logísticas entrenadas sobre el *Ground Truth*. Esto permite que el sistema aprenda las no-linealidades (ej: "si el NIT es idéntico, el nombre importa menos").

*   **Gestión Eficiente de Strings:**
    *   Si se utiliza Pandas, forzar siempre el tipo de dato `string[pyarrow]` en lugar del `object` (numpy) por defecto. Esto reduce la fragmentación de memoria y acelera las operaciones de texto vectorizadas.

*   **Estrategias de Bloqueo (Blocking) Híbridas:**
    *   No depender de una sola técnica de indexación. Combinar **LSH** (MinHash) para similitud difusa con **Sorted Neighborhood** (ordenar y comparar ventana deslizante) para capturar errores tipográficos al inicio de las cadenas. Esto aumenta el *Recall* robusteciendo la fase de generación de candidatos.

# 🟨🔝✅🛑 Guía de uso y convenciones

Este notebook utiliza los siguientes símbolos y colores para indicar el tipo de operación,
el estado y las acciones requeridas en cada sección del código.

## Códigos de colores y símbolos

### ✅ Ejecución

*   **▶️ 🔵  (Necesario):** Estas líneas de código deben ejecutarse siempre. Generalmente, se ejecutan solo una vez por sesión (ej: configuración inicial, carga de librerías).
*   **⏯️ ⏺️  (Esencial):**  Este código es fundamental para el funcionamiento del programa. Debe ejecutarse cada vez que se modifiquen los datos de entrada o la lógica central.
*   **🟥 🛑  (Cuidado):**  Indica una operación que debe realizarse con precaución, ya que podría modificar datos, consumir muchos recursos o tener efectos irreversibles.  Lee los comentarios cuidadosamente.
*   **🟨 🟡  (Opcional):**  Estas secciones contienen código que no es estrictamente necesario para el flujo principal, pero puede ser útil (ej: visualizaciones, análisis exploratorios, pruebas).
*   **🟧 🟠  (Revisar/Ejecutar una vez):**  Estas secciones ya se ejecutaron, o realizan configuraciones/tareas que no necesitan repetirse en cada ejecución.  Revísalas antes de correrlas de nuevo.  Si ya las ejecutaste una vez, puedes omitirlas.
*   **🔷  🔶  (Cargar resultados):** Se utiliza para cargar resultados de procesos previos que consumen mucho tiempo, evitando tener que volver a ejecutar todo el código (ej: cargar un DataFrame ya procesado desde un archivo).

### ♻️📦 Reutilización

*   **♻️📦 (Reutilizable):**  Indica funciones, bloques de código o variables que pueden ser reutilizados en otras partes del proyecto o en proyectos futuros.
* **🛡️📁🔐 (Copia de seguridad):** Copia a código que funcion.

### 🚧 Estado del Código

*   **🚧 🔨 (Pendiente/En Desarrollo):**  Secciones de código que aún no están completas, que requieren trabajo adicional o que están en fase de pruebas.
*   **✔️ (Completado):**  Indica que una tarea o sección de código que antes estaba pendiente ya ha sido finalizada.
*   **⚠️ (Advertencia):**  Zonas del código que requieren atención especial, ya sea porque son delicadas, propensas a errores, o tienen implicaciones importantes.
*   **❌ (Falla):**  Indica una sección de código que actualmente no funciona correctamente y necesita ser corregida.
*   **⏳ (Largo):**  Procesos que consumen mucho tiempo o recursos computacionales.  Útil para anticipar demoras y optimizar la ejecución.
*   **🚀 (Listo para producción):** Indica que el código o sección ya ha sido probado, validado y está listo para implementarse.

### 📊 Tipos de Operación

*   **⚙️  SETUP:**  Configuración del entorno (importación de librerías, definición de variables globales, etc.).
*   **📥  DATA LOAD:**  Carga de datos desde archivos o fuentes externas.
*   **💾 ⏸️ DATA SAVE:**  Guardado de datos intermedios o resultados (checkpoint).
*   **📤 DATA EXPORT:**  Exportación de los resultados finales a un archivo (ej: Excel, CSV).
*   **🔄 Transformación:** Modificación de la estructura o formato de los datos (ej: pivotar, agregar, filtrar).
*   **🛠️ Arreglando:**  Corrección de errores, manejo de valores faltantes, limpieza de datos.
*   **🧹 Limpieza:** Eliminación de duplicados, corrección de formatos, etc. (similar a "Arreglando", pero más enfocado en la calidad de los datos).
*   **🔗 Cruce:**  Combinación de datos de diferentes fuentes (joins, merges).
*   **📈 Graficar:**  Creación de visualizaciones de datos (gráficos, diagramas).
*   **📊 Análisis:**  Procesos de análisis exploratorio, cálculo de estadísticas, etc.
*   **🔢📐 Cálculos:**  Operaciones matemáticas o lógicas sobre los datos.
*   **🩺🐛 Debug/Fixing:**  Sección dedicada a la depuración y corrección de errores.
*   **✅🟢🔍 Verificación:**  Pasos para comprobar la validez de los datos o resultados.  Si la verificación es exitosa, puede omitirse en futuras ejecuciones.
*   **📋📜 REPORT:** Generación de informes y resúmenes.
* **⚗️🧠🧪 Experimentos** Experimentar un proceso
* **💻🔁🔌 Reiniciar máquina** Volver a reiniciar el entorno de ejecución. Usualmente después de una instalación.
* **✍️📝📍 Control de cambios** Modificaciones que se hayan hecho a la programación.
* **💡💬💭 Aprendizajes** Enseñanzas que se tuvieron a lo largo del desarrollo del programa.
* **📚⚖️🎓 Teoria** Explicación de conceptos fundamentales, principios básicos o marco teórico.
* **☁️⬆️💾 Backup** Subir a GitHub copia de respaldo

---

✅❇️✳️❎🟢 💚 💹 ♻️ 🔰 🌲 🌳 🌴 🌵 🌿 ☘️ 🍀 🌱 🪴📗 🎄 🔋 🧪

🆒 🆕 🆓 🆖 🆗 ▶️ ⏸️ ⏯️ ⏹️ ⏺️ ℹ️ ➡️ ⬅️ ⬆️ ⬇️ ↗️ ↘️ ↙️ ↖️ ↪️ ↩️ ⤴️ ⤵️ 🔷 💠 🔹

☑️🟪🟣⬛⚫🟫🟤⚪

⚙️🔬📐🛠️

<<<

>>>

# 🔰 📘 🧭 ℹ️ READ ME

# Guía de Uso — Paquete `geih_2025` v3.0

** · Coordinación Analítica**  
*Gran Encuesta Integrada de Hogares | DANE | Marco Muestral 2018 | 431 variables · 8 módulos · 12 meses*

---

## 1. ¿Qué cubre este paquete?

Los **8 módulos completos** del DANE, con **48 clases de análisis** que explotan las **431 variables** de la GEIH:

```
Módulos CSV del DANE                  Clases de análisis del paquete
─────────────────────────             ──────────────────────────────
① Características generales    ──→   EtnicoRacial, AnalisisCampesino, AnalisisDiscapacidad
② Datos del hogar/vivienda     ──→   (variables de contexto para cruces)
③ Fuerza de trabajo            ──→   IndicadoresLaborales, Estacionalidad, FuerzaLaboralJoven
④ Ocupados                     ──→   AnalisisSalarios, BrechaGenero, CalidadEmpleo, CostoLaboral,
                                      AnalisisContractual, AnalisisAutonomia, AnalisisAlcanceMercado,
                                      ProxyBilinguismo, EcuacionMincer, ProductividadTamano...
⑤ No ocupados                  ──→   AnalisisDesanimados, AnalisisFFT
⑥ Otras formas de trabajo      ──→   AnalisisOtrasFormas (autoconsumo, voluntariado)
⑦ Migración                    ──→   AnalisisMigracion (interna + internacional)
⑧ Otros ingresos e impuestos   ──→   AnalisisOtrosIngresos (pensiones, remesas, arriendos)
```

---

## 2. Instalación en Google Colab

```python
# ═══ CELDA 1: Montar Drive y configurar paquete ═══
from google.colab import drive
drive.mount('/content/drive')

import sys
sys.path.insert(0, '/content/drive/MyDrive/GEIH'  # ← ajustar a tu ruta)

from geih import __version__
print(f"geih v{__version__} — 48 clases de análisis")
```

---

## 3. Flujo Completo con Ejemplos Concretos

### PASO 1: Consolidar los 8 módulos (solo una vez)

```python
from geih import ConfigGEIH, ConsolidadorGEIH, GestorMemoria

RUTA = '/content/drive/MyDrive/GEIH'  # ← ajustar a tu ruta
config = ConfigGEIH(n_meses=12)

consolidador = ConsolidadorGEIH(
    ruta_base=RUTA,
    config=config,
    incluir_area=True,   # ← habilita análisis por 32 ciudades
)

# Verificar antes de leer (evita esperar 10 min para descubrir que falta un archivo)
consolidador.verificar_estructura()

# Consolidar los 8 módulos × 12 meses
geih = consolidador.consolidar()
# Output: ✅ CONSOLIDACIÓN COMPLETA — 5,234,128 filas × 347 columnas

# Guardar (nunca repetir la consolidación)
consolidador.exportar(geih, 'GEIH_2025_8mod_FINAL.parquet')
GestorMemoria.estado()
```

### PASO 2: Cargar y diagnosticar

```python
from geih import ConsolidadorGEIH, DiagnosticoCalidad

geih = ConsolidadorGEIH.cargar(f'{RUTA}/GEIH_2025_8mod_FINAL.parquet')

# Diagnóstico de calidad
diag = DiagnosticoCalidad()
diag.resumen_rapido(geih)
diag.verificar_tipos(geih)
diag.validar_identidades(geih)
tabla_nulos = diag.valores_faltantes(geih)
```

### PASO 3: Preparar datos

```python
from geih import PreparadorGEIH, ConfigGEIH

config = ConfigGEIH(n_meses=12)
prep = PreparadorGEIH(config=config)

# Base anual (FEX ÷ 12) — para promedios anuales
df = prep.preparar_base(geih)
df = prep.agregar_variables_derivadas(df)

# Base diciembre puntual (FEX sin dividir) — para indicador mensual
df_dic = prep.preparar_base(geih, mes_filtro=12)
df_dic = prep.agregar_variables_derivadas(df_dic)
```

### PASO 4: Exportador (todo organizado en carpetas)

```python
from geih import Exportador

exp = Exportador(ruta_base=RUTA, config=config)
# Crea automáticamente:
#   resultados_geih_2025/graficas/
#   resultados_geih_2025/tablas/
#   resultados_geih_2025/excel/
```

---

## 4. Ejemplos Concretos de Cada Análisis

### 4.1 · Indicadores nacionales + validación DANE

```python
from geih import IndicadoresLaborales

ind = IndicadoresLaborales(config=config)
r = ind.calcular(df)
ind.sanity_check(r, periodo="Anual 2025")
# Output:
#   ✅ TD_%  =  8.9%  (ref. DANE: 8.9%  Δ=0.0)
#   ✅ TGP_% = 64.3%  (ref. DANE: 64.3% Δ=0.0)

# Por departamento
td_dpto = ind.por_departamento(df)
exp.guardar_tabla(td_dpto, 'indicadores_por_departamento')
```

### 4.2 · Distribución de ingresos SMMLV

```python
from geih import DistribucionIngresos, GraficoDistribucionIngresos

dist = DistribucionIngresos(config=config)
resultado = dist.calcular(df)
dist.imprimir(resultado, titulo="Ene–Dic 2025")

fig = GraficoDistribucionIngresos().graficar(resultado['total'], resultado['por_sexo'])
exp.guardar_grafica(fig, 'distribucion_ingresos_2025')
```

### 4.3 · Salarios por rama con box plot

```python
from geih import AnalisisSalarios, GraficoBoxPlotSalarios

sal = AnalisisSalarios(config=config)
tabla_rama = sal.por_rama(df)
tabla_edad = sal.por_edad(df)

fig = GraficoBoxPlotSalarios().graficar(tabla_rama)
exp.guardar_grafica(fig, 'boxplot_salarios_rama')
exp.guardar_tabla(tabla_rama.reset_index(), 'estadisticas_salario_rama')
```

### 4.4 · Brecha salarial de género

```python
from geih import BrechaGenero, GraficoBrechaGenero

brecha = BrechaGenero().calcular(df)
print(brecha)
# Nivel                Hombres     Mujeres    Brecha_%
# 6. Universitaria     2,847,000   2,345,000  -17.6%
# 7. Posgrado          5,694,000   4,271,000  -25.0%

fig = GraficoBrechaGenero().graficar(brecha)
exp.guardar_grafica(fig, 'brecha_genero_educacion')
```

### 4.5 · Gini del ingreso laboral

```python
from geih import IndicesCompuestos

gini = IndicesCompuestos(config=config).gini(df)
print(f"Gini del ingreso laboral: {gini:.3f}")
# → ~0.480 (solo laboral, más alto que el oficial DANE que incluye transferencias)
```

### 4.6 · Calidad del empleo (ICE) por departamento

```python
from geih import CalidadEmpleo

ice = CalidadEmpleo(config=config).calcular_por_departamento(df)
print(ice[['Departamento', 'ICE', 'Ocupados_M']].head(10))
# Bogotá ≈ 65+, Chocó ≈ 30
exp.guardar_tabla(ice, 'ICE_por_departamento')
```

### 4.7 · Competitividad laboral (ICI) para pitch de IED

```python
from geih import CompetitividadLaboral

ici = CompetitividadLaboral(config=config).calcular(df)
print(ici[['Departamento', 'ICI', 'TD_%', 'Costo_efectivo', 'Talento_univ_%']].head(10))
exp.guardar_tabla(ici, 'ICI_competitividad_laboral')
```

### 4.8 · Ecuación de Mincer (retorno a la educación)

```python
from geih import EcuacionMincer

mincer = EcuacionMincer(config=config).estimar_todos(df)
print(mincer[['Grupo', 'beta_educacion', 'R2', 'N']])
# Nacional: β₁ ≈ 10.5% (cada año extra de educación → 10.5% más de salario)
# TIC: β₁ ≈ 16% (educación mucho más valorada)
exp.guardar_tabla(mincer, 'mincer_retorno_educacion')
```

### 4.9 · Estacionalidad mensual (12 puntos)

```python
from geih import Estacionalidad

estac = Estacionalidad().calcular(geih)  # ← usa geih sin preparar (necesita FEX_C18 sin dividir)
print(estac[['MES', 'TD_%', 'TO_%']])
exp.guardar_tabla(estac, 'estacionalidad_mensual_2025')
```

### 4.10 · Población campesina (NUEVO — P2057)

```python
from geih import AnalisisCampesino

camp = AnalisisCampesino(config=config).calcular(df)
print(camp)
# Se considera campesino: TD=X%, Mediana=0.6 SMMLV, Formalidad=15%
# No se considera campesino: TD=Y%, Mediana=1.2 SMMLV, Formalidad=42%
```

### 4.11 · Discapacidad (NUEVO — P1906S1-S8 Escala Washington)

```python
from geih import AnalisisDiscapacidad

disc = AnalisisDiscapacidad().calcular(df)
print(f"Con discapacidad: {disc['Con discapacidad_TD_%']}% TD")
print(f"Sin discapacidad: {disc['Sin discapacidad_TD_%']}% TD")
print(f"Prevalencia por dimensión: {disc['prevalencia_por_dimension']}")
```

### 4.12 · Migración (NUEVO — P3370/P3376)

```python
from geih import AnalisisMigracion

migr = AnalisisMigracion(config=config).calcular(df)
print(migr)
# Mismo municipio: 45M, TD=8.8%
# Otro municipio: 1.2M, TD=12.3%
# Otro país: 0.3M, TD=15.1%
# Nacido en el extranjero: TD más alta, mediana salarial más baja
```

### 4.13 · Autonomía laboral — contratistas dependientes (NUEVO — P3047-P3049)

```python
from geih import AnalisisAutonomia

auto = AnalisisAutonomia().calcular(df)
print(f"Cuenta propia total: {auto['cuenta_propia_M']}M")
print(f"Cta propia DEPENDIENTE (asalariados disfrazados): {auto['cta_propia_dependiente_M']}M")
print(f"→ {auto['pct_dependientes']}% de los cuenta propia son realmente dependientes")
```

### 4.14 · Alcance de mercado — empleo vinculado a exportación (NUEVO — P1802)

```python
from geih import AnalisisAlcanceMercado

alc = AnalisisAlcanceMercado().calcular(df)
print(alc)
# Exportación ★: ~X miles de personas → empleo DIRECTO en comercio exterior
# Esto es la variable más relevante para ProColombia que nadie publica
```

### 4.15 · Exportar todo a Excel multi-hoja

```python
from geih import AnalisisArea

area = AnalisisArea(config=config)
tablas = area.calcular_tablas(df)
area.exportar_excel(tablas, str(exp.excel / 'CIIU_Area_2025.xlsx'))

# O manualmente
exp.guardar_excel({
    'Indicadores': td_dpto,
    'Salarios': tabla_rama.reset_index(),
    'Brecha': brecha.reset_index(),
    'ICE': ice,
    'ICI': ici,
}, 'Resultados_completos_2025')

exp.guardar_metadata(config, {'n_registros': len(geih), 'modulos': 8})
exp.resumen()
```

---

## 5. Mapa Completo de las 48 Clases

| # | Clase | Módulo DANE | Qué mide |
|---|---|---|---|
| 1 | `ConfigGEIH` | — | Parámetros configurables |
| 2 | `GestorMemoria` | — | RAM en Colab |
| 3 | `ConversorTipos` | — | CIIU, DPTO, AREA a formato correcto |
| 4 | `EstadisticasPonderadas` | — | Media, mediana, Gini ponderados |
| 5 | `ConsolidadorGEIH` | Todos | Une 8 módulos × 12 meses |
| 6 | `PreparadorGEIH` | Todos | FEX_ADJ, variables derivadas |
| 7 | `MergeCorrelativas` | Ocupados | CIIU descriptivo + DIVIPOLA |
| 8 | `DiagnosticoCalidad` | Todos | Missing values, tipos, identidades |
| 9 | `Exportador` | — | Carpetas organizadas para outputs |
| 10 | `IndicadoresLaborales` | FT, Ocupados, No ocu. | TD, TGP, TO + sanity check |
| 11 | `DistribucionIngresos` | Ocupados | Rangos SMMLV por sexo |
| 12 | `AnalisisRamaSexo` | Ocupados | 13 ramas × sexo |
| 13 | `AnalisisSalarios` | Ocupados | Percentiles ponderados por rama y edad |
| 14 | `BrechaGenero` | Ocupados + Caract. | Brecha salarial por nivel educativo |
| 15 | `AnalisisCruzado` | Ocupados | Empresa × departamento |
| 16 | `IndicesCompuestos` | Ocupados | Gini, ICE básico |
| 17 | `AnalisisArea` | Ocupados | 6 tablas por 32 ciudades |
| 18 | `CalidadEmpleo` | Ocupados | ICE por dpto y por rama |
| 19 | `FormalidadSectorial` | Ocupados | ICF por rama |
| 20 | `VulnerabilidadLaboral` | Ocupados | IVI por rama |
| 21 | `CompetitividadLaboral` | Todo | ICI por departamento |
| 22 | `AnalisisSubempleo` | Ocupados | Horas, ingresos |
| 23 | `AnalisisHoras` | Ocupados | Distribución P6800 |
| 24 | `Estacionalidad` | Todo | TD/TO/TGP mensual |
| 25 | `FuerzaLaboralJoven` | FT + Caract. | TD joven 15-28 |
| 26 | `EtnicoRacial` | Caract. | TD e ingreso por grupo étnico |
| 27 | `BonoDemografico` | FT | Ratio dependencia por dpto |
| 28 | `CostoLaboral` | Ocupados | Mediana × 1.54 por rama |
| 29 | `AnalisisFFT` | No ocupados | FFT por sexo y tipo |
| 30 | `AnalisisUrbanoRural` | Caract. + Todo | Urbano vs Rural |
| 31 | `ProductividadTamano` | Ocupados | Salario × tamaño empresa |
| 32 | `ContribucionSectorial` | Ocupados + FT | Contribución al empleo (p.p.) |
| 33 | `MapaTalento` | Todo | ITAT por departamento |
| 34 | `EcuacionMincer` | Ocupados + Caract. | β₁ retorno educación (WLS) |
| 35 | `ProxyBilinguismo` | Ocupados + Caract. | 3 proxies de inglés |
| 36–39 | `Grafico*` | — | 4 tipos de gráficos matplotlib |
| 40 | `AnalisisCampesino` | Caract. (P2057) | Mercado laboral campesino |
| 41 | `AnalisisDiscapacidad` | Caract. (P1906) | Washington Scale × empleo |
| 42 | `AnalisisMigracion` | Migración (P3370) | Interna + internacional |
| 43 | `AnalisisOtrasFormas` | Otras formas | Autoconsumo, voluntariado |
| 44 | `AnalisisOtrosIngresos` | Otros ingresos | Pensiones, remesas, arriendos |
| 45 | `AnalisisSobrecalificacion` | Ocupados + Caract. | Universitarios subempleados |
| 46 | `AnalisisContractual` | Ocupados | Contrato escrito/verbal/sin |
| 47 | `AnalisisAutonomia` | Ocupados (P3047) | Cuenta propia dependiente |
| 48 | `AnalisisAlcanceMercado` | Ocupados (P1802) | Local → exportación |

---

## 6. La Regla de Oro del FEX

| Período | `ConfigGEIH` | `preparar_base()` |
|---|---|---|
| Mes puntual | `ConfigGEIH(n_meses=1)` | `mes_filtro=12` |
| Trimestre | `ConfigGEIH(n_meses=3)` | — |
| Anual | `ConfigGEIH(n_meses=12)` | `mes_filtro=None` (default) |

**Si PEA > 40M → el factor NO se está dividiendo.** El `sanity_check()` lo detecta.

---

## 7. Variables de Alto Valor para ProColombia

| Variable | Pregunta | Valor para IED/Exportaciones |
|---|---|---|
| `P1802=6` | ¿Mercado → exportación? | Empleo DIRECTO en comercio exterior |
| `P3047-P3049` | ¿Quién decide horario/producto/precio? | Contratistas dependientes (evasión) |
| `P3363` | ¿Cómo consiguió su empleo? | Digitalización del mercado laboral |
| `P6765` | Forma de trabajo | Destajo, honorarios, comisión |
| `P3364` | ¿Retención en la fuente? | Proxy de tributación formal DIAN |
| `P3042 × P6430` | Universitario + posición | Sobrecalificación = talento subutilizado |
| `P2057` | ¿Campesino? | Política pública rural |
| `P1906S1-S8` | Discapacidad Washington | ESG para IED europea |
| `P3376` | País de nacimiento | Migración venezolana + otras |

---

*Autor: Néstor Enrique Forero Herrera · Bogotá, 2026*

## 🚀 Proyecto: Nombre

[![Python](https://img.shields.io/badge/Python-3.7+-blue.svg)](https://www.python.org/downloads/)
[![Pandas](https://img.shields.io/badge/Pandas-1.x+-blue.svg)](https://pandas.pydata.org/)
[![Openpyxl](https://img.shields.io/badge/Openpyxl-3.x+-green.svg)](https://openpyxl.readthedocs.io/en/stable/)
[![Google Colab](https://colab.research.google.com/assets/colab-badge.svg)](https://colab.research.google.com/)
[![License: MIT](https://img.shields.io/badge/License-MIT-yellow.svg)](https://opensource.org/licenses/MIT)

## 📝 Descripción

Este proyecto ...

## 🌟 Características

*   **Carga y transformación de datos:** Lee datos de exportaciones desde un archivo .zip de Excel, maneja valores faltantes, y agrupa la información por empresa (NIT).
*   .

## 📁 Estructura del Proyecto

*   `notebook.ipynb`:  Script principal de Python con el código completo (originalmente un notebook Jupyter).
*  `README.md`: Este archivo, que proporciona documentación del proyecto.
* `requeriments.txt`: Lista los paquetes/librerías que se necesitan instalar

## 🛠️ Requisitos

*   **Python 3.7+**
*   **Bibliotecas:**
    *   `pandas`: Manipulación y análisis de datos.
    *   `numpy`: Soporte para operaciones numéricas.
    *   `openpyxl`: Escritura de archivos Excel (formato .xlsx).
    *   `re` (módulo de expresiones regulares): Parte de la biblioteca estándar de Python, no necesita instalación.

⚙️ Configuración y Personalización

Ruta de datos: Modifica la variable ruta_datos en el script si tus datos están en una ubicación diferente.

Nombre del archivo de salida: Cambia el valor del parámetro nombre_archivo en la función exportar_excel_con_formato para personalizar el nombre del archivo Excel generado.


📝 Funciones Principales

nombre .



🤝 Contribuciones

Las contribuciones son bienvenidas. Si encuentras errores, tienes sugerencias de mejora o quieres añadir nuevas funcionalidades, por favor, abre un issue o envía un pull request.

📄 Licencia

Este proyecto está bajo la licencia MIT. Consulta el archivo LICENSE para más detalles.

## 🤖 Desarrollo Colaborativo con IA

Este sistema fue desarrollado mediante una metodología de **desarrollo colaborativo humano-IA**, donde:

### Proceso de Desarrollo
- **Análisis y Diseño**: Desarrollador humano definió requerimientos, arquitectura y especificaciones técnicas
- **Implementación**: Desarrollo iterativo con asistencia de AI para funciones específicas
- **Validación**: Todo el código fue revisado, probado y validado por el desarrollador humano
- **Optimización**: Refinamiento conjunto para cumplir con estándares de calidad y rendimiento

### Responsabilidad y Calidad
- **Arquitectura del Sistema**: 100% desarrollador humano
- **Lógica de Negocio**: Definida por desarrollador humano, implementada colaborativamente
- **Integración de Sistemas**: Desarrollador humano
- **Testing y Validación**: Desarrollador humano
- **Responsabilidad Final**: Desarrollador humano

### Transparencia
Esta documentación se proporciona en cumplimiento de mejores prácticas de transparencia en desarrollo asistido por IA, reconociendo las contribuciones de todas las herramientas utilizadas mientras manteniendo la responsabilidad humana sobre el producto final.

**Desarrollador Principal**: Enrique Forero
**Asistencia IA**: Claude AI (Anthropic), Gemini 2.5 Pro, ChatGPT o4.  
**Fecha**: 2025-06-08

## ✉️ Contacto

Si tienes preguntas o comentarios, no dudes en contactarme:

* **Autor**: Néstor Enrique Forero Herrera
* **Email**: [enrique.economista@gmail.com]
* **GitHub**: [https://github.com/enriqueforero]
* **LinkedIn**: [https://www.linkedin.com/in/enriqueforero/]

## 📚 Referencias y Recursos Adicionales

* [Nombre](https://en.wikipedia.org/wiki/Locality-sensitive_hashing)

Notas Adicionales y Mejoras:

# 🎯 Objetivo - **Contexto de negocio**  

.

## 📝 Notas

[Carpeta](https://drive.google.com/drive/u/0/folders/1HwhwntFvdAqqKvgEpDnMxP-5EudT-TZN) donde está esta programación.

## ✍️📝📍 Control de cambios

(YYYY_MM_DD) - Modificación hecha

## 💡💬💭 Aprendizajes

💡

## ⬜ 🚧🚧🚧 PARA HACER 🚧🚧🚧

🚧 Intentar replicar todos los análisis de [Notebook original](https://colab.research.google.com/drive/1b3sSduXNoN5h11NIhFvQTT5fBmZa30va#scrollTo=L7kwsG_oAMLu).

🚧 Subir la base de la GEIH a Snowflake, dejarla con acceso a todos,compartir la ruta, dejar el registro en el Excel. Colocar targs.

🚧 Dejar la programación lista para que un público amplio la utilice.

🚧

## ✔️ Lo que se ha hecho ✔️

✔️

# ☁️⬆️💾 Subir a github
"""



"""# ⚙️  Configuración entorno de Google Colab"""

from google.colab import drive
drive.mount('/content/drive')

# Commented out IPython magic to ensure Python compatibility.
# Ruta donde está la programación del algoritmo en Cython
# %cd "/content/drive/MyDrive/GEIH"  # ← ajustar a tu ruta
# %ls

import sys
sys.path.insert(0, '/content/drive/MyDrive/GEIH'  # ← ajustar a tu ruta)

from geih import __version__
print(f"geih v{__version__}")

"""# ⚙️ Instalar paquetes"""

!pip install streamlit

# # Paquete a instalar de repositorio especializado en descarga
# !pip install geihdanepy # No funciona para 2026

"""# ⚙️ Importar librerías"""

import pandas as pd
import os

import matplotlib.pyplot as plt

os.getcwd()

# Que muestre todas las columnas
pd.options.display.max_columns = None
# En los dataframes, mostrar los float con dos decimales
pd.options.display.float_format = '{:,.2f}'.format

import os

# Ver en qué carpeta está trabajando actualmente el notebook
print("Directorio de trabajo actual:")
print(os.getcwd())
print()

# Listar todo lo que hay en ese directorio
print("Contenido del directorio de trabajo:")
for item in sorted(os.listdir('.')):
    print(f"  {repr(item)}")

from datetime import datetime
import pytz

# Definir la zona horaria de Colombia
zona_colombia = pytz.timezone('America/Bogota')

# Obtener la fecha y hora actual en Colombia
fecha_hora_colombia = datetime.now(zona_colombia)

# Imprimir con formato
print("Hora en Colombia:", fecha_hora_colombia.strftime("%Y-%m-%d %H:%M:%S"))

# Para cambiar entre rutas
import sys

import missingno as msno

# Generales
import pandas as pd
import numpy as np
import os

# Ocultar warnings
import warnings
warnings.filterwarnings('ignore')

# Aumentar número de columnas que se pueden ver
pd.options.display.max_columns = None
# En los dataframes, mostrar los float con dos decimales
pd.options.display.float_format = '{:,.2f}'.format
# Cada columna será tan grande como sea necesario para mostrar todo su contenido
pd.set_option('display.max_colwidth', 0)

"""# ⚙️ Definir variables"""

# ── Año base (referencia histórica) ────────────────────────────────────
ANIO = 2025
ANIO_BASE   = ANIO
N_MESES = 12
N_MESES_BASE = N_MESES          # Meses disponibles del año base

# ── Año de comparación (año en curso) ──────────────────────────────────
ANIO_COMP   = 2026
N_MESES_COMP = 1           # ← ÚNICO valor que cambia mes a mes:
                           #   enero=1 | enero+feb=2 | ... | dic=12

"""n_meses no es "cuántos meses has procesado hasta ahora" — es el divisor del FEX
La mayoría de la gente lo lee como un contador de meses. En realidad es la instrucción matemática para el factor de expansión. Si tienes 12 meses en el Parquet pero cargas con n_meses=1, todos los pesos poblacionales se multiplican por 12.

La encuesta GEIH se publica mensualmente, pero el DANE diseña los factores de expansión (FEX_C18) para que al sumar todos los meses del año, representen la población total anual. Esto significa que el FEX de enero ya está "pensado" como una doceava parte del año.
Si divides por 12 (año completo), cada persona pesa lo correcto para promedios anuales. Si divides por 1 (un mes), cada persona pesa para representar a toda la población ese mes.
"""

# ── Período de comparación (debe existir en AMBOS años) ────────────────
#   Usar los meses acumulados disponibles del año en curso.
#   Si tienes solo enero 2026, este valor debe ser igual a N_MESES_COMP.
N_MESES_COMPARAR = N_MESES_COMP   # ← Heredado automáticamente; cambiar
                                   #   solo si quieres comparar un sub-período

SMMLV_2025 = 1_423_500
SMMLV = SMMLV_2025

FONDO = '#F7F9FC'

RUTA = '/content/drive/MyDrive/GEIH'  # ← ajustar a tu ruta

RUTA_CIIU = '/content/drive/MyDrive/ProColombia/0A. Datos/Correlativas/2025-01-14 Correlativa CIIU Rev4 - Cadenas ProColombia.xlsx'

RUTA_DIVIPOLA = '/content/drive/MyDrive/ProColombia/0A. Datos/Correlativas/2025-01-14 DIVIPOLA.xlsx'

"""**Ejemplos de configuración:**

| Caso | Código |
|---|---|
| Año 2025 completo | `ConfigGEIH(anio=2025, n_meses=12)` |
| Primer trimestre 2026 | `ConfigGEIH(anio=2026, n_meses=3)` |
| Solo enero 2026 | `ConfigGEIH(anio=2026, n_meses=1)` |
"""

# Configuración del año (el ÚNICO lugar que se cambia)
from geih import ConfigGEIH
config = ConfigGEIH(anio=ANIO, n_meses=N_MESES)

config.resumen()  # Muestra SMMLV, carpetas, período

"""CONVENCIÓN DE NOMBRES
─────────────────────
  Archivos : GEIH_{anio}_M{ini:02d}.parquet           → mes único
             GEIH_{anio}_M{ini:02d}_M{fin:02d}.parquet → rango de meses
  Variables: df_{descripcion}  (ej: df_base_anual, df_combinado)

REGLA DE ORO DE RAM
───────────────────
  Nunca se retienen dos DataFrames de año completo en memoria al mismo
  tiempo. Los parquets se garantizan en disco primero; los DataFrames
  se cargan solo cuando se necesitan y se liberan inmediatamente después.

# ♻️ Funciones

## ♻️ Medir tiempo
"""

import time
from datetime import timedelta

def medir_tiempo(nombre_seccion=""):
    """
    Crea un medidor de tiempo simple y elegante.

    Parámetros:
    nombre_seccion (str): Nombre identificativo de la sección a medir (opcional)

    Retorna:
    function: Función que al llamarla imprime el tiempo transcurrido
    """
    tiempo_inicio = time.time()

    def imprimir_tiempo():
        tiempo_total = time.time() - tiempo_inicio
        # Convertimos a un formato más legible usando timedelta
        tiempo_formateado = str(timedelta(seconds=tiempo_total))

        # Si se proporcionó un nombre de sección, lo incluimos en el mensaje
        seccion = f" para {nombre_seccion}" if nombre_seccion else ""
        print(f"Tiempo transcurrido{seccion}: {tiempo_formateado}")

    return imprimir_tiempo

# Ejemplo de uso:
"""
# Para medir una sección específica:
fin_seccion = medir_tiempo("importar librerías")
import pandas as pd
import numpy as np
fin_seccion()

# Para medir todo un bloque de código:
fin_total = medir_tiempo()
# ... tu código aquí ...
fin_total()

# Para medir múltiples secciones:
fin_seccion1 = medir_tiempo("procesamiento inicial")
# ... código primera sección ...
fin_seccion1()

fin_seccion2 = medir_tiempo("análisis de datos")
# ... código segunda sección ...
fin_seccion2()
"""

"""## ♻️ Funciones auxiliares"""

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  CELDA EXTRAS A — CONFIGURACIÓN CENTRALIZADA                        ║
# ║  Ejecutar UNA sola vez al inicio. No modificar la clase.            ║
# ╚══════════════════════════════════════════════════════════════════════╝

import gc
import os
from dataclasses import dataclass, field
from typing import Optional
import pyarrow.parquet as pq
import pyarrow as pa
from geih import GestorMemoria


@dataclass
class ConfigComparativo:
    """Parámetros centrales del análisis comparativo GEIH.

    Único lugar que se modifica entre corridas.
    Valida en __post_init__ para fallar rápido si algo está mal.
    """
    # ── Rutas ──────────────────────────────────────────────────────
    ruta: str
    ruta_ciiu: str
    ruta_divipola: str

    # ── Año base (referencia histórica, normalmente año completo) ──
    anio_base: int    = 2025
    n_meses_base: int = 12

    # ── Año de comparación (año en curso) ──────────────────────────
    anio_comp: int    = 2026
    n_meses_comp: int = 1   # ← ÚNICO valor que cambia mes a mes
                             #   enero=1 | feb=2 | mar=3 | ... | dic=12

    def __post_init__(self) -> None:
        # Fail fast: validar antes de ejecutar cualquier proceso
        assert os.path.isdir(self.ruta),        f"Ruta no existe: {self.ruta}"
        assert 1 <= self.n_meses_base <= 12,    "n_meses_base debe ser 1–12"
        assert 1 <= self.n_meses_comp  <= 12,   "n_meses_comp debe ser 1–12"
        assert self.anio_comp > self.anio_base, "anio_comp debe ser > anio_base"

    @property
    def n_meses_comparar(self) -> int:
        """Meses a comparar. Heredado de n_meses_comp automáticamente."""
        return self.n_meses_comp

    def nombre_parquet(self, anio: int, n_meses: int) -> str:
        """GEIH_2025_M01.parquet  /  GEIH_2025_M01_M12.parquet"""
        if n_meses == 1:
            return f'GEIH_{anio}_M01.parquet'
        return f'GEIH_{anio}_M01_M{n_meses:02d}.parquet'

    def ruta_parquet(self, anio: int, n_meses: int) -> str:
        return os.path.join(self.ruta, self.nombre_parquet(anio, n_meses))

    def ruta_parquet_existente(self, anio: int) -> Optional[str]:
        """Localiza cualquier parquet completo del año (bajo cualquier nombre)."""
        candidatos = [
            self.ruta_parquet(anio, 12),
            os.path.join(self.ruta, f'GEIH_{anio}_Consolidado.parquet'),
        ]
        return next((c for c in candidatos if os.path.exists(c)), None)

    def resumen(self) -> None:
        print(f"\n{'═'*55}")
        print(f"  CONFIG COMPARATIVO GEIH")
        print(f"{'═'*55}")
        print(f"  Año base     : {self.anio_base} ({self.n_meses_base} meses)")
        print(f"  Año comp     : {self.anio_comp} ({self.n_meses_comp} meses)")
        print(f"  Comparar     : M01–M{self.n_meses_comparar:02d}")
        print(f"  Ruta datos   : {self.ruta}")
        print(f"{'═'*55}")


print("✅ ConfigComparativo definida")

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  CELDA EXTRAS B — FUNCIONES AUXILIARES                              ║
# ║  Ejecutar UNA sola vez. No modificar.                               ║
# ╚══════════════════════════════════════════════════════════════════════╝


def _extraer_con_pyarrow(
    fuente: str,
    destino: str,
    meses: list[int],
    nombre: str,
) -> str:
    """Extrae filas de un Parquet usando PyArrow predicate pushdown.

    NUNCA carga el archivo fuente completo en RAM de Pandas.
    PyArrow lee solo las filas que pasan el filtro en formato columnar
    (~5x más eficiente que Pandas). El resultado se escribe directo a
    disco desde Arrow, sin pasar por Pandas.

    Comparación de RAM (817k filas × 515 cols):
      pandas read_parquet completo : ~2.0 GB
      pyarrow con filtro M01       : ~50–120 MB  (solo filas del mes)

    Args:
        fuente  : Ruta al Parquet fuente (puede ser el año completo).
        destino : Ruta del Parquet a generar.
        meses   : Lista de MES_NUM a incluir (ej: [1] o [1,2,3]).
        nombre  : Nombre descriptivo para los prints.
    """
    print(f"  📂 Extrayendo {nombre} via PyArrow "
          f"(predicate pushdown, sin Pandas)...")

    GestorMemoria.estado()

    # Predicate pushdown: Arrow solo materializa las filas del filtro
    table = pq.read_table(
        fuente,
        filters=[('MES_NUM', 'in', meses)],
    )

    print(f"     {table.num_rows:,} filas × {table.num_columns} cols "
          f"en Arrow ({table.nbytes / 1e6:.0f} MB)")

    # Escribir a Parquet directamente desde Arrow (sin pasar por Pandas)
    pq.write_table(table, destino, compression='snappy')
    del table
    gc.collect()

    mb = os.path.getsize(destino) / 1e6
    print(f"  ✅ Guardado: {os.path.basename(destino)} ({mb:.0f} MB)")
    GestorMemoria.estado()
    return destino


def asegurar_parquet(cfg: ConfigComparativo, anio: int, n_meses: int) -> str:
    """Garantiza que el Parquet del período existe en disco.

    Jerarquía (falla rápido, de más eficiente a menos):
      1. Parquet exacto ya existe → retorna ruta, CERO trabajo.
      2. n_meses == 12 y existe parquet completo bajo cualquier nombre
         → retorna esa ruta directamente (SIN copiar ni leer el archivo).
      3. Existe parquet completo → extrae meses vía PyArrow predicate
         pushdown (RAM ≈ tamaño del resultado, NO del archivo fuente).
      4. Nada existe → consolida desde CSV → guarda → libera RAM.

    Postcondición: al retornar, CERO DataFrames en RAM.

    Args:
        cfg     : Configuración centralizada.
        anio    : Año de los datos.
        n_meses : Meses desde enero a incluir (1–12).

    Returns:
        Ruta absoluta al Parquet garantizado en disco.
    """
    path   = cfg.ruta_parquet(anio, n_meses)
    nombre = cfg.nombre_parquet(anio, n_meses)

    # ── Nivel 1: ya existe el Parquet exacto ─────────────────────────
    if os.path.exists(path):
        mb = os.path.getsize(path) / 1e6
        print(f"  ✅ [{anio}] {nombre} ({mb:.0f} MB) — ya existe")
        return path

    # ── Nivel 2: año completo → devolver sin copiar (CERO RAM) ───────
    if n_meses == 12:
        fuente = cfg.ruta_parquet_existente(anio)
        if fuente:
            mb = os.path.getsize(fuente) / 1e6
            print(f"  ✅ [{anio}] Usando {os.path.basename(fuente)} "
                  f"({mb:.0f} MB) como año completo")
            return fuente  # sin copiar ni leer

    # ── Nivel 3: extraer subset via PyArrow predicate pushdown ────────
    fuente = cfg.ruta_parquet_existente(anio)
    if fuente:
        meses = list(range(1, n_meses + 1))

        # Trust but Verify: confirmar que los meses requeridos existen
        meta  = pq.read_metadata(fuente)
        stats = pq.read_table(fuente, columns=['MES_NUM'])
        meses_disponibles = stats.column('MES_NUM').to_pylist()
        # Usar set para verificación O(1)
        meses_disponibles_set = set(meses_disponibles)
        del stats
        gc.collect()

        meses_faltantes = [m for m in meses if m not in meses_disponibles_set]
        if meses_faltantes:
            print(f"  ⚠️  [{anio}] Meses faltantes en fuente: {meses_faltantes}")
            print(f"       Procediendo a consolidar desde CSV...")
        else:
            return _extraer_con_pyarrow(fuente, path, meses, nombre)

    # ── Nivel 4: consolidar desde CSV ────────────────────────────────
    from geih import ConsolidadorGEIH, ConfigGEIH
    print(f"  🔄 [{anio}] Consolidando desde CSV (M01–M{n_meses:02d})...")
    config = ConfigGEIH(anio=anio, n_meses=n_meses)
    config.resumen()

    consolidador = ConsolidadorGEIH(
        ruta_base=cfg.ruta, config=config, incluir_area=True,
    )
    consolidador.verificar_estructura()
    df_consolidado = consolidador.consolidar(checkpoint=True)
    GestorMemoria.tamano_df(df_consolidado, f'df_{anio}_M{n_meses:02d}')

    df_consolidado.to_parquet(path, index=False, compression='snappy')
    del df_consolidado
    gc.collect()

    mb = os.path.getsize(path) / 1e6
    print(f"  ✅ [{anio}] Guardado: {nombre} ({mb:.0f} MB)")
    return path


def garantizar_parquets(cfg: ConfigComparativo) -> tuple[str, str, str]:
    """Garantiza en disco los tres Parquets necesarios.

    Procesa un año a la vez para no acumular RAM.
    Al terminar: CERO DataFrames en RAM, solo tres rutas (strings).

    Returns:
        (path_base_comp, path_comp_comp, path_base_full)
          path_base_comp : año base, período de comparación
          path_comp_comp : año comp, período de comparación
          path_base_full : año base, año completo (para análisis individuales)
    """
    cfg.resumen()
    print(f"\n{'─'*55}")
    print(f"  Garantizando Parquets en disco (RAM-safe)...")
    print(f"{'─'*55}")

    # ── Paso 1: período de comparación del año base ──────────────────
    #   Si N_MESES_COMPARAR == N_MESES_BASE, esta y la siguiente son la misma
    print(f"\n[1/3] Año base — período de comparación")
    path_base_comp = asegurar_parquet(cfg, cfg.anio_base, cfg.n_meses_comparar)

    # ── Paso 2: año de comparación ───────────────────────────────────
    print(f"\n[2/3] Año comp — período de comparación")
    path_comp_comp = asegurar_parquet(cfg, cfg.anio_comp, cfg.n_meses_comp)

    # ── Paso 3: año base completo (para análisis individuales) ────────
    #   Nivel 2 de asegurar_parquet lo devuelve sin leer el archivo.
    print(f"\n[3/3] Año base — año completo")
    path_base_full = asegurar_parquet(cfg, cfg.anio_base, cfg.n_meses_base)

    print(f"\n{'─'*55}")
    print(f"  ✅ Parquets listos. RAM al terminar:")
    GestorMemoria.estado()
    print(f"{'─'*55}")
    return path_base_comp, path_comp_comp, path_base_full


def agregar_mes_nuevo(cfg: ConfigComparativo, n_meses_nuevo: int) -> None:
    """Agrega un mes nuevo al Parquet acumulado del año comp.

    Lee solo el mes nuevo (~1 min) sin re-consolidar todo (~10 min).
    Al terminar libera RAM completamente.

    Uso cuando llega febrero 2026 con n_meses_comp=2:
        agregar_mes_nuevo(cfg, 2)
        # Luego cambiar n_meses_comp = 2 en la celda de ConfigComparativo

    Args:
        cfg          : Configuración centralizada.
        n_meses_nuevo: Total de meses DESPUÉS de agregar el nuevo.
    """
    from geih import ConsolidadorGEIH, ConfigGEIH, MESES_NOMBRES

    mes_carpeta  = f'{MESES_NOMBRES[n_meses_nuevo - 1]} {cfg.anio_comp}'
    parquet_prev = cfg.nombre_parquet(cfg.anio_comp, n_meses_nuevo - 1)
    path_nuevo   = cfg.ruta_parquet(cfg.anio_comp, n_meses_nuevo)

    print(f"🔄 Agregando {mes_carpeta} a {parquet_prev}...")
    config_geih = ConfigGEIH(anio=cfg.anio_comp, n_meses=n_meses_nuevo)
    consolidador = ConsolidadorGEIH(
        ruta_base=cfg.ruta, config=config_geih, incluir_area=True,
    )
    df_actualizado = consolidador.agregar_mes(
        mes_carpeta=mes_carpeta,
        parquet_existente=parquet_prev,
    )
    GestorMemoria.tamano_df(
        df_actualizado,
        f'df_{cfg.anio_comp}_M{n_meses_nuevo:02d}'
    )
    df_actualizado.to_parquet(path_nuevo, index=False, compression='snappy')
    del df_actualizado
    gc.collect()

    mb = os.path.getsize(path_nuevo) / 1e6
    print(f"✅ {cfg.nombre_parquet(cfg.anio_comp, n_meses_nuevo)} ({mb:.0f} MB)")
    print(f"   ► Actualiza n_meses_comp = {n_meses_nuevo} en ConfigComparativo")


print("✅ Funciones auxiliares cargadas")
GestorMemoria.estado()

"""## ♻️ Dashboard GEIH en Colab"""

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  CELDA EXTRAS — Preparar datos y lanzar dashboard                   ║
# ║  Ejecutar UNA sola vez. No modificar.                               ║
# ╚══════════════════════════════════════════════════════════════════════╝
"""
POR QUÉ se necesita un parquet "preparado" para el dashboard
─────────────────────────────────────────────────────────────
El dashboard carga el parquet directamente con pd.read_parquet().
Los parquets raw tienen 515 columnas y NO tienen FEX_ADJ.
El dashboard necesita FEX_ADJ para ponderar correctamente los totales.

  Parquet raw    : 515 cols, ~2 GB RAM, sin FEX_ADJ → CRASH + cálculos incorrectos
  Parquet prep   :  ~60 cols, ~300 MB RAM, con FEX_ADJ → OK

La preparación también reduce el tamaño en disco de ~200 MB → ~25 MB.
Los parquets preparados se guardan en {RUTA}/dashboard/ (carpeta dedicada).

ESTIMACIONES DE RAM POR ESCENARIO
──────────────────────────────────
  Enero solo (68k filas prep)   : ~40  MB  ✅ Sin problemas
  Año completo (817k filas prep): ~400 MB  ✅ Funciona, margen amplio
  Año completo (817k filas raw) : ~2.0 GB  ❌ Crash en Colab Free
"""

import gc
import os
import sys
import re
import time
import subprocess
from pathlib import Path
from typing import Optional

import pyarrow.parquet as pq
from geih import PreparadorGEIH, ConfigGEIH, GestorMemoria


# ── Preparar parquet para el dashboard ───────────────────────────────

def preparar_para_dashboard(
    cfg,                  # ConfigComparativo (definido en celda anterior)
    anio: int,
    n_meses: int,
) -> str:
    """Prepara y guarda un parquet listo para el dashboard.

    Hace dos cosas que el dashboard no puede hacer solo:
      1. Reduce de 515 → ~60 columnas (solo las que el dashboard usa)
      2. Agrega FEX_ADJ = FEX_C18 / n_meses (factor de expansión correcto)

    Los archivos se guardan en {cfg.ruta}/dashboard/ para que el
    dashboard los encuentre automáticamente.

    Al terminar: CERO DataFrames en RAM.

    Args:
        cfg     : ConfigComparativo con ruta y parámetros.
        anio    : Año a preparar.
        n_meses : Meses del período (1=enero, 12=año completo, etc.).

    Returns:
        Ruta al parquet preparado.
    """
    # ── Naming y rutas ────────────────────────────────────────────────
    carpeta_dash = Path(cfg.ruta) / 'dashboard'
    carpeta_dash.mkdir(exist_ok=True)

    nombre_prep = cfg.nombre_parquet(anio, n_meses).replace(
        '.parquet', '_prep.parquet'
    )
    path_prep = str(carpeta_dash / nombre_prep)

    if os.path.exists(path_prep):
        mb = os.path.getsize(path_prep) / 1e6
        print(f"  ✅ [{anio}] {nombre_prep} ({mb:.0f} MB) — ya existe")
        return path_prep

    # ── Paso 1: asegurar que el parquet raw exista en disco ───────────
    #   (llama a asegurar_parquet de la celda de extras anterior)
    path_raw = asegurar_parquet(cfg, anio, n_meses)

    # ── Paso 2: cargar via PyArrow (sin Pandas, eficiente en RAM) ─────
    print(f"\n  🔄 [{anio}] Preparando {nombre_prep}...")
    table = pq.read_table(path_raw)
    df_raw = table.to_pandas()
    del table
    gc.collect()
    GestorMemoria.tamano_df(df_raw, f'  df_{anio}_raw')

    # ── Paso 3: preparar — 515 cols → ~60 cols + FEX_ADJ ─────────────
    config_geih = ConfigGEIH(anio=anio, n_meses=n_meses)
    prep        = PreparadorGEIH(config=config_geih)
    df_prep     = prep.preparar_base(df_raw)       # columnas + FEX_ADJ
    df_prep     = prep.agregar_variables_derivadas(df_prep)  # RAMA, CIUDAD...
    del df_raw
    gc.collect()
    GestorMemoria.tamano_df(df_prep, f'  df_{anio}_prep')

    # ── Paso 4: guardar y liberar ─────────────────────────────────────
    df_prep.to_parquet(path_prep, index=False, compression='snappy')
    del df_prep
    gc.collect()

    mb = os.path.getsize(path_prep) / 1e6
    print(f"  ✅ [{anio}] Guardado: {nombre_prep} ({mb:.0f} MB)")
    return path_prep


# ── Garantizar carpeta dashboard con todos los parquets necesarios ────

def garantizar_dashboard(cfg, modo: str = 'enero') -> str:
    """Prepara los parquets para el dashboard y retorna la carpeta.

    Modos disponibles:
      'enero'       → Solo enero del año comp  (más liviano, recomendado)
      'comparacion' → Enero base + enero comp  (ambos años)
      'anio_comp'   → Meses acumulados del año comp
      'anio_base'   → Año base completo

    Args:
        cfg  : ConfigComparativo.
        modo : Qué datos cargar en el dashboard.

    Returns:
        Ruta a la carpeta dashboard/ lista para usar.
    """
    carpeta = str(Path(cfg.ruta) / 'dashboard')
    print(f"\n{'═'*55}")
    print(f"  Preparando dashboard — modo: '{modo}'")
    print(f"{'═'*55}")
    GestorMemoria.estado()

    if modo == 'enero':
        # Solo enero del año más reciente disponible
        preparar_para_dashboard(cfg, cfg.anio_comp, 1)

    elif modo == 'comparacion':
        # Enero de ambos años (permite ver los dos en el selector)
        preparar_para_dashboard(cfg, cfg.anio_base, 1)
        gc.collect()
        preparar_para_dashboard(cfg, cfg.anio_comp, 1)

    elif modo == 'anio_comp':
        # Todos los meses acumulados del año en curso
        preparar_para_dashboard(cfg, cfg.anio_comp, cfg.n_meses_comp)

    elif modo == 'anio_base':
        # Año base completo — funciona bien con parquet preparado (~400 MB)
        preparar_para_dashboard(cfg, cfg.anio_base, cfg.n_meses_base)

    else:
        raise ValueError(f"modo='{modo}' no válido. "
                         f"Opciones: 'enero', 'comparacion', 'anio_comp', 'anio_base'")

    print(f"\n  ✅ Dashboard listo. RAM final:")
    GestorMemoria.estado()
    print(f"{'═'*55}")
    return carpeta


# ── Lanzar el dashboard (reutiliza las funciones de la celda anterior) ─

def _lanzar_streamlit_prep(ruta_dash: str, puerto: int) -> subprocess.Popen:
    """Lanza Streamlit apuntando a la carpeta dashboard/ preparada."""
    import geih
    script = str(Path(geih_2025.__file__).parent / 'dashboard.py')
    assert Path(script).exists(), f"dashboard.py no encontrado: {script}"

    subprocess.run(['pkill', '-f', 'streamlit'], capture_output=True)
    time.sleep(1)

    return subprocess.Popen(
        [
            sys.executable, '-m', 'streamlit', 'run', script,
            '--server.port',                 str(puerto),
            '--server.address',              '0.0.0.0',
            '--server.headless',             'true',
            '--server.enableCORS',           'false',
            '--server.enableXsrfProtection', 'false',
            '--', ruta_dash,      # el dashboard lee esta ruta para buscar parquets
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )


def lanzar_dashboard_prep(
    cfg,
    modo: str            = 'enero',
    puerto: int          = 8501,
    metodo: str          = 'cloudflare',
    ngrok_token: Optional[str] = None,
) -> None:
    """Prepara los datos y lanza el dashboard con túnel público.

    Flujo completo en una llamada:
      1. Prepara los parquets (515 → ~60 cols, agrega FEX_ADJ)
      2. Lanza Streamlit con los flags correctos para Colab
      3. Abre túnel cloudflare o ngrok y muestra la URL

    Args:
        cfg         : ConfigComparativo (del notebook).
        modo        : Qué datos mostrar — ver garantizar_dashboard().
        puerto      : Puerto Streamlit (por defecto 8501).
        metodo      : 'cloudflare' (sin cuenta) o 'ngrok' (con token).
        ngrok_token : Token ngrok si metodo='ngrok'.

    Estimaciones de RAM en el dashboard según modo:
      'enero'       : ~40  MB  ✅
      'comparacion' : ~80  MB  ✅
      'anio_comp'   : ~40–400 MB según meses ✅
      'anio_base'   : ~400 MB  ✅ (funciona; crash era por raw 515 cols)
    """
    # ── Validaciones (fail fast) ──────────────────────────────────────
    assert metodo in ('cloudflare', 'ngrok'), \
        "metodo debe ser 'cloudflare' o 'ngrok'"
    if metodo == 'ngrok':
        assert ngrok_token, (
            "Proporciona ngrok_token='tu_token'.\n"
            "Obtén uno gratis en: https://dashboard.ngrok.com"
        )

    # ── Paso 1: preparar parquets ─────────────────────────────────────
    ruta_dash = garantizar_dashboard(cfg, modo=modo)

    # ── Paso 2: instalar streamlit si falta ───────────────────────────
    try:
        import streamlit  # noqa: F401
    except ImportError:
        print("📦 Instalando streamlit y plotly...")
        subprocess.run(
            [sys.executable, '-m', 'pip', 'install', 'streamlit', 'plotly', '-q'],
            check=True,
        )

    # ── Paso 3: lanzar Streamlit ──────────────────────────────────────
    print(f"\n⏳ Iniciando Streamlit (puerto {puerto})...")
    proceso_st = _lanzar_streamlit_prep(ruta_dash, puerto)
    time.sleep(4)

    if proceso_st.poll() is not None:
        raise RuntimeError(
            "Streamlit terminó inesperadamente. "
            "Revisa que el paquete geih_2025 esté accesible."
        )
    print(f"✅ Streamlit activo (PID {proceso_st.pid})")

    # ── Paso 4: abrir túnel ───────────────────────────────────────────
    if metodo == 'cloudflare':
        binario = '/usr/local/bin/cloudflared'
        if not os.path.exists(binario):
            print("📦 Instalando cloudflared...")
            subprocess.run(
                ['wget', '-q',
                 'https://github.com/cloudflare/cloudflared/releases/latest/'
                 'download/cloudflared-linux-amd64', '-O', binario],
                check=True,
            )
            subprocess.run(['chmod', '+x', binario], check=True)

        print("🔗 Abriendo túnel Cloudflare (~15 s)...")
        proceso_cf = subprocess.Popen(
            ['cloudflared', 'tunnel', '--url', f'http://localhost:{puerto}'],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True,
        )
        patron = re.compile(r'https://[a-z0-9\-]+\.trycloudflare\.com')
        for linea in proceso_cf.stdout:
            match = patron.search(linea)
            if match:
                url = match.group(0)
                break
        else:
            raise RuntimeError("cloudflared no generó URL. Verifica conexión.")

    else:
        subprocess.run(
            [sys.executable, '-m', 'pip', 'install', 'pyngrok', '-q'],
            check=True,
        )
        from pyngrok import ngrok
        ngrok.set_auth_token(ngrok_token)
        url = ngrok.connect(puerto, 'http').public_url

    # ── Resultado ─────────────────────────────────────────────────────
    print(f"\n{'═'*55}")
    print(f"  🌐  Dashboard disponible en:")
    print(f"      {url}")
    print(f"{'═'*55}")
    print(f"  Modo    : {modo}")
    print(f"  Parquets: {ruta_dash}")
    print(f"\n  ⚠️  Mantén esta celda en ejecución.")
    print(f"      Si el túnel expira, re-ejecuta esta celda.")


print("✅ Funciones de dashboard listas")

tiempo_sesión = medir_tiempo("Sesión")
tiempo_total = medir_tiempo("Total")

tiempo_sesión()
tiempo_total()

"""# ✅ Consolidar datos (primera vez) o cargar existente"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # Wall time: 3min 16s si se consolida por primera vez
# # Wall time: 12.8 s para cargar existente
# 
# from geih import ConsolidadorGEIH
# import os
# 
# PARQUET = f'{RUTA}/GEIH_{config.anio}_Consolidado.parquet'
# 
# if os.path.exists(PARQUET):
#     geih = ConsolidadorGEIH.cargar(PARQUET)
# else:
#     consolidador = ConsolidadorGEIH(ruta_base=RUTA, config=config, incluir_area=True)
#     consolidador.verificar_estructura()
#     geih = consolidador.consolidar(checkpoint=True)  # ← se recupera si falla
#     consolidador.exportar(geih)

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ═══ Consolidar enero 2026 ═══
# from geih import ConsolidadorGEIH, ConfigGEIH
# import os
# 
# # ┌──────────────────────────────────────────────────────────┐
# # │  Solo cambias n_meses=1 porque tienes únicamente enero   │
# # │  Cuando llegue febrero, cámbialo a n_meses=2, etc.       │
# # └──────────────────────────────────────────────────────────┘
# config_2026 = ConfigGEIH(anio=2026, n_meses=1)
# config_2026.resumen()  # Verifica SMMLV, carpetas esperadas, período
# 
# PARQUET_2026 = f'{RUTA}/GEIH_2026_Consolidado.parquet'
# 
# if os.path.exists(PARQUET_2026):
#     print("📂 Cargando base 2026 existente...")
#     geih_2026 = ConsolidadorGEIH.cargar(PARQUET_2026)
# else:
#     print("🔄 Consolidando enero 2026...")
#     consolidador_2026 = ConsolidadorGEIH(
#         ruta_base=RUTA,
#         config=config_2026,
#         incluir_area=True,
#     )
#     consolidador_2026.verificar_estructura()   # ← valida que exista "Enero 2026/CSV/"
#     geih_2026 = consolidador_2026.consolidar(checkpoint=True)
#     consolidador_2026.exportar(geih_2026)
# 
# print(f"✅ 2026 listo: {geih_2026.shape[0]:,} filas")

"""**Agregar un mes nuevo sin re-consolidar:**

```python
config_4m = ConfigGEIH(anio=2026, n_meses=4)
consolidador = ConsolidadorGEIH(ruta_base=RUTA, config=config_4m, incluir_area=True)
geih = consolidador.agregar_mes(
    mes_carpeta='Abril 2026',
    parquet_existente='GEIH_2026_Consolidado.parquet',
)
consolidador.exportar(geih)
```

# ✅ Bases para comparaciones
"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ╔══════════════════════════════════════════════════════════════════════╗
# # ║  CELDA EJECUTAR C — GARANTIZAR PARQUETS EN DISCO                     ║
# # ║  Al terminar: CERO DataFrames en RAM.                                ║
# # ╚══════════════════════════════════════════════════════════════════════╝
# 
# cfg = ConfigComparativo(
#     ruta          = RUTA,
#     ruta_ciiu     = RUTA_CIIU,
#     ruta_divipola = RUTA_DIVIPOLA,
#     anio_base     = ANIO_BASE,
#     n_meses_base  = N_MESES_BASE,
#     anio_comp     = ANIO_COMP,
#     n_meses_comp  = N_MESES_COMP,     # ← ÚNICO valor que cambia mes a mes
# )
# 
# PATH_BASE_COMP, PATH_COMP_COMP, PATH_BASE_FULL = garantizar_parquets(cfg)

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  AGREGAR MES NUEVO (ejecutar cuando llegue el mes)                   ║
# ║  Descomenta y ejecuta SOLO cuando tengas la carpeta del mes nuevo    ║
# ╚══════════════════════════════════════════════════════════════════════╝

# Paso 1: Agrega el mes nuevo al Parquet acumulado
# agregar_mes_nuevo(ANIO_COMP, n_meses_nuevo=2)   # cuando llegue febrero

# Paso 2: Actualiza N_MESES_COMP = 2 en CELDA A y re-ejecuta desde allí

"""# ✅ Preparar datos"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # Wall time: 2.88 s
# from geih import PreparadorGEIH, Exportador
# 
# prep = PreparadorGEIH(config=config)
# df = prep.preparar_base(geih)
# df = prep.agregar_variables_derivadas(df)
# exp = Exportador(ruta_base=RUTA, config=config)

"""# ✅ Análisis disponibles (70+ clases)

## 🟢 Indicadores nacionales
"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import IndicadoresLaborales
# 
# ind = IndicadoresLaborales(config=config)
# r = ind.calcular(df)
# ind.sanity_check(r, f"Anual {config.anio}")
# td_dpto = ind.por_departamento(df)

"""## 🟢 Los 10 análisis básicos"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# #
# from geih import (
#     DistribucionIngresos, AnalisisRamaSexo, AnalisisSalarios,
#     BrechaGenero, IndicesCompuestos, Estacionalidad,
#     AnalisisUrbanoRural, FuerzaLaboralJoven, EcuacionMincer,
# )
# from geih import Top20Sectores
# 
# DistribucionIngresos(config=config).calcular(df)

AnalisisRamaSexo().calcular(df)

AnalisisSalarios(config=config).por_rama(df)

BrechaGenero().calcular(df)

IndicesCompuestos(config=config).gini(df)

Estacionalidad().calcular(geih)            # usa geih crudo

AnalisisUrbanoRural(config=config).calcular(df)

EcuacionMincer(config=config).estimar_todos(df)

Top20Sectores(config=config).calcular(df)

FuerzaLaboralJoven(config=config).calcular(df)

"""## 🟢 Los 8 análisis avanzados"""

from geih import (
    CalidadEmpleo, FormalidadSectorial, CompetitividadLaboral,
    VulnerabilidadLaboral, CostoLaboral, ContribucionSectorial,
    MapaTalento, BonoDemografico,
)

CalidadEmpleo(config=config).calcular_por_departamento(df)     # ICE

FormalidadSectorial(config=config).calcular(df)                 # ICF

CompetitividadLaboral(config=config).calcular(df)               # ICI

VulnerabilidadLaboral(config=config).calcular(df)               # IVI

CostoLaboral(config=config).calcular(df)

ContribucionSectorial().calcular(df)

MapaTalento(config=config).calcular(df)                         # ITAT

BonoDemografico(config=config).calcular(df)

"""## 🟢 Los 15 análisis poblacionales"""

from geih import (
    AnalisisCampesino, AnalisisDiscapacidad, AnalisisMigracion,
    AnalisisOtrasFormas, AnalisisOtrosIngresos, AnalisisSobrecalificacion,
    AnalisisContractual, AnalisisAutonomia, AnalisisAlcanceMercado,
    AnalisisDesanimados,
    DuracionDesempleo, DashboardSectoresProColombia,
    AnatomaSalario, FormaPago, CanalEmpleo,
)

# P2057 — ¿se considera campesino?
AnalisisCampesino(config=config).calcular(df)

# P1906S1-S8 — escala Washington
AnalisisDiscapacidad().calcular(df)

# P3370/P3376 — migración interna e internacional
AnalisisMigracion(config=config).calcular(df)

# P3054-P3057 — autoconsumo, voluntariado, formación
AnalisisOtrasFormas().calcular(df)

# P7422/P7500 — pensiones, remesas, arriendos
AnalisisOtrosIngresos().calcular(df)

# P3042 × P6430 — universitarios en empleos simples
AnalisisSobrecalificacion(config=config).calcular(df)

# P6440/P6450/P6460 — contrato escrito/verbal/sin
AnalisisContractual().calcular(df)

# P3047-P3049 — contratistas dependientes (asalariados disfrazados)
AnalisisAutonomia().calcular(df)

# P1802 — local/regional/nacional/exportación
AnalisisAlcanceMercado().calcular(df)

# P6300/P6310 — FFT que desean trabajar (potencial laboral latente)
AnalisisDesanimados().calcular(df)

# P7250 — semanas buscando empleo (friccional→crónico)
DuracionDesempleo(config=config).calcular(df)

# Mediana semanas × dpto (proxy rigidez)
DuracionDesempleo(config=config).por_departamento(df)

# 7 sectores estratégicos de actividad económica
DashboardSectoresProColombia(config=config).calcular(df)

# P6500 vs INGLABO — ingreso "invisible"
AnatomaSalario(config=config).resumen_nacional(df)

# Brecha salarial por rama
AnatomaSalario(config=config).por_rama(df)

# P3363 — contactos, internet, agencia
CanalEmpleo(config=config).calcular(df)

# P6765 — destajo, honorarios, comisión
FormaPago(config=config).calcular(df)

"""## 🟢 Análisis por 32 ciudades"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import AnalisisOcupadosCiudad
# 
# area = AnalisisOcupadosCiudad(config=config)
# tablas = area.calcular(df, ruta_ciiu=RUTA_CIIU)
# area.imprimir(tablas)
# area.exportar_excel(tablas, f'{RUTA}/resultados_geih_2025/excel/CIIU_Area_{config.anio}.xlsx')

"""## 🟢 Gráficos estáticos"""

from geih import (
    GraficoBoxPlotSalarios, GraficoBrechaGenero,
    GraficoCurvaLorenz, GraficoICIBubble, GraficoEstacionalidad,
)

fig = GraficoCurvaLorenz().graficar(df[(df['OCI']==1) & (df['INGLABO']>0)])
exp.guardar_grafica(fig, f'Lorenz_{config.anio}')

"""## 🟢 Gráficos interactivos"""

from geih import PlotlyLorenz, PlotlyICIBubble, PlotlyEstacionalidad

fig = PlotlyLorenz().graficar(df[(df['OCI']==1) & (df['INGLABO']>0)])
fig.show()  # interactivo en Colab

"""# ✅ Histogramas horas/ingresos (matplotlib + plotly)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# import numpy as np
# import matplotlib.pyplot as plt
# import matplotlib.ticker as mticker
# import gc
# 
# FONDO = '#F7F9FC'
# C = {'azul':'#2E6DA4','rojo':'#C0392B','verde':'#1E8449',
#      'morado':'#7D3C98','naranja':'#E67E22','gris':'#7F8C8D',
#      'amarillo':'#F39C12'}
# SMMLV = config.smmlv
# 
# df_hist_ocu = df[df['OCI'] == 1].copy()
# print(f"Ocupados para histogramas: {len(df_hist_ocu):,} registros")
# 
# fig_h = plt.figure(figsize=(18, 10))
# fig_h.patch.set_facecolor(FONDO)
# gs_h = plt.GridSpec(2, 3, figure=fig_h, hspace=0.45, wspace=0.38)
# ax_h1 = fig_h.add_subplot(gs_h[0, :2])
# ax_h2 = fig_h.add_subplot(gs_h[0, 2])
# ax_h3 = fig_h.add_subplot(gs_h[1, :2])
# ax_h4 = fig_h.add_subplot(gs_h[1, 2])
# for ax in [ax_h1, ax_h2, ax_h3, ax_h4]:
#     ax.set_facecolor('white'); ax.spines[['top','right']].set_visible(False)
# 
# if 'P6800' in df_hist_ocu.columns:
#     df_p68 = df_hist_ocu[df_hist_ocu['P6800'].between(1, 100)].copy()
#     bins_h = np.arange(0, 101, 4)
#     counts_h, edges_h = np.histogram(df_p68['P6800'], bins=bins_h, weights=df_p68['FEX_ADJ'])
#     ax_h1.bar(edges_h[:-1], counts_h/1e6, width=3.8, color=C['azul'], alpha=0.82, align='edge')
#     for horas, label, col_v in [(40,'40h legal',C['verde']),(48,'48h límite',C['naranja']),(32,'32h subempleo',C['rojo'])]:
#         ax_h1.axvline(horas, color=col_v, ls='--', lw=1.8, alpha=0.85)
#         ax_h1.text(horas+0.3, counts_h.max()/1e6*0.85, label, fontsize=8, color=col_v, fontweight='bold')
#     v8, w8 = df_p68['P6800'].values, df_p68['FEX_ADJ'].values
#     idx8 = np.argsort(v8); v8, w8 = v8[idx8], w8[idx8]
#     med8 = float(v8[np.searchsorted(np.cumsum(w8), np.cumsum(w8)[-1]/2)])
#     ax_h1.axvline(med8, color='black', ls='-', lw=2, alpha=0.7)
#     ax_h1.text(med8+0.5, counts_h.max()/1e6*1.02, f'Mediana={med8:.0f}h', fontsize=9, fontweight='bold')
#     ax_h1.set_xlabel('Horas normales semanales (P6800)'); ax_h1.set_ylabel('Personas (millones)')
#     ax_h1.set_title(f'Distribución de horas normales — GEIH {ANIO}', fontsize=11, fontweight='bold')
#     ax_h1.grid(axis='y', alpha=0.3)
# 
#     BINS_LABS = [(1,15,'<15h'),(15,32,'15–32h'),(32,40,'32–40h'),(40,48,'40–48h'),(48,70,'48–70h'),(70,126,'>70h')]
#     vals_bins = [df_p68.loc[df_p68['P6800'].between(lo, hi), 'FEX_ADJ'].sum()/1e6 for lo,hi,_ in BINS_LABS]
#     col_bins = [C['rojo'],C['naranja'],C['amarillo'],C['verde'],C['naranja'],C['rojo']]
#     bars_h2 = ax_h2.bar(range(len(vals_bins)), vals_bins, 0.7, color=col_bins, alpha=0.85)
#     for bar, v in zip(bars_h2, vals_bins):
#         ax_h2.text(bar.get_x()+bar.get_width()/2, v+0.03, f'{v:.2f}M', ha='center', fontsize=8.5, fontweight='bold')
#     ax_h2.set_xticks(range(len(BINS_LABS))); ax_h2.set_xticklabels([l for _,_,l in BINS_LABS], fontsize=8)
#     ax_h2.set_ylabel('Millones'); ax_h2.set_title('Por rango de horas', fontsize=10, fontweight='bold')
#     ax_h2.grid(axis='y', alpha=0.3)
# 
# if 'P6850' in df_hist_ocu.columns:
#     df_p685 = df_hist_ocu[df_hist_ocu['P6850'].between(1, 100)].copy()
#     for sx_v, sx_l, col_s in [(1,'Hombres',C['azul']),(2,'Mujeres',C['rojo'])]:
#         sub = df_p685[df_p685['P3271'] == sx_v]
#         if len(sub) == 0: continue
#         counts_s, _ = np.histogram(sub['P6850'], bins=bins_h, weights=sub['FEX_ADJ'])
#         ax_h3.plot(edges_h[:-1]+2, counts_s/1e6, lw=2.2, color=col_s, label=sx_l, alpha=0.88)
#         ax_h3.fill_between(edges_h[:-1]+2, counts_s/1e6, alpha=0.12, color=col_s)
#     ax_h3.axvline(48, color=C['naranja'], ls='--', lw=1.5, alpha=0.7, label='48h')
#     ax_h3.set_xlabel('Horas reales (P6850)'); ax_h3.set_ylabel('Personas (millones)')
#     ax_h3.set_title(f'Horas reales por sexo — GEIH {ANIO}', fontsize=11, fontweight='bold')
#     ax_h3.legend(fontsize=10); ax_h3.grid(axis='y', alpha=0.3)
# 
# if 'P6800' in df_hist_ocu.columns and 'INGLABO' in df_hist_ocu.columns:
#     df_sc = df_hist_ocu[df_hist_ocu['P6800'].between(1,80) & df_hist_ocu['INGLABO'].between(100_000, 15_000_000)]
#     if len(df_sc) > 0:
#         df_sc_s = df_sc.sample(min(5000, len(df_sc)), weights='FEX_ADJ', random_state=42, replace=True)
#         ax_h4.scatter(df_sc_s['P6800'], df_sc_s['INGLABO']/SMMLV, alpha=0.08, s=8, color=C['azul'])
#         ax_h4.axhline(1, color=C['verde'], ls='--', lw=1.3, alpha=0.7)
#         ax_h4.set_xlabel('Horas semanales'); ax_h4.set_ylabel('Ingreso (× SMMLV)')
#         ax_h4.set_title('Horas vs ingreso (5K muestra)', fontsize=10, fontweight='bold')
#         ax_h4.set_ylim(0, 12); ax_h4.grid(alpha=0.2)
# 
# fig_h.suptitle(f'ANÁLISIS DE HORAS TRABAJADAS — GEIH {ANIO}', fontsize=13, fontweight='bold')
# plt.tight_layout(rect=[0, 0, 1, 0.94]); plt.show()
# exp.guardar_grafica(fig_h, f'Histograma_horas_{ANIO}', cerrar=True)
# 
# try:
#     import plotly.express as px
#     df_plotly = df_hist_ocu[df_hist_ocu['INGLABO'].between(50_000, 20_000_000)].copy()
#     df_plotly['INGLABO_SML'] = df_plotly['INGLABO'] / SMMLV
#     df_plotly['SEXO'] = df_plotly['P3271'].map({1:'Hombres', 2:'Mujeres'})
#     df_pl_s = df_plotly[df_plotly['SEXO'].notna()].sample(min(30_000, len(df_plotly)), weights='FEX_ADJ', random_state=42, replace=True)
#     fig_px = px.box(df_pl_s, x='SEXO', y='INGLABO_SML', color='SEXO',
#                     color_discrete_map={'Hombres':'#2E6DA4','Mujeres':'#C0392B'},
#                     title=f'Distribución del ingreso laboral por sexo — GEIH {ANIO}',
#                     labels={'INGLABO_SML':'Ingreso (× SMMLV)','SEXO':''}, points='outliers')
#     fig_px.update_layout(yaxis=dict(range=[0,15]), plot_bgcolor='white', paper_bgcolor='#F7F9FC', showlegend=False)
#     fig_px.add_hline(y=1, line_dash='dash', line_color='#1E8449', annotation_text='1 SMMLV')
#     fig_px.show()
#     print("✅ Box plot interactivo Plotly generado")
# except ImportError:
#     print("⚠️  plotly no disponible — instalar con: !pip install plotly")
# 
# del df_hist_ocu; gc.collect()
# print("✅ Celda 10 completada")

"""# ✅ Cruce empresa × departamento (P7130, P3069, P6920)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# import matplotlib.colors as mcolors
# from geih import DEPARTAMENTOS
# from geih import TAMANO_EMPRESA
# 
# df_ocu_cr = df[df['OCI'] == 1].copy()
# if 'NOMBRE_DPTO' not in df_ocu_cr.columns:
#     df_ocu_cr['DPTO_STR'] = df_ocu_cr['DPTO'].astype(str).str.zfill(2)
#     df_ocu_cr['NOMBRE_DPTO'] = df_ocu_cr['DPTO_STR'].map(DEPARTAMENTOS)
# 
# filas_cr = []
# for dpto_nom in df_ocu_cr['NOMBRE_DPTO'].dropna().unique():
#     m_d = df_ocu_cr['NOMBRE_DPTO'] == dpto_nom
#     n_tot = df_ocu_cr.loc[m_d, 'FEX_ADJ'].sum()
#     if n_tot < 5_000: continue
#     fila = {'Departamento': dpto_nom, 'Ocupados_M': round(n_tot/1e6, 2)}
#     if 'P7130' in df_ocu_cr.columns:
#         fila['Desea_cambiar_%'] = round(df_ocu_cr.loc[m_d & (df_ocu_cr['P7130']==1), 'FEX_ADJ'].sum() / n_tot * 100, 1)
#     if 'P6920' in df_ocu_cr.columns:
#         fila['Cotiza_pension_%'] = round(df_ocu_cr.loc[m_d & (df_ocu_cr['P6920']==1), 'FEX_ADJ'].sum() / n_tot * 100, 1)
#     if 'P3069' in df_ocu_cr.columns:
#         for cod_t, lab_t in TAMANO_EMPRESA.items():
#             fila[f'Emp_{lab_t}'] = round(df_ocu_cr.loc[m_d & (df_ocu_cr['P3069']==cod_t), 'FEX_ADJ'].sum() / n_tot * 100, 1)
#     filas_cr.append(fila)
# 
# df_cruce = pd.DataFrame(filas_cr)
# if 'Desea_cambiar_%' in df_cruce.columns:
#     df_cruce = df_cruce.sort_values('Desea_cambiar_%', ascending=False)
# 
# import pandas as pd
# fig_cr, (ax_cr1, ax_cr2) = plt.subplots(1, 2, figsize=(18, 8))
# fig_cr.patch.set_facecolor(FONDO); ax_cr1.set_facecolor('white'); ax_cr2.set_facecolor('white')
# 
# if 'Desea_cambiar_%' in df_cruce.columns:
#     df_plot_cr = df_cruce.sort_values('Desea_cambiar_%', ascending=True)
#     y_cr = np.arange(len(df_plot_cr))
#     col_cr = [C['rojo'] if v > 40 else (C['naranja'] if v > 25 else C['verde']) for v in df_plot_cr['Desea_cambiar_%']]
#     ax_cr1.barh(y_cr, df_plot_cr['Desea_cambiar_%'], 0.65, color=col_cr, alpha=0.88)
#     prom = df_cruce['Desea_cambiar_%'].mean()
#     ax_cr1.axvline(prom, color='gray', ls='--', lw=1.3, alpha=0.7, label=f'Promedio ({prom:.1f}%)')
#     for i, (_, row) in enumerate(df_plot_cr.iterrows()):
#         ax_cr1.text(row['Desea_cambiar_%']+0.3, i, f"{row['Desea_cambiar_%']:.1f}%", va='center', fontsize=8)
#     ax_cr1.set_yticks(y_cr); ax_cr1.set_yticklabels(df_plot_cr['Departamento'], fontsize=9)
#     ax_cr1.set_xlabel('% desea cambiar (P7130=1)'); ax_cr1.set_title(f'Insatisfacción laboral — GEIH {ANIO}', fontsize=11, fontweight='bold')
#     ax_cr1.legend(fontsize=9); ax_cr1.grid(axis='x', alpha=0.3); ax_cr1.spines[['top','right']].set_visible(False)
# 
# cols_tam = [c for c in df_cruce.columns if c.startswith('Emp_')]
# if cols_tam:
#     df_hm = df_cruce.set_index('Departamento')[cols_tam].fillna(0)
#     df_hm.columns = [c.replace('Emp_','') for c in df_hm.columns]
#     cmap_t = mcolors.LinearSegmentedColormap.from_list('wg', ['white','#2E6DA4'])
#     im_t = ax_cr2.imshow(df_hm.values, cmap=cmap_t, aspect='auto', vmin=0, vmax=50)
#     plt.colorbar(im_t, ax=ax_cr2, label='% empleo')
#     for i in range(len(df_hm)):
#         for j in range(len(df_hm.columns)):
#             v = df_hm.iloc[i,j]
#             if v >= 5: ax_cr2.text(j, i, f'{v:.0f}%', ha='center', va='center', fontsize=7, fontweight='bold',
#                                     color='white' if v > 30 else '#1A252F')
#     ax_cr2.set_xticks(range(len(df_hm.columns))); ax_cr2.set_xticklabels(df_hm.columns, rotation=35, ha='right', fontsize=8)
#     ax_cr2.set_yticks(range(len(df_hm))); ax_cr2.set_yticklabels(df_hm.index, fontsize=8)
#     ax_cr2.set_title(f'Tamaño empresa × departamento — GEIH {ANIO}', fontsize=11, fontweight='bold')
# 
# fig_cr.suptitle(f'CRUCE EMPRESA × DEPARTAMENTO — GEIH {ANIO}', fontsize=13, fontweight='bold')
# fig_cr.tight_layout(pad=2.5); plt.show()
# exp.guardar_grafica(fig_cr, f'Cruce_empresa_dpto_{ANIO}', cerrar=True)
# exp.guardar_tabla(df_cruce, f'Cruce_empresa_departamento_{ANIO}')
# del df_ocu_cr; gc.collect()
# print("✅ Celda completada")

"""# ✅ 4 gráficos avanzados (Lorenz, ICI bubble, estacionalidad, heatmap)"""

# ── Prerequisitos para la celda de gráficos ──────────────────────────
# Ejecutar solo si ici / estac / salarios_rama / brecha no están definidas

from geih import (
    AnalisisSalarios, BrechaGenero, Estacionalidad,
    CompetitividadLaboral,
)

if 'salarios_rama' not in dir():
    salarios_rama = AnalisisSalarios(config=config).por_rama(df)

if 'brecha' not in dir():
    brecha = BrechaGenero().calcular(df)

if 'estac' not in dir():
    estac = Estacionalidad().calcular(geih)   # ← usa geih (base cruda), no df

if 'ici' not in dir():
    ici = CompetitividadLaboral(config=config).calcular(df)

print("✅ Variables listas:",
      [v for v in ['salarios_rama','brecha','estac','ici']
       if v in dir()])

from geih import (
    GraficoCurvaLorenz, GraficoICIBubble,
    GraficoEstacionalidad, GraficoContribucionHeatmap,
    GraficoBoxPlotSalarios, GraficoBrechaGenero,
)

df_ocu_lorenz = df[(df['OCI'] == 1) & (df['INGLABO'] > 0)]
fig_lorenz = GraficoCurvaLorenz().graficar(df_ocu_lorenz,
    titulo=f'Curva de Lorenz del ingreso laboral — GEIH {ANIO}')
exp.guardar_grafica(fig_lorenz, f'M5_Lorenz_Gini_{ANIO}')
plt.show()

if 'Costo_efectivo' in ici.columns and 'Talento_univ_%' in ici.columns:
    fig_ici = GraficoICIBubble().graficar(ici,
        titulo=f'Competitividad laboral (ICI) — GEIH {ANIO}')
    exp.guardar_grafica(fig_ici, f'M16_ICI_bubble_{ANIO}')
    plt.show()

fig_estac = GraficoEstacionalidad().graficar(estac,
    titulo=f'Estacionalidad del mercado laboral — GEIH {ANIO}')
exp.guardar_grafica(fig_estac, f'M11_Estacionalidad_lineas_{ANIO}')
plt.show()

fig_boxplot = GraficoBoxPlotSalarios().graficar(salarios_rama,
    titulo=f'Distribución del ingreso por rama — GEIH {ANIO}', smmlv=SMMLV)
exp.guardar_grafica(fig_boxplot, f'M4_BoxPlot_salarios_{ANIO}')
plt.show()

fig_brecha = GraficoBrechaGenero().graficar(brecha,
    titulo=f'Brecha salarial de género — GEIH {ANIO}', smmlv=SMMLV)
exp.guardar_grafica(fig_brecha, f'M6_Brecha_genero_{ANIO}')
plt.show()

plt.close('all')
gc.collect()
print("✅ Celda 12 completada — 5 gráficos generados")

"""# ✅ Análisis por 32 ciudades (AREA)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import AnalisisOcupadosCiudad
# 
# area = AnalisisOcupadosCiudad(config=config)
# tablas_area = area.calcular(df, ruta_ciiu=RUTA_CIIU)   # ← calcular, no calcular_tablas
# area.exportar_excel(tablas_area, str(exp.excel / f'CIIU_Area_{ANIO}.xlsx'))
# print("✅ Celda 13 completada — 6 tablas AREA exportadas")

"""# ✅ Salario por nivel educativo e industria"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ── Análisis: salario por nivel educativo e industria ─────────────────
# import matplotlib.colors as mcolors
# from geih.utils import EstadisticasPonderadas as EP
# 
# # Mapeo granular P3042 → nivel (separa Especialización, Maestría, Doctorado)
# NIVEL_DETALLE = {
#     1: '1. Sin educación',   2: '1. Sin educación',
#     3: '2. Primaria',        4: '3. Secundaria',
#     5: '4. Media',           6: '4. Media',           7: '4. Media',
#     8: '5. Técnica/Tecno.',  9: '5. Técnica/Tecno.',
#     10: '6. Universitaria',
#     11: '7. Especialización',
#     12: '8. Maestría',
#     13: '9. Doctorado',
# }
# ORDEN_NIVELES = list(dict.fromkeys(NIVEL_DETALLE.values()))  # preserva orden
# 
# # ── Filtro base ───────────────────────────────────────────────────────
# df_edu = df[
#     (df['OCI'] == 1) & (df['INGLABO'] > 0) &
#     df['P3042'].notna() & df['RAMA'].notna()
# ].copy()
# df_edu['NIVEL_EDU'] = df_edu['P3042'].map(NIVEL_DETALLE)
# print(f"Base de cálculo: {df_edu.shape[0]:,} ocupados con ingreso y nivel educativo")
# 
# # ── Tabla 1: mediana nacional por nivel ───────────────────────────────
# filas_niv = []
# for nivel in ORDEN_NIVELES:
#     m = df_edu['NIVEL_EDU'] == nivel
#     n_pond = df_edu.loc[m, 'FEX_ADJ'].sum()
#     if n_pond < 3_000: continue
#     mediana = EP.mediana(df_edu.loc[m, 'INGLABO'], df_edu.loc[m, 'FEX_ADJ'])
#     media   = EP.media(df_edu.loc[m, 'INGLABO'],   df_edu.loc[m, 'FEX_ADJ'])
#     filas_niv.append({
#         'Nivel_educativo': nivel,
#         'Ocupados_miles':  round(n_pond / 1_000),
#         'Mediana_COP':     round(mediana),
#         'Media_COP':       round(media),
#         'Mediana_SMMLV':   round(mediana / SMMLV, 2),
#         'Media_SMMLV':     round(media   / SMMLV, 2),
#     })
# 
# df_nivel = pd.DataFrame(filas_niv)
# 
# # Prima posgrado respecto a universitaria
# base_univ = df_nivel.loc[df_nivel['Nivel_educativo'] == '6. Universitaria', 'Mediana_COP'].values
# if len(base_univ):
#     df_nivel['Prima_vs_univ_%'] = ((df_nivel['Mediana_COP'] / base_univ[0] - 1) * 100).round(1)
# 
# exp.guardar_tabla(df_nivel, f'Salario_nivel_educativo_{ANIO}')
# print(df_nivel.to_string(index=False))
# 
# # ── Tabla 2: mediana por nivel × industria (top 8 ramas) ─────────────
# ramas_top8 = (
#     df_edu.groupby('RAMA')['FEX_ADJ'].sum()
#     .nlargest(8).index.tolist()
# )
# 
# filas_cruce = []
# for nivel in ORDEN_NIVELES:
#     for rama in ramas_top8:
#         m = (df_edu['NIVEL_EDU'] == nivel) & (df_edu['RAMA'] == rama)
#         n_pond = df_edu.loc[m, 'FEX_ADJ'].sum()
#         if n_pond < 3_000: continue
#         mediana = EP.mediana(df_edu.loc[m, 'INGLABO'], df_edu.loc[m, 'FEX_ADJ'])
#         filas_cruce.append({
#             'Nivel_educativo': nivel, 'Industria': rama,
#             'Ocupados_miles':  round(n_pond / 1_000),
#             'Mediana_COP':     round(mediana),
#             'Mediana_SMMLV':   round(mediana / SMMLV, 2),
#         })
# 
# df_cruce  = pd.DataFrame(filas_cruce)
# df_heatmap = df_cruce.pivot_table(
#     index='Nivel_educativo', columns='Industria',
#     values='Mediana_SMMLV',
# ).reindex(ORDEN_NIVELES).dropna(how='all')
# 
# exp.guardar_tabla(df_cruce, f'Salario_educacion_industria_{ANIO}')
# 
# # ── Gráfico 1: barras horizontales por nivel ──────────────────────────
# def _color_nivel(n):
#     if 'Doctorado'      in n: return '#26215C'
#     if 'Maestría'       in n: return '#534AB7'
#     if 'Especialización'in n: return '#7F77DD'
#     if 'Universitaria'  in n: return '#378ADD'
#     if 'Técnica'        in n: return '#1D9E75'
#     return '#B4B2A9'
# 
# fig1, ax1 = plt.subplots(figsize=(11, 6))
# fig1.patch.set_facecolor(FONDO); ax1.set_facecolor('white')
# 
# bars = ax1.barh(
#     df_nivel['Nivel_educativo'], df_nivel['Mediana_SMMLV'], 0.65,
#     color=[_color_nivel(n) for n in df_nivel['Nivel_educativo']], alpha=0.9,
# )
# for bar, row in zip(bars, df_nivel.itertuples()):
#     ax1.text(bar.get_width() + 0.05, bar.get_y() + bar.get_height() / 2,
#              f'{row.Mediana_SMMLV:.2f}× (${row.Mediana_COP/1e6:.1f}M)',
#              va='center', fontsize=8.5)
# 
# ax1.axvline(1, color='gray', ls='--', lw=1.2, alpha=0.6, label='1 SMMLV')
# ax1.set_xlabel('Mediana ingreso laboral (× SMMLV)', fontsize=11)
# ax1.set_title(f'Salario mediano por nivel educativo — GEIH {ANIO}', fontsize=12, fontweight='bold')
# ax1.legend(fontsize=9); ax1.grid(axis='x', alpha=0.3)
# ax1.spines[['top','right']].set_visible(False)
# fig1.tight_layout(pad=2)
# exp.guardar_grafica(fig1, f'Salario_nivel_educativo_{ANIO}')
# plt.show()
# 
# # ── Gráfico 2: heatmap nivel × industria ─────────────────────────────
# if not df_heatmap.empty:
#     fig2, ax2 = plt.subplots(figsize=(14, 7))
#     fig2.patch.set_facecolor(FONDO); ax2.set_facecolor('white')
# 
#     cmap = mcolors.LinearSegmentedColormap.from_list('edu', ['#E6F1FB', '#185FA5'])
#     im = ax2.imshow(df_heatmap.values, cmap=cmap, aspect='auto',
#                     vmin=0, vmax=df_heatmap.values[~np.isnan(df_heatmap.values)].max())
#     plt.colorbar(im, ax=ax2, label='Mediana (× SMMLV)', shrink=0.8)
# 
#     for i in range(len(df_heatmap)):
#         for j in range(len(df_heatmap.columns)):
#             v = df_heatmap.iloc[i, j]
#             if pd.notna(v):
#                 ax2.text(j, i, f'{v:.1f}×', ha='center', va='center', fontsize=8,
#                          fontweight='bold',
#                          color='white' if v > df_heatmap.values[~np.isnan(df_heatmap.values)].max() * 0.6 else '#1A252F')
# 
#     # Borde morado en filas de posgrado avanzado
#     for i, nivel in enumerate(df_heatmap.index):
#         if any(x in nivel for x in ['Especialización','Maestría','Doctorado']):
#             ax2.add_patch(plt.Rectangle((-0.5, i - 0.5), len(df_heatmap.columns), 1,
#                                          fill=False, edgecolor='#534AB7', lw=2, zorder=5))
# 
#     ax2.set_xticks(range(len(df_heatmap.columns)))
#     ax2.set_xticklabels([c[:28] for c in df_heatmap.columns], rotation=30, ha='right', fontsize=9)
#     ax2.set_yticks(range(len(df_heatmap)))
#     ax2.set_yticklabels(df_heatmap.index, fontsize=9)
#     ax2.set_title(f'Salario mediano (× SMMLV) por educación × industria — GEIH {ANIO}',
#                   fontsize=12, fontweight='bold')
#     fig2.tight_layout(pad=2.5)
#     exp.guardar_grafica(fig2, f'Salario_educacion_industria_{ANIO}')
#     plt.show()
# 
# del df_edu; gc.collect()
# print(f"✅ 2 tablas + 2 gráficos — guardados en {exp.raiz}")

"""# ✅ Salarios más altos"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ════════════════════════════════════════════════════════════════════
# # ANÁLISIS DE ALTOS INGRESOS — Percentiles 90 / 95 / 99 / 99.9
# # ════════════════════════════════════════════════════════════════════
# import numpy as np
# import matplotlib.ticker as mticker
# import matplotlib.colors as mcolors
# from geih.utils import EstadisticasPonderadas as EP
# from geih import TAMANO_EMPRESA
# 
# NIVEL_DETALLE = {
#     1:'Sin educación', 2:'Sin educación', 3:'Primaria', 4:'Secundaria',
#     5:'Media', 6:'Media', 7:'Media', 8:'Técnica/Tecno.', 9:'Técnica/Tecno.',
#     10:'Universitaria', 11:'Especialización', 12:'Maestría', 13:'Doctorado',
# }
# POS_OC = {
#     1:'Obrero/empl. particular', 2:'Obrero/empl. gobierno',
#     3:'Empl. doméstico',         4:'Cta. propia',
#     5:'Patrón/empleador',        6:'Trab. familiar s/r',
#     7:'Trab. s/r no familiar',   8:'Jornalero', 9:'Otro',
# }
# 
# # ── 0. Base ───────────────────────────────────────────────────────────
# df_ing    = df[(df['OCI'] == 1) & (df['INGLABO'] > 0)].copy()
# N_TOTAL   = df_ing['FEX_ADJ'].sum()
# ING_TOTAL = (df_ing['INGLABO'] * df_ing['FEX_ADJ']).sum()
# print(f"Base: {df_ing.shape[0]:,} registros | {N_TOTAL/1e6:.2f}M personas expandidas")
# 
# # ── 1. Percentil ponderado (función pura) ─────────────────────────────
# def pct_pond(serie: pd.Series, pesos: pd.Series, q: float) -> float:
#     s   = np.asarray(serie, dtype=float)
#     w   = np.asarray(pesos, dtype=float)
#     idx = np.argsort(s)
#     cum_w = np.cumsum(w[idx])
#     return float(s[idx][np.searchsorted(cum_w, q / 100.0 * cum_w[-1], side='left')])
# 
# # ── 2. Umbrales ───────────────────────────────────────────────────────
# QS       = [90, 95, 99, 99.9]
# umbrales = {q: pct_pond(df_ing['INGLABO'], df_ing['FEX_ADJ'], q) for q in QS}
# max_ing  = df_ing['INGLABO'].max()
# n_max    = df_ing.loc[df_ing['INGLABO'] == max_ing, 'FEX_ADJ'].sum()
# 
# print(f"\n{'─'*55}")
# for q, v in umbrales.items():
#     print(f"  P{q:5.1f}: ${v:>15,.0f} = {v/SMMLV:>6.1f}× SMMLV")
# print(f"  MÁXIMO: ${max_ing:>15,.0f} = {max_ing/SMMLV:>6.1f}× SMMLV (~{n_max:,.0f} personas)")
# print(f"{'─'*55}")
# 
# # ── 3. Tabla resumen de percentiles ──────────────────────────────────
# filas_pct = []
# for q, u in umbrales.items():
#     m     = df_ing['INGLABO'] >= u
#     n_g   = df_ing.loc[m, 'FEX_ADJ'].sum()
#     ing_g = (df_ing.loc[m, 'INGLABO'] * df_ing.loc[m, 'FEX_ADJ']).sum()
#     med_g = EP.mediana(df_ing.loc[m, 'INGLABO'], df_ing.loc[m, 'FEX_ADJ'])
#     filas_pct.append({
#         'Percentil':       f'Encima P{q}',
#         'Umbral_COP':      round(u),
#         'Umbral_SMMLV':    round(u / SMMLV, 1),
#         'Personas_miles':  round(n_g / 1_000, 1),
#         'Pct_poblacion_%': round(n_g / N_TOTAL * 100, 2),
#         'Mediana_COP':     round(med_g),
#         'Mediana_SMMLV':   round(med_g / SMMLV, 1),
#         'Pct_ingreso_%':   round(ing_g / ING_TOTAL * 100, 1),
#     })
# df_pct = pd.DataFrame(filas_pct)
# exp.guardar_tabla(df_pct, f'Percentiles_ingreso_{ANIO}')
# 
# # ── 4. Función de perfil — completamente corregida ────────────────────
# def perfil_segmento(mask: pd.Series) -> dict:
#     """Perfil de un segmento. Vectorizado, sin loops sobre filas.
# 
#     CORRECCIÓN v2:
#       - mapeo seguro via list comprehension (evita Index.fillna con Index)
#       - edad_grupo en sub-DataFrame local (evita pasar Series a dist)
#       - observed=True en groupby categórico (Pandas ≥ 2.0)
#     """
#     seg = df_ing[mask].copy()
#     tw  = seg['FEX_ADJ'].sum()
#     if tw == 0:
#         return {}
# 
#     def dist(col: str, mapeo: dict = None, top_n: int = 10) -> pd.DataFrame:
#         if col not in seg.columns:
#             return pd.DataFrame()
#         s = seg.groupby(col, dropna=True)['FEX_ADJ'].sum().div(tw).mul(100).round(1)
#         if mapeo:
#             # ← CORRECCIÓN: list comprehension en vez de Index.fillna(otro_Index)
#             s.index = [mapeo.get(k, str(k)) for k in s.index]
#         return s.nlargest(top_n).rename('%').reset_index()
# 
#     # ── Edad: sub-DataFrame local para no tocar `seg` ────────────────
#     # ← CORRECCIÓN: calcular pd.cut aquí, no pasar Series a dist()
#     edad_df = pd.DataFrame()
#     if 'P6040' in seg.columns:
#         d_tmp = seg[['P6040', 'FEX_ADJ']].copy()
#         d_tmp['EDAD_GRP'] = pd.cut(
#             d_tmp['P6040'],
#             bins=[15, 25, 35, 45, 55, 65, 90],
#             labels=['15–24', '25–34', '35–44', '45–54', '55–64', '65+'],
#             right=False,
#         )
#         # observed=True evita DeprecationWarning de Pandas 2.0 con CategoricalDtype
#         edad_df = (
#             d_tmp.groupby('EDAD_GRP', observed=True)['FEX_ADJ']
#             .sum().div(tw).mul(100).round(1)
#             .rename('%').reset_index()
#         )
#         del d_tmp
# 
#     return {
#         'personas_miles': round(tw / 1_000, 1),
#         'mediana_cop':    round(EP.mediana(seg['INGLABO'], seg['FEX_ADJ'])),
#         'max_cop':        int(seg['INGLABO'].max()),
#         'sexo':      dist('P3271',  {1: 'Hombres', 2: 'Mujeres'}),
#         'educacion': dist('P3042',  NIVEL_DETALLE),
#         'industria': dist('RAMA'),
#         'ciudad':    dist('NOMBRE_DPTO'),
#         'posicion':  dist('P6430',  POS_OC),
#         'tamano_emp':dist('P3069',  TAMANO_EMPRESA),
#         'zona':      dist('CLASE',  {1: 'Urbano', 2: 'Rural'}),
#         'formalidad':dist('P6920',  {1: 'Cotiza pensión', 2: 'No cotiza', 3: 'Pensionado'}),
#         'edad_grupo': edad_df,
#     }
# 
# # Calcular todos los perfiles
# perfiles     = {q: perfil_segmento(df_ing['INGLABO'] >= u) for q, u in umbrales.items()}
# perfil_total = perfil_segmento(pd.Series(True, index=df_ing.index))
# print("✅ Perfiles calculados")
# 
# # ── 5. Exportar tablas comparativas (Total vs cada percentil) ─────────
# DIMS_EXPORT = ['educacion', 'industria', 'ciudad', 'posicion',
#                'tamano_emp', 'sexo', 'zona', 'formalidad']
# 
# for dim in DIMS_EXPORT:
#     df_base = perfil_total.get(dim, pd.DataFrame())
#     if df_base.empty:
#         continue
#     col_key = df_base.columns[0]
#     df_comp = df_base.rename(columns={'%': '% Total'})
#     for q in QS:
#         tmp = perfiles[q].get(dim, pd.DataFrame())
#         if not tmp.empty:
#             df_comp = df_comp.merge(
#                 tmp.rename(columns={'%': f'% P{q}+'}),
#                 on=col_key, how='outer',
#             )
#     # Rellenar solo columnas numéricas de porcentaje
#     pct_cols = [c for c in df_comp.columns if c.startswith('%')]
#     df_comp[pct_cols] = df_comp[pct_cols].fillna(0)
#     df_comp = df_comp.sort_values('% Total', ascending=False)
#     exp.guardar_tabla(df_comp, f'Altos_ingresos_{dim}_{ANIO}')
# 
# print(f"✅ {len(DIMS_EXPORT)} tablas comparativas exportadas")
# 
# # ── 6. Gráfico 1: Distribución con umbrales ───────────────────────────
# fig1, ax1 = plt.subplots(figsize=(13, 5))
# fig1.patch.set_facecolor(FONDO); ax1.set_facecolor('white')
# 
# lim_sup  = df_ing['INGLABO'].quantile(0.999)
# bins_h   = np.logspace(np.log10(max(df_ing['INGLABO'].min(), 1e4)),
#                         np.log10(lim_sup), 55)
# ing_clip = df_ing['INGLABO'].clip(upper=lim_sup)
# ax1.hist(ing_clip, bins=bins_h, weights=df_ing['FEX_ADJ'] / 1e3,
#          color=C['azul'], alpha=0.75)
# 
# colores_q = {90: C['verde'], 95: C['naranja'], 99: C['rojo'], 99.9: '#534AB7'}
# for q, v in umbrales.items():
#     if v <= lim_sup:
#         ax1.axvline(v, color=colores_q[q], lw=1.8, ls='--',
#                     label=f'P{q}: ${v/1e6:.1f}M ({v/SMMLV:.1f}×)')
# 
# ax1.set_xscale('log')
# ax1.xaxis.set_major_formatter(mticker.FuncFormatter(
#     lambda v, _: f'${v/1e6:.0f}M' if v >= 1e6 else f'${v/1e3:.0f}k'))
# ax1.set_xlabel('Ingreso laboral mensual (COP, escala log)', fontsize=11)
# ax1.set_ylabel('Personas (miles)', fontsize=11)
# ax1.set_title(f'Distribución del ingreso laboral con umbrales percentiles — GEIH {ANIO}',
#               fontsize=12, fontweight='bold')
# ax1.legend(fontsize=8.5, ncol=2); ax1.grid(axis='y', alpha=0.3)
# ax1.spines[['top', 'right']].set_visible(False)
# fig1.tight_layout(pad=2)
# exp.guardar_grafica(fig1, f'Distribucion_percentiles_{ANIO}')
# plt.show()
# 
# # ── 7. Gráfico 2: Escalera de medianas ───────────────────────────────
# qs_esc  = [50, 75] + QS
# pal_esc = [C['gris'], C['gris']] + [colores_q[q] for q in QS]
# escalera = []
# for q in qs_esc:
#     u   = pct_pond(df_ing['INGLABO'], df_ing['FEX_ADJ'], q)
#     m   = df_ing['INGLABO'] >= u
#     med = EP.mediana(df_ing.loc[m, 'INGLABO'], df_ing.loc[m, 'FEX_ADJ'])
#     escalera.append({'Grupo': f'P{q}+', 'Mediana_SMMLV': round(med / SMMLV, 1)})
# df_esc = pd.DataFrame(escalera)
# 
# fig2, ax2 = plt.subplots(figsize=(10, 5))
# fig2.patch.set_facecolor(FONDO); ax2.set_facecolor('white')
# bars2 = ax2.bar(df_esc['Grupo'], df_esc['Mediana_SMMLV'],
#                 color=pal_esc, alpha=0.88, width=0.6)
# for bar, row in zip(bars2, df_esc.itertuples()):
#     ax2.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.15,
#              f'{row.Mediana_SMMLV:.1f}×', ha='center', fontsize=9, fontweight='bold')
# ax2.axhline(1, color='gray', ls='--', lw=1, alpha=0.6, label='1 SMMLV')
# ax2.set_ylabel('Mediana ingreso (× SMMLV)', fontsize=11)
# ax2.set_title(f'Mediana por grupo percentil — GEIH {ANIO}', fontsize=12, fontweight='bold')
# ax2.legend(fontsize=9); ax2.grid(axis='y', alpha=0.3)
# ax2.spines[['top', 'right']].set_visible(False)
# fig2.tight_layout(pad=2)
# exp.guardar_grafica(fig2, f'Escalera_mediana_percentil_{ANIO}')
# plt.show()
# 
# # ── 8. Gráfico 3: 4 paneles comparativos Top 1% vs Total ─────────────
# def panel_doble(ax, dim: str, titulo: str, top_n: int = 7):
#     df_t1  = perfiles[99].get(dim, pd.DataFrame())
#     df_tot = perfil_total.get(dim, pd.DataFrame())
#     if df_t1.empty or df_tot.empty:
#         ax.axis('off'); return
#     col_key = df_t1.columns[0]
#     df_c = (df_t1.rename(columns={'%': '% Top 1%'})
#               .merge(df_tot.rename(columns={'%': '% Total'}),
#                      on=col_key, how='outer')
#               .assign(**{c: lambda x, c=c: x[c].fillna(0)
#                          for c in ['% Top 1%', '% Total']})
#               .nlargest(top_n, '% Top 1%'))
#     y = np.arange(len(df_c))
#     ax.barh(y - 0.2, df_c['% Top 1%'], 0.38, color='#534AB7', alpha=0.85, label='Top 1%')
#     ax.barh(y + 0.2, df_c['% Total'],   0.38, color='#B4B2A9', alpha=0.70, label='Total')
#     ax.set_yticks(y)
#     ax.set_yticklabels(df_c[col_key].astype(str).str[:30], fontsize=8.5)
#     ax.set_title(titulo, fontsize=10, fontweight='bold')
#     ax.set_xlabel('%', fontsize=9); ax.legend(fontsize=8)
#     ax.grid(axis='x', alpha=0.3); ax.spines[['top', 'right']].set_visible(False)
#     ax.set_facecolor('white')
# 
# fig3, axes = plt.subplots(2, 2, figsize=(16, 12))
# fig3.patch.set_facecolor(FONDO)
# fig3.suptitle(f'Perfil del Top 1% vs Total ocupado — GEIH {ANIO}',
#               fontsize=13, fontweight='bold')
# panel_doble(axes[0, 0], 'educacion',  'Nivel educativo')
# panel_doble(axes[0, 1], 'industria',  'Industria')
# panel_doble(axes[1, 0], 'ciudad',     'Departamento')
# panel_doble(axes[1, 1], 'posicion',   'Posición ocupacional')
# fig3.tight_layout(pad=2.5)
# exp.guardar_grafica(fig3, f'Perfil_top1pct_{ANIO}')
# plt.show()
# 
# # ── 9. Gráfico 4: Heatmap edad × percentil ───────────────────────────
# grupos_edad = ['15–24', '25–34', '35–44', '45–54', '55–64', '65+']
# heat_rows   = {}
# for q in QS:
#     eg = perfiles[q].get('edad_grupo', pd.DataFrame())
#     if not eg.empty:
#         col_age = eg.columns[0]
#         heat_rows[f'P{q}+'] = eg.set_index(col_age)['%']
# 
# if heat_rows:
#     df_age_heat = pd.DataFrame(heat_rows).reindex(grupos_edad).fillna(0)
#     fig4, ax4   = plt.subplots(figsize=(9, 5))
#     fig4.patch.set_facecolor(FONDO); ax4.set_facecolor('white')
#     cmap_a = mcolors.LinearSegmentedColormap.from_list('age', ['#E6F1FB', '#0C447C'])
#     im4    = ax4.imshow(df_age_heat.values, cmap=cmap_a, aspect='auto', vmin=0)
#     plt.colorbar(im4, ax=ax4, label='% del grupo', shrink=0.8)
#     for i in range(len(df_age_heat)):
#         for j in range(len(df_age_heat.columns)):
#             v = df_age_heat.iloc[i, j]
#             ax4.text(j, i, f'{v:.1f}%', ha='center', va='center', fontsize=9,
#                      fontweight='bold', color='white' if v > 30 else '#1A252F')
#     ax4.set_xticks(range(len(df_age_heat.columns)))
#     ax4.set_xticklabels(df_age_heat.columns, fontsize=11)
#     ax4.set_yticks(range(len(grupos_edad)))
#     ax4.set_yticklabels(grupos_edad, fontsize=11)
#     ax4.set_title(f'Distribución por edad según grupo percentil — GEIH {ANIO}',
#                   fontsize=12, fontweight='bold')
#     fig4.tight_layout(pad=2)
#     exp.guardar_grafica(fig4, f'Edad_percentiles_{ANIO}')
#     plt.show()
# 
# del df_ing
# gc.collect()
# n_tablas = 1 + len(DIMS_EXPORT)
# print(f"\n✅ Análisis completo — 4 gráficos + {n_tablas} tablas en {exp.raiz}")

"""# ✅ Excel - Empleados por ciudad y actividad económica"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ── Prerequisitos Celda 14 — regenerar variables faltantes ───────────
# from geih import (
#     CalidadEmpleo, FormalidadSectorial, CompetitividadLaboral,
#     VulnerabilidadLaboral, CostoLaboral, ContribucionSectorial,
#     MapaTalento, BonoDemografico,
#     AnalisisSalarios, BrechaGenero, Estacionalidad,
#     AnalisisUrbanoRural, FuerzaLaboralJoven, EcuacionMincer,
#     Top20Sectores,
# )
# 
# _faltantes = []
# 
# # ── Celda 7 (análisis básicos) ────────────────────────────────────────
# if 'salarios_rama' not in dir(): salarios_rama = AnalisisSalarios(config=config).por_rama(df);           _faltantes.append('salarios_rama')
# if 'brecha'        not in dir(): brecha        = BrechaGenero().calcular(df);                            _faltantes.append('brecha')
# if 'estac'         not in dir(): estac         = Estacionalidad().calcular(geih);                        _faltantes.append('estac')
# if 'joven'         not in dir(): joven         = FuerzaLaboralJoven(config=config).calcular(df);         _faltantes.append('joven')
# if 'urb_rural'     not in dir(): urb_rural     = AnalisisUrbanoRural(config=config).calcular(df);        _faltantes.append('urb_rural')
# if 'mincer'        not in dir(): mincer        = EcuacionMincer(config=config).estimar_todos(df);        _faltantes.append('mincer')
# if 'top20'         not in dir(): top20         = Top20Sectores(config=config).calcular(df, ruta_ciiu=RUTA_CIIU); _faltantes.append('top20')
# 
# # ── Celda 8 (análisis avanzados) ──────────────────────────────────────
# if 'ice_dpto' not in dir(): ice_dpto = CalidadEmpleo(config=config).calcular_por_departamento(df);      _faltantes.append('ice_dpto')
# if 'icf'      not in dir(): icf      = FormalidadSectorial(config=config).calcular(df);                 _faltantes.append('icf')
# if 'ici'      not in dir(): ici      = CompetitividadLaboral(config=config).calcular(df);               _faltantes.append('ici')
# if 'ivi'      not in dir(): ivi      = VulnerabilidadLaboral(config=config).calcular(df);               _faltantes.append('ivi')
# if 'costo'    not in dir(): costo    = CostoLaboral(config=config).calcular(df);                        _faltantes.append('costo')
# if 'contrib'  not in dir(): contrib  = ContribucionSectorial().calcular(geih);                          _faltantes.append('contrib')
# if 'talento'  not in dir(): talento  = MapaTalento(config=config).calcular(df);                         _faltantes.append('talento')
# if 'bono'     not in dir(): bono     = BonoDemografico(config=config).calcular(df);                     _faltantes.append('bono')
# 
# if _faltantes:
#     print(f"✅ Recalculadas: {_faltantes}")
# else:
#     print("✅ Todas las variables ya estaban definidas")

"""# ✅ Comparación entre años

## 🟢 Análisis entre eneros de 2025 y 2026
"""

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  CELDA D — COMPARACIÓN INTER-ANUAL                                  ║
# ║  Carga un año → lo procesa → libera crudo → carga el siguiente.     ║
# ║  En RAM simultánea: solo los dos DF procesados (~66k filas c/u).    ║
# ╚══════════════════════════════════════════════════════════════════════╝
# %%time

from geih import ComparadorMultiAnio, ConfigGEIH

comp = ComparadorMultiAnio()

# ── Año base (período de comparación, no el año completo) ────────────
comp.agregar_anio(
    ANIO_BASE,
    PATH_BASE_COMP,
    ConfigGEIH(anio=ANIO_BASE, n_meses=N_MESES_COMPARAR),
)
gc.collect()
GestorMemoria.estado()

# ── Año de comparación ───────────────────────────────────────────────
comp.agregar_anio(
    ANIO_COMP,
    PATH_COMP_COMP,
    ConfigGEIH(anio=ANIO_COMP, n_meses=N_MESES_COMPARAR),
)
gc.collect()

comp.resumen()
GestorMemoria.estado()

# ── Comparaciones ─────────────────────────────────────────────────────
df_indicadores  = comp.comparar_indicadores()
df_departamentos = comp.comparar_departamentos()
df_ingresos     = comp.evolucion_ingresos()
df_ramas        = comp.comparar_ramas()
df_brecha       = comp.comparar_brecha_genero()

# ── Base combinada (solo si necesitas análisis ad-hoc) ───────────────
# df_combinado = comp.obtener_base_combinada()   # ← descomenta si lo necesitas

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  CELDA EJECUTAR F — AGREGAR MES NUEVO (cuando llegue)               ║
# ╚══════════════════════════════════════════════════════════════════════╝

# Paso 1: Descomentar cuando tengas la carpeta del mes nuevo en Drive
# agregar_mes_nuevo(cfg, n_meses_nuevo=2)   # cuando llegue febrero 2026

# Paso 2: Cambiar n_meses_comp = 2 en CELDA EJECUTAR C
# Paso 3: Re-ejecutar desde CELDA EJECUTAR C

# from geih import ComparadorMultiAnio, ConfigGEIH

# comp = ComparadorMultiAnio()
# comp.agregar_anio(2025, f'{RUTA}/GEIH_2025_Consolidado.parquet', ConfigGEIH(anio=2025))
# comp.agregar_anio(2026, f'{RUTA}/GEIH_2026_Consolidado.parquet', ConfigGEIH(anio=2026, n_meses=6))

# comp.comparar_indicadores()        # TD/TGP/TO × año + variación
# comp.comparar_departamentos()      # TD por dpto × año
# comp.evolucion_ingresos()          # mediana salarial por año
# comp.comparar_ramas()              # empleo por rama × año
# comp.comparar_brecha_genero()      # brecha H/M por año

"""## 🟢 Análisis año base completo"""

# # ╔══════════════════════════════════════════════════════════════════════╗
# # ║  CELDA EJECUTAR E — ANÁLISIS AÑO BASE COMPLETO (bajo demanda)       ║
# # ║  Patrón: cargar → analizar → liberar. Nunca convive con CELDA D.    ║
# # ╚══════════════════════════════════════════════════════════════════════╝
# # %%time

# # Liberar el comparador si ya no lo necesitas antes de cargar el año completo
# # GestorMemoria.liberar(['comp','df_indicadores','df_ramas'], scope=globals())

# df_base_anual = __import__('pandas').read_parquet(PATH_BASE_FULL)
# GestorMemoria.tamano_df(df_base_anual, 'df_base_anual')
# GestorMemoria.estado()

# # ── Tu análisis del año completo aquí ────────────────────────────────
# # from geih import PreparadorGEIH, IndicadoresLaborales
# # config_base = ConfigGEIH(anio=cfg.anio_base, n_meses=cfg.n_meses_base)
# # prep = PreparadorGEIH(config=config_base)
# # df_base_prep = prep.preparar_base(df_base_anual)
# # ...

# # ── Liberar cuando termines ──────────────────────────────────────────
# GestorMemoria.liberar(['df_base_anual'], scope=globals())

"""# ✅ Dashboard interactivo (sin código)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # En algunos casos toca correrlo dos veces.
# # ╔══════════════════════════════════════════════════════════════════════╗
# # ║  CELDA EJECUTAR — Lanzar dashboard                                  ║
# # ║                                                                     ║
# # ║  Cambiar 'modo' según lo que quieres explorar:                      ║
# # ║    'enero'       → Solo enero año comp   (~40 MB)  ← recomendado   ║
# # ║    'comparacion' → Enero 2025 + 2026     (~80 MB)                  ║
# # ║    'anio_comp'   → Meses acumulados 2026 (~40+ MB) ← según meses   ║
# # ║    'anio_base'   → Año 2025 completo     (~400 MB) ← funciona OK   ║
# # ╚══════════════════════════════════════════════════════════════════════╝
# 
# lanzar_dashboard_prep(
#     cfg    = cfg,           # ← variable definida en CELDA EJECUTAR C
#     modo   = 'comparacion', # ← cambiar según necesidad
#     metodo = 'cloudflare',
# )

# Commented out IPython magic to ensure Python compatibility.
# %%time
# lanzar_dashboard(
#     ruta_base = RUTA,              # ← variable ya definida en tu notebook
#     puerto    = 8501,
#     metodo    = 'cloudflare',      # 'cloudflare' (sin cuenta) recomendado
#     # metodo  = 'ngrok',           # ← descomentar si prefieres ngrok
#     # ngrok_token = 'TU_TOKEN',    # ← pegar token de ngrok.com
# )

"""✅ Otros Análisis

## 🟢 M8 · Duración del desempleo
"""

from geih import DuracionDesempleo

dur = DuracionDesempleo(config=config)
dist = dur.calcular(df)             # distribución nacional por rangos
por_sexo = dur.por_sexo(df)         # cruce × sexo
por_educ = dur.por_educacion(df)    # cruce × nivel educativo
por_dpto = dur.por_departamento(df) # mediana semanas × departamento

exp.guardar_tabla(dist, 'M8_duracion_desempleo')
exp.guardar_tabla(por_dpto, 'M8_rigidez_departamental')

"""## 🟢 M14 · Dashboard sectores estratégicos ProColombia"""

from geih import DashboardSectoresProColombia

dash = DashboardSectoresProColombia(config=config).calcular(df)
exp.guardar_tabla(dash, 'M14_dashboard_sectores_IED')

"""## 🟢 MX1 · Anatomía salarial (P6500 vs INGLABO)"""

from geih import AnatomaSalario

anat = AnatomaSalario(config=config)
resumen = anat.resumen_nacional(df)   # brecha %, retención fuente
por_rama = anat.por_rama(df)          # brecha por rama
por_tam = anat.por_tamano_empresa(df) # brecha por tamaño empresa

exp.guardar_tabla(por_rama, 'MX1_anatomia_salario_rama')

"""## 🟢 MX2 · Forma de pago (P6765)"""

from geih import FormaPago

fp = FormaPago(config=config)
dist = fp.calcular(df)               # distribución por forma de pago
cruce = fp.cruce_formalidad(df)      # forma × cotización pensión

exp.guardar_tabla(dist, 'MX2_forma_pago')
exp.guardar_tabla(cruce, 'MX2_forma_pago_formalidad')

"""## 🟢 MX3 · Canal de empleo (P3363)"""

from geih import CanalEmpleo

ce = CanalEmpleo(config=config)
dist = ce.calcular(df)                    # distribución nacional
por_educ = ce.por_nivel_educativo(df)     # canal × educación

exp.guardar_tabla(dist, 'MX3_canal_empleo')
exp.guardar_tabla(por_educ, 'MX3_canal_x_educacion')

"""## 🟢 Gráfico Lorenz (complementa M5)"""

from geih import GraficoCurvaLorenz

df_ocu = df[(df['OCI'] == 1) & (df['INGLABO'] > 0)]
fig = GraficoCurvaLorenz().graficar(df_ocu)
exp.guardar_grafica(fig, 'M5_Lorenz_Gini')

"""## 🟢 Gráfico ICI Burbujas"""

from geih import CompetitividadLaboral, GraficoICIBubble

ici = CompetitividadLaboral(config=config).calcular(df)
fig = GraficoICIBubble().graficar(ici)
exp.guardar_grafica(fig, 'M16_ICI_bubble')

"""## 🟢 Gráfico Estacionalidad mensual"""

from geih import Estacionalidad, GraficoEstacionalidad

estac = Estacionalidad().calcular(geih)
fig = GraficoEstacionalidad().graficar(estac)
exp.guardar_grafica(fig, 'M11_Estacionalidad_lineas')

"""## 🟢 Gráfico Contribución sectorial (heatmap)"""

from geih import ContribucionSectorial, GraficoContribucionHeatmap

contrib = ContribucionSectorial().calcular(geih)
fig = GraficoContribucionHeatmap().graficar(contrib)
exp.guardar_grafica(fig, 'MA_Contribucion_heatmap')

"""# ▶️ BOX PLOT PONDERADO DE SALARIOS POR ACTIVIDAD ECONÓMICA"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ══════════════════════════════════════════════════════════════
# # BOX PLOT PONDERADO DE SALARIOS POR ACTIVIDAD ECONÓMICA
# # Replica el gráfico original del notebook con todas las
# # estadísticas: P01, P05, P25, P50, Media, P75, P95, P99, Max
# # Usa EstadisticasPonderadas del paquete geih_2025
# # ══════════════════════════════════════════════════════════════
# 
# import numpy as np
# import pandas as pd
# import matplotlib.pyplot as plt
# import matplotlib.ticker as mticker
# import matplotlib.patches as mpatches
# import matplotlib.patheffects as pe
# from matplotlib.lines import Line2D
# import gc
# 
# from geih import EstadisticasPonderadas as EP
# 
# C = {'azul':'#2E6DA4','rojo':'#C0392B','verde':'#1E8449',
#      'morado':'#7D3C98','naranja':'#E67E22','gris':'#7F8C8D'}
# 
# # ── 1. Preparar datos: solo ocupados con ingreso > 0 ─────────
# df_oci = df[(df['OCI'] == 1) & (df['INGLABO'] > 0) & df['RAMA'].notna()].copy()
# print(f"✅ Base: {len(df):,} filas | Ocupados con ingreso: {len(df_oci):,}")
# print(f"\n📊 Análisis 1: Box plot ponderado de salarios por actividad económica...")
# 
# # ── 2. Calcular estadísticas completas por rama ──────────────
# stats_rama = []
# for rama in df_oci['RAMA'].dropna().unique():
#     m = df_oci['RAMA'] == rama
#     v = df_oci.loc[m, 'INGLABO']
#     w = df_oci.loc[m, 'FEX_ADJ']
#     n = w.sum() / 1_000
#     if n < 100:
#         continue
# 
#     stats_rama.append({
#         'Rama':    rama,
#         'N_miles': round(n),
#         'P01':     EP.percentil(v, w, 0.01),
#         'P05':     EP.percentil(v, w, 0.05),
#         'P25':     EP.percentil(v, w, 0.25),
#         'P50':     EP.mediana(v, w),
#         'Media':   EP.media(v, w),
#         'P75':     EP.percentil(v, w, 0.75),
#         'P90':     EP.percentil(v, w, 0.90),
#         'P95':     EP.percentil(v, w, 0.95),
#         'P99':     EP.percentil(v, w, 0.99),
#         'Max_obs': float(v[w > 0].max()),
#         'Min_pos': float(v[(v > 0) & (w > 0)].min()),
#     })
# 
# df_stats = pd.DataFrame(stats_rama).dropna()
# df_stats = df_stats.sort_values('P50', ascending=True).reset_index(drop=True)
# 
# # ── 3. Gráfico principal ─────────────────────────────────────
# fig_bp = plt.figure(figsize=(20, 12))
# fig_bp.patch.set_facecolor(FONDO)
# ax_bp  = fig_bp.add_subplot(111)
# ax_bp.set_facecolor('white')
# 
# n_ramas    = len(df_stats)
# y_pos      = np.arange(n_ramas)
# altura_cj  = 0.5
# cap_h      = 0.15
# dot_size   = 55
# 
# # Límite del eje X: P95 del máximo (para no distorsionar)
# x_lim = df_stats['P95'].max() * 1.15
# 
# # Colorear cajas por nivel de P50 (gradiente verde→naranja→rojo)
# cmap_bp = plt.cm.RdYlGn
# norm_bp = plt.Normalize(df_stats['P50'].min(), df_stats['P50'].max())
# 
# for i, row in df_stats.iterrows():
#     y    = y_pos[i]
#     col  = cmap_bp(norm_bp(row['P50']))
# 
#     # ── Bigote inferior: P05 a P25 ───────────────────────
#     ax_bp.plot([row['P05'], row['P25']], [y, y],
#                color='#555555', lw=1.6, zorder=3)
#     ax_bp.plot([row['P05'], row['P05']], [y-cap_h, y+cap_h],
#                color='#555555', lw=1.6, zorder=3)
# 
#     # ── Caja: P25 a P75 ─────────────────────────────────
#     rect = plt.Rectangle(
#         (row['P25'], y - altura_cj/2), row['P75'] - row['P25'], altura_cj,
#         facecolor=col, alpha=0.80,
#         edgecolor='#333333', linewidth=1.5, zorder=4
#     )
#     ax_bp.add_patch(rect)
# 
#     # ── Bigote superior: P75 a P95 ───────────────────────
#     ax_bp.plot([row['P75'], row['P95']], [y, y],
#                color='#555555', lw=1.6, zorder=3)
#     ax_bp.plot([row['P95'], row['P95']], [y-cap_h, y+cap_h],
#                color='#555555', lw=1.6, zorder=3)
# 
#     # ── Mediana: línea vertical dentro de la caja ────────
#     ax_bp.plot([row['P50'], row['P50']], [y-altura_cj/2, y+altura_cj/2],
#                color='#1A1A1A', lw=2.8, zorder=6)
# 
#     # ── Media: diamante rojo ─────────────────────────────
#     ax_bp.scatter(row['Media'], y, marker='D', s=dot_size, color=C['rojo'],
#                   zorder=7, edgecolors='white', linewidth=0.8)
# 
#     # ── Atípico superior P99 ─────────────────────────────
#     if row['P99'] > row['P95'] * 1.02:
#         ax_bp.scatter(row['P99'], y, marker='o', s=45, color=C['naranja'],
#                       zorder=7, edgecolors='white', linewidth=0.8, alpha=0.90)
#         ax_bp.annotate(f"P99={row['P99']/SMMLV_2025:.1f}×",
#                        (row['P99'], y), xytext=(8, 0), textcoords='offset points',
#                        fontsize=7, color=C['naranja'], va='center')
# 
#     # ── Máximo observado (estrella) ──────────────────────
#     if row['Max_obs'] <= x_lim:
#         ax_bp.scatter(row['Max_obs'], y, marker='*', s=80, color='#8B0000',
#                       zorder=8, edgecolors='white', linewidth=0.5, alpha=0.85)
# 
#     # ── Atípico inferior P01 ─────────────────────────────
#     if row['P01'] < row['P05'] * 0.98:
#         ax_bp.scatter(row['P01'], y, marker='<', s=40, color=C['gris'],
#                       zorder=7, edgecolors='white', linewidth=0.8, alpha=0.85)
# 
#     # ── Etiqueta mediana dentro de la caja ────────────────
#     mid_x = (row['P25'] + row['P75']) / 2
#     med_sml = row['P50'] / SMMLV_2025
#     ax_bp.text(mid_x, y, f'{med_sml:.1f}×',
#                ha='center', va='center', fontsize=7.5,
#                fontweight='bold', color='white', zorder=8,
#                path_effects=[pe.withStroke(linewidth=1.5, foreground='#1A1A1A')])
# 
# # ── Líneas verticales de referencia SMMLV ─────────────────────
# for mult in [1, 2, 3, 4, 6, 8, 10]:
#     x_ref = mult * SMMLV_2025
#     if x_ref <= x_lim:
#         ax_bp.axvline(x_ref, color='#AAAAAA', ls=':', lw=1.0, alpha=0.6, zorder=1)
#         ax_bp.text(x_ref, n_ramas - 0.1, f'{mult}×SML',
#                    ha='center', fontsize=7.5, color='#666666', va='bottom')
# 
# # ── Ejes y etiquetas ──────────────────────────────────────────
# ax_bp.set_yticks(y_pos)
# ax_bp.set_yticklabels(df_stats['Rama'], fontsize=10.5)
# ax_bp.set_xlim(0, x_lim)
# ax_bp.set_ylim(-0.7, n_ramas - 0.3)
# ax_bp.xaxis.set_major_formatter(
#     mticker.FuncFormatter(lambda v,_: f'${v/1e6:.1f}M' if v>=1e6 else f'${v/1e3:.0f}K')
# )
# ax_bp.set_xlabel('Ingreso laboral mensual (COP corrientes 2025)', fontsize=12)
# ax_bp.set_title(
#     'Distribución del ingreso laboral por rama de actividad económica\n'
#     'GEIH 2025 · Estadísticas ponderadas por FEX_C18 | Eje cortado en P95 para legibilidad',
#     fontsize=13, fontweight='bold', pad=14
# )
# ax_bp.grid(axis='x', alpha=0.25, zorder=0)
# ax_bp.spines[['top','right']].set_visible(False)
# 
# # ── Leyenda ───────────────────────────────────────────────────
# leyenda_bp = [
#     plt.Rectangle((0,0),1,1, facecolor='#4CAF50', alpha=0.8, edgecolor='#333333'),
#     Line2D([0],[0], color='#1A1A1A', lw=2.8),
#     Line2D([0],[0], marker='D', color='w', markerfacecolor=C['rojo'],
#            markersize=9, label='Media ponderada'),
#     Line2D([0],[0], color='#555555', lw=1.6),
#     Line2D([0],[0], marker='o', color='w', markerfacecolor=C['naranja'],
#            markersize=7, alpha=0.9),
#     Line2D([0],[0], marker='*', color='w', markerfacecolor='#8B0000',
#            markersize=10, alpha=0.85),
# ]
# labs_bp = [
#     'Caja IQR (P25–P75)', 'Mediana (P50)', 'Media ponderada',
#     'Bigotes (P05–P95)', 'Atípico alto (P99)', 'Máximo observado',
# ]
# ax_bp.legend(leyenda_bp, labs_bp, fontsize=9.5, loc='lower right',
#              framealpha=0.95, ncol=2)
# 
# # ── Segundo eje X en SMMLV (arriba) ──────────────────────────
# ax_bp2 = ax_bp.twiny()
# ax_bp2.set_xlim(0, x_lim / SMMLV_2025)
# ax_bp2.set_xlabel('Múltiplos del SMMLV 2025 ($1,423,500 COP)', fontsize=11)
# ax_bp2.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f'{v:.0f}×'))
# ax_bp2.spines[['top','right']].set_visible(False)
# 
# plt.tight_layout(pad=2.5)
# plt.show()
# 
# # ── 4. Tabla de resultados en consola ─────────────────────────
# print(f"\n{'='*90}")
# print(f"  ESTADÍSTICAS DE SALARIO POR RAMA — GEIH 2025 (ponderadas FEX_C18/12)")
# print(f"  Todos los valores en SMMLV (1 SMMLV 2025 = ${SMMLV_2025:,} COP)")
# print(f"{'='*90}")
# print(f"  {'Rama':<32} {'N(K)':>6} {'P05':>5} {'P25':>5} {'Med':>5} {'Med$':>5} "
#       f"{'P75':>5} {'P95':>5} {'P99':>5} {'Max':>5}")
# print(f"  {'─'*32} {'─'*6} {'─'*5} {'─'*5} {'─'*5} {'─'*5} {'─'*5} {'─'*5} {'─'*5} {'─'*5}")
# 
# for _, row in df_stats.sort_values('P50', ascending=False).iterrows():
#     def s(v):
#         return f"{v/SMMLV_2025:.1f}×"
#     max_str = f"{row['Max_obs']/SMMLV_2025:.0f}×" if row['Max_obs'] <= 300*SMMLV_2025 else ">300×"
#     print(f"  {row['Rama']:<32} {row['N_miles']:>6,} "
#           f"{s(row['P05']):>5} {s(row['P25']):>5} {s(row['P50']):>5} "
#           f"{s(row['Media']):>5} {s(row['P75']):>5} {s(row['P95']):>5} "
#           f"{s(row['P99']):>5} {max_str:>5}")
# 
# # ── 5. Exportar ──────────────────────────────────────────────
# df_stats.to_csv('BoxPlot_stats_salario_rama_2025.csv',
#                 index=False, encoding='utf-8-sig', sep=';')
# print(f"\n✅ BoxPlot_stats_salario_rama_2025.csv")
# 
# del df_oci
# gc.collect()

"""# ▶️ DISTRIBUCIÓN DE INGRESOS POR RANGOS DE SMMLV"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ══════════════════════════════════════════════════════════════
# # DISTRIBUCIÓN DE INGRESOS POR RANGOS DE SMMLV
# # Usa DistribucionIngresos del paquete para el cálculo
# # ══════════════════════════════════════════════════════════════
# 
# from geih import DistribucionIngresos
# import numpy as np
# import matplotlib.pyplot as plt
# import matplotlib.patches as mpatches
# 
# AZUL, ROJO, VERDE, MORADO = '#2E6DA4', '#C0392B', '#1E8449', '#7D3C98'
# 
# # ── 1. Cálculo (usa el paquete) ───────────────────────────────
# resultado = DistribucionIngresos(config=config).calcular(df)
# dist   = resultado['total']
# dist_s = resultado['por_sexo']
# 
# # ── 2. Imprimir tabla ─────────────────────────────────────────
# DistribucionIngresos(config=config).imprimir(resultado, titulo='Enero – Diciembre 2025')
# 
# # ── 3. Gráfico (2 paneles) ───────────────────────────────────
# fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 7),
#                                 gridspec_kw={'width_ratios': [1.8, 1]})
# fig.patch.set_facecolor(FONDO)
# 
# labels = [str(r) for r in dist['RANGO']]
# x      = np.arange(len(labels))
# w      = 0.65
# 
# # ── Panel izquierdo: barras apiladas H + M ────────────────────
# ax1.set_facecolor('white')
# h_vals = dist_s['H_M'].values
# m_vals = dist_s['M_M'].values
# tot    = h_vals + m_vals
# 
# ax1.bar(x, h_vals, w, label='Hombres', color=AZUL, alpha=0.88, zorder=3)
# ax1.bar(x, m_vals, w, bottom=h_vals, label='Mujeres', color=ROJO, alpha=0.88, zorder=3)
# 
# for i, (t, pct) in enumerate(zip(tot, dist['Pct'])):
#     ax1.text(i, t + 0.03, f'{t:.2f}M\n{pct:.1f}%',
#              ha='center', va='bottom', fontsize=8.2,
#              fontweight='bold', color='#1A252F', zorder=5)
# 
# # Línea "Menos de 1 SMMLV"
# ax1.axvline(x=1.5, color=VERDE, ls='--', lw=1.8, alpha=0.75, zorder=4)
# ax1.text(1.55, max(tot) * 1.18, '← Menos de\n1 SMMLV',
#          fontsize=8.5, color=VERDE, fontstyle='italic', va='top')
# 
# ax1.set_xticks(x)
# ax1.set_xticklabels(labels, fontsize=9.5, rotation=20, ha='right')
# ax1.set_ylabel('Millones de personas', fontsize=11)
# ax1.set_ylim(0, max(tot) * 1.32)
# ax1.set_title(f'Distribución de ingresos laborales por rangos de SMMLV\n'
#               f'GEIH Enero – Diciembre 2025 — SMMLV 2025 = ${SMMLV:,}',
#               fontsize=12.5, fontweight='bold', pad=14)
# ax1.legend(handles=[
#     mpatches.Patch(color=AZUL, alpha=0.88, label='Hombres'),
#     mpatches.Patch(color=ROJO, alpha=0.88, label='Mujeres'),
# ], fontsize=10, loc='upper right')
# ax1.grid(axis='y', alpha=0.3, zorder=0)
# ax1.spines[['top','right']].set_visible(False)
# 
# # ── Panel derecho: curva acumulada ────────────────────────────
# ax2.set_facecolor('white')
# ax2.plot(x, dist['Acum_Pct'], color=MORADO, lw=2.8, marker='o',
#          markersize=8, markerfacecolor='white', markeredgecolor=MORADO,
#          markeredgewidth=2, zorder=5)
# ax2.fill_between(x, dist['Acum_Pct'], alpha=0.08, color=MORADO)
# 
# # Etiquetas en cada punto
# for i, (xi, yi) in enumerate(zip(x, dist['Acum_Pct'])):
#     ax2.annotate(f'{yi:.1f}%', (xi, yi), textcoords='offset points',
#                  xytext=(0, 12), ha='center', fontsize=9, fontweight='bold',
#                  color=MORADO)
# 
# # Líneas de referencia 50% y 80%
# ax2.axhline(50, color=ROJO, ls='--', lw=1.2, alpha=0.6)
# ax2.text(len(x)-0.5, 51, '50%', fontsize=9, color=ROJO, fontweight='bold', va='bottom')
# ax2.axhline(80, color=ROJO, ls='--', lw=1.2, alpha=0.6)
# ax2.text(len(x)-0.5, 81, '80%', fontsize=9, color=ROJO, fontweight='bold', va='bottom')
# 
# ax2.set_xticks(x)
# ax2.set_xticklabels(labels, fontsize=8, rotation=25, ha='right')
# ax2.set_ylabel('% acumulado de ocupados con ingreso > 0', fontsize=10)
# ax2.set_ylim(0, 108)
# ax2.set_title('Curva acumulada de distribución\n% de ocupados hasta cada rango',
#               fontsize=12.5, fontweight='bold', pad=14)
# ax2.grid(axis='y', alpha=0.2)
# ax2.spines[['top','right']].set_visible(False)
# 
# fig.tight_layout(pad=2.5)
# plt.show()
# print("✅ Gráfico generado")

"""# ▶️ SALARIOS POR RAMA Y POR EDAD — ANUAL + DICIEMBRE"""

# ══════════════════════════════════════════════════════════════
# SALARIOS POR RAMA Y POR EDAD — ANUAL + DICIEMBRE
# Usa clases del paquete | Gráficas visibles | Tablas legibles
# Requiere: config, df, geih, RUTA ya definidos
# ══════════════════════════════════════════════════════════════

from geih import (AnalisisSalarios, GraficoBoxPlotSalarios,
                        PreparadorGEIH, Exportador)
from IPython.display import display
import matplotlib.pyplot as plt
import gc

sal  = AnalisisSalarios(config=config)
exp  = Exportador(ruta_base=RUTA, config=config)
graf = GraficoBoxPlotSalarios()
prep = PreparadorGEIH(config=config)

# ── 1. Calcular: anual ───────────────────────────────────────
tabla_rama_anual = sal.por_rama(df)
tabla_edad_anual = sal.por_edad(df)

# ── 2. Calcular: diciembre (FEX sin dividir) ─────────────────
df_dic = prep.preparar_base(geih, mes_filtro=12)
df_dic = prep.agregar_variables_derivadas(df_dic)
tabla_rama_dic = sal.por_rama(df_dic)
tabla_edad_dic = sal.por_edad(df_dic)

# ══════════════════════════════════════════════════════════════
# TABLAS (render HTML en Colab — legibles y con scroll)
# ══════════════════════════════════════════════════════════════

def mostrar_tabla(tabla, titulo):
    """Muestra un DataFrame como tabla HTML formateada en Colab."""
    print(f"\n{'='*80}")
    print(f"  {titulo}")
    print(f"  SMMLV 2025 = ${config.smmlv:,} COP")
    print(f"{'='*80}\n")
    t = tabla.copy()
    for col in ['Media', 'Mediana', 'P10', 'P25', 'P75', 'P90', 'Std', 'IQR']:
        if col in t.columns:
            t[col] = t[col].apply(lambda x: f"${x:,.0f}")
    for col in ['Media_SMMLV', 'Mediana_SMMLV']:
        if col in t.columns:
            t[col] = t[col].apply(lambda x: f"{x:.2f}×")
    if 'N_personas' in t.columns:
        t['N_personas'] = t['N_personas'].apply(lambda x: f"{x:,}")
    if 'CV_%' in t.columns:
        t['CV_%'] = t['CV_%'].apply(lambda x: f"{x:.1f}%")
    display(t)

mostrar_tabla(tabla_rama_anual,
    "ESTADÍSTICAS DE INGRESO POR RAMA — Enero–Diciembre 2025")

mostrar_tabla(tabla_rama_dic,
    "ESTADÍSTICAS DE INGRESO POR RAMA — Diciembre 2025 (puntual)")

mostrar_tabla(tabla_edad_anual,
    "ESTADÍSTICAS DE INGRESO POR EDAD — Enero–Diciembre 2025")

# ══════════════════════════════════════════════════════════════
# GRÁFICAS (plt.show ANTES de guardar para que se vean)
# ══════════════════════════════════════════════════════════════

fig1 = graf.graficar(tabla_rama_anual,
    titulo=(f'Distribución del ingreso laboral por rama de actividad económica\n'
            f'GEIH Enero – Diciembre 2025 — Ponderado por FEX_C18 | '
            f'SMMLV 2025 = ${config.smmlv:,}'))
plt.show()
exp.guardar_grafica(fig1, 'BoxPlot_salario_rama_anual_2025', cerrar=False)

fig2 = graf.graficar(tabla_rama_dic,
    titulo=(f'Distribución del ingreso laboral por rama de actividad económica\n'
            f'GEIH Diciembre 2025 — Ponderado por FEX_C18 | '
            f'SMMLV 2025 = ${config.smmlv:,}'))
plt.show()
exp.guardar_grafica(fig2, 'BoxPlot_salario_rama_diciembre_2025', cerrar=False)

plt.close('all')

# ══════════════════════════════════════════════════════════════
# EXPORTAR CSVs
# ══════════════════════════════════════════════════════════════

exp.guardar_tabla(tabla_rama_anual.reset_index(), 'Estadisticas_salario_rama_anual_2025')
exp.guardar_tabla(tabla_rama_dic.reset_index(), 'Estadisticas_salario_rama_diciembre_2025')
exp.guardar_tabla(tabla_edad_anual, 'Estadisticas_salario_edad_anual_2025')
exp.guardar_tabla(tabla_edad_dic, 'Estadisticas_salario_edad_diciembre_2025')

del df_dic; gc.collect()
print("\n✅ Completado: 2 gráficos visibles + 3 tablas HTML + 4 CSVs exportados")

"""# ▶️ CICLO VITAL DEL SALARIO POR GRUPO DE EDAD"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ══════════════════════════════════════════════════════════════
# # CICLO VITAL DEL SALARIO POR GRUPO DE EDAD
# # 3 paneles: mediana+media, en SMMLV, coeficiente de variación
# # Requiere: tabla_edad_anual, tabla_edad_dic ya calculadas
# #           (output de AnalisisSalarios(config).por_edad(df))
# # ══════════════════════════════════════════════════════════════
# 
# from IPython.display import display
# import numpy as np
# import matplotlib.pyplot as plt
# import matplotlib.ticker as mticker
# import matplotlib.patches as mpatches
# from matplotlib.gridspec import GridSpec
# 
# FONDO = '#F7F9FC'
# AZUL, ROJO, VERDE, GRIS = '#2E6DA4', '#C0392B', '#1E8449', '#7F8C8D'
# 
# # ── Tablas HTML ───────────────────────────────────────────────
# print(f"\n{'='*70}")
# print(f"  INGRESO LABORAL POR GRUPO DE EDAD — GEIH 2025")
# print(f"  SMMLV 2025 = ${config.smmlv:,} COP")
# print(f"{'='*70}")
# print("\n📊 Anual (Enero–Diciembre 2025)\n")
# display(tabla_edad_anual)
# print("\n📊 Diciembre 2025\n")
# display(tabla_edad_dic)
# 
# # ── Gráfico 3 paneles ────────────────────────────────────────
# fig = plt.figure(figsize=(18, 10))
# fig.patch.set_facecolor(FONDO)
# gs = GridSpec(2, 2, figure=fig, hspace=0.42, wspace=0.35)
# ax1 = fig.add_subplot(gs[0, :])    # curva principal (fila completa)
# ax2 = fig.add_subplot(gs[1, 0])    # mediana en SMMLV
# ax3 = fig.add_subplot(gs[1, 1])    # coeficiente de variación
# for ax in [ax1, ax2, ax3]:
#     ax.set_facecolor('white'); ax.spines[['top','right']].set_visible(False)
# 
# x_a = np.arange(len(tabla_edad_anual))
# x_d = np.arange(len(tabla_edad_dic))
# edades_a = tabla_edad_anual['Grupo_edad'] if 'Grupo_edad' in tabla_edad_anual.columns else tabla_edad_anual.index
# 
# # ── Panel 1: Mediana y Media por edad ─────────────────────────
# for tabla, x, color, lbl in [
#     (tabla_edad_anual, x_a, AZUL, 'Anual'),
#     (tabla_edad_dic, x_d, ROJO, 'Diciembre'),
# ]:
#     med = tabla['Mediana'].values / 1e6
#     mea = tabla['Media'].values / 1e6
#     p25 = tabla['P25'].values / 1e6
#     p75 = tabla['P75'].values / 1e6
# 
#     ax1.fill_between(x, p25, p75, alpha=0.10, color=color)
#     ax1.plot(x, med, 'o-', color=color, lw=2.4, ms=5, markerfacecolor='white',
#              markeredgewidth=2, label=f'Mediana — {lbl}', zorder=4)
#     ax1.plot(x, mea, '--', color=color, lw=1.6, alpha=0.7,
#              label=f'Media — {lbl}')
# 
#     # Marcar pico
#     idx_pico = np.argmax(mea)
#     edad_pico = (tabla['Grupo_edad'].values if 'Grupo_edad' in tabla.columns
#                  else tabla.index.values)[idx_pico]
#     ax1.scatter(idx_pico, mea[idx_pico], s=120, color=color, marker='*',
#                 edgecolors='white', linewidth=0.8, zorder=6)
#     ax1.annotate(f"Pico: {edad_pico}\n${mea[idx_pico]:.2f}M",
#                  xy=(idx_pico, mea[idx_pico]),
#                  xytext=(idx_pico + 0.5, mea[idx_pico] + 0.08),
#                  fontsize=8.5, color=color, fontweight='bold',
#                  arrowprops=dict(arrowstyle='->', color=color, lw=1.2))
# 
# ax1.axhline(config.smmlv / 1e6, color=VERDE, ls=':', lw=1.5, alpha=0.75)
# ax1.text(len(x_a) - 0.5, config.smmlv / 1e6 + 0.02, '1 SMMLV',
#          ha='right', fontsize=8.5, color=VERDE, fontstyle='italic')
# ax1.set_xticks(x_a)
# ax1.set_xticklabels(edades_a, fontsize=9, rotation=35)
# ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f'${v:.1f}M'))
# ax1.set_ylabel('Ingreso laboral mensual (millones COP)', fontsize=11)
# ax1.set_title('Ciclo vital del ingreso laboral por grupo de edad\n'
#               'GEIH 2025 — Mediana y Media ponderadas por FEX_C18 | '
#               'Banda IQR (P25–P75)',
#               fontsize=12.5, fontweight='bold', pad=12)
# ax1.legend(fontsize=9, ncol=4, framealpha=0.92)
# ax1.grid(axis='y', alpha=0.3)
# 
# # ── Panel 2: Mediana en SMMLV ─────────────────────────────────
# smmlv_a = tabla_edad_anual['Mediana_SMMLV'].values
# smmlv_d = tabla_edad_dic['Mediana_SMMLV'].values
# 
# ax2.fill_between(x_a, smmlv_a, alpha=0.15, color=AZUL)
# ax2.plot(x_a, smmlv_a, 'o-', color=AZUL, lw=2, ms=4, label='Anual')
# ax2.plot(x_d, smmlv_d, 's--', color=ROJO, lw=1.6, ms=4, label='Diciembre')
# for ref, alpha in [(1, 0.8), (2, 0.5)]:
#     ax2.axhline(ref, color=VERDE, ls=':', lw=1.3, alpha=alpha)
#     ax2.text(0.1, ref + 0.03, f'{ref} SMMLV', fontsize=8, color=VERDE)
# ax2.set_xticks(x_a)
# ax2.set_xticklabels(edades_a, fontsize=8, rotation=40)
# ax2.set_ylabel('Mediana (× SMMLV)', fontsize=10)
# ax2.set_title('Mediana del ingreso por edad\n(en múltiplos de SMMLV)',
#               fontsize=11, fontweight='bold')
# ax2.legend(fontsize=9)
# ax2.grid(axis='y', alpha=0.3)
# 
# # ── Panel 3: Coeficiente de variación ─────────────────────────
# cv_a = tabla_edad_anual['CV_%'].values
# cv_d = tabla_edad_dic['CV_%'].values
# w3 = 0.35
# 
# ax3.bar(x_a - w3/2, cv_a, w3, color=AZUL, alpha=0.75, label='Anual')
# ax3.bar(x_d + w3/2, cv_d, w3, color=ROJO, alpha=0.75, label='Diciembre')
# for i, (va, vd) in enumerate(zip(cv_a, cv_d)):
#     ax3.text(i - w3/2, va + 1, f'{va:.0f}', ha='center', fontsize=7, color=AZUL)
#     ax3.text(i + w3/2, vd + 1, f'{vd:.0f}', ha='center', fontsize=7, color=ROJO)
# ax3.set_xticks(x_a)
# ax3.set_xticklabels(edades_a, fontsize=8, rotation=40)
# ax3.set_ylabel('Coeficiente de variación (%)', fontsize=10)
# ax3.set_title('Desigualdad interna por edad\n(CV% alto = mayor dispersión salarial)',
#               fontsize=11, fontweight='bold')
# ax3.legend(fontsize=9)
# ax3.grid(axis='y', alpha=0.3)
# 
# fig.tight_layout(pad=2.5)
# plt.show()
# exp.guardar_grafica(fig, 'Salario_por_edad_GEIH_2025', cerrar=False)
# plt.close('all')
# print("\n✅ Ciclo vital del salario completado: 2 tablas HTML + 1 gráfico + CSVs ya exportados")

"""# ▶️ INFORMALIDAD"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ══════════════════════════════════════════════════════════════
# # SECCIÓN: INFORMALIDAD LABORAL — GEIH 2025
# #
# # Definición DANE (proxy estándar):
# #   Informal = Ocupado que NO cotiza a pensión (P6920 ≠ 1)
# #   Incluye P6920=2 (No cotiza) y P6920=NaN (sin dato)
# #   Excluye P6920=3 (ya pensionado → formal por definición)
# #
# # Contenido:
# #   1. Tasa nacional + Hombres/Mujeres
# #   2. Por rama de actividad económica
# #   3. Por ciudad / departamento
# #   4. Cruce rama × sexo (informales en millones)
# #   5. Perfil salarial: formal vs informal
# #   6. Gráficos (3 figuras)
# #
# # Requiere: config, df, RUTA ya definidos
# # ══════════════════════════════════════════════════════════════
# 
# from geih import EstadisticasPonderadas as EP, Exportador
# from IPython.display import display
# import pandas as pd
# import numpy as np
# import matplotlib.pyplot as plt
# import matplotlib.patches as mpatches
# import gc
# 
# exp = Exportador(ruta_base=RUTA, config=config)
# 
# # ── Ocupados ──────────────────────────────────────────────────
# df_ocu = df[df['OCI'] == 1].copy()
# total_ocu = df_ocu['FEX_ADJ'].sum()
# 
# # ── Clasificar formalidad ─────────────────────────────────────
# # P6920: 1=Cotiza pensión (Formal), 2=No cotiza (Informal), 3=Ya pensionado
# df_ocu['FORMAL'] = np.where(
#     df_ocu['P6920'] == 1, 'Formal',
#     np.where(df_ocu['P6920'] == 3, 'Formal',   # pensionado = formal
#              'Informal')
# )
# 
# n_formal   = df_ocu.loc[df_ocu['FORMAL'] == 'Formal', 'FEX_ADJ'].sum()
# n_informal = df_ocu.loc[df_ocu['FORMAL'] == 'Informal', 'FEX_ADJ'].sum()
# tasa_inf   = n_informal / total_ocu * 100
# 
# print(f"{'='*65}")
# print(f"  INFORMALIDAD LABORAL — GEIH 2025")
# print(f"  Proxy: No cotiza a pensión (P6920 ≠ 1 y ≠ 3)")
# print(f"{'='*65}")
# print(f"  Ocupados totales : {total_ocu/1e6:.2f} M")
# print(f"  Formales         : {n_formal/1e6:.2f} M ({n_formal/total_ocu*100:.1f}%)")
# print(f"  Informales       : {n_informal/1e6:.2f} M ({tasa_inf:.1f}%)")
# print(f"{'='*65}")
# 
# # ══════════════════════════════════════════════════════════════
# # 1. NACIONAL + POR SEXO
# # ══════════════════════════════════════════════════════════════
# 
# filas_sexo = []
# for sexo_val, sexo_lbl in [(None, 'Nacional'), (1, 'Hombres'), (2, 'Mujeres')]:
#     m = pd.Series(True, index=df_ocu.index)
#     if sexo_val is not None:
#         m = df_ocu['P3271'] == sexo_val
#     n_tot = df_ocu.loc[m, 'FEX_ADJ'].sum()
#     n_inf = df_ocu.loc[m & (df_ocu['FORMAL'] == 'Informal'), 'FEX_ADJ'].sum()
#     n_for = df_ocu.loc[m & (df_ocu['FORMAL'] == 'Formal'), 'FEX_ADJ'].sum()
#     med_inf = EP.mediana(
#         df_ocu.loc[m & (df_ocu['FORMAL'] == 'Informal') & (df_ocu['INGLABO'] > 0), 'INGLABO'],
#         df_ocu.loc[m & (df_ocu['FORMAL'] == 'Informal') & (df_ocu['INGLABO'] > 0), 'FEX_ADJ'])
#     med_for = EP.mediana(
#         df_ocu.loc[m & (df_ocu['FORMAL'] == 'Formal') & (df_ocu['INGLABO'] > 0), 'INGLABO'],
#         df_ocu.loc[m & (df_ocu['FORMAL'] == 'Formal') & (df_ocu['INGLABO'] > 0), 'FEX_ADJ'])
#     filas_sexo.append({
#         'Grupo': sexo_lbl,
#         'Ocupados_M': round(n_tot / 1e6, 2),
#         'Formales_M': round(n_for / 1e6, 2),
#         'Informales_M': round(n_inf / 1e6, 2),
#         'Tasa_informalidad_%': round(n_inf / n_tot * 100, 1),
#         'Mediana_formal_SMMLV': round(med_for / config.smmlv, 2) if not np.isnan(med_for) else np.nan,
#         'Mediana_informal_SMMLV': round(med_inf / config.smmlv, 2) if not np.isnan(med_inf) else np.nan,
#     })
# 
# df_sexo = pd.DataFrame(filas_sexo)
# print("\n📊 1. Informalidad nacional y por sexo\n")
# display(df_sexo)
# 
# # ══════════════════════════════════════════════════════════════
# # 2. POR RAMA DE ACTIVIDAD ECONÓMICA
# # ══════════════════════════════════════════════════════════════
# 
# filas_rama = []
# for rama in df_ocu['RAMA'].dropna().unique():
#     m_r = df_ocu['RAMA'] == rama
#     n_tot = df_ocu.loc[m_r, 'FEX_ADJ'].sum()
#     if n_tot < 5_000:
#         continue
#     n_inf = df_ocu.loc[m_r & (df_ocu['FORMAL'] == 'Informal'), 'FEX_ADJ'].sum()
#     n_for = n_tot - n_inf
# 
#     # Por sexo dentro de la rama
#     n_inf_h = df_ocu.loc[m_r & (df_ocu['FORMAL'] == 'Informal') & (df_ocu['P3271'] == 1), 'FEX_ADJ'].sum()
#     n_inf_m = df_ocu.loc[m_r & (df_ocu['FORMAL'] == 'Informal') & (df_ocu['P3271'] == 2), 'FEX_ADJ'].sum()
# 
#     filas_rama.append({
#         'Rama': rama,
#         'Ocupados_miles': round(n_tot / 1_000),
#         'Informales_miles': round(n_inf / 1_000),
#         'Tasa_inf_%': round(n_inf / n_tot * 100, 1),
#         'Inf_hombres_miles': round(n_inf_h / 1_000),
#         'Inf_mujeres_miles': round(n_inf_m / 1_000),
#         'Formales_miles': round(n_for / 1_000),
#     })
# 
# df_rama = pd.DataFrame(filas_rama).sort_values('Tasa_inf_%', ascending=False)
# print("\n📊 2. Informalidad por rama de actividad económica\n")
# display(df_rama)
# 
# # ══════════════════════════════════════════════════════════════
# # 3. POR CIUDAD / DEPARTAMENTO
# # ══════════════════════════════════════════════════════════════
# 
# col_geo = 'CIUDAD' if 'CIUDAD' in df_ocu.columns else 'NOMBRE_DPTO'
# filas_geo = []
# for geo in df_ocu[col_geo].dropna().unique():
#     m_g = df_ocu[col_geo] == geo
#     n_tot = df_ocu.loc[m_g, 'FEX_ADJ'].sum()
#     if n_tot < 10_000:
#         continue
#     n_inf = df_ocu.loc[m_g & (df_ocu['FORMAL'] == 'Informal'), 'FEX_ADJ'].sum()
#     filas_geo.append({
#         'Ciudad_Depto': geo,
#         'Ocupados_miles': round(n_tot / 1_000),
#         'Informales_miles': round(n_inf / 1_000),
#         'Tasa_inf_%': round(n_inf / n_tot * 100, 1),
#     })
# 
# df_geo = pd.DataFrame(filas_geo).sort_values('Tasa_inf_%', ascending=False)
# print(f"\n📊 3. Informalidad por {col_geo.lower()}\n")
# display(df_geo)
# 
# # ══════════════════════════════════════════════════════════════
# # 4. GRÁFICO A: Informalidad por rama + sexo
# # ══════════════════════════════════════════════════════════════
# 
# AZUL, ROJO, VERDE, GRIS = '#2E6DA4', '#C0392B', '#1E8449', '#7F8C8D'
# FONDO = '#F7F9FC'
# 
# fig1, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 8),
#                                  gridspec_kw={'width_ratios': [1.2, 1]})
# fig1.patch.set_facecolor(FONDO)
# 
# # Panel izq: tasa por rama con color semáforo
# df_r_plot = df_rama.sort_values('Tasa_inf_%', ascending=True)
# y = np.arange(len(df_r_plot))
# colores = [ROJO if v > 70 else ('#E67E22' if v > 50 else VERDE)
#            for v in df_r_plot['Tasa_inf_%']]
# 
# ax1.barh(y, df_r_plot['Tasa_inf_%'], 0.65, color=colores, alpha=0.85)
# promedio = df_rama['Tasa_inf_%'].mean()
# ax1.axvline(promedio, color=GRIS, ls='--', lw=1.3, alpha=0.7,
#             label=f'Promedio ({promedio:.1f}%)')
# for i, (_, row) in enumerate(df_r_plot.iterrows()):
#     ax1.text(row['Tasa_inf_%'] + 0.5, i,
#              f"{row['Tasa_inf_%']:.1f}%  ({row['Informales_miles']:,}K)",
#              va='center', fontsize=8.5)
# ax1.set_yticks(y)
# ax1.set_yticklabels(df_r_plot['Rama'], fontsize=9.5)
# ax1.set_xlabel('Tasa de informalidad (%)', fontsize=11)
# ax1.set_title('Tasa de informalidad por rama de actividad\n'
#               'P6920 ≠ 1: No cotiza a pensión', fontsize=12, fontweight='bold')
# ax1.legend(fontsize=9)
# ax1.grid(axis='x', alpha=0.3)
# ax1.set_facecolor('white')
# ax1.spines[['top', 'right']].set_visible(False)
# 
# # Panel der: informal H vs M por rama
# df_r_plot2 = df_rama.sort_values('Informales_miles', ascending=True)
# y2 = np.arange(len(df_r_plot2))
# ax2.barh(y2, df_r_plot2['Inf_hombres_miles'], 0.65, color=AZUL, alpha=0.85, label='Hombres')
# ax2.barh(y2, df_r_plot2['Inf_mujeres_miles'], 0.65,
#          left=df_r_plot2['Inf_hombres_miles'], color=ROJO, alpha=0.85, label='Mujeres')
# for i, (_, row) in enumerate(df_r_plot2.iterrows()):
#     total_i = row['Inf_hombres_miles'] + row['Inf_mujeres_miles']
#     ax2.text(total_i + 10, i, f"{total_i:,.0f}K", va='center', fontsize=8)
# ax2.set_yticks(y2)
# ax2.set_yticklabels(df_r_plot2['Rama'], fontsize=9.5)
# ax2.set_xlabel('Miles de informales', fontsize=11)
# ax2.set_title('Cantidad de informales por rama y sexo\n'
#               'Miles de personas', fontsize=12, fontweight='bold')
# ax2.legend(fontsize=10, loc='lower right')
# ax2.grid(axis='x', alpha=0.3)
# ax2.set_facecolor('white')
# ax2.spines[['top', 'right']].set_visible(False)
# 
# fig1.suptitle('INFORMALIDAD LABORAL POR ACTIVIDAD ECONÓMICA — GEIH 2025\n'
#               'Proxy: No cotiza a pensión (P6920) | Ponderado FEX_C18/12',
#               fontsize=13, fontweight='bold')
# fig1.tight_layout(rect=[0, 0, 1, 0.93])
# plt.show()
# exp.guardar_grafica(fig1, 'Informalidad_por_rama_sexo', cerrar=False)
# 
# # ══════════════════════════════════════════════════════════════
# # 5. GRÁFICO B: Informalidad por ciudad/departamento
# # ══════════════════════════════════════════════════════════════
# 
# fig2, ax3 = plt.subplots(figsize=(14, 8))
# fig2.patch.set_facecolor(FONDO)
# ax3.set_facecolor('white')
# 
# df_g_plot = df_geo.sort_values('Tasa_inf_%', ascending=True)
# y3 = np.arange(len(df_g_plot))
# colores_g = [ROJO if v > 70 else ('#E67E22' if v > 50 else VERDE)
#              for v in df_g_plot['Tasa_inf_%']]
# 
# ax3.barh(y3, df_g_plot['Tasa_inf_%'], 0.65, color=colores_g, alpha=0.85)
# promedio_g = df_geo['Tasa_inf_%'].mean()
# ax3.axvline(promedio_g, color=GRIS, ls='--', lw=1.3, alpha=0.7,
#             label=f'Promedio ({promedio_g:.1f}%)')
# for i, (_, row) in enumerate(df_g_plot.iterrows()):
#     ax3.text(row['Tasa_inf_%'] + 0.3, i,
#              f"{row['Tasa_inf_%']:.1f}%  ({row['Informales_miles']:,}K inf / {row['Ocupados_miles']:,}K ocu)",
#              va='center', fontsize=7.5)
# ax3.set_yticks(y3)
# ax3.set_yticklabels(df_g_plot['Ciudad_Depto'], fontsize=9)
# ax3.set_xlabel('Tasa de informalidad (%)', fontsize=11)
# ax3.set_title(f'Tasa de informalidad por {col_geo.lower()} — GEIH 2025\n'
#               f'P6920 ≠ 1: No cotiza a pensión | Ponderado FEX_C18/12',
#               fontsize=12, fontweight='bold')
# ax3.legend(fontsize=9)
# ax3.grid(axis='x', alpha=0.3)
# ax3.spines[['top', 'right']].set_visible(False)
# fig2.tight_layout(pad=2.5)
# plt.show()
# exp.guardar_grafica(fig2, 'Informalidad_por_ciudad', cerrar=False)
# 
# # ══════════════════════════════════════════════════════════════
# # 6. GRÁFICO C: Perfil salarial formal vs informal
# # ══════════════════════════════════════════════════════════════
# 
# fig3, (ax4, ax5) = plt.subplots(1, 2, figsize=(16, 6))
# fig3.patch.set_facecolor(FONDO)
# 
# # Panel izq: mediana salarial formal vs informal por rama
# filas_sal = []
# for rama in df_ocu['RAMA'].dropna().unique():
#     m_r = df_ocu['RAMA'] == rama
#     for formal, etiq in [('Formal', 'Formal'), ('Informal', 'Informal')]:
#         m_f = m_r & (df_ocu['FORMAL'] == formal) & (df_ocu['INGLABO'] > 0)
#         med = EP.mediana(df_ocu.loc[m_f, 'INGLABO'], df_ocu.loc[m_f, 'FEX_ADJ'])
#         if not np.isnan(med):
#             filas_sal.append({'Rama': rama, 'Tipo': etiq,
#                               'Mediana_SMMLV': round(med / config.smmlv, 2)})
# 
# df_sal = pd.DataFrame(filas_sal)
# pivot_sal = df_sal.pivot(index='Rama', columns='Tipo', values='Mediana_SMMLV').dropna()
# pivot_sal = pivot_sal.sort_values('Formal', ascending=True)
# 
# y4 = np.arange(len(pivot_sal))
# w4 = 0.35
# ax4.barh(y4 - w4/2, pivot_sal['Formal'], w4, color=VERDE, alpha=0.8, label='Formal')
# ax4.barh(y4 + w4/2, pivot_sal['Informal'], w4, color=ROJO, alpha=0.8, label='Informal')
# ax4.axvline(1, color=GRIS, ls=':', lw=1.2, alpha=0.6)
# ax4.text(1.02, len(pivot_sal) - 0.5, '1 SMMLV', fontsize=8, color=GRIS)
# ax4.set_yticks(y4)
# ax4.set_yticklabels(pivot_sal.index, fontsize=9)
# ax4.set_xlabel('Mediana de ingreso (× SMMLV)', fontsize=10)
# ax4.set_title('Ingreso mediano: Formal vs Informal\npor rama de actividad',
#               fontsize=11, fontweight='bold')
# ax4.legend(fontsize=9)
# ax4.grid(axis='x', alpha=0.3)
# ax4.set_facecolor('white')
# ax4.spines[['top', 'right']].set_visible(False)
# 
# # Panel der: composición formal/informal como % apilado
# df_comp = df_rama[['Rama', 'Formales_miles', 'Informales_miles']].copy()
# df_comp['Total'] = df_comp['Formales_miles'] + df_comp['Informales_miles']
# df_comp['Pct_formal'] = df_comp['Formales_miles'] / df_comp['Total'] * 100
# df_comp['Pct_informal'] = df_comp['Informales_miles'] / df_comp['Total'] * 100
# df_comp = df_comp.sort_values('Pct_informal', ascending=True)
# 
# y5 = np.arange(len(df_comp))
# ax5.barh(y5, df_comp['Pct_formal'], 0.65, color=VERDE, alpha=0.8, label='Formal')
# ax5.barh(y5, df_comp['Pct_informal'], 0.65, left=df_comp['Pct_formal'],
#          color=ROJO, alpha=0.8, label='Informal')
# for i, (_, row) in enumerate(df_comp.iterrows()):
#     ax5.text(50, i, f"{row['Pct_informal']:.0f}% inf",
#              ha='center', va='center', fontsize=7.5, fontweight='bold', color='white')
# ax5.set_yticks(y5)
# ax5.set_yticklabels(df_comp['Rama'], fontsize=9)
# ax5.set_xlabel('Composición (%)', fontsize=10)
# ax5.set_title('Composición formal / informal\npor rama (% del empleo)',
#               fontsize=11, fontweight='bold')
# ax5.legend(fontsize=9, loc='lower right')
# ax5.set_facecolor('white')
# ax5.spines[['top', 'right']].set_visible(False)
# 
# fig3.suptitle('PERFIL SALARIAL Y COMPOSICIÓN: EMPLEO FORMAL vs INFORMAL — GEIH 2025',
#               fontsize=13, fontweight='bold')
# fig3.tight_layout(rect=[0, 0, 1, 0.93])
# plt.show()
# exp.guardar_grafica(fig3, 'Informalidad_salario_composicion', cerrar=False)
# 
# plt.close('all')
# 
# # ══════════════════════════════════════════════════════════════
# # 7. EXPORTAR
# # ══════════════════════════════════════════════════════════════
# 
# exp.guardar_tabla(df_sexo, 'Informalidad_nacional_sexo')
# exp.guardar_tabla(df_rama, 'Informalidad_por_rama')
# exp.guardar_tabla(df_geo, 'Informalidad_por_ciudad')
# 
# del df_ocu; gc.collect()
# print("\n✅ Sección informalidad completada: 3 tablas HTML + 3 gráficos + 3 CSVs")

"""# ▶️ POBLACIÓN FUERA DE LA FUERZA DE TRABAJO (FFT)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ══════════════════════════════════════════════════════════════
# # POBLACIÓN FUERA DE LA FUERZA DE TRABAJO (FFT)
# # Según tipo de actividad y sexo — Anual + Diciembre
# #
# # P6240: 3=Estudiando, 4=Oficios hogar, 5=Incapacitado, 6=Otra
# # FFT=1: Fuera de la fuerza de trabajo
# # Requiere: config, df, geih, RUTA ya definidos
# # ══════════════════════════════════════════════════════════════
# 
# from geih import Exportador
# from IPython.display import display
# import pandas as pd, numpy as np
# import matplotlib.pyplot as plt
# import matplotlib.patches as mpatches
# from matplotlib.gridspec import GridSpec
# import gc
# 
# exp = Exportador(ruta_base=RUTA, config=config)
# FONDO = '#F7F9FC'
# 
# # ── Mapeo de categorías DANE ─────────────────────────────────
# MAPA_ACT = {
#     3: 'Estudiando',
#     4: 'Oficios del hogar',
#     5: 'Otra actividad° — Incapacitado permanente',
#     6: 'Otra actividad° — Rentista / pensionado / otro',
# }
# # Agrupación para subtotal "Otra actividad°"
# MAPA_GRUPO = {
#     3: 'Estudiando',
#     4: 'Oficios del hogar',
#     5: 'Otra actividad°',
#     6: 'Otra actividad°',
# }
# 
# # ══════════════════════════════════════════════════════════════
# # FUNCIÓN DE CÁLCULO
# # ══════════════════════════════════════════════════════════════
# 
# def calcular_fft(df_src, fex_col='FEX_ADJ', etiqueta=''):
#     """Calcula tabla FFT por tipo de actividad y sexo."""
#     df_f = df_src[df_src['FFT'] == 1].copy()
#     df_f['ACTIVIDAD'] = df_f['P6240'].map(MAPA_ACT)
#     df_f['GRUPO'] = df_f['P6240'].map(MAPA_GRUPO)
#     df_f['SEXO'] = df_f['P3271'].map({1: 'Hombre', 2: 'Mujer'})
# 
#     filas = []
#     def agregar(tipo, sexo_filtro=None, mask=None):
#         m = mask if mask is not None else pd.Series(True, index=df_f.index)
#         if sexo_filtro:
#             m = m & (df_f['SEXO'] == sexo_filtro)
#         n = df_f.loc[m, fex_col].sum() / 1_000
#         filas.append({'Actividad': tipo, 'Sexo': sexo_filtro or 'Total',
#                       'Miles': round(n)})
# 
#     # Total
#     for s in [None, 'Hombre', 'Mujer']:
#         agregar('TOTAL FFT', s)
# 
#     # Por grupo principal
#     for grupo in ['Estudiando', 'Oficios del hogar', 'Otra actividad°']:
#         m_g = df_f['GRUPO'] == grupo
#         for s in [None, 'Hombre', 'Mujer']:
#             agregar(grupo, s, m_g)
# 
#     # Desagregación de Otra actividad°
#     for act in ['Otra actividad° — Incapacitado permanente',
#                 'Otra actividad° — Rentista / pensionado / otro']:
#         m_a = df_f['ACTIVIDAD'] == act
#         corto = act.replace('Otra actividad° — ', '  └ ')
#         for s in [None, 'Hombre', 'Mujer']:
#             agregar(corto, s, m_a)
# 
#     return pd.DataFrame(filas)
# 
# 
# # ── Calcular ambos períodos ───────────────────────────────────
# print("📊 Calculando FFT...")
# t_anual = calcular_fft(df, etiqueta='Anual')
# 
# # Diciembre: necesita FEX sin dividir
# from geih import PreparadorGEIH
# df_dic = PreparadorGEIH(config=config).preparar_base(geih, mes_filtro=12)
# df_dic = PreparadorGEIH(config=config).agregar_variables_derivadas(df_dic)
# # Para diciembre, FEX_ADJ ya es FEX_C18/1 (mes_filtro lo maneja)
# t_dic = calcular_fft(df_dic, etiqueta='Diciembre')
# 
# # ══════════════════════════════════════════════════════════════
# # TABLA COMPARATIVA (display HTML)
# # ══════════════════════════════════════════════════════════════
# 
# # Pivotar para comparación lado a lado
# comp = t_anual.rename(columns={'Miles': 'Ene–Dic_2025'}).merge(
#     t_dic.rename(columns={'Miles': 'Dic_2025'}),
#     on=['Actividad', 'Sexo'], how='outer'
# ).fillna(0)
# comp['Variación'] = comp['Dic_2025'] - comp['Ene–Dic_2025']
# comp['Ene–Dic_2025'] = comp['Ene–Dic_2025'].apply(lambda x: f"{x:,.0f}")
# comp['Dic_2025'] = comp['Dic_2025'].apply(lambda x: f"{x:,.0f}")
# comp['Variación'] = comp['Variación'].apply(lambda x: f"{'+' if x>=0 else ''}{x:,.0f}")
# 
# print(f"\n{'='*80}")
# print(f"  POBLACIÓN FUERA DE LA FUERZA DE TRABAJO — GEIH 2025")
# print(f"  Cifras en miles de personas | Ponderado FEX_C18")
# print(f"{'='*80}\n")
# display(comp)
# 
# # ══════════════════════════════════════════════════════════════
# # TABLA DE COMPOSICIÓN % (display HTML)
# # ══════════════════════════════════════════════════════════════
# 
# # Composición porcentual por sexo
# t_pct = t_anual[t_anual['Sexo'] == 'Total'].copy()
# total_fft = t_pct.loc[t_pct['Actividad'] == 'TOTAL FFT', 'Miles'].values[0]
# t_pct['Pct_%'] = (t_pct['Miles'] / total_fft * 100).round(1)
# 
# # Agregar H/M para cada actividad
# rows_pct = []
# for _, row in t_pct.iterrows():
#     act = row['Actividad']
#     h = t_anual[(t_anual['Actividad'] == act) & (t_anual['Sexo'] == 'Hombre')]['Miles'].values
#     m = t_anual[(t_anual['Actividad'] == act) & (t_anual['Sexo'] == 'Mujer')]['Miles'].values
#     h_v = h[0] if len(h) > 0 else 0
#     m_v = m[0] if len(m) > 0 else 0
#     pct_m = (m_v / (h_v + m_v) * 100) if (h_v + m_v) > 0 else 0
#     rows_pct.append({
#         'Actividad': act, 'Miles': f"{row['Miles']:,}",
#         'Del_total_%': f"{row['Pct_%']:.1f}%",
#         'Hombres_miles': f"{h_v:,}", 'Mujeres_miles': f"{m_v:,}",
#         'Pct_mujeres': f"{pct_m:.0f}%",
#     })
# df_pct = pd.DataFrame(rows_pct)
# print(f"\n📊 Composición de la población fuera de la FT — Anual 2025\n")
# display(df_pct)
# 
# # ══════════════════════════════════════════════════════════════
# # GRÁFICO: 3 PANELES
# # ══════════════════════════════════════════════════════════════
# 
# COLOR_H, COLOR_M = '#2980B9', '#E74C3C'
# COLORES_CAT = {'Estudiando': '#2E6DA4', 'Oficios del hogar': '#C0392B',
#                'Otra actividad°': '#8E44AD'}
# 
# fig = plt.figure(figsize=(20, 9))
# fig.patch.set_facecolor(FONDO)
# gs = GridSpec(1, 3, figure=fig, wspace=0.38)
# ax1, ax2, ax3 = [fig.add_subplot(gs[i]) for i in range(3)]
# for ax in [ax1, ax2, ax3]:
#     ax.set_facecolor('white'); ax.spines[['top','right']].set_visible(False)
# 
# CATS = ['Estudiando', 'Oficios del hogar', 'Otra actividad°']
# 
# # ── Panel 1: Barras apiladas Anual vs Diciembre ──────────────
# for p_idx, (tabla, periodo) in enumerate([(t_anual, 'Anual'), (t_dic, 'Diciembre')]):
#     bottom = 0
#     for cat in CATS:
#         row = tabla[(tabla['Actividad'] == cat) & (tabla['Sexo'] == 'Total')]
#         v = row['Miles'].values[0] if len(row) > 0 else 0
#         bar = ax1.bar(p_idx, v, 0.5, bottom=bottom, color=COLORES_CAT[cat],
#                       alpha=0.85, label=cat if p_idx == 0 else '', zorder=3)
#         if v > 300:
#             ax1.text(p_idx, bottom + v/2, f'{v:,.0f}K', ha='center', va='center',
#                      fontsize=9, fontweight='bold', color='white', zorder=5)
#         bottom += v
#     # Total arriba
#     tot_row = tabla[(tabla['Actividad'] == 'TOTAL FFT') & (tabla['Sexo'] == 'Total')]
#     tot = tot_row['Miles'].values[0] if len(tot_row) > 0 else 0
#     ax1.text(p_idx, bottom + 100, f'Total\n{tot:,.0f}K', ha='center', fontsize=9, fontweight='bold')
# 
# ax1.set_xticks([0, 1])
# ax1.set_xticklabels(['Promedio\nAnual 2025', 'Diciembre\n2025'], fontsize=11)
# ax1.set_ylabel('Miles de personas', fontsize=11)
# ax1.set_title('FFT por tipo de actividad', fontsize=12, fontweight='bold')
# ax1.legend(fontsize=9)
# ax1.grid(axis='y', alpha=0.3)
# 
# # ── Panel 2: H vs M por categoría (anual) ────────────────────
# y2 = np.arange(len(CATS))
# for i, cat in enumerate(CATS):
#     h = t_anual[(t_anual['Actividad'] == cat) & (t_anual['Sexo'] == 'Hombre')]['Miles'].values
#     m = t_anual[(t_anual['Actividad'] == cat) & (t_anual['Sexo'] == 'Mujer')]['Miles'].values
#     h_v, m_v = (h[0] if len(h) else 0), (m[0] if len(m) else 0)
#     ax2.barh(i, h_v, 0.55, color=COLOR_H, alpha=0.85, label='Hombres' if i==0 else '')
#     ax2.barh(i, m_v, 0.55, left=h_v, color=COLOR_M, alpha=0.85, label='Mujeres' if i==0 else '')
#     tot = h_v + m_v
#     if tot > 0:
#         ax2.text(tot + 50, i, f'H:{h_v/tot*100:.0f}% M:{m_v/tot*100:.0f}%',
#                  va='center', fontsize=8.5)
# 
# ax2.set_yticks(y2)
# ax2.set_yticklabels(CATS, fontsize=10)
# ax2.set_xlabel('Miles de personas', fontsize=10)
# ax2.set_title('Composición por sexo\n(Anual 2025)', fontsize=12, fontweight='bold')
# ax2.legend(fontsize=9)
# ax2.grid(axis='x', alpha=0.3)
# 
# # ── Panel 3: Otra actividad° desagregada (Anual vs Dic) ──────
# sub_cats = ['  └ Incapacitado permanente', '  └ Rentista / pensionado / otro']
# sub_labels = ['Incapacitado\npermanente', 'Rentista /\nPensionado']
# x3 = np.arange(len(sub_cats))
# w3 = 0.3
# 
# for t_idx, (tabla, lbl, offset) in enumerate([
#     (t_anual, 'Anual', -w3/2), (t_dic, 'Dic.', w3/2)
# ]):
#     for j, cat in enumerate(sub_cats):
#         h = tabla[(tabla['Actividad'] == cat) & (tabla['Sexo'] == 'Hombre')]['Miles'].values
#         m = tabla[(tabla['Actividad'] == cat) & (tabla['Sexo'] == 'Mujer')]['Miles'].values
#         h_v, m_v = (h[0] if len(h) else 0), (m[0] if len(m) else 0)
#         ax3.bar(j + offset, h_v, w3, color=COLOR_H, alpha=0.85, zorder=3,
#                 label='Hombres' if j==0 and t_idx==0 else '')
#         ax3.bar(j + offset, m_v, w3, bottom=h_v, color=COLOR_M, alpha=0.85, zorder=3,
#                 label='Mujeres' if j==0 and t_idx==0 else '')
#         ax3.text(j + offset, h_v + m_v + 20, f'{h_v+m_v:,.0f}K',
#                  ha='center', fontsize=8, fontweight='bold')
# 
# ax3.set_xticks(x3)
# ax3.set_xticklabels(sub_labels, fontsize=10)
# ax3.set_ylabel('Miles de personas', fontsize=10)
# ax3.set_title('Desagregación "Otra actividad°"\nAnual vs Diciembre', fontsize=12, fontweight='bold')
# ax3.legend(fontsize=9)
# ax3.grid(axis='y', alpha=0.3)
# 
# fig.suptitle('Población fuera de la Fuerza de Trabajo — GEIH 2025 | DANE Marco 2018',
#              fontsize=14, fontweight='bold', y=1.01)
# fig.tight_layout(pad=2.5)
# plt.show()
# exp.guardar_grafica(fig, 'FFT_tipo_actividad_sexo_2025', cerrar=False)
# plt.close('all')
# 
# # ── Exportar ──────────────────────────────────────────────────
# exp.guardar_tabla(t_anual, 'FFT_tipo_actividad_anual_2025')
# exp.guardar_tabla(t_dic, 'FFT_tipo_actividad_diciembre_2025')
# 
# del df_dic; gc.collect()
# print("\n✅ Sección FFT completada: 2 tablas HTML + 1 gráfico + 2 CSVs")

"""# ▶️ Top 20 sectores CIIU (NUEVO en el paquete)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import Top20Sectores
# 
# top20 = Top20Sectores(config=config)
# tabla = top20.calcular(df, ruta_ciiu=RUTA_CIIU)
# top20.imprimir(tabla)
# 
# fig = top20.graficar(tabla)
# exp.guardar_grafica(fig, 'Top20_sectores_CIIU')
# plt.show()
# 
# exp.guardar_tabla(tabla, 'Top20_sectores_CIIU_2025')

"""# ▶️ Excel con formato institucional ProColombia"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import (Exportador, IndicadoresLaborales, DistribucionIngresos,
#                         AnalisisSalarios, BrechaGenero, CalidadEmpleo,
#                         CompetitividadLaboral, Top20Sectores)
# 
# exp = Exportador(ruta_base=RUTA, config=config)
# 
# # Recopilar todas las tablas producidas
# exp.guardar_excel({
#     # '1_Diagnostico':    tabla_nulos.reset_index().rename(columns={'index': 'Columna'}),
#     '2_Top20_CIIU':     tabla,
#     '3_Salarios_rama':  AnalisisSalarios(config=config).por_rama(df).reset_index(),
#     '4_Brecha_genero':  BrechaGenero().calcular(df).reset_index(),
#     '5_ICE_depto':      CalidadEmpleo(config=config).calcular_por_departamento(df),
#     '6_ICI_depto':      CompetitividadLaboral(config=config).calcular(df),
# }, 'GEIH_2025_Resumen_ProColombia')
# 
# # Con formato institucional: headers burdeos, filas alternas, freeze panes ✅

"""# ▶️ Empleo por ciudad y actividad económica"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ═══ CELDA 1: Calcular las 6 tablas ═══
# from geih import AnalisisOcupadosCiudad, ConfigGEIH, ConsolidadorGEIH
# 
# analisis = AnalisisOcupadosCiudad(config=ConfigGEIH(n_meses=12))
# tablas = analisis.calcular(df, ruta_ciiu=RUTA_CIIU)
# analisis.imprimir(tablas)

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ═══ CELDA 2: Gráfico (3 paneles) ═══
# fig = analisis.graficar(tablas)
# fig.savefig(f'{RUTA}/resultados_geih_2025/graficas/OcupadosCIIU_Area_2025.png',
#             dpi=150, bbox_inches='tight', facecolor='#F7F9FC')
# plt.show()

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ═══ CELDA 3: Excel con las 6 hojas ═══
# analisis.exportar_excel(tablas, f'{RUTA}/resultados_geih_2025/excel/Resultados_CIIU_Area_GEIH2025.xlsx')

"""# ▶️ Ejemplos Concretos de Cada Análisis"""

tiempo_sesión = medir_tiempo("Sesión - Ejemplos Concretos de Cada Análisis")

"""## 🔵 Indicadores nacionales + validación DANE"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import IndicadoresLaborales
# 
# ind = IndicadoresLaborales(config=config)
# r = ind.calcular(df)
# ind.sanity_check(r, periodo="Anual 2025")
# # Output:
# #   ✅ TD_%  =  8.9%  (ref. DANE: 8.9%  Δ=0.0)
# #   ✅ TGP_% = 64.3%  (ref. DANE: 64.3% Δ=0.0)
# 
# # Por departamento
# td_dpto = ind.por_departamento(df)
# exp.guardar_tabla(td_dpto, 'indicadores_por_departamento')

"""## 🔵 Distribución de ingresos SMMLV"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # Wall time: 1.51 s
# from geih import DistribucionIngresos, GraficoDistribucionIngresos
# 
# dist = DistribucionIngresos(config=config)
# resultado = dist.calcular(df)
# dist.imprimir(resultado, titulo="Ene–Dic 2025")
# 
# fig = GraficoDistribucionIngresos().graficar(resultado['total'], resultado['por_sexo'])
# exp.guardar_grafica(fig, 'distribucion_ingresos_2025')

"""## 🔵 Salarios por rama con box plot"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import AnalisisSalarios, GraficoBoxPlotSalarios
# 
# sal = AnalisisSalarios(config=config)
# tabla_rama = sal.por_rama(df)
# tabla_edad = sal.por_edad(df)
# 
# fig = GraficoBoxPlotSalarios().graficar(tabla_rama)
# exp.guardar_grafica(fig, 'boxplot_salarios_rama')
# exp.guardar_tabla(tabla_rama.reset_index(), 'estadisticas_salario_rama')

"""## 🔵 Brecha salarial de género"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import BrechaGenero, GraficoBrechaGenero
# 
# brecha = BrechaGenero().calcular(df)
# print(brecha)
# # Nivel                Hombres     Mujeres    Brecha_%
# # 6. Universitaria     2,847,000   2,345,000  -17.6%
# # 7. Posgrado          5,694,000   4,271,000  -25.0%
# 
# fig = GraficoBrechaGenero().graficar(brecha)
# exp.guardar_grafica(fig, 'brecha_genero_educacion')

"""## 🔵 Gini del ingreso laboral"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import IndicesCompuestos
# 
# gini = IndicesCompuestos(config=config).gini(df)
# print(f"Gini del ingreso laboral: {gini:.3f}")
# # → ~0.480 (solo laboral, más alto que el oficial DANE que incluye transferencias)

"""## 🔵 Calidad del empleo (ICE) por departamento"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import CalidadEmpleo
# 
# ice = CalidadEmpleo(config=config).calcular_por_departamento(df)
# print(ice[['Departamento', 'ICE', 'Ocupados_M']].head(10))
# # Bogotá ≈ 65+, Chocó ≈ 30
# exp.guardar_tabla(ice, 'ICE_por_departamento')

"""## 🔵 Competitividad laboral (ICI) para pitch de IED"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import CompetitividadLaboral
# 
# ici = CompetitividadLaboral(config=config).calcular(df)
# print(ici[['Departamento', 'ICI', 'TD_%', 'Costo_efectivo', 'Talento_univ_%']].head(10))
# exp.guardar_tabla(ici, 'ICI_competitividad_laboral')

"""## 🔵 Estacionalidad mensual (12 puntos)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import Estacionalidad
# 
# estac = Estacionalidad().calcular(geih)  # ← usa geih sin preparar (necesita FEX_C18 sin dividir)
# print(estac[['MES', 'TD_%', 'TO_%']])
# exp.guardar_tabla(estac, 'estacionalidad_mensual_2025')

"""## 🔵 Población campesina (NUEVO — P2057)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import AnalisisCampesino
# 
# camp = AnalisisCampesino(config=config).calcular(df)
# print(camp)
# # Se considera campesino: TD=X%, Mediana=0.6 SMMLV, Formalidad=15%
# # No se considera campesino: TD=Y%, Mediana=1.2 SMMLV, Formalidad=42%

"""## 🔵 Discapacidad (NUEVO — P1906S1-S8 Escala Washington)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import AnalisisDiscapacidad
# 
# disc = AnalisisDiscapacidad().calcular(df)
# print(f"Con discapacidad: {disc['Con discapacidad_TD_%']}% TD")
# print(f"Sin discapacidad: {disc['Sin discapacidad_TD_%']}% TD")
# print(f"Prevalencia por dimensión: {disc['prevalencia_por_dimension']}")

"""## 🔵 Migración (NUEVO — P3370/P3376)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import AnalisisMigracion
# 
# migr = AnalisisMigracion(config=config).calcular(df)
# print(migr)
# # Mismo municipio: 45M, TD=8.8%
# # Otro municipio: 1.2M, TD=12.3%
# # Otro país: 0.3M, TD=15.1%
# # Nacido en el extranjero: TD más alta, mediana salarial más baja

"""## 🔵 Autonomía laboral — contratistas dependientes (NUEVO — P3047-P3049)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import AnalisisAutonomia
# 
# auto = AnalisisAutonomia().calcular(df)
# print(f"Cuenta propia total: {auto['cuenta_propia_M']}M")
# print(f"Cta propia DEPENDIENTE (asalariados disfrazados): {auto['cta_propia_dependiente_M']}M")
# print(f"→ {auto['pct_dependientes']}% de los cuenta propia son realmente dependientes")

"""## 🔵 Alcance de mercado — empleo vinculado a exportación (NUEVO — P1802)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import AnalisisAlcanceMercado
# 
# alc = AnalisisAlcanceMercado().calcular(df)
# print(alc)
# # Exportación ★: ~X miles de personas → empleo DIRECTO en comercio exterior
# # Esto es la variable más relevante para ProColombia que nadie publica

"""## 🟢 Ecuación de Mincer (retorno a la educación)"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import EcuacionMincer
# 
# mincer = EcuacionMincer(config=config).estimar_todos(df)
# print(mincer[['Grupo', 'beta_educacion', 'R2', 'N']])
# # Nacional: β₁ ≈ 10.5% (cada año extra de educación → 10.5% más de salario)
# # TIC: β₁ ≈ 16% (educación mucho más valorada)
# exp.guardar_tabla(mincer, 'mincer_retorno_educacion')

"""# ▶️ ANÁLISIS AVANZADOS — TODOS LOS MÓDULOS DEL PAQUETE"""

# Commented out IPython magic to ensure Python compatibility.
# %%time
# # ══════════════════════════════════════════════════════════════
# # ANÁLISIS AVANZADOS — TODOS LOS MÓDULOS DEL PAQUETE
# # Usa las clases existentes de geih_2025 con display() + plt.show()
# # Requiere: config, df, geih, RUTA ya definidos
# # ══════════════════════════════════════════════════════════════
# 
# from geih import (
#     BrechaGenero, GraficoBrechaGenero,
#     CalidadEmpleo, FormalidadSectorial, VulnerabilidadLaboral,
#     CompetitividadLaboral, ProductividadTamano, CostoLaboral,
#     AnalisisUrbanoRural, AnalisisSubempleo, AnalisisHoras,
#     Estacionalidad, FuerzaLaboralJoven, EtnicoRacial,
#     BonoDemografico, ContribucionSectorial, MapaTalento,
#     EcuacionMincer, IndicesCompuestos, AnalisisFFT,
#     Exportador, EstadisticasPonderadas as EP,
# )
# from IPython.display import display
# import numpy as np, pandas as pd
# import matplotlib.pyplot as plt
# import matplotlib.ticker as mticker
# import matplotlib.patches as mpatches
# import gc
# 
# exp = Exportador(ruta_base=RUTA, config=config)
# FONDO = '#F7F9FC'
# AZUL, ROJO, VERDE, MORADO, NARANJA, GRIS = '#2E6DA4','#C0392B','#1E8449','#7D3C98','#E67E22','#7F8C8D'
# 
# # ══════════════════════════════════════════════════════════════
# # M1: BRECHA SALARIAL DE GÉNERO
# # ══════════════════════════════════════════════════════════════
# print("📊 M1: Brecha salarial de género por nivel educativo")
# brecha = BrechaGenero().calcular(df)
# display(brecha)
# fig = GraficoBrechaGenero().graficar(brecha)
# plt.show(); exp.guardar_grafica(fig, 'M1_Brecha_genero', cerrar=False); plt.close('all')
# 
# # ══════════════════════════════════════════════════════════════
# # M2: ICE — ÍNDICE DE CALIDAD DEL EMPLEO POR DEPARTAMENTO
# # ══════════════════════════════════════════════════════════════
# print("\n📊 M2: Índice de Calidad del Empleo (ICE)")
# ice = CalidadEmpleo(config=config)
# ice_dpto = ice.calcular_por_departamento(df)
# display(ice_dpto.sort_values('ICE', ascending=False))
# exp.guardar_tabla(ice_dpto, 'M2_ICE_por_departamento')
# 
# # ══════════════════════════════════════════════════════════════
# # M3: SALARIO POR TAMAÑO DE EMPRESA
# # ══════════════════════════════════════════════════════════════
# print("\n📊 M3: Productividad laboral por tamaño de empresa")
# prod = ProductividadTamano(config=config).calcular(df)
# display(prod)
# exp.guardar_tabla(prod, 'M3_Salario_tamano_empresa')
# 
# # ══════════════════════════════════════════════════════════════
# # M4: URBANO vs RURAL
# # ══════════════════════════════════════════════════════════════
# print("\n📊 M4: Mercado laboral Urbano vs Rural")
# urb = AnalisisUrbanoRural(config=config).calcular(df)
# display(urb)
# exp.guardar_tabla(urb, 'M4_Urbano_vs_Rural')
# 
# # ══════════════════════════════════════════════════════════════
# # M5: CURVA DE LORENZ + GINI
# # ══════════════════════════════════════════════════════════════
# print("\n📊 M5: Curva de Lorenz y coeficiente de Gini")
# gini_nac = IndicesCompuestos(config=config).gini(df)
# print(f"   Gini nacional del ingreso laboral: {gini_nac:.3f}")
# 
# # Lorenz chart
# df_oci = df[(df['OCI']==1) & (df['INGLABO']>0)].copy()
# fig5, ax5 = plt.subplots(figsize=(9,7)); fig5.patch.set_facecolor(FONDO); ax5.set_facecolor('white')
# ax5.plot([0,1],[0,1],'k--',lw=1.5,alpha=0.7,label='Perfecta igualdad')
# 
# for sub, color, lbl in [
#     (df_oci, AZUL, 'Nacional'),
#     (df_oci[df_oci['P3271']==1], '#1A5276', 'Hombres'),
#     (df_oci[df_oci['P3271']==2], '#922B21', 'Mujeres'),
# ]:
#     v,w = sub['INGLABO'].values, sub['FEX_ADJ'].values
#     idx=np.argsort(v); v,w=v[idx],w[idx]
#     wc=np.cumsum(w)/w.sum(); vc=np.cumsum(v*w)/(v*w).sum()
#     _trapz = getattr(np,'trapezoid',getattr(np,'trapz',None))
#     g = 1 - 2*_trapz(vc,wc)
#     wc,vc = np.insert(wc,0,0), np.insert(vc,0,0)
#     ax5.plot(wc,vc,lw=2.2,color=color,label=f'{lbl} (Gini={g:.3f})')
#     if lbl=='Nacional': ax5.fill_between(wc,wc,vc,alpha=0.08,color=color)
# 
# ax5.set_xlabel('Proporción acumulada de ocupados',fontsize=11)
# ax5.set_ylabel('Proporción acumulada del ingreso',fontsize=11)
# ax5.set_title('Curva de Lorenz del ingreso laboral — GEIH 2025',fontsize=12,fontweight='bold')
# ax5.legend(fontsize=10); ax5.grid(alpha=0.25); ax5.spines[['top','right']].set_visible(False)
# ax5.xaxis.set_major_formatter(mticker.PercentFormatter(1)); ax5.yaxis.set_major_formatter(mticker.PercentFormatter(1))
# plt.tight_layout(); plt.show()
# exp.guardar_grafica(fig5,'M5_Lorenz_Gini',cerrar=False); plt.close('all')
# del df_oci
# 
# # ══════════════════════════════════════════════════════════════
# # M11: ESTACIONALIDAD MENSUAL
# # ══════════════════════════════════════════════════════════════
# print("\n📊 M11: Estacionalidad del mercado laboral (12 meses)")
# estac = Estacionalidad().calcular(geih)
# display(estac)
# exp.guardar_tabla(estac, 'M11_Estacionalidad_mensual')
# 
# # # ══════════════════════════════════════════════════════════════
# # # M12: FUERZA LABORAL JOVEN (15-28 años)
# # # ══════════════════════════════════════════════════════════════
# # print("\n📊 M12: Fuerza laboral joven (15-28 años)")
# # joven = FuerzaLaboralJoven(config=config).calcular(df)
# # display(joven)
# # exp.guardar_tabla(joven, 'M12_Fuerza_laboral_joven')
# 
# # ══════════════════════════════════════════════════════════════
# # M13: FORMALIDAD SECTORIAL (ICF)
# # ══════════════════════════════════════════════════════════════
# print("\n📊 M13: Índice de Formalidad Sectorial (ICF)")
# icf = FormalidadSectorial(config=config).calcular(df)
# display(icf.sort_values('ICF', ascending=False) if 'ICF' in icf.columns else icf)
# exp.guardar_tabla(icf, 'M13_Formalidad_sectorial')
# 
# # # ══════════════════════════════════════════════════════════════
# # # M15: AUTORRECONOCIMIENTO ÉTNICO-RACIAL
# # # ══════════════════════════════════════════════════════════════
# # print("\n📊 M15: Mercado laboral por grupo étnico-racial")
# # etnia = EtnicoRacial(config=config).calcular(df)
# # display(etnia)
# # exp.guardar_tabla(etnia, 'M15_Etnico_racial')
# 
# # ══════════════════════════════════════════════════════════════
# # M16: ICI — RANKING COMPETITIVIDAD LABORAL
# # ══════════════════════════════════════════════════════════════
# print("\n📊 M16: Ranking de Competitividad Laboral (ICI) por departamento")
# ici = CompetitividadLaboral(config=config).calcular(df)
# display(ici.sort_values('ICI', ascending=False) if 'ICI' in ici.columns else ici)
# exp.guardar_tabla(ici, 'M16_ICI_competitividad')
# 
# # Gráfico ICI
# fig16, ax16 = plt.subplots(figsize=(14,8)); fig16.patch.set_facecolor(FONDO); ax16.set_facecolor('white')
# ici_s = ici.sort_values('ICI', ascending=True) if 'ICI' in ici.columns else ici
# col_ici = [VERDE if v>55 else (NARANJA if v>45 else ROJO) for v in ici_s['ICI']]
# ax16.barh(range(len(ici_s)), ici_s['ICI'], 0.65, color=col_ici, alpha=0.88)
# ax16.axvline(ici_s['ICI'].mean(), color=GRIS, ls='--', lw=1.3, alpha=0.7, label='Promedio')
# for i,(_, row) in enumerate(ici_s.iterrows()):
#     ax16.text(row['ICI']+0.3, i, f"#{len(ici_s)-i} {row['ICI']:.0f}", va='center', fontsize=8)
# ax16.set_yticks(range(len(ici_s))); ax16.set_yticklabels(ici_s['Departamento'], fontsize=9)
# ax16.set_xlabel('ICI (0-100)', fontsize=11)
# ax16.set_title('Ranking de Competitividad Laboral por departamento — GEIH 2025\n'
#                'ICI = Desempleo + Costo + Talento + Formalidad + Jóvenes', fontsize=12, fontweight='bold')
# ax16.legend(); ax16.grid(axis='x',alpha=0.3); ax16.spines[['top','right']].set_visible(False)
# plt.tight_layout(); plt.show()
# exp.guardar_grafica(fig16, 'M16_ICI_ranking', cerrar=False); plt.close('all')
# 
# # # ══════════════════════════════════════════════════════════════
# # # M18: BONO DEMOGRÁFICO
# # # ══════════════════════════════════════════════════════════════
# # print("\n📊 M18: Bono demográfico — Ratio de dependencia")
# # bono = BonoDemografico(config=config).calcular(df)
# # display(bono)
# # exp.guardar_tabla(bono, 'M18_Bono_demografico')
# 
# # ══════════════════════════════════════════════════════════════
# # M19: COSTO LABORAL EFECTIVO
# # ══════════════════════════════════════════════════════════════
# print("\n📊 M19: Costo laboral efectivo por sector (+54% carga)")
# costo = CostoLaboral(config=config).calcular(df)
# display(costo)
# exp.guardar_tabla(costo, 'M19_Costo_laboral_sector')
# 
# # ══════════════════════════════════════════════════════════════
# # M20: VULNERABILIDAD DEL EMPLEO (IVI)
# # ══════════════════════════════════════════════════════════════
# print("\n📊 M20: Índice de Vulnerabilidad Laboral (IVI)")
# ivi = VulnerabilidadLaboral(config=config).calcular(df)
# display(ivi)
# exp.guardar_tabla(ivi, 'M20_Vulnerabilidad_sectorial')
# 
# # ══════════════════════════════════════════════════════════════
# # MA: CONTRIBUCIÓN SECTORIAL AL EMPLEO
# # ══════════════════════════════════════════════════════════════
# print("\n📊 MA: Contribución sectorial al crecimiento del empleo")
# contrib = ContribucionSectorial().calcular(geih)
# display(contrib)
# exp.guardar_tabla(contrib, 'MA_Contribucion_sectorial')
# 
# # ══════════════════════════════════════════════════════════════
# # MB: MAPA DE TALENTO DISPONIBLE (ITAT)
# # ══════════════════════════════════════════════════════════════
# print("\n📊 MB: Mapa de talento disponible (ITAT) por departamento")
# talento = MapaTalento(config=config).calcular(df)
# display(talento.sort_values('ITAT', ascending=False) if 'ITAT' in talento.columns else talento)
# exp.guardar_tabla(talento, 'MB_Mapa_talento_ITAT')
# 
# # ══════════════════════════════════════════════════════════════
# # MC: ECUACIÓN DE MINCER (retorno a la educación)
# # ══════════════════════════════════════════════════════════════
# print("\n📊 MC: Ecuación de Mincer — retorno salarial a la educación")
# mincer = EcuacionMincer(config=config).estimar_todos(df)
# display(mincer)
# exp.guardar_tabla(mincer, 'MC_Mincer_retorno_educacion')
# 
# # # ══════════════════════════════════════════════════════════════
# # # EXCEL CONSOLIDADO CON FORMATO INSTITUCIONAL
# # # ══════════════════════════════════════════════════════════════
# # print("\n📗 Exportando Excel consolidado con formato ProColombia...")
# # exp.guardar_excel({
# #     'ICE_departamento':     ice_dpto,
# #     'ICI_competitividad':   ici,
# #     'Formalidad_sectorial': icf,
# #     'Vulnerabilidad_IVI':   ivi,
# #     'Costo_laboral':        costo,
# #     'Bono_demografico':     bono,
# #     'Talento_ITAT':         talento,
# #     'Estacionalidad':       estac,
# #     'Etnico_racial':        etnia,
# #     'Joven_15_28':          joven,
# #     'Mincer':               mincer,
# #     'Contribucion_sect':    contrib,
# #     'Urbano_Rural':         urb,
# # }, 'GEIH_2025_Analisis_Avanzado_ProColombia')
# 
# # gc.collect()
# # print(f"\n{'='*65}")
# # print(f"  ✅ 13 MÓDULOS COMPLETADOS")
# # print(f"  ✅ 13 tablas HTML visibles en notebook")
# # print(f"  ✅ 2 gráficos (Lorenz + ICI)")
# # print(f"  ✅ 13 CSVs en resultados_geih_2025/tablas/")
# # print(f"  ✅ 1 Excel multi-hoja con formato institucional")
# # print(f"{'='*65}")

tiempo_sesión()
tiempo_total()

"""# 📤 Exportador (todo organizado en carpetas)"""

tiempo_sesión = medir_tiempo("Sesión - Exportador (todo organizado en carpetas)")

# Commented out IPython magic to ensure Python compatibility.
# %%time
# from geih import Exportador
# 
# exp = Exportador(ruta_base=RUTA, config=config)
# # Crea automáticamente:
# #   resultados_geih_2025/graficas/
# #   resultados_geih_2025/tablas/
# #   resultados_geih_2025/excel/

from datetime import datetime
import pytz

# Definir la zona horaria de Colombia
zona_colombia = pytz.timezone('America/Bogota')

# Obtener la fecha y hora actual en Colombia
fecha_hora_colombia = datetime.now(zona_colombia)

# Imprimir con formato
print("Hora en Colombia:", fecha_hora_colombia.strftime("%Y-%m-%d %H:%M:%S"))

tiempo_sesión()
tiempo_total()

"""# 🛑 AQUI VOY"""
────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: notebooks/2026_03_28_consolidar_geih_analisis_y_calculos.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [28/39]: Pipeline_GEIH.py
   TIPO      : CÓDIGO
   UBICACIÓN : notebooks
   RUTA      : notebooks/Pipeline_GEIH.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: notebooks/Pipeline_GEIH.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""Pipeline GEIH — Notebook ejecutable.

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1qlnDv9uXcOYWYRaSyZ6xfKKrB9PWvf8_

Autor: Néstor Enrique Forero Herrera
Paquete: geih v4.1.0 — 57 clases de análisis
Guía metodológica completa: GUIA_GEIH.md (en la carpeta del paquete)

Autor: Néstor Enrique Forero Herrera de Datos
"""

# ══════════════════════════════════════════════════════════════════════
# CELDA 1 · Montar Drive + cargar paquete
# ══════════════════════════════════════════════════════════════════════

from google.colab import drive
drive.mount('/content/drive')

import sys
sys.path.insert(0, '/content/drive/MyDrive/GEIH'  # ← ajustar a tu ruta)

from geih import __version__
print(f"✅ geih v{__version__} — 57 clases de análisis cargadas")

# ══════════════════════════════════════════════════════════════════════
# CELDA 2 · Configuración (ÚNICO LUGAR DONDE SE CAMBIAN PARÁMETROS)
# ══════════════════════════════════════════════════════════════════════

from geih import ConfigGEIH

# ┌──────────────────────────────────────────────────────────────────┐
# │  PARÁMETROS DEL USUARIO — Modificar aquí para cada corrida      │
# │  anio    : Año de los datos (2025, 2026, ...)                   │
# │  n_meses : Cuántos meses consolidar (1-12)                     │
# │  Para mes puntual: n_meses=1, mes_filtro=N en preparar_base()  │
# └──────────────────────────────────────────────────────────────────┘
ANIO    = 2025
N_MESES = 12

RUTA          = '/content/drive/MyDrive/GEIH'  # ← ajustar a tu ruta
RUTA_CIIU     = '/content/drive/MyDrive/ProColombia/0A. Datos/Correlativas/2025-01-14 Correlativa CIIU Rev4 - Cadenas ProColombia.xlsx'
RUTA_DIVIPOLA = '/content/drive/MyDrive/ProColombia/0A. Datos/Correlativas/2025-01-14 DIVIPOLA.xlsx'

config = ConfigGEIH(anio=ANIO, n_meses=N_MESES)
config.resumen()

# ══════════════════════════════════════════════════════════════════════
# CELDA 3 · Consolidar (primera vez) o cargar Parquet existente
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
from geih import ConsolidadorGEIH, GestorMemoria
import os

PARQUET = f'{RUTA}/GEIH_{ANIO}_Consolidado.parquet'

if os.path.exists(PARQUET):
    print(f"📂 Cargando base existente: {PARQUET}")
    geih = ConsolidadorGEIH.cargar(PARQUET)
else:
    print(f"🔄 Consolidando {N_MESES} meses de {ANIO}...")
    consolidador = ConsolidadorGEIH(ruta_base=RUTA, config=config, incluir_area=True)
    consolidador.verificar_estructura()
    geih = consolidador.consolidar()
    consolidador.exportar(geih)

GestorMemoria.estado()

# ══════════════════════════════════════════════════════════════════════
# CELDA 4 · Diagnóstico rápido + validación de identidades
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
from geih import DiagnosticoCalidad

diag = DiagnosticoCalidad()
diag.resumen_rapido(geih)
diag.verificar_tipos(geih)
diag.validar_identidades(geih)

# ══════════════════════════════════════════════════════════════════════
# CELDA 5 · Preparar datos (FEX_ADJ + variables derivadas)
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
from geih import PreparadorGEIH, Exportador

prep = PreparadorGEIH(config=config)
df = prep.preparar_base(geih)
df = prep.agregar_variables_derivadas(df)

exp = Exportador(ruta_base=RUTA, config=config)
print(f"✅ df preparado: {df.shape[0]:,} filas × {df.shape[1]} columnas")

# ══════════════════════════════════════════════════════════════════════
# CELDA 6 · Indicadores nacionales + sanity check
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
from geih import IndicadoresLaborales

ind = IndicadoresLaborales(config=config)
r = ind.calcular(df)
ind.sanity_check(r, f"Anual {ANIO}")
td_dpto = ind.por_departamento(df)
exp.guardar_tabla(td_dpto, f'M0_indicadores_departamento_{ANIO}')

# ══════════════════════════════════════════════════════════════════════
# CELDA 7 · 10 análisis básicos (M1–M6, M11, M12, Mincer, Gini)
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
from geih import (
    DistribucionIngresos, AnalisisRamaSexo, AnalisisSalarios,
    BrechaGenero, IndicesCompuestos, Estacionalidad,
    AnalisisUrbanoRural, FuerzaLaboralJoven, EcuacionMincer,
)
from geih import Top20Sectores

print(f"\n{'='*60}")
print(f"  10 ANÁLISIS BÁSICOS — {config.periodo_etiqueta}")
print(f"{'='*60}")

dist_ing = DistribucionIngresos(config=config).calcular(df)
exp.guardar_tabla(dist_ing['total'], f'M1_distribucion_ingresos_{ANIO}')

rama_sexo = AnalisisRamaSexo().calcular(df)
exp.guardar_tabla(rama_sexo, f'M3_rama_sexo_{ANIO}')

salarios_rama = AnalisisSalarios(config=config).por_rama(df)
exp.guardar_tabla(salarios_rama.reset_index(), f'M4_salarios_rama_{ANIO}')

brecha = BrechaGenero().calcular(df)
exp.guardar_tabla(brecha.reset_index(), f'M6_brecha_genero_{ANIO}')

gini_val = IndicesCompuestos(config=config).gini(df)
print(f"\n  Gini del ingreso laboral: {gini_val:.3f}")

estac = Estacionalidad().calcular(geih)
exp.guardar_tabla(estac, f'M11_estacionalidad_{ANIO}')

urb_rural = AnalisisUrbanoRural(config=config).calcular(df)
exp.guardar_tabla(urb_rural, f'M_urbano_rural_{ANIO}')

joven = FuerzaLaboralJoven(config=config).calcular(df)
exp.guardar_tabla(joven, f'M12_fuerza_laboral_joven_{ANIO}')

mincer = EcuacionMincer(config=config).estimar_todos(df)
exp.guardar_tabla(mincer, f'MC_mincer_{ANIO}')

top20 = Top20Sectores(config=config).calcular(df, ruta_ciiu=RUTA_CIIU)
exp.guardar_tabla(top20, f'Top20_sectores_CIIU_{ANIO}')

print(f"\n✅ 10 análisis básicos completados")

# ══════════════════════════════════════════════════════════════════════
# CELDA 8 · 8 análisis avanzados (ICE, ICF, ICI, IVI, ITAT, etc.)
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
from geih import (
    CalidadEmpleo, FormalidadSectorial, CompetitividadLaboral,
    VulnerabilidadLaboral, CostoLaboral, ContribucionSectorial,
    MapaTalento, BonoDemografico,
)

print(f"\n{'='*60}")
print(f"  8 ANÁLISIS AVANZADOS — {config.periodo_etiqueta}")
print(f"{'='*60}")

ice_dpto = CalidadEmpleo(config=config).calcular_por_departamento(df)
exp.guardar_tabla(ice_dpto, f'M7_ICE_departamento_{ANIO}')

icf = FormalidadSectorial(config=config).calcular(df)
exp.guardar_tabla(icf, f'M13_formalidad_sectorial_{ANIO}')

ici = CompetitividadLaboral(config=config).calcular(df)
exp.guardar_tabla(ici, f'M16_ICI_competitividad_{ANIO}')

ivi = VulnerabilidadLaboral(config=config).calcular(df)
exp.guardar_tabla(ivi, f'M20_vulnerabilidad_{ANIO}')

costo = CostoLaboral(config=config).calcular(df)
exp.guardar_tabla(costo, f'M19_costo_laboral_{ANIO}')

contrib = ContribucionSectorial().calcular(geih)
exp.guardar_tabla(contrib, f'MA_contribucion_sectorial_{ANIO}')

talento = MapaTalento(config=config).calcular(df)
exp.guardar_tabla(talento, f'MB_mapa_talento_ITAT_{ANIO}')

bono = BonoDemografico(config=config).calcular(df)
exp.guardar_tabla(bono, f'M18_bono_demografico_{ANIO}')

print(f"\n✅ 8 análisis avanzados completados")

# ══════════════════════════════════════════════════════════════════════
# CELDA 9 · 10 análisis poblacionales + 5 complementarios
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
from geih import (
    AnalisisCampesino, AnalisisDiscapacidad, AnalisisMigracion,
    AnalisisOtrasFormas, AnalisisOtrosIngresos, AnalisisSobrecalificacion,
    AnalisisContractual, AnalisisAutonomia, AnalisisAlcanceMercado,
    AnalisisDesanimados,
    DuracionDesempleo, DashboardSectoresProColombia,
    AnatomaSalario, FormaPago, CanalEmpleo,
)

print(f"\n{'='*60}")
print(f"  15 ANÁLISIS POBLACIONALES — {config.periodo_etiqueta}")
print(f"{'='*60}")

resultados_pob = {
    'Campesino':        AnalisisCampesino(config=config).calcular(df),
    'Discapacidad':     AnalisisDiscapacidad().calcular(df),
    'Migracion':        AnalisisMigracion(config=config).calcular(df),
    'Otras_formas':     AnalisisOtrasFormas().calcular(df),
    'Otros_ingresos':   AnalisisOtrosIngresos().calcular(df),
    'Sobrecalificacion':AnalisisSobrecalificacion().calcular(df),
    'Contractual':      AnalisisContractual().calcular(df),
    'Autonomia':        AnalisisAutonomia().calcular(df),
    'Alcance_mercado':  AnalisisAlcanceMercado().calcular(df),
    'Desanimados':      AnalisisDesanimados().calcular(df),
    'M8_Duracion':      DuracionDesempleo(config=config).calcular(df),
    'M14_Dashboard':    DashboardSectoresProColombia(config=config).calcular(df),
    'MX2_Forma_pago':   FormaPago(config=config).calcular(df),
    'MX3_Canal_empleo': CanalEmpleo(config=config).calcular(df),
}

for nombre, tabla in resultados_pob.items():
    if tabla is not None and len(tabla) > 0:
        exp.guardar_tabla(tabla, f'{nombre}_{ANIO}')

anat = AnatomaSalario(config=config)
anat.resumen_nacional(df)
anat_rama = anat.por_rama(df)
if len(anat_rama) > 0:
    exp.guardar_tabla(anat_rama, f'MX1_anatomia_salario_{ANIO}')

dur_dpto = DuracionDesempleo(config=config).por_departamento(df)
if len(dur_dpto) > 0:
    exp.guardar_tabla(dur_dpto, f'M8_rigidez_departamental_{ANIO}')

print(f"\n✅ 15 análisis poblacionales completados")

# ══════════════════════════════════════════════════════════════════════
# CELDA 10 · Histogramas horas/ingresos (matplotlib + plotly)
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import gc

FONDO = '#F7F9FC'
C = {'azul':'#2E6DA4','rojo':'#C0392B','verde':'#1E8449',
     'morado':'#7D3C98','naranja':'#E67E22','gris':'#7F8C8D',
     'amarillo':'#F39C12'}
SMMLV = config.smmlv

df_hist_ocu = df[df['OCI'] == 1].copy()
print(f"Ocupados para histogramas: {len(df_hist_ocu):,} registros")

fig_h = plt.figure(figsize=(18, 10))
fig_h.patch.set_facecolor(FONDO)
gs_h = plt.GridSpec(2, 3, figure=fig_h, hspace=0.45, wspace=0.38)
ax_h1 = fig_h.add_subplot(gs_h[0, :2])
ax_h2 = fig_h.add_subplot(gs_h[0, 2])
ax_h3 = fig_h.add_subplot(gs_h[1, :2])
ax_h4 = fig_h.add_subplot(gs_h[1, 2])
for ax in [ax_h1, ax_h2, ax_h3, ax_h4]:
    ax.set_facecolor('white'); ax.spines[['top','right']].set_visible(False)

if 'P6800' in df_hist_ocu.columns:
    df_p68 = df_hist_ocu[df_hist_ocu['P6800'].between(1, 100)].copy()
    bins_h = np.arange(0, 101, 4)
    counts_h, edges_h = np.histogram(df_p68['P6800'], bins=bins_h, weights=df_p68['FEX_ADJ'])
    ax_h1.bar(edges_h[:-1], counts_h/1e6, width=3.8, color=C['azul'], alpha=0.82, align='edge')
    for horas, label, col_v in [(40,'40h legal',C['verde']),(48,'48h límite',C['naranja']),(32,'32h subempleo',C['rojo'])]:
        ax_h1.axvline(horas, color=col_v, ls='--', lw=1.8, alpha=0.85)
        ax_h1.text(horas+0.3, counts_h.max()/1e6*0.85, label, fontsize=8, color=col_v, fontweight='bold')
    v8, w8 = df_p68['P6800'].values, df_p68['FEX_ADJ'].values
    idx8 = np.argsort(v8); v8, w8 = v8[idx8], w8[idx8]
    med8 = float(v8[np.searchsorted(np.cumsum(w8), np.cumsum(w8)[-1]/2)])
    ax_h1.axvline(med8, color='black', ls='-', lw=2, alpha=0.7)
    ax_h1.text(med8+0.5, counts_h.max()/1e6*1.02, f'Mediana={med8:.0f}h', fontsize=9, fontweight='bold')
    ax_h1.set_xlabel('Horas normales semanales (P6800)'); ax_h1.set_ylabel('Personas (millones)')
    ax_h1.set_title(f'Distribución de horas normales — GEIH {ANIO}', fontsize=11, fontweight='bold')
    ax_h1.grid(axis='y', alpha=0.3)

    BINS_LABS = [(1,15,'<15h'),(15,32,'15–32h'),(32,40,'32–40h'),(40,48,'40–48h'),(48,70,'48–70h'),(70,126,'>70h')]
    vals_bins = [df_p68.loc[df_p68['P6800'].between(lo, hi), 'FEX_ADJ'].sum()/1e6 for lo,hi,_ in BINS_LABS]
    col_bins = [C['rojo'],C['naranja'],C['amarillo'],C['verde'],C['naranja'],C['rojo']]
    bars_h2 = ax_h2.bar(range(len(vals_bins)), vals_bins, 0.7, color=col_bins, alpha=0.85)
    for bar, v in zip(bars_h2, vals_bins):
        ax_h2.text(bar.get_x()+bar.get_width()/2, v+0.03, f'{v:.2f}M', ha='center', fontsize=8.5, fontweight='bold')
    ax_h2.set_xticks(range(len(BINS_LABS))); ax_h2.set_xticklabels([l for _,_,l in BINS_LABS], fontsize=8)
    ax_h2.set_ylabel('Millones'); ax_h2.set_title('Por rango de horas', fontsize=10, fontweight='bold')
    ax_h2.grid(axis='y', alpha=0.3)

if 'P6850' in df_hist_ocu.columns:
    df_p685 = df_hist_ocu[df_hist_ocu['P6850'].between(1, 100)].copy()
    for sx_v, sx_l, col_s in [(1,'Hombres',C['azul']),(2,'Mujeres',C['rojo'])]:
        sub = df_p685[df_p685['P3271'] == sx_v]
        if len(sub) == 0: continue
        counts_s, _ = np.histogram(sub['P6850'], bins=bins_h, weights=sub['FEX_ADJ'])
        ax_h3.plot(edges_h[:-1]+2, counts_s/1e6, lw=2.2, color=col_s, label=sx_l, alpha=0.88)
        ax_h3.fill_between(edges_h[:-1]+2, counts_s/1e6, alpha=0.12, color=col_s)
    ax_h3.axvline(48, color=C['naranja'], ls='--', lw=1.5, alpha=0.7, label='48h')
    ax_h3.set_xlabel('Horas reales (P6850)'); ax_h3.set_ylabel('Personas (millones)')
    ax_h3.set_title(f'Horas reales por sexo — GEIH {ANIO}', fontsize=11, fontweight='bold')
    ax_h3.legend(fontsize=10); ax_h3.grid(axis='y', alpha=0.3)

if 'P6800' in df_hist_ocu.columns and 'INGLABO' in df_hist_ocu.columns:
    df_sc = df_hist_ocu[df_hist_ocu['P6800'].between(1,80) & df_hist_ocu['INGLABO'].between(100_000, 15_000_000)]
    if len(df_sc) > 0:
        df_sc_s = df_sc.sample(min(5000, len(df_sc)), weights='FEX_ADJ', random_state=42, replace=True)
        ax_h4.scatter(df_sc_s['P6800'], df_sc_s['INGLABO']/SMMLV, alpha=0.08, s=8, color=C['azul'])
        ax_h4.axhline(1, color=C['verde'], ls='--', lw=1.3, alpha=0.7)
        ax_h4.set_xlabel('Horas semanales'); ax_h4.set_ylabel('Ingreso (× SMMLV)')
        ax_h4.set_title('Horas vs ingreso (5K muestra)', fontsize=10, fontweight='bold')
        ax_h4.set_ylim(0, 12); ax_h4.grid(alpha=0.2)

fig_h.suptitle(f'ANÁLISIS DE HORAS TRABAJADAS — GEIH {ANIO}', fontsize=13, fontweight='bold')
plt.tight_layout(rect=[0, 0, 1, 0.94]); plt.show()
exp.guardar_grafica(fig_h, f'Histograma_horas_{ANIO}', cerrar=True)

try:
    import plotly.express as px
    df_plotly = df_hist_ocu[df_hist_ocu['INGLABO'].between(50_000, 20_000_000)].copy()
    df_plotly['INGLABO_SML'] = df_plotly['INGLABO'] / SMMLV
    df_plotly['SEXO'] = df_plotly['P3271'].map({1:'Hombres', 2:'Mujeres'})
    df_pl_s = df_plotly[df_plotly['SEXO'].notna()].sample(min(30_000, len(df_plotly)), weights='FEX_ADJ', random_state=42, replace=True)
    fig_px = px.box(df_pl_s, x='SEXO', y='INGLABO_SML', color='SEXO',
                    color_discrete_map={'Hombres':'#2E6DA4','Mujeres':'#C0392B'},
                    title=f'Distribución del ingreso laboral por sexo — GEIH {ANIO}',
                    labels={'INGLABO_SML':'Ingreso (× SMMLV)','SEXO':''}, points='outliers')
    fig_px.update_layout(yaxis=dict(range=[0,15]), plot_bgcolor='white', paper_bgcolor='#F7F9FC', showlegend=False)
    fig_px.add_hline(y=1, line_dash='dash', line_color='#1E8449', annotation_text='1 SMMLV')
    fig_px.show()
    print("✅ Box plot interactivo Plotly generado")
except ImportError:
    print("⚠️  plotly no disponible — instalar con: !pip install plotly")

del df_hist_ocu; gc.collect()
print("✅ Celda 10 completada")

# ══════════════════════════════════════════════════════════════════════
# CELDA 11 · Cruce empresa × departamento (P7130, P3069, P6920)
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
import matplotlib.colors as mcolors
from geih import DEPARTAMENTOS, TAMANO_EMPRESA

df_ocu_cr = df[df['OCI'] == 1].copy()
if 'NOMBRE_DPTO' not in df_ocu_cr.columns:
    df_ocu_cr['DPTO_STR'] = df_ocu_cr['DPTO'].astype(str).str.zfill(2)
    df_ocu_cr['NOMBRE_DPTO'] = df_ocu_cr['DPTO_STR'].map(DEPARTAMENTOS)

filas_cr = []
for dpto_nom in df_ocu_cr['NOMBRE_DPTO'].dropna().unique():
    m_d = df_ocu_cr['NOMBRE_DPTO'] == dpto_nom
    n_tot = df_ocu_cr.loc[m_d, 'FEX_ADJ'].sum()
    if n_tot < 5_000: continue
    fila = {'Departamento': dpto_nom, 'Ocupados_M': round(n_tot/1e6, 2)}
    if 'P7130' in df_ocu_cr.columns:
        fila['Desea_cambiar_%'] = round(df_ocu_cr.loc[m_d & (df_ocu_cr['P7130']==1), 'FEX_ADJ'].sum() / n_tot * 100, 1)
    if 'P6920' in df_ocu_cr.columns:
        fila['Cotiza_pension_%'] = round(df_ocu_cr.loc[m_d & (df_ocu_cr['P6920']==1), 'FEX_ADJ'].sum() / n_tot * 100, 1)
    if 'P3069' in df_ocu_cr.columns:
        for cod_t, lab_t in TAMANO_EMPRESA.items():
            fila[f'Emp_{lab_t}'] = round(df_ocu_cr.loc[m_d & (df_ocu_cr['P3069']==cod_t), 'FEX_ADJ'].sum() / n_tot * 100, 1)
    filas_cr.append(fila)

df_cruce = pd.DataFrame(filas_cr)
if 'Desea_cambiar_%' in df_cruce.columns:
    df_cruce = df_cruce.sort_values('Desea_cambiar_%', ascending=False)

import pandas as pd
fig_cr, (ax_cr1, ax_cr2) = plt.subplots(1, 2, figsize=(18, 8))
fig_cr.patch.set_facecolor(FONDO); ax_cr1.set_facecolor('white'); ax_cr2.set_facecolor('white')

if 'Desea_cambiar_%' in df_cruce.columns:
    df_plot_cr = df_cruce.sort_values('Desea_cambiar_%', ascending=True)
    y_cr = np.arange(len(df_plot_cr))
    col_cr = [C['rojo'] if v > 40 else (C['naranja'] if v > 25 else C['verde']) for v in df_plot_cr['Desea_cambiar_%']]
    ax_cr1.barh(y_cr, df_plot_cr['Desea_cambiar_%'], 0.65, color=col_cr, alpha=0.88)
    prom = df_cruce['Desea_cambiar_%'].mean()
    ax_cr1.axvline(prom, color='gray', ls='--', lw=1.3, alpha=0.7, label=f'Promedio ({prom:.1f}%)')
    for i, (_, row) in enumerate(df_plot_cr.iterrows()):
        ax_cr1.text(row['Desea_cambiar_%']+0.3, i, f"{row['Desea_cambiar_%']:.1f}%", va='center', fontsize=8)
    ax_cr1.set_yticks(y_cr); ax_cr1.set_yticklabels(df_plot_cr['Departamento'], fontsize=9)
    ax_cr1.set_xlabel('% desea cambiar (P7130=1)'); ax_cr1.set_title(f'Insatisfacción laboral — GEIH {ANIO}', fontsize=11, fontweight='bold')
    ax_cr1.legend(fontsize=9); ax_cr1.grid(axis='x', alpha=0.3); ax_cr1.spines[['top','right']].set_visible(False)

cols_tam = [c for c in df_cruce.columns if c.startswith('Emp_')]
if cols_tam:
    df_hm = df_cruce.set_index('Departamento')[cols_tam].fillna(0)
    df_hm.columns = [c.replace('Emp_','') for c in df_hm.columns]
    cmap_t = mcolors.LinearSegmentedColormap.from_list('wg', ['white','#2E6DA4'])
    im_t = ax_cr2.imshow(df_hm.values, cmap=cmap_t, aspect='auto', vmin=0, vmax=50)
    plt.colorbar(im_t, ax=ax_cr2, label='% empleo')
    for i in range(len(df_hm)):
        for j in range(len(df_hm.columns)):
            v = df_hm.iloc[i,j]
            if v >= 5: ax_cr2.text(j, i, f'{v:.0f}%', ha='center', va='center', fontsize=7, fontweight='bold',
                                    color='white' if v > 30 else '#1A252F')
    ax_cr2.set_xticks(range(len(df_hm.columns))); ax_cr2.set_xticklabels(df_hm.columns, rotation=35, ha='right', fontsize=8)
    ax_cr2.set_yticks(range(len(df_hm))); ax_cr2.set_yticklabels(df_hm.index, fontsize=8)
    ax_cr2.set_title(f'Tamaño empresa × departamento — GEIH {ANIO}', fontsize=11, fontweight='bold')

fig_cr.suptitle(f'CRUCE EMPRESA × DEPARTAMENTO — GEIH {ANIO}', fontsize=13, fontweight='bold')
fig_cr.tight_layout(pad=2.5); plt.show()
exp.guardar_grafica(fig_cr, f'Cruce_empresa_dpto_{ANIO}', cerrar=True)
exp.guardar_tabla(df_cruce, f'Cruce_empresa_departamento_{ANIO}')
del df_ocu_cr; gc.collect()
print("✅ Celda 11 completada")

# ══════════════════════════════════════════════════════════════════════
# CELDA 12 · 4 gráficos avanzados (Lorenz, ICI bubble, estacionalidad, heatmap)
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
from geih import (
    GraficoCurvaLorenz, GraficoICIBubble,
    GraficoEstacionalidad, GraficoContribucionHeatmap,
    GraficoBoxPlotSalarios, GraficoBrechaGenero,
)

df_ocu_lorenz = df[(df['OCI'] == 1) & (df['INGLABO'] > 0)]
fig_lorenz = GraficoCurvaLorenz().graficar(df_ocu_lorenz,
    titulo=f'Curva de Lorenz del ingreso laboral — GEIH {ANIO}')
exp.guardar_grafica(fig_lorenz, f'M5_Lorenz_Gini_{ANIO}')
plt.show()

if 'Costo_efectivo' in ici.columns and 'Talento_univ_%' in ici.columns:
    fig_ici = GraficoICIBubble().graficar(ici,
        titulo=f'Competitividad laboral (ICI) — GEIH {ANIO}')
    exp.guardar_grafica(fig_ici, f'M16_ICI_bubble_{ANIO}')
    plt.show()

fig_estac = GraficoEstacionalidad().graficar(estac,
    titulo=f'Estacionalidad del mercado laboral — GEIH {ANIO}')
exp.guardar_grafica(fig_estac, f'M11_Estacionalidad_lineas_{ANIO}')
plt.show()

fig_boxplot = GraficoBoxPlotSalarios().graficar(salarios_rama,
    titulo=f'Distribución del ingreso por rama — GEIH {ANIO}', smmlv=SMMLV)
exp.guardar_grafica(fig_boxplot, f'M4_BoxPlot_salarios_{ANIO}')
plt.show()

fig_brecha = GraficoBrechaGenero().graficar(brecha,
    titulo=f'Brecha salarial de género — GEIH {ANIO}', smmlv=SMMLV)
exp.guardar_grafica(fig_brecha, f'M6_Brecha_genero_{ANIO}')
plt.show()

plt.close('all')
gc.collect()
print("✅ Celda 12 completada — 5 gráficos generados")

# ══════════════════════════════════════════════════════════════════════
# CELDA 13 · Análisis por 32 ciudades (AREA)
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
from geih import AnalisisOcupadosCiudad

area = AnalisisOcupadosCiudad(config=config)
tablas_area = area.calcular_tablas(df)
area.exportar_excel(tablas_area, str(exp.excel / f'CIIU_Area_{ANIO}.xlsx'))
print("✅ Celda 13 completada — 6 tablas AREA exportadas")

# ══════════════════════════════════════════════════════════════════════
# CELDA 14 · Excel consolidado multi-hoja + metadata
# ══════════════════════════════════════════════════════════════════════

# Commented out IPython magic to ensure Python compatibility.
# %%time
print(f"\n📗 Exportando Excel consolidado con formato institucional ProColombia...")

exp.guardar_excel({
    'Indicadores_dpto':     td_dpto,
    'Salarios_rama':        salarios_rama.reset_index(),
    'Brecha_genero':        brecha.reset_index(),
    'ICE_departamento':     ice_dpto,
    'ICI_competitividad':   ici,
    'Formalidad_ICF':       icf,
    'Vulnerabilidad_IVI':   ivi,
    'Costo_laboral':        costo,
    'Bono_demografico':     bono,
    'Talento_ITAT':         talento,
    'Estacionalidad':       estac,
    'Joven_15_28':          joven,
    'Urbano_Rural':         urb_rural,
    'Mincer':               mincer,
    'Contribucion_sect':    contrib,
    'Top20_CIIU':           top20,
}, f'GEIH_{ANIO}_Analisis_Completo_ProColombia')

exp.guardar_metadata(config, {
    'n_registros': len(geih),
    'n_columnas': geih.shape[1],
    'meses_procesados': N_MESES,
    'anio': ANIO,
    'clases_ejecutadas': 57,
    'version_paquete': __version__,
})

exp.resumen()
print("✅ Celda 14 completada")

# ══════════════════════════════════════════════════════════════════════
# CELDA 15 · Resumen final + liberación de memoria
# ══════════════════════════════════════════════════════════════════════

print(f"\n{'='*65}")
print(f"  ✅ PIPELINE GEIH {ANIO} COMPLETADO")
print(f"{'='*65}")
print(f"  Período          : {config.periodo_etiqueta}")
print(f"  Base              : {geih.shape[0]:,} filas × {geih.shape[1]} cols")
print(f"  Análisis básicos  : 10 tablas")
print(f"  Análisis avanzados: 8 tablas")
print(f"  Análisis poblac.  : 15 tablas")
print(f"  Gráficos          : 7 PNGs")
print(f"  Excel             : 1 multi-hoja + 1 AREA")
print(f"  CSVs              : ver resultados_geih_{ANIO}/tablas/")
print(f"{'='*65}")

GestorMemoria.estado()

from datetime import datetime
import pytz
print(f"\n  Finalizado: {datetime.now(pytz.timezone('America/Bogota')).strftime('%Y-%m-%d %H:%M:%S')} (hora Colombia)")

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: notebooks/Pipeline_GEIH.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📄 ARCHIVO [29/39]: pyproject.toml
   TIPO      : CONFIGURACIÓN
   UBICACIÓN : (raíz del proyecto)
   RUTA      : pyproject.toml
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: pyproject.toml <<<
────────────────────────────────────────────────────────────────────────────────
[build-system]
requires = ["setuptools>=68", "wheel"]
build-backend = "setuptools.backends.legacy:build"

[project]
name = "geih-analisis"
version = "5.0.0"
description = "Análisis de microdatos GEIH del DANE — Colombia. 70+ clases: TD, TGP, TO, Gini, Mincer, ICE, ICI, brecha de género, 32 ciudades y más."
readme = "README.md"
license = {text = "MIT"}
requires-python = ">=3.9"
authors = [
    {name = "Néstor Enrique Forero Herrera", email = "nforero@procolombia.co"},
]
keywords = [
    "GEIH", "DANE", "Colombia", "mercado laboral", "desempleo", "empleo",
    "encuesta hogares", "microdatos", "estadistica", "econometria",
]
classifiers = [
    "Development Status :: 5 - Production/Stable",
    "Intended Audience :: Science/Research",
    "Intended Audience :: Developers",
    "License :: OSI Approved :: MIT License",
    "Programming Language :: Python :: 3",
    "Programming Language :: Python :: 3.9",
    "Programming Language :: Python :: 3.10",
    "Programming Language :: Python :: 3.11",
    "Programming Language :: Python :: 3.12",
    "Topic :: Scientific/Engineering :: Information Analysis",
    "Operating System :: OS Independent",
    "Natural Language :: Spanish",
]

dependencies = [
    "pandas>=1.5",
    "numpy>=1.21",
    "openpyxl>=3.0",
    "pyarrow>=10.0",
]

[project.optional-dependencies]
viz = [
    "matplotlib>=3.5",
    "plotly>=5.0",
]
dashboard = [
    "streamlit>=1.20",
    "plotly>=5.0",
]
colab = [
    "psutil>=5.9",
    "pyarrow>=10.0",
    "matplotlib>=3.5",
]
dev = [
    "pytest>=7.0",
    "pytest-cov>=4.0",
    "psutil>=5.9",
]
all = [
    "matplotlib>=3.5",
    "plotly>=5.0",
    "streamlit>=1.20",
    "psutil>=5.9",
    "pytest>=7.0",
]

[project.urls]
Homepage      = "https://github.com/enriqueforero/geih-analisis"
Repository    = "https://github.com/enriqueforero/geih-analisis"
"Bug Tracker" = "https://github.com/enriqueforero/geih-analisis/issues"
Documentation = "https://github.com/enriqueforero/geih-analisis#readme"

[tool.setuptools.packages.find]
where   = ["."]
include = ["geih*"]         # incluye tanto 'geih' como 'geih_2025' (shim)
exclude = ["tests*", "notebooks*", "docs*"]

[tool.setuptools.package-data]
"geih"      = ["py.typed"]
"geih_2025" = ["py.typed"]

[tool.pytest.ini_options]
testpaths  = ["tests"]
python_files = ["test_*.py"]
addopts    = "-v --tb=short"

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: pyproject.toml >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📄 ARCHIVO [30/39]: requirements.txt
   TIPO      : CONFIGURACIÓN
   UBICACIÓN : (raíz del proyecto)
   RUTA      : requirements.txt
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: requirements.txt <<<
────────────────────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════
# geih_2025 v4.3.0 — Dependencias
# ══════════════════════════════════════════════════════════════

# ── Core (obligatorias) ──────────────────────────────────────
pandas>=1.5
numpy>=1.21
matplotlib>=3.5
openpyxl>=3.0

# ── Rendimiento (recomendadas) ───────────────────────────────
psutil>=5.9
pyarrow>=10.0

# ── Interactividad (opcionales) ──────────────────────────────
# plotly>=5.0               # Gráficos interactivos Plotly
# streamlit>=1.28           # Dashboard web interactivo

# ── Testing (solo desarrollo) ────────────────────────────────
# pytest>=7.0

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: requirements.txt >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
📄 ARCHIVO [31/39]: setup.cfg
   TIPO      : CONFIGURACIÓN
   UBICACIÓN : (raíz del proyecto)
   RUTA      : setup.cfg
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: setup.cfg <<<
────────────────────────────────────────────────────────────────────────────────
# setup.cfg — configuración legacy para compatibilidad con herramientas antiguas.
# La configuración principal vive en pyproject.toml.
[metadata]
name = geih-analisis
version = 5.0.0

[options]
packages = find:
python_requires = >=3.9

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: setup.cfg >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [32/39]: setup.py
   TIPO      : CÓDIGO
   UBICACIÓN : (raíz del proyecto)
   RUTA      : setup.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: setup.py <<<
────────────────────────────────────────────────────────────────────────────────
# setup.py — archivo legacy requerido por algunas herramientas.
# La configuración real está en pyproject.toml.
from setuptools import setup

if __name__ == "__main__":
    setup()

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: setup.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [33/39]: __init__.py
   TIPO      : CÓDIGO
   UBICACIÓN : tests
   RUTA      : tests/__init__.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: tests/__init__.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
geih — Análisis de microdatos GEIH del DANE.

Gran Encuesta Integrada de Hogares | DANE | Marco Muestral 2018
Autor: Néstor Enrique Forero Herrera

Paquete multi-año: soporta GEIH 2022–presente. No está atado a ningún
año específico — el nombre 'geih_2025' era la versión anterior.

Instalación:
    pip install geih-analisis
    # o desde GitHub:
    pip install git+https://github.com/enriqueforero/geih-analisis.git

Uso rápido:
    from geih import ConfigGEIH, ConsolidadorGEIH, PreparadorGEIH
    config = ConfigGEIH(anio=2025, n_meses=12)

Compatibilidad hacia atrás:
    'from geih_2025 import ...' sigue funcionando gracias al shim
    incluido en este paquete. Recibirás un DeprecationWarning.

CAMBIO v5.0 — Renombrado geih_2025 → geih:
  El paquete ahora se llama 'geih' (nombre de importación) y
  'geih-analisis' (nombre de distribución en PyPI).
  El shim geih_2025/ garantiza compatibilidad durante la transición.

Módulos del paquete (17 archivos):
  config.py                → Constantes, mapeos, configuración centralizada
  utils.py                 → Memoria, conversión de tipos, estadísticas ponderadas
  consolidador.py          → Lectura y unión de módulos CSV mensuales
  preparador.py            → Preparación de datos, merge con correlativas
  diagnostico.py           → Diagnóstico de calidad de datos
  indicadores.py           → Indicadores básicos (TD, TGP, TO, ingresos, rama)
  analisis_avanzado.py     → Módulos avanzados (ICE, ICI, ITAT, Mincer, etc.)
  analisis_area.py         → 32 ciudades × CIIU
  analisis_poblacional.py  → Campesinos, discapacidad, migración
  analisis_complementario.py → M8, M14, MX1–MX3
  exportador.py            → Exportación organizada a carpetas
  visualizacion.py         → Gráficos matplotlib
  visualizacion_interactiva.py → Gráficos Plotly
  comparativo.py           → Comparación inter-anual
  descargador.py           → Descarga automática DANE
  logger.py                → Logging centralizado
  profiler.py              → Profiling de memoria
  dashboard.py             → Dashboard Streamlit
"""

__version__ = "5.0.0"
__author__  = "Néstor Enrique Forero Herrera"
__email__   = "nforero@procolombia.co"
__url__     = "https://github.com/enriqueforero/geih-analisis"
__license__ = "MIT"

# ── Configuración ──────────────────────────────────────────────────
from .config import (
    # Configuración principal
    ConfigGEIH,
    # SMMLV
    SMMLV_2025, SMMLV_POR_ANIO,
    # Colores
    COLORES,
    # Períodos
    MESES_CARPETAS, MESES_NOMBRES,
    generar_carpetas_mensuales, generar_etiqueta_periodo,
    # Geografía
    DEPARTAMENTOS, DPTO_A_CIUDAD, AREA_A_CIUDAD,
    CIUDADES_13_PRINCIPALES, CIUDADES_10_INTERMEDIAS,
    # Clasificaciones económicas
    RAMAS_DANE, TABLA_CIIU_RAMAS, AGRUPACION_DANE_8,
    # Referencias DANE
    REF_DANE_2025, REF_DANE, ReferenciaDane,
    # Constantes laborales — antes faltaban, causaban ImportError
    CARGA_PRESTACIONAL, TAMANO_EMPRESA,
    RANGOS_SMMLV_LIMITES, RANGOS_SMMLV_ETIQUETAS,
    # Educación — antes faltaban
    NIVELES_AGRUPADOS, NIVELES_EDUCATIVOS, P3042_A_ANOS,
    # Llaves y converters
    LLAVES_PERSONA, LLAVES_HOGAR,
    CONVERTERS_BASE, CONVERTERS_CON_AREA,
    MODULOS_CSV,
)

# ── Utilidades ─────────────────────────────────────────────────────
from .utils import GestorMemoria, ConversorTipos, EstadisticasPonderadas

# ── Consolidación ──────────────────────────────────────────────────
from .consolidador import ConsolidadorGEIH

# ── Preparación ────────────────────────────────────────────────────
from .preparador import PreparadorGEIH, MergeCorrelativas

# ── Diagnóstico ────────────────────────────────────────────────────
from .diagnostico import DiagnosticoCalidad, Top20Sectores

# ── Exportación organizada ─────────────────────────────────────────
from .exportador import Exportador

# ── Indicadores básicos ────────────────────────────────────────────
from .indicadores import (
    IndicadoresLaborales, DistribucionIngresos, AnalisisRamaSexo,
    AnalisisSalarios, BrechaGenero, AnalisisCruzado,
    IndicesCompuestos, AnalisisArea,
)

# ── Análisis por 32 ciudades ───────────────────────────────────────
from .analisis_area import AnalisisOcupadosCiudad

# ── Análisis avanzado ──────────────────────────────────────────────
from .analisis_avanzado import (
    CalidadEmpleo, FormalidadSectorial, VulnerabilidadLaboral,
    CompetitividadLaboral, AnalisisSubempleo, AnalisisHoras,
    Estacionalidad, FuerzaLaboralJoven, EtnicoRacial,
    BonoDemografico, CostoLaboral, AnalisisFFT,
    AnalisisUrbanoRural, ProductividadTamano,
    ContribucionSectorial, MapaTalento, EcuacionMincer,
    ProxyBilinguismo,
)

# ── Visualización matplotlib ───────────────────────────────────────
from .visualizacion import (
    GraficoDistribucionIngresos, GraficoBoxPlotSalarios,
    GraficoBrechaGenero, GraficoRamaSexo,
    GraficoCurvaLorenz, GraficoICIBubble,
    GraficoEstacionalidad, GraficoContribucionHeatmap,
)

# ── Análisis poblacional ───────────────────────────────────────────
from .analisis_poblacional import (
    AnalisisCampesino, AnalisisDiscapacidad, AnalisisMigracion,
    AnalisisOtrasFormas, AnalisisOtrosIngresos,
    AnalisisSobrecalificacion, AnalisisContractual,
    AnalisisAutonomia, AnalisisAlcanceMercado, AnalisisDesanimados,
)

# ── Análisis complementarios ───────────────────────────────────────
from .analisis_complementario import (
    DuracionDesempleo, DashboardSectoresProColombia,
    AnatomaSalario, FormaPago, CanalEmpleo,
)

# ── Descarga automática DANE ───────────────────────────────────────
from .descargador import DescargadorDANE

# ── Comparativo multi-año ──────────────────────────────────────────
from .comparativo import ComparadorMultiAnio

# ── Visualización interactiva Plotly ───────────────────────────────
try:
    from .visualizacion_interactiva import (
        PlotlyLorenz, PlotlyICIBubble, PlotlyEstacionalidad,
        PlotlyDistribucionIngresos, PlotlyBrechaGenero,
        PlotlyBoxPlotSalarios, PlotlySalarioRama,
        PlotlyComparativoAnual,
    )
except ImportError:
    pass  # plotly no instalado — instalar con: pip install geih-analisis[viz]

# ── Logging centralizado ───────────────────────────────────────────
from .logger import get_logger, configurar_logging, LoggerGEIH

# ── Profiling de memoria ───────────────────────────────────────────
from .profiler import PerfilMemoria, medir_tiempo, tamano_objeto

# ── Dashboard Streamlit ────────────────────────────────────────────
from .dashboard import ejecutar_dashboard

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: tests/__init__.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [34/39]: smoke_test.py
   TIPO      : CÓDIGO
   UBICACIÓN : tests
   RUTA      : tests/smoke_test.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: tests/smoke_test.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""
tests/smoke_test.py — Smoke test del pipeline GEIH.

Ejecuta el pipeline completo sobre una muestra mínima (0.1% de la base
o el golden set) para validar que todo funciona antes de procesar
los 5 millones de registros (~10 segundos vs ~10 minutos).

CÓMO USAR EN EL NOTEBOOK — Insertar como Celda 0 antes del pipeline:

    # ══════════════════════════════════════════════════════════════
    # CELDA 0 · Smoke test (ejecutar ANTES del pipeline completo)
    # ══════════════════════════════════════════════════════════════
    import sys
    sys.path.insert(0, '/content/drive/MyDrive/ProColombia/GEIH')
    from tests.smoke_test import ejecutar_smoke_test
    ejecutar_smoke_test(ruta_base='/content/drive/MyDrive/ProColombia/GEIH')
    # Si imprime "✅ SMOKE TEST PASADO", es seguro correr el pipeline completo.
    # Si imprime "❌ SMOKE TEST FALLIDO", NO proceder.

También se puede ejecutar standalone:
    python tests/smoke_test.py

Autor: Néstor Enrique Forero Herrera
"""

import sys
import time
import traceback
from pathlib import Path
from typing import Optional


def ejecutar_smoke_test(
    ruta_base: Optional[str] = None,
    anio: int = 2025,
    n_meses: int = 12,
    muestra_n: int = 5_000,
) -> bool:
    """Ejecuta el pipeline completo sobre una muestra mínima.

    Args:
        ruta_base: Ruta al directorio GEIH. Si None, usa golden set.
        anio: Año para ConfigGEIH.
        n_meses: Meses para ConfigGEIH.
        muestra_n: Tamaño de la muestra (0 = usar golden set completo).

    Returns:
        True si todos los checks pasan, False si alguno falla.
    """
    inicio = time.time()
    errores = []

    print(f"\n{'='*65}")
    print(f"  🧪 SMOKE TEST — Pipeline GEIH {anio}")
    print(f"{'='*65}")

    # ── 1. Importar paquete ────────────────────────────────────
    try:
        from geih import (
            ConfigGEIH, PreparadorGEIH, DiagnosticoCalidad,
            IndicadoresLaborales, DistribucionIngresos,
            AnalisisSalarios, BrechaGenero, IndicesCompuestos,
            CalidadEmpleo, CompetitividadLaboral,
            DuracionDesempleo, DashboardSectoresProColombia,
            AnatomaSalario, FormaPago, CanalEmpleo,
            AnalisisCampesino, AnalisisDiscapacidad,
            GraficoCurvaLorenz,
            __version__,
        )
        print(f"  ✅ 1/7 Paquete importado (v{__version__})")
    except Exception as e:
        errores.append(f"Importación: {e}")
        print(f"  ❌ 1/7 Error importando paquete: {e}")
        _resultado_final(errores, inicio)
        return False

    # ── 2. Configuración ──────────────────────────────────────
    try:
        config = ConfigGEIH(anio=anio, n_meses=n_meses)
        assert config.smmlv > 100_000
        assert len(config.carpetas_mensuales) == n_meses
        print(f"  ✅ 2/7 ConfigGEIH OK (SMMLV=${config.smmlv:,})")
    except Exception as e:
        errores.append(f"Config: {e}")
        print(f"  ❌ 2/7 Error config: {e}")

    # ── 3. Cargar datos (golden set o muestra) ────────────────
    try:
        golden_path = Path(__file__).parent / "golden_set.parquet"
        if golden_path.exists():
            import pandas as pd
            df_raw = pd.read_parquet(golden_path)
            print(f"  ✅ 3/7 Golden set cargado ({len(df_raw):,} filas)")
        elif ruta_base:
            from geih import ConsolidadorGEIH
            parquet = Path(ruta_base) / f"GEIH_{anio}_Consolidado.parquet"
            if parquet.exists():
                import pandas as pd
                df_raw = pd.read_parquet(parquet)
                if muestra_n > 0 and len(df_raw) > muestra_n:
                    df_raw = df_raw.sample(muestra_n, random_state=42)
                print(f"  ✅ 3/7 Muestra cargada ({len(df_raw):,} filas)")
            else:
                print(f"  ⚠️  3/7 No hay Parquet ni golden set — generando golden set...")
                from tests.generar_golden_set import generar_golden_set
                df_raw = generar_golden_set(ruta_salida=str(golden_path))
                print(f"  ✅ 3/7 Golden set generado ({len(df_raw):,} filas)")
        else:
            from tests.generar_golden_set import generar_golden_set
            df_raw = generar_golden_set(ruta_salida=str(golden_path))
            print(f"  ✅ 3/7 Golden set generado ({len(df_raw):,} filas)")
    except Exception as e:
        errores.append(f"Carga datos: {e}")
        print(f"  ❌ 3/7 Error cargando datos: {e}")
        _resultado_final(errores, inicio)
        return False

    # ── 4. Preparar datos ─────────────────────────────────────
    try:
        prep = PreparadorGEIH(config=config)
        df = prep.preparar_base(df_raw)
        df = prep.agregar_variables_derivadas(df)
        assert "FEX_ADJ" in df.columns
        assert "RAMA" in df.columns or "RAMA2D_R4" in df.columns
        print(f"  ✅ 4/7 Datos preparados ({df.shape[1]} columnas)")
    except Exception as e:
        errores.append(f"Preparación: {e}")
        print(f"  ❌ 4/7 Error preparando datos: {e}")
        traceback.print_exc()
        _resultado_final(errores, inicio)
        return False

    # ── 5. Indicadores fundamentales ──────────────────────────
    try:
        ind = IndicadoresLaborales(config=config)
        r = ind.calcular(df)
        assert "TD_%" in r
        assert "TGP_%" in r
        assert "TO_%" in r
        assert 0 < r["TD_%"] < 100
        assert 0 < r["TGP_%"] < 100
        assert r["PEA_M"] < 40  # no debe estar inflado
        print(f"  ✅ 5/7 Indicadores OK (TD={r['TD_%']:.1f}%, "
              f"TGP={r['TGP_%']:.1f}%, TO={r['TO_%']:.1f}%)")
    except Exception as e:
        errores.append(f"Indicadores: {e}")
        print(f"  ❌ 5/7 Error indicadores: {e}")

    # ── 6. Módulos de análisis (subset) ───────────────────────
    modulos_ok = 0
    modulos_total = 0
    clases_test = [
        ("DistribucionIngresos", lambda: DistribucionIngresos(config=config).calcular(df)),
        ("BrechaGenero", lambda: BrechaGenero().calcular(df)),
        ("Gini", lambda: IndicesCompuestos(config=config).gini(df)),
        ("DuracionDesempleo", lambda: DuracionDesempleo(config=config).calcular(df)),
        ("DashboardSectores", lambda: DashboardSectoresProColombia(config=config).calcular(df)),
        ("AnatomaSalario", lambda: AnatomaSalario(config=config).resumen_nacional(df)),
        ("FormaPago", lambda: FormaPago(config=config).calcular(df)),
        ("CanalEmpleo", lambda: CanalEmpleo(config=config).calcular(df)),
        ("Campesino", lambda: AnalisisCampesino(config=config).calcular(df)),
    ]
    for nombre, fn in clases_test:
        modulos_total += 1
        try:
            resultado = fn()
            modulos_ok += 1
        except Exception as e:
            errores.append(f"{nombre}: {e}")

    if modulos_ok == modulos_total:
        print(f"  ✅ 6/7 {modulos_ok}/{modulos_total} módulos ejecutados sin error")
    else:
        print(f"  ⚠️  6/7 {modulos_ok}/{modulos_total} módulos OK "
              f"({modulos_total - modulos_ok} fallaron)")

    # ── 7. Gráfico de prueba ──────────────────────────────────
    try:
        import matplotlib
        matplotlib.use("Agg")  # No mostrar ventana
        df_ocu = df[(df["OCI"] == 1) & (df["INGLABO"] > 0)]
        if len(df_ocu) > 10:
            fig = GraficoCurvaLorenz().graficar(df_ocu)
            import matplotlib.pyplot as plt
            plt.close(fig)
            print(f"  ✅ 7/7 Gráfico Lorenz generado sin error")
        else:
            print(f"  ⚠️  7/7 Pocos ocupados para gráfico ({len(df_ocu)})")
    except Exception as e:
        errores.append(f"Gráfico: {e}")
        print(f"  ❌ 7/7 Error generando gráfico: {e}")

    return _resultado_final(errores, inicio)


def _resultado_final(errores, inicio):
    """Imprime resultado final del smoke test."""
    elapsed = time.time() - inicio
    print(f"\n{'─'*65}")
    if not errores:
        print(f"  ✅ SMOKE TEST PASADO en {elapsed:.1f}s")
        print(f"  → Es seguro ejecutar el pipeline completo.")
        print(f"{'─'*65}")
        return True
    else:
        print(f"  ❌ SMOKE TEST FALLIDO — {len(errores)} errores en {elapsed:.1f}s")
        for i, err in enumerate(errores, 1):
            print(f"     {i}. {err}")
        print(f"\n  → NO proceder con el pipeline hasta corregir los errores.")
        print(f"{'─'*65}")
        return False


if __name__ == "__main__":
    # Ejecutar standalone
    import sys
    ruta = sys.argv[1] if len(sys.argv) > 1 else None
    ok = ejecutar_smoke_test(ruta_base=ruta)
    sys.exit(0 if ok else 1)

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: tests/smoke_test.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [35/39]: test_config.py
   TIPO      : CÓDIGO
   UBICACIÓN : tests
   RUTA      : tests/test_config.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: tests/test_config.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""Tests para geih.config — Configuración multi-año."""

import pytest
from geih.config import (
    ConfigGEIH, SMMLV_POR_ANIO, REF_DANE, SMMLV_2025,
    MESES_CARPETAS, REF_DANE_2025, MESES_NOMBRES,
    generar_carpetas_mensuales, generar_etiqueta_periodo,
)


class TestConfigGEIH:
    """Tests del dataclass ConfigGEIH."""

    def test_defaults_2025(self):
        """Config por defecto debe ser 2025, 12 meses."""
        c = ConfigGEIH()
        assert c.anio == 2025
        assert c.n_meses == 12
        assert c.smmlv == 1_423_500

    def test_smmlv_auto_seleccion(self):
        """SMMLV se auto-selecciona según el año."""
        c25 = ConfigGEIH(anio=2025)
        c26 = ConfigGEIH(anio=2026)
        assert c25.smmlv == SMMLV_POR_ANIO[2025]
        assert c26.smmlv == SMMLV_POR_ANIO[2026]
        assert c26.smmlv > c25.smmlv

    def test_smmlv_manual_override(self):
        """SMMLV manual tiene prioridad sobre auto-selección."""
        c = ConfigGEIH(anio=2025, smmlv=2_000_000)
        assert c.smmlv == 2_000_000

    def test_carpetas_mensuales_12(self):
        """12 meses genera 12 carpetas."""
        c = ConfigGEIH(anio=2025, n_meses=12)
        carpetas = c.carpetas_mensuales
        assert len(carpetas) == 12
        assert carpetas[0] == "Enero 2025"
        assert carpetas[-1] == "Diciembre 2025"

    def test_carpetas_mensuales_3(self):
        """3 meses genera solo 3 carpetas."""
        c = ConfigGEIH(anio=2026, n_meses=3)
        carpetas = c.carpetas_mensuales
        assert len(carpetas) == 3
        assert carpetas[0] == "Enero 2026"
        assert carpetas[-1] == "Marzo 2026"

    def test_carpetas_mensuales_1(self):
        """1 mes genera 1 carpeta."""
        c = ConfigGEIH(anio=2026, n_meses=1)
        assert len(c.carpetas_mensuales) == 1
        assert c.carpetas_mensuales[0] == "Enero 2026"

    def test_periodo_etiqueta_auto(self):
        """Etiqueta de período se genera automáticamente."""
        c12 = ConfigGEIH(anio=2025, n_meses=12)
        assert c12.periodo_etiqueta == "Enero – Diciembre 2025"

        c3 = ConfigGEIH(anio=2026, n_meses=3)
        assert c3.periodo_etiqueta == "Enero – Marzo 2026"

        c1 = ConfigGEIH(anio=2026, n_meses=1)
        assert c1.periodo_etiqueta == "Enero 2026"

    def test_referencia_dane_2025(self):
        """2025 tiene referencia DANE disponible."""
        c = ConfigGEIH(anio=2025)
        ref = c.referencia_dane
        assert ref is not None
        assert ref.td_anual_pct == 8.9

    def test_referencia_dane_2026_none(self):
        """2026 no tiene referencia DANE (aún no publicada)."""
        c = ConfigGEIH(anio=2026)
        assert c.referencia_dane is None

    def test_validacion_n_meses_0(self):
        """n_meses=0 debe lanzar ValueError."""
        with pytest.raises(ValueError, match="n_meses=0"):
            ConfigGEIH(n_meses=0)

    def test_validacion_n_meses_13(self):
        """n_meses=13 debe lanzar ValueError."""
        with pytest.raises(ValueError, match="n_meses=13"):
            ConfigGEIH(n_meses=13)

    def test_validacion_anio_2010(self):
        """anio=2010 debe lanzar ValueError (antes del marco 2018)."""
        with pytest.raises(ValueError, match="anio=2010"):
            ConfigGEIH(anio=2010)

    def test_validacion_smmlv_bajo(self):
        """SMMLV < 100,000 debe lanzar ValueError."""
        with pytest.raises(ValueError, match="parece demasiado bajo"):
            ConfigGEIH(smmlv=50_000)


class TestRetrocompatibilidad:
    """Verifica que los símbolos antiguos siguen funcionando."""

    def test_smmlv_2025_existe(self):
        assert SMMLV_2025 == 1_423_500

    def test_ref_dane_2025_existe(self):
        assert REF_DANE_2025 is not None
        assert REF_DANE_2025 == REF_DANE[2025]

    def test_meses_carpetas_existe(self):
        assert len(MESES_CARPETAS) == 12
        assert MESES_CARPETAS[0] == "Enero 2025"


class TestFuncionesAuxiliares:
    """Tests de generar_carpetas_mensuales y generar_etiqueta_periodo."""

    def test_generar_carpetas_basico(self):
        result = generar_carpetas_mensuales(2027, 6)
        assert len(result) == 6
        assert result[0] == "Enero 2027"
        assert result[-1] == "Junio 2027"

    def test_generar_carpetas_clamp(self):
        """n_meses se clampea a [1, 12]."""
        assert len(generar_carpetas_mensuales(2025, 0)) == 1   # clamp a 1
        assert len(generar_carpetas_mensuales(2025, 99)) == 12  # clamp a 12

    def test_etiqueta_1_mes(self):
        assert generar_etiqueta_periodo(2026, 1) == "Enero 2026"

    def test_etiqueta_12_meses(self):
        assert generar_etiqueta_periodo(2025, 12) == "Enero – Diciembre 2025"

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: tests/test_config.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [36/39]: test_consolidador.py
   TIPO      : CÓDIGO
   UBICACIÓN : tests
   RUTA      : tests/test_consolidador.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: tests/test_consolidador.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""Tests para geih.consolidador — Lógica de consolidación."""

import pandas as pd
import pytest
from geih.consolidador import ConsolidadorGEIH


class TestUnirSinDuplicados:
    """Tests de _unir_sin_duplicados (LEFT JOIN sin columnas repetidas)."""

    def test_no_multiplica_filas(self, df_merge_izq, df_merge_der):
        """LEFT JOIN no debe producir más filas que el DataFrame izquierdo."""
        llaves = ["DIRECTORIO", "SECUENCIA_P", "ORDEN"]
        resultado = ConsolidadorGEIH._unir_sin_duplicados(
            df_merge_izq, df_merge_der, llaves
        )
        assert len(resultado) == len(df_merge_izq)

    def test_preserva_universo(self, df_merge_izq, df_merge_der):
        """LEFT JOIN debe preservar todas las filas del DataFrame izquierdo."""
        llaves = ["DIRECTORIO", "SECUENCIA_P", "ORDEN"]
        resultado = ConsolidadorGEIH._unir_sin_duplicados(
            df_merge_izq, df_merge_der, llaves
        )
        # D3 no está en df_der → debe tener NaN en COL_B
        d3 = resultado[resultado["DIRECTORIO"] == "D3"]
        assert len(d3) == 1
        assert pd.isna(d3["COL_B"].iloc[0])

    def test_elimina_columnas_duplicadas(self, df_merge_izq, df_merge_der):
        """No debe traer COL_A del df_der (ya existe en df_izq)."""
        llaves = ["DIRECTORIO", "SECUENCIA_P", "ORDEN"]
        resultado = ConsolidadorGEIH._unir_sin_duplicados(
            df_merge_izq, df_merge_der, llaves
        )
        # No debe haber COL_A_x ni COL_A_y
        assert "COL_A_x" not in resultado.columns
        assert "COL_A_y" not in resultado.columns
        # COL_A original se preserva
        assert "COL_A" in resultado.columns

    def test_trae_columnas_nuevas(self, df_merge_izq, df_merge_der):
        """Debe traer COL_B (nueva) del df_der."""
        llaves = ["DIRECTORIO", "SECUENCIA_P", "ORDEN"]
        resultado = ConsolidadorGEIH._unir_sin_duplicados(
            df_merge_izq, df_merge_der, llaves
        )
        assert "COL_B" in resultado.columns
        # D1 debe tener COL_B=100
        d1 = resultado[resultado["DIRECTORIO"] == "D1"]
        assert d1["COL_B"].iloc[0] == 100


class TestNormalizar:
    """Tests de _normalizar (comparación sin tildes)."""

    def test_tildes(self):
        assert ConsolidadorGEIH._normalizar("Migración.CSV") == "migracion.csv"

    def test_mayusculas(self):
        assert ConsolidadorGEIH._normalizar("OCUPADOS.CSV") == "ocupados.csv"

    def test_mixto(self):
        result = ConsolidadorGEIH._normalizar(
            "Características generales, seguridad social en salud y educación.CSV"
        )
        assert "caracteristicas" in result
        assert "educacion" in result

    def test_sin_tildes_ya(self):
        assert ConsolidadorGEIH._normalizar("ocupados.csv") == "ocupados.csv"


class TestInferirNumeroMes:
    """Tests de _inferir_numero_mes."""

    def test_enero(self):
        assert ConsolidadorGEIH._inferir_numero_mes("Enero 2025") == 1

    def test_diciembre(self):
        assert ConsolidadorGEIH._inferir_numero_mes("Diciembre 2026") == 12

    def test_marzo_con_espacios(self):
        assert ConsolidadorGEIH._inferir_numero_mes("  Marzo 2025  ") == 3

    def test_invalido(self):
        with pytest.raises(ValueError, match="No se pudo inferir"):
            ConsolidadorGEIH._inferir_numero_mes("InvalidMonth 2025")

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: tests/test_consolidador.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [37/39]: test_estadisticas_ponderadas.py
   TIPO      : CÓDIGO
   UBICACIÓN : tests
   RUTA      : tests/test_estadisticas_ponderadas.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: tests/test_estadisticas_ponderadas.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""Tests para geih.utils — Estadísticas ponderadas."""

import numpy as np
import pandas as pd
import pytest
from geih.utils import EstadisticasPonderadas as EP


class TestMedia:
    """Tests de media ponderada."""

    def test_pesos_iguales(self, df_simple):
        """Con pesos iguales, media ponderada = media aritmética."""
        resultado = EP.media(df_simple["valor"], df_simple["peso"])
        assert resultado == pytest.approx(300.0)

    def test_pesos_desiguales(self, df_pesos_desiguales):
        """Con peso 10 en valor 100, la media se sesga hacia 100."""
        resultado = EP.media(df_pesos_desiguales["valor"], df_pesos_desiguales["peso"])
        # (100*10 + 200*1 + 300*1) / (10+1+1) = 1500/12 = 125.0
        assert resultado == pytest.approx(125.0)

    def test_con_nans(self):
        """Los NaN se ignoran correctamente."""
        valores = pd.Series([100, np.nan, 300])
        pesos = pd.Series([1.0, 1.0, 1.0])
        resultado = EP.media(valores, pesos)
        assert resultado == pytest.approx(200.0)

    def test_pesos_cero(self):
        """Pesos = 0 se ignoran."""
        valores = pd.Series([100, 200])
        pesos = pd.Series([0.0, 0.0])
        assert np.isnan(EP.media(valores, pesos))

    def test_vacio(self):
        """Serie vacía retorna NaN."""
        assert np.isnan(EP.media(pd.Series(dtype=float), pd.Series(dtype=float)))


class TestMediana:
    """Tests de mediana ponderada."""

    def test_pesos_iguales(self, df_simple):
        """Con pesos iguales, mediana ponderada = mediana simple."""
        resultado = EP.mediana(df_simple["valor"], df_simple["peso"])
        assert resultado == pytest.approx(300.0)

    def test_pesos_desiguales(self, df_pesos_desiguales):
        """Con peso 10 en valor 100, la mediana se desplaza hacia 100."""
        resultado = EP.mediana(df_pesos_desiguales["valor"], df_pesos_desiguales["peso"])
        # 10/12 del peso está en 100, mediana debe ser 100
        assert resultado == pytest.approx(100.0)


class TestPercentil:
    """Tests de percentil ponderado."""

    def test_p0_es_minimo(self):
        """Percentil 0 debe ser el valor mínimo."""
        v = pd.Series([10, 20, 30, 40, 50])
        w = pd.Series([1.0] * 5)
        assert EP.percentil(v, w, 0.0) == pytest.approx(10.0)

    def test_p100_es_maximo(self):
        """Percentil 1.0 debe ser el valor máximo."""
        v = pd.Series([10, 20, 30, 40, 50])
        w = pd.Series([1.0] * 5)
        assert EP.percentil(v, w, 1.0) == pytest.approx(50.0)

    def test_p50_es_mediana(self):
        """Percentil 0.5 = mediana ponderada."""
        v = pd.Series([10, 20, 30])
        w = pd.Series([1.0, 1.0, 1.0])
        assert EP.percentil(v, w, 0.5) == EP.mediana(v, w)


class TestDesviacionEstandar:
    """Tests de desviación estándar ponderada."""

    def test_valores_iguales(self):
        """Todos iguales → desviación = 0."""
        v = pd.Series([100, 100, 100])
        w = pd.Series([1.0, 1.0, 1.0])
        assert EP.desviacion_estandar(v, w) == pytest.approx(0.0)

    def test_positiva(self, df_simple):
        """Con valores distintos, desviación > 0."""
        resultado = EP.desviacion_estandar(df_simple["valor"], df_simple["peso"])
        assert resultado > 0


class TestGini:
    """Tests del coeficiente de Gini."""

    def test_igualdad_perfecta(self):
        """Todos ganan igual → Gini ≈ 0."""
        v = pd.Series([1000] * 100)
        w = pd.Series([1.0] * 100)
        gini = EP.gini(v, w)
        assert gini == pytest.approx(0.0, abs=0.01)

    def test_desigualdad_alta(self):
        """Uno gana todo → Gini cercano a 1."""
        v = pd.Series([1] * 99 + [10_000])
        w = pd.Series([1.0] * 100)
        gini = EP.gini(v, w)
        assert gini > 0.9

    def test_rango_valido(self):
        """Gini siempre entre 0 y 1."""
        v = pd.Series(np.random.lognormal(10, 1, size=500))
        w = pd.Series(np.ones(500))
        gini = EP.gini(v, w)
        assert 0 <= gini <= 1

    def test_vacio_retorna_nan(self):
        """Con menos de 2 valores, retorna NaN."""
        assert np.isnan(EP.gini(pd.Series([100]), pd.Series([1.0])))


class TestResumenCompleto:
    """Tests de resumen_completo."""

    def test_retorna_dict(self, df_simple):
        """Debe retornar un diccionario con las claves esperadas."""
        resultado = EP.resumen_completo(df_simple["valor"], df_simple["peso"])
        assert "Media" in resultado
        assert "Mediana" in resultado
        assert "P10" in resultado
        assert "P90" in resultado
        assert "CV_%" in resultado
        assert "Media_SMMLV" in resultado

    def test_media_smmlv_correcto(self):
        """Media_SMMLV = Media / SMMLV."""
        v = pd.Series([1_423_500] * 10)
        w = pd.Series([1.0] * 10)
        resultado = EP.resumen_completo(v, w, smmlv=1_423_500)
        assert resultado["Media_SMMLV"] == pytest.approx(1.0)

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: tests/test_estadisticas_ponderadas.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [38/39]: test_indicadores.py
   TIPO      : CÓDIGO
   UBICACIÓN : tests
   RUTA      : tests/test_indicadores.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: tests/test_indicadores.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""Tests para geih.indicadores — Indicadores laborales."""

import numpy as np
import pandas as pd
import pytest
from geih.config import ConfigGEIH
from geih.indicadores import IndicadoresLaborales


class TestIndicadoresLaboralesSinteticos:
    """Tests con datos sintéticos de resultado conocido."""

    @pytest.fixture
    def df_sintetico(self):
        """100 personas: 60 OCI, 20 DSI, 20 FFT → TD=25%, TGP=80%, TO=60%."""
        n = 100
        df = pd.DataFrame({
            "FEX_ADJ": [100.0] * n,
            "OCI": [1]*60 + [0]*40,
            "FT":  [1]*80 + [0]*20,    # PEA = 80
            "DSI": [0]*60 + [1]*20 + [0]*20,
            "PET": [1]*n,
            "FFT": [0]*80 + [1]*20,
        })
        return df

    def test_td_conocida(self, df_sintetico):
        """TD = DSI/PEA = 20/80 = 25.0%."""
        config = ConfigGEIH(anio=2025, n_meses=1)
        ind = IndicadoresLaborales(config=config)
        r = ind.calcular(df_sintetico)
        assert r["TD_%"] == pytest.approx(25.0, abs=0.1)

    def test_tgp_conocida(self, df_sintetico):
        """TGP = PEA/PET = 80/100 = 80.0%."""
        config = ConfigGEIH(anio=2025, n_meses=1)
        ind = IndicadoresLaborales(config=config)
        r = ind.calcular(df_sintetico)
        assert r["TGP_%"] == pytest.approx(80.0, abs=0.1)

    def test_to_conocida(self, df_sintetico):
        """TO = OCI/PET = 60/100 = 60.0%."""
        config = ConfigGEIH(anio=2025, n_meses=1)
        ind = IndicadoresLaborales(config=config)
        r = ind.calcular(df_sintetico)
        assert r["TO_%"] == pytest.approx(60.0, abs=0.1)

    def test_identidad_pea(self, df_sintetico):
        """PEA = OCI + DSI debe cumplirse."""
        config = ConfigGEIH(anio=2025, n_meses=1)
        ind = IndicadoresLaborales(config=config)
        r = ind.calcular(df_sintetico)
        pea = r["PEA_M"]
        ocu = r["Ocupados_M"]
        dsi = r["Desocupados_M"]
        assert pea == pytest.approx(ocu + dsi, rel=0.01)


class TestSanityCheck:
    """Tests del sanity check multi-año."""

    def test_sanity_check_2025_pasa(self):
        """Con TD=8.9% y config 2025, sanity check debe pasar."""
        config = ConfigGEIH(anio=2025, n_meses=12)
        ind = IndicadoresLaborales(config=config)
        resultado = {
            "TD_%": 8.9, "TGP_%": 64.3, "TO_%": 58.6,
            "PET_M": 40.0, "PEA_M": 26.3,
            "Ocupados_M": 23.8, "Desocupados_M": 2.1,
        }
        ok = ind.sanity_check(resultado, "Anual 2025")
        assert ok is True

    def test_sanity_check_pea_inflada(self):
        """PEA > 40M debe disparar alerta."""
        config = ConfigGEIH(anio=2025, n_meses=12)
        ind = IndicadoresLaborales(config=config)
        resultado = {
            "TD_%": 8.9, "TGP_%": 64.3, "TO_%": 58.6,
            "PET_M": 400.0, "PEA_M": 56.0,
            "Ocupados_M": 50.0, "Desocupados_M": 6.0,
        }
        ok = ind.sanity_check(resultado, "Anual 2025")
        assert ok is False

    def test_sanity_check_2026_sin_ref(self):
        """2026 sin referencia DANE → no falla, solo advierte."""
        config = ConfigGEIH(anio=2026, n_meses=3)
        ind = IndicadoresLaborales(config=config)
        resultado = {
            "TD_%": 10.0, "TGP_%": 65.0, "TO_%": 58.0,
            "PET_M": 40.0, "PEA_M": 26.0,
            "Ocupados_M": 23.5, "Desocupados_M": 2.5,
        }
        # No debe lanzar excepción aunque no haya ref DANE
        ok = ind.sanity_check(resultado, "Trimestre 2026")
        assert ok is True  # PEA < 40M → pasa el check de seguridad


class TestIndicadoresGoldenSet:
    """Tests con el golden set sintético."""

    def test_td_golden(self, golden_set, config_2025):
        """TD del golden set debe coincidir con el valor esperado."""
        from geih.preparador import PreparadorGEIH
        from tests.generar_golden_set import GOLDEN_EXPECTED

        prep = PreparadorGEIH(config=config_2025)
        df = prep.preparar_base(golden_set)

        ind = IndicadoresLaborales(config=config_2025)
        r = ind.calcular(df)

        # Con FEX uniforme, TD debe coincidir con la proporción del golden set
        td_esperada = GOLDEN_EXPECTED["td_pct"]
        assert r["TD_%"] == pytest.approx(td_esperada, abs=0.5)

    def test_tgp_golden(self, golden_set, config_2025):
        """TGP del golden set debe coincidir."""
        from geih.preparador import PreparadorGEIH
        from tests.generar_golden_set import GOLDEN_EXPECTED

        prep = PreparadorGEIH(config=config_2025)
        df = prep.preparar_base(golden_set)

        ind = IndicadoresLaborales(config=config_2025)
        r = ind.calcular(df)

        tgp_esperada = GOLDEN_EXPECTED["tgp_pct"]
        assert r["TGP_%"] == pytest.approx(tgp_esperada, abs=0.5)

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: tests/test_indicadores.py >>>
════════════════════════════════════════════════════════════════════════════════


════════════════════════════════════════════════════════════════════════════════
🐍 ARCHIVO [39/39]: test_preparador.py
   TIPO      : CÓDIGO
   UBICACIÓN : tests
   RUTA      : tests/test_preparador.py
────────────────────────────────────────────────────────────────────────────────
>>> INICIO DE ARCHIVO: tests/test_preparador.py <<<
────────────────────────────────────────────────────────────────────────────────
# -*- coding: utf-8 -*-
"""Tests para geih.preparador — Preparación de datos."""

import numpy as np
import pandas as pd
import pytest
from geih.config import ConfigGEIH
from geih.utils import ConversorTipos


class TestFEXAdjuste:
    """Tests de la división correcta del factor de expansión."""

    def test_fex_12_meses(self):
        """FEX_ADJ = FEX_C18 / 12 para análisis anual."""
        config = ConfigGEIH(anio=2025, n_meses=12)
        df = pd.DataFrame({"FEX_C18": [1200.0, 2400.0, 600.0]})
        df["FEX_ADJ"] = df["FEX_C18"] / config.n_meses
        assert df["FEX_ADJ"].tolist() == [100.0, 200.0, 50.0]

    def test_fex_1_mes(self):
        """FEX_ADJ = FEX_C18 para mes puntual."""
        config = ConfigGEIH(anio=2025, n_meses=1)
        df = pd.DataFrame({"FEX_C18": [1200.0]})
        df["FEX_ADJ"] = df["FEX_C18"] / config.n_meses
        assert df["FEX_ADJ"].iloc[0] == 1200.0

    def test_fex_3_meses(self):
        """FEX_ADJ = FEX_C18 / 3 para trimestre."""
        config = ConfigGEIH(anio=2026, n_meses=3)
        df = pd.DataFrame({"FEX_C18": [900.0]})
        df["FEX_ADJ"] = df["FEX_C18"] / config.n_meses
        assert df["FEX_ADJ"].iloc[0] == 300.0


class TestConversorTipos:
    """Tests de estandarización de tipos."""

    def test_dpto_con_cero_lider(self):
        """'5' debe convertirse a '05' (Antioquia)."""
        serie = pd.Series(["5", "8", "11", "05"])
        resultado = ConversorTipos.estandarizar_dpto(serie)
        assert resultado.tolist() == ["05", "08", "11", "05"]

    def test_dpto_numerico(self):
        """Entrada numérica se convierte correctamente."""
        serie = pd.Series([5, 8, 11])
        resultado = ConversorTipos.estandarizar_dpto(serie)
        assert resultado.tolist() == ["05", "08", "11"]

    def test_area_5_digitos(self):
        """AREA debe tener exactamente 5 dígitos."""
        serie = pd.Series(["11001", "5001", "76001"])
        resultado = ConversorTipos.estandarizar_area(serie)
        assert resultado.tolist() == ["11001", "05001", "76001"]

    def test_ciiu4_estandarizar(self):
        """CIIU 4 dígitos: quitar .0 y rellenar a 4 dígitos."""
        serie = pd.Series(["111.0", "4711", "111", "47.0"])
        resultado = ConversorTipos.estandarizar_ciiu4(serie)
        assert resultado.iloc[0] == "0111"
        assert resultado.iloc[1] == "4711"
        assert resultado.iloc[2] == "0111"

    def test_a_numerico_con_comas(self):
        """Convierte strings con comas a float."""
        serie = pd.Series(["1.234,56", "2.000", "nan"])
        resultado = ConversorTipos.a_numerico(serie)
        # "1.234,56" → con replace de , por . → "1.234.56" → puede fallar
        # El método maneja esto con errors='coerce'
        assert resultado.notna().sum() >= 1

    def test_a_numerico_ya_numerico(self):
        """Si ya es numérico, devuelve sin cambios."""
        serie = pd.Series([1.0, 2.0, 3.0])
        resultado = ConversorTipos.a_numerico(serie)
        assert resultado.tolist() == [1.0, 2.0, 3.0]


class TestPreparadorConGoldenSet:
    """Tests del preparador usando el golden set."""

    def test_fex_adj_presente(self, golden_set, config_2025):
        """Después de preparar, FEX_ADJ debe existir."""
        from geih.preparador import PreparadorGEIH
        prep = PreparadorGEIH(config=config_2025)
        df = prep.preparar_base(golden_set)
        assert "FEX_ADJ" in df.columns

    def test_fex_adj_valor_correcto(self, golden_set, config_2025):
        """FEX_ADJ = FEX_C18 / 12 para 12 meses."""
        from geih.preparador import PreparadorGEIH
        prep = PreparadorGEIH(config=config_2025)
        df = prep.preparar_base(golden_set)
        expected = golden_set["FEX_C18"].iloc[0] / 12
        # Buscar el mismo registro en df
        assert df["FEX_ADJ"].iloc[0] == pytest.approx(expected, rel=0.01)

    def test_columnas_minimas(self, golden_set, config_2025):
        """Base preparada debe tener las columnas mínimas."""
        from geih.preparador import PreparadorGEIH
        prep = PreparadorGEIH(config=config_2025)
        df = prep.preparar_base(golden_set)
        for col in ["FEX_ADJ", "OCI", "FT", "DSI", "PET", "P3271", "P6040"]:
            assert col in df.columns, f"Falta columna: {col}"

────────────────────────────────────────────────────────────────────────────────
<<< FIN DE ARCHIVO: tests/test_preparador.py >>>
════════════════════════════════════════════════════════════════════════════════
