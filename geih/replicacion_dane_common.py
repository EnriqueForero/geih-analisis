"""
geih.replicacion_dane_common — Infraestructura compartida de replicación DANE.

Este módulo provee la infraestructura común que usan
`replicacion_dane_informalidad` y `replicacion_dane_seguridad_social`:

- Dataclasses del contrato de paridad (RutasProyecto, PeriodoMovil,
  ParametrosValidacion, CeldaComparada, BloqueReplicado, ResultadoHoja,
  ResultadoReplicacion, LayoutHoja).
- `ResolvedTrimestre`: resuelve la última columna del Excel y decide qué
  meses cargar desde los microdatos.
- `ExcelLayoutParser`: lee la estructura real de una hoja (dominios,
  bloques, categorías) sin asumir heredabilidad entre condiciones
  Total/Formal/Informal.
- `LoaderGEIH`: decide si usar un parquet consolidado existente o
  consolidar solo los meses necesarios desde ZIP.
- `ValidadorParidad`: compara celda a celda contra el Excel y clasifica.

Principios del plan IT1 v2:
- Ninguna cifra del DANE se hardcodea; todo se lee del propio Excel.
- Las categorías se extraen por terna (dominio, condición) — no se
  asume que Total/Formal/Informal comparten listado.
- El resolver de trimestre soporta los 4 formatos observados del DANE:
  "Ene - mar", "Dic - feb", "Nov 25 - ene 26", "Dic 25 - feb 26".

Autor: Néstor Enrique Forero Herrera · ProColombia · 2026-04-16
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

__all__ = [
    # Dataclasses
    "RutasProyecto",
    "PeriodoMovil",
    "ParametrosValidacion",
    "CeldaComparada",
    "BloqueReplicado",
    "ResultadoHoja",
    "ResultadoReplicacion",
    "LayoutHoja",
    "BloqueLayout",
    # Clases funcionales
    "ResolvedTrimestre",
    "ExcelLayoutParser",
    "LoaderGEIH",
    "ValidadorParidad",
    # Helpers
    "normalizar_texto",
    "hash_archivo",
    "MESES_NOMBRE_A_NUM",
    "MESES_NUM_A_NOMBRE",
]


# ═════════════════════════════════════════════════════════════════════
# Helpers
# ═════════════════════════════════════════════════════════════════════

MESES_NOMBRE_A_NUM: dict[str, int] = {
    "ene": 1,
    "enero": 1,
    "feb": 2,
    "febrero": 2,
    "mar": 3,
    "marzo": 3,
    "abr": 4,
    "abril": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "junio": 6,
    "jul": 7,
    "julio": 7,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "sept": 9,
    "septiembre": 9,
    "oct": 10,
    "octubre": 10,
    "nov": 11,
    "noviembre": 11,
    "dic": 12,
    "diciembre": 12,
}

MESES_NUM_A_NOMBRE: dict[int, str] = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def normalizar_texto(s: Any) -> str:
    """Normaliza un label del Excel a una forma comparable.

    Aplica:
    - unicodedata NFKD para descomponer tildes.
    - strip de espacios al inicio/final.
    - colapso de espacios múltiples.
    - minúsculas.
    - remueve caracteres non-ascii restantes.

    Para comparar labels entre Excel y código sin depender de
    mayúsculas, tildes o espacios duplicados.

    Args:
        s: texto a normalizar (puede ser None, float NaN, etc.).

    Returns:
        str normalizado. Si s es None/NaN → "".
    """
    if s is None:
        return ""
    if isinstance(s, float) and np.isnan(s):
        return ""
    txt = str(s)
    # Descomponer tildes
    txt = unicodedata.normalize("NFKD", txt)
    # Remover acentos combinantes
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    # Minúsculas, strip, colapso de espacios
    txt = txt.lower().strip()
    txt = re.sub(r"\s+", " ", txt)
    # Remover símbolos decorativos del DANE (^, *, ^^)
    txt = re.sub(r"[\^\*]+$", "", txt).strip()
    return txt


def hash_archivo(ruta: Path, algoritmo: str = "sha256") -> str:
    """Calcula el hash de un archivo para metadata de trazabilidad."""
    h = hashlib.new(algoritmo)
    with open(ruta, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ═════════════════════════════════════════════════════════════════════
# Dataclasses del contrato de paridad
# ═════════════════════════════════════════════════════════════════════


@dataclass
class RutasProyecto:
    """Rutas del proyecto. Única fuente de verdad para paths."""

    ruta_excel: Path
    ruta_data: Path
    ruta_salida: Path
    ruta_parquet: Optional[Path] = None
    ruta_divipola: Optional[Path] = None
    ruta_ciiu_ac: Optional[Path] = None

    def __post_init__(self):
        # Convertir strings a Path
        for campo in [
            "ruta_excel",
            "ruta_data",
            "ruta_salida",
            "ruta_parquet",
            "ruta_divipola",
            "ruta_ciiu_ac",
        ]:
            val = getattr(self, campo)
            if val is not None and not isinstance(val, Path):
                setattr(self, campo, Path(val))

    def verificar(self) -> None:
        """Fail fast: valida existencia de archivos críticos."""
        faltan: list[str] = []
        if not self.ruta_excel.exists():
            faltan.append(f"Excel oficial: {self.ruta_excel}")
        if not self.ruta_data.exists():
            faltan.append(f"Carpeta data: {self.ruta_data}")
        if faltan:
            msg = "❌ Archivos faltantes:\n" + "\n".join(f"  • {f}" for f in faltan)
            raise FileNotFoundError(msg)
        self.ruta_salida.mkdir(parents=True, exist_ok=True)


@dataclass
class PeriodoMovil:
    """Representa un trimestre móvil o un mes único.

    Para hojas trimestrales el DANE presenta "trimestres móviles":
    3 meses deslizantes, por ejemplo "Dic 25 - feb 26" = Dic 2025,
    Ene 2026, Feb 2026.

    Para la hoja `Grandes dominios` el DANE publica mensualmente, por
    lo que `es_mensual=True` y solo se usa el último mes.
    """

    etiqueta_original: str  # Texto tal como aparece en el Excel
    meses: list[tuple[int, int]]  # [(año, mes), ...] ordenados cronológicamente
    es_mensual: bool = False

    @property
    def mes_final(self) -> tuple[int, int]:
        """(año, mes) del último mes del período."""
        return self.meses[-1]

    @property
    def mes_inicial(self) -> tuple[int, int]:
        """(año, mes) del primer mes del período."""
        return self.meses[0]

    @property
    def n_meses(self) -> int:
        return len(self.meses)

    @property
    def etiqueta_corta(self) -> str:
        """Forma canonizada 'dic25_feb26' útil para nombres de archivo."""
        ai, mi = self.mes_inicial
        af, mf = self.mes_final
        mi_s = MESES_NUM_A_NOMBRE[mi][:3].lower()
        mf_s = MESES_NUM_A_NOMBRE[mf][:3].lower()
        if ai == af:
            return f"{mi_s}_{mf_s}_{ai}"
        return f"{mi_s}{ai % 100}_{mf_s}{af % 100}"


@dataclass
class ParametrosValidacion:
    """Tolerancias y reglas del contrato de paridad (plan IT1 v2 §7.2)."""

    # Tolerancias de diagnóstico temprano (semáforos)
    tol_absoluto_relativo: float = 0.01  # ±1% para valores absolutos (miles)
    tol_proporcion_pp: float = 0.4  # ±0.4 p.p. para proporciones/%
    # Criterio de cierre (estricto, cercano a redondeo de presentación)
    tol_cierre_absoluto_relativo: float = 0.001  # 0.1%
    tol_cierre_proporcion_pp: float = 0.1  # 0.1 p.p.

    def clasificar(
        self,
        valor_excel: float,
        valor_calc: float,
        es_proporcion: bool,
    ) -> tuple[str, float]:
        """Devuelve (estado, diferencia).

        Estados posibles:
          - "OK_CIERRE": cumple criterio de cierre.
          - "OK_DIAGNOSTICO": cumple tolerancia de diagnóstico pero no de cierre.
          - "FAIL": fuera de tolerancias.
          - "NA": alguno de los valores es NaN.
        """
        if valor_excel is None or valor_calc is None:
            return ("NA", np.nan)
        if isinstance(valor_excel, float) and np.isnan(valor_excel):
            return ("NA", np.nan)
        if isinstance(valor_calc, float) and np.isnan(valor_calc):
            return ("NA", np.nan)

        diff = float(valor_calc) - float(valor_excel)

        if es_proporcion:
            # Medimos en puntos porcentuales (ya vienen en %)
            if abs(diff) <= self.tol_cierre_proporcion_pp:
                return ("OK_CIERRE", diff)
            if abs(diff) <= self.tol_proporcion_pp:
                return ("OK_DIAGNOSTICO", diff)
            return ("FAIL", diff)
        else:
            # Error relativo para absolutos
            if valor_excel == 0:
                # Evitar div/0: usar diferencia absoluta
                if abs(diff) <= 0.01:
                    return ("OK_CIERRE", diff)
                return ("FAIL", diff)
            err_rel = abs(diff / valor_excel)
            if err_rel <= self.tol_cierre_absoluto_relativo:
                return ("OK_CIERRE", diff)
            if err_rel <= self.tol_absoluto_relativo:
                return ("OK_DIAGNOSTICO", diff)
            return ("FAIL", diff)


@dataclass
class CeldaComparada:
    """Unidad mínima de paridad: una celda replicada y comparada.

    Clave única del contrato (plan IT1 v2 §7.1):
        (hoja, dominio, condición, categoría, trimestre)
    """

    hoja: str
    dominio: str
    condicion: str  # 'Total' | 'Formal' | 'Informal' | 'Miles' | '%' | etc.
    categoria: str
    trimestre: str
    valor_excel: Optional[float]
    valor_calculado: Optional[float]
    diferencia: Optional[float]
    estado: str  # OK_CIERRE | OK_DIAGNOSTICO | FAIL | NA
    es_proporcion: bool = False

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class BloqueLayout:
    """Layout de un bloque dentro de una hoja (resultado del parser).

    Un bloque corresponde a una combinación (dominio, condición) en
    hojas de informalidad. En hojas de seguridad social cada bloque
    adicional agrega la dimensión (miles|%).
    """

    hoja: str
    dominio: str
    condicion: str
    fila_titulo: int  # fila donde aparece el nombre del dominio
    fila_encabezado: int  # fila "Concepto"
    fila_meses: int  # fila con meses/trimestres
    fila_inicio_datos: int  # primera fila con categoría
    fila_fin_datos: int  # última fila con categoría
    col_categoria: int = 1  # columna de categorías (A=1, B=2)
    categorias: list[tuple[int, str]] = field(default_factory=list)
    # (fila_excel, etiqueta_original)


@dataclass
class LayoutHoja:
    """Layout completo de una hoja: una lista de bloques."""

    hoja: str
    bloques: list[BloqueLayout] = field(default_factory=list)
    col_ultimo_trimestre: int = -1
    etiqueta_ultimo_trimestre: str = ""
    ano_ultimo_trimestre: Optional[int] = None


@dataclass
class BloqueReplicado:
    """Tabla réplica calculada desde microdatos: bloque → categoría → valor."""

    hoja: str
    dominio: str
    condicion: str
    valores: dict[str, float] = field(default_factory=dict)
    es_proporcion: bool = False


@dataclass
class ResultadoHoja:
    """Resultado de replicar una hoja completa."""

    hoja: str
    estado: str  # CERRADA | ABIERTA_CON_CAUSA | ABIERTA_SIN_CAUSA
    trimestre: str
    n_celdas: int = 0
    n_ok_cierre: int = 0
    n_ok_diagnostico: int = 0
    n_fail: int = 0
    n_na: int = 0
    celdas: list[CeldaComparada] = field(default_factory=list)
    causa_raiz: Optional[str] = None

    @property
    def tasa_cierre(self) -> float:
        if self.n_celdas == 0:
            return 0.0
        return self.n_ok_cierre / self.n_celdas

    def resumen(self) -> str:
        return (
            f"{self.hoja:>40s} | {self.estado:22s} | "
            f"{self.n_ok_cierre:>4d} OK / {self.n_ok_diagnostico:>3d} DIAG / "
            f"{self.n_fail:>3d} FAIL / {self.n_na:>3d} NA = {self.n_celdas:>4d}"
        )


@dataclass
class ResultadoReplicacion:
    """Resultado de replicar un Excel completo."""

    ruta_excel: Path
    hash_excel: str
    timestamp: datetime
    trimestre: PeriodoMovil
    hojas: list[ResultadoHoja] = field(default_factory=list)
    hash_codigo: Optional[str] = None
    parametros: Optional[ParametrosValidacion] = None

    def celdas_df(self) -> pd.DataFrame:
        """Devuelve todas las CeldaComparada como DataFrame."""
        filas = []
        for hoja in self.hojas:
            for celda in hoja.celdas:
                filas.append(celda.to_dict())
        if not filas:
            return pd.DataFrame(
                columns=[
                    "hoja",
                    "dominio",
                    "condicion",
                    "categoria",
                    "trimestre",
                    "valor_excel",
                    "valor_calculado",
                    "diferencia",
                    "estado",
                    "es_proporcion",
                ]
            )
        return pd.DataFrame(filas)

    def reporte_consolidado(self) -> str:
        """Reporte imprimible con el resumen por hoja."""
        lineas = [
            "=" * 110,
            f"REPORTE DE PARIDAD — Trimestre objetivo: {self.trimestre.etiqueta_original}",
            f"Excel: {self.ruta_excel.name}  (hash: {self.hash_excel[:12]}...)",
            f"Timestamp: {self.timestamp:%Y-%m-%d %H:%M:%S}",
            "=" * 110,
            f"{'Hoja':>40s} | {'Estado':22s} | {'Cierre/Diag/Fail/NA = N'}",
            "-" * 110,
        ]
        for h in self.hojas:
            lineas.append(h.resumen())
        total_celdas = sum(h.n_celdas for h in self.hojas)
        total_cierre = sum(h.n_ok_cierre for h in self.hojas)
        total_fail = sum(h.n_fail for h in self.hojas)
        cerradas = sum(1 for h in self.hojas if h.estado == "CERRADA")
        lineas.append("-" * 110)
        lineas.append(
            f"TOTAL: {cerradas}/{len(self.hojas)} hojas cerradas | "
            f"{total_cierre}/{total_celdas} celdas en cierre | {total_fail} FAIL"
        )
        lineas.append("=" * 110)
        return "\n".join(lineas)


# ═════════════════════════════════════════════════════════════════════
# ResolvedTrimestre — parser de la última columna del Excel
# ═════════════════════════════════════════════════════════════════════


class ResolvedTrimestre:
    """Resuelve el período objetivo a replicar desde la última columna del Excel.

    Soporta los 4 formatos observados del DANE (plan IT1 v2 §2.1):

    - "Ene - mar"               → mensual sin año embebido (usar fila 12 como ancla)
    - "Dic - feb"               → cruza de año; inferir años por contexto
    - "Nov 25 - ene 26"         → año embebido de 2 dígitos
    - "Dic 25 - feb 26"         → año embebido de 2 dígitos
    - "Ene", "Feb", ..., "Dic"  → mensual (hoja Grandes dominios)

    Uso típico:
        r = ResolvedTrimestre.desde_excel("docs/anex-EISS.xlsx", hoja="Total nacional")
        print(r.periodo.etiqueta_original)
        print(r.periodo.meses)   # [(2025, 12), (2026, 1), (2026, 2)]
    """

    # Columna donde arrancan los datos (después de labels)
    COL_PRIMERA_DATO = 2  # B

    # Regex compilados
    _RE_TRIM_CON_ANIO = re.compile(
        r"^\s*([a-zñ]+)\s*(\d{2})\s*-\s*([a-zñ]+)\s*(\d{2})\s*$",
        re.IGNORECASE,
    )
    _RE_TRIM_SIN_ANIO = re.compile(
        r"^\s*([a-zñ]+)\s*-\s*([a-zñ]+)\s*$",
        re.IGNORECASE,
    )
    _RE_MES_SOLO = re.compile(r"^\s*([a-zñ]+)\s*$", re.IGNORECASE)

    def __init__(
        self,
        periodo: PeriodoMovil,
        columna_excel: int,
        hoja_fuente: str,
    ):
        self.periodo = periodo
        self.columna_excel = columna_excel
        self.hoja_fuente = hoja_fuente

    # ────────────────────────────────────────────────────────────────
    # Parsing de cadenas de período
    # ────────────────────────────────────────────────────────────────

    @classmethod
    def parsear_etiqueta(
        cls,
        etiqueta: str,
        ano_ancla: Optional[int] = None,
    ) -> PeriodoMovil:
        """Convierte una cadena de período del Excel en `PeriodoMovil`.

        Args:
            etiqueta: texto tal como aparece en la fila 13 del Excel.
            ano_ancla: año leído de la fila 12 del mismo bloque cuando
                la etiqueta no trae año embebido. Obligatorio si la
                etiqueta es "Ene - mar" o similar sin año.

        Returns:
            `PeriodoMovil` con la lista de meses resueltos.

        Raises:
            ValueError: si la etiqueta no matchea ningún formato conocido
                o si falta año y no se pasa ano_ancla.
        """
        if not etiqueta or not isinstance(etiqueta, str):
            raise ValueError(f"Etiqueta vacía o no string: {etiqueta!r}")

        txt = etiqueta.strip()
        txt_low = normalizar_texto(txt)

        # Caso 1: trimestre con año embebido — "Dic 25 - feb 26"
        m = cls._RE_TRIM_CON_ANIO.match(txt_low)
        if m:
            mes_i_str, yy_i, mes_f_str, yy_f = m.groups()
            mes_i = cls._mes_a_num(mes_i_str)
            mes_f = cls._mes_a_num(mes_f_str)
            anio_i = 2000 + int(yy_i)
            anio_f = 2000 + int(yy_f)
            meses = cls._construir_trimestre(anio_i, mes_i, anio_f, mes_f)
            return PeriodoMovil(etiqueta_original=txt, meses=meses, es_mensual=False)

        # Caso 2: trimestre sin año — "Ene - mar" o "Dic - feb"
        m = cls._RE_TRIM_SIN_ANIO.match(txt_low)
        if m:
            mes_i_str, mes_f_str = m.groups()
            mes_i = cls._mes_a_num(mes_i_str)
            mes_f = cls._mes_a_num(mes_f_str)
            if ano_ancla is None:
                raise ValueError(f"Etiqueta '{txt}' requiere ano_ancla (no tiene año embebido)")
            # Si cruza de año (Dic - feb), año inicial = ancla-1, final = ancla
            if mes_i > mes_f:
                anio_i = ano_ancla - 1
                anio_f = ano_ancla
            else:
                anio_i = ano_ancla
                anio_f = ano_ancla
            meses = cls._construir_trimestre(anio_i, mes_i, anio_f, mes_f)
            return PeriodoMovil(etiqueta_original=txt, meses=meses, es_mensual=False)

        # Caso 3: mes único — "Ene", "Feb", etc.
        m = cls._RE_MES_SOLO.match(txt_low)
        if m:
            mes_str = m.group(1)
            mes_num = cls._mes_a_num(mes_str)
            if ano_ancla is None:
                raise ValueError(f"Etiqueta '{txt}' requiere ano_ancla")
            return PeriodoMovil(
                etiqueta_original=txt,
                meses=[(ano_ancla, mes_num)],
                es_mensual=True,
            )

        raise ValueError(f"Formato de etiqueta no reconocido: {etiqueta!r}")

    @staticmethod
    def _mes_a_num(mes_txt: str) -> int:
        """Convierte 'ene', 'enero', 'Ene', etc. → 1..12."""
        k = normalizar_texto(mes_txt).strip(". ")
        if k in MESES_NOMBRE_A_NUM:
            return MESES_NOMBRE_A_NUM[k]
        # Reintentar con los primeros 3 chars
        k3 = k[:3]
        if k3 in MESES_NOMBRE_A_NUM:
            return MESES_NOMBRE_A_NUM[k3]
        raise ValueError(f"Mes no reconocido: {mes_txt!r}")

    @staticmethod
    def _construir_trimestre(
        anio_i: int,
        mes_i: int,
        anio_f: int,
        mes_f: int,
    ) -> list[tuple[int, int]]:
        """Genera la lista de (año, mes) del trimestre inclusive."""
        meses: list[tuple[int, int]] = []
        y, m = anio_i, mes_i
        # Iterar hasta llegar a (anio_f, mes_f) inclusive
        # Máximo 12 iteraciones por seguridad
        for _ in range(12):
            meses.append((y, m))
            if (y, m) == (anio_f, mes_f):
                break
            m += 1
            if m > 12:
                m = 1
                y += 1
        else:
            raise ValueError(f"No se pudo construir trimestre {anio_i}/{mes_i} → {anio_f}/{mes_f}")
        return meses

    # ────────────────────────────────────────────────────────────────
    # Construcción desde un Excel real
    # ────────────────────────────────────────────────────────────────

    @classmethod
    def desde_excel(
        cls,
        ruta_excel: Path,
        hoja: str = "Total nacional",
        fila_meses: int = 13,
        fila_anos: int = 12,
        override: Optional[str] = None,
    ) -> ResolvedTrimestre:
        """Lee la última columna visible de una hoja del Excel.

        Args:
            ruta_excel: path al archivo Excel oficial.
            hoja: nombre de la hoja donde buscar la última columna.
                Por defecto 'Total nacional' (hoja trimestral base).
            fila_meses: fila donde están las etiquetas de trimestre
                (típicamente 13).
            fila_anos: fila donde están los años (típicamente 12).
            override: si se pasa una etiqueta literal, se usa esa en vez
                de leer del Excel. Útil para replicar trimestres históricos.

        Returns:
            `ResolvedTrimestre` completo.
        """
        ruta_excel = Path(ruta_excel)
        wb = load_workbook(ruta_excel, read_only=True, data_only=True)
        try:
            if hoja not in wb.sheetnames:
                raise KeyError(
                    f"Hoja '{hoja}' no existe en {ruta_excel.name}. "
                    f"Disponibles: {wb.sheetnames}"
                )
            ws = wb[hoja]
            # Autodetección si fila_meses está vacía
            fila_meses_real = fila_meses
            v_test = ws.cell(fila_meses_real, 2).value
            if v_test is None:
                # Buscar primera fila entre 10 y 20 con etiqueta de período
                for f in range(10, 21):
                    v = ws.cell(f, 2).value
                    if v is not None and str(v).strip():
                        txt = str(v).strip().lower()
                        if re.match(r"^[a-zñ]{3,}(\s*-\s*[a-zñ]{3,})?", txt):
                            fila_meses_real = f
                            break
            fila_anos_real = fila_meses_real - 1

            # Encontrar última columna con dato en fila de meses
            max_col = ws.max_column
            ultima_col = 0
            for c in range(max_col, cls.COL_PRIMERA_DATO - 1, -1):
                v = ws.cell(fila_meses_real, c).value
                if v is not None and str(v).strip():
                    ultima_col = c
                    break
            if ultima_col == 0:
                raise ValueError(f"No se encontró columna con datos en '{hoja}'")

            etiqueta = str(ws.cell(fila_meses_real, ultima_col).value).strip()

            # Buscar año ancla recorriendo hacia atrás en fila_anos
            ano_ancla: Optional[int] = None
            for c in range(ultima_col, cls.COL_PRIMERA_DATO - 1, -1):
                v = ws.cell(fila_anos_real, c).value
                if v is not None:
                    try:
                        ano_ancla = int(v)
                        break
                    except (ValueError, TypeError):
                        try:
                            ano_ancla = int(str(v).strip())
                            break
                        except ValueError:
                            continue

            etiqueta_a_parsear = override if override else etiqueta
            periodo = cls.parsear_etiqueta(etiqueta_a_parsear, ano_ancla=ano_ancla)

            return cls(
                periodo=periodo,
                columna_excel=ultima_col,
                hoja_fuente=hoja,
            )
        finally:
            wb.close()


# ═════════════════════════════════════════════════════════════════════
# ExcelLayoutParser
# ═════════════════════════════════════════════════════════════════════


class ExcelLayoutParser:
    """Lee el layout real de una hoja del Excel oficial del DANE.

    Detecta bloques por reconocimiento de títulos de dominio y
    extrae categorías por terna (dominio, condición) respetando que
    distintas condiciones pueden tener distintos listados (plan IT1
    v2 §5.2 - riesgo 7).

    Un bloque típico tiene esta estructura:

        R 11: Total Nacional             ← título del dominio
        R 12: Concepto       2021  ...   ← encabezado (fila_anos)
        R 13:                Ene - mar   ← meses (fila_meses)
        R 14: Población ocupada   ...    ← inicio de datos
        R 15: Formal              ...    ← puede marcar sub-condición
        R 16: Informal            ...
        ...
        R 22:                            ← línea vacía = fin de bloque
    """

    # Títulos de dominio reconocidos (normalizados)
    TITULOS_DOMINIO_NORM: set[str] = {
        "total nacional",
        "cabeceras",
        "centros poblados y rural disperso",
        "13 ciudades y a.m.",
        "13 ciudades y areas metropolitanas",
        "10 ciudades",
        "23 ciudades y a.m.",
        "23 ciudades y areas metropolitanas",
        "total nacional - hombres",
        "total nacional - mujeres",
        "cabeceras - hombres",
        "cabeceras - mujeres",
        "centros poblados y rural disperso - hombres",
        "centros poblados y rural disperso - mujeres",
        "13 ciudades y areas metropolitanas - hombres",
        "13 ciudades y areas metropolitanas - mujeres",
        "23 ciudades y areas metropolitanas - hombres",
        "23 ciudades y areas metropolitanas - mujeres",
        "13 ciudades y a.m. - hombres",
        "13 ciudades y a.m. - mujeres",
        "23 ciudades y a.m. - hombres",
        "23 ciudades y a.m. - mujeres",
    }

    # Palabras clave que identifican una condición (Total / Formal / Informal)
    # en hojas de desagregación. Aparecen dentro de un bloque de dominio.
    CONDICIONES_NORM: dict[str, str] = {
        "poblacion ocupada": "Total",
        "ocupada total": "Total",
        "formal": "Formal",
        "informal": "Informal",
    }

    def __init__(
        self,
        ruta_excel: Path,
        fila_meses: int = 13,
        fila_anos: int = 12,
    ):
        self.ruta_excel = Path(ruta_excel)
        self.fila_meses = fila_meses
        self.fila_anos = fila_anos
        self._wb = None

    def __enter__(self):
        self._wb = load_workbook(self.ruta_excel, read_only=True, data_only=True)
        return self

    def __exit__(self, *a):
        if self._wb is not None:
            self._wb.close()
            self._wb = None

    # ────────────────────────────────────────────────────────────────

    def _abrir(self):
        if self._wb is None:
            self._wb = load_workbook(self.ruta_excel, read_only=True, data_only=True)
        return self._wb

    def _ws(self, nombre: str) -> Worksheet:
        wb = self._abrir()
        if nombre not in wb.sheetnames:
            raise KeyError(f"Hoja '{nombre}' no existe")
        return wb[nombre]

    @staticmethod
    def _es_titulo_dominio(txt: str) -> bool:
        """¿El string parece un título de dominio (aislado en col A)?"""
        norm = normalizar_texto(txt)
        if not norm:
            return False
        # Match exacto con catálogo
        # if norm in ExcelLayoutParser.TITULOS_DOMINIO_NORM:
        #     return True
        # Match por prefijo: "Bogotá D.C.", "Medellín A.M.", etc.
        # Son ciudades en hoja "Ciudades" y "Prop informalidad".
        # Regla heurística: contiene "a.m." o es una de las 23 ciudades
        # conocidas. Mejor: devolver False aquí y que el caller use
        # métodos especializados por hoja (más seguro).
        return norm in ExcelLayoutParser.TITULOS_DOMINIO_NORM  # False

    def encontrar_ultima_columna(
        self,
        hoja: str,
    ) -> tuple[int, str, Optional[int]]:
        """Devuelve (columna, etiqueta_trimestre, ano_ancla) de la última
        columna con dato en la fila de meses.

        Autodetecta la fila de meses si la fila configurada
        (self.fila_meses) está vacía: busca en filas 10–20 la primera
        que contenga una etiqueta de período válida.
        """
        ws = self._ws(hoja)
        max_col = ws.max_column
        # Autodetección de fila_meses si la actual está vacía
        fila_meses_real = self.fila_meses
        if not self._fila_tiene_etiquetas_periodo(ws, fila_meses_real):
            fila_meses_real = self._autodetectar_fila_meses(ws)
        fila_anos_real = fila_meses_real - 1

        col = 0
        for c in range(max_col, 1, -1):
            v = ws.cell(fila_meses_real, c).value
            if v is not None and str(v).strip():
                col = c
                break
        if col == 0:
            raise ValueError(f"No hay dato de período en ninguna fila de '{hoja}'")
        etiqueta = str(ws.cell(fila_meses_real, col).value).strip()

        # Retroceder para encontrar último año
        ano: Optional[int] = None
        for c in range(col, 1, -1):
            v = ws.cell(fila_anos_real, c).value
            if v is not None:
                try:
                    ano = int(v)
                    break
                except (ValueError, TypeError):
                    try:
                        ano = int(str(v).strip())
                        break
                    except ValueError:
                        continue
        return col, etiqueta, ano

    @staticmethod
    def _fila_tiene_etiquetas_periodo(ws: Worksheet, fila: int) -> bool:
        """¿La fila tiene al menos un valor que parece etiqueta de período?"""
        # Escanear primeras 10 columnas
        for c in range(2, 12):
            v = ws.cell(fila, c).value
            if v is not None and str(v).strip():
                txt = normalizar_texto(str(v))
                # Match con formato mes / trimestre
                if re.match(r"^[a-z]{3,}(\s*-\s*[a-z]{3,})?", txt):
                    return True
        return False

    @staticmethod
    def _autodetectar_fila_meses(ws: Worksheet) -> int:
        """Busca la primera fila entre 10 y 20 con etiquetas de período."""
        for fila in range(10, 21):
            if ExcelLayoutParser._fila_tiene_etiquetas_periodo(ws, fila):
                return fila
        raise ValueError("No se encontró fila con etiquetas de período entre 10 y 20")

    # ────────────────────────────────────────────────────────────────
    # Parser genérico por bloques de dominio (hojas de SS principalmente)
    # ────────────────────────────────────────────────────────────────

    def parsear_bloques_por_dominio(
        self,
        hoja: str,
        col_categoria: int = 1,
        filas_encabezado_relativo: tuple[int, int] = (1, 2),
    ) -> LayoutHoja:
        """Identifica bloques de dominio en una hoja.

        Un bloque arranca cuando la columna A tiene un título de dominio
        conocido; termina cuando aparece el siguiente título o se
        alcanza el final de datos. Dentro del bloque, se extraen todas
        las categorías no vacías de col_categoria.

        Este parser NO distingue sub-condiciones (Formal/Informal); es
        adecuado para hojas como `Total nacional` o seguridad social
        donde el bloque se toma como un único listado vertical.

        Para hojas tipo `Sexo`, `Ramas`, `Posición ocupacional` use
        `parsear_bloques_por_dominio_y_condicion`.
        """
        ws = self._ws(hoja)
        layout = LayoutHoja(hoja=hoja)
        col_ult, etiqueta, ano = self.encontrar_ultima_columna(hoja)
        layout.col_ultimo_trimestre = col_ult
        layout.etiqueta_ultimo_trimestre = etiqueta
        layout.ano_ultimo_trimestre = ano

        n_filas = ws.max_row
        fila = 1
        bloques: list[BloqueLayout] = []

        while fila <= n_filas:
            v = ws.cell(fila, col_categoria).value
            if v is not None and self._es_titulo_dominio(str(v)):
                dominio = str(v).strip()
                bloque = BloqueLayout(
                    hoja=hoja,
                    dominio=dominio,
                    condicion="Miles",  # default, caller puede sobrescribir
                    fila_titulo=fila,
                    fila_encabezado=fila + filas_encabezado_relativo[0],
                    fila_meses=fila + filas_encabezado_relativo[1],
                    fila_inicio_datos=fila + filas_encabezado_relativo[1] + 1,
                    fila_fin_datos=fila,
                    col_categoria=col_categoria,
                )
                # Extraer categorías hasta la próxima fila vacía o próximo dominio
                i = bloque.fila_inicio_datos
                ultima_con_dato = i - 1
                while i <= n_filas:
                    v_i = ws.cell(i, col_categoria).value
                    if v_i is None or not str(v_i).strip():
                        # Fila vacía: ¿fin de bloque o solo espacio?
                        # Si la siguiente fila es otro título, fin.
                        if i + 1 <= n_filas:
                            nxt = ws.cell(i + 1, col_categoria).value
                            if nxt is not None and self._es_titulo_dominio(str(nxt)):
                                break
                        # Si varias filas vacías consecutivas, fin
                        if i + 1 > n_filas or not ws.cell(i + 1, col_categoria).value:
                            # Chequear si hay título más abajo
                            # (típicamente bloque de distribución % en SS)
                            salto = i + 2
                            if salto <= n_filas:
                                nxt2 = ws.cell(salto, col_categoria).value
                                if nxt2 is not None and self._es_titulo_dominio(str(nxt2)):
                                    break
                        i += 1
                        continue
                    # Es una categoría válida
                    etq = str(v_i).strip()
                    # Excluir filas de metadata (Fuente, Nota, etc.)
                    if etq.lower().startswith(("fuente", "nota", "actualizado")):
                        break
                    # Excluir encabezados internos
                    if etq.lower() in ("concepto",):
                        i += 1
                        continue
                    bloque.categorias.append((i, etq))
                    ultima_con_dato = i
                    i += 1
                bloque.fila_fin_datos = ultima_con_dato
                bloques.append(bloque)
                fila = ultima_con_dato + 1
            else:
                fila += 1

        layout.bloques = bloques
        return layout

    # ────────────────────────────────────────────────────────────────
    # Parser especializado: hojas con sub-condiciones Total/Formal/Informal
    # ────────────────────────────────────────────────────────────────

    def parsear_bloques_por_dominio_y_condicion(
        self,
        hoja: str,
        col_categoria: int = 1,
        filas_encabezado_relativo: tuple[int, int] = (1, 2),
    ) -> LayoutHoja:
        """Para hojas donde cada dominio contiene bloques Total/Formal/Informal.

        Aplicable a: Sexo, Educación, Ramas, Posición ocupacional,
        Lugar de trabajo, Tamaño de empresa.

        Estructura esperada dentro de cada bloque de dominio:
            Población ocupada     → inicio de sub-bloque Total
              categoría 1
              categoría 2
              ...
            Formal                → inicio de sub-bloque Formal
              categoría 1
              ...
            Informal              → inicio de sub-bloque Informal
              categoría 1
              ...
        """
        layout_base = self.parsear_bloques_por_dominio(
            hoja, col_categoria, filas_encabezado_relativo
        )
        self._ws(hoja)

        bloques_expandidos: list[BloqueLayout] = []
        for b in layout_base.bloques:
            # Escanear las categorías del bloque y dividir por condición
            sub_bloque: Optional[BloqueLayout] = None
            for fila_i, etq in b.categorias:
                etq_norm = normalizar_texto(etq)
                # ¿Es un marcador de condición?
                if etq_norm in ("poblacion ocupada",):
                    # inicia condición Total
                    if sub_bloque is not None:
                        sub_bloque.fila_fin_datos = fila_i - 1
                        bloques_expandidos.append(sub_bloque)
                    sub_bloque = BloqueLayout(
                        hoja=hoja,
                        dominio=b.dominio,
                        condicion="Total",
                        fila_titulo=b.fila_titulo,
                        fila_encabezado=b.fila_encabezado,
                        fila_meses=b.fila_meses,
                        fila_inicio_datos=fila_i + 1,
                        fila_fin_datos=fila_i,
                        col_categoria=col_categoria,
                    )
                elif etq_norm == "formal":
                    if sub_bloque is not None:
                        sub_bloque.fila_fin_datos = fila_i - 1
                        bloques_expandidos.append(sub_bloque)
                    sub_bloque = BloqueLayout(
                        hoja=hoja,
                        dominio=b.dominio,
                        condicion="Formal",
                        fila_titulo=b.fila_titulo,
                        fila_encabezado=b.fila_encabezado,
                        fila_meses=b.fila_meses,
                        fila_inicio_datos=fila_i + 1,
                        fila_fin_datos=fila_i,
                        col_categoria=col_categoria,
                    )
                elif etq_norm == "informal":
                    if sub_bloque is not None:
                        sub_bloque.fila_fin_datos = fila_i - 1
                        bloques_expandidos.append(sub_bloque)
                    sub_bloque = BloqueLayout(
                        hoja=hoja,
                        dominio=b.dominio,
                        condicion="Informal",
                        fila_titulo=b.fila_titulo,
                        fila_encabezado=b.fila_encabezado,
                        fila_meses=b.fila_meses,
                        fila_inicio_datos=fila_i + 1,
                        fila_fin_datos=fila_i,
                        col_categoria=col_categoria,
                    )
                else:
                    # Es una categoría regular
                    if sub_bloque is None:
                        # No había marcador de condición, usar Total como default
                        sub_bloque = BloqueLayout(
                            hoja=hoja,
                            dominio=b.dominio,
                            condicion="Total",
                            fila_titulo=b.fila_titulo,
                            fila_encabezado=b.fila_encabezado,
                            fila_meses=b.fila_meses,
                            fila_inicio_datos=fila_i,
                            fila_fin_datos=fila_i,
                            col_categoria=col_categoria,
                        )
                    sub_bloque.categorias.append((fila_i, etq))
                    sub_bloque.fila_fin_datos = fila_i
            if sub_bloque is not None and sub_bloque.categorias:
                bloques_expandidos.append(sub_bloque)

        layout_base.bloques = bloques_expandidos
        return layout_base


# ═════════════════════════════════════════════════════════════════════
# LoaderGEIH — decide parquet vs zip
# ═════════════════════════════════════════════════════════════════════


class LoaderGEIH:
    """Carga de microdatos GEIH decidiendo entre parquet consolidado o ZIP.

    Principio: minimizar I/O y memoria.
    1. Si existe un parquet consolidado y abarca los meses requeridos,
       cargarlo y filtrar.
    2. Si no, consolidar solo los meses requeridos desde ZIP y persistir
       un parquet checkpoint nombrado según el trimestre.
    """

    def __init__(
        self,
        ruta_data: Path,
        ruta_parquet: Optional[Path] = None,
        consolidador=None,  # instancia de ConsolidadorGEIH (inyectada)
        preparador=None,  # instancia de PreparadorGEIH (inyectada)
    ):
        self.ruta_data = Path(ruta_data)
        self.ruta_parquet = Path(ruta_parquet) if ruta_parquet else None
        self._consolidador = consolidador
        self._preparador = preparador

    def cargar(
        self,
        periodo: PeriodoMovil,
        forzar: bool = False,
        solo_ocupados: bool = True,
    ) -> pd.DataFrame:
        """Carga los microdatos del período como DataFrame preparado.

        Args:
            periodo: `PeriodoMovil` con la lista de meses a cargar.
            forzar: si True, ignora parquet existente y re-consolida.
            solo_ocupados: si True, filtra OCI==1 (para hojas de informalidad).

        Returns:
            DataFrame con columnas del preparador + INFORMAL.
        """
        # Path del parquet checkpoint
        if self.ruta_parquet and self.ruta_parquet.exists() and not forzar:
            print(f"♻️  Cargando parquet existente: {self.ruta_parquet.name}")
            df = pd.read_parquet(self.ruta_parquet)
            # Verificar que cubre los meses
            if "MES_NUM" in df.columns and "PER" in df.columns:
                meses_presentes = set(
                    zip(
                        df["PER"].astype(int).tolist(),
                        df["MES_NUM"].astype(int).tolist(),
                    )
                )
                meses_faltantes = set(periodo.meses) - meses_presentes
                if meses_faltantes:
                    print(
                        f"⚠️  Parquet no cubre: {sorted(meses_faltantes)}. "
                        f"Consolidando desde ZIP."
                    )
                    df = self._consolidar_desde_zip(periodo, solo_ocupados)
            return df

        # Consolidar desde ZIP
        return self._consolidar_desde_zip(periodo, solo_ocupados)

    def _consolidar_desde_zip(
        self,
        periodo: PeriodoMovil,
        solo_ocupados: bool,
    ) -> pd.DataFrame:
        """Consolida los meses del período desde los ZIPs del DANE.

        Adaptador: traduce `PeriodoMovil.meses` (lista de `(año, mes)`)
        a la API real del `ConsolidadorGEIH`, que recibe nombres de
        carpeta como `["Diciembre 2025", "Enero 2026", "Febrero 2026"]`.

        Si los ZIPs de distintos años están mezclados en `ruta_data`
        (p.ej. `Diciembre 2025.zip`, `Enero 2026.zip`, `Febrero 2026.zip`),
        el consolidador solo ve esa misma raíz y cada ZIP se resuelve
        por nombre de carpeta.
        """
        if self._consolidador is None or self._preparador is None:
            raise RuntimeError(
                "LoaderGEIH requiere consolidador y preparador inyectados. "
                "Uso: LoaderGEIH(ruta_data, consolidador=..., preparador=...)"
            )

        # Importar aquí para evitar import circular
        from .config import MESES_NOMBRES

        # Traducir (año, mes) → "Diciembre 2025"
        carpetas: list[str] = []
        for anio, mes in periodo.meses:
            nombre_mes = MESES_NOMBRES[mes - 1]
            carpetas.append(f"{nombre_mes} {anio}")

        print(f"🔄 Consolidando {len(carpetas)} carpeta(s): {carpetas}")

        # Asegurar que el consolidador apunta a la ruta correcta
        # (por si fue inyectado con una ruta distinta)
        if hasattr(self._consolidador, "ruta_base"):
            self._consolidador.ruta_base = self.ruta_data

        # Llamar al API real: consolidar(carpetas=[...], checkpoint=False)
        df_raw = self._consolidador.consolidar(
            carpetas=carpetas,
            checkpoint=False,
        )
        print(f"   Raw consolidado: {len(df_raw):,} registros")

        # El ConsolidadorGEIH asigna MES_NUM secuencialmente (1, 2, 3, ...)
        # según el ORDEN de las carpetas procesadas, NO según el mes del
        # calendario. Para el trimestre "Dic 25 – Feb 26" las carpetas van
        # en orden cronológico y MES_NUM recibe 1=Dic, 2=Ene, 3=Feb.
        #
        # Reescribimos MES_NUM y creamos PER para tener el mapeo calendario
        # correcto, que es lo que el notebook de replicación necesita.
        mes_num_original_a_real = {i + 1: mes for i, (_, mes) in enumerate(periodo.meses)}
        mes_num_original_a_anio = {i + 1: anio for i, (anio, _) in enumerate(periodo.meses)}
        if "MES_NUM" in df_raw.columns:
            df_raw["PER"] = df_raw["MES_NUM"].map(mes_num_original_a_anio).astype(int)
            df_raw["MES_NUM"] = df_raw["MES_NUM"].map(mes_num_original_a_real).astype(int)

        # Ajustar n_meses del preparador para que FEX_ADJ sea correcto
        if hasattr(self._preparador, "config"):
            self._preparador.config.n_meses = periodo.n_meses

        # Preparar (aplica FEX_ADJ, informalidad, variables derivadas).
        # Pasar PER + AREA como columnas extra. AREA viene como float
        # (ej. 5.0, 11.0, 76.0) en el raw y hay que convertirla a string
        # de 2 dígitos antes de usarla para DOMINIO/CIUDAD.
        df = self._preparador.preparar_base(
            df_raw,
            solo_ocupados=solo_ocupados,
            columnas_extra=["PER", "AREA"],
        )

        # ── Rederivar DOMINIO y CIUDAD usando AREA (2 dígitos) ────────
        # Bug del preparador: hace `df["AREA"].astype(str).str.zfill(2)`
        # pero AREA viene como float (5.0), así que `astype(str)` da
        # "5.0" y zfill(2) queda "5.0" — no matchea con los sets de
        # DPTOS_13_CIUDADES que están en formato 2-dígitos string ("05").
        # Fix: convertir a Int64 primero y luego zfill(2).
        if "AREA" in df.columns and "CLASE" in df.columns:
            from .config import (
                AREA_GEIH_A_CIUDAD,
                DPTO_A_CIUDAD,
                DPTOS_10_CIUDADES,
                DPTOS_13_CIUDADES,
            )

            # AREA numérica → int → string 2 dígitos
            area_num = pd.to_numeric(df["AREA"], errors="coerce")
            area_str = area_num.dropna().astype(int).astype(str).str.zfill(2)
            area_str = area_str.reindex(df.index)
            clase_num = pd.to_numeric(df["CLASE"], errors="coerce")

            # DOMINIO según Boletín DANE oficial
            es_13 = area_str.isin(DPTOS_13_CIUDADES)
            es_10 = area_str.isin(DPTOS_10_CIUDADES)
            nuevo_dominio = pd.Series("otras_cab", index=df.index, dtype="object")
            nuevo_dominio[(clase_num == 1) & es_10] = "10_ciudades"
            nuevo_dominio[(clase_num == 1) & es_13] = "13_AM"
            nuevo_dominio[clase_num == 2] = "rural"
            df["DOMINIO"] = nuevo_dominio

            # CIUDAD por AREA (2 dígitos) — solo para registros con AREA
            # no-NaN (rurales quedan sin ciudad).
            ciudad_por_area = area_str.map(AREA_GEIH_A_CIUDAD)
            # Fallback a DPTO solo para cabeceras (CLASE=1); rurales quedan NaN
            if "DPTO_STR" in df.columns:
                ciudad_por_dpto = df["DPTO_STR"].map(DPTO_A_CIUDAD)
                ciudad_final = ciudad_por_area.fillna(ciudad_por_dpto)
            else:
                ciudad_final = ciudad_por_area
            # En zona rural nullificar CIUDAD
            ciudad_final = ciudad_final.where(clase_num == 1, other=None)
            df["CIUDAD"] = ciudad_final

        # Persistir checkpoint si se configuró parquet
        if self.ruta_parquet:
            self.ruta_parquet.parent.mkdir(parents=True, exist_ok=True)
            df.to_parquet(self.ruta_parquet, index=False)
            print(f"💾 Checkpoint: {self.ruta_parquet.name}")

        return df


# ═════════════════════════════════════════════════════════════════════
# ValidadorParidad — compara tabla réplica contra Excel
# ═════════════════════════════════════════════════════════════════════


class ValidadorParidad:
    """Compara `BloqueReplicado` contra valores leídos del Excel oficial.

    Produce `CeldaComparada` por cada categoría y calcula métricas
    agregadas para el `ResultadoHoja`.
    """

    def __init__(self, parametros: Optional[ParametrosValidacion] = None):
        self.params = parametros or ParametrosValidacion()

    def comparar_bloque(
        self,
        bloque_calc: BloqueReplicado,
        layout_bloque: BloqueLayout,
        ws: Worksheet,
        columna_excel: int,
        trimestre_etiqueta: str,
    ) -> list[CeldaComparada]:
        """Compara un bloque replicado contra las celdas del Excel.

        Matching por normalización de labels: la categoría de
        `bloque_calc.valores` se normaliza y se busca en las
        categorías del `layout_bloque`.
        """
        celdas: list[CeldaComparada] = []

        # Índice normalizado → (fila, etiqueta_original)
        idx_excel = {normalizar_texto(etq): (fila, etq) for fila, etq in layout_bloque.categorias}

        categorias_excel_usadas: set[str] = set()

        for cat_calc, valor_calc in bloque_calc.valores.items():
            cat_norm = normalizar_texto(cat_calc)
            if cat_norm not in idx_excel:
                # La categoría calculada no existe en el Excel:
                # se registra como celda NA para traza
                celda = CeldaComparada(
                    hoja=bloque_calc.hoja,
                    dominio=bloque_calc.dominio,
                    condicion=bloque_calc.condicion,
                    categoria=cat_calc,
                    trimestre=trimestre_etiqueta,
                    valor_excel=None,
                    valor_calculado=valor_calc,
                    diferencia=None,
                    estado="NA",
                    es_proporcion=bloque_calc.es_proporcion,
                )
                celdas.append(celda)
                continue

            fila_excel, etq_original = idx_excel[cat_norm]
            categorias_excel_usadas.add(cat_norm)
            valor_excel = ws.cell(fila_excel, columna_excel).value
            # Convertir a float
            try:
                valor_excel_f = float(valor_excel) if valor_excel is not None else None
            except (ValueError, TypeError):
                valor_excel_f = None

            estado, diff = self.params.clasificar(
                valor_excel=valor_excel_f,
                valor_calc=valor_calc,
                es_proporcion=bloque_calc.es_proporcion,
            )

            celda = CeldaComparada(
                hoja=bloque_calc.hoja,
                dominio=bloque_calc.dominio,
                condicion=bloque_calc.condicion,
                categoria=etq_original,  # usar etiqueta como viene en el Excel
                trimestre=trimestre_etiqueta,
                valor_excel=valor_excel_f,
                valor_calculado=float(valor_calc) if valor_calc is not None else None,
                diferencia=diff,
                estado=estado,
                es_proporcion=bloque_calc.es_proporcion,
            )
            celdas.append(celda)

        # Categorías del Excel que no fueron replicadas
        for cat_norm, (fila_excel, etq_original) in idx_excel.items():
            if cat_norm in categorias_excel_usadas:
                continue
            valor_excel = ws.cell(fila_excel, columna_excel).value
            try:
                valor_excel_f = float(valor_excel) if valor_excel is not None else None
            except (ValueError, TypeError):
                valor_excel_f = None
            celda = CeldaComparada(
                hoja=bloque_calc.hoja,
                dominio=bloque_calc.dominio,
                condicion=bloque_calc.condicion,
                categoria=etq_original,
                trimestre=trimestre_etiqueta,
                valor_excel=valor_excel_f,
                valor_calculado=None,
                diferencia=None,
                estado="NA",
                es_proporcion=bloque_calc.es_proporcion,
            )
            celdas.append(celda)

        return celdas

    @staticmethod
    def construir_resultado_hoja(
        hoja: str,
        trimestre: str,
        celdas: list[CeldaComparada],
        causa_raiz: Optional[str] = None,
    ) -> ResultadoHoja:
        """Agrega métricas a partir de la lista de celdas."""
        n_cierre = sum(1 for c in celdas if c.estado == "OK_CIERRE")
        n_diag = sum(1 for c in celdas if c.estado == "OK_DIAGNOSTICO")
        n_fail = sum(1 for c in celdas if c.estado == "FAIL")
        n_na = sum(1 for c in celdas if c.estado == "NA")

        # Estado global
        if n_fail == 0 and n_na == 0 and n_diag == 0:
            estado = "CERRADA"
        elif n_fail == 0 and n_diag > 0 and n_na == 0:
            estado = "CERRADA"  # diagnóstico es aceptable como cierre
        elif causa_raiz is not None:
            estado = "ABIERTA_CON_CAUSA"
        else:
            estado = "ABIERTA_SIN_CAUSA"

        return ResultadoHoja(
            hoja=hoja,
            estado=estado,
            trimestre=trimestre,
            n_celdas=len(celdas),
            n_ok_cierre=n_cierre,
            n_ok_diagnostico=n_diag,
            n_fail=n_fail,
            n_na=n_na,
            celdas=celdas,
            causa_raiz=causa_raiz,
        )
