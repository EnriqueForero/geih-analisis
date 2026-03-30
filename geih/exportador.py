# -*- coding: utf-8 -*-
"""
geih.exportador — Gestión de exportaciones con estructura de carpetas.

Organiza TODOS los outputs (gráficas, tablas CSV, Excel) en una
estructura predecible. Nunca deja archivos sueltos en la raíz.

Estructura de salida:
    resultados_geih_{anio}/
    ├── graficas/          ← PNGs de matplotlib
    ├── tablas/            ← CSVs de resultados
    ├── excel/             ← Archivos Excel multi-hoja
    └── metadata.json      ← Parámetros usados en la corrida

Autor: Néstor Enrique Forero Herrera
"""

__all__ = [
    "Exportador",
]


import json
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any

import pandas as pd
import matplotlib.pyplot as plt

from .config import ConfigGEIH


class Exportador:
    """Gestiona la exportación organizada de todos los resultados.

    Crea la estructura de carpetas automáticamente y provee métodos
    para guardar gráficas, tablas y archivos Excel en el lugar correcto.

    Uso típico:
        exp = Exportador(ruta_base='/content/drive/MyDrive/GEIH')
        exp.guardar_grafica(fig, 'BoxPlot_salarios_rama')
        exp.guardar_tabla(df, 'estadisticas_rama_2025')
        exp.guardar_excel({'Hoja1': df1, 'Hoja2': df2}, 'resultados_2025')
        exp.guardar_metadata(config, {'n_registros': 5_000_000})
    """

    def __init__(
        self,
        ruta_base: str = ".",
        nombre_carpeta: Optional[str] = None,
        config: Optional["ConfigGEIH"] = None,
    ):
        """
        Args:
            ruta_base     : Directorio padre donde se crea la carpeta de resultados.
            nombre_carpeta: Nombre de la carpeta raíz. Si None, se auto-genera
                            como 'resultados_geih_{anio}' usando config.
                            Si config también es None, usa 'resultados_geih'.
            config        : ConfigGEIH para inferir el año automáticamente.

        Ejemplos:
            Exportador(ruta_base=RUTA, config=config)
            # → resultados_geih_2025/ si config.anio==2025
            # → resultados_geih_2026/ si config.anio==2026

            Exportador(ruta_base=RUTA, nombre_carpeta="mis_resultados")
            # → mis_resultados/ (nombre explícito, compatible con v4.x)
        """
        if nombre_carpeta is None:
            anio = config.anio if config is not None else ""
            nombre_carpeta = f"resultados_geih_{anio}" if anio else "resultados_geih"
        self.raiz = Path(ruta_base) / nombre_carpeta
        self.graficas = self.raiz / "graficas"
        self.tablas = self.raiz / "tablas"
        self.excel = self.raiz / "excel"

        # Crear toda la estructura
        for carpeta in [self.graficas, self.tablas, self.excel]:
            carpeta.mkdir(parents=True, exist_ok=True)

        print(f"📂 Exportador listo → {self.raiz}")

    def guardar_grafica(
        self,
        fig: plt.Figure,
        nombre: str,
        dpi: int = 160,
        fondo: str = "#F7F9FC",
        cerrar: bool = True,
    ) -> Path:
        """Guarda una figura matplotlib como PNG.

        Args:
            fig: Figura de matplotlib.
            nombre: Nombre del archivo (sin extensión).
            dpi: Resolución (160 es bueno para presentaciones).
            fondo: Color de fondo del archivo exportado.
            cerrar: Si True, cierra la figura después de guardar (libera RAM).

        Returns:
            Path al archivo guardado.
        """
        ruta = self.graficas / f"{nombre}.png"
        fig.savefig(ruta, dpi=dpi, bbox_inches="tight", facecolor=fondo)
        if cerrar:
            plt.close(fig)
        print(f"   📊 {ruta.name}")
        return ruta

    def guardar_tabla(
        self,
        df: pd.DataFrame,
        nombre: str,
        sep: str = ";",
        encoding: str = "utf-8-sig",
    ) -> Path:
        """Guarda un DataFrame como CSV.

        Args:
            df: DataFrame a exportar.
            nombre: Nombre del archivo (sin extensión).

        Returns:
            Path al archivo guardado.
        """
        ruta = self.tablas / f"{nombre}.csv"
        df.to_csv(ruta, index=False, sep=sep, encoding=encoding)
        print(f"   📋 {ruta.name} ({len(df):,} filas)")
        return ruta

    def guardar_excel(
        self,
        hojas: Dict[str, pd.DataFrame],
        nombre: str,
        formato_institucional: bool = True,
    ) -> Path:
        """Guarda múltiples DataFrames como hojas de un Excel con formato.

        Aplica formato institucional:
        - Headers con color burdeos corporativo
        - Filas alternas
        - Ancho de columnas automático
        - Primera fila congelada

        Args:
            hojas: Dict {nombre_hoja: DataFrame}.
            nombre: Nombre del archivo (sin extensión).
            formato_institucional: Si True, aplica formato institucional con estilos predefinidos.

        Returns:
            Path al archivo guardado.
        """
        ruta = self.excel / f"{nombre}.xlsx"

        with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
            for hoja, df in hojas.items():
                if df is None or len(df) == 0:
                    continue
                df.to_excel(writer, sheet_name=hoja[:31], index=False)

            if formato_institucional:
                self._aplicar_formato(writer)

        print(f"   📗 {ruta.name} ({len(hojas)} hojas)")
        return ruta

    @staticmethod
    def _aplicar_formato(writer) -> None:
        """Aplica formato institucional a todas las hojas."""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return  # openpyxl no disponible, salir silenciosamente

        header_fill = PatternFill("solid", fgColor="8B1A4A")  # burdeos corporativo
        header_font = Font(bold=True, color="FFFFFF", size=10)
        fill_alt = PatternFill("solid", fgColor="F5E8EF")      # rosa claro alterno

        for ws in writer.sheets.values():
            # Headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Ancho automático de columnas
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = 0
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

            # Filas alternas
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if i % 2 == 0:
                    for cell in row:
                        cell.fill = fill_alt

            # Congelar primera fila
            ws.freeze_panes = "A2"

    def guardar_metadata(
        self,
        config: ConfigGEIH,
        stats: Optional[Dict[str, Any]] = None,
    ) -> Path:
        """Guarda metadata de la corrida para trazabilidad.

        Args:
            config: Configuración usada.
            stats: Estadísticas adicionales (n_registros, etc.)

        Returns:
            Path al archivo metadata.json.
        """
        meta = {
            "timestamp": datetime.now().isoformat(),
            "config": {
                "n_meses": config.n_meses,
                "smmlv": config.smmlv,
                "periodo": config.periodo_etiqueta,
                "random_seed": config.random_seed,
            },
            "stats": stats or {},
        }
        ruta = self.raiz / "metadata.json"
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2, ensure_ascii=False)
        print(f"   📋 metadata.json")
        return ruta

    def resumen(self) -> None:
        """Imprime un resumen de todos los archivos exportados."""
        print(f"\n{'='*60}")
        print(f"  RESUMEN DE EXPORTACIÓN")
        print(f"  {self.raiz}")
        print(f"{'='*60}")
        for carpeta, icono in [
            (self.graficas, "📊"),
            (self.tablas, "📋"),
            (self.excel, "📗"),
        ]:
            archivos = sorted(carpeta.glob("*"))
            if archivos:
                print(f"\n  {icono} {carpeta.name}/ ({len(archivos)} archivos)")
                for a in archivos:
                    size = a.stat().st_size / 1024
                    print(f"     {a.name:<50} {size:>7.0f} KB")
        # Metadata
        meta_path = self.raiz / "metadata.json"
        if meta_path.exists():
            print(f"\n  📋 metadata.json")
