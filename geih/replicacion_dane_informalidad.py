"""
geih.replicacion_dane_informalidad — Replicación de las 10 hojas de
informalidad del anexo oficial DANE `anex-GEIHEISS-*.xlsx`.

Hojas replicadas (en orden de prioridad del plan IT1 v2):

    1. Total nacional
    2. Grandes dominios       (mensual)
    3. Prop informalidad
    4. Ciudades
    5. Sexo
    6. Educación
    7. Ramas de actividad CIIU 4 A.C
    8. Posición ocupacional
    9. Lugar de trabajo
    10. Tamaño de empresa

Todas las cifras se calculan desde microdatos GEIH preparados. El
módulo no hardcodea ninguna cifra del DANE; las etiquetas y estructuras
se leen del propio Excel usando `ExcelLayoutParser`.

Convenciones del DataFrame de entrada:
- `FEX_ADJ`: factor de expansión ajustado por n meses del período.
- `OCI == 1`: población ocupada.
- `INFORMAL`: Int8 (1=informal, 0=formal, <NA>).
- `DOMINIO`: string con valores {'13_AM', '10_ciudades', 'otras_cab',
  'rural'}.
- `CIUDAD`: nombre canónico de ciudad (13 A.M. + 10 intermedias).
- `SEXO`: 'Hombres' | 'Mujeres'.
- `RAMA`: nombre de rama según `RAMAS_DANE` del config.
- `POSICION_OCU`: nombre canónico según `POSICION_OCUPACIONAL`.
- `NIVEL_GRUPO`: nivel agrupado según `NIVELES_AGRUPADOS`.
- `MES_NUM`, `PER`: año/mes para filtrar Grandes dominios.
- `P3069`: tamaño de empresa (código 1-10).
- `P6880`: lugar de trabajo (código 1-11).

Autor: Néstor Enrique Forero Herrera · ProColombia · 2026-04-16
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .replicacion_dane_common import (
    BloqueLayout,
    BloqueReplicado,
    CeldaComparada,
    ExcelLayoutParser,
    ParametrosValidacion,
    PeriodoMovil,
    ResultadoHoja,
    ResultadoReplicacion,
    ValidadorParidad,
    hash_archivo,
    normalizar_texto,
)

__all__ = [
    "ReplicadorInformalidad",
    "replicar_informalidad_excel",
    "HOJAS_INFORMALIDAD",
    "MAPEO_DOMINIO_EXCEL",
    "MAPEO_NIVEL_EDU",
    "MAPEO_TAMANO_EMPRESA",
    "MAPEO_LUGAR_TRABAJO",
]


# ═════════════════════════════════════════════════════════════════════
# Constantes y mapeos
# ═════════════════════════════════════════════════════════════════════

HOJAS_INFORMALIDAD: list[str] = [
    "Total nacional",
    "Grandes dominios",  # con espacio al final en el Excel
    "Prop informalidad",
    "Ciudades",
    "Sexo",
    "Educación",  # con espacio al final en el Excel
    "Ramas de actividad CIIU 4 A.C",
    "Posición ocupacional",
    "Lugar de trabajo",
    "Tamaño de empresa",
]

# Nombres alternativos que aparecen con espacios trailing en el Excel
ALIAS_HOJA: dict[str, str] = {
    "Grandes dominios": "Grandes dominios ",
    "Educación": "Educación ",
}


# DOMINIO del Excel → función que aplica filtro sobre df["DOMINIO"]
# Devuelve una máscara booleana del largo de df.
MAPEO_DOMINIO_EXCEL: dict[str, callable] = {
    "total nacional": lambda s: s.notna(),  # todos los dominios
    "cabeceras": lambda s: s.isin(["13_AM", "10_ciudades", "otras_cab"]),
    "centros poblados y rural disperso": lambda s: s == "rural",
    "13 ciudades y a.m.": lambda s: s == "13_AM",
    "13 ciudades y areas metropolitanas": lambda s: s == "13_AM",
    "23 ciudades y a.m.": lambda s: s.isin(["13_AM", "10_ciudades"]),
    "23 ciudades y areas metropolitanas": lambda s: s.isin(["13_AM", "10_ciudades"]),
    "10 ciudades": lambda s: s == "10_ciudades",
}


# Nivel educativo del preparador → etiqueta Excel
MAPEO_NIVEL_EDU: dict[str, str] = {
    "1. Sin educación": "Ninguno",
    "2. Primaria": "Básica primaria",
    "3. Secundaria": "Básica secundaria^",
    "4. Media": "Educación media*^",
    "5. Técnica/Tecno.": "Técnica profesional y Tecnológica^^",
    "6. Universitaria": "Universitaria",
    "7. Posgrado": "Posgrado",
}


# Tamaño de empresa: P3069 → etiqueta Excel (categorización DANE)
# Micro: 1-10 personas (P3069 1..4)
# Pequeña: 11-50 (P3069 5..7)
# Mediana: 51-200 (P3069 8..9)
# Grande: 201+ (P3069 10)
MAPEO_TAMANO_EMPRESA: dict[int, str] = {
    1: "Microempresa",
    2: "Microempresa",
    3: "Microempresa",
    4: "Microempresa",
    5: "Empresa pequeña",
    6: "Empresa pequeña",
    7: "Empresa pequeña",
    8: "Empresa mediana",
    9: "Empresa mediana",
    10: "Empresa grande",
}


# Lugar de trabajo: P6880 → etiqueta Excel
MAPEO_LUGAR_TRABAJO: dict[int, str] = {
    1: "En esta vivienda",
    2: "En otras viviendas",
    3: "En kiosco - caseta",
    4: "En un vehículo",
    5: "De puerta en puerta",
    6: "Sitio al descubierto en la calle",
    7: "Local fijo, oficina, fábrica, etc",
    8: "En el campo o área rural",
    9: "En una obra en construcción",
    10: "En una mina o cantera",
    11: "Otro",
}


# ═════════════════════════════════════════════════════════════════════
# Replicador principal
# ═════════════════════════════════════════════════════════════════════


class ReplicadorInformalidad:
    """Replica las 10 hojas de informalidad del Excel oficial.

    Uso:
        rep = ReplicadorInformalidad(ruta_excel)
        resultado = rep.replicar(df_preparado, periodo)
        print(resultado.reporte_consolidado())
    """

    def __init__(
        self,
        ruta_excel: Path,
        parametros: Optional[ParametrosValidacion] = None,
    ):
        self.ruta_excel = Path(ruta_excel)
        self.params = parametros or ParametrosValidacion()
        self.validador = ValidadorParidad(self.params)

    # ────────────────────────────────────────────────────────────────
    # API principal
    # ────────────────────────────────────────────────────────────────

    def replicar(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        hojas: Optional[list[str]] = None,
    ) -> ResultadoReplicacion:
        """Replica las hojas y compara contra el Excel.

        Args:
            df: DataFrame preparado con las columnas del preparador.
                Debe contener los 3 meses del trimestre para hojas
                trimestrales. Para Grandes dominios se filtrará el mes
                final internamente.
            periodo: `PeriodoMovil` con los meses del trimestre objetivo.
            hojas: subconjunto de hojas a replicar. Si None, todas.

        Returns:
            `ResultadoReplicacion` con los resultados por hoja.
        """
        self._verificar_columnas(df)

        if hojas is None:
            hojas = HOJAS_INFORMALIDAD

        resultados: list[ResultadoHoja] = []
        wb = load_workbook(self.ruta_excel, read_only=False, data_only=True)
        try:
            for hoja in hojas:
                try:
                    res = self._replicar_hoja(df, periodo, hoja, wb)
                    resultados.append(res)
                    print(res.resumen())
                except Exception as e:
                    print(f"❌ Error replicando '{hoja}': {e}")
                    res = ResultadoHoja(
                        hoja=hoja,
                        estado="ABIERTA_CON_CAUSA",
                        trimestre=periodo.etiqueta_original,
                        causa_raiz=f"Excepción: {type(e).__name__}: {e}",
                    )
                    resultados.append(res)
        finally:
            wb.close()

        from datetime import datetime as _dt

        return ResultadoReplicacion(
            ruta_excel=self.ruta_excel,
            hash_excel=hash_archivo(self.ruta_excel),
            timestamp=_dt.now(),
            trimestre=periodo,
            hojas=resultados,
            parametros=self.params,
        )

    # ────────────────────────────────────────────────────────────────
    # Dispatcher por hoja
    # ────────────────────────────────────────────────────────────────

    def _replicar_hoja(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        hoja: str,
        wb,
    ) -> ResultadoHoja:
        """Dispatcher: delega a método específico por hoja."""
        nombre_real = ALIAS_HOJA.get(hoja, hoja)
        if nombre_real not in wb.sheetnames:
            return ResultadoHoja(
                hoja=hoja,
                estado="ABIERTA_CON_CAUSA",
                trimestre=periodo.etiqueta_original,
                causa_raiz=f"Hoja '{nombre_real}' no existe en Excel",
            )
        ws = wb[nombre_real]

        if hoja == "Total nacional":
            return self._replicar_total_nacional(df, periodo, ws, nombre_real)
        if hoja == "Grandes dominios":
            return self._replicar_grandes_dominios(df, periodo, ws, nombre_real)
        if hoja == "Prop informalidad":
            return self._replicar_prop_informalidad(df, periodo, ws, nombre_real)
        if hoja == "Ciudades":
            return self._replicar_ciudades(df, periodo, ws, nombre_real)
        if hoja == "Sexo":
            return self._replicar_sexo(df, periodo, ws, nombre_real)
        if hoja == "Educación":
            return self._replicar_educacion(df, periodo, ws, nombre_real)
        if hoja == "Ramas de actividad CIIU 4 A.C":
            return self._replicar_ramas(df, periodo, ws, nombre_real)
        if hoja == "Posición ocupacional":
            return self._replicar_posicion(df, periodo, ws, nombre_real)
        if hoja == "Lugar de trabajo":
            return self._replicar_lugar(df, periodo, ws, nombre_real)
        if hoja == "Tamaño de empresa":
            return self._replicar_tamano(df, periodo, ws, nombre_real)

        return ResultadoHoja(
            hoja=hoja,
            estado="ABIERTA_CON_CAUSA",
            trimestre=periodo.etiqueta_original,
            causa_raiz="Hoja no soportada por este replicador",
        )

    # ────────────────────────────────────────────────────────────────
    # Helpers numéricos
    # ────────────────────────────────────────────────────────────────

    @staticmethod
    def _verificar_columnas(df: pd.DataFrame) -> None:
        """Fail fast: verifica columnas críticas del DataFrame preparado."""
        requeridas = {"FEX_ADJ", "OCI", "INFORMAL", "DOMINIO"}
        faltan = requeridas - set(df.columns)
        if faltan:
            raise ValueError(
                f"DataFrame no tiene columnas requeridas: {sorted(faltan)}. "
                f"¿Aplicó PreparadorGEIH.preparar_base(derivar=True)?"
            )

    def _mask_dominio(self, df: pd.DataFrame, dominio_excel: str) -> pd.Series:
        """Devuelve máscara booleana para el dominio del Excel."""
        key = normalizar_texto(dominio_excel)
        if key in MAPEO_DOMINIO_EXCEL:
            return MAPEO_DOMINIO_EXCEL[key](df["DOMINIO"])
        # Dominios por ciudad (Bogotá D.C., Medellín A.M., etc.) van
        # por CIUDAD == <nombre>. El matching usa normalizacion.
        if "CIUDAD" in df.columns:
            ciudad_norm = df["CIUDAD"].fillna("").astype(str).map(normalizar_texto)
            # Remover " a.m." del Excel para comparar con CIUDAD
            key_ciudad = key.replace(" a.m.", "").strip()
            mask = ciudad_norm == key_ciudad
            if mask.any():
                return mask
        # Default: nadie
        return pd.Series(False, index=df.index)

    @staticmethod
    def _suma_fex(df_sub: pd.DataFrame) -> float:
        """Suma ponderada en miles."""
        if len(df_sub) == 0:
            return 0.0
        return float(df_sub["FEX_ADJ"].sum()) / 1000.0

    def _total_formal_informal(
        self,
        df: pd.DataFrame,
        mask_dominio: pd.Series,
    ) -> tuple[float, float, float]:
        """Devuelve (pob_ocupada, formal, informal) en miles para un dominio."""
        df_ocu = df[mask_dominio & (df["OCI"] == 1)]
        total = self._suma_fex(df_ocu)
        formal = self._suma_fex(df_ocu[df_ocu["INFORMAL"] == 0])
        informal = self._suma_fex(df_ocu[df_ocu["INFORMAL"] == 1])
        return total, formal, informal

    # ────────────────────────────────────────────────────────────────
    # Replicadores por hoja
    # ────────────────────────────────────────────────────────────────

    def _replicar_total_nacional(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        """Replica hoja 'Total nacional': 3 dominios × 3 valores (Total, Formal, Informal)."""
        parser = ExcelLayoutParser(self.ruta_excel)
        with parser:
            layout = parser.parsear_bloques_por_dominio(hoja_nombre)
        col = layout.col_ultimo_trimestre
        celdas: list[CeldaComparada] = []

        # Para cada bloque de dominio, calcular los 3 valores
        for b in layout.bloques:
            mask = self._mask_dominio(df, b.dominio)
            total, formal, informal = self._total_formal_informal(df, mask)
            valores = {
                "Población ocupada": total,
                "Formal": formal,
                "Informal": informal,
            }
            bloque_rep = BloqueReplicado(
                hoja="Total nacional",
                dominio=b.dominio,
                condicion="Miles",
                valores=valores,
                es_proporcion=False,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    b,
                    ws,
                    col,
                    layout.etiqueta_ultimo_trimestre,
                )
            )

        return self.validador.construir_resultado_hoja(
            hoja="Total nacional",
            trimestre=layout.etiqueta_ultimo_trimestre,
            celdas=celdas,
        )

    def _replicar_grandes_dominios(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        """Replica 'Grandes dominios' (mensual): filtra mes final del período.

        El Excel presenta 5 dominios mensualmente: Total nacional,
        13 A.M., 10 ciudades, 23 A.M., Rural. Como los microdatos llegan
        con los 3 meses del trimestre, debemos filtrar solo el mes final.
        Además el divisor FEX_ADJ fue hecho para 3 meses; al filtrar a
        un solo mes hay que **multiplicar por 3** para recuperar el FEX
        mensual.
        """
        if "MES_NUM" not in df.columns or "PER" not in df.columns:
            return ResultadoHoja(
                hoja="Grandes dominios",
                estado="ABIERTA_CON_CAUSA",
                trimestre=periodo.etiqueta_original,
                causa_raiz="df no tiene MES_NUM/PER (requeridas para filtrar mes)",
            )

        anio_f, mes_f = periodo.mes_final
        mask_mes = (df["MES_NUM"] == mes_f) & (df["PER"] == anio_f)
        if not mask_mes.any():
            return ResultadoHoja(
                hoja="Grandes dominios",
                estado="ABIERTA_CON_CAUSA",
                trimestre=periodo.etiqueta_original,
                causa_raiz=f"No hay datos para {anio_f}-{mes_f} en df",
            )
        # Re-escalar FEX: el FEX_ADJ viene dividido por 3 meses, aquí
        # queremos el FEX mensual del mes final
        df_mes = df[mask_mes].copy()
        df_mes["FEX_ADJ"] = df_mes["FEX_ADJ"] * periodo.n_meses

        parser = ExcelLayoutParser(self.ruta_excel)
        with parser:
            layout = parser.parsear_bloques_por_dominio(hoja_nombre)
        col = layout.col_ultimo_trimestre

        celdas: list[CeldaComparada] = []
        for b in layout.bloques:
            mask = self._mask_dominio(df_mes, b.dominio)
            total, formal, informal = self._total_formal_informal(df_mes, mask)
            valores = {
                "Población ocupada": total,
                "Formal": formal,
                "Informal": informal,
            }
            bloque_rep = BloqueReplicado(
                hoja="Grandes dominios",
                dominio=b.dominio,
                condicion="Miles",
                valores=valores,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    b,
                    ws,
                    col,
                    layout.etiqueta_ultimo_trimestre,
                )
            )

        return self.validador.construir_resultado_hoja(
            hoja="Grandes dominios",
            trimestre=layout.etiqueta_ultimo_trimestre,
            celdas=celdas,
        )

    def _replicar_prop_informalidad(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        """Hoja 'Prop informalidad': % informalidad por dominio/ciudad."""
        # La estructura es una sola columna A con la lista:
        #   Total nacional
        #   13 Ciudades y A.M.
        #   23 Ciudades y A.M.
        #   Bogotá D.C.
        #   Medellín A.M.
        #   ... (23 ciudades)
        col_ult = self._encontrar_col_ult(ws, fila_meses=12)
        celdas: list[CeldaComparada] = []
        etiqueta_trim = str(ws.cell(12, col_ult).value or "").strip()

        # Leer las categorías de col A, filas 13 en adelante
        categorias: list[tuple[int, str]] = []
        for r in range(13, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is None:
                # Posible fin de bloque
                continue
            s = str(v).strip()
            if not s or s.lower().startswith(("fuente", "nota", "actualizado")):
                break
            categorias.append((r, s))

        valores_calc: dict[str, float] = {}
        for _fila, etq in categorias:
            mask = self._mask_dominio_prop_informalidad(df, etq)
            if not mask.any():
                continue
            df_ocu = df[mask & (df["OCI"] == 1)]
            total = self._suma_fex(df_ocu)
            informal = self._suma_fex(df_ocu[df_ocu["INFORMAL"] == 1])
            prop = 100 * informal / total if total > 0 else np.nan
            valores_calc[etq] = prop

        # Construir bloque y layout sintético
        bloque_layout = BloqueLayout(
            hoja="Prop informalidad",
            dominio="(todos)",
            condicion="%",
            fila_titulo=11,
            fila_encabezado=11,
            fila_meses=12,
            fila_inicio_datos=13,
            fila_fin_datos=categorias[-1][0] if categorias else 13,
            col_categoria=1,
            categorias=categorias,
        )
        bloque_rep = BloqueReplicado(
            hoja="Prop informalidad",
            dominio="(todos)",
            condicion="%",
            valores=valores_calc,
            es_proporcion=True,
        )
        celdas = self.validador.comparar_bloque(
            bloque_rep,
            bloque_layout,
            ws,
            col_ult,
            etiqueta_trim,
        )
        return self.validador.construir_resultado_hoja(
            hoja="Prop informalidad",
            trimestre=etiqueta_trim,
            celdas=celdas,
        )

    def _mask_dominio_prop_informalidad(
        self,
        df: pd.DataFrame,
        etiqueta: str,
    ) -> pd.Series:
        """Resuelve la máscara de dominio para una fila de Prop informalidad.

        Acepta los 3 agregados y las 23 ciudades individuales.
        """
        key = normalizar_texto(etiqueta)
        # Agregados
        if key in MAPEO_DOMINIO_EXCEL:
            return MAPEO_DOMINIO_EXCEL[key](df["DOMINIO"])
        # Ciudad: "Bogotá D.C." → CIUDAD == "Bogotá D.C."
        # "Medellín A.M." → CIUDAD == "Medellín A.M."
        if "CIUDAD" not in df.columns:
            return pd.Series(False, index=df.index)
        ciudad_norm = df["CIUDAD"].fillna("").astype(str).map(normalizar_texto)
        return ciudad_norm == key

    @staticmethod
    def _encontrar_col_ult(ws, fila_meses: int = 13) -> int:
        """Encuentra última columna con dato en fila_meses."""
        max_col = ws.max_column
        for c in range(max_col, 1, -1):
            v = ws.cell(fila_meses, c).value
            if v is not None and str(v).strip():
                return c
        raise ValueError(f"Sin datos en fila {fila_meses}")

    def _replicar_ciudades(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        """Hoja 'Ciudades': col A=dominio/ciudad, col B=Total/Formal/Informal.

        Estructura:
            A         | B               | ...
            13 ciud.. | Población ocupada
                      | Formal
                      | Informal
            Bogotá    | Población ocupada
                      | Formal
                      | Informal
            ...
        """
        col_ult = self._encontrar_col_ult(ws, fila_meses=12)
        etiqueta_trim = str(ws.cell(12, col_ult).value or "").strip()

        # Escanear filas 13+ para construir tripletes (dominio, fila_total,
        # fila_formal, fila_informal)
        dominios_tripletes: list[tuple[str, int, int, int]] = []
        dominio_actual: Optional[str] = None
        fila_total = fila_formal = fila_informal = 0
        for r in range(13, ws.max_row + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if a is not None and str(a).strip():
                s = str(a).strip()
                if s.lower().startswith(("fuente", "nota", "actualizado")):
                    break
                dominio_actual = s
            if b is not None:
                cat = normalizar_texto(b)
                if cat == "poblacion ocupada":
                    fila_total = r
                elif cat == "formal":
                    fila_formal = r
                elif cat == "informal":
                    fila_informal = r
                    if dominio_actual and fila_total and fila_formal:
                        dominios_tripletes.append(
                            (dominio_actual, fila_total, fila_formal, fila_informal)
                        )
                        fila_total = fila_formal = 0

        celdas: list[CeldaComparada] = []
        for dominio, f_t, f_f, f_i in dominios_tripletes:
            mask = self._mask_dominio_prop_informalidad(df, dominio)
            if not mask.any():
                # registrar NA
                for fila, cat in [(f_t, "Población ocupada"), (f_f, "Formal"), (f_i, "Informal")]:
                    v = ws.cell(fila, col_ult).value
                    celdas.append(
                        CeldaComparada(
                            hoja="Ciudades",
                            dominio=dominio,
                            condicion="Miles",
                            categoria=cat,
                            trimestre=etiqueta_trim,
                            valor_excel=float(v) if v is not None else None,
                            valor_calculado=None,
                            diferencia=None,
                            estado="NA",
                        )
                    )
                continue
            df_ocu = df[mask & (df["OCI"] == 1)]
            total = self._suma_fex(df_ocu)
            formal = self._suma_fex(df_ocu[df_ocu["INFORMAL"] == 0])
            informal = self._suma_fex(df_ocu[df_ocu["INFORMAL"] == 1])
            for fila, cat, valor_calc in [
                (f_t, "Población ocupada", total),
                (f_f, "Formal", formal),
                (f_i, "Informal", informal),
            ]:
                v = ws.cell(fila, col_ult).value
                try:
                    ve = float(v) if v is not None else None
                except (TypeError, ValueError):
                    ve = None
                estado, diff = self.params.clasificar(ve, valor_calc, False)
                celdas.append(
                    CeldaComparada(
                        hoja="Ciudades",
                        dominio=dominio,
                        condicion="Miles",
                        categoria=cat,
                        trimestre=etiqueta_trim,
                        valor_excel=ve,
                        valor_calculado=valor_calc,
                        diferencia=diff,
                        estado=estado,
                    )
                )

        return self.validador.construir_resultado_hoja(
            hoja="Ciudades",
            trimestre=etiqueta_trim,
            celdas=celdas,
        )

    def _replicar_sexo(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        """Hoja 'Sexo': por dominio, bloques Total/Formal/Informal + Hombres/Mujeres con
        sub-bloques por sexo.

        Estructura observada:
            Total nacional
              Población ocupada (total)
              Formal
              Informal
              Hombres  (población ocupada hombres)
                Formal (formal hombres)
                Informal (informal hombres)
              Mujeres (población ocupada mujeres)
                Formal
                Informal
            13 Ciudades y A.M.
              ... (idéntica estructura)
            23 Ciudades y A.M.
              ...
            Centros poblados y rural disperso
              ...
        """
        col_ult = self._encontrar_col_ult(ws, fila_meses=13)
        etiqueta_trim = str(ws.cell(13, col_ult).value or "").strip()

        # Dominios objetivo en esta hoja (en orden del Excel)
        dominios_a_leer = [
            "Total nacional",
            "13 Ciudades y áreas metropolitanas",
            "23 Ciudades y áreas metropolitanas",
            "Centros poblados y rural disperso",
        ]

        celdas: list[CeldaComparada] = []
        # Escanear manualmente la hoja por bloques de dominio
        n_filas = ws.max_row
        fila = 1
        dominio_idx = 0
        while fila <= n_filas and dominio_idx < len(dominios_a_leer):
            v = ws.cell(fila, 1).value
            if v is None:
                fila += 1
                continue
            s_norm = normalizar_texto(v)
            dominio_esperado = dominios_a_leer[dominio_idx]
            if s_norm == normalizar_texto(dominio_esperado):
                # Este es el inicio del bloque. Buscar las filas dentro:
                celdas.extend(
                    self._leer_bloque_sexo(df, ws, col_ult, etiqueta_trim, fila, dominio_esperado)
                )
                dominio_idx += 1
                fila += 12  # saltar hacia abajo (tamaño típico del bloque)
            else:
                fila += 1

        return self.validador.construir_resultado_hoja(
            hoja="Sexo",
            trimestre=etiqueta_trim,
            celdas=celdas,
        )

    def _leer_bloque_sexo(
        self,
        df,
        ws,
        col_ult,
        etiqueta_trim,
        fila_titulo,
        dominio,
    ) -> list[CeldaComparada]:
        """Lee un bloque de la hoja Sexo y calcula las 9 celdas esperadas."""
        # Las 9 celdas (valores + labels en col A) empiezan 3 filas debajo del título
        # (R+1 Concepto, R+2 meses, R+3 Población ocupada)
        # Mapear qué fila tiene qué etiqueta
        mapa_filas: dict[str, int] = {}
        sexo_actual: Optional[str] = None
        for r in range(fila_titulo + 3, fila_titulo + 15):
            if r > ws.max_row:
                break
            v = ws.cell(r, 1).value
            if v is None:
                continue
            s = str(v).strip()
            s_norm = normalizar_texto(s)
            if s_norm == "poblacion ocupada":
                mapa_filas["Población ocupada (total)"] = r
                sexo_actual = None
            elif s_norm == "formal":
                if sexo_actual is None:
                    mapa_filas["Formal"] = r
                else:
                    mapa_filas[f"Formal - {sexo_actual}"] = r
            elif s_norm == "informal":
                if sexo_actual is None:
                    mapa_filas["Informal"] = r
                else:
                    mapa_filas[f"Informal - {sexo_actual}"] = r
            elif s_norm == "hombres":
                mapa_filas["Hombres"] = r
                sexo_actual = "Hombres"
            elif s_norm == "mujeres":
                mapa_filas["Mujeres"] = r
                sexo_actual = "Mujeres"
            elif s_norm.startswith(("fuente", "nota", "actualizado")):
                break
            elif not s:
                # línea vacía — puede ser separador inter-dominio
                # Si ya tenemos valores, salir
                if len(mapa_filas) >= 9:
                    break
                continue

        mask = self._mask_dominio(df, dominio)
        df_ocu = df[mask & (df["OCI"] == 1)]

        def _v(submask):
            return self._suma_fex(df_ocu[submask])

        # Calcular los 9 valores
        valores_calc = {
            "Población ocupada (total)": _v(df_ocu.index == df_ocu.index),  # todos
            "Formal": _v(df_ocu["INFORMAL"] == 0),
            "Informal": _v(df_ocu["INFORMAL"] == 1),
            "Hombres": _v(df_ocu["SEXO"] == "Hombres"),
            "Formal - Hombres": _v((df_ocu["SEXO"] == "Hombres") & (df_ocu["INFORMAL"] == 0)),
            "Informal - Hombres": _v((df_ocu["SEXO"] == "Hombres") & (df_ocu["INFORMAL"] == 1)),
            "Mujeres": _v(df_ocu["SEXO"] == "Mujeres"),
            "Formal - Mujeres": _v((df_ocu["SEXO"] == "Mujeres") & (df_ocu["INFORMAL"] == 0)),
            "Informal - Mujeres": _v((df_ocu["SEXO"] == "Mujeres") & (df_ocu["INFORMAL"] == 1)),
        }

        # Comparar contra Excel
        celdas: list[CeldaComparada] = []
        for categoria, fila in mapa_filas.items():
            v = ws.cell(fila, col_ult).value
            try:
                ve = float(v) if v is not None else None
            except (TypeError, ValueError):
                ve = None
            vc = valores_calc.get(categoria)
            estado, diff = self.params.clasificar(ve, vc, False)
            celdas.append(
                CeldaComparada(
                    hoja="Sexo",
                    dominio=dominio,
                    condicion="Miles",
                    categoria=categoria,
                    trimestre=etiqueta_trim,
                    valor_excel=ve,
                    valor_calculado=vc,
                    diferencia=diff,
                    estado=estado,
                )
            )
        return celdas

    # ── Hojas estándar con parser por dominio y condición ────────────

    def _replicar_educacion(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        return self._replicar_con_categorias(
            df,
            periodo,
            ws,
            hoja_nombre,
            "Educación",
            columna_categoria="NIVEL_GRUPO",
            mapeo_valor_a_excel=MAPEO_NIVEL_EDU,
        )

    def _replicar_ramas(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        """Ramas tiene una peculiaridad: en 13/23 A.M. agrupa varias ramas
        pequeñas en "Otras ramas*"."""
        hoja = "Ramas de actividad CIIU 4 A.C"
        parser = ExcelLayoutParser(self.ruta_excel)
        with parser:
            layout = parser.parsear_bloques_por_dominio_y_condicion(hoja_nombre)
        col = layout.col_ultimo_trimestre
        celdas: list[CeldaComparada] = []

        # Categorías que son "Otras ramas*" agrupan estas:
        otras_ramas_miembros = {
            normalizar_texto("Agricultura, ganadería, caza, silvicultura y pesca"),
            normalizar_texto("Suministro de electricidad, gas, agua y gestión de desechos"),
        }

        for b in layout.bloques:
            mask = self._mask_dominio(df, b.dominio)
            df_ocu = df[mask & (df["OCI"] == 1)]
            if b.condicion == "Total":
                df_sub = df_ocu
            elif b.condicion == "Formal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 0]
            elif b.condicion == "Informal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 1]
            else:
                continue

            valores: dict[str, float] = {}
            # Detectar si el bloque tiene "Otras ramas*"
            {normalizar_texto(e) for _, e in b.categorias}

            for _fila_xl, etq in b.categorias:
                etq_norm = normalizar_texto(etq)
                if etq_norm == "no informa":
                    # RAMA es None para estos
                    if "RAMA" in df_sub.columns:
                        valor = self._suma_fex(df_sub[df_sub["RAMA"].isna()])
                    else:
                        valor = 0.0
                    valores[etq] = valor
                elif etq_norm == "otras ramas":
                    # Agrupar AGRI + SUMI
                    if "RAMA" in df_sub.columns:
                        serie_rama_norm = (
                            df_sub["RAMA"].fillna("").astype(str).map(normalizar_texto)
                        )
                        mask_otras = serie_rama_norm.isin(otras_ramas_miembros)
                        valor = self._suma_fex(df_sub[mask_otras])
                    else:
                        valor = 0.0
                    valores[etq] = valor
                else:
                    # Match directo por nombre de rama
                    if "RAMA" in df_sub.columns:
                        serie_rama_norm = (
                            df_sub["RAMA"].fillna("").astype(str).map(normalizar_texto)
                        )
                        valor = self._suma_fex(df_sub[serie_rama_norm == etq_norm])
                    else:
                        valor = 0.0
                    valores[etq] = valor

            bloque_rep = BloqueReplicado(
                hoja=hoja,
                dominio=b.dominio,
                condicion=b.condicion,
                valores=valores,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    b,
                    ws,
                    col,
                    layout.etiqueta_ultimo_trimestre,
                )
            )

        return self.validador.construir_resultado_hoja(
            hoja=hoja,
            trimestre=layout.etiqueta_ultimo_trimestre,
            celdas=celdas,
        )

    def _replicar_posicion(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        """Posición ocupacional: en A.M. agrupa 'Jornalero' dentro de 'Obrero particular^'."""
        hoja = "Posición ocupacional"
        parser = ExcelLayoutParser(self.ruta_excel)
        with parser:
            layout = parser.parsear_bloques_por_dominio_y_condicion(hoja_nombre)
        col = layout.col_ultimo_trimestre

        # Nombres canónicos del preparador
        nombre_obrero_particular = "obrero, empleado particular"
        nombre_jornalero = "jornalero o peon"  # normalizado (sin tilde)
        nombre_otro = "otro"
        nombre_tfsr_otros = "trabajador sin remuneracion en empresas de otros hogares"

        celdas: list[CeldaComparada] = []
        for b in layout.bloques:
            mask = self._mask_dominio(df, b.dominio)
            df_ocu = df[mask & (df["OCI"] == 1)]
            if b.condicion == "Total":
                df_sub = df_ocu
            elif b.condicion == "Formal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 0]
            elif b.condicion == "Informal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 1]
            else:
                continue

            # Detectar si este bloque agrupa Jornalero en Obrero^
            {normalizar_texto(e) for _, e in b.categorias}
            agrupa_jornalero = any("^" in e for _, e in b.categorias)

            serie_pos_norm = df_sub["POSICION_OCU"].fillna("").astype(str).map(normalizar_texto)

            valores: dict[str, float] = {}
            for _fila_xl, etq in b.categorias:
                etq_norm = normalizar_texto(etq)
                if etq_norm in (
                    "obrero, empleado particular",
                    "empleado particular",  # variantes
                ):
                    mask_cat = serie_pos_norm == nombre_obrero_particular
                    if agrupa_jornalero:
                        mask_cat = mask_cat | (serie_pos_norm == nombre_jornalero)
                    valores[etq] = self._suma_fex(df_sub[mask_cat])
                elif etq_norm == "otro":
                    # DANE agrupa TFSR en empresas de otros hogares (P6430=8) con "Otro" (P6430=9)
                    mask_otro = (serie_pos_norm == nombre_otro) | (
                        serie_pos_norm == nombre_tfsr_otros
                    )
                    valores[etq] = self._suma_fex(df_sub[mask_otro])
                else:
                    # Match directo
                    valores[etq] = self._suma_fex(df_sub[serie_pos_norm == etq_norm])

            bloque_rep = BloqueReplicado(
                hoja=hoja,
                dominio=b.dominio,
                condicion=b.condicion,
                valores=valores,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    b,
                    ws,
                    col,
                    layout.etiqueta_ultimo_trimestre,
                )
            )

        return self.validador.construir_resultado_hoja(
            hoja=hoja,
            trimestre=layout.etiqueta_ultimo_trimestre,
            celdas=celdas,
        )

    def _replicar_lugar(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        """Lugar de trabajo: mapeo P6880 → etiqueta Excel."""
        hoja = "Lugar de trabajo"
        parser = ExcelLayoutParser(self.ruta_excel)
        with parser:
            layout = parser.parsear_bloques_por_dominio_y_condicion(hoja_nombre)
        col = layout.col_ultimo_trimestre
        celdas: list[CeldaComparada] = []

        if "P6880" not in df.columns:
            return ResultadoHoja(
                hoja=hoja,
                estado="ABIERTA_CON_CAUSA",
                trimestre=layout.etiqueta_ultimo_trimestre,
                causa_raiz="df no tiene P6880 (lugar de trabajo)",
            )

        # Construir etiqueta de lugar en el df
        serie_lugar = (
            pd.to_numeric(df["P6880"], errors="coerce")
            .map(MAPEO_LUGAR_TRABAJO)
            .fillna("")
            .astype(str)
            .map(normalizar_texto)
        )

        for b in layout.bloques:
            mask = self._mask_dominio(df, b.dominio)
            df_ocu = df[mask & (df["OCI"] == 1)]
            if b.condicion == "Total":
                df_sub = df_ocu
            elif b.condicion == "Formal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 0]
            elif b.condicion == "Informal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 1]
            else:
                continue

            lugar_sub = serie_lugar.loc[df_sub.index]

            valores: dict[str, float] = {}
            for _, etq in b.categorias:
                etq_norm = normalizar_texto(etq)
                valores[etq] = self._suma_fex(df_sub[lugar_sub == etq_norm])

            bloque_rep = BloqueReplicado(
                hoja=hoja,
                dominio=b.dominio,
                condicion=b.condicion,
                valores=valores,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    b,
                    ws,
                    col,
                    layout.etiqueta_ultimo_trimestre,
                )
            )

        return self.validador.construir_resultado_hoja(
            hoja=hoja,
            trimestre=layout.etiqueta_ultimo_trimestre,
            celdas=celdas,
        )

    def _replicar_tamano(
        self,
        df: pd.DataFrame,
        periodo: PeriodoMovil,
        ws,
        hoja_nombre: str,
    ) -> ResultadoHoja:
        """Tamaño de empresa: agrupación DANE (Micro/Pequeña/Mediana/Grande)."""
        hoja = "Tamaño de empresa"
        parser = ExcelLayoutParser(self.ruta_excel)
        with parser:
            layout = parser.parsear_bloques_por_dominio_y_condicion(hoja_nombre)
        col = layout.col_ultimo_trimestre
        celdas: list[CeldaComparada] = []

        if "P3069" not in df.columns:
            return ResultadoHoja(
                hoja=hoja,
                estado="ABIERTA_CON_CAUSA",
                trimestre=layout.etiqueta_ultimo_trimestre,
                causa_raiz="df no tiene P3069 (tamaño de empresa)",
            )

        serie_tam = (
            pd.to_numeric(df["P3069"], errors="coerce")
            .map(MAPEO_TAMANO_EMPRESA)
            .fillna("")
            .astype(str)
            .map(normalizar_texto)
        )

        for b in layout.bloques:
            mask = self._mask_dominio(df, b.dominio)
            df_ocu = df[mask & (df["OCI"] == 1)]
            if b.condicion == "Total":
                df_sub = df_ocu
            elif b.condicion == "Formal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 0]
            elif b.condicion == "Informal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 1]
            else:
                continue

            tam_sub = serie_tam.loc[df_sub.index]

            valores: dict[str, float] = {}
            for _, etq in b.categorias:
                etq_norm = normalizar_texto(etq)
                valores[etq] = self._suma_fex(df_sub[tam_sub == etq_norm])

            bloque_rep = BloqueReplicado(
                hoja=hoja,
                dominio=b.dominio,
                condicion=b.condicion,
                valores=valores,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    b,
                    ws,
                    col,
                    layout.etiqueta_ultimo_trimestre,
                )
            )

        return self.validador.construir_resultado_hoja(
            hoja=hoja,
            trimestre=layout.etiqueta_ultimo_trimestre,
            celdas=celdas,
        )

    def _replicar_con_categorias(
        self,
        df,
        periodo,
        ws,
        hoja_nombre,
        hoja_final,
        columna_categoria: str,
        mapeo_valor_a_excel: dict[str, str],
    ) -> ResultadoHoja:
        """Plantilla genérica: usa una columna del df y un mapeo a etiquetas Excel."""
        parser = ExcelLayoutParser(self.ruta_excel)
        with parser:
            layout = parser.parsear_bloques_por_dominio_y_condicion(hoja_nombre)
        col = layout.col_ultimo_trimestre
        celdas: list[CeldaComparada] = []

        if columna_categoria not in df.columns:
            return ResultadoHoja(
                hoja=hoja_final,
                estado="ABIERTA_CON_CAUSA",
                trimestre=layout.etiqueta_ultimo_trimestre,
                causa_raiz=f"df no tiene columna {columna_categoria}",
            )

        # Precomputar el label Excel para cada fila del df
        serie_cat_excel = (
            df[columna_categoria]
            .map(mapeo_valor_a_excel)
            .fillna("")
            .astype(str)
            .map(normalizar_texto)
        )

        for b in layout.bloques:
            mask = self._mask_dominio(df, b.dominio)
            df_ocu = df[mask & (df["OCI"] == 1)]
            if b.condicion == "Total":
                df_sub = df_ocu
            elif b.condicion == "Formal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 0]
            elif b.condicion == "Informal":
                df_sub = df_ocu[df_ocu["INFORMAL"] == 1]
            else:
                continue

            cat_sub = serie_cat_excel.loc[df_sub.index]

            valores: dict[str, float] = {}
            for _, etq in b.categorias:
                etq_norm = normalizar_texto(etq)
                valores[etq] = self._suma_fex(df_sub[cat_sub == etq_norm])

            bloque_rep = BloqueReplicado(
                hoja=hoja_final,
                dominio=b.dominio,
                condicion=b.condicion,
                valores=valores,
            )
            celdas.extend(
                self.validador.comparar_bloque(
                    bloque_rep,
                    b,
                    ws,
                    col,
                    layout.etiqueta_ultimo_trimestre,
                )
            )

        return self.validador.construir_resultado_hoja(
            hoja=hoja_final,
            trimestre=layout.etiqueta_ultimo_trimestre,
            celdas=celdas,
        )


# ═════════════════════════════════════════════════════════════════════
# Función pública de alto nivel
# ═════════════════════════════════════════════════════════════════════


def replicar_informalidad_excel(
    ruta_excel: Path,
    df_preparado: pd.DataFrame,
    periodo: PeriodoMovil,
    parametros: Optional[ParametrosValidacion] = None,
    hojas: Optional[list[str]] = None,
) -> ResultadoReplicacion:
    """API funcional para replicar las hojas de informalidad.

    Args:
        ruta_excel: Path al anexo oficial.
        df_preparado: DataFrame ya pasado por PreparadorGEIH.
        periodo: `PeriodoMovil` con los meses del trimestre objetivo.
        parametros: tolerancias (opcional).
        hojas: subset de hojas (opcional).

    Returns:
        `ResultadoReplicacion` con un `ResultadoHoja` por cada hoja.
    """
    rep = ReplicadorInformalidad(ruta_excel, parametros)
    return rep.replicar(df_preparado, periodo, hojas)
