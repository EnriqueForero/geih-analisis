"""
Lector programático del anexo Excel del Boletín DANE GEIH.

Diseñado para el anexo `anex-GEIH-feb2026.xlsx` y compatibles de otros meses.
La idea es **no hardcodear cifras**: siempre leerlas del Excel para que el
notebook pedagógico funcione para cualquier período que el usuario descargue.

Uso:
    lector = LectorBoletinDANE('docs/anex-GEIH-feb2026.xlsx')
    ref = lector.indicadores_nacionales(anio=2026, mes='Feb')
    # ref = {'TD': 9.2350, 'TGP': 64.7146, 'TO': 58.7383, ...}
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import openpyxl


class LectorBoletinDANE:
    """Extrae cifras de referencia del anexo Excel del Boletín DANE.

    El anexo tiene ~21 hojas, todas con el mismo patrón:
      - Filas 0-10: encabezados institucionales
      - Fila 11 (o 12): años (sparse: solo la primera columna de cada año)
      - Fila 12 (o 13): meses (Ene, Feb, …, Dic) o trimestres (Ene-Mar, Feb-Abr,…)
      - Filas siguientes: indicadores (TGP, TO, TD, poblaciones)

    Este lector abstrae esa geometría para que el notebook siga leyendo
    cifras de referencia incluso si el DANE publica otro mes.
    """

    def __init__(self, ruta: str | Path):
        self.ruta = Path(ruta)
        if not self.ruta.exists():
            raise FileNotFoundError(f"No existe el anexo Excel: {self.ruta}")
        self._wb = openpyxl.load_workbook(
            self.ruta, read_only=True, data_only=True
        )
        # Cache de hojas cargadas en memoria como lista de filas
        self._cache: dict[str, list[tuple]] = {}

    # ────────────────────────────────────────────────────────────────
    # Helpers internos
    # ────────────────────────────────────────────────────────────────

    def _rows(self, sheet_name: str) -> list[tuple]:
        if sheet_name not in self._cache:
            if sheet_name not in self._wb.sheetnames:
                raise KeyError(
                    f"Hoja '{sheet_name}' no existe en {self.ruta.name}. "
                    f"Hojas disponibles: {self._wb.sheetnames}"
                )
            sheet = self._wb[sheet_name]
            self._cache[sheet_name] = list(sheet.iter_rows(values_only=True))
        return self._cache[sheet_name]

    @staticmethod
    def _find_col(
        year_row: tuple,
        month_row: tuple,
        target_year: int,
        target_label: str,
    ) -> Optional[int]:
        """Devuelve el índice de columna (0-based) para (año, mes/trim)."""
        current_year = None
        target_norm = " ".join(str(target_label).split()).lower()
        for i in range(1, len(year_row)):
            if year_row[i] is not None:
                current_year = year_row[i]
            if current_year == target_year and month_row[i]:
                m_norm = " ".join(str(month_row[i]).split()).lower()
                if m_norm == target_norm:
                    return i
        return None

    def _extract_block(
        self,
        sheet_name: str,
        year_row_idx: int,
        month_row_idx: int,
        data_range: tuple[int, int],
        anio: int,
        mes: str,
    ) -> dict[str, float]:
        """Extrae {etiqueta: valor} para un bloque de filas en un período."""
        rows = self._rows(sheet_name)
        year_row = rows[year_row_idx]
        month_row = rows[month_row_idx]
        col = self._find_col(year_row, month_row, anio, mes)
        if col is None:
            raise ValueError(
                f"No se encontró ({anio}, '{mes}') en hoja '{sheet_name}'. "
                f"Verifique que el período exista en el anexo."
            )
        resultado: dict[str, float] = {}
        for r_idx in range(data_range[0], data_range[1]):
            lbl = rows[r_idx][0]
            if lbl and isinstance(lbl, str):
                clave = lbl.strip()
                val = rows[r_idx][col]
                if val is not None:
                    resultado[clave] = float(val)
        return resultado

    # ────────────────────────────────────────────────────────────────
    # Métodos públicos — extracciones de alto nivel
    # ────────────────────────────────────────────────────────────────

    def indicadores_nacionales(
        self, anio: int, mes: str
    ) -> dict[str, float]:
        """Indicadores mensuales Total nacional (TGP, TO, TD, PET, Ocupados, …).

        Ejemplo: ``lector.indicadores_nacionales(2026, 'Feb')``
        """
        return self._extract_block(
            "Total nacional",
            year_row_idx=11,
            month_row_idx=12,
            data_range=(13, 30),
            anio=anio,
            mes=mes,
        )

    def indicadores_cabeceras(
        self, anio: int, mes: str
    ) -> dict[str, float]:
        """Bloque Total Cabeceras de la hoja Total nacional."""
        return self._extract_block(
            "Total nacional",
            year_row_idx=11,
            month_row_idx=12,
            data_range=(35, 52),
            anio=anio,
            mes=mes,
        )

    def indicadores_13_ciudades(
        self, anio: int, mes: str
    ) -> dict[str, float]:
        """Indicadores mensuales Total 13 ciudades y A.M."""
        return self._extract_block(
            "Total 13 ciudades A.M.",
            year_row_idx=11,
            month_row_idx=12,
            data_range=(13, 30),
            anio=anio,
            mes=mes,
        )

    def indicadores_nacional_trim(
        self, anio: int, trim: str
    ) -> dict[str, float]:
        """Indicadores trimestrales Total nacional.

        ``trim`` es la etiqueta del Excel (ej: 'Dic 25 - Feb 26').
        """
        return self._extract_block(
            "Total nacional Trim",
            year_row_idx=11,
            month_row_idx=12,
            data_range=(13, 30),
            anio=anio,
            mes=trim,
        )

    def ramas_nacional(self, anio: int, mes: str) -> dict[str, float]:
        """Ocupados por rama CIIU (Total nacional, 13 ramas), en miles."""
        return self._extract_block(
            "Ocupados TN_T13_rama",
            year_row_idx=12,
            month_row_idx=13,
            data_range=(15, 30),
            anio=anio,
            mes=mes,
        )

    def posicion_ocupacional(
        self, anio: int, mes: str
    ) -> dict[str, float]:
        """Ocupados por posición ocupacional (Total nacional), en miles."""
        return self._extract_block(
            "Ocupados TN_posición",
            year_row_idx=12,
            month_row_idx=13,
            data_range=(15, 25),
            anio=anio,
            mes=mes,
        )

    def iml_hombres(self, anio: int, mes: str) -> dict[str, float]:
        """Indicadores Total Nacional — Hombres (bloque 1 de la hoja IML)."""
        return self._extract_block(
            "Total_nacional_IML_Sexo",
            year_row_idx=11,
            month_row_idx=12,
            data_range=(13, 27),
            anio=anio,
            mes=mes,
        )

    def iml_mujeres(self, anio: int, mes: str) -> dict[str, float]:
        """Indicadores Total Nacional — Mujeres (bloque 2 de la hoja IML)."""
        return self._extract_block(
            "Total_nacional_IML_Sexo",
            year_row_idx=11,
            month_row_idx=12,
            data_range=(32, 46),
            anio=anio,
            mes=mes,
        )

    def td_23_ciudades(
        self, anio: int, trim: str
    ) -> dict[str, float]:
        """TD trimestral por cada una de las 23 ciudades."""
        rows = self._rows("Total 23 ciudades A.M. Trim")
        year_row = rows[11]
        month_row = rows[12]
        col = self._find_col(year_row, month_row, anio, trim)
        if col is None:
            raise ValueError(
                f"No se encontró ({anio}, '{trim}') en '23 ciudades A.M. Trim'"
            )
        # Bloques: header at [9, 28, 47, ..., 446], each +19 rows, TD at header+7
        bases = [28, 47, 66, 85, 104, 123, 142, 161, 180, 199, 218, 237,
                 256, 275, 294, 313, 332, 351, 370, 389, 408, 427, 446]
        resultado: dict[str, float] = {}
        for b in bases:
            ciudad = str(rows[b][0]).strip() if rows[b][0] else f"row_{b}"
            td_val = rows[b + 7][col]
            if td_val is not None:
                resultado[ciudad] = float(td_val)
        return resultado

    # ────────────────────────────────────────────────────────────────
    # Introspección
    # ────────────────────────────────────────────────────────────────

    def hojas(self) -> list[str]:
        return list(self._wb.sheetnames)

    def __repr__(self) -> str:
        return f"LectorBoletinDANE({self.ruta.name}, {len(self._wb.sheetnames)} hojas)"
